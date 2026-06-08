# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

try:
    import FinanceDataReader as fdr
except Exception:
    fdr = None


BASE_DIR = Path("D:/Study/주식_데일리")
FILE_STEM = "데일리_기업스크리닝"
SHEET_MAIN = "주도주 찾기"
SHEET_DAILY = "데일리데이터"
SHEET_52W = "52주신고가"


@dataclass(frozen=True)
class ScreeningFile:
    path: Path
    date_key: str


def normalize_date(value: str) -> str:
    digits = re.sub(r"\D", "", str(value))
    if not re.fullmatch(r"20\d{6}", digits):
        raise ValueError(f"date must be YYYYMMDD: {value}")
    return digits


def file_date_key(path: Path) -> str:
    m = re.search(r"(20\d{6})", path.name)
    return m.group(1) if m else ""


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def iter_stock_rows(sheet: Any, start_row: int = 2, max_empty_streak: int = 300):
    empty = 0
    row = start_row
    while True:
        code = str(sheet.cell(row=row, column=3).value or "").strip()
        name = str(sheet.cell(row=row, column=4).value or "").strip()
        if code and name:
            empty = 0
            yield row, code, name
        else:
            empty += 1
            if empty >= max_empty_streak:
                break
        row += 1


def list_screening_files(base_dir: Path = BASE_DIR, require_52w: bool = True) -> list[ScreeningFile]:
    rows: list[ScreeningFile] = []
    for path in sorted(base_dir.glob(f"*_{FILE_STEM}.xls*")):
        date_key = file_date_key(path)
        if not date_key:
            continue
        try:
            wb = load_workbook(path, read_only=True, data_only=False, keep_links=False)
            try:
                names = set(wb.sheetnames)
                if SHEET_MAIN not in names:
                    continue
                if require_52w and SHEET_52W not in names:
                    continue
            finally:
                wb.close()
        except Exception:
            continue
        rows.append(ScreeningFile(path=path, date_key=date_key))
    return rows


def apply_s_weighted_score(files: list[ScreeningFile], w_today: float = 0.35, w_1w: float = 0.45, w_1m: float = 0.2) -> dict[str, int]:
    changed_cells = 0
    touched_files = 0
    skipped_files = 0
    for item in files:
        try:
            keep_vba = item.path.suffix.lower() == ".xlsm"
            wb = load_workbook(item.path, read_only=False, data_only=False, keep_vba=keep_vba, keep_links=False)
            try:
                ws = wb[SHEET_MAIN]
                changed = 0
                for row, _code, _name in iter_stock_rows(ws):
                    o_val = to_number(ws.cell(row=row, column=15).value)
                    q_val = to_number(ws.cell(row=row, column=17).value)
                    r_val = to_number(ws.cell(row=row, column=18).value)
                    if o_val is None or o_val == -100000:
                        s_val = -100000
                    else:
                        q_num = q_val if q_val is not None else o_val
                        r_num = r_val if r_val is not None else o_val
                        s_val = round(o_val * w_today + r_num * w_1w + q_num * w_1m, 2)
                    cell = ws.cell(row=row, column=19)
                    if cell.value != s_val:
                        cell.value = s_val
                        changed += 1
                if changed:
                    wb.save(item.path)
                    touched_files += 1
                    changed_cells += changed
            finally:
                wb.close()
        except Exception:
            skipped_files += 1
            continue
    return {"files": len(files), "touched_files": touched_files, "changed_cells": changed_cells, "skipped_files": skipped_files}


def update_daily_sheet_with_fdr(target_path: Path) -> int:
    if fdr is None:
        return 0
    keep_vba = target_path.suffix.lower() == ".xlsm"
    wb = load_workbook(target_path, read_only=False, data_only=False, keep_vba=keep_vba, keep_links=False)
    updated = 0
    try:
        if SHEET_DAILY not in wb.sheetnames:
            return 0
        ws = wb[SHEET_DAILY]
        target_date = datetime.strptime(file_date_key(target_path), "%Y%m%d")
        # 데일리데이터 시트는 종목코드/종목명 위치가 주도주 시트와 다를 수 있어 직접 스캔한다.
        row = 2
        empty = 0
        while True:
            code = str(ws.cell(row=row, column=2).value or "").strip()
            name = str(ws.cell(row=row, column=3).value or "").strip()
            if not code and not name:
                empty += 1
                if empty >= 300:
                    break
                row += 1
                continue
            empty = 0
            norm_code = re.sub(r"\D", "", code).zfill(6)
            if len(norm_code) != 6 or norm_code == "000000":
                row += 1
                continue
            try:
                frame = fdr.DataReader(norm_code, target_date.strftime("%Y-%m-%d"), target_date.strftime("%Y-%m-%d"))
            except Exception:
                row += 1
                continue
            if frame is None or frame.empty:
                row += 1
                continue
            close = to_number(frame.iloc[-1].get("Close"))
            volume = to_number(frame.iloc[-1].get("Volume"))
            change = to_number(frame.iloc[-1].get("Change"))
            if close is not None:
                ws.cell(row=row, column=12).value = close
                updated += 1
            if volume is not None:
                ws.cell(row=row, column=10).value = volume
            if change is not None:
                ws.cell(row=row, column=7).value = change
            row += 1
        if updated:
            wb.save(target_path)
    finally:
        wb.close()
    return updated


def create_next_daily_file(target_date: str, base_dir: Path = BASE_DIR) -> Path:
    target = normalize_date(target_date)
    target_name_prefix = f"{target}_{FILE_STEM}"
    existing = sorted(base_dir.glob(f"{target_name_prefix}.xls*"))
    if existing:
        return existing[0]

    files = list_screening_files(base_dir=base_dir, require_52w=True)
    if not files:
        raise FileNotFoundError("No source screening file found")
    source = files[-1].path
    ext = source.suffix
    target_path = base_dir / f"{target_name_prefix}{ext}"
    shutil.copy2(source, target_path)
    return target_path


def run_build_for_date(target_date: str, base_dir: Path = BASE_DIR) -> dict[str, Any]:
    target_path = create_next_daily_file(target_date, base_dir=base_dir)
    fdr_updated = update_daily_sheet_with_fdr(target_path)

    all_files = list_screening_files(base_dir=base_dir, require_52w=True)
    score_stats = apply_s_weighted_score(all_files)
    return {
        "target_file": str(target_path),
        "fdr_daily_rows_updated": fdr_updated,
        "s_score_files": score_stats["files"],
        "s_score_touched_files": score_stats["touched_files"],
        "s_score_changed_cells": score_stats["changed_cells"],
        "s_score_skipped_files": score_stats.get("skipped_files", 0),
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build and update daily screening files.")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p1 = sub.add_parser("apply-s-score")
    p1.add_argument("--base-dir", default=str(BASE_DIR))
    p1.add_argument("--w-today", type=float, default=0.35)
    p1.add_argument("--w-1w", type=float, default=0.45)
    p1.add_argument("--w-1m", type=float, default=0.2)

    p2 = sub.add_parser("build-date")
    p2.add_argument("--base-dir", default=str(BASE_DIR))
    p2.add_argument("--date", required=True)
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.cmd == "apply-s-score":
        files = list_screening_files(base_dir=Path(args.base_dir), require_52w=True)
        out = apply_s_weighted_score(files, w_today=args.w_today, w_1w=args.w_1w, w_1m=args.w_1m)
        print(f"files={out['files']}")
        print(f"touched_files={out['touched_files']}")
        print(f"changed_cells={out['changed_cells']}")
        print(f"skipped_files={out.get('skipped_files', 0)}")
    elif args.cmd == "build-date":
        out = run_build_for_date(args.date, base_dir=Path(args.base_dir))
        print(f"target_file={out['target_file']}")
        print(f"fdr_daily_rows_updated={out['fdr_daily_rows_updated']}")
        print(f"s_score_files={out['s_score_files']}")
        print(f"s_score_touched_files={out['s_score_touched_files']}")
        print(f"s_score_changed_cells={out['s_score_changed_cells']}")
        print(f"s_score_skipped_files={out['s_score_skipped_files']}")


if __name__ == "__main__":
    main()
