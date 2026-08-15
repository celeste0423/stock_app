# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCREENING_DIR = PROJECT_ROOT / "data" / "screening" / "legacy"
FILE_STEM = "데일리_기업스크리닝"
SCREENING_SHEET = "주도주 찾기"
HIGH52_SHEET = "52주신고가"
DB_PATH = PROJECT_ROOT / "outputs" / "screening_avg_cache.sqlite3"


@dataclass(frozen=True)
class ScreeningFile:
    path: Path
    date_key: str


def normalize_date(value: str | None) -> str:
    if not value:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if not re.fullmatch(r"20\d{6}", digits):
        raise ValueError(f"date must be YYYYMMDD: {value}")
    return digits


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def file_date(path: Path) -> str:
    match = re.search(r"(20\d{6})", path.name)
    return match.group(1) if match else ""


def workbook_has_required_sheets(path: Path, require_high52_sheet: bool = True) -> bool:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except Exception:
        return False
    try:
        names = set(workbook.sheetnames)
        if SCREENING_SHEET not in names:
            return False
        if require_high52_sheet and HIGH52_SHEET not in names:
            return False
        return True
    finally:
        workbook.close()


def iter_stock_rows(sheet: Any, *, start_row: int = 2, max_empty_streak: int = 300):
    empty = 0
    row = start_row
    while True:
        code = str(sheet.cell(row=row, column=3).value or "").strip()
        name = str(sheet.cell(row=row, column=4).value or "").strip()
        if code and name:
            empty = 0
            yield row, code, name
        else:
            empty += 1
            if empty >= max_empty_streak:
                break
        row += 1


def list_screening_files(
    *,
    base_dir: Path = SCREENING_DIR,
    start_date: str | None = None,
    end_date: str | None = None,
    require_high52_sheet: bool = True,
) -> list[ScreeningFile]:
    start = normalize_date(start_date) if start_date else ""
    end = normalize_date(end_date) if end_date else ""
    result: list[ScreeningFile] = []
    for path in sorted(base_dir.glob(f"*_{FILE_STEM}.xls*")):
        date_key = file_date(path)
        if not date_key:
            continue
        if start and date_key < start:
            continue
        if end and date_key > end:
            continue
        if not workbook_has_required_sheets(path, require_high52_sheet=require_high52_sheet):
            continue
        result.append(ScreeningFile(path=path, date_key=date_key))
    return result


def init_db(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        PRAGMA journal_mode=WAL;
        PRAGMA synchronous=NORMAL;
        CREATE TABLE IF NOT EXISTS scores (
            file_date TEXT NOT NULL,
            stock_code TEXT NOT NULL,
            score REAL NOT NULL,
            PRIMARY KEY (file_date, stock_code)
        );
        CREATE INDEX IF NOT EXISTS idx_scores_code_date ON scores(stock_code, file_date);
        """
    )
    conn.commit()


def rebuild_score_table(conn: sqlite3.Connection, files: list[ScreeningFile]) -> dict[str, int]:
    conn.execute("DELETE FROM scores")
    inserted = 0
    per_file: dict[str, int] = {}
    skipped = 0
    for item in files:
        try:
            workbook = load_workbook(item.path, read_only=True, data_only=True, keep_links=False)
            try:
                sheet = workbook[SCREENING_SHEET]
                rows = []
                for row_idx, code, _name in iter_stock_rows(sheet):
                    score = to_number(sheet.cell(row=row_idx, column=15).value)
                    if score is None or score == -100000:
                        continue
                    rows.append((item.date_key, code, score))
                if rows:
                    conn.executemany(
                        "INSERT OR REPLACE INTO scores(file_date, stock_code, score) VALUES (?, ?, ?)",
                        rows,
                    )
                per_file[item.date_key] = len(rows)
                inserted += len(rows)
            finally:
                workbook.close()
        except Exception:
            skipped += 1
            continue
    conn.commit()
    return {"inserted_rows": inserted, "files": len(files), "rows_by_date_count": len(per_file), "skipped_files": skipped}


def avg_for_code(conn: sqlite3.Connection, code: str, current_date: str, days: int) -> float | None:
    cursor = conn.execute(
        """
        SELECT AVG(score)
        FROM scores
        WHERE stock_code = ?
          AND file_date < ?
          AND file_date >= strftime('%Y%m%d', date(substr(?,1,4)||'-'||substr(?,5,2)||'-'||substr(?,7,2), ?))
        """,
        (code, current_date, current_date, current_date, current_date, f"-{days} day"),
    )
    value = cursor.fetchone()[0]
    return float(value) if value is not None else None


def set_recalc_on_open(workbook: Any) -> None:
    if hasattr(workbook, "calculation"):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        workbook.calculation.calcMode = "auto"


def apply_averages_to_excel(conn: sqlite3.Connection, files: list[ScreeningFile]) -> dict[str, int]:
    changed_cells = 0
    changed_files = 0
    eligible_rows = 0
    skipped_files = 0
    for item in files:
        try:
            keep_vba = item.path.suffix.lower() == ".xlsm"
            workbook = load_workbook(item.path, read_only=False, data_only=False, keep_vba=keep_vba, keep_links=False)
            try:
                sheet = workbook[SCREENING_SHEET]
                sheet.cell(row=1, column=17).value = "1M 평균"
                sheet.cell(row=1, column=18).value = "1W 평균"
                file_changed = 0
                for row_idx, code, _ in iter_stock_rows(sheet):
                    eligible_rows += 1
                    today = to_number(sheet.cell(row=row_idx, column=15).value)
                    if today is None or today == -100000:
                        today = 0.0
                    avg_1m = avg_for_code(conn, code, item.date_key, 30)
                    avg_1w = avg_for_code(conn, code, item.date_key, 7)
                    q_val = round(avg_1m if avg_1m is not None else today, 2)
                    r_val = round(avg_1w if avg_1w is not None else today, 2)
                    q_cell = sheet.cell(row=row_idx, column=17)
                    r_cell = sheet.cell(row=row_idx, column=18)
                    if q_cell.value != q_val:
                        q_cell.value = q_val
                        file_changed += 1
                    if r_cell.value != r_val:
                        r_cell.value = r_val
                        file_changed += 1
                if file_changed:
                    set_recalc_on_open(workbook)
                    workbook.save(item.path)
                    changed_files += 1
                    changed_cells += file_changed
            finally:
                workbook.close()
        except Exception:
            skipped_files += 1
            continue
    return {"changed_cells": changed_cells, "changed_files": changed_files, "eligible_rows": eligible_rows, "skipped_files": skipped_files}


def run_pipeline(
    *,
    base_dir: Path = SCREENING_DIR,
    db_path: Path = DB_PATH,
    start_date: str | None = None,
    end_date: str | None = None,
    require_high52_sheet: bool = True,
) -> dict[str, Any]:
    files = list_screening_files(
        base_dir=base_dir,
        start_date=start_date,
        end_date=end_date,
        require_high52_sheet=require_high52_sheet,
    )
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        init_db(conn)
        load_stats = rebuild_score_table(conn, files)
        apply_stats = apply_averages_to_excel(conn, files)
    finally:
        conn.close()
    return {
        "files": len(files),
        "db_rows": load_stats["inserted_rows"],
        "changed_files": apply_stats["changed_files"],
        "changed_cells": apply_stats["changed_cells"],
        "eligible_rows": apply_stats["eligible_rows"],
        "skipped_read_files": load_stats["skipped_files"],
        "skipped_write_files": apply_stats["skipped_files"],
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Compute 1M/1W averages from previous screening files via SQLite and write to Excel.")
    parser.add_argument("--base-dir", default=str(SCREENING_DIR))
    parser.add_argument("--db-path", default=str(DB_PATH))
    parser.add_argument("--start-date", default=None)
    parser.add_argument("--end-date", default=None)
    parser.add_argument("--include-no-52w", action="store_true")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    result = run_pipeline(
        base_dir=Path(args.base_dir),
        db_path=Path(args.db_path),
        start_date=args.start_date,
        end_date=args.end_date,
        require_high52_sheet=not args.include_no_52w,
    )
    print(f"files={result['files']}")
    print(f"db_rows={result['db_rows']}")
    print(f"changed_files={result['changed_files']}")
    print(f"changed_cells={result['changed_cells']}")
    print(f"eligible_rows={result['eligible_rows']}")
    print(f"skipped_read_files={result['skipped_read_files']}")
    print(f"skipped_write_files={result['skipped_write_files']}")


if __name__ == "__main__":
    main()
