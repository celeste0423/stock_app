# -*- coding: utf-8 -*-
from __future__ import annotations

import math
import re
import shutil
import sys
from copy import copy
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from statistics import median
from typing import Any

import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


SCREENING_DIR = Path("D:/Study/\uc8fc\uc2dd_\ub370\uc77c\ub9ac")
APP_DIR = Path(__file__).resolve().parents[1]
VENDOR_DIR = APP_DIR / "backend" / "vendor"
if VENDOR_DIR.exists():
    sys.path.insert(0, str(VENDOR_DIR))
try:
    import FinanceDataReader as fdr
except Exception:
    fdr = None

TARGET_DATE = "20260528"
OUTPUT_SUFFIX = "_\ud14c\uc2a4\ud2b8"
FILE_STEM = "\ub370\uc77c\ub9ac_\uae30\uc5c5\uc2a4\ud06c\ub9ac\ub2dd"
SCREENING_SHEET = "\uc8fc\ub3c4\uc8fc \ucc3e\uae30"
DAILY_SHEET = "\ub370\uc77c\ub9ac\ub370\uc774\ud130"
SETTINGS_SHEET = "\uc810\uc218\uc124\uc815"

K = {
    "rank": "\uc21c\uc704",
    "code": "\uc885\ubaa9\ucf54\ub4dc",
    "name": "\uc885\ubaa9\uba85",
    "industry": "\uc5c5\uc885\uad6c\ubd84",
    "target": "\ub300\uc0c1",
    "change": "\ub4f1\ub77d\ub960",
    "strength": "\uccb4\uacb0\uac15\ub3c4",
    "trading_value": "\uac70\ub798\ub300\uae08",
    "volume": "\uac70\ub798\ub7c9",
    "market_cap": "\uc2dc\uac00\ucd1d\uc561",
    "close": "\ud604\uc7ac\uac00",
    "score": "\uc810\uc218",
    "note": "\ube44\uace0",
    "avg_1m": "1M \ud3c9\uade0",
    "avg_1w": "1W \ud3c9\uade0",
    "stronger": "\uac15\ud574\uc9c4\uc815\ub3c4",
    "liquidity_metric": "\uac70\ub798\ub300\uae08^2/\uc2dc\ucd1d",
}

DEFAULT_Z_WEIGHT = 1.0
DEFAULT_A5D_WEIGHT = 0.1
DEFAULT_LIQUIDITY_MULTIPLIER = 200.0
DEFAULT_52W_WEIGHT = 10.0
DEFAULT_1W_WEIGHT = 0.5
DEFAULT_MIN_SCORE = 0.0
DEFAULT_MAX_SCORE = 100.0


def source_file() -> Path:
    return SCREENING_DIR / f"{TARGET_DATE}_{FILE_STEM}.xlsm"


def output_file() -> Path:
    return SCREENING_DIR / f"{TARGET_DATE}_{FILE_STEM}{OUTPUT_SUFFIX}.xlsm"


@dataclass
class DailyPoint:
    date_key: str
    close: float
    change_pct: float | None
    trading_value_100m: float | None
    market_cap_100m: float | None


def normalize_code(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.zfill(6) if digits else ""


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "nan", "None"}:
        return None
    cleaned = re.sub(r"[^0-9.+\-]", "", text.replace(",", ""))
    if cleaned in {"", "+", "-", "."}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_change_to_pct(value: Any) -> float | None:
    number = to_float(value)
    if number is None:
        return None
    return number * 100.0 if abs(number) <= 2 else number


def robust_tanh(values: list[float | None], scale: float = 2.0) -> list[float | None]:
    valid = np.array([float(value) for value in values if value is not None and math.isfinite(float(value))], dtype=float)
    if valid.size == 0:
        return [None for _ in values]
    med = float(np.median(valid))
    mad = float(np.median(np.abs(valid - med)))
    denom = 1.4826 * mad
    if not math.isfinite(denom) or denom <= 1e-9:
        denom = float(np.std(valid)) or 1.0
    output: list[float | None] = []
    for value in values:
        if value is None or not math.isfinite(float(value)):
            output.append(None)
        else:
            output.append(float(np.tanh(((float(value) - med) / denom) / scale)))
    return output


def clamp(value: float, low: float, high: float) -> float:
    return max(low, min(high, value))


def calculate_formula_raw(
    z_sor: float | None,
    mom_1w: float | None,
    a5d_norm: float | None,
    high52_signal: float | None,
    liquidity_metric: float | None,
) -> float | None:
    if z_sor is None or liquidity_metric is None:
        return None
    momentum_factor = max(0.05, 1.0 + DEFAULT_Z_WEIGHT * float(z_sor))
    recent_factor = max(0.05, 1.0 + DEFAULT_1W_WEIGHT * math.tanh(float(mom_1w or 0.0) / 20.0))
    anomaly_factor = max(0.25, 1.0 + DEFAULT_A5D_WEIGHT * math.tanh(float(a5d_norm or 0.0) / 2.0))
    high52_factor = max(0.25, 1.0 + DEFAULT_52W_WEIGHT * float(high52_signal or 0.0))
    liquidity_factor = math.log1p(max(0.0, float(liquidity_metric)) * DEFAULT_LIQUIDITY_MULTIPLIER)
    return math.log1p(momentum_factor * recent_factor * anomaly_factor * high52_factor * max(0.001, liquidity_factor))


def calibrate_score_scale(rows: list[dict[str, Any]]) -> tuple[float, float]:
    pairs = [
        (float(row["formula_raw"]), float(row["old_score"]))
        for row in rows
        if row.get("formula_raw") is not None
        and row.get("old_score") is not None
        and math.isfinite(float(row["formula_raw"]))
        and math.isfinite(float(row["old_score"]))
        and float(row["old_score"]) > -99999
    ]
    if len(pairs) < 10:
        return 40.0, 20.0
    raw = np.array([item[0] for item in pairs], dtype=float)
    old = np.array([item[1] for item in pairs], dtype=float)
    # The old sheet intentionally has a long negative tail. Calibrate against the
    # actionable upper band so 50+ point candidates stay close to the old scale.
    raw_p90, raw_p995 = np.percentile(raw, [90, 99.5])
    old_p90, old_p995 = np.percentile(old, [90, 99.5])
    if abs(raw_p995 - raw_p90) <= 1e-9:
        return 40.0, float(np.median(old) - 40.0 * np.median(raw))
    scale = float((old_p995 - old_p90) / (raw_p995 - raw_p90))
    offset = float(old_p90 - scale * raw_p90)
    return scale, offset


def screening_files() -> list[Path]:
    pattern = re.compile(rf"^(20\d{{6}})_{re.escape(FILE_STEM)}\.xlsm$", re.I)
    rows: list[tuple[str, Path]] = []
    for path in SCREENING_DIR.glob(f"*_{FILE_STEM}.xlsm"):
        match = pattern.match(path.name)
        if match and match.group(1) <= TARGET_DATE:
            rows.append((match.group(1), path))
    rows.sort(key=lambda item: item[0])
    return [path for _date, path in rows[-90:]]


def previous_screening_file() -> Path | None:
    previous = [path for path in screening_files() if re.search(r"(20\d{6})", path.name).group(1) < TARGET_DATE]
    return previous[-1] if previous else None


def header_map(sheet: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[1]:
        text = str(cell.value or "").strip()
        if text and text not in headers:
            headers[text] = cell.column
    return headers


def header_index_from_values(values: tuple[Any, ...]) -> dict[str, int]:
    headers: dict[str, int] = {}
    for index, value in enumerate(values):
        text = str(value or "").strip()
        if text and text not in headers:
            headers[text] = index
    return headers


def load_daily_history() -> dict[str, list[DailyPoint]]:
    history: dict[str, list[DailyPoint]] = {}
    for path in screening_files():
        date_match = re.search(r"(20\d{6})", path.name)
        if not date_match:
            continue
        date_key = date_match.group(1)
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            if DAILY_SHEET not in workbook.sheetnames:
                continue
            sheet = workbook[DAILY_SHEET]
            rows_iter = sheet.iter_rows(values_only=True)
            header_values = next(rows_iter, None)
            if not header_values:
                continue
            headers = header_index_from_values(header_values)
            code_col = headers.get(K["code"])
            close_col = headers.get(K["close"])
            change_col = headers.get(K["change"])
            trading_col = headers.get(K["trading_value"])
            market_cap_col = headers.get(K["market_cap"])
            if code_col is None or close_col is None:
                continue
            for values in rows_iter:
                code = normalize_code(values[code_col] if code_col < len(values) else None)
                close = to_float(values[close_col] if close_col < len(values) else None)
                if not code or close is None or close <= 0:
                    continue
                history.setdefault(code, []).append(
                    DailyPoint(
                        date_key=date_key,
                        close=close,
                        change_pct=normalize_change_to_pct(values[change_col]) if change_col is not None and change_col < len(values) else None,
                        trading_value_100m=to_float(values[trading_col]) if trading_col is not None and trading_col < len(values) else None,
                        market_cap_100m=to_float(values[market_cap_col]) if market_cap_col is not None and market_cap_col < len(values) else None,
                    )
                )
        finally:
            workbook.close()
    for points in history.values():
        points.sort(key=lambda point: point.date_key)
    return history


def fetch_fdr_points(code: str) -> list[DailyPoint]:
    if fdr is None:
        return []
    target = datetime.strptime(TARGET_DATE, "%Y%m%d")
    start = (target - timedelta(days=430)).strftime("%Y-%m-%d")
    end = target.strftime("%Y-%m-%d")
    try:
        frame = fdr.DataReader(code, start, end)
    except Exception:
        return []
    if frame is None or frame.empty or "Close" not in frame.columns:
        return []
    points: list[DailyPoint] = []
    for date_index, row in frame.iterrows():
        close = to_float(row.get("Close"))
        if close is None or close <= 0:
            continue
        change = row.get("Change") if "Change" in frame.columns else None
        points.append(
            DailyPoint(
                date_key=date_index.strftime("%Y%m%d"),
                close=close,
                change_pct=normalize_change_to_pct(change),
                trading_value_100m=None,
                market_cap_100m=None,
            )
        )
    return points


def supplement_history_with_fdr(rows: list[dict[str, Any]], history: dict[str, list[DailyPoint]]) -> int:
    if fdr is None:
        print("FinanceDataReader is not available; 3M uses workbook history only.", flush=True)
        return 0
    updated = 0
    for index, row in enumerate(rows, start=1):
        code = row["code"]
        if len(history.get(code, [])) >= 260:
            continue
        points = fetch_fdr_points(code)
        if len(points) >= 64:
            history[code] = points
            updated += 1
        if index % 100 == 0:
            print(f"FDR supplement {index}/{len(rows)} updated={updated}", flush=True)
    return updated


def load_previous_annotations() -> dict[str, dict[str, Any]]:
    previous = previous_screening_file()
    if not previous:
        return {}
    workbook = load_workbook(previous, read_only=False, data_only=False, keep_vba=True, keep_links=False)
    try:
        if SCREENING_SHEET not in workbook.sheetnames:
            return {}
        sheet = workbook[SCREENING_SHEET]
        annotations: dict[str, dict[str, Any]] = {}
        for row_index in range(2, sheet.max_row + 1):
            code = normalize_code(sheet.cell(row=row_index, column=3).value)
            if not code:
                continue
            annotations[code] = {
                "note": sheet.cell(row=row_index, column=16).value,
                "name_fill": copy(sheet.cell(row=row_index, column=4).fill),
            }
        return annotations
    finally:
        workbook.close()


def apply_previous_annotations(rows: list[dict[str, Any]], annotations: dict[str, dict[str, Any]]) -> None:
    for row in rows:
        previous = annotations.get(row["code"])
        if not previous:
            continue
        row["note"] = previous.get("note")
        row["name_fill"] = previous.get("name_fill")


def load_current_rows() -> list[dict[str, Any]]:
    workbook = load_workbook(source_file(), read_only=True, data_only=True, keep_vba=True, keep_links=False)
    try:
        sheet = workbook[SCREENING_SHEET]
        rows: list[dict[str, Any]] = []
        for values in sheet.iter_rows(min_row=2, max_col=19, values_only=True):
            code = normalize_code(values[2] if len(values) > 2 else None)
            name = values[3] if len(values) > 3 else None
            if not code or not name:
                continue
            rows.append(
                {
                    "rank": values[0] if len(values) > 0 else None,
                    "high52": values[1] if len(values) > 1 else None,
                    "code": code,
                    "name": name,
                    "industry": values[4] if len(values) > 4 else None,
                    "market": values[5] if len(values) > 5 else None,
                    "change": values[6] if len(values) > 6 else None,
                    "strength": values[7] if len(values) > 7 else None,
                    "trading_text": values[8] if len(values) > 8 else None,
                    "volume": values[9] if len(values) > 9 else None,
                    "market_cap_text": values[10] if len(values) > 10 else None,
                    "close": values[11] if len(values) > 11 else None,
                    "trading_value_100m": to_float(values[12] if len(values) > 12 else None),
                    "market_cap_100m": to_float(values[13] if len(values) > 13 else None),
                    "old_score": to_float(values[14] if len(values) > 14 else None),
                    "note": values[15] if len(values) > 15 else None,
                    "avg_1m": values[16] if len(values) > 16 else None,
                    "avg_1w": values[17] if len(values) > 17 else None,
                    "stronger": values[18] if len(values) > 18 else None,
                }
            )
        return rows
    finally:
        workbook.close()


def filter_large_cap_rows(rows: list[dict[str, Any]], minimum_market_cap_100m: float = 2000.0) -> list[dict[str, Any]]:
    return [row for row in rows if (to_float(row.get("market_cap_100m")) or 0.0) >= minimum_market_cap_100m]


def high52_decay_signal(closes: list[float]) -> float:
    if len(closes) < 252:
        return 0.0
    last_signal_index: int | None = None
    for index in range(251, len(closes)):
        window_high = max(closes[index - 251 : index + 1])
        if closes[index] >= window_high:
            last_signal_index = index
    if last_signal_index is None:
        return 0.0
    days_since = len(closes) - 1 - last_signal_index
    return round(max(0.0, 1.0 - 0.2 * days_since), 3)


def calculate_metrics(rows: list[dict[str, Any]], history: dict[str, list[DailyPoint]]) -> tuple[float, float]:
    mom_3m_values: list[float | None] = []
    mom_1w_values: list[float | None] = []
    sortino_values: list[float | None] = []
    a5d_values: list[float | None] = []

    for row in rows:
        points = history.get(row["code"], [])
        closes = [point.close for point in points]
        returns = [(closes[i] / closes[i - 1] - 1.0) * 100.0 for i in range(1, len(closes)) if closes[i - 1] > 0]
        current_change_pct = normalize_change_to_pct(row.get("change"))
        ivol = float(np.std(returns[-20:], ddof=1)) if len(returns) >= 20 else None
        sigma = None
        if ivol is not None and ivol > 1e-9 and current_change_pct is not None:
            sigma = clamp(abs(current_change_pct) / ivol - 1.0, -1.0, 1.0)

        current_close = closes[-1] if closes else to_float(row.get("close"))
        mom_3m = (current_close / closes[-64] - 1.0) * 100.0 if current_close and len(closes) >= 64 and closes[-64] > 0 else None
        mom_1w = (current_close / closes[-6] - 1.0) * 100.0 if current_close and len(closes) >= 6 and closes[-6] > 0 else None
        high52_signal = high52_decay_signal(closes)

        sortino = None
        if returns:
            sample = np.array(returns[-63:], dtype=float)
            downside = sample[sample < 0]
            downside_std = float(np.std(downside, ddof=1)) if downside.size >= 2 else None
            if downside_std and downside_std > 1e-9:
                sortino = float(np.mean(sample) / downside_std * math.sqrt(252))

        a5d = None
        if len(closes) >= 12:
            five_day_returns = [(closes[i] / closes[i - 5] - 1.0) * 100.0 for i in range(5, len(closes)) if closes[i - 5] > 0]
            if len(five_day_returns) >= 8:
                recent = five_day_returns[-1]
                baseline = five_day_returns[:-1][-60:]
                std = float(np.std(baseline, ddof=1)) if len(baseline) >= 5 else 0.0
                center = float(median(baseline)) if baseline else 0.0
                if std > 1e-9:
                    a5d = (recent - center) / std

        row.update({"IVol": ivol, "1sigma": sigma, "3M": mom_3m, "1W": mom_1w, "52W": high52_signal, "Sor_raw": sortino, "A_5d": a5d})
        mom_3m_values.append(mom_3m)
        mom_1w_values.append(mom_1w)
        sortino_values.append(sortino)
        a5d_values.append(a5d)

    mom_3m_norm = robust_tanh(mom_3m_values)
    mom_1w_norm = robust_tanh(mom_1w_values)
    sortino_norm = robust_tanh(sortino_values)
    a5d_norm = robust_tanh(a5d_values)

    for index, row in enumerate(rows):
        z = None
        if mom_3m_norm[index] is not None or mom_1w_norm[index] is not None:
            z = 0.65 * float(mom_3m_norm[index] or 0.0) + 0.35 * float(mom_1w_norm[index] or 0.0)
            z = clamp(z, -1.0, 1.0)
        sor = sortino_norm[index]
        z_sor = None
        if z is not None or sor is not None:
            z_sor = (float(z or 0.0) + float(sor or 0.0)) / 2.0
            z_sor = clamp(z_sor, -1.0, 1.0)
        trading_value = to_float(row.get("trading_value_100m"))
        market_cap = to_float(row.get("market_cap_100m"))
        liquidity_metric = (trading_value * trading_value / market_cap) if trading_value is not None and market_cap and market_cap > 0 else None
        a_norm = float(a5d_norm[index] or 0.0)
        formula_raw = calculate_formula_raw(z_sor, row.get("1W"), a_norm, row.get("52W"), liquidity_metric)
        row.update(
            {
                "Z": z,
                "Sor": sor,
                "Z_Sor": z_sor,
                "A_5d_norm": a_norm,
                "liquidity_metric": liquidity_metric,
                "formula_raw": formula_raw,
            }
        )

    scale, offset = calibrate_score_scale(rows)
    for row in rows:
        raw = row.get("formula_raw")
        row["sort_score"] = clamp(offset + scale * float(raw), DEFAULT_MIN_SCORE, DEFAULT_MAX_SCORE) if raw is not None else -100000
    return scale, offset


def copy_column_style(sheet: Any, source_col: int, target_col: int, max_row: int) -> None:
    sheet.column_dimensions[sheet.cell(row=1, column=target_col).column_letter].width = 12
    for row_index in range(1, max_row + 1):
        source = sheet.cell(row=row_index, column=source_col)
        target = sheet.cell(row=row_index, column=target_col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)


def round_or_none(value: Any, digits: int = 2) -> Any:
    number = to_float(value)
    if number is None or not math.isfinite(number):
        return None
    return round(number, digits)


def score_formula(row_index: int) -> str:
    mom_1w_cell = f"Q{row_index}"
    z_sor_cell = f"T{row_index}"
    a5d_cell = f"U{row_index}"
    high52_cell = f"V{row_index}"
    liquidity_cell = f"W{row_index}"
    s = f"'{SETTINGS_SHEET}'!"
    return (
        f"=IFERROR(MAX({s}$B$9,MIN({s}$B$10,"
        f"{s}$B$8+{s}$B$7*LN(1+"
        f"MAX(0.05,1+{s}$B$2*{z_sor_cell})*"
        f"MAX(0.05,1+{s}$B$6*TANH({mom_1w_cell}/20))*"
        f"MAX(0.25,1+{s}$B$3*TANH({a5d_cell}/2))*"
        f"MAX(0.25,1+{s}$B$5*{high52_cell})*"
        f"LN(1+MAX(0,{liquidity_cell})*{s}$B$4)"
        f"))),-100000)"
    )


def write_settings_sheet(workbook: Any, scale: float, offset: float) -> None:
    if SETTINGS_SHEET in workbook.sheetnames:
        del workbook[SETTINGS_SHEET]
    sheet = workbook.create_sheet(SETTINGS_SHEET)
    rows = [
        ("\uc2e0\uaddc \uc8fc\ub3c4\uc8fc \uc810\uc218 \uc124\uc815", None),
        ("Z_Sor \uac00\uc911\uce58", DEFAULT_Z_WEIGHT),
        ("A_5d \uac00\uc911\uce58", DEFAULT_A5D_WEIGHT),
        ("\uac70\ub798\ub300\uae08^2/\uc2dc\ucd1d \ubc30\uc728", DEFAULT_LIQUIDITY_MULTIPLIER),
        ("52W \uc2e0\uace0 \uac00\uc911\uce58", DEFAULT_52W_WEIGHT),
        ("1W \ucd94\uc138 \uac00\uc911\uce58", DEFAULT_1W_WEIGHT),
        ("\uc810\uc218 \uc2a4\ucf00\uc77c", round(scale, 6)),
        ("\uc810\uc218 \uc624\ud504\uc14b", round(offset, 6)),
        ("\ucd5c\uc800\uc810", DEFAULT_MIN_SCORE),
        ("\ucd5c\uace0\uc810", DEFAULT_MAX_SCORE),
        (None, None),
        ("\uc218\uc2dd", "score = offset + scale * LN(1 + momentum * anomaly * liquidity)"),
        ("\uc815\ub82c", "\uc774 \uc2dc\ud2b8\uc758 B2:B8 \uac12\uc744 \ubc14\uafb8\uace0 \uc810\uc218\ub97c \uc7ac\uacc4\uc0b0\ud55c \ub4a4 \uc810\uc218\ub85c \uc815\ub82c\ud558\uba74 \ub429\ub2c8\ub2e4."),
    ]
    for row_index, (label, value) in enumerate(rows, start=1):
        sheet.cell(row=row_index, column=1).value = label
        sheet.cell(row=row_index, column=2).value = value
    sheet["A1"].font = Font(bold=True, size=14)
    sheet["A1"].fill = PatternFill("solid", fgColor="EAF2FF")
    for row_index in range(2, 11):
        sheet.cell(row=row_index, column=1).font = Font(bold=True)
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 22
    sheet.column_dimensions["C"].width = 70


def write_test_workbook(rows: list[dict[str, Any]], scale: float, offset: float) -> None:
    source = source_file()
    output = output_file()
    if not source.exists():
        raise FileNotFoundError(source)
    shutil.copy2(source, output)
    workbook = load_workbook(output, keep_vba=True, keep_links=False)
    try:
        sheet = workbook[SCREENING_SHEET]
        sheet.delete_cols(8, 1)
        sheet.insert_cols(14, 10)
        for column in range(14, 24):
            copy_column_style(sheet, 13, column, sheet.max_row)

        headers = [
            K["rank"], "52\uc2e0\uace0", K["code"], K["name"], K["industry"], K["target"],
            K["change"], K["trading_value"], K["volume"], K["market_cap"], K["close"],
            K["trading_value"], K["market_cap"], "IVol", "1\u03c3", "3M", "1W", "Z", "Sor", "Z_Sor",
            "A_5d", "52W", K["liquidity_metric"], K["score"], K["note"], K["avg_1m"], K["avg_1w"], K["stronger"],
        ]
        for column, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=column).value = header

        sorted_rows = sorted(
            rows,
            key=lambda item: (
                float(item.get("sort_score") if item.get("sort_score") is not None else -100000),
                float(item.get("liquidity_metric") or 0.0),
            ),
            reverse=True,
        )
        for row_index, row in enumerate(sorted_rows, start=2):
            values = [
                row_index - 1,
                row.get("high52"),
                row.get("code"),
                row.get("name"),
                row.get("industry"),
                row.get("market"),
                row.get("change"),
                row.get("trading_text"),
                row.get("volume"),
                row.get("market_cap_text"),
                row.get("close"),
                round_or_none(row.get("trading_value_100m"), 4),
                round_or_none(row.get("market_cap_100m"), 2),
                round_or_none(row.get("IVol"), 2),
                round_or_none(row.get("1sigma"), 3),
                round_or_none(row.get("3M"), 2),
                round_or_none(row.get("1W"), 2),
                round_or_none(row.get("Z"), 3),
                round_or_none(row.get("Sor"), 3),
                round_or_none(row.get("Z_Sor"), 3),
                round_or_none(row.get("A_5d"), 2),
                round_or_none(row.get("52W"), 3),
                round_or_none(row.get("liquidity_metric"), 3),
                score_formula(row_index),
                row.get("note"),
                row.get("avg_1m"),
                row.get("avg_1w"),
                row.get("stronger"),
            ]
            for column, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column).value = value
            if row.get("name_fill") is not None:
                sheet.cell(row=row_index, column=4).fill = copy(row["name_fill"])

        for row_index in range(len(sorted_rows) + 2, sheet.max_row + 1):
            for column in range(1, 30):
                sheet.cell(row=row_index, column=column).value = None

        width_overrides = {
            3: 10, 4: 16, 14: 9, 15: 8, 16: 9, 17: 9, 18: 8, 19: 8, 20: 9, 21: 9,
            22: 7, 23: 14, 24: 10, 25: 20, 26: 9, 27: 9, 28: 11,
        }
        for column, width in width_overrides.items():
            sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = width
        for row_index in range(2, len(sorted_rows) + 2):
            sheet.cell(row=row_index, column=24).number_format = "0.00"

        write_settings_sheet(workbook, scale, offset)
        if hasattr(workbook, "calculation"):
            workbook.calculation.fullCalcOnLoad = True
            workbook.calculation.forceFullCalc = True
            workbook.calculation.calcMode = "auto"
        workbook.save(output)
    finally:
        workbook.close()


def main() -> None:
    global OUTPUT_SUFFIX, TARGET_DATE
    if len(sys.argv) >= 2:
        requested_date = sys.argv[1].strip()
        if not re.fullmatch(r"20\d{6}", requested_date):
            raise ValueError("date argument must be YYYYMMDD, e.g. 20260521")
        TARGET_DATE = requested_date
    if len(sys.argv) >= 3:
        OUTPUT_SUFFIX = sys.argv[2].strip()
        if not OUTPUT_SUFFIX.startswith("_"):
            OUTPUT_SUFFIX = f"_{OUTPUT_SUFFIX}"
    source = source_file()
    if not source.exists():
        raise FileNotFoundError(f"source workbook not found: {source}")
    print(f"source={source}", flush=True)
    print("loading daily history...", flush=True)
    history = load_daily_history()
    print(f"loaded history codes={len(history)}", flush=True)
    print("loading current rows...", flush=True)
    rows = load_current_rows()
    annotations = load_previous_annotations()
    apply_previous_annotations(rows, annotations)
    before_filter = len(rows)
    rows = filter_large_cap_rows(rows)
    print(f"loaded rows={before_filter} large_cap_rows={len(rows)}", flush=True)
    print("supplementing 3M history with FinanceDataReader...", flush=True)
    fdr_count = supplement_history_with_fdr(rows, history)
    print(f"FDR supplemented codes={fdr_count}", flush=True)
    print("calculating metrics...", flush=True)
    scale, offset = calculate_metrics(rows, history)
    print("writing workbook...", flush=True)
    write_test_workbook(rows, scale, offset)
    valid_scores = [float(row["sort_score"]) for row in rows if to_float(row.get("sort_score")) is not None and float(row["sort_score"]) > -99999]
    print(f"created={output_file()}")
    print(f"rows={len(rows)} valid_scores={len(valid_scores)}")
    if valid_scores:
        print(f"score_min={min(valid_scores):.2f} score_median={np.median(valid_scores):.2f} score_max={max(valid_scores):.2f}")
    print(f"scale={scale:.6f} offset={offset:.6f}")


if __name__ == "__main__":
    main()
