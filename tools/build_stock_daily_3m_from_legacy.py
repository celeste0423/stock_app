from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


LEGACY_DIR = Path("D:/Study/주식_데일리")
OUT_DIR = Path("D:/Study/Stock_Daily")
SECTOR_DB_PATH = Path("D:/Study/stock app/backend/sector_database.json")
SHEET_NAME = "주도주 찾기"


@dataclass
class LegacyFile:
    path: Path
    date_key: str
    dt: datetime


def _to_float(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return default
    s = str(v).strip().replace(",", "")
    if not s:
        return default
    try:
        return float(s)
    except Exception:
        return default


def _norm_code(v: Any) -> str:
    d = re.sub(r"\D", "", str(v or ""))
    return d.zfill(6) if d else ""


def _read_sector_map() -> dict[str, str]:
    import json

    if not SECTOR_DB_PATH.exists():
        return {}
    try:
        payload = json.loads(SECTOR_DB_PATH.read_text(encoding="utf-8"))
        smap = payload.get("stock_map", {}) if isinstance(payload, dict) else {}
        out: dict[str, str] = {}
        for key, item in smap.items():
            if not isinstance(item, dict):
                continue
            code = _norm_code(item.get("stock_code") or key)
            sec = str(item.get("sector") or "").strip()
            if code and sec:
                out[code] = sec
        return out
    except Exception:
        return {}


def _legacy_files_last_3m() -> list[LegacyFile]:
    rows: list[LegacyFile] = []
    for p in sorted(LEGACY_DIR.glob("*_데일리_기업스크리닝.xls*")):
        m = re.search(r"(20\d{6})", p.name)
        if not m:
            continue
        key = m.group(1)
        try:
            dt = datetime.strptime(key, "%Y%m%d")
        except Exception:
            continue
        rows.append(LegacyFile(path=p, date_key=key, dt=dt))
    if not rows:
        return []
    max_dt = max(r.dt for r in rows)
    start_dt = max_dt - timedelta(days=92)
    return [r for r in rows if r.dt >= start_dt]


def _read_one_legacy(file: LegacyFile) -> pd.DataFrame:
    wb = load_workbook(file.path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            return pd.DataFrame()
        ws = wb[SHEET_NAME]
        recs: list[dict[str, Any]] = []
        empty = 0
        r = 2
        while True:
            code = _norm_code(ws.cell(r, 3).value)
            name = str(ws.cell(r, 4).value or "").strip()
            if not code and not name:
                empty += 1
                if empty >= 250:
                    break
                r += 1
                continue
            empty = 0
            if not code or not name:
                r += 1
                continue
            recs.append(
                {
                    "date_key": file.date_key,
                    "date": file.dt,
                    "rank_raw": _to_float(ws.cell(r, 1).value, 0.0),
                    "code": code,
                    "name": name,
                    "industry": str(ws.cell(r, 5).value or "").strip(),
                    "change": _to_float(ws.cell(r, 7).value, 0.0),  # decimal return (e.g. 0.0931)
                    "amount_100m": _to_float(ws.cell(r, 13).value, 0.0),
                    "marcap_100m": _to_float(ws.cell(r, 14).value, 0.0),
                    "close": _to_float(ws.cell(r, 12).value, 0.0),
                    "note": str(ws.cell(r, 16).value or "").strip(),
                }
            )
            r += 1
        return pd.DataFrame(recs)
    finally:
        wb.close()


def _sortino_norm(returns: np.ndarray) -> float:
    if returns.size == 0:
        return 0.5
    downside = np.minimum(returns, 0.0)
    downside_dev = float(np.sqrt(np.mean(np.square(downside))))
    if downside_dev <= 1e-8:
        downside_dev = 1e-8
    ratio = float(np.mean(returns) / downside_dev)
    return float(1.0 / (1.0 + math.exp(-ratio)))


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    sector_map = _read_sector_map()
    files = _legacy_files_last_3m()
    if not files:
        print("[ERROR] 최근 3개월 대상 파일이 없습니다.")
        return 1

    frames: list[pd.DataFrame] = []
    for f in files:
        df = _read_one_legacy(f)
        if not df.empty:
            frames.append(df)
    if not frames:
        print("[ERROR] 원본 시트에서 읽은 행이 없습니다.")
        return 1

    all_df = pd.concat(frames, ignore_index=True)
    all_df = all_df.sort_values(["code", "date"]).reset_index(drop=True)

    # 52주 신고가(종가 기준, 최근 252거래일)
    all_df["rolling_max_252"] = (
        all_df.groupby("code", sort=False)["close"]
        .transform(lambda s: s.rolling(window=252, min_periods=1).max())
    )
    all_df["is_52w_high"] = np.where(all_df["close"] >= all_df["rolling_max_252"], 1, 0)

    # 점수: LOG((M^2/N^0.8)*(1.1+G)^4,1.1) + IF(52주신고가,+4,-4) - 13
    M = np.maximum(all_df["amount_100m"].astype(float), 0.0)
    N = np.maximum(all_df["marcap_100m"].astype(float), 0.0)
    G = all_df["change"].astype(float)  # decimal return
    core = np.where(N > 0, (np.power(M, 2.0) / np.power(N, 0.8)) * np.power(1.1 + G, 4.0), np.nan)
    core = np.where(core > 0, core, np.nan)
    log_term = np.log(core) / np.log(1.1)
    bonus = np.where(all_df["is_52w_high"] == 1, 4.0, -4.0)
    all_df["score"] = np.where(np.isfinite(log_term), log_term + bonus - 13.0, -100000.0)

    # 1W/1M 평균 점수 (당일 제외)
    all_df["avg_1w"] = (
        all_df.groupby("code", sort=False)["score"]
        .transform(lambda s: s.shift(1).rolling(window=7, min_periods=1).mean())
    )
    all_df["avg_1m"] = (
        all_df.groupby("code", sort=False)["score"]
        .transform(lambda s: s.shift(1).rolling(window=30, min_periods=1).mean())
    )
    all_df["avg_1w"] = all_df["avg_1w"].fillna(all_df["score"])
    all_df["avg_1m"] = all_df["avg_1m"].fillna(all_df["score"])

    # Sortino 정규화(60거래일, 당일 포함)
    def _roll_sortino(s: pd.Series) -> pd.Series:
        vals = s.to_numpy(dtype=float)
        out = np.zeros_like(vals, dtype=float)
        for i in range(len(vals)):
            st = max(0, i - 59)
            out[i] = _sortino_norm(vals[st : i + 1])
        return pd.Series(out, index=s.index)

    all_df["sortino_norm"] = all_df.groupby("code", sort=False)["change"].apply(_roll_sortino).reset_index(level=0, drop=True)

    # 종합점수: ((오늘*0.35)+(1W*0.45)+(1M*0.2))*(Sortino^0.8)
    blended = (all_df["score"] * 0.35) + (all_df["avg_1w"] * 0.45) + (all_df["avg_1m"] * 0.2)
    all_df["final_score"] = blended * np.power(np.clip(all_df["sortino_norm"], 1e-6, None), 0.8)

    all_df["sector"] = all_df["code"].map(lambda c: sector_map.get(c, ""))
    all_df["change_pct"] = all_df["change"] * 100.0

    # 일자별 파일 출력
    for date_key, dfg in all_df.groupby("date_key", sort=True):
        dfg = dfg[dfg["marcap_100m"] >= 2000.0].copy()
        if dfg.empty:
            continue
        dfg = dfg.sort_values(["final_score", "score"], ascending=False).reset_index(drop=True)
        dfg["rank"] = np.arange(1, len(dfg) + 1)

        out = pd.DataFrame(
            {
                "순위": dfg["rank"].astype(int),
                "섹터": dfg["sector"].fillna(""),
                "종목코드": dfg["code"],
                "종목 이름": dfg["name"],
                "업종": dfg["industry"],
                "시총 (억원)": dfg["marcap_100m"].round(0).astype(int),
                "거래대금 (억원)": dfg["amount_100m"].round(0).astype(int),
                "등락률": dfg["change_pct"].round(2),
                "점수": dfg["score"].round(2),
                "1W 평균 점수": dfg["avg_1w"].round(2),
                "1M 평균 점수": dfg["avg_1m"].round(2),
                "20일 기준 Sortino 정규화 점수": dfg["sortino_norm"].round(4),
                "종합 점수": dfg["final_score"].round(2),
                "비고": dfg["note"].fillna(""),
            }
        )
        out_path = OUT_DIR / f"{date_key}_데일리_기업스크리닝.xlsx"
        out.to_excel(out_path, index=False)
        print(f"[DONE] {out_path}")

    print(f"[SUMMARY] generated_dates={all_df['date_key'].nunique()}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

