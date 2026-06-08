# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import re
import sqlite3
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


DEFAULT_BASE_DIR = Path("D:/Study/주식_데일리")
DEFAULT_DB_PATH = Path("D:/Study/stock app/outputs/screening_avg_cache.sqlite3")
SHEET_NAME = "주도주 찾기"


def normalize_date(value: str | None) -> str:
    if not value:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if not re.fullmatch(r"20\d{6}", digits):
        raise ValueError(f"Invalid date: {value}")
    return digits


def parse_date_from_name(name: str) -> str:
    match = re.search(r"(20\d{6})", name)
    return match.group(1) if match else ""


def list_screening_files(base_dir: Path, start_date: str | None, end_date: str | None) -> list[Path]:
    start = normalize_date(start_date) if start_date else ""
    end = normalize_date(end_date) if end_date else ""
    files: list[Path] = []
    for path in sorted(base_dir.glob("*_데일리_기업스크리닝.xls*")):
        date_key = parse_date_from_name(path.name)
        if not date_key:
            continue
        if start and date_key < start:
            continue
        if end and date_key > end:
            continue
        files.append(path)
    return files


def build_avg_table(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        DROP TABLE IF EXISTS avg_scores;
        CREATE TABLE avg_scores (
          file_date TEXT NOT NULL,
          stock_code TEXT NOT NULL,
          avg_1w REAL,
          avg_1m REAL,
          PRIMARY KEY (file_date, stock_code)
        );

        INSERT INTO avg_scores(file_date, stock_code, avg_1w, avg_1m)
        SELECT cur.file_date,
               cur.stock_code,
               (
                 SELECT AVG(prev.score)
                 FROM scores prev
                 WHERE prev.stock_code = cur.stock_code
                   AND prev.file_date < cur.file_date
                   AND prev.file_date >= strftime('%Y%m%d', date(substr(cur.file_date,1,4)||'-'||substr(cur.file_date,5,2)||'-'||substr(cur.file_date,7,2), '-7 day'))
               ) AS avg_1w,
               (
                 SELECT AVG(prev.score)
                 FROM scores prev
                 WHERE prev.stock_code = cur.stock_code
                   AND prev.file_date < cur.file_date
                   AND prev.file_date >= strftime('%Y%m%d', date(substr(cur.file_date,1,4)||'-'||substr(cur.file_date,5,2)||'-'||substr(cur.file_date,7,2), '-30 day'))
               ) AS avg_1m
        FROM scores cur;

        CREATE INDEX IF NOT EXISTS idx_avg_scores_date ON avg_scores(file_date);
        """
    )
    conn.commit()


def write_qr_from_db(conn: sqlite3.Connection, files: Iterable[Path]) -> dict[str, int]:
    changed_files = 0
    changed_cells = 0
    scanned_rows = 0
    for path in files:
        file_date = parse_date_from_name(path.name)
        if not file_date:
            continue
        rows = conn.execute(
            "SELECT stock_code, avg_1m, avg_1w FROM avg_scores WHERE file_date = ?",
            (file_date,),
        ).fetchall()
        if not rows:
            continue
        avg_map = {str(code).strip(): (a1m, a1w) for code, a1m, a1w in rows}

        keep_vba = path.suffix.lower() == ".xlsm"
        wb = load_workbook(path, keep_vba=keep_vba, data_only=False, read_only=False)
        ws = wb[SHEET_NAME]
        file_changed = 0

        ws.cell(1, 17).value = "1M 평균"
        ws.cell(1, 18).value = "1W 평균"

        row = 2
        empty = 0
        while empty < 300:
            code = str(ws.cell(row, 3).value or "").strip()
            if not code:
                empty += 1
                row += 1
                continue
            empty = 0
            scanned_rows += 1
            o_val = ws.cell(row, 15).value
            try:
                o_num = float(o_val)
            except Exception:
                o_num = 0.0
            a1m, a1w = avg_map.get(code, (None, None))
            q_val = round(float(a1m) if a1m is not None else o_num, 2)
            r_val = round(float(a1w) if a1w is not None else o_num, 2)

            q_cell = ws.cell(row, 17)
            r_cell = ws.cell(row, 18)
            if q_cell.value != q_val:
                q_cell.value = q_val
                file_changed += 1
            if r_cell.value != r_val:
                r_cell.value = r_val
                file_changed += 1
            row += 1

        if file_changed:
            wb.save(path)
            changed_files += 1
            changed_cells += file_changed
        wb.close()

    return {"changed_files": changed_files, "changed_cells": changed_cells, "scanned_rows": scanned_rows}


def main() -> None:
    parser = argparse.ArgumentParser(description="Write Q/R (1M/1W averages) from DB scores table.")
    parser.add_argument("--base-dir", default=str(DEFAULT_BASE_DIR))
    parser.add_argument("--db-path", default=str(DEFAULT_DB_PATH))
    parser.add_argument("--start-date", default=None)
    parser.add_argument("--end-date", default=None)
    args = parser.parse_args()

    base_dir = Path(args.base_dir)
    db_path = Path(args.db_path)
    files = list_screening_files(base_dir, args.start_date, args.end_date)
    if not files:
        print('{"status":"no_files"}')
        return

    conn = sqlite3.connect(db_path)
    try:
        build_avg_table(conn)
        result = write_qr_from_db(conn, files)
    finally:
        conn.close()
    print(result)


if __name__ == "__main__":
    main()

