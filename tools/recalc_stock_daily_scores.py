from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


STOCK_DAILY_DIR = Path("D:/Study/Stock_Daily")
FORMULA_CONFIG_PATH = STOCK_DAILY_DIR / "score_formula_config.json"

DEFAULT_FORMULA_CONFIG: dict[str, Any] = {
    "score_formula": {
        "amount_power": 2.0,
        "marcap_power": 0.8,
        "return_base": 1.0,
        "return_power": 4.0,
        "log_base": 1.1,
        "bonus_if_52w_high": 4.0,
        "bonus_if_not_52w_high": -4.0,
        "offset": -13.0,
        "invalid_fill": -100000.0,
    },
    "final_score_formula": {
        "weight_today": 0.35,
        "weight_1w": 0.45,
        "weight_3m": 0.2,
        "sortino_power": 2.0,
        "sortino_floor": 1e-6,
    },
}


def _n(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return default
    s = str(v).strip().replace(",", "")
    if not s:
        return default
    try:
        return float(s)
    except Exception:
        return default


def _load_formula_config() -> dict[str, Any]:
    if not FORMULA_CONFIG_PATH.exists():
        FORMULA_CONFIG_PATH.write_text(json.dumps(DEFAULT_FORMULA_CONFIG, ensure_ascii=False, indent=2), encoding="utf-8")
        return json.loads(json.dumps(DEFAULT_FORMULA_CONFIG))
    try:
        raw = json.loads(FORMULA_CONFIG_PATH.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            raise ValueError("invalid")
    except Exception:
        FORMULA_CONFIG_PATH.write_text(json.dumps(DEFAULT_FORMULA_CONFIG, ensure_ascii=False, indent=2), encoding="utf-8")
        return json.loads(json.dumps(DEFAULT_FORMULA_CONFIG))
    merged = json.loads(json.dumps(DEFAULT_FORMULA_CONFIG))
    for k in ("score_formula", "final_score_formula"):
        if isinstance(raw.get(k), dict):
            merged[k].update(raw[k])
    return merged


def _core_value(amount_100m: float, marcap_100m: float, change_pct: float, score_cfg: dict[str, Any]) -> float:
    if marcap_100m <= 0:
        return float("nan")
    amount_power = float(score_cfg.get("amount_power", 2.0))
    marcap_power = float(score_cfg.get("marcap_power", 0.8))
    return_base = float(score_cfg.get("return_base", 1.1))
    return_power = float(score_cfg.get("return_power", 4.0))
    g = change_pct / 100.0
    try:
        core = ((max(amount_100m, 0.0) ** amount_power) / (marcap_100m ** marcap_power)) * ((return_base + g) ** return_power)
    except Exception:
        return float("nan")
    if not math.isfinite(core) or core <= 0:
        return float("nan")
    return core


def _infer_is_52w_from_old_score(
    old_score: float,
    amount_100m: float,
    marcap_100m: float,
    change_pct: float,
    score_cfg: dict[str, Any],
) -> int:
    # old_score ~= log(core, log_base) + bonus(+4/-4) + offset
    log_base = float(score_cfg.get("log_base", 1.1))
    offset = float(score_cfg.get("offset", -13.0))
    core = _core_value(amount_100m, marcap_100m, change_pct, score_cfg)
    if not math.isfinite(core) or core <= 0:
        return 0
    log_term = math.log(core) / math.log(log_base)
    bonus_est = old_score - log_term - offset
    return 1 if bonus_est >= 0 else 0


def _recalc_one_file(path: Path, cfg: dict[str, Any]) -> tuple[int, int]:
    score_cfg = cfg["score_formula"]
    final_cfg = cfg["final_score_formula"]
    invalid_fill = float(score_cfg.get("invalid_fill", -100000.0))

    wb = load_workbook(path)
    ws = wb.active
    changed = 0
    total = 0

    # 고정 열 순서 (Stock_Daily 표준)
    # 1 순위 / 2 섹터 / 3 종목코드 / 4 종목명 / 5 업종 / 6 시총 / 7 거래대금 / 8 등락률
    # 9 점수 / 10 1W / 11 1M / 12 sortino / 13 종합 / 14 비고
    for r in range(2, ws.max_row + 1):
        code = str(ws.cell(r, 3).value or "").strip()
        if not code:
            continue
        total += 1
        marcap_100m = _n(ws.cell(r, 6).value, 0.0)
        amount_100m = _n(ws.cell(r, 7).value, 0.0)
        change_pct = _n(ws.cell(r, 8).value, 0.0)
        old_score = _n(ws.cell(r, 9).value, invalid_fill)
        avg_1w = _n(ws.cell(r, 10).value, old_score if math.isfinite(old_score) else 0.0)
        avg_1m = _n(ws.cell(r, 11).value, old_score if math.isfinite(old_score) else 0.0)
        sortino = _n(ws.cell(r, 12).value, 0.5)

        core = _core_value(amount_100m, marcap_100m, change_pct, score_cfg)
        if not math.isfinite(core):
            new_score = invalid_fill
        else:
            is_52w = _infer_is_52w_from_old_score(old_score, amount_100m, marcap_100m, change_pct, score_cfg)
            log_base = float(score_cfg.get("log_base", 1.1))
            bonus_if_52w_high = float(score_cfg.get("bonus_if_52w_high", 4.0))
            bonus_if_not_52w_high = float(score_cfg.get("bonus_if_not_52w_high", -4.0))
            offset = float(score_cfg.get("offset", -13.0))
            bonus = bonus_if_52w_high if is_52w == 1 else bonus_if_not_52w_high
            new_score = (math.log(core) / math.log(log_base)) + bonus + offset

        weight_today = float(final_cfg.get("weight_today", 0.35))
        weight_1w = float(final_cfg.get("weight_1w", 0.45))
        weight_3m = float(final_cfg.get("weight_3m", final_cfg.get("weight_1m", 0.2)))
        sortino_power = float(final_cfg.get("sortino_power", 2.0))
        sortino_floor = float(final_cfg.get("sortino_floor", 1e-6))
        sortino_center = 0.6

        composite = (new_score * weight_today) + (avg_1w * weight_1w) + (avg_1m * weight_3m)
        sortino_multiplier = math.exp(sortino_power * (sortino - sortino_center))
        if composite >= 0:
            final_score = composite * sortino_multiplier
        else:
            final_score = composite / sortino_multiplier

        ws.cell(r, 9).value = round(new_score, 2)
        ws.cell(r, 13).value = round(final_score, 2)
        changed += 1

    wb.save(path)
    wb.close()
    return changed, total


def main() -> int:
    cfg = _load_formula_config()
    files = sorted(
        p for p in STOCK_DAILY_DIR.glob("*.xlsx")
        if re.match(r"^\d{8}_", p.name)
    )
    if not files:
        print("[ERROR] 대상 파일 없음")
        return 1

    total_files = len(files)
    total_rows = 0
    total_changed = 0
    print(f"[INFO] files={total_files}")
    for i, path in enumerate(files, start=1):
        changed, rows = _recalc_one_file(path, cfg)
        total_rows += rows
        total_changed += changed
        if i % 50 == 0 or i == total_files:
            print(f"[PROGRESS] {i}/{total_files} files, rows={total_rows}")

    print(f"[DONE] files={total_files}, rows={total_rows}, updated_rows={total_changed}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
