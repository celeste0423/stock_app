from __future__ import annotations

import argparse
import json
import math
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
from pykrx import stock as pykrx_stock


PROJECT_ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = PROJECT_ROOT / "data" / "screening" / "current"
LEGACY_DAILY_DIR = PROJECT_ROOT / "data" / "screening" / "legacy"
SECTOR_DB_PATH = PROJECT_ROOT / "backend" / "sector_database.json"
NOTE_DB_PATH = PROJECT_ROOT / "backend" / "stock_note_database.json"
CACHE_DIR = OUT_DIR / ".cache"
SNAPSHOT_CACHE_DIR = CACHE_DIR / "snapshots"
NAME_CACHE_PATH = CACHE_DIR / "ticker_name_cache.json"
PANEL_CACHE_DIR = CACHE_DIR / "panels"
FORMULA_CONFIG_PATH = OUT_DIR / "score_formula_config.json"


@dataclass
class RunConfig:
    start_date: str
    end_date: str
    min_marcap_100m: float = 2000.0
    lookback_days: int = 320
    skip_existing: bool = True


DEFAULT_FORMULA_CONFIG: dict[str, Any] = {
    "score_formula": {
        "amount_power": 2.0,
        "marcap_power": 0.8,
        "return_base": 1.0,
        "return_power": 4.0,
        "log_base": 1.1,
        "bonus_if_52w_high": 4.0,
        "bonus_if_not_52w_high": -4.0,
        "offset": -13.0,
        "invalid_fill": -100000.0,
    },
    "final_score_formula": {
        "weight_today": 0.35,
        "weight_1w": 0.45,
        "weight_3m": 0.2,
        "sortino_power": 2.0,
        "sortino_floor": 1e-6,
    },
}


def _to_date(s: str) -> datetime:
    return datetime.strptime(re.sub(r"\D", "", s), "%Y%m%d")


def _to_yyyymmdd(d: datetime) -> str:
    return d.strftime("%Y%m%d")


def _normalize_code(v: Any) -> str:
    d = re.sub(r"\D", "", str(v or ""))
    return d.zfill(6) if d else ""


def _ensure_formula_config() -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if not FORMULA_CONFIG_PATH.exists():
        FORMULA_CONFIG_PATH.write_text(
            json.dumps(DEFAULT_FORMULA_CONFIG, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return json.loads(json.dumps(DEFAULT_FORMULA_CONFIG))
    try:
        raw = json.loads(FORMULA_CONFIG_PATH.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            raise ValueError("formula config must be object")
    except Exception:
        FORMULA_CONFIG_PATH.write_text(
            json.dumps(DEFAULT_FORMULA_CONFIG, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return json.loads(json.dumps(DEFAULT_FORMULA_CONFIG))

    merged = json.loads(json.dumps(DEFAULT_FORMULA_CONFIG))
    for section in ("score_formula", "final_score_formula"):
        src = raw.get(section)
        if isinstance(src, dict):
            merged[section].update(src)
    return merged


def _load_sector_map() -> dict[str, str]:
    if not SECTOR_DB_PATH.exists():
        return {}
    try:
        payload = json.loads(SECTOR_DB_PATH.read_text(encoding="utf-8"))
        smap = payload.get("stock_map", {}) if isinstance(payload, dict) else {}
        out: dict[str, str] = {}
        for key, item in smap.items():
            if not isinstance(item, dict):
                continue
            code = _normalize_code(item.get("stock_code") or key)
            sec = str(item.get("sector") or "").strip()
            if code and sec:
                out[code] = sec
        return out
    except Exception:
        return {}


def _load_note_map() -> dict[str, str]:
    note_map: dict[str, str] = {}
    # 1) dedicated note DB (preferred)
    if NOTE_DB_PATH.exists():
        try:
            payload = json.loads(NOTE_DB_PATH.read_text(encoding="utf-8"))
            if isinstance(payload, dict):
                for k, v in payload.items():
                    code = _normalize_code(k)
                    note = str(v or "").strip()
                    if code and note:
                        note_map[code] = note
        except Exception:
            pass

    # 2) sector DB fallback (if note fields exist)
    if SECTOR_DB_PATH.exists():
        try:
            payload = json.loads(SECTOR_DB_PATH.read_text(encoding="utf-8"))
            smap = payload.get("stock_map", {}) if isinstance(payload, dict) else {}
            if isinstance(smap, dict):
                for k, item in smap.items():
                    if not isinstance(item, dict):
                        continue
                    code = _normalize_code(item.get("stock_code") or k)
                    note = str(
                        item.get("note")
                        or item.get("memo")
                        or item.get("remark")
                        or ""
                    ).strip()
                    if code and note and code not in note_map:
                        note_map[code] = note
        except Exception:
            pass
    return note_map


def _load_legacy_note_map_for_day(day: str) -> dict[str, str]:
    note_map: dict[str, str] = {}
    if not LEGACY_DAILY_DIR.exists():
        return note_map
    files = sorted(LEGACY_DAILY_DIR.glob(f"{day}_*"))
    if not files:
        return note_map
    file_path = files[0]
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        # 관례적으로 2번째 시트가 "주도주 찾기"
        ws = wb[wb.sheetnames[1]] if len(wb.sheetnames) > 1 else wb[wb.sheetnames[0]]
        r = 2
        while True:
            code_raw = ws.cell(row=r, column=3).value  # 종목코드
            if code_raw is None:
                break
            code = _normalize_code(code_raw)
            note = str(ws.cell(row=r, column=16).value or "").strip()  # 비고
            if code and note:
                note_map[code] = note
            r += 1
        wb.close()
    except Exception:
        return {}
    return note_map


def _valid_trading_days(start: str, end: str) -> list[str]:
    all_days = pykrx_stock.get_market_ohlcv_by_date(start, end, "005930")
    if all_days is None or all_days.empty:
        return []
    return [d.strftime("%Y%m%d") for d in all_days.index]


def _fetch_snapshot_for_day(date_key: str) -> pd.DataFrame:
    parts: list[pd.DataFrame] = []
    for market in ("KOSPI", "KOSDAQ"):
        ohlcv = pykrx_stock.get_market_ohlcv_by_ticker(date_key, market=market)
        cap = pykrx_stock.get_market_cap_by_ticker(date_key, market=market)
        if ohlcv is None or ohlcv.empty:
            continue
        ohlcv = ohlcv.copy()
        ohlcv.index = ohlcv.index.astype(str).str.zfill(6)
        cap = cap.copy() if cap is not None else pd.DataFrame(index=ohlcv.index)
        cap.index = cap.index.astype(str).str.zfill(6)
        if "시가총액" not in ohlcv.columns and "시가총액" in cap.columns:
            merged = ohlcv.join(cap[["시가총액"]], how="left")
        else:
            merged = ohlcv
        merged["Code"] = merged.index
        merged["market"] = market
        parts.append(merged)
    if not parts:
        return pd.DataFrame()
    df = pd.concat(parts, axis=0, ignore_index=True)
    return df


def _snapshot_cache_path(day: str) -> Path:
    return SNAPSHOT_CACHE_DIR / f"{day}.pkl"


def _load_snapshot_from_cache(day: str) -> pd.DataFrame:
    p = _snapshot_cache_path(day)
    if not p.exists():
        return pd.DataFrame()
    try:
        df = pd.read_pickle(p)
        return df if isinstance(df, pd.DataFrame) else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def _save_snapshot_to_cache(day: str, df: pd.DataFrame) -> None:
    SNAPSHOT_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    p = _snapshot_cache_path(day)
    df.to_pickle(p)


def _load_name_cache() -> dict[str, str]:
    if not NAME_CACHE_PATH.exists():
        return {}
    try:
        obj = json.loads(NAME_CACHE_PATH.read_text(encoding="utf-8"))
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def _save_name_cache(name_map: dict[str, str]) -> None:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    NAME_CACHE_PATH.write_text(json.dumps(name_map, ensure_ascii=False), encoding="utf-8")


def _build_history_panel(days: list[str]) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    close_dict: dict[str, pd.Series] = {}
    change_dict: dict[str, pd.Series] = {}
    amount_dict: dict[str, pd.Series] = {}
    marcap_dict: dict[str, pd.Series] = {}
    for d in days:
        snap = _load_snapshot_from_cache(d)
        loaded_from = "cache"
        if snap.empty:
            snap = _fetch_snapshot_for_day(d)
            loaded_from = "api"
            if not snap.empty:
                _save_snapshot_to_cache(d, snap)
        if snap.empty:
            continue
        snap = snap.set_index("Code")
        close_dict[d] = pd.to_numeric(snap.get("종가"), errors="coerce")
        change_dict[d] = pd.to_numeric(snap.get("등락률"), errors="coerce")
        amount_dict[d] = pd.to_numeric(snap.get("거래대금"), errors="coerce") / 100_000_000.0
        marcap_dict[d] = pd.to_numeric(snap.get("시가총액"), errors="coerce") / 100_000_000.0
        print(f"[LOAD] {d} snapshot loaded ({len(snap)}) [{loaded_from}]")

    close = pd.DataFrame(close_dict).sort_index(axis=1)
    change = pd.DataFrame(change_dict).sort_index(axis=1)
    amount = pd.DataFrame(amount_dict).sort_index(axis=1)
    marcap = pd.DataFrame(marcap_dict).sort_index(axis=1)
    return close, change, amount, marcap


def _panel_cache_paths(first_day: str, last_day: str, lookback_days: int) -> dict[str, Path]:
    PANEL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    key = f"{first_day}_{last_day}_{lookback_days}"
    return {
        "close": PANEL_CACHE_DIR / f"{key}_close.pkl",
        "change": PANEL_CACHE_DIR / f"{key}_change.pkl",
        "amount": PANEL_CACHE_DIR / f"{key}_amount.pkl",
        "marcap": PANEL_CACHE_DIR / f"{key}_marcap.pkl",
    }


def _load_panel_cache(first_day: str, last_day: str, lookback_days: int) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame] | None:
    paths = _panel_cache_paths(first_day, last_day, lookback_days)
    if not all(p.exists() for p in paths.values()):
        return None
    try:
        close = pd.read_pickle(paths["close"])
        change = pd.read_pickle(paths["change"])
        amount = pd.read_pickle(paths["amount"])
        marcap = pd.read_pickle(paths["marcap"])
        return close, change, amount, marcap
    except Exception:
        return None


def _save_panel_cache(
    first_day: str,
    last_day: str,
    lookback_days: int,
    close: pd.DataFrame,
    change: pd.DataFrame,
    amount: pd.DataFrame,
    marcap: pd.DataFrame,
) -> None:
    paths = _panel_cache_paths(first_day, last_day, lookback_days)
    close.to_pickle(paths["close"])
    change.to_pickle(paths["change"])
    amount.to_pickle(paths["amount"])
    marcap.to_pickle(paths["marcap"])


def _score_formula(
    m_100m: pd.Series,
    n_100m: pd.Series,
    g_pct: pd.Series,
    is_52w: pd.Series,
    cfg_formula: dict[str, Any],
) -> pd.Series:
    m = pd.to_numeric(m_100m, errors="coerce")
    n = pd.to_numeric(n_100m, errors="coerce")
    g = pd.to_numeric(g_pct, errors="coerce") / 100.0

    amount_power = float(cfg_formula.get("amount_power", 2.0))
    marcap_power = float(cfg_formula.get("marcap_power", 0.8))
    return_base = float(cfg_formula.get("return_base", 1.1))
    return_power = float(cfg_formula.get("return_power", 4.0))
    log_base = float(cfg_formula.get("log_base", 1.1))
    bonus_if_52w_high = float(cfg_formula.get("bonus_if_52w_high", 4.0))
    bonus_if_not_52w_high = float(cfg_formula.get("bonus_if_not_52w_high", -4.0))
    offset = float(cfg_formula.get("offset", -13.0))
    invalid_fill = float(cfg_formula.get("invalid_fill", -100000.0))

    core = (np.power(m, amount_power) / np.power(n, marcap_power)) * np.power(return_base + g, return_power)
    core = core.where((core > 0) & np.isfinite(core))
    score = np.log(core) / math.log(log_base)
    bonus = np.where(is_52w.fillna(0).astype(int) == 1, bonus_if_52w_high, bonus_if_not_52w_high)
    score = score + bonus + offset
    return score.fillna(invalid_fill)


def _build_for_day(
    target_day: str,
    panel_days: list[str],
    close_df: pd.DataFrame,
    change_df: pd.DataFrame,
    amount_df: pd.DataFrame,
    marcap_df: pd.DataFrame,
    sector_map: dict[str, str],
    min_marcap_100m: float,
    name_cache: dict[str, str],
    note_map: dict[str, str],
    formula_cfg: dict[str, Any],
) -> pd.DataFrame:
    day_idx = panel_days.index(target_day)
    target_col = target_day
    target_marcap = marcap_df.get(target_col, pd.Series(dtype=float))

    universe = target_marcap[target_marcap >= min_marcap_100m].dropna().index
    if len(universe) == 0:
        return pd.DataFrame()

    close_u = close_df.loc[universe]
    change_u = change_df.loc[universe]
    amount_u = amount_df.loc[universe]
    marcap_u = marcap_df.loc[universe]

    close_window = close_u.iloc[:, max(0, day_idx - 251) : day_idx + 1]
    max_52w = close_window.max(axis=1)
    is_52w = (close_u[target_col] >= max_52w).astype(int)

    score_today = _score_formula(
        amount_u[target_col],
        marcap_u[target_col],
        change_u[target_col],
        is_52w,
        formula_cfg.get("score_formula", {}),
    )

    score_hist: list[pd.Series] = []
    hist_start = max(0, day_idx - 30)
    for i in range(hist_start, day_idx):
        d = panel_days[i]
        win_i = close_u.iloc[:, max(0, i - 251) : i + 1]
        is52_i = (close_u[d] >= win_i.max(axis=1)).astype(int)
        s = _score_formula(
            amount_u[d],
            marcap_u[d],
            change_u[d],
            is52_i,
            formula_cfg.get("score_formula", {}),
        )
        score_hist.append(s.rename(d))

    if score_hist:
        score_hist_df = pd.concat(score_hist, axis=1)
        avg_1w = score_hist_df.iloc[:, -7:].mean(axis=1)
        avg_3m = score_hist_df.mean(axis=1)
    else:
        avg_1w = score_today.copy()
        avg_3m = score_today.copy()

    ret60 = close_u.pct_change(axis=1).iloc[:, max(0, day_idx - 59) : day_idx + 1]
    downside = ret60.clip(upper=0)
    dd = np.sqrt((downside**2).mean(axis=1)).replace(0, 1e-8)
    ratio = ret60.mean(axis=1) / dd
    sortino = 1.0 / (1.0 + np.exp(-ratio))
    sortino = sortino.fillna(0.5)

    final_cfg = formula_cfg.get("final_score_formula", {})
    weight_today = float(final_cfg.get("weight_today", 0.35))
    weight_1w = float(final_cfg.get("weight_1w", 0.45))
    weight_3m = float(final_cfg.get("weight_3m", final_cfg.get("weight_1m", 0.2)))
    sortino_power = float(final_cfg.get("sortino_power", 2.0))
    sortino_floor = float(final_cfg.get("sortino_floor", 1e-6))
    sortino_center = 0.6
    composite = (score_today * weight_today) + (avg_1w * weight_1w) + (avg_3m * weight_3m)
    sortino_multiplier = np.exp(sortino_power * (sortino - sortino_center))
    final_score = composite * sortino_multiplier if composite >= 0 else composite / sortino_multiplier

    def _name_of(code: str) -> str:
        n = name_cache.get(code)
        if n:
            return n
        try:
            n = pykrx_stock.get_market_ticker_name(code) or code
        except Exception:
            n = code
        name_cache[code] = n
        return n

    records = pd.DataFrame(
        {
            "종목코드": universe,
            "섹터": [sector_map.get(c, "") for c in universe],
            "종목 이름": [_name_of(c) for c in universe],
            "업종": "",
            "시총 (억원)": marcap_u[target_col].round(0).astype("Int64"),
            "거래대금 (억원)": amount_u[target_col].round(0).astype("Int64"),
            "등락률": change_u[target_col].round(2),
            "점수": score_today.round(2),
            "1W 평균 점수": avg_1w.round(2),
            "1M 평균 점수": avg_1m.round(2),
            "60일 기준 Sortino 정규화 점수": sortino.round(4),
            "종합 점수": final_score.round(2),
            "비고": [note_map.get(c, "") for c in universe],
        }
    )
    records = records.sort_values(["종합 점수", "점수"], ascending=False).reset_index(drop=True)
    records.insert(0, "순위", np.arange(1, len(records) + 1))
    return records


def run(cfg: RunConfig) -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    SNAPSHOT_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    PANEL_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    sector_map = _load_sector_map()
    note_map = _load_note_map()
    name_cache = _load_name_cache()
    formula_cfg = _ensure_formula_config()
    print(f"[INFO] formula config loaded: {FORMULA_CONFIG_PATH}")

    target_days = _valid_trading_days(cfg.start_date, cfg.end_date)
    if not target_days:
        print("[ERROR] 타겟 거래일 조회 실패")
        return 1

    # 대상 구간 전체를 안전하게 포함해 히스토리 패널을 구성한다.
    hist_days = _valid_trading_days(cfg.start_date, target_days[-1])
    if not hist_days:
        print("[ERROR] 히스토리 거래일 조회 실패")
        return 1
    panel_days = hist_days[-cfg.lookback_days :]
    if target_days and target_days[0] not in panel_days:
        panel_days = hist_days

    print(f"[INFO] panel days: {panel_days[0]} ~ {panel_days[-1]} ({len(panel_days)})")
    panel_cached = _load_panel_cache(panel_days[0], panel_days[-1], cfg.lookback_days)
    if panel_cached is not None:
        close_df, change_df, amount_df, marcap_df = panel_cached
        print("[INFO] history panel loaded from panel cache")
    else:
        close_df, change_df, amount_df, marcap_df = _build_history_panel(panel_days)
        _save_panel_cache(panel_days[0], panel_days[-1], cfg.lookback_days, close_df, change_df, amount_df, marcap_df)
        print("[INFO] history panel built and cached")

    for td in target_days:
        out_path = OUT_DIR / f"{td}_데일리_기업스크리닝.xlsx"
        if cfg.skip_existing and out_path.exists():
            print(f"[SKIP] {td} already exists")
            continue
        out_df = _build_for_day(
            td,
            panel_days,
            close_df,
            change_df,
            amount_df,
            marcap_df,
            sector_map,
            cfg.min_marcap_100m,
            name_cache,
            note_map,
            formula_cfg,
        )
        if out_df.empty:
            print(f"[SKIP] {td} empty")
            continue
        out_df.to_excel(out_path, index=False)
        # 등락률 시각 강조: +빨강 / -파랑
        wb = load_workbook(out_path)
        ws = wb.active
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        try:
            change_col = header.index("등락률") + 1
            score_col = header.index("점수") + 1
            red_font = Font(color="00C00000")
            blue_font = Font(color="000000CC")
            default_font = Font(color="00000000")
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=change_col)
                try:
                    v = float(cell.value)
                except Exception:
                    v = 0.0
                if v > 0:
                    cell.font = red_font
                elif v < 0:
                    cell.font = blue_font
                else:
                    cell.font = default_font

            # 당일 점수 조건부 색상: 낮음(초록) - 중간(연노랑) - 높음(빨강)
            if ws.max_row >= 2:
                score_col_letter = ws.cell(row=1, column=score_col).column_letter
                score_range = f"{score_col_letter}2:{score_col_letter}{ws.max_row}"
                ws.conditional_formatting.add(
                    score_range,
                    ColorScaleRule(
                        start_type="min",
                        start_color="63BE7B",
                        mid_type="percentile",
                        mid_value=50,
                        mid_color="FFEB84",
                        end_type="max",
                        end_color="F8696B",
                    ),
                )
        except ValueError:
            pass
        wb.save(out_path)
        wb.close()
        print(f"[DONE] {out_path} rows={len(out_df)}")
    _save_name_cache(name_cache)
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--start", required=True, help="YYYYMMDD")
    parser.add_argument("--end", required=True, help="YYYYMMDD")
    parser.add_argument("--no-skip-existing", action="store_true", help="기존 파일도 덮어쓰기")
    args = parser.parse_args()
    cfg = RunConfig(
        start_date=re.sub(r"\D", "", args.start),
        end_date=re.sub(r"\D", "", args.end),
        skip_existing=not bool(args.no_skip_existing),
    )
    return run(cfg)


if __name__ == "__main__":
    raise SystemExit(main())
