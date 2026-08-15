# -*- coding: utf-8 -*-
from __future__ import annotations

import math
import re
import sqlite3
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCREENING_DIR = PROJECT_ROOT / "data" / "screening" / "legacy"
FILE_STEM = "\ub370\uc77c\ub9ac_\uae30\uc5c5\uc2a4\ud06c\ub9ac\ub2dd"
SCREENING_SHEET = "\uc8fc\ub3c4\uc8fc \ucc3e\uae30"
DB_PATH = SCREENING_DIR / "screening_score_history.sqlite"
FILE_PATTERN = re.compile(rf"^(20\d{{6}})_{re.escape(FILE_STEM)}\.xlsm$", re.I)


def normalize_code(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.zfill(6) if digits else ""


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        return float(value)
    text = str(value).strip()
    if not text or text in {"-", "nan", "None"}:
        return None
    cleaned = re.sub(r"[^0-9.+\-]", "", text.replace(",", ""))
    if cleaned in {"", "+", "-", "."}:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def list_screening_files() -> list[tuple[str, Path]]:
    rows: list[tuple[str, Path]] = []
    for path in SCREENING_DIR.glob(f"*_{FILE_STEM}.xlsm"):
        match = FILE_PATTERN.match(path.name)
        if match:
            rows.append((match.group(1), path))
    return sorted(rows, key=lambda item: item[0])


def header_map(sheet: Any) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[1]:
        text = str(cell.value or "").strip()
        if text and text not in headers:
            headers[text] = cell.column
    return headers


def reset_db(connection: sqlite3.Connection) -> None:
    connection.executescript(
        """
        DROP TABLE IF EXISTS screening_scores;
        DROP TABLE IF EXISTS screening_files;
        CREATE TABLE screening_files (
            file_date TEXT PRIMARY KEY,
            file_name TEXT NOT NULL,
            row_count INTEGER NOT NULL,
            score_count INTEGER NOT NULL
        );
        CREATE TABLE screening_scores (
            file_date TEXT NOT NULL,
            stock_code TEXT NOT NULL,
            score REAL NOT NULL,
            PRIMARY KEY (file_date, stock_code)
        );
        CREATE INDEX idx_screening_scores_code_date ON screening_scores(stock_code, file_date);
        """
    )
    connection.commit()


def load_scores_to_db(connection: sqlite3.Connection, files: list[tuple[str, Path]]) -> dict[str, dict[str, float]]:
    scores_by_date: dict[str, dict[str, float]] = {}
    for file_date, path in files:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        row_count = 0
        score_count = 0
        scores: dict[str, float] = {}
        try:
            if SCREENING_SHEET not in workbook.sheetnames:
                print(f"skip no sheet: {path.name}", flush=True)
                continue
            sheet = workbook[SCREENING_SHEET]
            headers = header_map(sheet)
            code_col = headers.get("\uc885\ubaa9\ucf54\ub4dc") or 3
            score_col = headers.get("\uc810\uc218") or 15
            for row_index in range(2, sheet.max_row + 1):
                code = normalize_code(sheet.cell(row=row_index, column=code_col).value)
                if not code:
                    continue
                row_count += 1
                score = to_float(sheet.cell(row=row_index, column=score_col).value)
                if score is None or score == -100000:
                    continue
                scores[code] = score
                score_count += 1
        finally:
            workbook.close()
        scores_by_date[file_date] = scores
        connection.execute(
            "INSERT INTO screening_files(file_date, file_name, row_count, score_count) VALUES (?, ?, ?, ?)",
            (file_date, path.name, row_count, score_count),
        )
        connection.executemany(
            "INSERT OR REPLACE INTO screening_scores(file_date, stock_code, score) VALUES (?, ?, ?)",
            [(file_date, code, score) for code, score in scores.items()],
        )
        connection.commit()
        print(f"loaded {file_date}: rows={row_count} scores={score_count}", flush=True)
    return scores_by_date


def average_score(
    scores_by_date: dict[str, dict[str, float]],
    file_dates: list[str],
    current_date: str,
    stock_code: str,
    lookback_days: int,
) -> float:
    end = datetime.strptime(current_date, "%Y%m%d").date()
    start = end - timedelta(days=lookback_days)
    values: list[float] = []
    for file_date in file_dates:
        row_date = datetime.strptime(file_date, "%Y%m%d").date()
        if start <= row_date <= end:
            score = scores_by_date.get(file_date, {}).get(stock_code)
            if score is not None and math.isfinite(score) and score != -100000:
                values.append(float(score))
    return round(sum(values) / len(values), 2) if values else 0.0


def update_workbook_averages(
    path: Path,
    file_date: str,
    file_dates: list[str],
    scores_by_date: dict[str, dict[str, float]],
) -> dict[str, Any]:
    workbook = load_workbook(path, keep_vba=True, keep_links=False)
    updated = 0
    try:
        if SCREENING_SHEET not in workbook.sheetnames:
            return {"file": path.name, "updated": 0, "error": f"'{SCREENING_SHEET}' 시트 없음"}
        sheet = workbook[SCREENING_SHEET]
        headers = header_map(sheet)
        code_col = headers.get("\uc885\ubaa9\ucf54\ub4dc") or 3
        q_col = 17
        r_col = 18
        sheet.cell(row=1, column=q_col).value = "1M \ud3c9\uade0"
        sheet.cell(row=1, column=r_col).value = "1W \ud3c9\uade0"
        for row_index in range(2, sheet.max_row + 1):
            code = normalize_code(sheet.cell(row=row_index, column=code_col).value)
            if not code:
                continue
            sheet.cell(row=row_index, column=q_col).value = average_score(scores_by_date, file_dates, file_date, code, 30)
            sheet.cell(row=row_index, column=r_col).value = average_score(scores_by_date, file_dates, file_date, code, 7)
            updated += 1
        workbook.save(path)
        return {"file": path.name, "updated": updated, "error": ""}
    finally:
        workbook.close()


def main() -> None:
    files = list_screening_files()
    if not files:
        raise FileNotFoundError(f"screening files not found: {SCREENING_DIR}")
    print(f"files={len(files)}", flush=True)
    print(f"db={DB_PATH}", flush=True)
    with sqlite3.connect(DB_PATH) as connection:
        reset_db(connection)
        scores_by_date = load_scores_to_db(connection, files)
    file_dates = [file_date for file_date, _path in files]
    results: list[dict[str, Any]] = []
    for file_date, path in files:
        try:
            result = update_workbook_averages(path, file_date, file_dates, scores_by_date)
        except PermissionError as exc:
            result = {"file": path.name, "updated": 0, "error": f"locked: {exc}"}
        except Exception as exc:
            result = {"file": path.name, "updated": 0, "error": str(exc)}
        results.append(result)
        print(f"updated {result['file']}: rows={result['updated']} error={result['error']}", flush=True)
    failed = [item for item in results if item.get("error")]
    print(f"done updated_files={len(results) - len(failed)} failed={len(failed)}", flush=True)
    if failed:
        for item in failed:
            print(f"FAILED {item['file']}: {item['error']}", flush=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
