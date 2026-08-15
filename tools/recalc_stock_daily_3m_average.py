from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
STOCK_DAILY_DIR = PROJECT_ROOT / "data" / "screening" / "current"
FORMULA_CONFIG_PATH = STOCK_DAILY_DIR / "score_formula_config.json"
WINDOW_3M = 60

DEFAULT_FORMULA_CONFIG: dict[str, Any] = {
    "final_score_formula": {
        "weight_today": 0.35,
        "weight_1w": 0.45,
        "weight_3m": 0.2,
        "sortino_power": 0.8,
        "sortino_floor": 1e-6,
    },
}


def _n(v: Any, default: float = 0.0) -> float:
    if v is None:
        return default
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return default
    s = str(v).strip().replace(",", "")
    if not s:
        return default
    try:
        return float(s)
    except Exception:
        return default


def _load_formula_config() -> dict[str, Any]:
    if not FORMULA_CONFIG_PATH.exists():
        FORMULA_CONFIG_PATH.write_text(json.dumps(DEFAULT_FORMULA_CONFIG, ensure_ascii=False, indent=2), encoding="utf-8")
        return json.loads(json.dumps(DEFAULT_FORMULA_CONFIG))
    try:
        raw = json.loads(FORMULA_CONFIG_PATH.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            raise ValueError("invalid")
    except Exception:
        raw = {}
    merged = json.loads(json.dumps(DEFAULT_FORMULA_CONFIG))
    if isinstance(raw.get("final_score_formula"), dict):
        merged["final_score_formula"].update(raw["final_score_formula"])
    return merged


def _normalize_code(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.zfill(6) if digits else ""


def list_target_files() -> list[Path]:
    files = sorted(
        p for p in STOCK_DAILY_DIR.glob("*.xlsx")
        if re.match(r"^\d{8}_", p.name)
    )
    return files


def build_score_maps(files: list[Path]) -> tuple[list[dict[str, float]], list[str]]:
    score_maps: list[dict[str, float]] = []
    all_codes: set[str] = set()
    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        score_map: dict[str, float] = {}
        for r in range(2, ws.max_row + 1):
            code = _normalize_code(ws.cell(r, 3).value)
            if not code:
                continue
            score = _n(ws.cell(r, 9).value, -100000.0)
            if not math.isfinite(score):
                continue
            score_map[code] = float(score)
            all_codes.add(code)
        wb.close()
        score_maps.append(score_map)
    return score_maps, sorted(all_codes)


def build_rolling_3m(score_maps: list[dict[str, float]], all_codes: list[str]) -> dict[str, np.ndarray]:
    n = len(score_maps)
    rolling: dict[str, np.ndarray] = {}
    for code in all_codes:
        arr = np.full(n, np.nan, dtype=float)
        for idx, sm in enumerate(score_maps):
            if code in sm:
                arr[idx] = sm[code]
        series = pd.Series(arr)
        rolled = series.rolling(WINDOW_3M, min_periods=1).mean().to_numpy()
        rolling[code] = rolled
    return rolling


def recalc_file(
    path: Path,
    file_index: int,
    rolling_3m: dict[str, np.ndarray],
    final_cfg: dict[str, Any],
) -> int:
    weight_today = float(final_cfg.get("weight_today", 0.35))
    weight_1w = float(final_cfg.get("weight_1w", 0.45))
    weight_3m = float(final_cfg.get("weight_3m", final_cfg.get("weight_1m", 0.2)))
    sortino_power = float(final_cfg.get("sortino_power", 0.8))
    sortino_floor = float(final_cfg.get("sortino_floor", 1e-6))

    wb = load_workbook(path)
    ws = wb.active
    ws.cell(1, 11).value = "3M 평균 점수"
    changed = 0

    for r in range(2, ws.max_row + 1):
        code = _normalize_code(ws.cell(r, 3).value)
        if not code:
            continue
        score_today = _n(ws.cell(r, 9).value, -100000.0)
        avg_1w = _n(ws.cell(r, 10).value, score_today)
        sortino = _n(ws.cell(r, 12).value, 0.5)

        avg_3m = score_today
        series = rolling_3m.get(code)
        if series is not None and file_index < len(series):
            v = series[file_index]
            if v is not None and np.isfinite(v):
                avg_3m = float(v)

        composite = (score_today * weight_today) + (avg_1w * weight_1w) + (avg_3m * weight_3m)
        if composite >= 0:
            final_score = composite * (max(sortino, sortino_floor) ** sortino_power)
        else:
            final_score = composite * (max(2.0 - sortino, sortino_floor) ** sortino_power)

        ws.cell(r, 11).value = round(avg_3m, 2)
        ws.cell(r, 13).value = round(final_score, 2)
        changed += 1

    wb.save(path)
    wb.close()
    return changed


def main() -> int:
    cfg = _load_formula_config()
    final_cfg = cfg.get("final_score_formula", {})
    files = list_target_files()
    if not files:
        print("[ERROR] target files not found.")
        return 1

    print(f"[INFO] files={len(files)}")
    score_maps, all_codes = build_score_maps(files)
    print(f"[INFO] unique_codes={len(all_codes)}")
    rolling_3m = build_rolling_3m(score_maps, all_codes)

    total_rows = 0
    for idx, path in enumerate(files):
        changed = recalc_file(path, idx, rolling_3m, final_cfg)
        total_rows += changed
        if (idx + 1) % 50 == 0 or (idx + 1) == len(files):
            print(f"[PROGRESS] {idx + 1}/{len(files)} files, updated_rows={total_rows}")

    print(f"[DONE] files={len(files)}, updated_rows={total_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
