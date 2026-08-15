# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import re
import subprocess
import tempfile
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SCREENING_DIR = PROJECT_ROOT / "data" / "screening" / "legacy"
FILE_STEM = "데일리_기업스크리닝"
SCREENING_SHEET = "주도주 찾기"
HIGH52_SHEET = "52주신고가"


@dataclass(frozen=True)
class EditResult:
    path: Path
    changed_cells: int
    eligible_rows: int


def normalize_date(value: str | None) -> str:
    if not value:
        return ""
    digits = re.sub(r"\D", "", str(value))
    if not re.fullmatch(r"20\d{6}", digits):
        raise ValueError(f"date must be YYYYMMDD: {value}")
    return digits


def screening_file_date(path: Path) -> str:
    match = re.search(r"(20\d{6})", path.name)
    return match.group(1) if match else ""


def file_date_as_datetime(path: Path) -> datetime:
    return datetime.strptime(screening_file_date(path), "%Y%m%d")


def workbook_has_sheets(path: Path, sheet_names: Iterable[str]) -> bool:
    try:
        workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    except Exception:
        return False
    try:
        existing = set(workbook.sheetnames)
        return all(name in existing for name in sheet_names)
    finally:
        workbook.close()


def iter_screening_files(
    base_dir: Path = SCREENING_DIR,
    *,
    start_date: str | None = None,
    end_date: str | None = None,
    require_high52_sheet: bool = True,
) -> list[Path]:
    start = normalize_date(start_date) if start_date else ""
    end = normalize_date(end_date) if end_date else ""
    paths: list[Path] = []
    for path in sorted(base_dir.glob(f"*_{FILE_STEM}.xls*")):
        date_key = screening_file_date(path)
        if not date_key:
            continue
        if start and date_key < start:
            continue
        if end and date_key > end:
            continue
        if require_high52_sheet and not workbook_has_sheets(path, [SCREENING_SHEET, HIGH52_SHEET]):
            continue
        paths.append(path)
    return paths


def row_has_stock_identity(sheet: Any, row_index: int) -> bool:
    code = str(sheet.cell(row=row_index, column=3).value or "").strip()
    name = str(sheet.cell(row=row_index, column=4).value or "").strip()
    return bool(code and name)


def iter_stock_rows(sheet: Any, *, start_row: int = 2, max_empty_streak: int = 300) -> Iterable[tuple[int, str, str]]:
    empty_streak = 0
    row_index = start_row
    while True:
        code = str(sheet.cell(row=row_index, column=3).value or "").strip()
        name = str(sheet.cell(row=row_index, column=4).value or "").strip()
        if code and name:
            empty_streak = 0
            yield row_index, code, name
        else:
            empty_streak += 1
            if empty_streak >= max_empty_streak:
                break
        row_index += 1


def to_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def set_workbook_recalc_on_open(workbook: Any) -> None:
    if not hasattr(workbook, "calculation"):
        return
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"


def update_screening_column(
    path: Path,
    *,
    column: int | str,
    value_factory: Callable[[int, Any], Any],
    sheet_name: str = SCREENING_SHEET,
    start_row: int = 2,
    row_filter: Callable[[Any, int], bool] = row_has_stock_identity,
) -> EditResult:
    keep_vba = path.suffix.lower() == ".xlsm"
    workbook = load_workbook(path, read_only=False, data_only=False, keep_vba=keep_vba, keep_links=False)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f"sheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
        col_index = column if isinstance(column, int) else sheet[column + "1"].column
        changed = 0
        eligible = 0
        for row_index in range(start_row, sheet.max_row + 1):
            if row_filter and not row_filter(sheet, row_index):
                continue
            eligible += 1
            next_value = value_factory(row_index, sheet)
            cell = sheet.cell(row=row_index, column=col_index)
            if cell.value != next_value:
                cell.value = next_value
                changed += 1
        if changed:
            set_workbook_recalc_on_open(workbook)
            workbook.save(path)
        return EditResult(path=path, changed_cells=changed, eligible_rows=eligible)
    finally:
        workbook.close()


def update_many_screening_column(
    paths: Iterable[Path],
    *,
    column: int | str,
    value_factory: Callable[[int, Any], Any],
    sheet_name: str = SCREENING_SHEET,
) -> list[EditResult]:
    results: list[EditResult] = []
    for path in paths:
        results.append(
            update_screening_column(
                path,
                column=column,
                value_factory=value_factory,
                sheet_name=sheet_name,
            )
        )
    return results


def score_o_formula_without_strength(row_index: int, high52_bonus: int = 4) -> str:
    return (
        f"=IFERROR(LOG(M{row_index}*M{row_index}*100000*100000/(N{row_index})^0.8/10000000000 * "
        f"(1.1+G{row_index})^4,1.1)+ IF(COUNTIF('52주신고가'!C:C, D{row_index}) > 0, "
        f"{-abs(int(high52_bonus))}, {abs(int(high52_bonus))})-13,-100000)"
    )


def _collect_scores_from_workbook(path: Path) -> dict[str, float]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if SCREENING_SHEET not in workbook.sheetnames:
            return {}
        sheet = workbook[SCREENING_SHEET]
        score_map: dict[str, float] = {}
        for row_index, stock_code, _ in iter_stock_rows(sheet):
            score = to_number(sheet.cell(row=row_index, column=15).value)
            if score is None or score == -100000:
                continue
            score_map[stock_code] = score
        return score_map
    finally:
        workbook.close()


def apply_avg_qr_from_history(
    *,
    base_dir: Path = SCREENING_DIR,
    start_date: str | None = None,
    end_date: str | None = None,
    require_high52_sheet: bool = True,
) -> dict[str, Any]:
    paths = iter_screening_files(
        base_dir,
        start_date=start_date,
        end_date=end_date,
        require_high52_sheet=require_high52_sheet,
    )
    if not paths:
        return {"files": 0, "touched_files": 0, "changed_cells": 0, "eligible_rows": 0}

    dated_paths = [(path, file_date_as_datetime(path)) for path in paths]
    scores_by_path = {path: _collect_scores_from_workbook(path) for path, _ in dated_paths}

    changed_cells_total = 0
    eligible_rows_total = 0
    touched_files = 0

    for target_path, target_date in dated_paths:
        monthly_sum: dict[str, float] = {}
        monthly_count: dict[str, int] = {}
        weekly_sum: dict[str, float] = {}
        weekly_count: dict[str, int] = {}

        for source_path, source_date in dated_paths:
            if source_date >= target_date:
                continue
            day_diff = (target_date - source_date).days
            if day_diff > 30:
                continue
            for code, score in scores_by_path.get(source_path, {}).items():
                monthly_sum[code] = monthly_sum.get(code, 0.0) + score
                monthly_count[code] = monthly_count.get(code, 0) + 1
                if day_diff <= 7:
                    weekly_sum[code] = weekly_sum.get(code, 0.0) + score
                    weekly_count[code] = weekly_count.get(code, 0) + 1

        keep_vba = target_path.suffix.lower() == ".xlsm"
        workbook = load_workbook(target_path, read_only=False, data_only=False, keep_vba=keep_vba, keep_links=False)
        try:
            sheet = workbook[SCREENING_SHEET]
            sheet.cell(row=1, column=17).value = "1M 평균"
            sheet.cell(row=1, column=18).value = "1W 평균"
            file_changed = 0
            for row_index, stock_code, _ in iter_stock_rows(sheet):
                eligible_rows_total += 1
                today_score = to_number(sheet.cell(row=row_index, column=15).value)
                if today_score is None or today_score == -100000:
                    today_score = 0.0

                if monthly_count.get(stock_code, 0):
                    avg_1m = round(monthly_sum[stock_code] / monthly_count[stock_code], 2)
                else:
                    avg_1m = round(today_score, 2)

                if weekly_count.get(stock_code, 0):
                    avg_1w = round(weekly_sum[stock_code] / weekly_count[stock_code], 2)
                else:
                    avg_1w = round(today_score, 2)

                q_cell = sheet.cell(row=row_index, column=17)
                r_cell = sheet.cell(row=row_index, column=18)
                if q_cell.value != avg_1m:
                    q_cell.value = avg_1m
                    file_changed += 1
                if r_cell.value != avg_1w:
                    r_cell.value = avg_1w
                    file_changed += 1

            if file_changed:
                set_workbook_recalc_on_open(workbook)
                workbook.save(target_path)
                touched_files += 1
                changed_cells_total += file_changed
        finally:
            workbook.close()

    return {
        "files": len(paths),
        "touched_files": touched_files,
        "changed_cells": changed_cells_total,
        "eligible_rows": eligible_rows_total,
    }


def apply_avg_qr_for_target_date(
    *,
    target_date: str,
    base_dir: Path = SCREENING_DIR,
    require_high52_sheet: bool = True,
) -> dict[str, Any]:
    target = normalize_date(target_date)
    paths = iter_screening_files(
        base_dir,
        start_date=None,
        end_date=target,
        require_high52_sheet=require_high52_sheet,
    )
    target_path = next((p for p in paths if screening_file_date(p) == target), None)
    if target_path is None:
        return {"target_date": target, "changed_cells": 0, "eligible_rows": 0, "updated": False, "reason": "target_not_found"}

    target_dt = file_date_as_datetime(target_path)
    score_maps = {p: _collect_scores_from_workbook(p) for p in paths}
    monthly_sum: dict[str, float] = {}
    monthly_count: dict[str, int] = {}
    weekly_sum: dict[str, float] = {}
    weekly_count: dict[str, int] = {}

    for source_path in paths:
        source_dt = file_date_as_datetime(source_path)
        if source_dt >= target_dt:
            continue
        day_diff = (target_dt - source_dt).days
        if day_diff > 30:
            continue
        for code, score in score_maps.get(source_path, {}).items():
            monthly_sum[code] = monthly_sum.get(code, 0.0) + score
            monthly_count[code] = monthly_count.get(code, 0) + 1
            if day_diff <= 7:
                weekly_sum[code] = weekly_sum.get(code, 0.0) + score
                weekly_count[code] = weekly_count.get(code, 0) + 1

    keep_vba = target_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(target_path, read_only=False, data_only=False, keep_vba=keep_vba, keep_links=False)
    try:
        sheet = workbook[SCREENING_SHEET]
        sheet.cell(row=1, column=17).value = "1M 평균"
        sheet.cell(row=1, column=18).value = "1W 평균"
        changed = 0
        eligible = 0
        for row_index, code, _ in iter_stock_rows(sheet):
            eligible += 1
            today_score = to_number(sheet.cell(row=row_index, column=15).value)
            if today_score is None or today_score == -100000:
                today_score = 0.0

            avg_1m = round(monthly_sum[code] / monthly_count[code], 2) if monthly_count.get(code, 0) else round(today_score, 2)
            avg_1w = round(weekly_sum[code] / weekly_count[code], 2) if weekly_count.get(code, 0) else round(today_score, 2)

            q_cell = sheet.cell(row=row_index, column=17)
            r_cell = sheet.cell(row=row_index, column=18)
            if q_cell.value != avg_1m:
                q_cell.value = avg_1m
                changed += 1
            if r_cell.value != avg_1w:
                r_cell.value = avg_1w
                changed += 1

        if changed:
            set_workbook_recalc_on_open(workbook)
            workbook.save(target_path)
        return {"target_date": target, "changed_cells": changed, "eligible_rows": eligible, "updated": changed > 0, "path": str(target_path)}
    finally:
        workbook.close()


def recalculate_with_excel(paths: Iterable[Path]) -> tuple[list[Path], list[tuple[Path, str]]]:
    path_list = [str(path.resolve()) for path in paths]
    if not path_list:
        return [], []
    script = r'''
param([string]$PayloadPath)
$ErrorActionPreference = "Continue"
$paths = Get-Content -LiteralPath $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json
$ok = @()
$failed = @()
foreach ($target in $paths) {
  $excel = $null
  $workbook = $null
  try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $excel.AskToUpdateLinks = $false
    $workbook = $excel.Workbooks.Open([string]$target, 0, $false)
    $excel.CalculateFullRebuild()
    $workbook.Save()
    $ok += [string]$target
  } catch {
    $failed += [PSCustomObject]@{ path = [string]$target; error = $_.Exception.Message }
  } finally {
    if ($null -ne $workbook) {
      try { $workbook.Close($true) | Out-Null } catch {}
      try { [Runtime.InteropServices.Marshal]::ReleaseComObject($workbook) | Out-Null } catch {}
    }
    if ($null -ne $excel) {
      try { $excel.Quit() | Out-Null } catch {}
      try { [Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null } catch {}
    }
    [GC]::Collect()
    [GC]::WaitForPendingFinalizers()
  }
}
[PSCustomObject]@{ ok = $ok; failed = $failed } | ConvertTo-Json -Depth 4 -Compress
'''
    with tempfile.TemporaryDirectory(prefix="screening-excel-recalc-") as temp_dir:
        temp_path = Path(temp_dir)
        payload_path = temp_path / "paths.json"
        script_path = temp_path / "recalc.ps1"
        payload_path.write_text(json.dumps(path_list, ensure_ascii=False), encoding="utf-8")
        script_path.write_text(script, encoding="utf-8-sig")
        result = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                str(payload_path),
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=max(120, len(path_list) * 45),
        )
    if result.returncode != 0 and not result.stdout.strip():
        raise RuntimeError(result.stderr.strip() or "Excel recalculation failed")
    payload = json.loads(result.stdout.strip().splitlines()[-1])
    ok_paths = [Path(item) for item in payload.get("ok", [])]
    failed = [
        (Path(item.get("path", "")), str(item.get("error", "")))
        for item in payload.get("failed", [])
        if isinstance(item, dict)
    ]
    return ok_paths, failed


def apply_o_score_formula(
    *,
    base_dir: Path = SCREENING_DIR,
    start_date: str | None = None,
    end_date: str | None = None,
    high52_bonus: int = 4,
    recalc: bool = False,
) -> dict[str, Any]:
    paths = iter_screening_files(
        base_dir,
        start_date=start_date,
        end_date=end_date,
        require_high52_sheet=True,
    )
    results = update_many_screening_column(
        paths,
        column="O",
        value_factory=lambda row_index, _sheet: score_o_formula_without_strength(row_index, high52_bonus=high52_bonus),
    )
    recalc_ok: list[Path] = []
    recalc_failed: list[tuple[Path, str]] = []
    if recalc:
        recalc_ok, recalc_failed = recalculate_with_excel(paths)
    return {
        "files": len(results),
        "changed_cells": sum(item.changed_cells for item in results),
        "eligible_rows": sum(item.eligible_rows for item in results),
        "recalc_ok": len(recalc_ok),
        "recalc_failed": [(str(path), error) for path, error in recalc_failed],
        "results": results,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Batch-edit daily screening Excel workbooks.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    score_parser = subparsers.add_parser("apply-o-score", help="Set O-column score formula without execution strength.")
    score_parser.add_argument("--base-dir", default=str(SCREENING_DIR))
    score_parser.add_argument("--start-date", default=None)
    score_parser.add_argument("--end-date", default=None)
    score_parser.add_argument("--high52-bonus", type=int, default=4)
    score_parser.add_argument("--recalc", action="store_true", help="Open each workbook in Excel and save recalculated values.")

    avg_parser = subparsers.add_parser("apply-avg-qr", help="Fill Q/R columns with 1M/1W historical averages from prior files.")
    avg_parser.add_argument("--base-dir", default=str(SCREENING_DIR))
    avg_parser.add_argument("--start-date", default=None)
    avg_parser.add_argument("--end-date", default=None)
    avg_parser.add_argument("--include-no-52w", action="store_true", help="Include files without 52주신고가 sheet.")

    avg_one_parser = subparsers.add_parser("apply-avg-qr-date", help="Fill Q/R columns for one target date using previous files.")
    avg_one_parser.add_argument("--target-date", required=True)
    avg_one_parser.add_argument("--base-dir", default=str(SCREENING_DIR))
    avg_one_parser.add_argument("--include-no-52w", action="store_true", help="Include files without 52주신고가 sheet.")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    if args.command == "apply-o-score":
        payload = apply_o_score_formula(
            base_dir=Path(args.base_dir),
            start_date=args.start_date,
            end_date=args.end_date,
            high52_bonus=args.high52_bonus,
            recalc=args.recalc,
        )
        print(f"files={payload['files']}")
        print(f"changed_cells={payload['changed_cells']}")
        print(f"eligible_rows={payload['eligible_rows']}")
        print(f"recalc_ok={payload['recalc_ok']}")
        if payload["recalc_failed"]:
            print("recalc_failed=" + json.dumps(payload["recalc_failed"], ensure_ascii=False))
    elif args.command == "apply-avg-qr":
        payload = apply_avg_qr_from_history(
            base_dir=Path(args.base_dir),
            start_date=args.start_date,
            end_date=args.end_date,
            require_high52_sheet=not args.include_no_52w,
        )
        print(f"files={payload['files']}")
        print(f"touched_files={payload['touched_files']}")
        print(f"changed_cells={payload['changed_cells']}")
        print(f"eligible_rows={payload['eligible_rows']}")
    elif args.command == "apply-avg-qr-date":
        payload = apply_avg_qr_for_target_date(
            target_date=args.target_date,
            base_dir=Path(args.base_dir),
            require_high52_sheet=not args.include_no_52w,
        )
        print(f"target_date={payload['target_date']}")
        print(f"updated={payload.get('updated', False)}")
        print(f"changed_cells={payload.get('changed_cells', 0)}")
        print(f"eligible_rows={payload.get('eligible_rows', 0)}")
        if payload.get("reason"):
            print(f"reason={payload['reason']}")


if __name__ == "__main__":
    main()
