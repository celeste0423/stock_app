from __future__ import annotations

import json
import math
import mimetypes
import os
import re
import zipfile
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import time
import asyncio
import base64
import csv
import hashlib
import html as html_lib
import threading
import uuid
import zlib
import xml.etree.ElementTree as ET
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from difflib import SequenceMatcher
from email.utils import parsedate_to_datetime
from functools import lru_cache
from io import StringIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any, Callable
from urllib.parse import parse_qs, quote, unquote, urlparse
from zoneinfo import ZoneInfo

CURRENT_DIR = Path(__file__).resolve().parent
VENDOR_DIR = CURRENT_DIR / "vendor"
if str(VENDOR_DIR) not in sys.path:
    sys.path.insert(0, str(VENDOR_DIR))

import FinanceDataReader as fdr
import numpy as np
import OpenDartReader
import pandas as pd
try:
    from pykrx import stock as pykrx_stock
    PYKRX_IMPORT_ERROR: Exception | None = None
except Exception as exc:
    pykrx_stock = None
    PYKRX_IMPORT_ERROR = exc
import requests
import uvicorn
from bs4 import BeautifulSoup
from fastapi import FastAPI, HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from telethon import TelegramClient
from telethon.errors import SessionPasswordNeededError
from telethon.tl.types import PeerChannel


BASE_DIR = CURRENT_DIR.parent
FRONTEND_DIR = BASE_DIR / "frontend"
STATIC_DIR = FRONTEND_DIR / "static"
VENDOR_FRONTEND_DIR = FRONTEND_DIR / "vendor"
EXPORT_DIR = BASE_DIR / "outputs"
WORKSPACE_NODE_EXE = Path(
    os.getenv(
        "STOCK_DASHBOARD_NODE_EXE",
        "C:/Users/jyeob/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/bin/node.exe",
    )
)
WORKSPACE_NODE_MODULES = Path(
    os.getenv(
        "STOCK_DASHBOARD_NODE_MODULES",
        "C:/Users/jyeob/.cache/codex-runtimes/codex-primary-runtime/dependencies/node/node_modules",
    )
)
PORTFOLIO_EXPORT_SCRIPT = CURRENT_DIR / "portfolio_export_builder.mjs"

PORTFOLIO_PATH = Path(
    os.getenv(
        "STOCK_DASHBOARD_PORTFOLIO_PATH",
        "C:/Users/jyeob/OneDrive - SK Hynix Inc/Cloud/\ud22c\uc790/\uc8fc\uc2dd/\ube44\uc911_\ub370\uc77c\ub9ac.xlsx",
    )
)
SCREENING_DIR = Path(os.getenv("STOCK_DASHBOARD_SCREENING_DIR", "D:/Study/Stock_Daily"))
SCREENING_FAST_DB_PATH = Path(os.getenv("STOCK_DASHBOARD_SCREENING_FAST_DB_PATH", str(CURRENT_DIR / "stock_daily_fast.sqlite")))
SCREENING_FAST_PARQUET_PATH = Path(os.getenv("STOCK_DASHBOARD_SCREENING_FAST_PARQUET_PATH", str(CURRENT_DIR / "stock_daily_fast.parquet")))
US_SCREENING_FAST_DB_PATH = Path(os.getenv("STOCK_DASHBOARD_US_SCREENING_FAST_DB_PATH", str(CURRENT_DIR / "us_stock_daily_fast.sqlite")))
US_SCREENING_FAST_PARQUET_PATH = Path(os.getenv("STOCK_DASHBOARD_US_SCREENING_FAST_PARQUET_PATH", str(CURRENT_DIR / "us_stock_daily_fast.parquet")))
ASIA_SCREENING_FAST_DB_PATH = Path(os.getenv("STOCK_DASHBOARD_ASIA_SCREENING_FAST_DB_PATH", str(CURRENT_DIR / "asia_stock_daily_fast.sqlite")))
ASIA_SCREENING_FAST_PARQUET_PATH = Path(os.getenv("STOCK_DASHBOARD_ASIA_SCREENING_FAST_PARQUET_PATH", str(CURRENT_DIR / "asia_stock_daily_fast.parquet")))
SCREENING_SQL_ONLY = str(os.getenv("STOCK_DASHBOARD_SCREENING_SQL_ONLY", "1")).strip().lower() in {"1", "true", "yes", "y", "on"}
STATE_DIR = Path(os.getenv("STOCK_DASHBOARD_STATE_DIR", str(CURRENT_DIR)))
SETTINGS_PATH = STATE_DIR / "local_settings.json"
SECTOR_DB_PATH = STATE_DIR / "sector_database.json"
SECTOR_DB_BACKUP_DIR = STATE_DIR / "sector_database_backups"
SCREENING_CACHE_PATH = STATE_DIR / "screening_cache.json"
US_SCREENING_CACHE_PATH = STATE_DIR / "us_screening_cache.json"
ASIA_SCREENING_CACHE_PATH = STATE_DIR / "asia_screening_cache.json"
SCREENING_CALENDAR_CACHE_VERSION = 6
US_SCREENING_CALENDAR_CACHE_VERSION = 1
ASIA_SCREENING_CALENDAR_CACHE_VERSION = 1
TELEGRAM_SESSION_DIR = STATE_DIR / "telegram_session"
TELEGRAM_SESSION_FILE = TELEGRAM_SESSION_DIR / "user"
TELEGRAM_CODE_FILE = TELEGRAM_SESSION_DIR / "login_state.json"
TELEGRAM_ATTACHMENT_DIR = TELEGRAM_SESSION_DIR / "attachments"
TELEGRAM_LOCK = threading.Lock()
TELEGRAM_EARNINGS_LOCK = threading.Lock()
TELEGRAM_SEARCH_JOBS: dict[str, dict[str, Any]] = {}
TELEGRAM_EARNINGS_SEARCH_JOBS: dict[str, dict[str, Any]] = {}
TELEGRAM_EARNINGS_CHANNEL_NAME = "Awake-실시간 주식 공시 정리채널"
TELEGRAM_EARNINGS_CHANNEL_USERNAME = "darthacking"
TELEGRAM_EARNINGS_CHANNEL_ID = -1001066938528
TELEGRAM_EARNINGS_DIALOG_CACHE: dict[str, Any] = {}
TRADE_DATA_CACHE_PATH = STATE_DIR / "trade_import_export_cache.json"
TRADE_SNAPSHOT_HISTORY_PATH = STATE_DIR / "trade_snapshot_history.json"
DRAM_PRICE_HISTORY_PATH = STATE_DIR / "dram_price_history.json"
SSD_PRICE_HISTORY_PATH = STATE_DIR / "ssd_price_history.json"
TOURIST_VISITOR_CACHE_PATH = STATE_DIR / "tourist_inbound_visitors_cache.json"
ECONOMY_CYCLE_CACHE_PATH = STATE_DIR / "economy_cycle_clock_cache.json"
FRED_PRICE_CACHE_DIR = STATE_DIR / "fred_price_cache"
TRADINGECONOMICS_PRICE_CACHE_DIR = STATE_DIR / "tradingeconomics_price_cache"
SMM_PRICE_CACHE_DIR = STATE_DIR / "smm_price_cache"
KOMIS_PRICE_CACHE_DIR = STATE_DIR / "komis_price_cache"
DART_EARNINGS_TREND_CACHE_DIR = STATE_DIR / "dart_earnings_trend_cache"
KIS_TOKEN_CACHE_PATH = STATE_DIR / "kis_token_cache.json"
MARKET_INVESTOR_FLOW_CACHE_PATH = STATE_DIR / "market_investor_flow_cache.json"
SIGNAL_RADAR_CACHE_PATH = STATE_DIR / "signal_radar_cache.json"
KIND_BUSINESS_SEGMENT_CACHE_DIR = STATE_DIR / "kind_business_segment_cache"
GLOBAL_INDICES_PAYLOAD_CACHE_DIR = STATE_DIR / "global_indices_payload_cache"
CHART_PREVIEW_CACHE_DIR = STATE_DIR / "stock_chart_preview_cache"
GLOBAL_COMPANY_AI_CACHE_PATH = STATE_DIR / "global_company_ai_cache.json"
GLOBAL_INDICES_PAYLOAD_CACHE: dict[str, tuple[float, dict[str, Any]]] = {}
GLOBAL_INDICES_PAYLOAD_CACHE_VERSION = "v3"
PORTFOLIO_PERFORMANCE_CACHE: dict[str, Any] = {}
PORTFOLIO_PERFORMANCE_CACHE_LOCK = threading.Lock()
PORTFOLIO_PERFORMANCE_CACHE_TTL_SECONDS = 180
REAL_ESTATE_DB_PATH = STATE_DIR / "real_estate_building.json"
REAL_ESTATE_PRICE_CACHE_PATH = STATE_DIR / "real_estate_price_cache.json"
REAL_ESTATE_TRADE_CACHE_PATH = STATE_DIR / "real_estate_trade_cache.json"
MARKET_CALENDAR_PATH = STATE_DIR / "market_calendar_events.json"
MARKET_CALENDAR_AUTO_CACHE_PATH = STATE_DIR / "market_calendar_auto_cache.json"
MARKET_CALENDAR_AUTO_CACHE_TTL_SECONDS = 6 * 60 * 60
MARKET_CALENDAR_MIN_KR_MARCAP = 2000 * 100000000
STOCK_ALERT_HOLDINGS_SNAPSHOT_PATH = STATE_DIR / "stock_alert_holdings_snapshot.json"
REAL_ESTATE_EXCEL_PATH = Path(
    os.getenv(
        "STOCK_DASHBOARD_REAL_ESTATE_EXCEL_PATH",
        "D:/Study/\uc0c1\uac00_\uad00\ub9ac_\ub370\uc774\ud130/\uc548\uc554\ud574\ub9c1\ud134 \uc0c1\uac00 \uad00\ub9ac.xlsx",
    )
)
REAL_ESTATE_DATA_DIR = REAL_ESTATE_EXCEL_PATH.parent
REAL_ESTATE_BANK_IMPORT_DIR = Path(
    os.getenv(
        "STOCK_DASHBOARD_REAL_ESTATE_BANK_IMPORT_DIR",
        str(REAL_ESTATE_DATA_DIR / "계좌입출금내역"),
    )
)
TRADE_DATA_VERSION = 3
TOURIST_VISITOR_DATA_VERSION = 1
KIS_MOCK_BASE_URL = "https://openapivts.koreainvestment.com:29443"
KIS_REAL_BASE_URL = "https://openapi.koreainvestment.com:9443"
SEC_USER_AGENT = os.getenv("STOCK_DASHBOARD_SEC_USER_AGENT", "stock-dashboard local app jyeob@example.com")
GLOBAL_COMPANY_ALIASES = {
    "애플": "AAPL",
    "아이폰": "AAPL",
    "마이크로소프트": "MSFT",
    "마소": "MSFT",
    "엔비디아": "NVDA",
    "엔비디아코퍼레이션": "NVDA",
    "테슬라": "TSLA",
    "알파벳": "GOOGL",
    "구글": "GOOGL",
    "아마존": "AMZN",
    "메타": "META",
    "페이스북": "META",
    "넷플릭스": "NFLX",
    "브로드컴": "AVGO",
    "퀄컴": "QCOM",
    "amd": "AMD",
    "어드밴스드마이크로디바이시스": "AMD",
    "인텔": "INTC",
    "코카콜라": "KO",
    "펩시": "PEP",
    "월마트": "WMT",
    "코스트코": "COST",
    "맥도날드": "MCD",
    "스타벅스": "SBUX",
    "나이키": "NKE",
    "디즈니": "DIS",
    "오라클": "ORCL",
    "세일즈포스": "CRM",
    "어도비": "ADBE",
    "팔란티어": "PLTR",
    "슈퍼마이크로컴퓨터": "SMCI",
    "슈마컴": "SMCI",
    "마이크론": "MU",
    "버크셔": "BRK-B",
    "버크셔해서웨이": "BRK-B",
    "일라이릴리": "LLY",
    "노보노디스크": "NVO",
    "tsmcf": "TSM",
    "tsmc": "TSM",
    "대만반도체": "TSM",
    "대만반도체제조": "TSM",
    "타이완세미컨덕터": "TSM",
    "토요타": "TM",
    "도요타": "TM",
    "텐센트": "0700.HK",
    "알리바바": "BABA",
    "asml": "ASML",
}
TELEGRAM_DISCLOSURE_CATEGORIES = {
    "earnings": {
        "label": "실적",
        "tokens": ["매출액", "영업익", "영업이익", "순이익", "순익", "최근 실적 추이", "잠정실적", "재무제표 종류"],
    },
    "orders": {
        "label": "수주",
        "tokens": ["수주", "공급계약", "단일판매", "판매ㆍ공급계약", "판매공급계약", "공급 계약"],
    },
    "warning": {
        "label": "투자경고",
        "tokens": ["투자경고", "투자주의", "투자위험", "단기과열", "매매거래정지", "관리종목", "불성실공시"],
    },
    "investment": {
        "label": "투자·증자",
        "tokens": ["시설투자", "신규시설투자", "타법인", "유상증자", "무상증자", "전환사채", "신주인수권", "자금조달", "회사분할", "분할결정", "합병", "분할합병", "영업양수", "영업양도"],
    },
    "ownership": {
        "label": "지분 변화",
        "tokens": ["주식등의대량보유상황보고서", "대량보유", "임원ㆍ주요주주", "임원·주요주주", "소유상황보고서", "보고전", "보고후", "보유목적", "장내매수", "장내매도"],
    },
    "shareholder": {
        "label": "배당·자사주",
        "tokens": ["배당", "자사주", "자기주식", "주식소각", "현금배당"],
    },
    "all": {"label": "전체", "tokens": []},
}

PORTFOLIO_SHEET = "\uc8fc\uc2dd\ube44\uc911"
SCREENING_SHEET = "\uc8fc\ub3c4\uc8fc \ucc3e\uae30"
SCREENING_SCORE_CACHE_VERSION = "s-score-v3"
SCREENING_SCORE_COLUMN_INDEX = 18  # Excel S column, zero-based for pandas iloc.
SCREENING_SCORE_COLUMN_NAME = "\uc885\ud569 \uc810\uc218"

NAME_ALIASES = {
    "\uc0bc\uc804": "\uc0bc\uc131\uc804\uc790",
    "\ub124\uc774\ubc84": "NAVER",
    "\ub2c9\uc2a4": "SK\ud558\uc774\ub2c9\uc2a4",
    "\uc5d8\uc77c\ub809": "LS ELECTRIC",
    "\uc6d0\uc775ips": "\uc6d0\uc775IPS",
    "kb\uae08\uc735": "KB\uae08\uc735",
    "\ud0a4\uc6c0": "\ud0a4\uc6c0\uc99d\uad8c",
    "lig\ub125\uc2a4\uc6d0": "LIG\ub514\ud39c\uc2a4\uc564\uc5d0\uc5b4\ub85c\uc2a4\ud398\uc774\uc2a4",
}

HEADER_TARGET = "\uc774\ud6c4"
HEADER_PREV = "\uc774\uc804"
HEADER_WEIGHT = "\ube44\uc911"
ROW_TOTAL = "\uacc4"
ROW_SEED = "\uc2dc\ub4dc"
ROW_SECTOR = "\uc139\ud130"
RECENT_SCREENING_LOOKBACK = 20
SCREENING_MIN_MARKET_CAP_100M = 2000.0

THEME_RULES = [
    {"theme": "\uc6d0\uc804", "keywords": ["\uc6d0\uc804", "smr", "\uc6d0\uc790\ub825", "\ud300\ucf54\ub9ac\uc544", "\ud55c\uc804", "\ub450\uc0b0\uc5d0\ub108\ube4c\ub9ac\ud2f0"]},
    {"theme": "\uc804\uc120\u00b7\uc804\ub825 \uc778\ud504\ub77c", "keywords": ["\uc804\uc120", "\uc804\ub825", "\ub098\ud504\ud0c0", "\uc804\ub825\ub9dd", "\uc804\ub825\uae30\uae30", "\ucd08\uace0\uc555", "\ubcc0\uc555\uae30"]},
    {"theme": "\ubc18\ub3c4\uccb4\u00b7HBM", "keywords": ["hbm", "\ubc18\ub3c4\uccb4", "\uba54\ubaa8\ub9ac", "\uc0bc\uc131\uc804\uc790", "sk\ud558\uc774\ub2c9\uc2a4", "\uc601\uc775", "\uc2e4\uc801", "\uc804\uacf5\uc815"]},
    {"theme": "AI \uc11c\ubc84 \uae30\ud310", "keywords": ["\uae30\ud310", "pcb", "fc-bga", "\uc1fc\ud2f0\uc9c0", "ai \uc11c\ubc84"]},
    {"theme": "\uc870\uc120\u00b7LNG", "keywords": ["lng", "\uc870\uc120", "\uc5d4\uc9c4", "\uc120\ubc15", "\uac00\uc2a4\uc804", "\uc6b4\ubc18\uc120", "\ud574\uc6b4"]},
    {"theme": "SpaceX\u00b7\uc6b0\uc8fc \ud22c\uc790", "keywords": ["spacex", "xai", "\uc6b0\uc8fc", "\uc138\ubbf8\ud30c\uc774\ube0c", "\ubca4\ucc98\ud22c\uc790", "\ubc1c\uc0ac\uccb4"]},
    {"theme": "\ubc29\uc0b0", "keywords": ["\ubc29\uc0b0", "\ucc9c\uad81", "\ubbf8\uc0ac\uc77c", "\ubc29\uacf5", "\uad6d\ubc29", "\uc218\uc8fc"]},
    {"theme": "\ud0dc\uc591\uad11\u00b7\uc5d0\ub108\uc9c0", "keywords": ["\ud0dc\uc591\uad11", "\uc5d0\ub108\uc9c0", "\ud398\ub85c\ube0c\uc2a4\uce74\uc774\ud2b8", "\ud48d\ub825", "\uc5f0\ub8cc\uc804\uc9c0", "2\ucc28\uc804\uc9c0"]},
    {"theme": "\ubc14\uc774\uc624\u00b7\ud5ec\uc2a4\ucf00\uc5b4", "keywords": ["\ubc14\uc774\uc624", "\uc81c\uc57d", "\ud5ec\uc2a4\ucf00\uc5b4", "\uc2e0\uc57d", "\uc758\ub8cc", "\ud53c\ubd80\ubbf8\uc6a9"]},
    {"theme": "\ub85c\ubd07\u00b7\uc790\ub3d9\ud654", "keywords": ["\ub85c\ubd07", "\uc790\ub3d9\ud654", "\ud734\uba38\ub178\uc774\ub4dc"]},
]

STOCK_THEME_RULES = [
    {"theme": "\uc804\uc120\u00b7\uc804\ub825 \uc778\ud504\ub77c", "keywords": ["\ub300\uc6d0\uc804\uc120", "\ub300\ud55c\uc804\uc120", "\uac00\uc628\uc804\uc120", "kbi\uba54\ud0c8", "ls electric", "ls\uc5d0\ucf54\uc5d0\ub108\uc9c0", "\uc138\uba85\uc804\uae30", "\uacc4\uc591\uc804\uae30"]},
    {"theme": "\uc6d0\uc804", "keywords": ["\uc6b0\ub9ac\uae30\uc220", "\ubcf4\uc131\ud30c\uc6cc\ud14d", "\ub300\uc6b0\uac74\uc124", "\ub450\uc0b0\uc5d0\ub108\ube4c\ub9ac\ud2f0", "\ud604\ub300\uac74\uc124", "\uc6b0\uc9c4"]},
    {"theme": "SpaceX\u00b7\uc6b0\uc8fc \ud22c\uc790", "keywords": ["\uc544\uc8fcib\ud22c\uc790", "\ubbf8\ub798\uc5d0\uc14b\ubca4\ucc98\ud22c\uc790", "\ud55c\uad6d\ud56d\uacf5\uc6b0\uc8fc"]},
    {"theme": "\uc870\uc120\u00b7LNG", "keywords": ["\uc0bc\uc131\uc911\uacf5\uc5c5", "hd\ud604\ub300\uc911\uacf5\uc5c5", "hj\uc911\uacf5\uc5c5", "\ud55c\ud654\uc5d4\uc9c4", "\ud765\uc544\ud574\uc6b4"]},
    {"theme": "AI \uc11c\ubc84 \uae30\ud310", "keywords": ["\uc774\uc218\ud398\ud0c0\uc2dc\uc2a4", "\ub300\ub355\uc804\uc790", "\ud574\uc131\ub514\uc5d0\uc2a4"]},
    {"theme": "\ubc18\ub3c4\uccb4\u00b7HBM", "keywords": ["\uc0bc\uc131\uc804\uc790", "sk\ud558\uc774\ub2c9\uc2a4", "db\ud558\uc774\ud14d", "\ud55c\ubbf8\ubc18\ub3c4\uccb4", "\uc0ac\ud53c\uc5d4\ubc18\ub3c4\uccb4"]},
]

NEWS_IMPORTANT_KEYWORDS = [
    "실적",
    "영업익",
    "매출",
    "순이익",
    "어닝",
    "수주",
    "계약",
    "공급",
    "납품",
    "투자",
    "증설",
    "공장",
    "양산",
    "생산",
    "인수",
    "합병",
    "매각",
    "지분",
    "자회사",
    "공시",
    "유상증자",
    "무상증자",
    "자사주",
    "배당",
    "특허",
    "허가",
    "승인",
    "임상",
    "fda",
    "정책",
    "정부",
    "규제",
    "제재",
    "협력",
    "파트너",
    "mou",
    "출시",
    "신제품",
    "수출",
]

NEWS_NOISE_KEYWORDS = [
    "특징주",
    "급등",
    "급락",
    "상승",
    "하락",
    "강세",
    "약세",
    "주가",
    "시황",
    "마감",
    "장중",
    "증시",
    "코스피",
    "코스닥",
    "거래량",
    "신고가",
    "52주",
    "투자주의",
    "목표가",
]

NEWS_BLOCKED_SOURCE_KEYWORDS = [
    "주달",
    "네이버 프리미엄콘텐츠",
    "네이버프리미엄콘텐츠",
    "프리미엄콘텐츠",
    "블로그",
    "blog",
    "카페",
    "cafe",
    "포스트",
    "post",
    "브런치",
    "brunch",
    "티스토리",
    "tistory",
    "인플루언서",
    "contents.premium.naver.com",
    "m.blog.naver.com",
    "blog.naver.com",
    "cafe.naver.com",
    "post.naver.com",
]

STRATEGY_INDEXES = {
    "KS11": {"name": "KOSPI", "symbol": "KS11"},
    "KQ11": {"name": "KOSDAQ", "symbol": "KQ11"},
    "IXIC": {"name": "NASDAQ", "symbol": "IXIC"},
    "US500": {"name": "S&P 500", "symbol": "US500"},
    "DJI": {"name": "Dow Jones", "symbol": "DJI"},
    "RUT": {"name": "Russell 2000", "symbol": "RUT"},
}

GLOBAL_INDEX_ITEMS = [
    {"symbol": "US500", "name": "S&P 500", "group": "국가별 지수"},
    {"symbol": "IXIC", "name": "Nasdaq", "group": "국가별 지수"},
    {"symbol": "DJI", "name": "Dow Jones", "group": "국가별 지수"},
    {"symbol": "RUT", "name": "Russell 2000", "group": "국가별 지수"},
    {"symbol": "KS11", "name": "KOSPI", "group": "국가별 지수"},
    {"symbol": "KQ11", "name": "KOSDAQ", "group": "국가별 지수"},
    {"symbol": "N225", "name": "Nikkei 225", "group": "국가별 지수"},
    {"symbol": "HSI", "name": "Hang Seng", "group": "국가별 지수"},
    {"symbol": "^TWII", "name": "Taiwan Weighted", "group": "국가별 지수"},
    {"symbol": "SSEC", "name": "Shanghai Composite", "group": "국가별 지수"},
    {"symbol": "^NSEI", "name": "Nifty 50", "group": "국가별 지수"},
    {"symbol": "GDAXI", "name": "DAX", "group": "국가별 지수"},
    {"symbol": "FTSE", "name": "FTSE 100", "group": "국가별 지수"},
    {"symbol": "GC=F", "name": "Gold", "group": "원자재"},
    {"symbol": "SI=F", "name": "Silver", "group": "원자재"},
    {"symbol": "CL=F", "name": "WTI Oil", "group": "원자재"},
    {"symbol": "BZ=F", "name": "Brent Oil", "group": "원자재"},
    {
        "symbol": "CMCU3",
        "fetch_symbol": "HG=F",
        "name": "LME Copper 3M",
        "group": "원자재",
        "source": "COMEX Copper proxy, USD/t 환산",
        "source_symbol": "HG=F",
        "price_multiplier": 2204.62262185,
    },
    {
        "symbol": "CMAL3",
        "fetch_symbol": "ALI=F",
        "name": "LME Aluminium 3M",
        "group": "원자재",
        "source": "CME Aluminium proxy, USD/t",
        "source_symbol": "ALI=F",
    },
    {"symbol": "ZC=F", "name": "Corn", "group": "원자재"},
    {"symbol": "ZS=F", "name": "Soybean", "group": "원자재"},
    {"symbol": "NG=F", "name": "Natural Gas", "group": "원자재"},
    {
        "symbol": "LCM",
        "fetch_symbol": "EASTMONEY:225.lcm",
        "name": "Lithium Carbonate",
        "group": "원자재",
        "source": "Eastmoney GFEX lithium carbonate continuous futures",
        "source_symbol": "GFEX 225.lcm",
    },
    {
        "symbol": "LI2S",
        "fetch_symbol": "SMM:202508060001",
        "name": "Lithium Sulfide",
        "group": "원자재",
        "source": "SMM Battery-Grade Lithium Sulfide VAT included, yuan/kg",
        "source_symbol": "SMM 202508060001",
    },
    {
        "symbol": "PSM",
        "fetch_symbol": "EASTMONEY:225.psm",
        "name": "Polysilicon",
        "group": "원자재",
        "source": "Eastmoney GFEX polysilicon continuous futures",
        "source_symbol": "GFEX 225.psm",
    },
    {
        "symbol": "FIBEROPTIC",
        "fetch_symbol": "BLS:WPU102603",
        "name": "Fiber/Communication Cable PPI",
        "group": "원자재",
        "source": "BLS Producer Price Index: Communication and Energy Wire and Cable proxy",
        "source_symbol": "WPU102603",
    },
    {
        "symbol": "MOLYBDENUM",
        "fetch_symbol": "TRADINGECONOMICS:MOLYBDEN:COM",
        "name": "Molybdenum",
        "group": "원자재",
        "source": "Trading Economics CFD benchmark, CNY/Kg",
        "source_symbol": "MOLYBDEN:COM",
    },
    {"symbol": "BTC-USD", "name": "Bitcoin", "group": "비트코인"},
    {"symbol": "^IRX", "name": "US 13W Yield", "group": "국채금리"},
    {"symbol": "^FVX", "name": "US 5Y Yield", "group": "국채금리"},
    {"symbol": "^TNX", "name": "US 10Y Yield", "group": "국채금리"},
    {"symbol": "^TYX", "name": "US 30Y Yield", "group": "국채금리"},
    {"symbol": "KRW=X", "name": "USD/KRW", "group": "환율", "source": "Yahoo Finance", "source_symbol": "KRW=X"},
    {"symbol": "JPY=X", "name": "USD/JPY", "group": "환율", "source": "Yahoo Finance", "source_symbol": "JPY=X"},
    {"symbol": "CNY=X", "name": "USD/CNY", "group": "환율", "source": "Yahoo Finance", "source_symbol": "CNY=X"},
    {"symbol": "EURUSD=X", "name": "EUR/USD", "group": "환율", "source": "Yahoo Finance", "source_symbol": "EURUSD=X"},
    {"symbol": "DX-Y.NYB", "name": "Dollar Index", "group": "환율", "source": "Yahoo Finance", "source_symbol": "DX-Y.NYB"},
]

KOREA_STOCK_ETF_ITEMS = [
    ("495850", "KODEX KOREA Value-up", "코리아 밸류업"),
    ("494670", "TIGER Shipbuilding TOP10", "조선 TOP10"),
    ("491820", "HANARO Electric Power Capex", "전력 설비 투자"),
    ("487240", "KODEX AI Electric Power Core Facilities", "AI 전력 핵심 설비"),
    ("475050", "ACE KPOP Focus", "KPOP 포커스"),
    ("471990", "KODEX AI Semiconductor Core Equipment", "AI 반도체 핵심 장비"),
    ("466940", "TIGER Bank High Dividend Plus TOP10", "은행 고배당 플러스"),
    ("463250", "TIGER K-Defense Industry & Space", "K-방산 및 우주"),
    ("449450", "PLUS K-Defense Industry", "K-방산"),
    ("445290", "KODEX Robot Active", "로봇 액티브"),
    ("438900", "HANARO K-Food", "K-푸드"),
    ("434730", "HANARO Nuclear Power", "원자력"),
    ("421320", "PLUS Aerospace & UAM", "우주항공 및 UAM"),
    ("395160", "KODEX AI Semiconductor TOP2 Plus", "AI 반도체 TOP2 플러스"),
    ("385510", "KODEX Renewable Energy Active", "신재생 에너지 액티브"),
    ("364970", "TIGER KRX Bio K-New Deal", "바이오 K-뉴딜"),
    ("329200", "TIGER REITs Real Estate Infra", "리츠 부동산 인프라"),
    ("307520", "TIGER Holdings Company", "지주회사"),
    ("305720", "KODEX Secondary Battery Industry", "2차전지 산업"),
    ("300950", "KODEX Game Industry", "게임 산업"),
    ("292150", "TIGER TOP10", "국내 TOP10"),
    ("266390", "KODEX Consumer Discretionary", "경기 소비재"),
    ("266360", "KODEX K-Content", "K-콘텐츠"),
    ("261140", "TIGER Preferred Stock", "우선주"),
    ("252650", "KODEX 200 Equalweight", "코스피 200 동일가중"),
    ("244580", "KODEX BIO", "바이오"),
    ("229200", "KODEX KOSDAQ150", "코스닥 150"),
    ("228810", "TIGER Media Contents", "미디어 콘텐츠"),
    ("228800", "TIGER Travel Leisure", "여행 레저"),
    ("228790", "TIGER Cosmetics", "화장품"),
    ("161510", "PLUS High Dividend", "고배당"),
    ("157490", "TIGER Software", "소프트웨어"),
    ("150460", "TIGER China Consumer", "중국 소비재"),
    ("143860", "TIGER Health Care", "헬스케어"),
    ("140710", "KODEX Transportation", "운송"),
    ("140700", "KODEX Insurance", "보험"),
    ("139280", "TIGER Consumer Staples", "필수 소비재"),
    ("139270", "TIGER Financials", "금융"),
    ("139260", "TIGER IT", "IT"),
    ("139230", "TIGER 200 Heavy Industry", "200 중공업"),
    ("138540", "TIGER Hyundai Motor Group", "현대차 그룹"),
    ("117700", "KODEX Construction", "건설"),
    ("117680", "KODEX Steels", "철강"),
    ("117460", "KODEX Energy & Chemicals", "에너지 화학"),
    ("102970", "KODEX Securities", "증권"),
    ("102960", "KODEX Machinery & Equipment", "기계 장비"),
    ("102780", "KODEX Samsung Group", "삼성그룹"),
    ("091180", "KODEX Autos", "자동차"),
    ("091170", "KODEX Banks", "은행"),
    ("069500", "KODEX 200", "코스피 200"),
    ("0105E0", "SOL Korea High Dividend", "국내 고배당"),
]

GLOBAL_INDEX_ITEMS.extend(
    {
        "symbol": f"KETF-{code}",
        "fetch_symbol": code,
        "name": f"{theme} · {etf_name}",
        "group": "한국주식ETF",
        "source": "FinanceDataReader KRX ETF",
        "source_symbol": code,
    }
    for code, etf_name, theme in KOREA_STOCK_ETF_ITEMS
)

GLOBAL_PPI_ITEMS = [
    ("PCU3344183344189", "PCB Assemblies / Loaded Boards PPI"),
    ("PCU334418334418", "Printed Circuit Assembly Mfg PPI"),
    ("PCU334412334412", "Bare Printed Circuit Board Mfg PPI"),
    ("PCU334413334413", "Semiconductor Device Mfg PPI"),
    ("PCU3344133344131", "Integrated Circuit Packages PPI"),
    ("PCU334413334413A", "Other Semiconductor Devices / Wafers PPI"),
    ("PCU3344173344170", "Electronic Connectors PPI"),
    ("PCU33441K33441K", "Capacitor/Resistor/Coil/Transformer PPI"),
    ("WPU117839", "IC Packages incl. Microprocessors PPI"),
]

GLOBAL_INDEX_ITEMS.extend(
    {
        "symbol": f"PPI-{series_id}",
        "fetch_symbol": f"BLS:{series_id}",
        "name": name,
        "group": "PPI",
        "source": "BLS API / FRED PPI series",
        "source_symbol": series_id,
        "frequency": "monthly",
    }
    for series_id, name in GLOBAL_PPI_ITEMS
)

KOMIS_MINERAL_PRICE_ITEMS = [
    ("HP001", "MNRL0002", "니켈", "Nickel"),
    ("HP001", "MNRL0008", "동", "Copper"),
    ("HP001", "MNRL0023", "아연", "Zinc"),
    ("HP001", "MNRL0009", "알루미늄", "Aluminum"),
    ("HP001", "MNRL0022", "연", "Lead"),
    ("HP001", "MNRL0016", "주석", "Tin"),
    ("HP002", "MNRL1064", "가돌리늄", "Gadolinium"),
    ("HP002", "MNRL0024", "갈륨", "Gallium"),
    ("HP002", "MNRL0035", "게르마늄", "Germanium"),
    ("HP002", "MNRL0010", "규소", "Silicon"),
    ("HP002", "MNRL1001", "네오디뮴", "Neodymium"),
    ("HP002", "MNRL0007", "니오븀", "Niobium"),
    ("HP002", "MNRL1004", "디스프로슘", "Dysprosium"),
    ("HP002", "MNRL1003", "란탄", "Lanthanum"),
    ("HP002", "MNRL1055", "루테튬", "Lutetium"),
    ("HP002", "MNRL0001", "리튬", "Lithium"),
    ("HP002", "MNRL0011", "마그네슘", "Magnesium"),
    ("HP002", "MNRL0004", "망간", "Manganese"),
    ("HP002", "MNRL0012", "몰리브덴", "Molybdenum"),
    ("HP002", "MNRL0013", "바나듐", "Vanadium"),
    ("HP002", "MNRL1062", "사마륨", "Samarium"),
    ("HP002", "MNRL1002", "세륨", "Cerium"),
    ("HP002", "MNRL0029", "셀레늄", "Selenium"),
    ("HP002", "MNRL1053", "스칸듐", "Scandium"),
    ("HP002", "MNRL0028", "스트론튬", "Strontium"),
    ("HP002", "MNRL0019", "안티모니", "Antimony"),
    ("HP002", "MNRL1067", "에르븀", "Erbium"),
    ("HP002", "MNRL1063", "유로퓸", "Europium"),
    ("HP002", "MNRL1054", "이트륨", "Yttrium"),
    ("HP002", "MNRL0025", "인듐", "Indium"),
    ("HP002", "MNRL0027", "지르코늄", "Zirconium"),
    ("HP002", "MNRL0020", "창연", "Bismuth"),
    ("HP002", "MNRL0003", "코발트", "Cobalt"),
    ("HP002", "MNRL0021", "크롬", "Chromium"),
    ("HP002", "MNRL0026", "탄탈륨", "Tantalum"),
    ("HP002", "MNRL1005", "터븀", "Terbium"),
    ("HP002", "MNRL0018", "텅스텐", "Tungsten"),
    ("HP002", "MNRL0017", "티타늄", "Titanium"),
    ("HP002", "MNRL1056", "프라세오디뮴", "Praseodymium"),
    ("HP002", "MNRL1068", "홀뮴", "Holmium"),
    ("HP003", "MNRL0031", "우라늄", "Uranium"),
    ("HP003", "MNRL0032", "유연탄", "Coal"),
    ("HP003", "MNRL1011", "철광석", "Iron Ore"),
    ("HP004", "MNRL0046", "금", "Gold"),
    ("HP004", "MNRL0014", "백금", "Platinum"),
    ("HP004", "MNRL0047", "은", "Silver"),
    ("HP004", "MNRL0015", "팔라듐", "Palladium"),
    ("HP004", "MNRL0005", "흑연", "Graphite"),
]

GLOBAL_INDEX_ITEMS.extend(
    {
        "symbol": f"KOMIS-{hp}-{code}",
        "fetch_symbol": f"KOMIS:{hp}:{code}",
        "name": f"{korean_name} ({english_name})",
        "group": "KOMIS 광물가격",
        "source": "KOMIS 광물자원가격",
        "source_symbol": code,
    }
    for hp, code, korean_name, english_name in KOMIS_MINERAL_PRICE_ITEMS
)

STRATEGY_TYPES = {
    "ma20_cross": {
        "name": "20일선 돌파",
        "description": "종가가 20일 이동평균선을 상향 돌파하면 매수, 하향 돌파하면 매도합니다. 매수/매도마다 0.2% 비용을 반영합니다.",
    },
    "ma20_cross_mdd7": {
        "name": "20일선 돌파 + MDD 7% 매도",
        "description": "20일선 상향 돌파 시 매수하고, 20일선 하향 이탈 또는 매수 후 최고 NAV 대비 7% 이상 하락 시 매도합니다. 매수/매도마다 0.2% 비용을 반영합니다.",
    },
    "golden_cross": {
        "name": "골든크로스",
        "description": "20일 이동평균선이 60일 이동평균선을 상향 돌파하면 매수, 하향 돌파하면 매도합니다. 매수/매도마다 0.2% 비용을 반영합니다.",
    },
    "rsi_rebound": {
        "name": "RSI 반등",
        "description": "RSI(14)가 30선을 상향 돌파하면 매수, 70선 아래로 재진입하면 매도합니다. 매수/매도마다 0.2% 비용을 반영합니다.",
    },
    "leader_top10_score70": {
        "name": "주도주 Top10·70점 리밸런싱",
        "description": "종합점수 상위 10위이면서 70점 이상에 처음 진입한 종목을 매수하고, 보유 종목 점수가 70점 이하로 내려가면 매도합니다. 보유 종목은 매일 점수 비율로 리밸런싱하며 회전 비용 0.2%를 반영합니다.",
    },
    "leader_top5_score80": {
        "name": "주도주 Top5·80점 리밸런싱",
        "description": "종합점수 상위 5위이면서 80점 이상에 처음 진입한 종목을 매수하고, 보유 종목 점수가 80점 미만으로 내려가면 매도합니다. 보유 종목은 매일 점수 비율로 리밸런싱하며 회전 비용 0.2%를 반영합니다.",
    },
    "leader_top5_score70": {
        "name": "주도주 Top5·70점 리밸런싱",
        "description": "종합점수 상위 5위이면서 70점 이상에 처음 진입한 종목을 매수하고, 보유 종목 점수가 70점 미만으로 내려가면 매도합니다. 보유 종목은 매일 점수 비율로 리밸런싱하며 회전 비용 0.2%를 반영합니다.",
    },
    "leader_all_score80": {
        "name": "주도주 전종목·80점 리밸런싱",
        "description": "종합점수 80점 이상인 종목을 모두 매수하고, 보유 종목 점수가 80점 미만으로 내려가면 매도합니다. 보유 종목은 매일 점수 비율로 리밸런싱하며 회전 비용 0.2%를 반영합니다.",
    },
    "leader_all_score70": {
        "name": "주도주 전종목·70점 리밸런싱",
        "description": "종합점수 70점 이상인 종목을 모두 매수하고, 보유 종목 점수가 70점 이하로 내려가면 매도합니다. 보유 종목은 매일 점수 비율로 리밸런싱하며 회전 비용 0.2%를 반영합니다.",
    },
    "leader_all_score55": {
        "name": "주도주 전종목·55점 리밸런싱",
        "description": "종합점수 55점 이상인 종목을 모두 매수하고, 보유 종목 점수가 55점 이하로 내려가면 매도합니다. 보유 종목은 매일 점수 비율로 리밸런싱하며 회전 비용 0.2%를 반영합니다.",
    },
    "leader_custom": {
        "name": "주도주 점수 직접입력",
        "description": "매수/매도 점수와 TopN을 직접 입력해 주도주 리밸런싱 백테스트를 수행합니다.",
    },
}

LEADER_STRATEGY_PRESETS: dict[str, dict[str, float | int]] = {
    "leader_top10_score70": {"top_n": 10, "entry_threshold": 70.0, "exit_threshold": 70.0},
    "leader_top5_score80": {"top_n": 5, "entry_threshold": 80.0, "exit_threshold": 80.0},
    "leader_top5_score70": {"top_n": 5, "entry_threshold": 70.0, "exit_threshold": 70.0},
    "leader_all_score80": {"top_n": 10000, "entry_threshold": 80.0, "exit_threshold": 80.0},
    "leader_all_score70": {"top_n": 10000, "entry_threshold": 70.0, "exit_threshold": 70.0},
    "leader_all_score55": {"top_n": 10000, "entry_threshold": 55.0, "exit_threshold": 55.0},
}
TELEGRAM_EARNINGS_REPORT_INCLUDE_TOKENS = [
    "실적",
    "잠정",
    "손익구조",
    "영업(잠정)",
    "영업실적",
    "매출액또는손익구조",
    "연결재무제표기준영업",
    "재무제표기준영업",
    "사업보고서",
    "분기보고서",
    "반기보고서",
]
TELEGRAM_EARNINGS_REPORT_EXCLUDE_TOKENS = [
    "기업가치제고",
    "기업가치 제고",
    "밸류업",
    "value-up",
    "주주환원",
    "기업설명회",
    "ir",
    "투자설명",
    "경영현황",
    "공정공시대상정보",
]
TRADE_IMPORT_EXPORT_ITEMS = [
    {
        "key": "ramen",
        "name": "라면",
        "hs_codes": ["1902301010"],
        "companies": ["삼양식품", "농심"],
        "unit": "천달러",
        "sample_values": {
            "2021": [50000, 48000, 58000, 56000, 50000, 52000, 51000, 50000, 58000, 59000, 65000, 66000],
            "2022": [56000, 52000, 70000, 57000, 76000, 66000, 57000, 60000, 62000, 60000, 66000, 70000],
            "2023": [62000, 60000, 74000, 74000, 74000, 88000, 76000, 85000, 90000, 87000, 90000, 76000],
            "2024": [85000, 92000, 90000, 108000, 106000, 103000, 108000, 100000, 102000, 116000, 116000, 110000],
            "2025": [106000, 120000, 115000, 134000, 125000, 125000, 130000, 114000, 147000, 128000, 124000, 139000],
            "2026": [128000, 140000, 164000, 181000, None, None, None, None, None, None, None, None],
        },
    },
    {
        "key": "cosmetics",
        "name": "화장품",
        "hs_codes": ["3304"],
        "companies": [],
        "unit": "천달러",
        "sample_values": {
            "2021": [590000, 500000, 660000, 690000, 680000, 560000, 500000, 540000, 680000, 620000, 700000, 490000],
            "2022": [410000, 440000, 530000, 550000, 610000, 540000, 480000, 540000, 570000, 550000, 500000, 440000],
            "2023": [350000, 500000, 590000, 530000, 620000, 590000, 480000, 620000, 630000, 570000, 580000, 520000],
            "2024": [630000, 540000, 610000, 540000, 680000, 600000, 610000, 640000, 710000, 800000, 700000, 650000],
            "2025": [560000, 670000, 740000, 770000, 730000, 700000, 720000, 640000, 870000, 690000, 700000, 810000],
            "2026": [780000, 700000, 910000, 1040000, None, None, None, None, None, None, None, None],
        },
    },
    {"key": "dram_memory", "name": "D램/메모리반도체", "hs_codes": ["854232"], "companies": ["삼성전자", "SK하이닉스"], "unit": "천달러"},
    {"key": "flash_memory", "name": "플래시 메모리", "hs_codes": ["854232"], "companies": ["삼성전자", "SK하이닉스"], "unit": "천달러"},
    {"key": "gas_scrubber", "name": "스크러버/가스정화장비", "hs_codes": ["842139"], "companies": ["GST", "유니셈"], "unit": "천달러"},
    {"key": "dry_etch_equipment", "name": "건식각 장비", "hs_codes": ["845612", "848620"], "companies": ["브이엠"], "unit": "천달러"},
    {"key": "blank_mask", "name": "블랭크마스크/포토마스크", "hs_codes": ["370590", "370199"], "companies": ["에스앤에스텍"], "unit": "천달러"},
    {"key": "test_socket", "name": "러버소켓/테스트소켓", "hs_codes": ["853669", "853690"], "companies": ["ISC", "티에스이"], "unit": "천달러"},
    {"key": "pellicle", "name": "펠리클", "hs_codes": ["392690"], "companies": ["에프엔에스텍", "에프에스티"], "unit": "천달러"},
    {"key": "probe_card_board", "name": "프로브카드/인터페이스보드", "hs_codes": ["853400", "903090"], "companies": ["리노공업", "티에스이"], "unit": "천달러"},
    {"key": "plug_socket", "name": "플러그/소켓", "hs_codes": ["853669"], "companies": ["리노공업"], "unit": "천달러"},
    {"key": "pi_film", "name": "PI필름", "hs_codes": ["392099"], "companies": ["PI첨단소재"], "unit": "천달러"},
    {"key": "ccl", "name": "CCL/동박적층판", "hs_codes": ["741021", "741011"], "companies": ["두산"], "unit": "천달러"},
    {"key": "mlcc", "name": "MLCC", "hs_codes": ["853224"], "companies": ["삼성전기"], "unit": "천달러"},
    {"key": "printed_circuit", "name": "인쇄회로기판", "hs_codes": ["853400"], "companies": ["이수페타시스", "대덕전자", "심텍", "코리아써키트"], "unit": "천달러"},
    {"key": "lead_frame", "name": "리드프레임", "hs_codes": ["8541901000", "8542901010", "8542902010", "8542903010"], "companies": ["해성디에스"], "unit": "천달러"},
    {"key": "ess_battery", "name": "ESS/리튬이온 축전지", "hs_codes": ["850760"], "companies": ["삼성SDI", "LG에너지솔루션", "서진시스템"], "unit": "천달러"},
    {"key": "copper_foil", "name": "동박", "hs_codes": ["741011", "741021"], "companies": ["SKC", "롯데에너지머티리얼즈", "솔루스첨단소재"], "unit": "천달러"},
    {"key": "aluminum_foil", "name": "알루미늄박", "hs_codes": ["7607"], "companies": ["삼아알미늄", "롯데케미칼", "DI동일", "동원시스템즈"], "unit": "천달러"},
    {"key": "battery_pouch_film", "name": "알루미늄 파우치 필름", "hs_codes": ["760720", "392190"], "companies": ["율촌화학", "롯데알미늄", "동원시스템즈"], "unit": "천달러"},
    {"key": "battery_case", "name": "배터리 케이스/부품", "hs_codes": ["850790"], "companies": ["신성에스티"], "unit": "천달러"},
    {"key": "small_transformer", "name": "소형 변압기", "hs_codes": ["850431", "850432"], "companies": ["제룡전기", "산일전기"], "unit": "천달러"},
    {"key": "high_voltage_transformer", "name": "초고압 변압기", "hs_codes": ["850433", "850434"], "companies": ["HD현대일렉트릭", "효성중공업", "LS ELECTRIC", "일진전기"], "unit": "천달러"},
    {"key": "power_cable", "name": "전력선/케이블", "hs_codes": ["8544"], "companies": ["일진전기"], "unit": "천달러"},
    {"key": "circuit_breaker", "name": "자동차단기", "hs_codes": ["853620", "853529"], "companies": ["HD현대일렉트릭", "LS ELECTRIC"], "unit": "천달러"},
    {"key": "switchboard", "name": "배전반/고압배전반", "hs_codes": ["853710", "853720"], "companies": ["LS ELECTRIC", "HD현대일렉트릭", "일진전기"], "unit": "천달러"},
    {"key": "oligo", "name": "올리고/핵산계 원료", "hs_codes": ["293499"], "companies": ["에스티팜"], "unit": "천달러"},
    {"key": "health_supplement", "name": "건기식", "hs_codes": ["210690"], "companies": ["케어젠", "노바렉스", "서흥"], "unit": "천달러"},
    {"key": "wound_dressing", "name": "창상피복재", "hs_codes": ["300510", "300590"], "companies": ["티앤엘"], "unit": "천달러"},
    {"key": "hemostat", "name": "지혈제", "hs_codes": ["300610"], "companies": ["넥스트바이오메디컬"], "unit": "천달러"},
    {"key": "toxin", "name": "톡신", "hs_codes": ["300290", "300215"], "companies": ["휴젤", "대웅제약", "메디톡스"], "unit": "천달러"},
    {"key": "filler", "name": "필러", "hs_codes": ["300490", "330499"], "companies": ["휴젤", "메디톡스", "대웅제약", "휴메딕스", "파마리서치"], "unit": "천달러"},
    {"key": "mask_pack", "name": "마스크팩", "hs_codes": ["330499"], "companies": ["제닉", "에이피알"], "unit": "천달러"},
    {"key": "skin_care_cosmetics", "name": "기초화장품", "hs_codes": ["330499"], "companies": [], "unit": "천달러"},
    {"key": "makeup_cosmetics", "name": "메이크업용 화장품", "hs_codes": ["330410", "330420", "330491"], "companies": [], "unit": "천달러"},
    {"key": "dermatology_laser", "name": "피부과용 기기/레이저", "hs_codes": ["901890", "901320"], "companies": ["클래시스", "원텍", "이루다"], "unit": "천달러"},
    {"key": "home_beauty_device", "name": "가정용 미용기기", "hs_codes": ["854370", "901910"], "companies": ["에이피알", "파마리서치"], "unit": "천달러"},
    {"key": "boiler_water_heater", "name": "보일러/온수기", "hs_codes": ["840310", "841911", "841919", "851610"], "companies": ["경동나비엔", "귀뚜라미"], "unit": "천달러"},
    {"key": "epoxy", "name": "에폭시", "hs_codes": ["390730"], "companies": ["국도화학"], "unit": "천달러"},
    {"key": "pipe_fitting", "name": "피팅", "hs_codes": ["7307"], "companies": ["태광", "성광벤드"], "unit": "천달러"},
    {"key": "tobacco", "name": "담배", "hs_codes": ["2402"], "companies": ["KT&G"], "unit": "천달러"},
    {"key": "music_album", "name": "음반", "hs_codes": ["852349"], "companies": [], "unit": "천달러"},
]
TRADE_REGION_OPTIONS = [
    {"code": "all", "sido_code": "", "name": "전국"},
    {"code": "11000", "sido_code": "11", "name": "서울"},
    {"code": "41000", "sido_code": "41", "name": "경기"},
    {"code": "28000", "sido_code": "28", "name": "인천"},
    {"code": "26000", "sido_code": "26", "name": "부산"},
    {"code": "31000", "sido_code": "31", "name": "울산"},
    {"code": "27000", "sido_code": "27", "name": "대구"},
    {"code": "29000", "sido_code": "29", "name": "광주"},
    {"code": "30000", "sido_code": "30", "name": "대전"},
    {"code": "36000", "sido_code": "36", "name": "세종"},
    {"code": "42000", "sido_code": "42", "name": "강원"},
    {"code": "43000", "sido_code": "43", "name": "충북"},
    {"code": "44000", "sido_code": "44", "name": "충남"},
    {"code": "45000", "sido_code": "45", "name": "전북"},
    {"code": "46000", "sido_code": "46", "name": "전남"},
    {"code": "47000", "sido_code": "47", "name": "경북"},
    {"code": "48000", "sido_code": "48", "name": "경남"},
    {"code": "50000", "sido_code": "50", "name": "제주"},
]
MARKET_YTD_UNIVERSE_LIMIT = 300
STRATEGY_TRADE_FEE_RATE = 0.002
STRATEGY_MDD_STOP_RATE = 0.07


def normalize_text(value: Any) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return ""
    return re.sub(r"\s+", "", str(value).strip()).lower()


def to_float(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_korean_number(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def safe_copy_to_temp(source_path: Path) -> Path:
    temp_dir = Path(tempfile.gettempdir()) / "stock_dashboard_cache"
    temp_dir.mkdir(parents=True, exist_ok=True)
    temp_path = temp_dir / f"{source_path.stem}_{uuid.uuid4().hex[:8]}{source_path.suffix}"
    last_error: Exception | None = None
    for _ in range(8):
        try:
            shutil.copy2(source_path, temp_path)
            return temp_path
        except PermissionError as exc:
            last_error = exc
            time.sleep(0.4)
        except Exception:
            raise
    if last_error is not None:
        raise PermissionError(
            f"엑셀 파일이 다른 프로세스(Excel/동기화)에서 잠겨 읽을 수 없습니다. "
            f"파일을 닫고 다시 시도하세요: {source_path}"
        ) from last_error
    return temp_path


def excel_engine_for_path(path: Path) -> str | None:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return "openpyxl"
    if suffix == ".xls":
        return "xlrd"
    return None


def sanitize_filename(value: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", value).strip(" .")
    return cleaned or "attachment"


def parse_date_label(value: Any) -> datetime | None:
    if value is None or pd.isna(value):
        return None
    match = re.search(r"(\d{4})[./-](\d{2})[./-](\d{2})", str(value))
    if not match:
        return None
    return datetime.strptime("-".join(match.groups()), "%Y-%m-%d")


def parse_iso_datetime(value: Any) -> datetime | None:
    try:
        if not value:
            return None
        parsed = datetime.fromisoformat(str(value))
        if parsed.tzinfo is not None:
            parsed = parsed.astimezone().replace(tzinfo=None)
        return parsed
    except Exception:
        return None


def load_settings() -> dict[str, Any]:
    if SETTINGS_PATH.exists():
        return json.loads(SETTINGS_PATH.read_text(encoding="utf-8-sig"))
    return {}


def save_settings(settings: dict[str, Any]) -> None:
    SETTINGS_PATH.parent.mkdir(parents=True, exist_ok=True)
    SETTINGS_PATH.write_text(json.dumps(settings, ensure_ascii=False, indent=2), encoding="utf-8")


def load_global_company_ai_cache() -> dict[str, Any]:
    if not GLOBAL_COMPANY_AI_CACHE_PATH.exists():
        return {"items": {}, "loaded_at": ""}
    try:
        payload = json.loads(GLOBAL_COMPANY_AI_CACHE_PATH.read_text(encoding="utf-8"))
        return {
            "items": payload.get("items", {}) if isinstance(payload.get("items"), dict) else {},
            "loaded_at": str(payload.get("loaded_at", "")),
        }
    except Exception:
        return {"items": {}, "loaded_at": ""}


def save_global_company_ai_cache(cache: dict[str, Any]) -> dict[str, Any]:
    cache["loaded_at"] = datetime.now().isoformat(timespec="seconds")
    temp_path = GLOBAL_COMPANY_AI_CACHE_PATH.with_name(f"{GLOBAL_COMPANY_AI_CACHE_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
    temp_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(GLOBAL_COMPANY_AI_CACHE_PATH)
    return cache


def get_openai_api_key() -> str:
    settings = load_settings()
    public_data = settings.get("public_data") if isinstance(settings.get("public_data"), dict) else {}
    return (
        str(os.getenv("OPENAI_API_KEY", "")).strip()
        or str(settings.get("openai_api_key") or "").strip()
        or str(settings.get("openai_key") or "").strip()
        or str(public_data.get("openai_api_key") or "").strip()
        or str(public_data.get("openai_key") or "").strip()
    )


def krx_credentials_signature(user_id: str, password: str) -> str:
    return hashlib.sha1(f"{str(user_id).strip()}::{str(password).strip()}".encode("utf-8")).hexdigest()


def should_attempt_krx_login(user_id: str, password: str) -> bool:
    user_id = str(user_id or "").strip()
    password = str(password or "").strip()
    if not user_id or not password:
        return False
    settings = load_settings()
    state = settings.get("krx_login_state", {}) if isinstance(settings.get("krx_login_state"), dict) else {}
    failed_signature = str(state.get("failed_signature") or "").strip()
    return failed_signature != krx_credentials_signature(user_id, password)


def record_krx_login_result(user_id: str, password: str, success: bool, error: str = "") -> None:
    user_id = str(user_id or "").strip()
    password = str(password or "").strip()
    settings = load_settings()
    if success:
        if "krx_login_state" in settings:
            settings.pop("krx_login_state", None)
            save_settings(settings)
        return
    if not user_id or not password:
        return
    settings["krx_login_state"] = {
        "failed_signature": krx_credentials_signature(user_id, password),
        "failed_at": datetime.now().isoformat(timespec="seconds"),
        "last_error": str(error or "").strip()[:500],
    }
    save_settings(settings)


def get_krx_settings() -> dict[str, str]:
    settings = load_settings()
    krx = settings.get("krx", {}) if isinstance(settings.get("krx"), dict) else {}
    user_id = str(os.getenv("KRX_ID", "") or krx.get("id", "")).strip()
    password = str(os.getenv("KRX_PW", "") or krx.get("password", "")).strip()
    if user_id and password and not should_attempt_krx_login(user_id, password):
        return {"id": "", "password": ""}
    return {"id": user_id, "password": password}


DEFAULT_MARKET_CALENDAR_EVENTS: list[dict[str, Any]] = [
    {"date": "2026-05-25", "title": "한국 증시 휴장", "category": "휴장", "market": "KR", "importance": "high", "note": "사용자 입력 일정"},
    {"date": "2026-05-25", "title": "미국 증시 휴장", "category": "휴장", "market": "US", "importance": "high", "note": "사용자 입력 일정"},
    {"date": "2026-05-26", "title": "미국 5월 소비자신뢰지수", "category": "경제지표", "market": "US", "importance": "high", "note": "Conference Board Consumer Confidence"},
    {"date": "2026-05-27", "title": "삼성전자·SK하이닉스 단일종목 레버리지 ETF 상장", "category": "상장/ETF", "market": "KR", "importance": "medium"},
    {"date": "2026-05-27", "title": "마벨 실적 발표", "category": "실적", "market": "US", "time": "장후", "importance": "medium"},
    {"date": "2026-05-27", "title": "세일즈포스 실적 발표", "category": "실적", "market": "US", "time": "장후", "importance": "medium"},
    {"date": "2026-05-27", "title": "Snowflake 실적 발표", "category": "실적", "market": "US", "time": "장후", "importance": "medium"},
    {"date": "2026-05-27", "title": "Synopsys 실적 발표", "category": "실적", "market": "US", "time": "장후", "importance": "medium"},
    {"date": "2026-05-28", "title": "한국은행 금통위 기준금리 결정", "category": "중앙은행", "market": "KR", "importance": "high"},
    {"date": "2026-05-28", "title": "미국 1분기 GDP", "category": "경제지표", "market": "US", "importance": "high"},
    {"date": "2026-05-28", "title": "미국 4월 PCE", "category": "경제지표", "market": "US", "importance": "high"},
    {"date": "2026-05-28", "title": "코스트코 실적 발표", "category": "실적", "market": "US", "time": "장후", "importance": "medium"},
    {"date": "2026-05-28", "title": "델 실적 발표", "category": "실적", "market": "US", "time": "장후", "importance": "medium"},
    {"date": "2026-05-28", "title": "MongoDB 실적 발표", "category": "실적", "market": "US", "time": "장후", "importance": "medium"},
    {"date": "2026-05-29", "title": "미국 4월 상품 무역수지", "category": "경제지표", "market": "US", "importance": "medium"},
]


def normalize_market_calendar_event(raw: dict[str, Any]) -> dict[str, Any] | None:
    try:
        date_text = datetime.strptime(str(raw.get("date") or "")[:10], "%Y-%m-%d").date().isoformat()
    except Exception:
        return None
    title = str(raw.get("title") or "").strip()
    if not title:
        return None
    event_id = str(raw.get("id") or "").strip() or hashlib.sha1(
        f"{date_text}|{title}|{raw.get('market') or ''}|{raw.get('time') or ''}".encode("utf-8")
    ).hexdigest()[:16]
    return {
        "id": event_id,
        "date": date_text,
        "title": title,
        "category": str(raw.get("category") or "기타").strip() or "기타",
        "market": str(raw.get("market") or "").strip(),
        "time": str(raw.get("time") or "").strip(),
        "importance": str(raw.get("importance") or "medium").strip() or "medium",
        "note": str(raw.get("note") or "").strip(),
        "source": str(raw.get("source") or "manual").strip() or "manual",
        "url": str(raw.get("url") or "").strip(),
    }


def load_market_calendar_events() -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    if MARKET_CALENDAR_PATH.exists():
        try:
            payload = json.loads(MARKET_CALENDAR_PATH.read_text(encoding="utf-8"))
            source = payload.get("events", []) if isinstance(payload, dict) else payload
            for raw in source if isinstance(source, list) else []:
                if isinstance(raw, dict):
                    normalized = normalize_market_calendar_event(raw)
                    if normalized:
                        events.append(normalized)
        except Exception:
            events = []
    if not events:
        events = [item for item in (normalize_market_calendar_event({**event, "source": "manual_seed"}) for event in DEFAULT_MARKET_CALENDAR_EVENTS) if item]
        save_market_calendar_events(events)
    return sorted(events, key=lambda item: (item.get("date", ""), item.get("time", ""), item.get("title", "")))


def save_market_calendar_events(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized = [item for item in (normalize_market_calendar_event(event) for event in events) if item]
    MARKET_CALENDAR_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_path = MARKET_CALENDAR_PATH.with_name(f"{MARKET_CALENDAR_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
    temp_path.write_text(json.dumps({"events": normalized}, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(MARKET_CALENDAR_PATH)
    return normalized


def market_calendar_event_key(event: dict[str, Any]) -> str:
    return "|".join(
        [
            str(event.get("date") or ""),
            normalize_text(event.get("title")),
            str(event.get("market") or ""),
            str(event.get("time") or ""),
            str(event.get("source") or ""),
        ]
    )


def market_calendar_date_range(start_date: date, end_date: date) -> list[date]:
    days = max(0, min((end_date - start_date).days, 75))
    return [start_date + timedelta(days=offset) for offset in range(days + 1)]


def kind_stock_code_from_onclick(onclick: str) -> str:
    match = re.search(r"(?:companysummary_open|openDisclsChart|fnPopStockPrices)\('([^']+)'", str(onclick or ""))
    if not match:
        return ""
    raw_code = re.sub(r"\D", "", match.group(1))
    if len(raw_code) == 5:
        return raw_code + "0"
    return raw_code.zfill(6) if raw_code else ""


def kind_disclosure_category(title: str) -> str:
    text = normalize_text(title)
    if any(token in text for token in ["잠정실적", "영업실적", "매출액", "영업이익", "분기보고서", "반기보고서", "사업보고서"]):
        return "실적/보고서"
    if any(token in text for token in ["단일판매", "공급계약", "수주", "계약체결"]):
        return "수주/계약"
    if any(token in text for token in ["시설투자", "신규시설", "타법인", "취득", "양수", "증설"]):
        return "투자"
    if any(token in text for token in ["유상증자", "무상증자", "전환사채", "신주인수권", "자금조달"]):
        return "자금조달"
    if any(token in text for token in ["합병", "분할", "영업양수", "영업양도"]):
        return "합병/분할"
    if any(token in text for token in ["배당", "자기주식", "자사주", "주식소각"]):
        return "주주환원"
    if any(token in text for token in ["주주총회", "기준일", "명의개서"]):
        return "주주일정"
    if any(token in text for token in ["투자주의", "투자경고", "단기과열", "거래정지", "관리종목"]):
        return "시장조치"
    return "공시"


def kind_disclosure_importance(title: str, marcap: float | None) -> str:
    text = normalize_text(title)
    if any(token in text for token in ["잠정실적", "영업실적", "단일판매", "공급계약", "시설투자", "합병", "분할", "유상증자", "거래정지"]):
        return "high"
    if marcap and marcap >= 10000 * 100000000:
        return "high"
    if any(token in text for token in ["투자주의", "단기과열"]):
        return "low"
    return "medium"


def fetch_kind_disclosures_for_date(target_date: date, min_marcap: float = MARKET_CALENDAR_MIN_KR_MARCAP) -> list[dict[str, Any]]:
    response = requests.post(
        "https://kind.krx.co.kr/disclosure/todaydisclosure.do?method=searchTodayDisclosureSub",
        data={
            "currentPageSize": "100",
            "pageIndex": "1",
            "orderMode": "0",
            "orderStat": "D",
            "forward": "todaydisclosure_sub",
            "todayFlag": "N",
            "selDate": target_date.isoformat(),
        },
        headers={
            "User-Agent": news_headers()["User-Agent"],
            "Referer": "https://kind.krx.co.kr/disclosure/todaydisclosure.do?method=searchTodayDisclosureMain",
        },
        timeout=15,
    )
    response.raise_for_status()
    if not response.text.strip():
        return []
    soup = BeautifulSoup(response.text, "html.parser")
    events: list[dict[str, Any]] = []
    for row in soup.select("tbody tr"):
        cells = row.select("td")
        if len(cells) < 3:
            continue
        time_text = clean_news_text(cells[0].get_text(" "))
        company_link = cells[1].select_one("a")
        title_link = cells[2].select_one("a")
        company = clean_news_text(company_link.get_text(" ") if company_link else cells[1].get_text(" "))
        title = clean_news_text(title_link.get_text(" ") if title_link else cells[2].get_text(" "))
        if not company or not title:
            continue
        code = kind_stock_code_from_onclick(company_link.get("onclick") if company_link else "")
        if not code and len(cells) >= 5:
            for link in cells[-1].select("a"):
                code = kind_stock_code_from_onclick(link.get("onclick") or "")
                if code:
                    break
        listing_row = find_listing_row_by_code(code) if code else resolve_stock_payload(name=company)
        marcap = to_float((listing_row or {}).get("marcap"))
        if marcap is None or marcap < min_marcap:
            continue
        accept_no = ""
        if title_link:
            match = re.search(r"openDisclsViewer\('([^']+)'", str(title_link.get("onclick") or ""))
            accept_no = match.group(1) if match else ""
        source_url = (
            "https://kind.krx.co.kr/common/disclsviewer.do?method=search&acptno=" + quote(accept_no)
            if accept_no
            else "https://kind.krx.co.kr/disclosure/todaydisclosure.do?method=searchTodayDisclosureMain"
        )
        events.append(
            {
                "date": target_date.isoformat(),
                "time": time_text,
                "title": f"{company} · {title}",
                "category": kind_disclosure_category(title),
                "market": "KR",
                "importance": kind_disclosure_importance(title, marcap),
                "note": f"KIND 공시 · 시총 {marcap / 100000000:.0f}억원 이상",
                "source": "KIND",
                "url": source_url,
                "id": hashlib.sha1(f"KIND|{target_date}|{company}|{title}|{accept_no}".encode("utf-8")).hexdigest()[:16],
            }
        )
    return events


def investing_country_to_market(country_text: str, currency_text: str) -> str:
    source = f"{country_text} {currency_text}".upper()
    if "UNITED STATES" in source or "USD" in source:
        return "US"
    if "SOUTH KOREA" in source or "KRW" in source:
        return "KR"
    if "CHINA" in source or "CNY" in source:
        return "CN"
    if "JAPAN" in source or "JPY" in source:
        return "JP"
    if "EUR" in source or "EURO" in source:
        return "EU"
    return "Global"


def fetch_investing_calendar(start_date: date, end_date: date) -> list[dict[str, Any]]:
    response = requests.post(
        "https://www.investing.com/economic-calendar/Service/getCalendarFilteredData",
        data={
            "country[]": ["5", "4", "35", "11", "25", "32", "6", "72"],
            "importance[]": ["3"],
            "timeZone": "8",
            "timeFilter": "timeRemain",
            "currentTab": "custom",
            "limit_from": "0",
            "dateFrom": start_date.isoformat(),
            "dateTo": end_date.isoformat(),
        },
        headers={
            "User-Agent": news_headers()["User-Agent"],
            "X-Requested-With": "XMLHttpRequest",
            "Referer": "https://www.investing.com/economic-calendar/",
        },
        timeout=20,
    )
    response.raise_for_status()
    payload = response.json()
    soup = BeautifulSoup(str(payload.get("data") or ""), "html.parser")
    events: list[dict[str, Any]] = []
    current_date: date | None = None
    for row in soup.select("tr"):
        day_cell = row.select_one("td.theDay")
        if day_cell:
            try:
                current_date = datetime.strptime(clean_news_text(day_cell.get_text(" ")), "%A, %B %d, %Y").date()
            except Exception:
                current_date = None
            continue
        if current_date is None:
            continue
        cells = row.select("td")
        if len(cells) < 4:
            continue
        time_text = clean_news_text(cells[0].get_text(" "))
        currency = clean_news_text(cells[1].get_text(" "))
        impact = clean_news_text(cells[2].get("title") or cells[2].get_text(" "))
        event_cell = cells[3]
        title = clean_news_text(event_cell.get_text(" "))
        if not title:
            continue
        country_span = cells[1].select_one("[title]")
        country = clean_news_text(country_span.get("title") if country_span else "")
        market = investing_country_to_market(country, currency)
        event_id = str(row.get("id") or "").replace("eventRowId_", "")
        events.append(
            {
                "date": current_date.isoformat(),
                "time": "" if time_text.lower() == "all day" else time_text,
                "title": title,
                "category": "휴장" if "holiday" in normalize_text(title + impact) else "경제지표",
                "market": market,
                "importance": "high",
                "note": "Investing.com economic calendar" + (f" · {currency}" if currency else ""),
                "source": "Investing.com",
                "url": "https://www.investing.com/economic-calendar/",
                "id": hashlib.sha1(f"INVESTING|{current_date}|{event_id}|{title}|{market}".encode("utf-8")).hexdigest()[:16],
            }
        )
    return events


def load_market_calendar_auto_cache() -> dict[str, Any]:
    if not MARKET_CALENDAR_AUTO_CACHE_PATH.exists():
        return {"events": [], "ranges": {}, "errors": []}
    try:
        payload = json.loads(MARKET_CALENDAR_AUTO_CACHE_PATH.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else {"events": [], "ranges": {}, "errors": []}
    except Exception:
        return {"events": [], "ranges": {}, "errors": []}


def save_market_calendar_auto_cache(payload: dict[str, Any]) -> dict[str, Any]:
    MARKET_CALENDAR_AUTO_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    temp_path = MARKET_CALENDAR_AUTO_CACHE_PATH.with_name(f"{MARKET_CALENDAR_AUTO_CACHE_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(MARKET_CALENDAR_AUTO_CACHE_PATH)
    return payload


def refresh_market_calendar_auto_events(start_date: date, end_date: date, force: bool = False) -> dict[str, Any]:
    cache = load_market_calendar_auto_cache()
    range_key = f"{start_date.isoformat()}:{end_date.isoformat()}"
    ranges = cache.get("ranges") if isinstance(cache.get("ranges"), dict) else {}
    range_meta = ranges.get(range_key) if isinstance(ranges.get(range_key), dict) else {}
    fetched_at = parse_iso_datetime(range_meta.get("fetched_at")) if range_meta else None
    if (
        not force
        and fetched_at
        and (datetime.now() - fetched_at).total_seconds() < MARKET_CALENDAR_AUTO_CACHE_TTL_SECONDS
    ):
        return cache

    errors: list[str] = []
    fetched_events: list[dict[str, Any]] = []
    try:
        kind_end_date = min(end_date, date.today())
        if start_date <= kind_end_date:
            for target_date in market_calendar_date_range(start_date, kind_end_date):
                fetched_events.extend(fetch_kind_disclosures_for_date(target_date))
    except Exception as exc:
        errors.append(f"KIND: {exc}")
    try:
        if start_date <= end_date:
            fetched_events.extend(fetch_investing_calendar(start_date, end_date))
    except Exception as exc:
        errors.append(f"Investing.com: {exc}")

    normalized_new = [item for item in (normalize_market_calendar_event(event) for event in fetched_events) if item]
    existing = [
        item for item in (normalize_market_calendar_event(event) for event in cache.get("events", []))
        if item and not (start_date.isoformat() <= str(item.get("date") or "") <= end_date.isoformat())
    ]
    merged_by_key: dict[str, dict[str, Any]] = {}
    for event in existing + normalized_new:
        merged_by_key[market_calendar_event_key(event)] = event
    ranges[range_key] = {
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "event_count": len(normalized_new),
        "errors": errors,
    }
    payload = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "events": sorted(merged_by_key.values(), key=lambda item: (item.get("date", ""), item.get("time", ""), item.get("title", ""))),
        "ranges": ranges,
        "errors": errors,
    }
    return save_market_calendar_auto_cache(payload)


def market_calendar_payload(start: str | None = None, end: str | None = None, refresh: bool = False) -> dict[str, Any]:
    today = date.today()
    try:
        start_date = datetime.strptime(str(start or ""), "%Y-%m-%d").date()
    except Exception:
        start_date = today.replace(day=1)
    try:
        end_date = datetime.strptime(str(end or ""), "%Y-%m-%d").date()
    except Exception:
        end_date = start_date + timedelta(days=45)
    auto_cache = refresh_market_calendar_auto_events(start_date, end_date, force=refresh)
    manual_events = load_market_calendar_events()
    auto_events = [
        item for item in (normalize_market_calendar_event(event) for event in auto_cache.get("events", []))
        if item
    ]
    merged_by_key: dict[str, dict[str, Any]] = {}
    for event in manual_events + auto_events:
        merged_by_key[market_calendar_event_key(event)] = event
    events = [
        event for event in merged_by_key.values()
        if start_date.isoformat() <= str(event.get("date") or "") <= end_date.isoformat()
    ]
    events.sort(key=lambda item: (item.get("date", ""), item.get("time", ""), item.get("importance") != "high", item.get("title", "")))
    by_date: dict[str, list[dict[str, Any]]] = {}
    for event in events:
        by_date.setdefault(str(event.get("date")), []).append(event)
    return {
        "start": start_date.isoformat(),
        "end": end_date.isoformat(),
        "events": events,
        "by_date": by_date,
        "auto": {
            "updated_at": auto_cache.get("updated_at") or "",
            "event_count": len(auto_events),
            "errors": auto_cache.get("errors") or [],
            "min_kr_market_cap_100m": round(MARKET_CALENDAR_MIN_KR_MARCAP / 100000000.0, 0),
        },
        "sources": [
            {
                "name": "KIND",
                "type": "국내 기업 공시",
                "note": "시가총액 2000억원 이상 기업 공시만 자동 반영",
                "url": "https://kind.krx.co.kr/disclosure/todaydisclosure.do?method=searchTodayDisclosureMain",
            },
            {
                "name": "Investing.com",
                "type": "해외 주요 경제 일정",
                "note": "High importance 경제지표와 휴장 일정 자동 반영",
                "url": "https://www.investing.com/economic-calendar/",
            },
            {
                "name": "수동 일정",
                "type": "직접 추가",
                "note": "자동 소스에서 빠진 ETF 상장, 내부 체크포인트 보강",
                "url": "/api/market-calendar.ics",
            },
        ],
    }


def escape_ics_text(value: Any) -> str:
    return str(value or "").replace("\\", "\\\\").replace(";", "\\;").replace(",", "\\,").replace("\n", "\\n")


def build_market_calendar_ics(events: list[dict[str, Any]]) -> str:
    stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Stock Dashboard//Market Calendar//KO",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "X-WR-CALNAME:Stock Dashboard 증시 일정",
    ]
    for event in events:
        try:
            event_date = datetime.strptime(str(event.get("date") or ""), "%Y-%m-%d").date()
        except Exception:
            continue
        next_date = event_date + timedelta(days=1)
        uid = f"{event.get('id')}@stock-dashboard.local"
        description = " / ".join(
            part for part in [
                event.get("category"),
                event.get("market"),
                event.get("time"),
                event.get("note"),
                event.get("source"),
            ] if part
        )
        lines.extend(
            [
                "BEGIN:VEVENT",
                f"UID:{escape_ics_text(uid)}",
                f"DTSTAMP:{stamp}",
                f"DTSTART;VALUE=DATE:{event_date.strftime('%Y%m%d')}",
                f"DTEND;VALUE=DATE:{next_date.strftime('%Y%m%d')}",
                f"SUMMARY:{escape_ics_text(event.get('title'))}",
                f"DESCRIPTION:{escape_ics_text(description)}",
                "END:VEVENT",
            ]
        )
    lines.append("END:VCALENDAR")
    return "\r\n".join(lines) + "\r\n"


def normalize_sector_db(data: Any) -> dict[str, Any]:
    source = data if isinstance(data, dict) else {}
    stock_map: dict[str, dict[str, str]] = {}
    for key, item in (source.get("stock_map") or {}).items():
        if not isinstance(item, dict):
            continue
        stock_name = str(item.get("stock_name") or "").strip()
        sector = str(item.get("sector") or "").strip()
        stock_code = str(item.get("stock_code") or "").strip()
        item_key = str(key or "").strip() or stock_code or normalize_text(stock_name)
        if stock_name and sector and item_key:
            stock_map[item_key] = {
                "stock_code": stock_code,
                "stock_name": stock_name,
                "sector": sector,
            }
    sector_set = {item["sector"] for item in stock_map.values() if item.get("sector")}
    sectors: list[str] = []
    for sector in source.get("sectors") or source.get("sector_order") or []:
        sector_name = str(sector or "").strip()
        if sector_name and sector_name in sector_set and sector_name not in sectors:
            sectors.append(sector_name)
    for sector in sorted(sector_set):
        if sector not in sectors:
            sectors.append(sector)
    source_stock_order = source.get("stock_order") if isinstance(source.get("stock_order"), dict) else {}
    stock_order: dict[str, list[str]] = {}
    for sector in sectors:
        sector_codes = [
            str(item.get("stock_code") or key or "").strip()
            for key, item in stock_map.items()
            if str(item.get("sector") or "").strip() == sector
        ]
        ordered: list[str] = []
        for code in source_stock_order.get(sector, []) if isinstance(source_stock_order.get(sector), list) else []:
            code_text = str(code or "").strip()
            if code_text and code_text in sector_codes and code_text not in ordered:
                ordered.append(code_text)
        for code in sector_codes:
            if code and code not in ordered:
                ordered.append(code)
        if ordered:
            stock_order[sector] = ordered
    return {"stock_map": stock_map, "sectors": sectors, "stock_order": stock_order}


def load_sector_db() -> dict[str, Any]:
    if not SECTOR_DB_PATH.exists():
        return {"stock_map": {}, "sectors": [], "stock_order": {}}
    try:
        return normalize_sector_db(json.loads(SECTOR_DB_PATH.read_text(encoding="utf-8")))
    except Exception:
        return {"stock_map": {}, "sectors": [], "stock_order": {}}


def backup_sector_db() -> None:
    if not SECTOR_DB_PATH.exists():
        return
    try:
        current = normalize_sector_db(json.loads(SECTOR_DB_PATH.read_text(encoding="utf-8")))
    except Exception:
        return
    if not current.get("stock_map") and not current.get("sectors"):
        return
    SECTOR_DB_BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    backup_path = SECTOR_DB_BACKUP_DIR / f"sector_database_{stamp}.json"
    backup_path.write_text(json.dumps(current, ensure_ascii=False, indent=2), encoding="utf-8")


def save_sector_db(data: dict[str, Any]) -> dict[str, Any]:
    normalized = normalize_sector_db(data)
    backup_sector_db()
    SECTOR_DB_PATH.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
    return normalized


def sector_db_key(stock_code: str | None, stock_name: str | None) -> str:
    code = str(stock_code or "").strip()
    if code:
        return code
    return normalize_text(str(stock_name or "").strip())


def resolve_sector_for_stock(stock_code: str | None, stock_name: str | None, db: dict[str, Any] | None = None) -> str:
    sector_db = db or load_sector_db()
    stock_map = sector_db.get("stock_map", {})
    keys = [sector_db_key(stock_code, stock_name), normalize_text(str(stock_name or "").strip())]
    for key in keys:
        if key and key in stock_map:
            return str(stock_map[key].get("sector") or "").strip()
    normalized_name = normalize_text(str(stock_name or "").strip())
    if normalized_name:
        for item in stock_map.values():
            if normalize_text(item.get("stock_name", "")) == normalized_name:
                return str(item.get("sector") or "").strip()
    return ""


def load_screening_cache() -> dict[str, Any]:
    if not SCREENING_CACHE_PATH.exists():
        return {"summaries": {}, "recent_leaders": {}, "loaded_at": ""}
    try:
        payload = json.loads(SCREENING_CACHE_PATH.read_text(encoding="utf-8"))
        return {
            "summaries": payload.get("summaries", {}) if isinstance(payload.get("summaries"), dict) else {},
            "recent_leaders": payload.get("recent_leaders", {}) if isinstance(payload.get("recent_leaders"), dict) else {},
            "calendar": payload.get("calendar", {}) if isinstance(payload.get("calendar"), dict) else {},
            "loaded_at": str(payload.get("loaded_at", "")),
        }
    except Exception:
        return {"summaries": {}, "recent_leaders": {}, "loaded_at": ""}


def save_screening_cache(cache: dict[str, Any]) -> dict[str, Any]:
    cache["loaded_at"] = datetime.now().isoformat(timespec="seconds")
    temp_path = SCREENING_CACHE_PATH.with_name(f"{SCREENING_CACHE_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
    temp_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(SCREENING_CACHE_PATH)
    return cache


def load_us_screening_cache() -> dict[str, Any]:
    if not US_SCREENING_CACHE_PATH.exists():
        return {"summaries": {}, "recent_leaders": {}, "loaded_at": ""}
    try:
        payload = json.loads(US_SCREENING_CACHE_PATH.read_text(encoding="utf-8"))
        return {
            "summaries": payload.get("summaries", {}) if isinstance(payload.get("summaries"), dict) else {},
            "recent_leaders": payload.get("recent_leaders", {}) if isinstance(payload.get("recent_leaders"), dict) else {},
            "calendar": payload.get("calendar", {}) if isinstance(payload.get("calendar"), dict) else {},
            "loaded_at": str(payload.get("loaded_at", "")),
        }
    except Exception:
        return {"summaries": {}, "recent_leaders": {}, "loaded_at": ""}


def save_us_screening_cache(cache: dict[str, Any]) -> dict[str, Any]:
    cache["loaded_at"] = datetime.now().isoformat(timespec="seconds")
    temp_path = US_SCREENING_CACHE_PATH.with_name(f"{US_SCREENING_CACHE_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
    temp_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(US_SCREENING_CACHE_PATH)
    return cache


def load_asia_screening_cache() -> dict[str, Any]:
    if not ASIA_SCREENING_CACHE_PATH.exists():
        return {"summaries": {}, "recent_leaders": {}, "loaded_at": ""}
    try:
        payload = json.loads(ASIA_SCREENING_CACHE_PATH.read_text(encoding="utf-8"))
        return {
            "summaries": payload.get("summaries", {}) if isinstance(payload.get("summaries"), dict) else {},
            "recent_leaders": payload.get("recent_leaders", {}) if isinstance(payload.get("recent_leaders"), dict) else {},
            "calendar": payload.get("calendar", {}) if isinstance(payload.get("calendar"), dict) else {},
            "loaded_at": str(payload.get("loaded_at", "")),
        }
    except Exception:
        return {"summaries": {}, "recent_leaders": {}, "loaded_at": ""}


def save_asia_screening_cache(cache: dict[str, Any]) -> dict[str, Any]:
    cache["loaded_at"] = datetime.now().isoformat(timespec="seconds")
    temp_path = ASIA_SCREENING_CACHE_PATH.with_name(f"{ASIA_SCREENING_CACHE_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
    temp_path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    temp_path.replace(ASIA_SCREENING_CACHE_PATH)
    return cache


def screening_data_version_token() -> str:
    # 점수 재계산(SQL 갱신) 시 DB mtime이 바뀌므로 캐시 키를 자동 무효화한다.
    try:
        if SCREENING_FAST_DB_PATH.exists():
            return f"db{int(SCREENING_FAST_DB_PATH.stat().st_mtime)}"
    except Exception:
        pass
    try:
        files = sorted(
            (p for p in SCREENING_DIR.glob("*.xls*") if re.match(r"^(20\d{6})_", p.name)),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if files:
            return f"xlsx{int(files[0].stat().st_mtime)}"
    except Exception:
        pass
    return "na"


def screening_cache_key(file_date: str, min_score: float) -> str:
    token = screening_data_version_token()
    return f"{SCREENING_SCORE_CACHE_VERSION}|{token}|{file_date}|{float(min_score):.4f}"


def screening_summary_payload_cache_key(file_date: str, min_score: float, recent_limit: int) -> str:
    token = screening_data_version_token()
    return f"{SCREENING_SCORE_CACHE_VERSION}|payload|{token}|{file_date}|{float(min_score):.4f}|{int(recent_limit)}"


def parse_screening_cache_key(key: Any) -> tuple[str, str]:
    parts = str(key).split("|")
    if len(parts) == 4 and parts[0] == SCREENING_SCORE_CACHE_VERSION:
        return parts[2], parts[3]
    if len(parts) == 3 and parts[0] == SCREENING_SCORE_CACHE_VERSION:
        # legacy format fallback
        return parts[1], parts[2]
    return "", ""


def get_dart_api_key() -> str:
    return os.getenv("DART_API_KEY", "").strip() or str(load_settings().get("dart_api_key", "")).strip()


def get_kis_settings(environment: str | None = None) -> dict[str, Any]:
    settings = load_settings()
    kis = settings.get("kis", {}) if isinstance(settings.get("kis"), dict) else {}
    target_environment = str(
        environment or os.getenv("KIS_ENVIRONMENT", "") or kis.get("active_environment", "") or kis.get("environment", "mock")
    ).strip().lower()
    if target_environment not in {"mock", "real"}:
        target_environment = "mock"
    profile = kis.get(target_environment, {}) if isinstance(kis.get(target_environment), dict) else {}
    app_key = str(os.getenv("KIS_APP_KEY", "") or profile.get("app_key", "") or kis.get("app_key", "")).strip()
    app_secret = str(os.getenv("KIS_APP_SECRET", "") or profile.get("app_secret", "") or kis.get("app_secret", "")).strip()
    account_no = str(os.getenv("KIS_ACCOUNT_NO", "") or profile.get("account_no", "") or kis.get("account_no", "")).strip()
    account_product_code = str(
        os.getenv("KIS_ACCOUNT_PRODUCT_CODE", "") or profile.get("account_product_code", "") or kis.get("account_product_code", "")
    ).strip()
    return {
        "environment": target_environment,
        "app_key": app_key,
        "app_secret": app_secret,
        "account_no": account_no,
        "account_product_code": account_product_code,
        "base_url": KIS_MOCK_BASE_URL if target_environment == "mock" else KIS_REAL_BASE_URL,
    }


def mask_secret(value: str, visible: int = 4) -> str:
    text = str(value or "")
    if not text:
        return ""
    if len(text) <= visible * 2:
        return "*" * len(text)
    return text[:visible] + "*" * max(4, len(text) - visible * 2) + text[-visible:]


def get_kis_access_token(force_refresh: bool = False) -> dict[str, Any]:
    kis = get_kis_settings()
    if not kis["app_key"] or not kis["app_secret"]:
        raise ValueError("한국투자증권 API Key/Secret이 설정되어 있지 않습니다.")

    now = datetime.now()
    if not force_refresh and KIS_TOKEN_CACHE_PATH.exists():
        try:
            cached = json.loads(KIS_TOKEN_CACHE_PATH.read_text(encoding="utf-8"))
            if (
                cached.get("environment") == kis["environment"]
                and cached.get("app_key") == kis["app_key"]
                and cached.get("access_token")
            ):
                expires_at = datetime.fromisoformat(str(cached.get("expires_at")))
                if expires_at > now + timedelta(minutes=5):
                    return cached
        except Exception:
            pass

    response = requests.post(
        kis["base_url"].rstrip("/") + "/oauth2/tokenP",
        headers={"content-type": "application/json; charset=utf-8"},
        json={
            "grant_type": "client_credentials",
            "appkey": kis["app_key"],
            "appsecret": kis["app_secret"],
        },
        timeout=20,
    )
    response.raise_for_status()
    payload = response.json()
    access_token = str(payload.get("access_token") or "").strip()
    if not access_token:
        raise ValueError(str(payload.get("msg1") or payload.get("error_description") or "한국투자증권 토큰을 받지 못했습니다."))
    expires_in = int(to_float(payload.get("expires_in")) or 86400)
    token_payload = {
        "environment": kis["environment"],
        "app_key": kis["app_key"],
        "access_token": access_token,
        "token_type": payload.get("token_type", "Bearer"),
        "expires_at": (now + timedelta(seconds=max(60, expires_in - 60))).isoformat(timespec="seconds"),
    }
    KIS_TOKEN_CACHE_PATH.write_text(json.dumps(token_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return token_payload


def kis_status_payload(check_token: bool = False) -> dict[str, Any]:
    kis = get_kis_settings()
    payload = {
        "configured": bool(kis["app_key"] and kis["app_secret"]),
        "environment": kis["environment"],
        "base_url": kis["base_url"],
        "app_key_masked": mask_secret(kis["app_key"]),
        "has_app_secret": bool(kis["app_secret"]),
        "has_account": bool(kis["account_no"] and kis["account_product_code"]),
        "account_no_masked": mask_secret(kis["account_no"], 2),
        "message": "한국투자증권 모의투자 API 설정이 저장되어 있습니다." if kis["app_key"] and kis["app_secret"] else "한국투자증권 API 설정이 필요합니다.",
    }
    if check_token and payload["configured"]:
        token = get_kis_access_token()
        payload["token_ok"] = bool(token.get("access_token"))
        payload["token_expires_at"] = token.get("expires_at")
    return payload


def stock_alert_settings() -> dict[str, str]:
    settings = load_settings()
    stock_alert = settings.get("stock_alert") if isinstance(settings.get("stock_alert"), dict) else {}
    return {
        "github_repository": str(os.getenv("STOCK_ALERT_GITHUB_REPOSITORY") or stock_alert.get("github_repository") or "").strip(),
        "github_token": str(os.getenv("STOCK_ALERT_GITHUB_TOKEN") or stock_alert.get("github_token") or "").strip(),
        "telegram_bot_token": str(os.getenv("TELEGRAM_BOT_TOKEN") or stock_alert.get("telegram_bot_token") or "").strip(),
        "telegram_chat_id": str(os.getenv("TELEGRAM_CHAT_ID") or stock_alert.get("telegram_chat_id") or "").strip(),
    }


def stock_alert_status_payload() -> dict[str, Any]:
    settings = stock_alert_settings()
    snapshot: dict[str, Any] = {}
    if STOCK_ALERT_HOLDINGS_SNAPSHOT_PATH.exists():
        try:
            snapshot = json.loads(STOCK_ALERT_HOLDINGS_SNAPSHOT_PATH.read_text(encoding="utf-8"))
        except Exception:
            snapshot = {}
    return {
        "repository": settings["github_repository"],
        "configured": bool(settings["github_repository"] and settings["github_token"]),
        "has_token": bool(settings["github_token"]),
        "token_masked": mask_secret(settings["github_token"]),
        "telegram": {
            "has_bot_token": bool(settings["telegram_bot_token"]),
            "bot_token_masked": mask_secret(settings["telegram_bot_token"]),
            "chat_id": settings["telegram_chat_id"],
            "configured": bool(settings["telegram_bot_token"] and settings["telegram_chat_id"]),
        },
        "secret_name": "STOCK_ALERT_HOLDINGS_JSON",
        "snapshot": {
            "updated_at": snapshot.get("updated_at") or "",
            "source_date": snapshot.get("source_date") or "",
            "holding_count": len(snapshot.get("holdings") or []),
        },
    }


def is_public_web_mode() -> bool:
    return os.getenv("STOCK_DASHBOARD_PUBLIC_WEB", "").strip().lower() in {"1", "true", "yes", "on"}


def public_web_lock_response(feature: str = "This feature") -> JSONResponse:
    return JSONResponse(
        {
            "error": f"{feature} is locked in public web mode for security reasons.",
            "locked": True,
            "reason": "public_web_security",
        },
        status_code=403,
    )


def get_telegram_settings() -> dict[str, Any]:
    settings = load_settings()
    telegram = settings.get("telegram", {})
    return {
        "api_id": int(telegram.get("api_id", 0) or 0),
        "api_hash": str(telegram.get("api_hash", "")).strip(),
        "phone": str(telegram.get("phone", "")).strip(),
    }


def save_telegram_settings(api_id: int, api_hash: str, phone: str) -> None:
    settings = load_settings()
    settings["telegram"] = {
        "api_id": int(api_id),
        "api_hash": api_hash.strip(),
        "phone": phone.strip(),
    }
    save_settings(settings)


def get_telegram_session_db_path(base_path: Path) -> Path:
    return base_path if base_path.suffix else base_path.with_suffix(".session")


def build_telegram_client(session_path: Path | None = None) -> TelegramClient:
    telegram = get_telegram_settings()
    if not telegram["api_id"] or not telegram["api_hash"]:
        raise ValueError("Telegram API settings are missing. Please enter API ID, API Hash, and phone number first.")
    TELEGRAM_SESSION_DIR.mkdir(parents=True, exist_ok=True)
    target = session_path or TELEGRAM_SESSION_FILE
    return TelegramClient(str(target), telegram["api_id"], telegram["api_hash"])


def build_telegram_readonly_client() -> tuple[TelegramClient, Path | None]:
    source_db = get_telegram_session_db_path(TELEGRAM_SESSION_FILE)
    if not source_db.exists():
        return build_telegram_client(), None

    temp_dir = Path(tempfile.mkdtemp(prefix="telegram-session-copy-", dir=str(TELEGRAM_SESSION_DIR)))
    target_base = temp_dir / "user"
    target_db = get_telegram_session_db_path(target_base)
    shutil.copy2(source_db, target_db)

    journal_path = Path(str(source_db) + "-journal")
    if journal_path.exists():
        shutil.copy2(journal_path, Path(str(target_db) + "-journal"))

    return build_telegram_client(target_base), temp_dir


def save_telegram_login_state(phone: str, phone_code_hash: str) -> None:
    TELEGRAM_SESSION_DIR.mkdir(parents=True, exist_ok=True)
    TELEGRAM_CODE_FILE.write_text(
        json.dumps({"phone": phone, "phone_code_hash": phone_code_hash}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_telegram_login_state() -> dict[str, Any]:
    if TELEGRAM_CODE_FILE.exists():
        return json.loads(TELEGRAM_CODE_FILE.read_text(encoding="utf-8"))
    return {}


def clear_telegram_login_state() -> None:
    if TELEGRAM_CODE_FILE.exists():
        TELEGRAM_CODE_FILE.unlink()


@dataclass
class PositionRow:
    sector: str
    stock_name: str
    stock_code: str | None
    resolved_name: str
    prev_weight: float
    target_weight: float
    note: str


class TelegramSendCodeRequest(BaseModel):
    api_id: int
    api_hash: str
    phone: str


class TelegramVerifyCodeRequest(BaseModel):
    phone: str
    code: str
    password: str | None = None


class TelegramSearchRequest(BaseModel):
    keywords: list[str]
    chat_ids: list[int] | None = None
    has_file: bool | None = None
    match_mode: str = "any"
    exact_phrase: bool = False
    start_date: str | None = None
    end_date: str | None = None
    limit: int | None = 200


class TelegramEarningsSearchRequest(BaseModel):
    company: str
    category: str | None = "earnings"
    limit: int | None = 30
    offset_id: int | None = None


class TelegramMarketEarningsRequest(BaseModel):
    days: int | None = 1095
    limit: int | None = 100
    scan_limit: int | None = 5000


class TelegramUiStateRequest(BaseModel):
    state: dict[str, Any]


class TradingViewOpenRequest(BaseModel):
    stock_code: str | None = None
    stock_name: str | None = None


class SectorAssignmentRequest(BaseModel):
    stock_code: str | None = None
    stock_name: str
    sector: str | None = None


class SectorStockItem(BaseModel):
    code: str | None = None
    name: str | None = None


class SectorGroupInput(BaseModel):
    sector: str
    stocks: list[SectorStockItem]


class SectorDatabaseSaveRequest(BaseModel):
    groups: list[SectorGroupInput]


class SectorWatchOrderRequest(BaseModel):
    sectors: list[str] = []
    stocks_by_sector: dict[str, list[str]] = {}


class SectorSnapshotRequest(BaseModel):
    groups: list[SectorGroupInput]


class SectorMarketCapChartRequest(BaseModel):
    sector: str
    stocks: list[SectorStockItem]
    months: int | None = 3


class ThemeReloadRequest(BaseModel):
    file_date: str | None = None
    min_score: float = 50.0
    recent_limit: int = RECENT_SCREENING_LOOKBACK
    reload_all: bool = False
    region: str | None = None


class ThemeTestExcelRequest(BaseModel):
    file_date: str | None = None
    suffix: str = "_test_52w"


class ThemeBuildTodayExcelRequest(BaseModel):
    min_score: float = 0.0
    recent_limit: int = 20
    region: str | None = None


class ThemeNoteUpdateRequest(BaseModel):
    file_date: str | None = None
    stock_code: str | None = None
    stock_name: str
    note: str | None = None
    close_open_excel: bool = False
    write_open_excel: bool = False


class GlobalCompanyDetailRequest(BaseModel):
    symbol: str


class RealEstateSaveRequest(BaseModel):
    data: dict[str, Any]


class RealEstateWaterTelegramSyncRequest(BaseModel):
    chat_name: str | None = "엄마"
    month: str | None = None
    limit: int | None = 300


class RealEstateElectricTelegramSyncRequest(BaseModel):
    chat_name: str | None = "엄마"
    month: str | None = None
    limit: int | None = 500


class MarketCalendarEventRequest(BaseModel):
    date: str
    title: str
    category: str | None = "기타"
    market: str | None = ""
    time: str | None = ""
    importance: str | None = "medium"
    note: str | None = ""
    source: str | None = "manual"


class StockAlertGitHubSettingsRequest(BaseModel):
    repository: str
    token: str


class StockAlertTelegramSettingsRequest(BaseModel):
    bot_token: str
    chat_id: str | None = ""


@lru_cache(maxsize=1)
def get_listing_table() -> pd.DataFrame:
    try:
        base_date = date.today().strftime("%Y%m%d")
        biz_date = pykrx_stock.get_nearest_business_day_in_a_week(base_date)
        rows: list[dict[str, Any]] = []

        for market in ("KOSPI", "KOSDAQ", "KONEX"):
            tickers = pykrx_stock.get_market_ticker_list(date=biz_date, market=market)
            if not tickers:
                continue

            cap_frame = pykrx_stock.get_market_cap_by_ticker(date=biz_date, market=market)
            cap_map: dict[str, dict[str, float | None]] = {}
            if cap_frame is not None and not cap_frame.empty:
                for ticker, rec in cap_frame.iterrows():
                    code = str(ticker).zfill(6)
                    cap_map[code] = {
                        "Close": to_float(rec.get("종가")),
                        "Marcap": to_float(rec.get("시가총액")),
                        "Stocks": to_float(rec.get("상장주식수")),
                    }

            for ticker in tickers:
                code = str(ticker).zfill(6)
                name = pykrx_stock.get_market_ticker_name(code) or code
                cap_row = cap_map.get(code, {})
                rows.append(
                    {
                        "Code": code,
                        "Name": name,
                        "Market": market,
                        "Close": cap_row.get("Close"),
                        "Marcap": cap_row.get("Marcap"),
                        "Stocks": cap_row.get("Stocks"),
                    }
                )

        if rows:
            listing = pd.DataFrame(rows)
            listing["normalized"] = listing["Name"].map(normalize_text)
            return listing
    except Exception:
        pass

    source = fdr.StockListing("KRX").copy()
    keep_columns = [column for column in ["Code", "Name", "Market", "Close", "Marcap", "Stocks"] if column in source.columns]
    listing = source[keep_columns].copy()
    for column in ["Close", "Marcap", "Stocks"]:
        if column not in listing.columns:
            listing[column] = np.nan
    listing["normalized"] = listing["Name"].map(normalize_text)
    return listing


@lru_cache(maxsize=1)
def get_name_lookup() -> dict[str, dict[str, str]]:
    listing = get_listing_table()
    lookup: dict[str, dict[str, str]] = {}
    for _, row in listing.iterrows():
        lookup[row["normalized"]] = {
            "code": str(row["Code"]).zfill(6),
            "name": row["Name"],
        }
    return lookup


@lru_cache(maxsize=1)
def get_screening_stock_lookup() -> dict[str, dict[str, dict[str, Any]]]:
    by_code: dict[str, dict[str, Any]] = {}
    by_name: dict[str, dict[str, Any]] = {}
    if not SCREENING_FAST_DB_PATH.exists():
        return {"by_code": by_code, "by_name": by_name}
    try:
        with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
            rows = conn.execute(
                """
                SELECT stock_code,
                       stock_name,
                       MAX(file_date_key) AS latest_file_date_key,
                       MAX(COALESCE(market_cap_100m, 0)) AS market_cap_100m
                FROM screening_rows
                WHERE stock_code IS NOT NULL
                  AND TRIM(stock_code) <> ''
                  AND stock_name IS NOT NULL
                  AND TRIM(stock_name) <> ''
                GROUP BY stock_code, stock_name
                """
            ).fetchall()
    except Exception:
        return {"by_code": by_code, "by_name": by_name}

    for stock_code, stock_name, latest_file_date_key, market_cap_100m in rows:
        code = str(stock_code or "").strip().zfill(6)
        name = str(stock_name or "").strip()
        normalized_name = normalize_text(name)
        if not code or not name or not normalized_name:
            continue
        item = {
            "code": code,
            "name": name,
            "latest_file_date_key": str(latest_file_date_key or ""),
            "market_cap_100m": to_float(market_cap_100m) or 0.0,
        }
        existing_code = by_code.get(code)
        if (
            existing_code is None
            or item["latest_file_date_key"] > str(existing_code.get("latest_file_date_key") or "")
            or (
                item["latest_file_date_key"] == str(existing_code.get("latest_file_date_key") or "")
                and item["market_cap_100m"] > float(existing_code.get("market_cap_100m") or 0)
            )
        ):
            by_code[code] = item
        existing_name = by_name.get(normalized_name)
        if (
            existing_name is None
            or item["latest_file_date_key"] > str(existing_name.get("latest_file_date_key") or "")
            or (
                item["latest_file_date_key"] == str(existing_name.get("latest_file_date_key") or "")
                and item["market_cap_100m"] > float(existing_name.get("market_cap_100m") or 0)
            )
        ):
            by_name[normalized_name] = item
    return {"by_code": by_code, "by_name": by_name}


def resolve_stock_from_screening_cache(name: str) -> tuple[str | None, str]:
    raw = str(name or "").strip()
    normalized = normalize_text(raw)
    if not normalized:
        return None, ""
    lookup = get_screening_stock_lookup()
    by_code = lookup.get("by_code") or {}
    by_name = lookup.get("by_name") or {}

    code_candidate = re.sub(r"\D", "", raw)
    if len(code_candidate) in {5, 6}:
        item = by_code.get(code_candidate.zfill(6))
        if item:
            return str(item.get("code") or "").zfill(6), str(item.get("name") or raw)

    exact = by_name.get(normalized)
    if exact:
        return str(exact.get("code") or "").zfill(6), str(exact.get("name") or raw)

    best_item: dict[str, Any] | None = None
    best_rank: tuple[int, float, str] | None = None
    for key, item in by_name.items():
        if normalized not in key and key not in normalized:
            continue
        rank = (
            0 if key.startswith(normalized) else 1,
            -float(item.get("market_cap_100m") or 0),
            str(item.get("name") or ""),
        )
        if best_rank is None or rank < best_rank:
            best_item = item
            best_rank = rank
    if best_item:
        return str(best_item.get("code") or "").zfill(6), str(best_item.get("name") or raw)
    return None, raw


def resolve_stock(name: str) -> tuple[str | None, str]:
    normalized = normalize_text(name)
    if not normalized:
        return None, ""

    code_candidate = re.sub(r"\D", "", str(name or ""))
    if len(code_candidate) in {5, 6}:
        listing_row = find_listing_row_by_code(code_candidate)
        if listing_row:
            return listing_row["code"], listing_row["name"]

    cached_code, cached_name = resolve_stock_from_screening_cache(name)
    if cached_code:
        return cached_code, cached_name

    candidate = NAME_ALIASES.get(normalized, name)
    normalized_candidate = normalize_text(candidate)
    lookup = get_name_lookup()

    if normalized_candidate in lookup:
        item = lookup[normalized_candidate]
        return item["code"], item["name"]

    for key, item in lookup.items():
        if normalized_candidate in key or key in normalized_candidate:
            return item["code"], item["name"]

    return None, name


def find_listing_row_by_code(code: str) -> dict[str, Any] | None:
    normalized_code = str(code or "").strip().zfill(6)
    if not normalized_code:
        return None
    listing = get_listing_table()
    matched = listing[listing["Code"].astype(str).str.zfill(6) == normalized_code]
    if matched.empty:
        return None
    row = matched.iloc[0]
    return {
        "code": str(row["Code"]).zfill(6),
        "name": str(row["Name"]),
        "market": str(row.get("Market", "") or ""),
        "close": to_float(row.get("Close")),
        "marcap": to_float(row.get("Marcap")),
        "stocks": to_float(row.get("Stocks")),
    }


def resolve_stock_payload(code: str | None = None, name: str | None = None) -> dict[str, Any] | None:
    if code:
        row = find_listing_row_by_code(code)
        if row:
            return row
    resolved_code, resolved_name = resolve_stock(str(name or "").strip())
    if not resolved_code:
        return None
    row = find_listing_row_by_code(resolved_code)
    if not row:
        return None
    row["name"] = resolved_name or row["name"]
    return row


def enrich_portfolio_item_metadata(item: dict[str, Any], sector_db: dict[str, Any] | None = None) -> dict[str, Any]:
    enriched = dict(item)
    code = normalize_stock_code_value(enriched.get("stock_code"))
    listing_row = find_listing_row_by_code(code) if code else None

    if listing_row:
        enriched["stock_code"] = listing_row["code"]
        enriched["resolved_name"] = listing_row["name"]
        raw_stock_name = str(enriched.get("stock_name") or "").strip()
        if not raw_stock_name or normalize_stock_code_value(raw_stock_name) == listing_row["code"]:
            enriched["stock_name"] = listing_row["name"]
    else:
        raw_name = str(enriched.get("resolved_name") or enriched.get("stock_name") or "").strip()
        resolved_code, resolved_name = resolve_stock(raw_name)
        if resolved_code:
            enriched["stock_code"] = resolved_code
            enriched["resolved_name"] = resolved_name
            raw_stock_name = str(enriched.get("stock_name") or "").strip()
            if not raw_stock_name or normalize_stock_code_value(raw_stock_name) == resolved_code:
                enriched["stock_name"] = resolved_name

    manual_sector = resolve_sector_for_stock(
        enriched.get("stock_code"),
        enriched.get("resolved_name") or enriched.get("stock_name"),
        sector_db,
    )
    if manual_sector:
        enriched["sector"] = re.sub(r"\s*\(\s*-?\d+(?:\.\d+)?%\s*\)\s*$", "", manual_sector).strip() or manual_sector
    elif enriched.get("sector"):
        enriched["sector"] = re.sub(r"\s*\(\s*-?\d+(?:\.\d+)?%\s*\)\s*$", "", str(enriched.get("sector") or "")).strip()
    return enriched


def global_stock_headers() -> dict[str, str]:
    return {"User-Agent": SEC_USER_AGENT, "Accept": "application/json,text/plain,*/*"}


def normalize_global_symbol(symbol: str) -> str:
    text = str(symbol or "").strip().upper()
    if not text:
        return ""
    return text


def sec_symbol_key(symbol: str) -> str:
    return normalize_global_symbol(symbol).replace(".", "-")


def sec_symbol_to_cik(symbol: str) -> str | None:
    target = sec_symbol_key(symbol)
    if not target:
        return None
    for item in get_global_company_tickers():
        if sec_symbol_key(item.get("symbol", "")) == target:
            cik = str(item.get("cik") or "").strip()
            return cik.zfill(10) if cik else None
    return None


@lru_cache(maxsize=1)
def get_global_company_tickers() -> list[dict[str, Any]]:
    url = "https://www.sec.gov/files/company_tickers.json"
    response = requests.get(url, headers=global_stock_headers(), timeout=20)
    response.raise_for_status()
    data = response.json()
    items: list[dict[str, Any]] = []
    for raw in data.values():
        symbol = normalize_global_symbol(raw.get("ticker", ""))
        title = str(raw.get("title") or "").strip()
        cik = str(raw.get("cik_str") or "").strip()
        if symbol and title and cik:
            items.append({"symbol": symbol, "name": title, "cik": cik.zfill(10), "source": "SEC"})
    return items


def yahoo_search_global_companies(query: str, limit: int = 8) -> list[dict[str, Any]]:
    try:
        response = requests.get(
            "https://query1.finance.yahoo.com/v1/finance/search",
            params={"q": query, "quotesCount": limit, "newsCount": 0},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=8,
        )
        response.raise_for_status()
        quotes = response.json().get("quotes", [])
    except Exception:
        return []
    items: list[dict[str, Any]] = []
    for quote in quotes:
        if str(quote.get("quoteType") or "").upper() not in {"EQUITY", "ETF"}:
            continue
        symbol = normalize_global_symbol(quote.get("symbol", ""))
        name = str(quote.get("longname") or quote.get("shortname") or "").strip()
        if not symbol or not name:
            continue
        items.append(
            {
                "symbol": symbol,
                "name": name,
                "exchange": str(quote.get("exchDisp") or quote.get("exchange") or "").strip(),
                "sector": str(quote.get("sectorDisp") or quote.get("sector") or "").strip(),
                "source": "Yahoo",
            }
        )
    return items


def search_global_companies(query: str, limit: int = 12) -> list[dict[str, Any]]:
    raw_query = str(query or "").strip()
    normalized_query = normalize_text(raw_query)
    if not raw_query:
        return []
    candidates: list[dict[str, Any]] = []
    alias_symbol = GLOBAL_COMPANY_ALIASES.get(normalized_query) or GLOBAL_COMPANY_ALIASES.get(raw_query.lower())
    if alias_symbol:
        candidates.append({"symbol": normalize_global_symbol(alias_symbol), "name": raw_query, "source": "Alias"})
    query_upper = normalize_global_symbol(raw_query)
    for item in get_global_company_tickers():
        symbol = normalize_global_symbol(item.get("symbol", ""))
        name = str(item.get("name") or "")
        if (
            query_upper and symbol.startswith(query_upper)
            or normalized_query and normalized_query in normalize_text(name)
            or normalized_query and normalize_text(symbol) == normalized_query
        ):
            candidates.append(
                {
                    "symbol": symbol,
                    "name": name,
                    "cik": item.get("cik"),
                    "exchange": "US",
                    "source": "SEC",
                }
            )
        if len(candidates) >= limit * 2:
            break
    candidates.extend(yahoo_search_global_companies(raw_query, limit=limit))
    seen: set[str] = set()
    results: list[dict[str, Any]] = []
    for item in candidates:
        symbol = normalize_global_symbol(item.get("symbol", ""))
        if not symbol or symbol in seen:
            continue
        seen.add(symbol)
        sec_cik = item.get("cik") or sec_symbol_to_cik(symbol)
        results.append(
            {
                "symbol": symbol,
                "name": item.get("name") or symbol,
                "cik": sec_cik,
                "exchange": item.get("exchange") or "",
                "sector": item.get("sector") or "",
                "source": item.get("source") or "",
            }
        )
        if len(results) >= limit:
            break
    return results


@lru_cache(maxsize=64)
def load_sec_company_facts(cik: str) -> dict[str, Any]:
    padded_cik = str(cik or "").strip().zfill(10)
    response = requests.get(
        f"https://data.sec.gov/api/xbrl/companyfacts/CIK{padded_cik}.json",
        headers=global_stock_headers(),
        timeout=30,
    )
    response.raise_for_status()
    return response.json()


@lru_cache(maxsize=128)
def load_yahoo_chart_meta(symbol: str) -> dict[str, Any]:
    normalized = normalize_global_symbol(symbol)
    response = requests.get(
        f"https://query1.finance.yahoo.com/v8/finance/chart/{normalized}",
        params={"range": "1y", "interval": "1d"},
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=10,
    )
    response.raise_for_status()
    result = (response.json().get("chart", {}).get("result") or [{}])[0]
    meta = result.get("meta", {}) or {}
    quote = ((result.get("indicators") or {}).get("quote") or [{}])[0]
    opens = [to_float(value) for value in quote.get("open", [])]
    closes = [to_float(value) for value in quote.get("close", [])]
    volumes = [to_float(value) for value in quote.get("volume", [])]
    valid_opens = [value for value in opens if value is not None]
    valid_closes = [value for value in closes if value is not None and value > 0]
    valid_volumes = [value for value in volumes if value is not None]
    first_close = valid_closes[0] if valid_closes else None
    last_close = valid_closes[-1] if valid_closes else None
    avg_volume_3m = float(np.mean(valid_volumes[-63:])) if valid_volumes else None
    one_year_return_pct = (
        (last_close / first_close - 1) * 100
        if first_close and last_close and first_close > 0
        else None
    )
    return {
        "symbol": normalized,
        "currency": meta.get("currency") or "USD",
        "exchange": meta.get("fullExchangeName") or meta.get("exchangeName") or "",
        "price": to_float(meta.get("regularMarketPrice") or meta.get("previousClose")),
        "previous_close": to_float(meta.get("chartPreviousClose") or meta.get("previousClose")),
        "open": valid_opens[-1] if valid_opens else None,
        "day_low": to_float(meta.get("regularMarketDayLow")),
        "day_high": to_float(meta.get("regularMarketDayHigh")),
        "fifty_two_week_low": to_float(meta.get("fiftyTwoWeekLow")),
        "fifty_two_week_high": to_float(meta.get("fiftyTwoWeekHigh")),
        "volume": to_float(meta.get("regularMarketVolume")),
        "avg_volume_3m": avg_volume_3m,
        "one_year_return_pct": one_year_return_pct,
        "market_time": meta.get("regularMarketTime"),
    }


@lru_cache(maxsize=1)
def get_usd_krw_rate() -> dict[str, Any]:
    try:
        response = requests.get("https://api.frankfurter.app/latest", params={"from": "USD", "to": "KRW"}, timeout=8)
        response.raise_for_status()
        payload = response.json()
        rate = to_float((payload.get("rates") or {}).get("KRW"))
        if rate:
            return {"rate": rate, "date": payload.get("date"), "source": "Frankfurter"}
    except Exception:
        pass
    meta = load_yahoo_chart_meta("KRW=X")
    return {"rate": to_float(meta.get("price")) or 0, "date": "", "source": "Yahoo"}


@lru_cache(maxsize=64)
def get_currency_rate(from_currency: str, to_currency: str = "USD") -> dict[str, Any]:
    source = str(from_currency or "").strip().upper()
    target = str(to_currency or "").strip().upper()
    if not source or not target:
        return {"rate": 0, "date": "", "source": ""}
    if source == target:
        return {"rate": 1.0, "date": "", "source": "identity"}
    try:
        response = requests.get(
            "https://api.frankfurter.app/latest",
            params={"from": source, "to": target},
            timeout=8,
        )
        response.raise_for_status()
        payload = response.json()
        rate = to_float((payload.get("rates") or {}).get(target))
        if rate:
            return {"rate": rate, "date": payload.get("date"), "source": "Frankfurter"}
    except Exception:
        pass
    yahoo_symbol = f"{source}{target}=X"
    try:
        meta = load_yahoo_chart_meta(yahoo_symbol)
        rate = to_float(meta.get("price"))
        if rate:
            return {"rate": rate, "date": "", "source": "Yahoo"}
    except Exception:
        pass
    return {"rate": 0, "date": "", "source": ""}


def convert_currency_value(value: Any, from_currency: str, to_currency: str = "USD") -> float | None:
    number = to_float(value)
    if number is None:
        return None
    source = str(from_currency or to_currency or "").strip().upper()
    target = str(to_currency or "USD").strip().upper()
    if source == target:
        return number
    rate = to_float(get_currency_rate(source, target).get("rate"))
    if not rate:
        return None
    return number * rate


def fact_values(company_facts: dict[str, Any], tag_names: list[str]) -> list[dict[str, Any]]:
    facts = company_facts.get("facts", {}).get("us-gaap", {})
    best_rows: list[dict[str, Any]] = []
    for tag in tag_names:
        rows = facts.get(tag, {}).get("units", {}).get("USD", [])
        normalized_rows = [row for row in rows if row.get("val") is not None]
        if len(normalized_rows) > len(best_rows):
            best_rows = normalized_rows
    return best_rows


def quarterly_fact_map(company_facts: dict[str, Any], tag_names: list[str]) -> dict[str, dict[str, Any]]:
    rows = fact_values(company_facts, tag_names)
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        frame = str(row.get("frame") or "")
        if not re.fullmatch(r"CY\d{4}Q[1-4]", frame):
            continue
        value = to_float(row.get("val"))
        if value is None:
            continue
        filed = str(row.get("filed") or "")
        current = result.get(frame)
        if not current or filed >= str(current.get("filed") or ""):
            result[frame] = {
                "value": value,
                "filed": filed,
                "form": row.get("form"),
                "end": row.get("end"),
                "fy": row.get("fy"),
                "fp": row.get("fp"),
            }
    return result


def annual_fact_map(company_facts: dict[str, Any], tag_names: list[str]) -> dict[str, dict[str, Any]]:
    rows = fact_values(company_facts, tag_names)
    result: dict[str, dict[str, Any]] = {}
    for row in rows:
        frame = str(row.get("frame") or "")
        if not re.fullmatch(r"CY\d{4}", frame):
            continue
        value = to_float(row.get("val"))
        if value is None:
            continue
        filed = str(row.get("filed") or "")
        current = result.get(frame)
        if not current or filed >= str(current.get("filed") or ""):
            result[frame] = {
                "value": value,
                "filed": filed,
                "form": row.get("form"),
                "end": row.get("end"),
                "fy": row.get("fy"),
                "fp": row.get("fp"),
            }
    return result


def latest_shares_outstanding(company_facts: dict[str, Any]) -> float | None:
    facts = company_facts.get("facts", {}).get("dei", {})
    rows = facts.get("EntityCommonStockSharesOutstanding", {}).get("units", {}).get("shares", [])
    latest: dict[str, Any] | None = None
    for row in rows:
        value = to_float(row.get("val"))
        filed = str(row.get("filed") or row.get("end") or "")
        if value is None or value <= 0:
            continue
        if not latest or filed >= str(latest.get("filed") or ""):
            latest = {"value": value, "filed": filed}
    return to_float((latest or {}).get("value"))


YAHOO_FINANCIAL_TYPES = [
    "quarterlyTotalRevenue",
    "quarterlyOperatingIncome",
    "quarterlyNetIncome",
    "quarterlyGrossProfit",
    "quarterlyEbitda",
    "annualTotalRevenue",
    "annualOperatingIncome",
    "annualNetIncome",
    "annualGrossProfit",
    "annualEbitda",
    "trailingMarketCap",
    "quarterlyMarketCap",
    "annualMarketCap",
    "quarterlyDilutedAverageShares",
    "annualDilutedAverageShares",
    "quarterlyBasicAverageShares",
    "annualBasicAverageShares",
]


@lru_cache(maxsize=128)
def load_yahoo_financial_timeseries(symbol: str) -> dict[str, Any]:
    normalized = normalize_global_symbol(symbol)
    response = requests.get(
        f"https://query1.finance.yahoo.com/ws/fundamentals-timeseries/v1/finance/timeseries/{normalized}",
        params={
            "symbol": normalized,
            "type": ",".join(YAHOO_FINANCIAL_TYPES),
            "period1": 0,
            "period2": int(time.time()),
        },
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=20,
    )
    response.raise_for_status()
    return response.json()


def yahoo_timeseries_items(payload: dict[str, Any], type_name: str) -> list[dict[str, Any]]:
    for item in payload.get("timeseries", {}).get("result", []) or []:
        if type_name in item:
            rows = item.get(type_name) or []
            if isinstance(rows, list):
                return rows
    return []


def yahoo_reported_value(row: dict[str, Any], to_currency: str = "USD") -> float | None:
    reported = row.get("reportedValue") or {}
    value = to_float(reported.get("raw"))
    currency = str(row.get("currencyCode") or to_currency or "USD").strip().upper()
    return convert_currency_value(value, currency, to_currency)


def yahoo_latest_value(payload: dict[str, Any], type_names: list[str], to_currency: str = "USD") -> float | None:
    best: dict[str, Any] | None = None
    for type_name in type_names:
        for row in yahoo_timeseries_items(payload, type_name):
            value = to_float((row.get("reportedValue") or {}).get("raw"))
            if value is None:
                continue
            if not best or str(row.get("asOfDate") or "") >= str(best.get("asOfDate") or ""):
                best = row
    return yahoo_reported_value(best, to_currency) if best else None


def yahoo_statement_map(payload: dict[str, Any], type_name: str, to_currency: str = "USD") -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in yahoo_timeseries_items(payload, type_name):
        as_of = str(row.get("asOfDate") or "")
        if not as_of:
            continue
        value = yahoo_reported_value(row, to_currency)
        if value is None:
            continue
        result[as_of] = {
            "value": value,
            "filed": as_of,
            "form": "Yahoo",
            "currency": str(row.get("currencyCode") or "").upper(),
        }
    return result


def quarter_label_from_date(as_of: str) -> str:
    try:
        parsed = datetime.fromisoformat(str(as_of)[:10])
        quarter = (parsed.month - 1) // 3 + 1
        return f"{parsed.year}.{quarter}Q"
    except Exception:
        return str(as_of or "")


def annual_label_from_date(as_of: str) -> str:
    return str(as_of or "")[:4]


def convert_meta_money_fields_to_usd(meta: dict[str, Any]) -> dict[str, Any]:
    currency = str(meta.get("currency") or "USD").upper()
    converted = dict(meta)
    for key in [
        "price",
        "previous_close",
        "open",
        "day_low",
        "day_high",
        "fifty_two_week_low",
        "fifty_two_week_high",
    ]:
        converted[key] = convert_currency_value(meta.get(key), currency, "USD")
    converted["local_currency"] = currency
    converted["currency"] = "USD"
    return converted


def build_yahoo_global_company_detail(symbol: str, company: dict[str, Any] | None = None) -> dict[str, Any]:
    normalized_symbol = normalize_global_symbol(symbol)
    meta = convert_meta_money_fields_to_usd(load_yahoo_chart_meta(normalized_symbol))
    fx = get_usd_krw_rate()
    rate = to_float(fx.get("rate")) or 0
    timeseries = load_yahoo_financial_timeseries(normalized_symbol)
    search_company = company or {}
    if (
        not search_company.get("name")
        or normalize_global_symbol(search_company.get("name", "")) == normalized_symbol
        or search_company.get("source") == "SEC"
    ):
        try:
            search_company = next(
                (
                    item
                    for item in yahoo_search_global_companies(normalized_symbol, limit=6)
                    if normalize_global_symbol(item.get("symbol", "")) == normalized_symbol
                ),
                search_company,
            )
        except Exception:
            pass

    revenue_map = yahoo_statement_map(timeseries, "quarterlyTotalRevenue")
    operating_map = yahoo_statement_map(timeseries, "quarterlyOperatingIncome")
    net_map = yahoo_statement_map(timeseries, "quarterlyNetIncome")
    gross_profit_map = yahoo_statement_map(timeseries, "quarterlyGrossProfit")
    ebitda_map = yahoo_statement_map(timeseries, "quarterlyEbitda")
    frames = sorted(set(revenue_map) | set(operating_map) | set(net_map), reverse=True)
    quarters = []
    for frame in frames[:12]:
        revenue = revenue_map.get(frame, {})
        operating = operating_map.get(frame, {})
        net = net_map.get(frame, {})
        gross_profit = gross_profit_map.get(frame, {})
        ebitda = ebitda_map.get(frame, {})
        quarters.append(
            {
                "frame": frame,
                "label": quarter_label_from_date(frame),
                "revenue": revenue.get("value"),
                "operating_income": operating.get("value"),
                "net_income": net.get("value"),
                "gross_profit": gross_profit.get("value"),
                "ebitda": ebitda.get("value"),
                "filed": frame,
                "form": "Yahoo",
            }
        )

    annual_revenue_map = yahoo_statement_map(timeseries, "annualTotalRevenue")
    annual_operating_map = yahoo_statement_map(timeseries, "annualOperatingIncome")
    annual_net_map = yahoo_statement_map(timeseries, "annualNetIncome")
    annual_gross_profit_map = yahoo_statement_map(timeseries, "annualGrossProfit")
    annual_ebitda_map = yahoo_statement_map(timeseries, "annualEbitda")
    annual_frames = sorted(set(annual_revenue_map) | set(annual_operating_map) | set(annual_net_map), reverse=True)
    annuals = []
    for frame in annual_frames[:8]:
        revenue = annual_revenue_map.get(frame, {})
        operating = annual_operating_map.get(frame, {})
        net = annual_net_map.get(frame, {})
        gross_profit = annual_gross_profit_map.get(frame, {})
        ebitda = annual_ebitda_map.get(frame, {})
        annuals.append(
            {
                "frame": frame,
                "label": annual_label_from_date(frame),
                "revenue": revenue.get("value"),
                "operating_income": operating.get("value"),
                "net_income": net.get("value"),
                "gross_profit": gross_profit.get("value"),
                "ebitda": ebitda.get("value"),
                "filed": frame,
                "form": "Yahoo",
            }
        )

    market_cap = yahoo_latest_value(timeseries, ["trailingMarketCap", "quarterlyMarketCap", "annualMarketCap"])
    shares = yahoo_latest_value(timeseries, ["quarterlyDilutedAverageShares", "quarterlyBasicAverageShares", "annualDilutedAverageShares", "annualBasicAverageShares"])
    latest_quarter = quarters[0] if quarters else {}
    revenue = to_float(latest_quarter.get("revenue"))
    gross_profit = to_float(latest_quarter.get("gross_profit"))
    operating_income = to_float(latest_quarter.get("operating_income"))
    net_income = to_float(latest_quarter.get("net_income"))
    return {
        "symbol": normalized_symbol,
        "name": search_company.get("name") or normalized_symbol,
        "cik": "",
        "exchange": meta.get("exchange") or search_company.get("exchange") or "",
        "sector": search_company.get("sector") or "",
        "currency": "USD",
        "local_currency": meta.get("local_currency") or "",
        "price": meta.get("price"),
        "shares_outstanding": shares,
        "market_cap": market_cap,
        "market_cap_billion": (market_cap / 1_000_000_000) if market_cap else None,
        "stats": {
            "previous_close": meta.get("previous_close"),
            "open": meta.get("open"),
            "day_low": meta.get("day_low"),
            "day_high": meta.get("day_high"),
            "fifty_two_week_low": meta.get("fifty_two_week_low"),
            "fifty_two_week_high": meta.get("fifty_two_week_high"),
            "volume": meta.get("volume"),
            "avg_volume_3m": meta.get("avg_volume_3m"),
            "one_year_return_pct": meta.get("one_year_return_pct"),
            "shares_outstanding": shares,
            "latest_revenue": revenue,
            "latest_net_income": net_income,
            "latest_ebitda": latest_quarter.get("ebitda"),
            "gross_margin_pct": (gross_profit / revenue * 100) if revenue and gross_profit is not None else None,
            "operating_margin_pct": (operating_income / revenue * 100) if revenue and operating_income is not None else None,
            "net_margin_pct": (net_income / revenue * 100) if revenue and net_income is not None else None,
        },
        "usd_krw": rate,
        "fx_date": fx.get("date") or "",
        "fx_source": fx.get("source") or "",
        "quarters": quarters,
        "annuals": annuals,
        "source": {
            "financials": "Yahoo Finance fundamentals-timeseries",
            "price": "Yahoo Finance chart",
            "fx": fx.get("source") or "",
        },
    }


def build_global_company_detail(symbol: str) -> dict[str, Any]:
    raw_symbol = str(symbol or "").strip()
    alias_symbol = GLOBAL_COMPANY_ALIASES.get(normalize_text(raw_symbol)) or GLOBAL_COMPANY_ALIASES.get(raw_symbol.lower())
    normalized_symbol = normalize_global_symbol(alias_symbol or raw_symbol)
    if not normalized_symbol:
        raise ValueError("검색할 티커를 입력해 주세요.")
    search_items = search_global_companies(normalized_symbol, limit=5)
    company = next((item for item in search_items if normalize_global_symbol(item.get("symbol", "")) == normalized_symbol), None)
    cik = (company or {}).get("cik") or sec_symbol_to_cik(normalized_symbol)
    if not cik:
        return build_yahoo_global_company_detail(normalized_symbol, company)

    facts = load_sec_company_facts(str(cik))
    meta = load_yahoo_chart_meta(normalized_symbol)
    fx = get_usd_krw_rate()
    rate = to_float(fx.get("rate")) or 0
    revenue_map = quarterly_fact_map(facts, ["Revenues", "SalesRevenueNet", "RevenueFromContractWithCustomerExcludingAssessedTax"])
    operating_map = quarterly_fact_map(facts, ["OperatingIncomeLoss"])
    net_map = quarterly_fact_map(facts, ["NetIncomeLoss", "ProfitLoss"])
    gross_profit_map = quarterly_fact_map(facts, ["GrossProfit"])
    ebitda_map = quarterly_fact_map(facts, ["EarningsBeforeInterestTaxesDepreciationAmortization"])
    frames = sorted(set(revenue_map) | set(operating_map) | set(net_map), reverse=True)
    quarters = []
    for frame in frames[:12]:
        year = frame[2:6]
        quarter = frame[-1]
        revenue = revenue_map.get(frame, {})
        operating = operating_map.get(frame, {})
        net = net_map.get(frame, {})
        gross_profit = gross_profit_map.get(frame, {})
        ebitda = ebitda_map.get(frame, {})
        filed_dates = [str(item.get("filed") or "") for item in [revenue, operating, net] if item.get("filed")]
        quarters.append(
            {
                "frame": frame,
                "label": f"{year}.{quarter}Q",
                "revenue": revenue.get("value"),
                "operating_income": operating.get("value"),
                "net_income": net.get("value"),
                "gross_profit": gross_profit.get("value"),
                "ebitda": ebitda.get("value"),
                "filed": max(filed_dates) if filed_dates else "",
                "form": revenue.get("form") or operating.get("form") or net.get("form") or "",
            }
        )

    annual_revenue_map = annual_fact_map(facts, ["Revenues", "SalesRevenueNet", "RevenueFromContractWithCustomerExcludingAssessedTax"])
    annual_operating_map = annual_fact_map(facts, ["OperatingIncomeLoss"])
    annual_net_map = annual_fact_map(facts, ["NetIncomeLoss", "ProfitLoss"])
    annual_gross_profit_map = annual_fact_map(facts, ["GrossProfit"])
    annual_ebitda_map = annual_fact_map(facts, ["EarningsBeforeInterestTaxesDepreciationAmortization"])
    annual_frames = sorted(set(annual_revenue_map) | set(annual_operating_map) | set(annual_net_map), reverse=True)
    annuals = []
    for frame in annual_frames[:8]:
        year = frame[2:6]
        revenue = annual_revenue_map.get(frame, {})
        operating = annual_operating_map.get(frame, {})
        net = annual_net_map.get(frame, {})
        gross_profit = annual_gross_profit_map.get(frame, {})
        ebitda = annual_ebitda_map.get(frame, {})
        filed_dates = [str(item.get("filed") or "") for item in [revenue, operating, net] if item.get("filed")]
        annuals.append(
            {
                "frame": frame,
                "label": year,
                "revenue": revenue.get("value"),
                "operating_income": operating.get("value"),
                "net_income": net.get("value"),
                "gross_profit": gross_profit.get("value"),
                "ebitda": ebitda.get("value"),
                "filed": max(filed_dates) if filed_dates else "",
                "form": revenue.get("form") or operating.get("form") or net.get("form") or "",
            }
        )

    if not quarters:
        return build_yahoo_global_company_detail(normalized_symbol, company)

    shares = latest_shares_outstanding(facts)
    price = to_float(meta.get("price"))
    try:
        yahoo_market_cap = yahoo_latest_value(load_yahoo_financial_timeseries(normalized_symbol), ["trailingMarketCap", "quarterlyMarketCap", "annualMarketCap"])
    except Exception:
        yahoo_market_cap = None
    market_cap = yahoo_market_cap or (price * shares if price and shares else None)
    latest_quarter = quarters[0] if quarters else {}
    revenue = to_float(latest_quarter.get("revenue"))
    gross_profit = to_float(latest_quarter.get("gross_profit"))
    operating_income = to_float(latest_quarter.get("operating_income"))
    net_income = to_float(latest_quarter.get("net_income"))
    return {
        "symbol": normalized_symbol,
        "name": facts.get("entityName") or (company or {}).get("name") or normalized_symbol,
        "cik": cik,
        "exchange": meta.get("exchange") or (company or {}).get("exchange") or "",
        "sector": (company or {}).get("sector") or "",
        "currency": meta.get("currency") or "USD",
        "price": price,
        "shares_outstanding": shares,
        "market_cap": market_cap,
        "market_cap_billion": (market_cap / 1_000_000_000) if market_cap else None,
        "stats": {
            "previous_close": meta.get("previous_close"),
            "open": meta.get("open"),
            "day_low": meta.get("day_low"),
            "day_high": meta.get("day_high"),
            "fifty_two_week_low": meta.get("fifty_two_week_low"),
            "fifty_two_week_high": meta.get("fifty_two_week_high"),
            "volume": meta.get("volume"),
            "avg_volume_3m": meta.get("avg_volume_3m"),
            "one_year_return_pct": meta.get("one_year_return_pct"),
            "shares_outstanding": shares,
            "latest_revenue": revenue,
            "latest_net_income": net_income,
            "latest_ebitda": latest_quarter.get("ebitda"),
            "gross_margin_pct": (gross_profit / revenue * 100) if revenue and gross_profit is not None else None,
            "operating_margin_pct": (operating_income / revenue * 100) if revenue and operating_income is not None else None,
            "net_margin_pct": (net_income / revenue * 100) if revenue and net_income is not None else None,
        },
        "usd_krw": rate,
        "fx_date": fx.get("date") or "",
        "fx_source": fx.get("source") or "",
        "quarters": quarters,
        "annuals": annuals,
        "source": {
            "financials": "SEC companyfacts",
            "price": "Yahoo Finance chart",
            "fx": fx.get("source") or "",
        },
    }


def global_company_ai_cache_key(symbol: str) -> str:
    return f"v1|{normalize_global_symbol(symbol)}"


def normalize_global_company_ai_brief(payload: Any, symbol: str, detail: dict[str, Any]) -> dict[str, Any]:
    source = payload if isinstance(payload, dict) else {}
    history_rows = []
    for item in source.get("history", []) if isinstance(source.get("history"), list) else []:
        if not isinstance(item, dict):
            continue
        year = str(item.get("year") or "").strip()
        event = str(item.get("event") or "").strip()
        if year or event:
            history_rows.append({"year": year, "event": event})
    segment_rows = []
    total_pct = 0.0
    finite_pct_count = 0
    for item in source.get("business_segments", []) if isinstance(source.get("business_segments"), list) else []:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "").strip()
        description = str(item.get("description") or "").strip()
        confidence = str(item.get("confidence") or "medium").strip().lower()
        share_pct = to_float(item.get("share_pct"))
        if share_pct is not None:
            share_pct = round(float(max(0.0, min(100.0, share_pct))), 1)
            total_pct += share_pct
            finite_pct_count += 1
        if name or description:
            segment_rows.append(
                {
                    "name": name,
                    "share_pct": share_pct,
                    "description": description,
                    "confidence": confidence if confidence in {"high", "medium", "low"} else "medium",
                }
            )
    return {
        "symbol": normalize_global_symbol(symbol),
        "company_name": str(detail.get("name") or source.get("company_name") or symbol),
        "overview": str(source.get("overview") or "").strip(),
        "history": history_rows[:8],
        "business_segments": segment_rows[:8],
        "revenue_mix_note": str(source.get("revenue_mix_note") or "").strip(),
        "risks": [str(item).strip() for item in source.get("risks", []) if str(item).strip()][:5] if isinstance(source.get("risks"), list) else [],
        "has_segment_percentages": bool(finite_pct_count),
        "segment_pct_total": round(total_pct, 1) if finite_pct_count else None,
    }


def build_global_company_ai_request_payload(detail: dict[str, Any]) -> dict[str, Any]:
    quarters = [item for item in (detail.get("quarters") or []) if isinstance(item, dict)][:6]
    annuals = [item for item in (detail.get("annuals") or []) if isinstance(item, dict)][:4]
    stats = detail.get("stats") if isinstance(detail.get("stats"), dict) else {}
    return {
        "symbol": str(detail.get("symbol") or ""),
        "name": str(detail.get("name") or ""),
        "exchange": str(detail.get("exchange") or ""),
        "sector": str(detail.get("sector") or ""),
        "currency": str(detail.get("currency") or ""),
        "market_cap_billion": round(float(to_float(detail.get("market_cap_billion")) or 0.0), 2) if to_float(detail.get("market_cap_billion")) is not None else None,
        "price": to_float(detail.get("price")),
        "latest_revenue": to_float(stats.get("latest_revenue")),
        "latest_net_income": to_float(stats.get("latest_net_income")),
        "gross_margin_pct": to_float(stats.get("gross_margin_pct")),
        "operating_margin_pct": to_float(stats.get("operating_margin_pct")),
        "net_margin_pct": to_float(stats.get("net_margin_pct")),
        "quarters": [
            {
                "label": str(item.get("label") or ""),
                "revenue": to_float(item.get("revenue")),
                "operating_income": to_float(item.get("operating_income")),
                "net_income": to_float(item.get("net_income")),
                "filed": str(item.get("filed") or ""),
                "form": str(item.get("form") or ""),
            }
            for item in quarters
        ],
        "annuals": [
            {
                "label": str(item.get("label") or ""),
                "revenue": to_float(item.get("revenue")),
                "operating_income": to_float(item.get("operating_income")),
                "net_income": to_float(item.get("net_income")),
                "filed": str(item.get("filed") or ""),
                "form": str(item.get("form") or ""),
            }
            for item in annuals
        ],
        "source": detail.get("source") if isinstance(detail.get("source"), dict) else {},
    }


def request_openai_global_company_brief(detail: dict[str, Any]) -> dict[str, Any]:
    api_key = get_openai_api_key()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY가 설정되어 있지 않습니다.")
    compact_payload = build_global_company_ai_request_payload(detail)
    schema = {
        "name": "global_company_ai_brief",
        "strict": True,
        "schema": {
            "type": "object",
            "properties": {
                "company_name": {"type": "string"},
                "overview": {"type": "string"},
                "history": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "year": {"type": "string"},
                            "event": {"type": "string"},
                        },
                        "required": ["year", "event"],
                        "additionalProperties": False,
                    },
                },
                "business_segments": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "properties": {
                            "name": {"type": "string"},
                            "share_pct": {"type": ["number", "null"]},
                            "description": {"type": "string"},
                            "confidence": {"type": "string"},
                        },
                        "required": ["name", "share_pct", "description", "confidence"],
                        "additionalProperties": False,
                    },
                },
                "revenue_mix_note": {"type": "string"},
                "risks": {"type": "array", "items": {"type": "string"}},
            },
            "required": ["company_name", "overview", "history", "business_segments", "revenue_mix_note", "risks"],
            "additionalProperties": False,
        },
    }
    system_prompt = (
        "You are an equity research assistant. Return concise Korean JSON only. "
        "Summarize the company's history and business areas for an investor UI. "
        "Do not invent precise revenue mix percentages unless reasonably well-known from public information. "
        "If uncertain, set share_pct to null and mention the uncertainty in revenue_mix_note. "
        "Keep overview to 3-5 sentences, history to 4-6 items, business_segments to 3-6 items, and risks to 2-4 bullets."
    )
    user_prompt = (
        "다음 기업 데이터를 바탕으로 연혁과 사업분야를 요약해 주세요.\n"
        "출력은 지정된 JSON 스키마만 따르세요.\n"
        f"{json.dumps(compact_payload, ensure_ascii=False)}"
    )
    model_candidates = [
        str(os.getenv("OPENAI_GLOBAL_COMPANY_MODEL", "")).strip(),
        "gpt-5-mini",
        "gpt-4.1-mini",
        "gpt-4o-mini",
    ]
    tried_errors: list[str] = []
    for model_name in [item for item in model_candidates if item]:
        try:
            response = requests.post(
                "https://api.openai.com/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": model_name,
                    "temperature": 0.2,
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt},
                    ],
                    "response_format": {
                        "type": "json_schema",
                        "json_schema": schema,
                    },
                },
                timeout=90,
            )
            response.raise_for_status()
            payload = response.json()
            content = (((payload.get("choices") or [{}])[0].get("message") or {}).get("content") or "").strip()
            if not content:
                raise RuntimeError("OpenAI 응답 본문이 비어 있습니다.")
            parsed = json.loads(content)
            normalized = normalize_global_company_ai_brief(parsed, str(detail.get("symbol") or ""), detail)
            normalized["model"] = model_name
            normalized["generated_at"] = datetime.now().isoformat(timespec="seconds")
            return normalized
        except Exception as exc:
            tried_errors.append(f"{model_name}: {exc}")
            continue
    raise RuntimeError(" / ".join(tried_errors) if tried_errors else "OpenAI 기업 브리프 생성 실패")


def build_global_company_ai_brief(symbol: str, force_refresh: bool = False) -> dict[str, Any]:
    normalized_symbol = normalize_global_symbol(symbol)
    if not normalized_symbol:
        raise ValueError("검색할 티커를 입력해 주세요.")
    cache = load_global_company_ai_cache()
    cache_key = global_company_ai_cache_key(normalized_symbol)
    cached = cache.setdefault("items", {}).get(cache_key)
    if not force_refresh and isinstance(cached, dict):
        return cached
    detail = build_global_company_detail(normalized_symbol)
    ai_brief = request_openai_global_company_brief(detail)
    payload = {
        "symbol": normalized_symbol,
        "company_name": str(detail.get("name") or normalized_symbol),
        "brief": ai_brief,
        "cache_source": "openai",
        "cached_at": datetime.now().isoformat(timespec="seconds"),
    }
    cache["items"][cache_key] = payload
    save_global_company_ai_cache(cache)
    return payload


def normalize_stock_code_value(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (int, np.integer)):
        return str(int(value)).zfill(6)
    if isinstance(value, (float, np.floating)):
        return str(int(value)).zfill(6)
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text.split(".", 1)[0]
    digits = re.sub(r"\D", "", text)
    return digits.zfill(6) if digits else ""


def upsert_sector_assignment(stock_code: str | None, stock_name: str, sector: str | None) -> dict[str, Any]:
    db = load_sector_db()
    stock_map = dict(db.get("stock_map", {}))
    normalized_sector = str(sector or "").strip()
    resolved_code = str(stock_code or "").strip()
    resolved_name = str(stock_name or "").strip()
    if not resolved_code and resolved_name:
        resolved_code, resolved_name_from_lookup = resolve_stock(resolved_name)
        resolved_name = resolved_name_from_lookup or resolved_name
    key = sector_db_key(resolved_code, resolved_name)
    if not key:
        return save_sector_db(db)
    if normalized_sector:
        stock_map[key] = {
            "stock_code": resolved_code or "",
            "stock_name": resolved_name,
            "sector": normalized_sector,
        }
    else:
        stock_map.pop(key, None)
    return save_sector_db({"stock_map": stock_map, "sectors": db.get("sectors", []), "stock_order": db.get("stock_order", {})})


def save_sector_groups_to_db(groups: list[SectorGroupInput]) -> dict[str, Any]:
    if not groups:
        return load_sector_db()
    current_db = load_sector_db()
    stock_map: dict[str, dict[str, str]] = dict(current_db.get("stock_map", {}))
    sectors: list[str] = [str(item or "").strip() for item in current_db.get("sectors", []) if str(item or "").strip()]
    stock_order: dict[str, list[str]] = dict(current_db.get("stock_order", {}))
    for group in groups:
        sector = str(group.sector or "").strip()
        if not sector:
            continue
        if sector not in sectors:
            sectors.append(sector)
        group_order: list[str] = []
        for stock in group.stocks:
            raw_name = str(stock.name or stock.code or "").strip()
            if not raw_name:
                continue
            resolved = resolve_stock_payload(stock.code, stock.name)
            stock_code = str((resolved or {}).get("code") or stock.code or "").strip()
            stock_name = str((resolved or {}).get("name") or raw_name).strip()
            key = sector_db_key(stock_code, stock_name)
            if key:
                stock_map[key] = {
                    "stock_code": stock_code,
                    "stock_name": stock_name,
                    "sector": sector,
                }
                if stock_code and stock_code not in group_order:
                    group_order.append(stock_code)
        if group_order:
            previous_order = [str(code or "").strip() for code in stock_order.get(sector, []) if str(code or "").strip()]
            stock_order[sector] = group_order + [code for code in previous_order if code not in group_order]
    if not stock_map:
        return load_sector_db()
    return save_sector_db({"stock_map": stock_map, "sectors": sectors, "stock_order": stock_order})


def sector_db_groups() -> list[dict[str, Any]]:
    db = load_sector_db()
    grouped: dict[str, list[dict[str, str]]] = {}
    for sector in db.get("sectors", []):
        sector_name = str(sector or "").strip()
        if sector_name:
            grouped.setdefault(sector_name, [])
    for item in db.get("stock_map", {}).values():
        sector = str(item.get("sector") or "").strip()
        stock_name = str(item.get("stock_name") or "").strip()
        if not sector or not stock_name:
            continue
        grouped.setdefault(sector, []).append(
            {
                "code": str(item.get("stock_code") or ""),
                "name": stock_name,
            }
        )
    stock_order = db.get("stock_order", {}) if isinstance(db.get("stock_order"), dict) else {}
    result = []
    for sector in db.get("sectors", []):
        sector_name = str(sector or "").strip()
        if not sector_name or sector_name not in grouped:
            continue
        order = [str(code or "").strip() for code in stock_order.get(sector_name, []) if str(code or "").strip()]
        order_index = {code: index for index, code in enumerate(order)}
        stocks = sorted(
            grouped.get(sector_name, []),
            key=lambda item: (order_index.get(str(item.get("code") or "").strip(), 999999), item.get("name") or ""),
        )
        result.append({"sector": sector_name, "stocks": stocks})
    return result


def autocomplete_stocks(query: str, limit: int = 12) -> list[dict[str, Any]]:
    listing = get_listing_table().copy()
    needle = normalize_text(query)
    if not needle:
        top = listing.sort_values("Marcap", ascending=False).head(limit)
        return [
            {
                "code": str(row["Code"]).zfill(6),
                "name": str(row["Name"]),
                "market": str(row.get("Market", "") or ""),
            }
            for _, row in top.iterrows()
        ]

    listing["code_text"] = listing["Code"].astype(str).str.zfill(6)
    listing["rank"] = 4
    listing.loc[listing["code_text"].str.startswith(needle), "rank"] = 0
    listing.loc[listing["normalized"].str.startswith(needle), "rank"] = 1
    listing.loc[listing["normalized"].str.contains(needle, na=False, regex=False), "rank"] = 2
    listing.loc[listing["code_text"].str.contains(needle, na=False, regex=False), "rank"] = listing["rank"].clip(upper=3)
    filtered = listing[
        listing["normalized"].str.contains(needle, na=False, regex=False) | listing["code_text"].str.contains(needle, na=False, regex=False)
    ].copy()
    filtered = filtered.sort_values(["rank", "Marcap"], ascending=[True, False]).head(limit)
    return [
        {
            "code": str(row["Code"]).zfill(6),
            "name": str(row["Name"]),
            "market": str(row.get("Market", "") or ""),
        }
        for _, row in filtered.iterrows()
    ]


def parse_kind_disclosure_rows(html: str) -> list[dict[str, Any]]:
    soup = BeautifulSoup(html or "", "html.parser")
    rows: list[dict[str, Any]] = []
    for tr in soup.select("table.list tbody tr"):
        cells = tr.find_all("td")
        if len(cells) < 4:
            continue
        accepted_at = cells[1].get_text(" ", strip=True)
        company_name = cells[2].get_text(" ", strip=True)
        report_link = None
        for link in cells[3].find_all("a"):
            onclick = str(link.get("onclick") or "")
            if "openDisclsViewer" in onclick:
                report_link = link
                break
        if report_link is None:
            continue
        title = report_link.get_text(" ", strip=True) or str(report_link.get("title") or "").strip()
        onclick = str(report_link.get("onclick") or "")
        match = re.search(r"openDisclsViewer\('([^']+)'", onclick)
        if not match:
            continue
        acpt_no = match.group(1).strip()
        rows.append(
            {
                "accepted_at": accepted_at,
                "company_name": company_name,
                "title": title,
                "acpt_no": acpt_no,
                "url": f"https://kind.krx.co.kr/common/disclsviewer.do?method=search&acptno={quote(acpt_no)}",
            }
        )
    return rows


def kind_report_title_matches(title: str, report_scope: str = "business") -> bool:
    text = normalize_text(title)
    if report_scope == "periodic":
        return any(token in text for token in ("분기보고서", "반기보고서", "사업보고서"))
    return "사업보고서" in text and "분기보고서" not in text and "반기보고서" not in text


def find_latest_kind_report(company: str, report_scope: str = "business") -> dict[str, Any]:
    query = str(company or "").strip()
    if not query:
        raise ValueError("기업명을 입력해 주세요.")

    code, resolved_name = resolve_stock_from_screening_cache(query)
    if not code:
        try:
            code, resolved_name = resolve_stock(query)
        except Exception:
            code, resolved_name = None, query
    if not code:
        try:
            matches = autocomplete_stocks(query, limit=1)
        except Exception:
            matches = []
        if matches:
            code = str(matches[0].get("code") or "").zfill(6)
            resolved_name = str(matches[0].get("name") or resolved_name or query)
    if not code:
        raise LookupError(f"KIND에서 조회할 종목코드를 찾지 못했습니다: {query}")

    today = datetime.now().date()
    from_date = (today - timedelta(days=365 * 5 + 10)).strftime("%Y-%m-%d")
    to_date = today.strftime("%Y-%m-%d")
    session = requests.Session()
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
        ),
        "Referer": "https://kind.krx.co.kr/disclosure/searchdisclosurebycorp.do?method=searchDisclosureByCorpMain",
        "X-Requested-With": "XMLHttpRequest",
    }
    endpoint = "https://kind.krx.co.kr/disclosure/searchdisclosurebycorp.do"
    for page_index in range(1, 6):
        data = {
            "method": "searchDisclosureByCorpSub",
            "currentPageSize": "100",
            "pageIndex": str(page_index),
            "searchCodeType": "char",
            "orderIndex": "1",
            "repIsuSrtCd": f"A{code}",
            "allRepIsuSrtCd": "",
            "forward": "searchdisclosurebycorp_sub",
            "searchMode": "",
            "kosreq": "",
            "outsvcno": "",
            "orderMode": "1",
            "orderStat": "D",
            "reportNm": "",
            "reportCd": "",
            "searchCorpName": resolved_name or query,
            "fromDate": from_date,
            "toDate": to_date,
            "reportNmTemp": "",
            "lastReport": "T",
        }
        response = session.post(endpoint, data=data, headers=headers, timeout=15)
        response.raise_for_status()
        rows = parse_kind_disclosure_rows(response.text)
        if not rows:
            break
        for row in rows:
            title = str(row.get("title") or "")
            if kind_report_title_matches(title, report_scope):
                return {
                    **row,
                    "query": query,
                    "stock_code": code,
                    "stock_name": resolved_name or query,
                    "source": "KIND",
                }
    report_label = "정기보고서" if report_scope == "periodic" else "사업보고서"
    raise LookupError(f"KIND에서 최근 5년 내 {report_label}를 찾지 못했습니다: {resolved_name or query}")


def find_latest_kind_business_report(company: str) -> dict[str, Any]:
    return find_latest_kind_report(company, "business")


def find_latest_kind_periodic_report(company: str) -> dict[str, Any]:
    return find_latest_kind_report(company, "periodic")


def kind_viewer_headers(referer: str | None = None) -> dict[str, str]:
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
        ),
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.7,en;q=0.6",
    }
    if referer:
        headers["Referer"] = referer
    return headers


def kind_main_document_url(report: dict[str, Any]) -> dict[str, Any]:
    report_url = str(report.get("url") or "").strip()
    if not report_url:
        raise ValueError("KIND 사업보고서 URL이 없습니다.")
    session = requests.Session()
    viewer_response = session.get(report_url, headers=kind_viewer_headers("https://kind.krx.co.kr/"), timeout=20)
    viewer_response.raise_for_status()
    viewer_response.encoding = "utf-8"
    soup = BeautifulSoup(viewer_response.text, "html.parser")
    main_option = None
    for option in soup.select("select#mainDoc option"):
        value = str(option.get("value") or "").strip()
        label = option.get_text(" ", strip=True)
        if value and "사업보고서" in label:
            main_option = option
            break
    if main_option is None:
        raise LookupError("KIND 사업보고서 본문 문서를 찾지 못했습니다.")
    doc_no = str(main_option.get("value") or "").split("|")[0].strip()
    if not doc_no:
        raise LookupError("KIND 사업보고서 본문 문서번호가 비어 있습니다.")
    content_response = session.post(
        "https://kind.krx.co.kr/common/disclsviewer.do",
        data={"method": "searchContents", "docNo": doc_no, "acptNo": report.get("acpt_no") or ""},
        headers={**kind_viewer_headers(report_url), "X-Requested-With": "XMLHttpRequest"},
        timeout=20,
    )
    content_response.raise_for_status()
    content_response.encoding = "utf-8"
    match = re.search(r"setPath\('([^']*)','([^']*)','([^']*)'", content_response.text)
    if not match:
        raise LookupError("KIND 사업보고서 본문 경로를 찾지 못했습니다.")
    toc_url, document_url, server_path = [item.strip() for item in match.groups()]
    return {
        "doc_no": doc_no,
        "toc_url": toc_url,
        "document_url": document_url,
        "server_path": server_path,
    }


def parse_segment_percent(value: Any) -> float | None:
    text = str(value or "").replace(",", "").strip()
    if not text or text in {"-", "nan", "None"}:
        return None
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    value = float(match.group(0))
    if "△" in text or text.startswith("("):
        value = -abs(value)
    return value


def parse_business_segment_amount_ratio(cells: list[str]) -> tuple[int | None, float | None, int | None]:
    for index, cell in enumerate(cells):
        text = str(cell or "").strip()
        if "%" not in text:
            continue
        ratio_text = text
        parenthesis_match = re.search(r"\(([^()]*(?:%|％)[^()]*)\)", text)
        if parenthesis_match:
            ratio_text = parenthesis_match.group(1)
        ratio = parse_segment_percent(ratio_text)
        if ratio is None:
            continue
        amount = None
        before_parenthesis = text.split("(", 1)[0].strip()
        if before_parenthesis and before_parenthesis != text:
            amount = clean_numeric_text(before_parenthesis)
        if amount is None:
            for amount_index in range(index - 1, -1, -1):
                numeric = clean_numeric_text(cells[amount_index])
                if numeric is not None and abs(float(numeric)) > 0:
                    amount = numeric
                    break
        if amount is not None:
            return int(round(float(amount))), round(float(ratio), 2), index

    for index in range(1, len(cells)):
        amount = clean_numeric_text(cells[index - 1])
        ratio = clean_numeric_text(cells[index])
        if amount is None or ratio is None:
            continue
        amount_value = float(amount)
        ratio_value = float(ratio)
        if abs(amount_value) >= 1000 and 0 < ratio_value <= 100:
            return int(round(amount_value)), round(ratio_value, 2), index
    return None, None, None


def merge_business_segment_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    order: list[str] = []
    for row in rows:
        segment = str(row.get("segment") or "").strip()
        if not segment:
            continue
        if segment not in grouped:
            grouped[segment] = {**row}
            grouped[segment]["amount_million_krw"] = int(row.get("amount_million_krw") or 0)
            grouped[segment]["ratio_pct"] = float(row.get("ratio_pct") or 0.0)
            order.append(segment)
            continue
        target = grouped[segment]
        target["amount_million_krw"] = int(target.get("amount_million_krw") or 0) + int(row.get("amount_million_krw") or 0)
        target["ratio_pct"] = round(float(target.get("ratio_pct") or 0.0) + float(row.get("ratio_pct") or 0.0), 2)
        for field in ["sales_type", "items", "usage", "brand"]:
            current = str(target.get(field) or "").strip()
            addition = str(row.get(field) or "").strip()
            if addition and addition not in current:
                target[field] = (current + " / " + addition).strip(" /")
    return [grouped[key] for key in order]


def business_segment_table_score(frame: pd.DataFrame) -> int:
    column_text = " ".join(
        " ".join([str(part) for part in column if str(part) != "nan"]) if isinstance(column, tuple) else str(column)
        for column in frame.columns
    )
    text = column_text + " " + " ".join(
        str(value)
        for value in frame.fillna("").astype(str).values.flatten().tolist()
        if str(value).strip()
    )
    score = 0
    for token, weight in [
        ("사업부문", 12),
        ("부 문", 10),
        ("부문", 8),
        ("매출유형", 8),
        ("주요 제품", 8),
        ("품 목", 6),
        ("품목", 5),
        ("구체적용도", 6),
        ("주요상표", 6),
        ("주요 생산", 8),
        ("판매제품", 8),
        ("금 액", 8),
        ("매출액", 12),
        ("비 율", 12),
        ("비율", 12),
        ("비중", 12),
    ]:
        if token in text:
            score += weight
    if "매입유형" in text or "매입액" in text:
        score -= 30
    if "주요 매출처" in text:
        score -= 20
    return score


def parse_kind_business_segment_rows(document_html: str) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    try:
        tables = pd.read_html(StringIO(document_html))
    except Exception:
        tables = []
    best_rows: list[dict[str, Any]] = []
    best_score = -999
    best_index = -1
    for table_index, frame in enumerate(tables):
        if frame.empty or frame.shape[1] < 4:
            continue
        score = business_segment_table_score(frame)
        if score < 35:
            continue
        rows: list[dict[str, Any]] = []
        for _, raw_row in frame.fillna("").iterrows():
            cells = [re.sub(r"\s+", " ", str(value)).strip() for value in raw_row.tolist()]
            joined = " ".join(cells)
            if not joined:
                continue
            if any(token in joined for token in ["사업부문 매출유형", "매출유형 품", "매출액 비율"]):
                continue
            if cells[0] in {"사업부문", "부 문", "부문", "합 계", "합계", "소 계", "소계", "총 계", "총계", "계"} or cells[0].startswith("※"):
                continue
            amount, ratio, ratio_idx = parse_business_segment_amount_ratio(cells)
            if ratio is None or amount is None:
                continue
            segment = cells[0]
            if not segment or segment in {"사업부문", "매출유형", "소 계", "소계"}:
                continue
            amount_idx = max(1, int(ratio_idx or 1) - 1)
            descriptors = [cell for cell in cells[1:amount_idx] if cell and cell not in {"매출유형", "품목", "품 목", "구체적용도", "주요상표등", "구분"}]
            sales_type = descriptors[0] if len(descriptors) >= 2 else ""
            items = descriptors[1] if len(descriptors) >= 2 else (descriptors[0] if descriptors else "")
            usage = descriptors[2] if len(descriptors) >= 3 else ""
            brand = descriptors[3] if len(descriptors) >= 4 else ""
            row = {
                "segment": segment,
                "sales_type": sales_type,
                "items": items,
                "usage": usage,
                "brand": brand,
                "amount_million_krw": int(round(float(amount))),
                "ratio_pct": round(float(ratio), 2),
            }
            if row["segment"] and row["ratio_pct"] > 0:
                rows.append(row)
        rows = merge_business_segment_rows(rows)
        if not rows:
            continue
        ratio_sum = sum(float(row.get("ratio_pct") or 0) for row in rows)
        if 85 <= ratio_sum <= 115:
            score += 25
        score += min(len(rows), 8)
        if score > best_score:
            best_score = score
            best_rows = rows
            best_index = table_index
    total_amount = sum(int(row.get("amount_million_krw") or 0) for row in best_rows)
    total_ratio = sum(float(row.get("ratio_pct") or 0) for row in best_rows)
    return best_rows, {
        "table_index": best_index,
        "table_score": best_score if best_rows else None,
        "unit": "백만원",
        "total_amount_million_krw": total_amount,
        "total_ratio_pct": round(total_ratio, 2) if best_rows else None,
    }


def load_kind_business_segments(company: str) -> dict[str, Any]:
    query = str(company or "").strip()
    if not query:
        raise ValueError("기업명을 입력해 주세요.")
    KIND_BUSINESS_SEGMENT_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    code, resolved_name = resolve_stock_from_screening_cache(query)
    if not code:
        try:
            code, resolved_name = resolve_stock(query)
        except Exception:
            code, resolved_name = None, query
    cache_seed = ("v3|" + (code or normalize_search_text(resolved_name or query) or query)).encode("utf-8")
    cache_path = KIND_BUSINESS_SEGMENT_CACHE_DIR / f"{hashlib.sha1(cache_seed).hexdigest()[:16]}.json"
    if cache_path.exists():
        try:
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
            if time.time() - float(cached.get("_cached_at") or 0) < 24 * 60 * 60:
                return {key: value for key, value in cached.items() if key != "_cached_at"}
        except Exception:
            pass
    report = find_latest_kind_business_report(query)
    document_meta = kind_main_document_url(report)
    document_url = document_meta["document_url"]
    response = requests.get(document_url, headers=kind_viewer_headers(report.get("url")), timeout=35)
    response.raise_for_status()
    response.encoding = "utf-8"
    rows, table_meta = parse_kind_business_segment_rows(response.text)
    if not rows:
        raise LookupError("최신 사업보고서에서 사업부문별 매출 비중 표를 찾지 못했습니다.")
    rows.sort(key=lambda item: float(item.get("ratio_pct") or 0), reverse=True)
    payload = {
        "query": query,
        "stock_code": report.get("stock_code") or code,
        "stock_name": report.get("stock_name") or resolved_name or query,
        "report_title": report.get("title"),
        "accepted_at": report.get("accepted_at"),
        "kind_url": report.get("url"),
        "document_url": document_url,
        "source": "KIND 최신 사업보고서",
        "segments": rows,
        "summary": table_meta,
    }
    cache_path.write_text(json.dumps({**payload, "_cached_at": time.time()}, ensure_ascii=False, indent=2), encoding="utf-8")
    return payload


def news_headers() -> dict[str, str]:
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
        ),
        "Accept-Language": "ko-KR,ko;q=0.9,en-US;q=0.7,en;q=0.6",
    }


def clean_news_text(value: Any) -> str:
    text = html_lib.unescape(str(value or ""))
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\[[^\]]{1,24}\]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_news_datetime(value: Any) -> datetime | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        parsed = parsedate_to_datetime(text)
        if parsed.tzinfo:
            return parsed.astimezone(timezone(timedelta(hours=9))).replace(tzinfo=None)
        return parsed
    except Exception:
        pass
    now = datetime.now()
    relative = re.search(r"(\d+)\s*(분|시간|일)\s*전", text)
    if relative:
        amount = int(relative.group(1))
        unit = relative.group(2)
        if unit == "분":
            return now - timedelta(minutes=amount)
        if unit == "시간":
            return now - timedelta(hours=amount)
        return now - timedelta(days=amount)
    match = re.search(r"(\d{4})[.년/-]\s*(\d{1,2})[.월/-]\s*(\d{1,2})", text)
    if match:
        try:
            return datetime(int(match.group(1)), int(match.group(2)), int(match.group(3)))
        except ValueError:
            return None
    return None


def decode_google_news_url(url: str) -> str:
    text = str(url or "").strip()
    if not text:
        return ""
    parsed = urlparse(text)
    query = parse_qs(parsed.query)
    for key in ("url", "u"):
        if query.get(key):
            return unquote(query[key][0])
    return text


def fetch_google_news(stock_name: str, days: int, limit: int) -> list[dict[str, Any]]:
    query = f'"{stock_name}" (실적 OR 수주 OR 계약 OR 공급 OR 투자 OR 공시 OR 증설 OR 인수 OR 임상 OR 정책) when:{days}d'
    response = requests.get(
        "https://news.google.com/rss/search",
        params={"q": query, "hl": "ko", "gl": "KR", "ceid": "KR:ko"},
        headers=news_headers(),
        timeout=12,
    )
    response.raise_for_status()
    root = ET.fromstring(response.content)
    items: list[dict[str, Any]] = []
    for item in root.findall(".//item")[: max(limit * 2, 20)]:
        raw_title = clean_news_text(item.findtext("title"))
        source = clean_news_text(item.findtext("source"))
        title = raw_title
        if source and title.endswith(" - " + source):
            title = title[: -(len(source) + 3)].strip()
        summary = clean_news_text(item.findtext("description"))
        published = parse_news_datetime(item.findtext("pubDate"))
        items.append(
            {
                "title": title,
                "summary": summary,
                "url": decode_google_news_url(item.findtext("link") or ""),
                "source": source or "Google News",
                "published_at": published.isoformat(timespec="seconds") if published else "",
                "source_type": "Google",
            }
        )
    return items


def fetch_naver_news(stock_name: str, days: int, limit: int) -> list[dict[str, Any]]:
    query = f"{stock_name} 실적 수주 계약 투자 공시"
    items: list[dict[str, Any]] = []
    for start in (1, 11, 21):
        if len(items) >= limit * 2:
            break
        response = requests.get(
            "https://search.naver.com/search.naver",
            params={"where": "news", "query": query, "sort": 1, "start": start},
            headers=news_headers(),
            timeout=10,
        )
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")
        for area in soup.select(".news_area"):
            title_link = area.select_one("a.news_tit")
            if not title_link:
                continue
            info_nodes = [clean_news_text(node.get_text(" ")) for node in area.select(".info_group .info")]
            date_text = ""
            source = ""
            for info in info_nodes:
                if re.search(r"전$|\d{4}[.년/-]", info):
                    date_text = info
                elif info and "네이버뉴스" not in info:
                    source = source or info
            published = parse_news_datetime(date_text)
            if published and published < datetime.now() - timedelta(days=days):
                continue
            items.append(
                {
                    "title": clean_news_text(title_link.get("title") or title_link.get_text(" ")),
                    "summary": clean_news_text((area.select_one(".news_dsc") or area).get_text(" ")),
                    "url": str(title_link.get("href") or "").strip(),
                    "source": source or "Naver News",
                    "published_at": published.isoformat(timespec="seconds") if published else "",
                    "source_type": "Naver",
                }
            )
    return items


def news_fingerprint(item: dict[str, Any], stock_name: str) -> str:
    text = normalize_text(item.get("title"))
    text = text.replace(normalize_text(stock_name), "")
    for token in NEWS_NOISE_KEYWORDS + NEWS_IMPORTANT_KEYWORDS:
        text = text.replace(normalize_text(token), "")
    text = re.sub(r"\d+", "", text)
    return text[:80]


def is_blocked_news_source(item: dict[str, Any]) -> bool:
    corpus = " ".join(
        [
            str(item.get("source") or ""),
            str(item.get("source_type") or ""),
            str(item.get("url") or ""),
            str(item.get("title") or ""),
        ]
    ).lower()
    normalized_corpus = normalize_text(corpus)
    return any(normalize_text(keyword) in normalized_corpus for keyword in NEWS_BLOCKED_SOURCE_KEYWORDS)


def classify_news_item(item: dict[str, Any], stock_name: str) -> tuple[bool, int, list[str]]:
    if is_blocked_news_source(item):
        return False, 0, []
    title = clean_news_text(item.get("title"))
    summary = clean_news_text(item.get("summary"))
    corpus = f"{title} {summary}".lower()
    normalized_corpus = normalize_text(corpus)
    normalized_stock = normalize_text(stock_name)
    if normalized_stock and normalized_stock not in normalized_corpus:
        return False, 0, []
    important_hits = [kw for kw in NEWS_IMPORTANT_KEYWORDS if normalize_text(kw) in normalized_corpus]
    noise_hits = [kw for kw in NEWS_NOISE_KEYWORDS if normalize_text(kw) in normalized_corpus]
    if not important_hits and noise_hits:
        return False, 0, noise_hits[:3]
    if not important_hits and len(title) < 24:
        return False, 0, []
    score = len(important_hits) * 3 - min(len(noise_hits), 3)
    if any(kw in important_hits for kw in ["실적", "영업익", "수주", "계약", "공시", "투자", "증설", "임상", "허가"]):
        score += 3
    return score > 0, max(score, 1), important_hits[:5]


def dedupe_news_items(items: list[dict[str, Any]], stock_name: str, limit: int) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []
    fingerprints: list[str] = []
    for raw in items:
        keep, score, tags = classify_news_item(raw, stock_name)
        if not keep:
            continue
        fingerprint = news_fingerprint(raw, stock_name) or normalize_text(raw.get("title"))[:80]
        duplicate = False
        for previous in fingerprints:
            if previous == fingerprint or SequenceMatcher(None, previous, fingerprint).ratio() >= 0.78:
                duplicate = True
                break
        if duplicate:
            continue
        fingerprints.append(fingerprint)
        item = dict(raw)
        item["importance_score"] = score
        item["reason_tags"] = tags
        item["published_date"] = str(item.get("published_at") or "")[:10]
        selected.append(item)
        if len(selected) >= limit:
            break
    return selected


def search_stock_news(query: str, limit: int = 30, days: int = 365) -> dict[str, Any]:
    raw_query = str(query or "").strip()
    if not raw_query:
        return {"query": "", "stock": None, "items": [], "filtered_count": 0}
    code_match = re.search(r"(\d{6})", raw_query)
    clean_query = re.sub(r"\(?\b\d{6}\b\)?", "", raw_query).strip()
    if code_match:
        resolved = resolve_stock_payload(code_match.group(1), clean_query)
        code = str((resolved or {}).get("code") or code_match.group(1)).zfill(6)
        resolved_name = str((resolved or {}).get("name") or clean_query).strip()
    else:
        code, resolved_name = resolve_stock(clean_query or raw_query)
    stock_name = resolved_name or clean_query or raw_query
    days = max(7, min(int(days or 365), 1825))
    limit = max(1, min(int(limit or 30), 80))
    collected: list[dict[str, Any]] = []
    errors: list[str] = []
    for fetcher in (fetch_naver_news, fetch_google_news):
        try:
            collected.extend(fetcher(stock_name, days, limit))
        except Exception as exc:
            errors.append(str(exc))
    collected.sort(key=lambda item: str(item.get("published_at") or ""), reverse=True)
    filtered = dedupe_news_items(collected, stock_name, limit)
    return {
        "query": raw_query,
        "stock": {"code": code or "", "name": stock_name},
        "items": filtered,
        "raw_count": len(collected),
        "filtered_count": len(filtered),
        "days": days,
        "errors": errors[:3],
    }


def find_tradingview_executable() -> Path | None:
    candidates = [
        Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "TradingView" / "TradingView.exe",
        Path(os.environ.get("LOCALAPPDATA", "")) / "TradingView" / "TradingView.exe",
        Path(os.environ.get("ProgramFiles", "")) / "TradingView" / "TradingView.exe",
        Path(os.environ.get("ProgramFiles(x86)", "")) / "TradingView" / "TradingView.exe",
    ]
    candidates.extend(Path("C:/Program Files/WindowsApps").glob("*TradingView*/TradingView.exe"))
    for candidate in candidates:
        try:
            if candidate.exists():
                return candidate
        except OSError:
            continue
    return None


def build_tradingview_symbol(stock_code: str | None, stock_name: str | None) -> tuple[str, str, str]:
    code = re.sub(r"\D", "", str(stock_code or "")).zfill(6)
    if code and code != "000000":
        return code, f"KRX:{code}", f"https://www.tradingview.com/chart/?symbol={quote(f'KRX:{code}', safe='')}"
    resolved_code, resolved_name = resolve_stock(str(stock_name or "").strip())
    if resolved_code:
        code = str(resolved_code).zfill(6)
        return code, f"KRX:{code}", f"https://www.tradingview.com/chart/?symbol={quote(f'KRX:{code}', safe='')}"
    name = str(stock_name or "").strip()
    if not name:
        raise ValueError("TradingView에서 열 종목을 찾을 수 없습니다.")
    return "", name, f"https://www.tradingview.com/chart/?symbol={quote(name, safe='')}"


@lru_cache(maxsize=512)
def load_stock_chart_preview_cached(stock_code: str, months: int = 3) -> dict[str, Any]:
    raw_code = str(stock_code or "").strip()
    numeric_code = re.sub(r"\D", "", raw_code).zfill(6)
    is_kr_code = bool(numeric_code and numeric_code != "000000" and re.fullmatch(r"\d{6}", numeric_code))
    code = numeric_code if is_kr_code else raw_code.upper()
    if not code:
        raise ValueError("차트 데이터를 가져올 종목코드를 찾을 수 없습니다.")
    months = max(1, min(int(months or 3), 12))
    disk_cached = load_chart_preview_disk_cache(code, months)
    if isinstance(disk_cached, dict):
        return disk_cached
    end_date = date.today()
    start_date = end_date - timedelta(days=max(45, months * 38))

    frame = fdr.DataReader(code, start_date.isoformat(), end_date.isoformat())
    if frame is None or frame.empty:
        if is_kr_code and SCREENING_FAST_DB_PATH.exists():
            ensure_screening_db_indexes()
            with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
                sql_rows = conn.execute(
                    """
                    SELECT file_date_key, close_price
                    FROM daily_close_cache
                    WHERE stock_code = ?
                      AND file_date_key BETWEEN ? AND ?
                    ORDER BY file_date_key ASC
                    """,
                    (code, start_date.strftime("%Y%m%d"), end_date.strftime("%Y%m%d")),
                ).fetchall()
            rows = []
            for file_date_key, close_price in sql_rows[-90:]:
                close = to_float(close_price)
                date_key = str(file_date_key or "")
                if not close or not re.fullmatch(r"20\d{6}", date_key):
                    continue
                date_text = f"{date_key[:4]}-{date_key[4:6]}-{date_key[6:]}"
                rows.append(
                    {
                        "date": date_text,
                        "open": close,
                        "high": close,
                        "low": close,
                        "close": close,
                        "volume": 0,
                    }
                )
            if not rows:
                raise ValueError("차트 데이터가 없습니다.")
        else:
            raise ValueError("차트 데이터가 없습니다.")
    else:
        frame = frame.reset_index().tail(90)
        date_column = None
        for candidate in ("Date", "date", "index"):
            if candidate in frame.columns:
                date_column = candidate
                break
        if date_column is None and len(frame.columns):
            date_column = str(frame.columns[0])
        rows = []
        for _, row in frame.iterrows():
            raw_date = row.get(date_column) if date_column else None
            if isinstance(raw_date, (datetime, date)):
                date_text = raw_date.strftime("%Y-%m-%d")
            else:
                date_text = str(raw_date)[:10]
            rows.append(
                {
                    "date": date_text,
                    "open": to_float(row.get("Open")),
                    "high": to_float(row.get("High")),
                    "low": to_float(row.get("Low")),
                    "close": to_float(row.get("Close")),
                    "volume": to_float(row.get("Volume")),
                }
            )
    valid_rows = [item for item in rows if all(item.get(key) is not None for key in ["open", "high", "low", "close"])]
    if not valid_rows:
        raise ValueError("유효한 차트 데이터가 없습니다.")
    first_close = float(valid_rows[0]["close"] or 0)
    last_close = float(valid_rows[-1]["close"] or 0)
    return_pct = ((last_close / first_close) - 1) * 100 if first_close else 0.0
    resolved = find_listing_row_by_code(code) if is_kr_code else {}
    payload = {
        "stock_code": code,
        "stock_name": resolved.get("name", "") if isinstance(resolved, dict) else "",
        "months": months,
        "rows": valid_rows,
        "summary": {
            "start_date": valid_rows[0]["date"],
            "end_date": valid_rows[-1]["date"],
            "last_close": round(last_close, 2),
            "return_pct": round(return_pct, 2),
            "point_count": len(valid_rows),
        },
    }
    save_chart_preview_disk_cache(code, months, payload)
    return payload


@lru_cache(maxsize=512)
def build_stock_sector_entry_markers(
    stock_code: str | None,
    start_date: str | None,
    end_date: str | None,
) -> list[dict[str, Any]]:
    code = normalize_stock_code_value(stock_code)
    if not code or not start_date or not end_date:
        return []
    try:
        start_dt = datetime.strptime(str(start_date), "%Y-%m-%d").date() - timedelta(days=7)
        end_dt = datetime.strptime(str(end_date), "%Y-%m-%d").date()
    except Exception:
        return []
    summaries = screening_backtest_source_summaries(start_date=start_dt, end_date=end_dt)
    if not summaries:
        return []
    sector_db = load_sector_db()
    markers: list[dict[str, Any]] = []
    min_score = 50.0
    trading_rank_limit = 20
    min_avg_score = 55.0
    min_strong_count = 2
    min_stock_count = 1
    for index_no, summary in enumerate(summaries):
        current_date = str(summary.get("file_date") or "")
        if not current_date or current_date < start_date or current_date > end_date:
            continue
        window = summaries[max(0, index_no - 4) : index_no + 1]
        trading_totals: dict[str, float] = {}
        for window_summary in window:
            for signal in build_sector_rotation_signals(window_summary, min_score=min_score, sector_db=sector_db):
                sector = str(signal.get("sector") or "").strip()
                if sector:
                    trading_totals[sector] = trading_totals.get(sector, 0.0) + float(signal.get("total_trading_value_100m") or 0.0)
        window_count = max(1, len(window))
        trading_rank_rows = sorted(
            [{"sector": sector, "avg_5d_trading_value_100m": value / window_count} for sector, value in trading_totals.items()],
            key=lambda item: item["avg_5d_trading_value_100m"],
            reverse=True,
        )
        rank_map = {
            item["sector"]: {"rank": rank + 1, "avg_5d_trading_value_100m": item["avg_5d_trading_value_100m"]}
            for rank, item in enumerate(trading_rank_rows)
        }
        for signal in build_sector_rotation_signals(summary, min_score=min_score, sector_db=sector_db):
            sector = str(signal.get("sector") or "").strip()
            rank_info = rank_map.get(sector, {})
            trading_rank = int(rank_info.get("rank") or 9999)
            if trading_rank > trading_rank_limit:
                continue
            if float(signal.get("avg_score") or 0.0) < min_avg_score:
                continue
            if int(signal.get("strong_count") or 0) < min_strong_count:
                continue
            if int(signal.get("stock_count") or 0) < min_stock_count:
                continue
            signal_stock_codes = {
                normalize_stock_code_value(stock.get("stock_code"))
                for stock in signal.get("stocks", [])
                if isinstance(stock, dict)
            }
            if code not in signal_stock_codes:
                continue
            entry_score = (
                float(signal.get("strength_score") or 0.0)
                + max(0.0, (trading_rank_limit + 1 - trading_rank)) * 2.0
                + min(float(signal.get("turnover_ratio_pct") or 0.0), 10.0) * 1.2
            )
            signal_level = "강한 진입" if trading_rank <= 5 and float(signal.get("avg_score") or 0) >= 70 else "진입"
            markers.append(
                {
                    "date": current_date,
                    "sector": sector,
                    "signal_level": signal_level,
                    "trading_rank": trading_rank,
                    "avg_score": signal.get("avg_score"),
                    "strong_count": signal.get("strong_count"),
                    "entry_score": round(entry_score, 2),
                }
            )
            break
    return markers


@lru_cache(maxsize=512)
def build_sector_entry_markers_for_sector(
    sector_name: str | None,
    start_date: str | None,
    end_date: str | None,
) -> list[dict[str, Any]]:
    sector_name = str(sector_name or "").strip()
    if not sector_name or not start_date or not end_date:
        return []
    try:
        start_dt = datetime.strptime(str(start_date), "%Y-%m-%d").date() - timedelta(days=7)
        end_dt = datetime.strptime(str(end_date), "%Y-%m-%d").date()
    except Exception:
        return []
    summaries = screening_backtest_source_summaries(start_date=start_dt, end_date=end_dt)
    if not summaries:
        return []
    sector_db = load_sector_db()
    markers: list[dict[str, Any]] = []
    min_score = 50.0
    trading_rank_limit = 20
    min_avg_score = 55.0
    min_strong_count = 2
    min_stock_count = 1
    for index_no, summary in enumerate(summaries):
        current_date = str(summary.get("file_date") or "")
        if not current_date or current_date < start_date or current_date > end_date:
            continue
        window = summaries[max(0, index_no - 4) : index_no + 1]
        trading_totals: dict[str, float] = {}
        for window_summary in window:
            for signal in build_sector_rotation_signals(window_summary, min_score=min_score, sector_db=sector_db):
                sector = str(signal.get("sector") or "").strip()
                if sector:
                    trading_totals[sector] = trading_totals.get(sector, 0.0) + float(signal.get("total_trading_value_100m") or 0.0)
        window_count = max(1, len(window))
        trading_rank_rows = sorted(
            [{"sector": sector, "avg_5d_trading_value_100m": value / window_count} for sector, value in trading_totals.items()],
            key=lambda item: item["avg_5d_trading_value_100m"],
            reverse=True,
        )
        rank_map = {
            item["sector"]: {"rank": rank + 1, "avg_5d_trading_value_100m": item["avg_5d_trading_value_100m"]}
            for rank, item in enumerate(trading_rank_rows)
        }
        for signal in build_sector_rotation_signals(summary, min_score=min_score, sector_db=sector_db):
            sector = str(signal.get("sector") or "").strip()
            if sector != sector_name:
                continue
            rank_info = rank_map.get(sector, {})
            trading_rank = int(rank_info.get("rank") or 9999)
            if trading_rank > trading_rank_limit:
                continue
            if float(signal.get("avg_score") or 0.0) < min_avg_score:
                continue
            if int(signal.get("strong_count") or 0) < min_strong_count:
                continue
            if int(signal.get("stock_count") or 0) < min_stock_count:
                continue
            entry_score = (
                float(signal.get("strength_score") or 0.0)
                + max(0.0, (trading_rank_limit + 1 - trading_rank)) * 2.0
                + min(float(signal.get("turnover_ratio_pct") or 0.0), 10.0) * 1.2
            )
            markers.append(
                {
                    "date": current_date,
                    "sector": sector,
                    "signal_level": "강한 진입" if trading_rank <= 5 and float(signal.get("avg_score") or 0) >= 70 else "진입",
                    "trading_rank": trading_rank,
                    "avg_score": signal.get("avg_score"),
                    "strong_count": signal.get("strong_count"),
                    "entry_score": round(entry_score, 2),
                }
            )
            break
    return markers


def load_stock_chart_preview(stock_code: str | None = None, stock_name: str | None = None, months: int = 3) -> dict[str, Any]:
    fallback_code = str(stock_code or "").strip()
    is_us_like_code = bool(re.fullmatch(r"[A-Za-z][A-Za-z0-9.\-]{0,11}", fallback_code))
    resolved = None if is_us_like_code else resolve_stock_payload(stock_code, stock_name)
    code = str((resolved or {}).get("code") or fallback_code or "").strip()
    if not code:
        raise ValueError("차트 데이터를 가져올 종목을 찾을 수 없습니다.")
    payload = load_stock_chart_preview_cached(code, months)
    payload = {
        **payload,
        "rows": [dict(row) for row in payload.get("rows", [])],
        "summary": dict(payload.get("summary", {})),
    }
    if is_us_like_code and stock_name:
        payload["stock_name"] = str(stock_name).strip()
    elif resolved and resolved.get("name"):
        payload["stock_name"] = resolved["name"]
    summary = payload.get("summary") or {}
    numeric_code = re.sub(r"\D", "", code).zfill(6)
    if numeric_code and numeric_code != "000000" and re.fullmatch(r"\d{6}", numeric_code):
        payload["entry_markers"] = build_stock_sector_entry_markers(
            numeric_code,
            str(summary.get("start_date") or ""),
            str(summary.get("end_date") or ""),
        )
    else:
        payload["entry_markers"] = []
    return payload


def build_sector_market_cap_chart(request: SectorMarketCapChartRequest) -> dict[str, Any]:
    sector = str(request.sector or "").strip()
    if not sector:
        raise ValueError("섹터명을 찾지 못했습니다.")
    months = max(1, min(int(request.months or 3), 12))
    end_date = date.today()
    start_date = end_date - timedelta(days=max(45, months * 38))
    resolved_items: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    for item in request.stocks or []:
        resolved = resolve_stock_payload(item.code, item.name)
        if not resolved:
            continue
        code = str(resolved.get("code") or "").zfill(6)
        if not code or code == "000000" or code in seen_codes:
            continue
        seen_codes.add(code)
        resolved_items.append(resolved)
    if not resolved_items:
        raise ValueError("섹터에 포함된 종목을 찾지 못했습니다.")

    series_by_date: dict[str, dict[str, float]] = {}
    included: list[dict[str, Any]] = []
    skipped: list[str] = []
    for item in resolved_items:
        code = str(item.get("code") or "").zfill(6)
        name = str(item.get("name") or code)
        shares = to_float(item.get("stocks"))
        close_from_listing = to_float(item.get("close"))
        marcap_from_listing = to_float(item.get("marcap"))
        if (not shares or shares <= 0) and marcap_from_listing and close_from_listing:
            shares = marcap_from_listing / close_from_listing
        if not shares or shares <= 0:
            skipped.append(name)
            continue
        try:
            frame = fdr.DataReader(code, start_date.isoformat(), end_date.isoformat())
        except Exception:
            skipped.append(name)
            continue
        if frame is None or frame.empty or "Close" not in frame.columns:
            skipped.append(name)
            continue
        frame = frame.reset_index()
        point_count = 0
        for _, row in frame.iterrows():
            close = to_float(row.get("Close"))
            if close is None:
                continue
            raw_date = row.get("Date")
            if isinstance(raw_date, (datetime, date)):
                date_text = raw_date.strftime("%Y-%m-%d")
            else:
                date_text = str(raw_date)[:10]
            bucket = series_by_date.setdefault(date_text, {"market_cap": 0.0, "stock_count": 0})
            bucket["market_cap"] += close * shares
            bucket["stock_count"] += 1
            point_count += 1
        if point_count:
            included.append({
                "stock_code": code,
                "stock_name": name,
                "shares": round(float(shares), 2),
            })
        else:
            skipped.append(name)

    rows: list[dict[str, Any]] = []
    for date_text in sorted(series_by_date):
        bucket = series_by_date[date_text]
        if bucket["market_cap"] <= 0:
            continue
        rows.append({
            "date": date_text,
            "market_cap_krw": round(bucket["market_cap"], 0),
            "market_cap_100m": round(bucket["market_cap"] / 100000000.0, 1),
            "stock_count": int(bucket["stock_count"]),
        })
    if not rows:
        raise ValueError("섹터 시가총액 차트 데이터를 만들 수 없습니다.")
    first_value = float(rows[0]["market_cap_krw"] or 0)
    last_value = float(rows[-1]["market_cap_krw"] or 0)
    return_pct = ((last_value / first_value) - 1) * 100 if first_value else 0.0
    return {
        "sector": sector,
        "months": months,
        "unit": "억원",
        "source": "FinanceDataReader 종가 × 상장주식수 합산",
        "rows": rows,
        "entry_markers": build_sector_entry_markers_for_sector(sector, rows[0]["date"], rows[-1]["date"]),
        "stocks": included,
        "skipped": skipped,
        "summary": {
            "start_date": rows[0]["date"],
            "end_date": rows[-1]["date"],
            "start_market_cap_100m": round(first_value / 100000000.0, 1),
            "last_market_cap_100m": round(last_value / 100000000.0, 1),
            "return_pct": round(return_pct, 2),
            "stock_count": len(included),
            "point_count": len(rows),
        },
    }


def calculate_rsi(series: pd.Series, period: int = 14) -> pd.Series:
    delta = series.diff()
    gain = delta.clip(lower=0)
    loss = -delta.clip(upper=0)
    avg_gain = gain.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    avg_loss = loss.ewm(alpha=1 / period, adjust=False, min_periods=period).mean()
    rs = avg_gain / avg_loss.replace(0, np.nan)
    rsi = 100 - (100 / (1 + rs))
    return rsi.fillna(50)


def safe_strategy_date(value: str | None, fallback: date) -> date:
    try:
        return datetime.strptime(str(value or ""), "%Y-%m-%d").date()
    except Exception:
        return fallback


def load_strategy_price_frame(symbol: str, start_date: date, end_date: date) -> pd.DataFrame:
    fetch_start = start_date - timedelta(days=320)
    frame = fdr.DataReader(symbol, fetch_start.isoformat(), end_date.isoformat())
    if frame is None or frame.empty:
        raise ValueError("지수 데이터를 가져오지 못했습니다.")
    frame = frame.copy().reset_index()
    if "Date" not in frame.columns:
        frame = frame.rename(columns={frame.columns[0]: "Date"})
    frame["Date"] = pd.to_datetime(frame["Date"]).dt.date
    for column in ["Open", "High", "Low", "Close"]:
        if column not in frame.columns:
            frame[column] = frame["Close"] if "Close" in frame.columns else np.nan
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    frame = frame.dropna(subset=["Date", "Close"]).sort_values("Date")
    frame["ma20"] = frame["Close"].rolling(20).mean()
    frame["ma60"] = frame["Close"].rolling(60).mean()
    frame["ma200"] = frame["Close"].rolling(200).mean()
    frame["rsi14"] = calculate_rsi(frame["Close"], 14)
    return frame[frame["Date"] >= start_date].reset_index(drop=True)


def load_eastmoney_futures_price_frame(secid: str, start_date: date, end_date: date) -> pd.DataFrame:
    start_text = start_date.strftime("%Y%m%d")
    end_text = end_date.strftime("%Y%m%d")
    response = requests.get(
        "https://push2his.eastmoney.com/api/qt/stock/kline/get",
        params={
            "secid": secid,
            "fields1": "f1,f2,f3,f4,f5,f6",
            "fields2": "f51,f52,f53,f54,f55,f56,f57,f58,f59,f60,f61",
            "klt": "101",
            "fqt": "0",
            "beg": start_text,
            "end": end_text,
        },
        headers={
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://quote.eastmoney.com/",
        },
        timeout=15,
    )
    response.raise_for_status()
    payload = response.json()
    klines = ((payload.get("data") or {}).get("klines") or [])
    rows = []
    for raw in klines:
        parts = str(raw).split(",")
        if len(parts) < 7:
            continue
        rows.append(
            {
                "Date": pd.to_datetime(parts[0], errors="coerce").date(),
                "Open": to_float(parts[1]),
                "Close": to_float(parts[2]),
                "High": to_float(parts[3]),
                "Low": to_float(parts[4]),
                "Volume": to_float(parts[5]) or 0,
                "Amount": to_float(parts[6]) or 0,
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        raise ValueError("Eastmoney 선물 가격 데이터가 없습니다.")
    frame = frame.dropna(subset=["Date", "Close"]).sort_values("Date")
    return frame.reset_index(drop=True)


def load_fred_price_frame(series_id: str, start_date: date, end_date: date) -> pd.DataFrame:
    safe_series = re.sub(r"[^A-Za-z0-9_-]", "", series_id)
    cache_path = FRED_PRICE_CACHE_DIR / f"{safe_series}.csv"
    csv_text = ""
    try:
        response = requests.get(
            "https://fred.stlouisfed.org/graph/fredgraph.csv",
            params={"id": series_id, "cosd": start_date.isoformat(), "coed": end_date.isoformat()},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=8,
        )
        response.raise_for_status()
        csv_text = response.text
        FRED_PRICE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(csv_text, encoding="utf-8")
    except Exception:
        if cache_path.exists():
            csv_text = cache_path.read_text(encoding="utf-8")
        else:
            raise

    rows = []
    reader = csv.DictReader(StringIO(csv_text))
    for row in reader:
        value = to_float(row.get(series_id))
        date_value = pd.to_datetime(row.get("observation_date") or row.get("DATE"), errors="coerce")
        if value is None or pd.isna(date_value):
            continue
        row_date = date_value.date()
        if row_date < start_date or row_date > end_date:
            continue
        rows.append(
            {
                "Date": row_date,
                "Open": value,
                "High": value,
                "Low": value,
                "Close": value,
                "Volume": 0,
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        raise ValueError(f"FRED 가격 데이터가 없습니다: {series_id}")
    frame = frame.dropna(subset=["Date", "Close"]).sort_values("Date")
    return frame.reset_index(drop=True)


ECONOMY_CYCLE_CATEGORY_META = {
    "fundamental": {
        "label": "펀더멘탈 강도",
        "description": "수출, 소비, 생산, 심리처럼 실물 경기 방향을 보여주는 지표입니다.",
    },
    "liquidity": {
        "label": "유동성 강도",
        "description": "시장을 밀어 올리는 자금 공급과 단기 유동성 압력을 보는 지표입니다.",
    },
    "breadth": {
        "label": "내부 마켓 Breadth",
        "description": "소수 대형주 착시가 아니라 시장 내부가 같이 강한지 확인하는 지표입니다.",
    },
    "risk": {
        "label": "위험 및 비용",
        "description": "신용 스프레드, 실질금리, 단기자금 비용처럼 시장의 브레이크를 보는 지표입니다.",
    },
}


ECONOMY_BREADTH_SYMBOLS = [
    "SPY",
    "QQQ",
    "IWM",
    "ACWI",
    "EWY",
    "EEM",
    "FEZ",
    "EWJ",
    "FXI",
]


ECONOMY_CYCLE_INDICATORS = [
    {
        "key": "korea_cli",
        "name": "한국 OECD 경기선행지수",
        "oecd_flow": "OECD.SDD.STES,DSD_STES@DF_CLI",
        "oecd_key": "KOR.M.LI.IX._Z.AA.IX._Z.H",
        "kind": "normal_100",
        "group": "선행",
        "source": "OECD",
        "description": "OECD CLI amplitude adjusted, 100이 장기 기준선입니다.",
    },
    {
        "key": "korea_manufacturing",
        "name": "제조업 생산",
        "oecd_flow": "OECD.SDD.STES,DSD_KEI@DF_KEI",
        "oecd_key": "KOR.M.PRVM.GR.F.Y.GY",
        "kind": "growth",
        "group": "실물",
        "source": "OECD",
        "description": "전년동월비 성장률을 최근 5년 분포로 표준화합니다.",
    },
    {
        "key": "korea_retail",
        "name": "소매판매",
        "oecd_flow": "OECD.SDD.STES,DSD_KEI@DF_KEI",
        "oecd_key": "KOR.M.TOVM.GR.G47.Y.GY",
        "kind": "growth",
        "group": "내수",
        "source": "OECD",
        "description": "전년동월비 성장률을 최근 5년 분포로 표준화합니다.",
    },
    {
        "key": "korea_export",
        "name": "수출액",
        "oecd_flow": "OECD.SDD.STES,DSD_KEI@DF_KEI",
        "oecd_key": "KOR.M.EX.GR._T.Y.GY",
        "kind": "growth",
        "group": "수출",
        "source": "OECD",
        "description": "수출액 전년동월비 성장률을 표준화합니다.",
    },
    {
        "key": "korea_consumer",
        "name": "소비자기대",
        "oecd_flow": "OECD.SDD.STES,DSD_STES@DF_CLI",
        "oecd_key": "KOR.M.CCICP.IX._Z.AA.IX._Z.H",
        "kind": "normal_100",
        "group": "심리",
        "source": "OECD",
        "description": "표준화 소비자기대지수입니다.",
    },
    {
        "key": "korea_business",
        "name": "기업심리",
        "oecd_flow": "OECD.SDD.STES,DSD_STES@DF_CLI",
        "oecd_key": "KOR.M.BCICP.IX._Z.AA.IX._Z.H",
        "kind": "normal_100",
        "group": "심리",
        "source": "OECD",
        "description": "표준화 기업심리지수입니다.",
    },
    {
        "key": "usa_cli",
        "name": "미국 OECD 경기선행지수",
        "oecd_flow": "OECD.SDD.STES,DSD_STES@DF_CLI",
        "oecd_key": "USA.M.LI.IX._Z.AA.IX._Z.H",
        "kind": "normal_100",
        "group": "글로벌",
        "source": "OECD",
        "description": "미국 경기 방향을 함께 비교합니다.",
    },
    {
        "key": "china_cli",
        "name": "중국 OECD 경기선행지수",
        "oecd_flow": "OECD.SDD.STES,DSD_STES@DF_CLI",
        "oecd_key": "CHN.M.LI.IX._Z.AA.IX._Z.H",
        "kind": "normal_100",
        "group": "글로벌",
        "source": "OECD",
        "description": "중국 경기 방향을 함께 비교합니다.",
    },
    {
        "key": "us_m2_yoy",
        "name": "미국 M2 통화량 YoY",
        "fred": "M2SL",
        "csv_url": "https://eco3min.fr/dataset/us-m2-money-supply.csv",
        "date_column": "date",
        "value_column": "m2_billions",
        "kind": "raw_yoy",
        "group": "유동성",
        "category_key": "liquidity",
        "source": "FRED/Eco3min",
        "description": "미국 M2 전년동월비입니다. 유동성 증가율이 개선되면 위험자산 환경을 우호적으로 봅니다.",
        "series_points": 60,
    },
    {
        "key": "us_rrp_balance",
        "name": "미국 역레포(RRP) 잔고",
        "csv_url": "https://eco3min.fr/wp-content/uploads/2026/03/us-net-liquidity-index-2003-present.csv",
        "date_column": "date",
        "value_column": "rrp_t",
        "kind": "risk_inverse",
        "group": "유동성",
        "category_key": "liquidity",
        "source": "FRED/Eco3min",
        "description": "RRP 잔고가 줄어들수록 시중 유동성 방출로 해석해 우호적으로 계산합니다.",
        "display_unit": "조$",
        "series_points": 180,
    },
    {
        "key": "us_tga_balance",
        "name": "미국 TGA 잔고",
        "csv_url": "https://eco3min.fr/wp-content/uploads/2026/03/us-net-liquidity-index-2003-present.csv",
        "date_column": "date",
        "value_column": "tga_t",
        "kind": "risk_inverse",
        "group": "유동성",
        "category_key": "liquidity",
        "source": "FRED/Eco3min",
        "description": "재무부 일반계정(TGA)이 줄어들수록 유동성 방출로 해석해 우호적으로 계산합니다.",
        "display_unit": "조$",
        "series_points": 180,
    },
    {
        "key": "us_net_liquidity_yoy",
        "name": "미국 순유동성 YoY",
        "csv_url": "https://eco3min.fr/wp-content/uploads/2026/03/us-net-liquidity-index-2003-present.csv",
        "date_column": "date",
        "value_column": "net_liq_yoy_pct",
        "kind": "growth",
        "group": "유동성",
        "category_key": "liquidity",
        "source": "FRED/Eco3min",
        "description": "글로벌 신용자극지수의 직접 데이터 대신 Fed balance sheet - TGA - RRP로 계산한 순유동성 YoY를 프록시로 사용합니다.",
        "display_unit": "YoY %",
        "series_points": 180,
    },
    {
        "key": "global_etf_above_ma200",
        "name": "글로벌 ETF 200일선 상회율",
        "breadth_symbols": ECONOMY_BREADTH_SYMBOLS,
        "breadth_metric": "above_ma200",
        "kind": "level_neutral",
        "neutral": 50,
        "group": "Breadth",
        "category_key": "breadth",
        "source": "Yahoo Finance 계산",
        "description": "주요 글로벌 ETF 바스켓 중 200일선 위에 있는 비율입니다. 50%를 기준선으로 봅니다.",
        "display_unit": "%",
        "series_points": 180,
    },
    {
        "key": "global_etf_new_high_low",
        "name": "글로벌 ETF 신고가-신저가",
        "breadth_symbols": ECONOMY_BREADTH_SYMBOLS,
        "breadth_metric": "new_high_low",
        "kind": "level_neutral",
        "neutral": 0,
        "group": "Breadth",
        "category_key": "breadth",
        "source": "Yahoo Finance 계산",
        "description": "주요 ETF 바스켓의 52주 신고가 비율에서 신저가 비율을 뺀 프록시 지표입니다.",
        "display_unit": "%p",
        "series_points": 180,
    },
    {
        "key": "usdkrw",
        "name": "원/달러 환율",
        "yahoo_symbol": "KRW=X",
        "kind": "risk_inverse",
        "group": "위험선호",
        "category_key": "risk",
        "source": "Yahoo Finance",
        "description": "원/달러 환율이 낮아지거나 하락세일수록 위험선호 개선으로 계산합니다.",
        "display_unit": "원",
        "series_points": 120,
    },
    {
        "key": "us_high_yield_spread",
        "name": "미국 하이일드 스프레드",
        "fred": "BAMLH0A0HYM2",
        "csv_url": "https://eco3min.fr/dataset/us-high-yield-spread.csv",
        "date_column": "date",
        "value_column": "hy_spread",
        "kind": "risk_inverse",
        "group": "위험선호",
        "category_key": "risk",
        "source": "FRED/Eco3min",
        "description": "스프레드가 낮아지거나 축소될수록 신용위험 완화와 위험선호 개선으로 계산합니다.",
        "display_unit": "%p",
        "series_points": 120,
    },
    {
        "key": "us_financial_conditions",
        "name": "미국 금융여건지수(NFCI)",
        "csv_url": "https://eco3min.fr/dataset/financial-conditions-index.csv",
        "date_column": "date",
        "value_column": "nfci",
        "kind": "risk_inverse",
        "group": "위험선호",
        "category_key": "risk",
        "source": "FRED/Eco3min",
        "description": "TED/SOFR-FF 같은 단기자금 압력을 포함한 금융여건 종합 프록시입니다. 상승할수록 긴축적으로 봅니다.",
        "display_unit": "지수",
        "series_points": 180,
    },
    {
        "key": "us_real_fed_funds",
        "name": "미국 실질 정책금리",
        "csv_url": "https://eco3min.fr/dataset/real-fed-funds-rate.csv",
        "date_column": "date",
        "value_column": "real_fed_funds",
        "kind": "risk_inverse",
        "group": "위험선호",
        "category_key": "risk",
        "source": "FRED/Eco3min",
        "description": "Fed Funds에서 CPI YoY를 뺀 실질 정책금리입니다. 상승할수록 조달 비용 부담이 커지는 것으로 계산합니다.",
        "display_unit": "%",
        "series_points": 180,
    },
]


def load_oecd_sdmx_csv_frame(flow: str, key: str, start_period: str) -> pd.DataFrame:
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", f"{flow}_{key}_{start_period}")[:180]
    cache_path = STATE_DIR / "oecd_cycle_cache" / f"{safe_name}.csv"
    csv_text = ""
    url = f"https://sdmx.oecd.org/public/rest/data/{flow}/{key}"
    try:
        response = requests.get(
            url,
            params={"startPeriod": start_period, "format": "csvfile"},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=18,
        )
        response.raise_for_status()
        csv_text = response.text
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(csv_text, encoding="utf-8")
    except Exception:
        if cache_path.exists():
            csv_text = cache_path.read_text(encoding="utf-8")
        else:
            raise
    frame = pd.read_csv(StringIO(csv_text))
    if frame.empty:
        raise ValueError("OECD 시계열 데이터가 비어 있습니다.")
    rows = []
    for _, row in frame.iterrows():
        period = str(row.get("TIME_PERIOD") or "")
        value = to_float(row.get("OBS_VALUE"))
        if value is None or not period:
            continue
        parsed_date = pd.to_datetime(period + "-01" if re.fullmatch(r"\d{4}-\d{2}", period) else period, errors="coerce")
        if pd.isna(parsed_date):
            continue
        rows.append({"Date": parsed_date.date(), "Close": value})
    result = pd.DataFrame(rows)
    if result.empty:
        raise ValueError("OECD 시계열 값을 찾지 못했습니다.")
    return result.dropna(subset=["Date", "Close"]).sort_values("Date").reset_index(drop=True)


def load_economy_external_csv_frame(config: dict[str, Any], start_date: date, end_date: date) -> pd.DataFrame:
    url = str(config.get("csv_url") or "")
    if not url:
        raise ValueError("외부 CSV URL이 없습니다.")
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", url)[:180]
    cache_path = STATE_DIR / "economy_external_csv_cache" / f"{safe_name}.csv"
    csv_text = ""
    try:
        response = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
        response.raise_for_status()
        csv_text = response.text
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(csv_text, encoding="utf-8")
    except Exception:
        if cache_path.exists():
            csv_text = cache_path.read_text(encoding="utf-8")
        else:
            raise
    frame = pd.read_csv(StringIO(csv_text))
    date_column = str(config.get("date_column") or "date")
    value_column = str(config.get("value_column") or "value")
    if date_column not in frame.columns or value_column not in frame.columns:
        raise ValueError(f"CSV 컬럼을 찾지 못했습니다: {date_column}, {value_column}")
    rows = []
    for _, row in frame.iterrows():
        row_date_value = pd.to_datetime(row.get(date_column), errors="coerce")
        close = to_float(row.get(value_column))
        if close is None or pd.isna(row_date_value):
            continue
        row_date = row_date_value.date()
        if row_date < start_date or row_date > end_date:
            continue
        rows.append({"Date": row_date, "Close": close})
    result = pd.DataFrame(rows)
    if result.empty:
        raise ValueError("외부 CSV 시계열 데이터가 없습니다.")
    return result.dropna(subset=["Date", "Close"]).sort_values("Date").reset_index(drop=True)


def load_fred_spread_frame(config: dict[str, Any], start_date: date, end_date: date) -> pd.DataFrame:
    series_ids = list(config.get("fred_spread") or [])
    if len(series_ids) != 2:
        raise ValueError("FRED 스프레드 설정은 2개 시리즈가 필요합니다.")
    left = load_fred_price_frame(str(series_ids[0]), start_date, end_date)
    right = load_fred_price_frame(str(series_ids[1]), start_date, end_date)
    left_frame = left[["Date", "Close"]].copy()
    right_frame = right[["Date", "Close"]].copy()
    left_frame["Date"] = pd.to_datetime(left_frame["Date"], errors="coerce")
    right_frame["Date"] = pd.to_datetime(right_frame["Date"], errors="coerce")
    left_frame = left_frame.dropna(subset=["Date"]).sort_values("Date")
    right_frame = right_frame.dropna(subset=["Date"]).sort_values("Date")
    merged = pd.merge_asof(
        left_frame.rename(columns={"Close": "left_close"}),
        right_frame.rename(columns={"Close": "right_close"}),
        on="Date",
        direction="backward",
    )
    merged["Close"] = pd.to_numeric(merged["left_close"], errors="coerce") - pd.to_numeric(merged["right_close"], errors="coerce")
    merged = merged.dropna(subset=["Close"])
    if merged.empty:
        raise ValueError("FRED 스프레드 계산 데이터가 없습니다.")
    return pd.DataFrame({"Date": merged["Date"].dt.date, "Close": merged["Close"]}).reset_index(drop=True)


def load_market_breadth_frame(config: dict[str, Any], start_date: date, end_date: date) -> pd.DataFrame:
    raw_symbols = config.get("breadth_symbols") or []
    if isinstance(raw_symbols, str):
        raw_symbols = [raw_symbols]
    symbols = [str(symbol).strip() for symbol in raw_symbols if str(symbol).strip()]
    if not symbols:
        raise ValueError("Breadth 계산 대상 심볼이 없습니다.")
    fetch_start = start_date - timedelta(days=430)
    rows: list[dict[str, Any]] = []
    metric = str(config.get("breadth_metric") or "above_ma200")
    for symbol in symbols:
        try:
            frame = load_global_index_price_frame(symbol, fetch_start, end_date)
        except Exception:
            continue
        if frame.empty:
            continue
        series = frame[["Date", "Close"]].copy()
        series["Date"] = pd.to_datetime(series["Date"], errors="coerce")
        series["Close"] = pd.to_numeric(series["Close"], errors="coerce")
        series = series.dropna(subset=["Date", "Close"]).sort_values("Date").reset_index(drop=True)
        if len(series) < 160:
            continue
        series["ma200"] = series["Close"].rolling(200, min_periods=140).mean()
        series["high252"] = series["Close"].rolling(252, min_periods=160).max()
        series["low252"] = series["Close"].rolling(252, min_periods=160).min()
        series = series[series["Date"].dt.date >= start_date]
        for _, row in series.iterrows():
            close = to_float(row.get("Close"))
            ma200 = to_float(row.get("ma200"))
            high252 = to_float(row.get("high252"))
            low252 = to_float(row.get("low252"))
            if close is None:
                continue
            above = 1.0 if ma200 is not None and close > ma200 else 0.0
            high_signal = 1.0 if high252 is not None and close >= high252 * 0.995 else 0.0
            low_signal = 1.0 if low252 is not None and close <= low252 * 1.005 else 0.0
            rows.append(
                {
                    "Date": pd.to_datetime(row["Date"]).date(),
                    "above_ma200": above,
                    "new_high": high_signal,
                    "new_low": low_signal,
                }
            )
    frame = pd.DataFrame(rows)
    if frame.empty:
        raise ValueError("Breadth 계산 데이터가 없습니다.")
    grouped = frame.groupby("Date", as_index=False).mean(numeric_only=True)
    if metric == "new_high_low":
        grouped["Close"] = (grouped["new_high"] - grouped["new_low"]) * 100.0
    else:
        grouped["Close"] = grouped["above_ma200"] * 100.0
    result = grouped[["Date", "Close"]].dropna(subset=["Close"]).sort_values("Date").reset_index(drop=True)
    if result.empty:
        raise ValueError("Breadth 계산 결과가 없습니다.")
    return result


def economy_category_key(config: dict[str, Any]) -> str:
    explicit = str(config.get("category_key") or "").strip()
    if explicit in ECONOMY_CYCLE_CATEGORY_META:
        return explicit
    group = str(config.get("group") or "")
    if "유동성" in group:
        return "liquidity"
    if "Breadth" in group:
        return "breadth"
    if "위험" in group:
        return "risk"
    return "fundamental"


def economy_phase_from_xy(x_value: float, y_value: float) -> str:
    if x_value >= 0 and y_value >= 0:
        return "상승"
    if x_value >= 0 and y_value < 0:
        return "둔화"
    if x_value < 0 and y_value < 0:
        return "하강"
    return "회복"


def economy_phase_comment(phase: str) -> str:
    return {
        "회복": "기준선 아래에서 모멘텀이 개선되는 구간입니다. 주식시장에는 보통 가장 관심을 둘 만한 초기 회복 신호로 해석합니다.",
        "상승": "기준선 위에서 모멘텀이 계속 개선되는 구간입니다. 경기와 위험자산 환경이 가장 우호적인 쪽입니다.",
        "둔화": "기준선 위에 있지만 모멘텀이 꺾이는 구간입니다. 주도주 압축과 리스크 관리가 중요해지는 구간입니다.",
        "하강": "기준선 아래에서 모멘텀도 악화되는 구간입니다. 현금/방어 비중을 높이는 판단이 필요한 구간입니다.",
    }.get(phase, "")


def clip_cycle_value(value: float, limit: float = 4.0) -> float:
    if not math.isfinite(value):
        return 0.0
    return max(-limit, min(limit, value))


def build_economy_cycle_indicator(config: dict[str, Any], start_date: date, end_date: date) -> dict[str, Any] | None:
    if config.get("oecd_flow") and config.get("oecd_key"):
        frame = load_oecd_sdmx_csv_frame(str(config["oecd_flow"]), str(config["oecd_key"]), f"{start_date.year}-01")
        frame = frame[(pd.to_datetime(frame["Date"], errors="coerce").dt.date >= start_date) & (pd.to_datetime(frame["Date"], errors="coerce").dt.date <= end_date)]
    elif config.get("csv_url"):
        frame = load_economy_external_csv_frame(config, start_date, end_date)
    elif config.get("yahoo_symbol"):
        frame = load_global_index_price_frame(str(config["yahoo_symbol"]), start_date, end_date)
    elif config.get("fred_spread"):
        frame = load_fred_spread_frame(config, start_date, end_date)
    elif config.get("breadth_symbols"):
        frame = load_market_breadth_frame(config, start_date, end_date)
    else:
        frame = load_fred_price_frame(str(config["fred"]), start_date, end_date)
    if frame.empty:
        return None
    series = frame[["Date", "Close"]].copy()
    series["Date"] = pd.to_datetime(series["Date"], errors="coerce")
    series["value"] = pd.to_numeric(series["Close"], errors="coerce")
    series = series.dropna(subset=["Date", "value"]).sort_values("Date").reset_index(drop=True)
    if len(series) < 6:
        return None

    kind = str(config.get("kind") or "normal_100")
    if kind == "raw_yoy":
        periods = int(to_float(config.get("yoy_periods")) or 12)
        series["metric"] = series["value"].pct_change(max(1, periods)) * 100.0
    else:
        series["metric"] = series["value"]
    series = series.dropna(subset=["metric"]).reset_index(drop=True)
    if len(series) < 6:
        return None

    latest = series.iloc[-1]
    latest_metric = float(latest["metric"])
    previous_metric = float(series.iloc[-2]["metric"]) if len(series) >= 2 else latest_metric
    three_month_metric = float(series.iloc[-4]["metric"]) if len(series) >= 4 else previous_metric
    window = series.tail(min(60, len(series)))
    std_value = float(window["metric"].std()) if len(window) >= 3 else 0.0
    if not math.isfinite(std_value) or std_value <= 1e-9:
        std_value = 1.0

    if kind == "normal_100":
        x_value = (latest_metric - 100.0) * 1.2
        y_value = (latest_metric - three_month_metric) * 4.0
        display_value = latest_metric
        display_unit = "지수"
    elif kind == "risk_inverse":
        mean_value = float(window["metric"].mean())
        x_value = ((mean_value - latest_metric) / std_value) * 1.2
        y_value = ((three_month_metric - latest_metric) / std_value) * 1.4
        display_value = latest_metric
        display_unit = str(config.get("display_unit") or "")
    elif kind == "level_neutral":
        neutral = float(to_float(config.get("neutral")) if to_float(config.get("neutral")) is not None else 0.0)
        x_value = ((latest_metric - neutral) / std_value) * 1.2
        y_value = ((latest_metric - three_month_metric) / std_value) * 1.4
        display_value = latest_metric
        display_unit = str(config.get("display_unit") or "")
    else:
        mean_value = float(window["metric"].mean())
        x_value = ((latest_metric - mean_value) / std_value) * 1.2
        y_value = ((latest_metric - three_month_metric) / std_value) * 1.4
        display_value = latest_metric
        display_unit = str(config.get("display_unit") or "YoY %")

    x_value = clip_cycle_value(x_value)
    y_value = clip_cycle_value(y_value)
    phase = economy_phase_from_xy(x_value, y_value)
    latest_date = latest["Date"].date()
    stale_days = (date.today() - latest_date).days
    category_key = economy_category_key(config)
    category = ECONOMY_CYCLE_CATEGORY_META.get(category_key, ECONOMY_CYCLE_CATEGORY_META["fundamental"])
    rows = []
    series_points = int(to_float(config.get("series_points")) or 36)
    for _, row in series.tail(max(12, min(240, series_points))).iterrows():
        metric = to_float(row.get("metric"))
        raw_value = to_float(row.get("value"))
        row_date = row.get("Date")
        if metric is None or pd.isna(row_date):
            continue
        rows.append(
            {
                "date": pd.to_datetime(row_date).date().isoformat(),
                "value": round(raw_value if raw_value is not None else metric, 4),
                "metric": round(metric, 4),
            }
        )

    return {
        "key": config["key"],
        "name": config["name"],
        "group": config.get("group", ""),
        "category_key": category_key,
        "category": category["label"],
        "category_description": category["description"],
        "fred": config.get("fred", ""),
        "oecd_key": config.get("oecd_key", ""),
        "source": config.get("source", "FRED"),
        "kind": kind,
        "description": config.get("description", ""),
        "latest_date": latest_date.isoformat(),
        "latest_value": round(display_value, 4),
        "display_unit": display_unit,
        "mom_change": round(latest_metric - previous_metric, 4),
        "three_month_change": round(latest_metric - three_month_metric, 4),
        "favorable_three_month_change": round(
            (three_month_metric - latest_metric) if kind == "risk_inverse" else (latest_metric - three_month_metric),
            4,
        ),
        "x": round(x_value, 3),
        "y": round(y_value, 3),
        "phase": phase,
        "phase_comment": economy_phase_comment(phase),
        "stale": stale_days > 240,
        "stale_days": stale_days,
        "series": rows,
    }


def build_economy_cycle_payload(force_refresh: bool = False) -> dict[str, Any]:
    if not force_refresh and ECONOMY_CYCLE_CACHE_PATH.exists():
        try:
            cached = json.loads(ECONOMY_CYCLE_CACHE_PATH.read_text(encoding="utf-8"))
            cached_at = datetime.fromisoformat(str(cached.get("loaded_at", "")).replace("Z", "+00:00"))
            if datetime.now(cached_at.tzinfo) - cached_at < timedelta(hours=12):
                return cached
        except Exception:
            pass

    end_date = date.today()
    start_date = end_date - timedelta(days=365 * 10)
    indicators: list[dict[str, Any]] = []
    errors: list[str] = []
    for config in ECONOMY_CYCLE_INDICATORS:
        try:
            indicator = build_economy_cycle_indicator(config, start_date, end_date)
            if indicator:
                indicators.append(indicator)
        except Exception as exc:
            errors.append(f"{config.get('name')}: {exc}")

    phase_order = ["회복", "상승", "둔화", "하강"]
    phase_counts = {phase: 0 for phase in phase_order}
    for item in indicators:
        phase_counts[item.get("phase", "")] = phase_counts.get(item.get("phase", ""), 0) + 1
    category_counts: dict[str, int] = {meta["label"]: 0 for meta in ECONOMY_CYCLE_CATEGORY_META.values()}
    for item in indicators:
        category_label = str(item.get("category") or "")
        category_counts[category_label] = category_counts.get(category_label, 0) + 1

    core_indicators = [item for item in indicators if item.get("group") != "글로벌"]
    average_x = float(np.mean([float(item.get("x") or 0) for item in core_indicators])) if core_indicators else 0.0
    average_y = float(np.mean([float(item.get("y") or 0) for item in core_indicators])) if core_indicators else 0.0
    current_phase = economy_phase_from_xy(average_x, average_y)
    sorted_phases = sorted(phase_counts.items(), key=lambda item: item[1], reverse=True)
    dominant_phase = sorted_phases[0][0] if sorted_phases and sorted_phases[0][1] > 0 else current_phase

    latest_dates = [item.get("latest_date") for item in indicators if item.get("latest_date")]
    payload = {
        "loaded_at": datetime.now(timezone.utc).isoformat(),
        "source_label": "OECD Data Explorer / FRED",
        "method_note": "100 기준 지표는 기준선 대비 위치와 3개월 변화, 성장률·유동성 지표는 최근 5년 z-score와 3개월 모멘텀으로 계산합니다. RRP/TGA·환율·스프레드·실질금리는 낮아질수록 우호적으로 보도록 방향을 반대로 적용합니다. 글로벌 신용자극과 Breadth는 무료 공개 데이터 기반 프록시입니다.",
        "latest_date": max(latest_dates) if latest_dates else "",
        "current_phase": current_phase,
        "dominant_phase": dominant_phase,
        "phase_counts": phase_counts,
        "category_counts": category_counts,
        "categories": [
            {"key": key, "label": meta["label"], "description": meta["description"]}
            for key, meta in ECONOMY_CYCLE_CATEGORY_META.items()
        ],
        "average_x": round(average_x, 3),
        "average_y": round(average_y, 3),
        "summary": economy_phase_comment(current_phase),
        "indicators": indicators,
        "errors": errors,
        "sources": [
            {"label": "OECD Composite Leading Indicators", "url": "https://www.oecd.org/en/data/datasets/oecd-composite-leading-indicators-clis.html"},
            {"label": "OECD API guide", "url": "https://www.oecd.org/en/data/insights/data-explainers/2024/09/api.html"},
            {"label": "OECD Composite Leading Indicator", "url": "https://www.oecd.org/en/data/indicators/composite-leading-indicator-cli.html"},
            {"label": "FRED Economic Data", "url": "https://fred.stlouisfed.org/"},
        ],
    }
    try:
        ECONOMY_CYCLE_CACHE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass
    return payload


def load_bls_price_frame(series_id: str, start_date: date, end_date: date) -> pd.DataFrame:
    payload = {
        "seriesid": [series_id],
        "startyear": str(start_date.year),
        "endyear": str(end_date.year),
    }
    response = requests.post(
        "https://api.bls.gov/publicAPI/v2/timeseries/data/",
        json=payload,
        headers={"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"},
        timeout=20,
    )
    response.raise_for_status()
    body = response.json()
    series_payload = ((body.get("Results") or {}).get("series") or [{}])[0]
    rows = []
    for row in series_payload.get("data") or []:
        period = str(row.get("period") or "")
        if not re.fullmatch(r"M\d{2}", period):
            continue
        value = to_float(row.get("value"))
        if value is None:
            continue
        row_date = date(int(row.get("year")), int(period[1:]), 1)
        if row_date < start_date.replace(day=1) or row_date > end_date:
            continue
        rows.append(
            {
                "Date": row_date,
                "Open": value,
                "High": value,
                "Low": value,
                "Close": value,
                "Volume": 0,
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        raise ValueError(f"BLS 가격 데이터가 없습니다: {series_id}")
    frame = frame.dropna(subset=["Date", "Close"]).sort_values("Date")
    return frame.reset_index(drop=True)


def decode_tradingeconomics_chart_payload(text: str) -> dict[str, Any]:
    payload = json.loads(text)
    if isinstance(payload, dict):
        return payload
    if not isinstance(payload, str):
        raise ValueError("Trading Economics 응답 형식이 올바르지 않습니다.")
    raw = base64.b64decode(payload)
    key = b"tradingeconomics-charts-core-api-key"
    decoded = bytearray(raw)
    for index in range(len(decoded)):
        decoded[index] ^= key[index % len(key)]
    return json.loads(zlib.decompress(bytes(decoded), 15 + 32).decode("utf-8"))


def load_tradingeconomics_market_price_frame(symbol: str, start_date: date, end_date: date) -> pd.DataFrame:
    safe_symbol = re.sub(r"[^A-Za-z0-9_-]", "_", symbol.upper())
    cache_path = TRADINGECONOMICS_PRICE_CACHE_DIR / f"{safe_symbol}.json"
    body: dict[str, Any] | None = None
    try:
        response = requests.get(
            f"https://d3ii0wo49og5mi.cloudfront.net/markets/{quote(symbol)}",
            params={
                "d1": start_date.isoformat(),
                "d2": end_date.isoformat(),
                "interval": "1d",
                "ohlc": "0",
            },
            headers={
                "User-Agent": "Mozilla/5.0",
                "Referer": "https://tradingeconomics.com/commodity/molybden",
            },
            timeout=20,
        )
        response.raise_for_status()
        body = decode_tradingeconomics_chart_payload(response.text)
        TRADINGECONOMICS_PRICE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(json.dumps(body, ensure_ascii=False), encoding="utf-8")
    except Exception:
        if cache_path.exists():
            body = json.loads(cache_path.read_text(encoding="utf-8"))
        else:
            raise

    rows = []
    series_list = (body or {}).get("series") or []
    points = (series_list[0] or {}).get("data") if series_list else []
    for point in points or []:
        if not isinstance(point, list) or len(point) < 2:
            continue
        close = to_float(point[1])
        timestamp = to_float(point[0])
        if close is None or timestamp is None:
            continue
        row_date = datetime.fromtimestamp(timestamp, timezone.utc).date()
        if row_date < start_date or row_date > end_date:
            continue
        rows.append(
            {
                "Date": row_date,
                "Open": close,
                "High": close,
                "Low": close,
                "Close": close,
                "Volume": 0,
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        raise ValueError(f"Trading Economics 가격 데이터가 없습니다: {symbol}")
    frame = frame.dropna(subset=["Date", "Close"]).drop_duplicates(subset=["Date"], keep="last").sort_values("Date")
    return frame.reset_index(drop=True)


def load_smm_spot_price_frame(product_id: str, start_date: date, end_date: date) -> pd.DataFrame:
    safe_product_id = re.sub(r"[^A-Za-z0-9_-]", "_", str(product_id))
    cache_path = SMM_PRICE_CACHE_DIR / f"{safe_product_id}.json"
    body: dict[str, Any] | None = None
    try:
        response = requests.get(
            "https://www-old.metal.com/api/spotcenter/get_mixen_history_prices",
            params={
                "token": "",
                "id": product_id,
                "beginDate": start_date.isoformat(),
                "endDate": end_date.isoformat(),
                "currency_type": "CNY",
            },
            headers={
                "User-Agent": "Mozilla/5.0",
                "Referer": f"https://www-old.metal.com/lithium/{product_id}",
            },
            timeout=20,
        )
        response.raise_for_status()
        body = response.json()
        SMM_PRICE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(json.dumps(body, ensure_ascii=False), encoding="utf-8")
    except Exception:
        if cache_path.exists():
            body = json.loads(cache_path.read_text(encoding="utf-8"))
        else:
            raise

    rows = []
    for row in (body or {}).get("data") or []:
        row_date_value = pd.to_datetime(row.get("RenewDate"), errors="coerce")
        close = to_float(row.get("Average"))
        if pd.isna(row_date_value) or close is None:
            continue
        row_date = row_date_value.date()
        if row_date < start_date or row_date > end_date:
            continue
        rows.append(
            {
                "Date": row_date,
                "Open": close,
                "High": to_float(row.get("Highs")) or close,
                "Low": to_float(row.get("Low")) or close,
                "Close": close,
                "Volume": 0,
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        raise ValueError(f"SMM 가격 데이터가 없습니다: {product_id}")
    frame = frame.dropna(subset=["Date", "Close"]).drop_duplicates(subset=["Date"], keep="last").sort_values("Date")
    return frame.reset_index(drop=True)


def komis_referer_for_hp(hp_code: str) -> str:
    page_by_hp = {
        "HP001": "BaseMetals",
        "HP002": "MinorMetals",
        "HP003": "IronOre",
        "HP004": "EtcMnrl",
    }
    page = page_by_hp.get(str(hp_code).upper(), "BaseMetals")
    return f"https://www.komis.or.kr/Komis/RsrcPrice/{page}"


def load_komis_mineral_price_frame(symbol: str, start_date: date, end_date: date) -> pd.DataFrame:
    parts = str(symbol).split(":")
    if len(parts) < 3:
        raise ValueError("KOMIS 심볼 형식이 올바르지 않습니다. 예: KOMIS:HP002:MNRL0001")
    hp_code = parts[1].upper()
    mineral_code = parts[2].upper()
    preferred_criterion = parts[3] if len(parts) >= 4 and parts[3] else ""
    safe_symbol = re.sub(r"[^A-Za-z0-9_-]", "_", "_".join(parts[1:4]))
    cache_path = KOMIS_PRICE_CACHE_DIR / f"{safe_symbol}.json"
    body: dict[str, Any] | None = None
    start_month = start_date.strftime("%Y%m")
    end_month = end_date.strftime("%Y%m")
    headers = {
        "User-Agent": "Mozilla/5.0",
        "Referer": komis_referer_for_hp(hp_code),
        "X-Requested-With": "XMLHttpRequest",
    }
    if cache_path.exists() and (time.time() - cache_path.stat().st_mtime) < 6 * 60 * 60:
        body = json.loads(cache_path.read_text(encoding="utf-8"))
    try:
        if body is None:
            criterion = preferred_criterion
            if not criterion:
                criterion_response = requests.post(
                    "https://www.komis.or.kr/Komis/RsrcPrice/ajax/getMnrlPriceCrtr",
                    data={"HP000": hp_code, "mnrkndUnqCd": mineral_code},
                    headers=headers,
                    timeout=15,
                )
                criterion_response.raise_for_status()
                criterion_rows = (criterion_response.json().get("data") or [])
                if not criterion_rows:
                    raise ValueError("KOMIS 가격기준 데이터가 없습니다.")
                criterion = str(criterion_rows[0].get("cdKey") or "")
            response = requests.post(
                "https://www.komis.or.kr/Komis/RsrcPrice/ajax/getChartData",
                data={
                    "srchMnrkndUnqCd": mineral_code,
                    "srchPrcCrtr": criterion,
                    "srchAvgOpt": "DAY",
                    "srchStartDate": start_month,
                    "srchEndDate": end_month,
                    "srchField": "month",
                    "srchCompareMnrkndUnqCd": "",
                    "srchComparePrcCrtr": "",
                    "lmeInvt": "N",
                    "HP000": hp_code,
                },
                headers=headers,
                timeout=20,
            )
            response.raise_for_status()
            body = response.json()
            KOMIS_PRICE_CACHE_DIR.mkdir(parents=True, exist_ok=True)
            cache_path.write_text(json.dumps(body, ensure_ascii=False), encoding="utf-8")
    except Exception:
        if cache_path.exists():
            body = json.loads(cache_path.read_text(encoding="utf-8"))
        else:
            raise

    data = (body or {}).get("data") or {}
    if data.get("errCode"):
        raise ValueError(f"KOMIS 가격 데이터 오류: {data.get('errCode')}")
    xaxis = data.get("xaxis") or []
    series = data.get("series") or []
    values = (series[0] or {}).get("data") if series else []
    rows = []
    for index, raw_date in enumerate(xaxis):
        close = to_float(values[index] if index < len(values) else None)
        row_date_value = pd.to_datetime(str(raw_date).replace(".", "-"), errors="coerce")
        if close is None or pd.isna(row_date_value):
            continue
        row_date = row_date_value.date()
        if row_date < start_date or row_date > end_date:
            continue
        rows.append(
            {
                "Date": row_date,
                "Open": close,
                "High": close,
                "Low": close,
                "Close": close,
                "Volume": 0,
            }
        )
    frame = pd.DataFrame(rows)
    if frame.empty:
        raise ValueError(f"KOMIS 가격 데이터가 없습니다: {mineral_code}")
    frame = frame.dropna(subset=["Date", "Close"]).drop_duplicates(subset=["Date"], keep="last").sort_values("Date")
    return frame.reset_index(drop=True)


def load_global_index_price_frame(symbol: str, start_date: date, end_date: date, multiplier: float = 1.0) -> pd.DataFrame:
    fetch_start = start_date - timedelta(days=260)
    if symbol.upper().startswith("EASTMONEY:"):
        secid = symbol.split(":", 1)[1]
        frame = load_eastmoney_futures_price_frame(secid, fetch_start, end_date)
        frame["ma20"] = frame["Close"].rolling(20).mean()
        frame["ma200"] = frame["Close"].rolling(200).mean()
        return frame[frame["Date"] >= fetch_start].reset_index(drop=True)
    if symbol.upper().startswith("FRED:"):
        series_id = symbol.split(":", 1)[1]
        frame = load_fred_price_frame(series_id, fetch_start, end_date)
        frame["ma20"] = frame["Close"].rolling(20).mean()
        frame["ma200"] = frame["Close"].rolling(200).mean()
        return frame[frame["Date"] >= fetch_start].reset_index(drop=True)
    if symbol.upper().startswith("BLS:"):
        series_id = symbol.split(":", 1)[1]
        frame = load_bls_price_frame(series_id, fetch_start, end_date)
        frame["ma20"] = frame["Close"].rolling(20).mean()
        frame["ma200"] = frame["Close"].rolling(200).mean()
        return frame[frame["Date"] >= fetch_start].reset_index(drop=True)
    if symbol.upper().startswith("TRADINGECONOMICS:"):
        market_symbol = symbol.split(":", 1)[1]
        frame = load_tradingeconomics_market_price_frame(market_symbol, fetch_start, end_date)
        frame["ma20"] = frame["Close"].rolling(20).mean()
        frame["ma200"] = frame["Close"].rolling(200).mean()
        return frame[frame["Date"] >= fetch_start].reset_index(drop=True)
    if symbol.upper().startswith("SMM:"):
        product_id = symbol.split(":", 1)[1]
        frame = load_smm_spot_price_frame(product_id, fetch_start, end_date)
        frame["ma20"] = frame["Close"].rolling(20).mean()
        frame["ma200"] = frame["Close"].rolling(200).mean()
        return frame[frame["Date"] >= fetch_start].reset_index(drop=True)
    if symbol.upper().startswith("KOMIS:"):
        frame = load_komis_mineral_price_frame(symbol, fetch_start, end_date)
        frame["ma20"] = frame["Close"].rolling(20).mean()
        frame["ma200"] = frame["Close"].rolling(200).mean()
        return frame[frame["Date"] >= fetch_start].reset_index(drop=True)
    try:
        frame = fdr.DataReader(symbol, fetch_start.isoformat(), end_date.isoformat())
    except Exception:
        frame = pd.DataFrame()
    if frame is None or frame.empty:
        period1 = int(datetime.combine(fetch_start, datetime.min.time()).timestamp())
        period2 = int(datetime.combine(end_date + timedelta(days=1), datetime.min.time()).timestamp())
        response = requests.get(
            f"https://query1.finance.yahoo.com/v8/finance/chart/{quote(symbol)}",
            params={"period1": period1, "period2": period2, "interval": "1d", "events": "history"},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=12,
        )
        response.raise_for_status()
        result = ((response.json().get("chart") or {}).get("result") or [{}])[0]
        timestamps = result.get("timestamp") or []
        quote_payload = (((result.get("indicators") or {}).get("quote") or [{}])[0])
        closes = quote_payload.get("close") or []
        opens = quote_payload.get("open") or closes
        highs = quote_payload.get("high") or closes
        lows = quote_payload.get("low") or closes
        volumes = quote_payload.get("volume") or []
        rows = []
        for index, ts in enumerate(timestamps):
            close = to_float(closes[index] if index < len(closes) else None)
            if close is None:
                continue
            rows.append(
                {
                    "Date": datetime.fromtimestamp(ts).date(),
                    "Open": to_float(opens[index] if index < len(opens) else close) or close,
                    "High": to_float(highs[index] if index < len(highs) else close) or close,
                    "Low": to_float(lows[index] if index < len(lows) else close) or close,
                    "Close": close,
                    "Volume": to_float(volumes[index] if index < len(volumes) else None) or 0,
                }
            )
        frame = pd.DataFrame(rows)
    frame = frame.copy().reset_index()
    if "Date" not in frame.columns:
        frame = frame.rename(columns={frame.columns[0]: "Date"})
    frame["Date"] = pd.to_datetime(frame["Date"]).dt.date
    for column in ["Open", "High", "Low", "Close", "Volume"]:
        if column not in frame.columns:
            frame[column] = frame["Close"] if "Close" in frame.columns else np.nan
        frame[column] = pd.to_numeric(frame[column], errors="coerce")
    price_multiplier = to_float(multiplier) or 1.0
    if price_multiplier != 1.0:
        for column in ["Open", "High", "Low", "Close"]:
            frame[column] = frame[column] * price_multiplier
    frame = frame.dropna(subset=["Date", "Close"]).sort_values("Date")
    frame["ma20"] = frame["Close"].rolling(20).mean()
    frame["ma200"] = frame["Close"].rolling(200).mean()
    return frame[frame["Date"] >= fetch_start].reset_index(drop=True)


def pct_change_between(current: float | None, previous: float | None) -> float | None:
    if current is None or previous is None or previous == 0:
        return None
    return (current / previous - 1) * 100


def indexed_return_series(frame: pd.DataFrame, start_date: date) -> list[dict[str, Any]]:
    window = frame[frame["Date"] >= start_date].reset_index(drop=True)
    if window.empty:
        return []
    base_close = to_float(window.iloc[0].get("Close"))
    if not base_close:
        return []
    rows = []
    for _, row in window.iterrows():
        close = to_float(row.get("Close"))
        if close is None:
            continue
        rows.append({"date": row["Date"].isoformat(), "return_pct": round((close / base_close - 1) * 100, 3), "close": round(close, 4)})
    return rows


def build_global_indices_payload(group: str | None = None) -> dict[str, Any]:
    today = date.today()
    selected_group = str(group or "").strip()
    if selected_group == "PPI":
        chart_start = today - timedelta(days=365 * 3)
    elif selected_group == "한국주식ETF":
        chart_start = today - timedelta(days=365)
    else:
        chart_start = today - timedelta(days=94)
    fetch_start = today - timedelta(days=420)
    if selected_group == "PPI":
        fetch_start = today - timedelta(days=365 * 6)
    elif selected_group == "한국주식ETF":
        fetch_start = today - timedelta(days=620)
    source_items = [
        item for item in GLOBAL_INDEX_ITEMS
        if not selected_group or (item.get("group") or "기타") == selected_group
    ]
    items: list[dict[str, Any]] = [None] * len(source_items)  # type: ignore[list-item]
    errors: list[str] = []

    def build_item_row(index: int, item: dict[str, Any]) -> tuple[int, dict[str, Any], str | None]:
        symbol = item["symbol"]
        fetch_symbol = item.get("fetch_symbol") or symbol
        price_multiplier = to_float(item.get("price_multiplier")) or 1.0
        try:
            frame = load_global_index_price_frame(fetch_symbol, fetch_start, today, multiplier=price_multiplier)
            if frame.empty:
                raise ValueError("가격 데이터가 없습니다.")
            last = frame.iloc[-1]
            last_close = to_float(last.get("Close"))
            closes = frame["Close"].dropna().tolist()
            def lag_close(days: int) -> float | None:
                if len(closes) <= days:
                    return None
                return float(closes[-days - 1])
            is_monthly = str(item.get("frequency") or "").lower() == "monthly"
            year_frame = frame[frame["Date"] >= date(today.year, 1, 1)]
            ytd_base = to_float(year_frame.iloc[0].get("Close")) if not year_frame.empty else None
            ma20 = to_float(last.get("ma20"))
            ma200 = to_float(last.get("ma200"))
            row = {
                **item,
                "fetch_symbol": fetch_symbol,
                "source": item.get("source") or "FinanceDataReader / Yahoo Finance",
                "price_multiplier": price_multiplier,
                "last_date": last["Date"].isoformat(),
                "last_close": round(last_close or 0, 4),
                "volume": int(to_float(last.get("Volume")) or 0),
                "return_1w_pct": None if last_close is None else round(pct_change_between(last_close, lag_close(1 if is_monthly else 5)) or 0, 2),
                "return_1m_pct": None if last_close is None else round(pct_change_between(last_close, lag_close(3 if is_monthly else 21)) or 0, 2),
                "return_3m_pct": None if last_close is None else round(pct_change_between(last_close, lag_close(12 if is_monthly else 63)) or 0, 2),
                "return_ytd_pct": None if last_close is None else round(pct_change_between(last_close, ytd_base) or 0, 2),
                "return_1y_pct": None if last_close is None else round(pct_change_between(last_close, lag_close(12 if is_monthly else 252)) or 0, 2),
                "ma20_gap_pct": None if last_close is None else round(pct_change_between(last_close, ma20) or 0, 2),
                "ma200_gap_pct": None if last_close is None else round(pct_change_between(last_close, ma200) or 0, 2),
                "series": indexed_return_series(frame, chart_start),
            }
            return index, row, None
        except Exception as exc:
            return index, {**item, "fetch_symbol": fetch_symbol, "error": str(exc), "series": []}, f"{symbol}({fetch_symbol}): {exc}"

    if selected_group == "한국주식ETF":
        max_workers = 6
    else:
        max_workers = 8
    if len(source_items) > 8 and selected_group != "PPI":
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = [executor.submit(build_item_row, index, item) for index, item in enumerate(source_items)]
            for future in as_completed(futures):
                index, row, error = future.result()
                items[index] = row
                if error:
                    errors.append(error)
    else:
        for index, item in enumerate(source_items):
            row_index, row, error = build_item_row(index, item)
            items[row_index] = row
            if error:
                errors.append(error)
    items = [item for item in items if item is not None]
    valid_items = [item for item in items if item.get("series")]
    return {
        "as_of": max((item.get("last_date") or "" for item in valid_items), default=today.isoformat()),
        "start_date": chart_start.isoformat(),
        "group": selected_group or None,
        "groups": list(dict.fromkeys((item.get("group") or "기타") for item in GLOBAL_INDEX_ITEMS)),
        "items": items,
        "errors": errors,
        "source": "FinanceDataReader / Yahoo Finance / FRED / BLS / Eastmoney / Trading Economics / SMM / KOMIS. PPI 탭은 FRED의 BLS Producer Price Index 월간 데이터를 사용합니다.",
    }


def strategy_signal(strategy: str, previous: pd.Series | None, current: pd.Series, position: int) -> str:
    if previous is None:
        return ""
    close = to_float(current.get("Close"))
    prev_close = to_float(previous.get("Close"))
    ma20 = to_float(current.get("ma20"))
    prev_ma20 = to_float(previous.get("ma20"))
    ma60 = to_float(current.get("ma60"))
    prev_ma60 = to_float(previous.get("ma60"))
    rsi = to_float(current.get("rsi14"))
    prev_rsi = to_float(previous.get("rsi14"))
    if strategy in {"ma20_cross", "ma20_cross_mdd7"}:
        if position == 0 and prev_close is not None and prev_ma20 is not None and close is not None and ma20 is not None:
            if prev_close <= prev_ma20 and close > ma20:
                return "buy"
        if position == 1 and prev_close is not None and prev_ma20 is not None and close is not None and ma20 is not None:
            if prev_close >= prev_ma20 and close < ma20:
                return "sell"
    elif strategy == "golden_cross":
        if prev_ma20 is not None and prev_ma60 is not None and ma20 is not None and ma60 is not None:
            if position == 0 and prev_ma20 <= prev_ma60 and ma20 > ma60:
                return "buy"
            if position == 1 and prev_ma20 >= prev_ma60 and ma20 < ma60:
                return "sell"
    elif strategy == "rsi_rebound":
        if prev_rsi is not None and rsi is not None:
            if position == 0 and prev_rsi <= 30 and rsi > 30:
                return "buy"
            if position == 1 and prev_rsi >= 70 and rsi < 70:
                return "sell"
    return ""


def build_leader_top10_score70_backtest(
    index_key: str,
    start_date: date,
    end_date: date,
    top_n: int = 10,
    entry_threshold: float = 70.0,
    exit_threshold: float = 70.0,
    allocation_mode: str = "score_weight",
) -> dict[str, Any]:
    summaries = screening_backtest_source_summaries(start_date=start_date, end_date=end_date)
    if len(summaries) < 2:
        raise ValueError("주도주 기반 전략 백테스트에 필요한 캐시 데이터가 부족합니다.")
    filtered = summaries
    if len(filtered) < 2:
        raise ValueError("선택 구간에 주도주 데이터가 부족합니다. 시작일/종료일을 다시 확인해 주세요.")
    strategy_start_date = datetime.strptime(str(filtered[0].get("file_date") or ""), "%Y-%m-%d").date()
    strategy_end_date = datetime.strptime(str(filtered[-1].get("file_date") or ""), "%Y-%m-%d").date()
    all_stock_codes = sorted(
        {
            normalize_stock_code_value(raw.get("stock_code"))
            for summary in filtered
            for raw in summary.get("qualified_stocks", [])
            if isinstance(raw, dict) and normalize_stock_code_value(raw.get("stock_code"))
        }
    )
    close_map = load_screening_close_map(strategy_start_date, strategy_end_date, all_stock_codes)
    benchmark_close_map = load_strategy_benchmark_close_map(index_key, strategy_start_date, strategy_end_date)

    def stock_close_on_date(stock_code: str | None, target_date: date | str) -> float | None:
        code = normalize_stock_code_value(stock_code)
        if not code:
            return None
        date_key = pd.to_datetime(target_date).strftime("%Y%m%d")
        value = to_float((close_map.get(date_key) or {}).get(code))
        return float(value) if value not in (None, 0) else None

    nav = 100.0
    benchmark_nav = 100.0
    previous_weights: dict[str, float] = {}
    holdings: set[str] = set()
    rows: list[dict[str, Any]] = []
    signals: list[dict[str, Any]] = []
    total_turnover = 0.0
    total_fee_pct = 0.0
    trade_count = 0
    position_state: dict[str, dict[str, float]] = {}
    previous_benchmark_close = to_float(benchmark_close_map.get(strategy_start_date.isoformat()))

    strict_all_mode = int(top_n or 0) >= 9999
    allocation_mode_key = str(allocation_mode or "score_weight").strip().lower()
    if allocation_mode_key not in {"score_weight", "fixed_20"}:
        allocation_mode_key = "score_weight"

    def build_target_weights(summary: dict[str, Any], current_holdings: set[str]) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]], list[str], list[str], dict[str, float]]:
        current_rows: dict[str, dict[str, Any]] = {}
        scored_rows: list[dict[str, Any]] = []
        for raw in summary.get("qualified_stocks", []):
            if not isinstance(raw, dict):
                continue
            key = sector_rotation_stock_key(raw)
            if not key:
                continue
            score = to_float(raw.get("score"))
            if score is None:
                continue
            current_rows[key] = raw
            scored_rows.append(raw)
        scored_rows.sort(key=lambda item: float(to_float(item.get("score")) or -99999), reverse=True)
        if strict_all_mode:
            candidate_pool = scored_rows
        else:
            candidate_pool = scored_rows[: max(1, int(top_n or 10))]
        entry_candidates = {
            sector_rotation_stock_key(item)
            for item in candidate_pool
            if item
            and to_float(item.get("score")) is not None
            and (
                float(to_float(item.get("score")) or 0.0) > float(entry_threshold)
                if strict_all_mode
                else float(to_float(item.get("score")) or 0.0) >= float(entry_threshold)
            )
        }
        buys: list[str] = []
        sells: list[str] = []
        next_holdings = set(current_holdings)
        for stock_key in sorted(current_holdings):
            row = current_rows.get(stock_key)
            score = to_float(row.get("score")) if row else None
            if score is None or float(score) <= float(exit_threshold):
                sells.append(stock_key)
                next_holdings.discard(stock_key)
        for stock_key in sorted(entry_candidates):
            if stock_key not in next_holdings:
                next_holdings.add(stock_key)
                buys.append(stock_key)
        target_weights: dict[str, float] = {}
        scored_holdings = []
        for stock_key in next_holdings:
            row = current_rows.get(stock_key)
            score = float(to_float(row.get("score")) or 0.0) if row else 0.0
            if score > 0:
                scored_holdings.append((stock_key, score))
        if scored_holdings:
            if allocation_mode_key == "fixed_20":
                for stock_key, _score in scored_holdings:
                    target_weights[stock_key] = 0.2
            else:
                score_sum = sum(score for _, score in scored_holdings)
                if score_sum > 0:
                    for stock_key, score in scored_holdings:
                        target_weights[stock_key] = score / score_sum
                else:
                    equal_weight = 1.0 / len(scored_holdings)
                    for stock_key, _score in scored_holdings:
                        target_weights[stock_key] = equal_weight
        return current_rows, scored_rows, buys, sells, target_weights

    def build_holdings_payload(target_weights: dict[str, float], current_rows: dict[str, dict[str, Any]], current_dt: date) -> tuple[list[dict[str, Any]], list[dict[str, float | str]]]:
        holdings_payload: list[dict[str, Any]] = []
        sector_weight_map: dict[str, float] = {}
        for stock_key, weight in sorted(target_weights.items(), key=lambda item: item[1], reverse=True):
            current_row = current_rows.get(stock_key) or {}
            stock_name = str(current_row.get("resolved_name") or current_row.get("stock_name") or stock_key)
            stock_code = str(current_row.get("stock_code") or "")
            score_val = float(to_float(current_row.get("score")) or 0.0)
            sector_name = str(current_row.get("manual_sector") or current_row.get("theme") or "").strip()
            if not sector_name:
                sector_name = "기타"
            sector_weight_map[sector_name] = sector_weight_map.get(sector_name, 0.0) + weight * 100.0
            cur_price = float(stock_close_on_date(stock_code, current_dt) or 0.0)
            state = position_state.get(stock_key)
            if not state:
                state = {"avg_buy_price": cur_price, "stock_code": stock_code}
                position_state[stock_key] = state
            avg_buy_price = float(state.get("avg_buy_price") or cur_price)
            avg_buy_return_pct = ((cur_price / avg_buy_price) - 1.0) * 100.0 if avg_buy_price else 0.0
            holdings_payload.append(
                {
                    "stock_key": stock_key,
                    "stock_code": stock_code,
                    "stock_name": stock_name,
                    "score": round(score_val, 2),
                    "weight_pct": round(weight * 100.0, 2),
                    "daily_contribution_pct": None,
                    "alpha_contribution_pct": None,
                    "avg_buy_price": round(avg_buy_price, 2),
                    "avg_buy_return_pct": round(avg_buy_return_pct, 2),
                    "sector": sector_name,
                    "status": "유지",
                }
            )
        sector_weights = [
            {"sector": sector, "weight_pct": round(weight_pct, 2)}
            for sector, weight_pct in sorted(sector_weight_map.items(), key=lambda item: item[1], reverse=True)
        ]
        return holdings_payload, sector_weights

    first_summary = filtered[0]
    first_date = datetime.strptime(str(first_summary.get("file_date") or ""), "%Y-%m-%d").date()
    first_rows, _first_scored, first_buys, first_sells, first_weights = build_target_weights(first_summary, holdings)
    first_turnover = sum(abs(first_weights.get(key, 0.0)) for key in first_weights)
    first_fee = first_turnover * STRATEGY_TRADE_FEE_RATE
    total_turnover += first_turnover
    total_fee_pct += first_fee * 100.0
    if first_buys or first_sells:
        trade_count += len(first_buys) + len(first_sells)
    nav *= (1.0 - first_fee)
    holdings = set(first_weights.keys())
    for stock_key, target_weight in first_weights.items():
        row_for_code = first_rows.get(stock_key) or {}
        stock_code_for_price = str(row_for_code.get("stock_code") or "")
        buy_price = stock_close_on_date(stock_code_for_price, first_date)
        if buy_price not in (None, 0):
            position_state[stock_key] = {"avg_buy_price": float(buy_price), "stock_code": stock_code_for_price}
    first_holdings_payload, first_sector_weights = build_holdings_payload(first_weights, first_rows, first_date)
    for item in first_holdings_payload:
        if item["stock_key"] in first_buys:
            item["status"] = "신규"
    base_benchmark_close = to_float(benchmark_close_map.get(first_date.isoformat())) or 0.0
    rows.append(
        {
            "date": first_date.isoformat(),
            "close": round(float(base_benchmark_close), 4),
            "ma20": None,
            "ma60": None,
            "ma200": None,
            "rsi14": None,
            "benchmark_return_pct": 0.0,
            "strategy_return_pct": round(nav - 100.0, 3),
            "benchmark_daily_return_pct": 0.0,
            "position": 1 if first_weights else 0,
            "signal": "buy" if first_buys else "",
            "signal_date": first_date.isoformat(),
            "daily_return_pct": round(-first_fee * 100.0, 3),
            "exposure_pct": round(sum(first_weights.values()) * 100.0, 2),
            "turnover_pct": round(first_turnover * 100.0, 2),
            "fee_pct": round(first_fee * 100.0, 3),
            "holdings_count": len(first_weights),
            "holdings": first_holdings_payload,
            "sector_weights": first_sector_weights,
            "sector_weight_sum_pct": round(sum(item["weight_pct"] for item in first_sector_weights), 2),
            "entry_exit": {
                "buy": [str((first_rows.get(key) or {}).get("resolved_name") or (first_rows.get(key) or {}).get("stock_name") or key) for key in first_buys],
                "sell": [],
                "sell_details": [],
            },
            "buy_count": len(first_buys),
            "sell_count": 0,
            "signal_reason": "초기 편입",
        }
    )
    previous_weights = first_weights

    for index_no in range(1, len(filtered)):
        previous_summary = filtered[index_no - 1]
        current_summary = filtered[index_no]
        previous_date = datetime.strptime(str(previous_summary.get("file_date") or ""), "%Y-%m-%d").date()
        current_date = datetime.strptime(str(current_summary.get("file_date") or ""), "%Y-%m-%d").date()
        current_rows, _current_scored, buys, sells, target_weights = build_target_weights(current_summary, holdings)
        sell_state_snapshot = {stock_key: dict(position_state.get(stock_key) or {}) for stock_key in sells}
        benchmark_close = to_float(benchmark_close_map.get(current_date.isoformat()))
        benchmark_daily_return = 0.0
        if previous_benchmark_close not in (None, 0) and benchmark_close not in (None, 0):
            benchmark_daily_return = float(benchmark_close / previous_benchmark_close - 1.0)
            benchmark_nav *= (1.0 + benchmark_daily_return)
        if benchmark_close not in (None, 0):
            previous_benchmark_close = benchmark_close

        holdings_payload_before: list[dict[str, Any]] = []
        portfolio_move = 0.0
        for stock_key, weight in sorted(previous_weights.items(), key=lambda item: item[1], reverse=True):
            source_row = (previous_summary.get("qualified_stocks") or [])
            prev_row = next((item for item in source_row if sector_rotation_stock_key(item) == stock_key), {})
            stock_code = normalize_stock_code_value(prev_row.get("stock_code"))
            prev_close = stock_close_on_date(stock_code, previous_date)
            curr_close = stock_close_on_date(stock_code, current_date)
            daily_stock_return = None
            if prev_close not in (None, 0) and curr_close not in (None, 0):
                daily_stock_return = (float(curr_close) / float(prev_close) - 1.0) * 100.0
                portfolio_move += weight * (float(curr_close) / float(prev_close) - 1.0)
            else:
                fallback_change = to_float((current_rows.get(stock_key) or {}).get("change_pct"))
                if fallback_change is not None:
                    daily_stock_return = float(fallback_change)
                    portfolio_move += weight * (float(fallback_change) / 100.0)
            state = position_state.get(stock_key) or {}
            avg_buy_price = to_float(state.get("avg_buy_price"))
            avg_buy_return_pct = None
            if avg_buy_price not in (None, 0) and curr_close not in (None, 0):
                avg_buy_return_pct = ((float(curr_close) / float(avg_buy_price)) - 1.0) * 100.0
            holdings_payload_before.append(
                {
                    "stock_key": stock_key,
                    "stock_code": stock_code,
                    "stock_name": str((current_rows.get(stock_key) or prev_row).get("stock_name") or stock_key),
                    "score": round(float(to_float((current_rows.get(stock_key) or prev_row).get("score")) or 0.0), 2),
                    "weight_pct": round(weight * 100.0, 2),
                    "daily_contribution_pct": round(weight * float(daily_stock_return or 0.0), 3) if daily_stock_return is not None else None,
                    "alpha_contribution_pct": round(weight * (float(daily_stock_return or 0.0) - benchmark_daily_return * 100.0), 3) if daily_stock_return is not None else None,
                    "avg_buy_price": round(float(avg_buy_price), 2) if avg_buy_price is not None else None,
                    "avg_buy_return_pct": round(float(avg_buy_return_pct), 2) if avg_buy_return_pct is not None else None,
                    "sector": str((current_rows.get(stock_key) or prev_row).get("manual_sector") or (current_rows.get(stock_key) or prev_row).get("theme") or "기타"),
                    "status": "유지",
                }
            )
        turnover = sum(abs(target_weights.get(key, 0.0) - previous_weights.get(key, 0.0)) for key in (set(target_weights) | set(previous_weights)))
        fee = turnover * STRATEGY_TRADE_FEE_RATE
        total_turnover += turnover
        total_fee_pct += fee * 100.0
        if buys or sells:
            trade_count += len(buys) + len(sells)
        period_return = (1.0 + portfolio_move) * (1.0 - fee) - 1.0
        nav *= (1.0 + period_return)

        sell_details: list[dict[str, Any]] = []
        for stock_key in sells:
            current_row = current_rows.get(stock_key) or {}
            state = sell_state_snapshot.get(stock_key) or {}
            sell_code = str(current_row.get("stock_code") or state.get("stock_code") or "")
            sell_name = str(current_row.get("resolved_name") or current_row.get("stock_name") or stock_key)
            avg_buy_price = to_float(state.get("avg_buy_price"))
            sell_price = stock_close_on_date(sell_code, current_date)
            sell_return_pct = None
            if avg_buy_price not in (None, 0) and sell_price not in (None, 0):
                sell_return_pct = ((float(sell_price) / float(avg_buy_price)) - 1.0) * 100.0
            sell_details.append(
                {
                    "stock_key": stock_key,
                    "stock_code": sell_code,
                    "stock_name": sell_name,
                    "avg_buy_price": round(float(avg_buy_price), 2) if avg_buy_price is not None else None,
                    "sell_price": round(float(sell_price), 2) if sell_price is not None else None,
                    "sell_return_pct": round(float(sell_return_pct), 2) if sell_return_pct is not None else None,
                }
            )
            position_state.pop(stock_key, None)

        rebalance_keys = set(previous_weights.keys()) | set(target_weights.keys())
        for stock_key in rebalance_keys:
            row_for_code = current_rows.get(stock_key) or {}
            stock_code_for_price = str(row_for_code.get("stock_code") or "")
            rebalance_price = float(stock_close_on_date(stock_code_for_price, current_date) or 0.0)
            prev_weight = float(previous_weights.get(stock_key, 0.0) or 0.0)
            target_weight = float(target_weights.get(stock_key, 0.0) or 0.0)
            state = position_state.get(stock_key)
            if target_weight <= 0.0:
                position_state.pop(stock_key, None)
                continue
            if state is None or prev_weight <= 0.0:
                position_state[stock_key] = {"avg_buy_price": rebalance_price, "stock_code": stock_code_for_price}
                continue
            avg_buy_price = float(state.get("avg_buy_price") or rebalance_price or 0.0)
            if target_weight > prev_weight and rebalance_price > 0:
                add_weight = target_weight - prev_weight
                new_avg_buy_price = ((avg_buy_price * prev_weight) + (rebalance_price * add_weight)) / target_weight
                state["avg_buy_price"] = float(new_avg_buy_price)
            state["stock_code"] = stock_code_for_price
            position_state[stock_key] = state

        holdings = set(target_weights.keys())
        next_holdings_payload, next_sector_weights = build_holdings_payload(target_weights, current_rows, current_date)
        for item in next_holdings_payload:
            if item["stock_key"] in buys:
                item["status"] = "신규"
        signal = "buy" if buys else ("sell" if sells else "")
        signal_reason = ""
        if buys or sells:
            signal_reason = "매수 " + str(len(buys)) + " / 매도 " + str(len(sells)) + " · 점수비중 리밸런싱"
            signals.append(
                {
                    "date": current_date.isoformat(),
                    "type": signal or "rebalance",
                    "price": round(float(benchmark_close or 0.0), 4),
                    "strategy_return_pct": round(nav - 100.0, 3),
                    "benchmark_return_pct": round(benchmark_nav - 100.0, 3),
                    "reason": signal_reason,
                }
            )

        rows.append(
            {
                "date": current_date.isoformat(),
                "close": round(float(benchmark_close or 0.0), 4),
                "ma20": None,
                "ma60": None,
                "ma200": None,
                "rsi14": None,
                "benchmark_return_pct": round(benchmark_nav - 100.0, 3),
                "strategy_return_pct": round(nav - 100.0, 3),
                "benchmark_daily_return_pct": round(benchmark_daily_return * 100.0, 3),
                "position": 1 if target_weights else 0,
                "signal": signal,
                "signal_date": previous_date.isoformat(),
                "daily_return_pct": round(period_return * 100.0, 3),
                "exposure_pct": round(sum(target_weights.values()) * 100.0, 2),
                "turnover_pct": round(turnover * 100.0, 2),
                "fee_pct": round(fee * 100.0, 3),
                "holdings_count": len(target_weights),
                "holdings": next_holdings_payload,
                "holdings_before_close": holdings_payload_before,
                "sector_weights": next_sector_weights,
                "sector_weight_sum_pct": round(sum(item["weight_pct"] for item in next_sector_weights), 2),
                "entry_exit": {
                    "buy": [str((current_rows.get(key) or {}).get("resolved_name") or (current_rows.get(key) or {}).get("stock_name") or key) for key in buys],
                    "sell": [str((current_rows.get(key) or {}).get("resolved_name") or (current_rows.get(key) or {}).get("stock_name") or key) for key in sells],
                    "sell_details": sell_details,
                },
                "buy_count": len(buys),
                "sell_count": len(sells),
                "signal_reason": signal_reason,
            }
        )
        previous_weights = target_weights

    final_strategy = rows[-1]["strategy_return_pct"] if rows else 0.0
    final_benchmark = rows[-1]["benchmark_return_pct"] if rows else 0.0
    summary = {
        "index_return_pct": round(final_benchmark, 2),
        "strategy_return_pct": round(final_strategy, 2),
        "excess_return_pct": round(final_strategy - final_benchmark, 2),
        "signal_count": len(signals),
        "trade_count": int(trade_count),
        "win_rate_pct": None,
        "trade_fee_rate_pct": round(STRATEGY_TRADE_FEE_RATE * 100.0, 2),
        "total_fee_pct_points": round(total_fee_pct, 2),
        "avg_exposure_pct": round(float(np.mean([float(row.get("exposure_pct") or 0.0) for row in rows])) if rows else 0.0, 2),
        "avg_holdings_count": round(float(np.mean([float(row.get("holdings_count") or 0.0) for row in rows])) if rows else 0.0, 2),
    }
    return {
        "rows": rows,
        "signals": signals,
        "summary": summary,
        "start_date": rows[0]["date"] if rows else str(filtered[0].get("file_date") or ""),
        "end_date": rows[-1]["date"] if rows else str(filtered[-1].get("file_date") or ""),
    }


def build_strategy_backtest(
    index: str = "KS11",
    strategy: str = "ma20_cross",
    start: str | None = None,
    end: str | None = None,
    top_n: int | None = None,
    entry_threshold: float | None = None,
    exit_threshold: float | None = None,
    allocation_mode: str = "score_weight",
) -> dict[str, Any]:
    index_key = str(index or "KS11").upper()
    if index_key not in STRATEGY_INDEXES:
        index_key = "KS11"
    strategy_key = str(strategy or "ma20_cross").strip()
    if strategy_key not in STRATEGY_TYPES:
        strategy_key = "ma20_cross"
    if strategy_key in LEADER_STRATEGY_PRESETS or strategy_key == "leader_custom":
        preset = LEADER_STRATEGY_PRESETS.get(strategy_key) or {}
        effective_top_n = int(top_n if top_n is not None else (preset.get("top_n") or 10))
        effective_entry_threshold = float(entry_threshold if entry_threshold is not None else (preset.get("entry_threshold") or 70.0))
        effective_exit_threshold = float(exit_threshold if exit_threshold is not None else (preset.get("exit_threshold") or effective_entry_threshold))
        effective_allocation_mode = str(allocation_mode or "score_weight").strip().lower()
        if effective_allocation_mode not in {"score_weight", "fixed_20"}:
            effective_allocation_mode = "score_weight"
        top10_payload = build_leader_top10_score70_backtest(
            index_key=index_key,
            start_date=safe_strategy_date(start, date.today() - timedelta(days=365)),
            end_date=safe_strategy_date(end, date.today()),
            top_n=effective_top_n,
            entry_threshold=effective_entry_threshold,
            exit_threshold=effective_exit_threshold,
            allocation_mode=effective_allocation_mode,
        )
        allocation_label = "점수 비중" if effective_allocation_mode == "score_weight" else "종목당 20%"
        return {
            "index": index_key,
            "index_name": STRATEGY_INDEXES[index_key]["name"],
            "strategy": strategy_key,
            "strategy_name": (
                f"주도주 점수 전략 (매수 {effective_entry_threshold:g} / 매도 {effective_exit_threshold:g} / "
                + (f"Top{effective_top_n}" if effective_top_n < 9999 else "전종목")
                + f" / {allocation_label})"
            ) if strategy_key == "leader_custom" else STRATEGY_TYPES[strategy_key]["name"],
            "strategy_description": (
                "종합점수가 매수 기준 이상이면 편입하고, 보유 종목 점수가 매도 기준 이하로 내려가면 편출합니다. "
                + ("보유 종목은 매일 점수 비율로 리밸런싱하며 "
                   if effective_allocation_mode == "score_weight"
                   else "보유 종목은 종목당 20%씩 배분하며, 5종목 미만은 현금 보유, 5종목 초과는 레버리지 사용으로 가정하고 ")
                + "회전 비용 0.2%를 반영합니다."
            ) if strategy_key == "leader_custom" else STRATEGY_TYPES[strategy_key]["description"],
            "start_date": top10_payload["start_date"],
            "end_date": top10_payload["end_date"],
            "rows": top10_payload["rows"],
            "signals": top10_payload["signals"],
            "summary": top10_payload["summary"],
            "params": {
                "top_n": effective_top_n,
                "entry_threshold": effective_entry_threshold,
                "exit_threshold": effective_exit_threshold,
                "allocation_mode": effective_allocation_mode,
            },
            "available_indexes": [
                {"key": key, "name": value["name"]} for key, value in STRATEGY_INDEXES.items()
            ],
            "available_strategies": [
                {"key": key, "name": value["name"], "description": value["description"]}
                for key, value in STRATEGY_TYPES.items()
            ],
        }
    today = date.today()
    end_date = safe_strategy_date(end, today)
    start_date = safe_strategy_date(start, end_date - timedelta(days=365))
    if start_date >= end_date:
        start_date = end_date - timedelta(days=365)
    frame = load_strategy_price_frame(STRATEGY_INDEXES[index_key]["symbol"], start_date, end_date)
    if frame.empty:
        raise ValueError("선택한 기간에 지수 데이터가 없습니다.")

    base_close = float(frame["Close"].iloc[0])
    strategy_nav = 100.0
    position = 0
    rows: list[dict[str, Any]] = []
    signals: list[dict[str, Any]] = []
    previous_row: pd.Series | None = None
    wins = 0
    trades = 0
    entry_nav: float | None = None
    position_peak_nav: float | None = None
    total_fee_pct = 0.0

    for index_no, row in frame.iterrows():
        close = float(row["Close"])
        if index_no > 0 and previous_row is not None:
            previous_close = float(previous_row["Close"])
            daily_return = (close / previous_close - 1) if previous_close else 0.0
            if position == 1:
                strategy_nav *= 1 + daily_return
                position_peak_nav = max(position_peak_nav or strategy_nav, strategy_nav)
        signal = strategy_signal(strategy_key, previous_row, row, position)
        signal_reason = ""
        if (
            strategy_key == "ma20_cross_mdd7"
            and position == 1
            and signal != "sell"
            and position_peak_nav
            and strategy_nav <= position_peak_nav * (1 - STRATEGY_MDD_STOP_RATE)
        ):
            signal = "sell"
            signal_reason = "MDD 7% 매도"
        if signal == "buy" and position == 0:
            strategy_nav *= (1 - STRATEGY_TRADE_FEE_RATE)
            total_fee_pct += STRATEGY_TRADE_FEE_RATE * 100
            position = 1
            entry_nav = strategy_nav
            position_peak_nav = strategy_nav
            signal_reason = "매수 수수료 0.2%"
        elif signal == "sell" and position == 1:
            strategy_nav *= (1 - STRATEGY_TRADE_FEE_RATE)
            total_fee_pct += STRATEGY_TRADE_FEE_RATE * 100
            position = 0
            trades += 1
            if entry_nav is not None and strategy_nav > entry_nav:
                wins += 1
            entry_nav = None
            position_peak_nav = None
            if not signal_reason:
                signal_reason = "매도 수수료 0.2%"

        benchmark_return = (close / base_close - 1) * 100
        strategy_return = strategy_nav - 100
        date_text = row["Date"].isoformat()
        row_payload = {
            "date": date_text,
            "close": round(close, 4),
            "ma20": round(float(row["ma20"]), 4) if pd.notna(row["ma20"]) else None,
            "ma60": round(float(row["ma60"]), 4) if pd.notna(row["ma60"]) else None,
            "ma200": round(float(row["ma200"]), 4) if pd.notna(row["ma200"]) else None,
            "rsi14": round(float(row["rsi14"]), 2) if pd.notna(row["rsi14"]) else None,
            "benchmark_return_pct": round(benchmark_return, 3),
            "strategy_return_pct": round(strategy_return, 3),
            "position": position,
            "signal": signal,
        }
        if signal:
            signals.append(
                {
                    "date": date_text,
                    "type": signal,
                    "price": row_payload["close"],
                    "strategy_return_pct": row_payload["strategy_return_pct"],
                    "benchmark_return_pct": row_payload["benchmark_return_pct"],
                    "reason": signal_reason,
                }
            )
        rows.append(row_payload)
        previous_row = row

    closes = [float(row["close"]) for row in rows]
    final_strategy = rows[-1]["strategy_return_pct"]
    final_benchmark = rows[-1]["benchmark_return_pct"]
    summary = {
        "index_return_pct": round(final_benchmark, 2),
        "strategy_return_pct": round(final_strategy, 2),
        "excess_return_pct": round(final_strategy - final_benchmark, 2),
        "signal_count": len(signals),
        "trade_count": trades,
        "win_rate_pct": round(wins / trades * 100, 1) if trades else None,
        "trade_fee_rate_pct": round(STRATEGY_TRADE_FEE_RATE * 100, 2),
        "total_fee_pct_points": round(total_fee_pct, 2),
        "max_close": round(max(closes), 2) if closes else None,
        "min_close": round(min(closes), 2) if closes else None,
    }
    return {
        "index": index_key,
        "index_name": STRATEGY_INDEXES[index_key]["name"],
        "strategy": strategy_key,
        "strategy_name": STRATEGY_TYPES[strategy_key]["name"],
        "strategy_description": STRATEGY_TYPES[strategy_key]["description"],
        "start_date": rows[0]["date"],
        "end_date": rows[-1]["date"],
        "rows": rows,
        "signals": signals,
        "summary": summary,
        "available_indexes": [
            {"key": key, "name": value["name"]} for key, value in STRATEGY_INDEXES.items()
        ],
        "available_strategies": [
            {"key": key, "name": value["name"], "description": value["description"]}
            for key, value in STRATEGY_TYPES.items()
        ],
    }


@lru_cache(maxsize=32)
def _screening_backtest_source_summaries_cached(start_key: str, end_key: str) -> tuple[dict[str, Any], ...]:
    ensure_screening_db_indexes()
    if not SCREENING_FAST_DB_PATH.exists():
        return ()
    with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
        query = """
            SELECT
                file_date,
                file_date_key,
                stock_code,
                stock_name,
                sector,
                industry,
                market_cap_100m,
                trading_value_100m,
                change_pct,
                score_o,
                score_s,
                sortino_norm,
                avg_1w,
                avg_1m,
                avg_3m,
                note
            FROM screening_rows
            WHERE file_date_key BETWEEN ? AND ?
            ORDER BY file_date_key ASC, score_s DESC, stock_code ASC
        """
        frame = pd.read_sql_query(query, conn, params=[start_key, end_key])
    if frame.empty:
        return ()
    grouped: list[dict[str, Any]] = []
    for file_date, group in frame.groupby("file_date", sort=True):
        qualified_stocks: list[dict[str, Any]] = []
        for _, row in group.iterrows():
            qualified_stocks.append(
                {
                    "file_date": str(row.get("file_date") or ""),
                    "file_date_key": str(row.get("file_date_key") or ""),
                    "stock_code": normalize_stock_code_value(row.get("stock_code")),
                    "stock_name": str(row.get("stock_name") or ""),
                    "resolved_name": str(row.get("stock_name") or ""),
                    "manual_sector": str(row.get("sector") or ""),
                    "theme": str(row.get("sector") or ""),
                    "industry": str(row.get("industry") or ""),
                    "market_cap_100m": to_float(row.get("market_cap_100m")),
                    "trading_value_100m": to_float(row.get("trading_value_100m")),
                    "change_pct": to_float(row.get("change_pct")) or 0.0,
                    "score_o": to_float(row.get("score_o")) or 0.0,
                    "score": to_float(row.get("score_s")) or 0.0,
                    "sortino_norm": to_float(row.get("sortino_norm")) or 0.0,
                    "avg_1w": to_float(row.get("avg_1w")) or 0.0,
                    "avg_1m": to_float(row.get("avg_1m")) or 0.0,
                    "avg_3m": to_float(row.get("avg_3m")) or 0.0,
                    "note": str(row.get("note") or ""),
                }
            )
        grouped.append(
            {
                "file_date": str(file_date or ""),
                "file_date_key": datetime.strptime(str(file_date), "%Y-%m-%d").strftime("%Y%m%d") if file_date else "",
                "qualified_stocks": qualified_stocks,
            }
        )
    return tuple(grouped)


def screening_backtest_source_summaries(
    start_date: date | None = None,
    end_date: date | None = None,
) -> list[dict[str, Any]]:
    start_key = (start_date or date(2000, 1, 1)).strftime("%Y%m%d")
    end_key = (end_date or date.today()).strftime("%Y%m%d")
    return list(_screening_backtest_source_summaries_cached(start_key, end_key))


@lru_cache(maxsize=1)
def ensure_screening_db_indexes() -> bool:
    if not SCREENING_FAST_DB_PATH.exists():
        return False
    with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
        conn.execute("CREATE INDEX IF NOT EXISTS idx_screening_date ON screening_rows(file_date_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_screening_code_date ON screening_rows(stock_code, file_date_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_screening_date_score_code ON screening_rows(file_date_key, score_s DESC, stock_code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_close_cache_date ON daily_close_cache(file_date_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_close_cache_code_date ON daily_close_cache(stock_code, file_date_key)")
        conn.commit()
    return True


def screening_available_file_entries(limit: int | None = None) -> list[dict[str, str]]:
    ensure_screening_db_indexes()
    if not SCREENING_FAST_DB_PATH.exists():
        return []
    query = """
        SELECT file_date_key, file_name
        FROM file_meta
        ORDER BY file_date_key DESC
    """
    if limit is not None:
        query += f" LIMIT {max(1, int(limit))}"
    with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
        rows = conn.execute(query).fetchall()
    result: list[dict[str, str]] = []
    for file_date_key, file_name in rows:
        date_key = str(file_date_key or "")
        if not re.fullmatch(r"20\d{6}", date_key):
            continue
        result.append(
            {
                "file_date_key": date_key,
                "file_date": f"{date_key[:4]}-{date_key[4:6]}-{date_key[6:]}",
                "file_name": str(file_name or f"{date_key}_데일리_기업스크리닝.xlsx"),
            }
        )
    return result


def load_screening_summaries_for_dates(file_dates: list[str]) -> list[dict[str, Any]]:
    ensure_screening_db_indexes()
    if not SCREENING_FAST_DB_PATH.exists():
        return []
    normalized_dates = []
    for file_date in file_dates:
        digits = re.sub(r"\D", "", str(file_date or ""))
        if re.fullmatch(r"20\d{6}", digits):
            normalized_dates.append(digits)
    normalized_dates = sorted(set(normalized_dates))
    if not normalized_dates:
        return []
    placeholders = ",".join(["?"] * len(normalized_dates))
    with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
        existing_columns = {
            str(row[1])
            for row in conn.execute("PRAGMA table_info(screening_rows)").fetchall()
            if row and len(row) > 1
        }
        rank_select = "rank" if "rank" in existing_columns else "NULL AS rank"
        query = f"""
        SELECT
            file_date,
            file_date_key,
            {rank_select},
            stock_code,
            stock_name,
            sector,
            industry,
            market_cap_100m,
            trading_value_100m,
            change_pct,
            score_o,
            score_s,
            sortino_norm,
            is_52w_high,
            avg_1w,
            avg_1m,
            avg_3m,
            note
        FROM screening_rows
        WHERE file_date_key IN ({placeholders})
        ORDER BY file_date_key ASC, score_s DESC, stock_code ASC
    """
        frame = pd.read_sql_query(query, conn, params=normalized_dates)
    if frame.empty:
        return []
    grouped: list[dict[str, Any]] = []
    for file_date, group in frame.groupby("file_date", sort=True):
        qualified_stocks: list[dict[str, Any]] = []
        for index_no, (_, row) in enumerate(group.iterrows(), start=1):
            qualified_stocks.append(
                {
                    "file_date": str(row.get("file_date") or ""),
                    "file_date_key": str(row.get("file_date_key") or ""),
                    "rank": int(to_float(row.get("rank")) or index_no),
                    "stock_code": normalize_stock_code_value(row.get("stock_code")),
                    "stock_name": str(row.get("stock_name") or ""),
                    "resolved_name": str(row.get("stock_name") or ""),
                    "manual_sector": str(row.get("sector") or ""),
                    "theme": str(row.get("sector") or ""),
                    "industry": str(row.get("industry") or ""),
                    "market_cap_100m": to_float(row.get("market_cap_100m")),
                    "trading_value_100m": to_float(row.get("trading_value_100m")),
                    "change_pct": to_float(row.get("change_pct")) or 0.0,
                    "score_o": to_float(row.get("score_o")) or 0.0,
                    "score": to_float(row.get("score_s")) or 0.0,
                    "sortino_norm": to_float(row.get("sortino_norm")) or 0.0,
                    "is_52w_high": int(to_float(row.get("is_52w_high")) or 0),
                    "avg_1w": to_float(row.get("avg_1w")) or 0.0,
                    "avg_1m": to_float(row.get("avg_1m")) or 0.0,
                    "avg_3m": to_float(row.get("avg_3m")) or 0.0,
                    "note": str(row.get("note") or ""),
                }
            )
        grouped.append(
            {
                "file_date": str(file_date or ""),
                "file_date_key": datetime.strptime(str(file_date), "%Y-%m-%d").strftime("%Y%m%d") if file_date else "",
                "qualified_stocks": qualified_stocks,
            }
        )
    return grouped


def load_screening_close_map(
    start_date: date,
    end_date: date,
    stock_codes: list[str],
) -> dict[str, dict[str, float]]:
    if not SCREENING_FAST_DB_PATH.exists():
        return {}
    normalized_codes = sorted({normalize_stock_code_value(code) for code in stock_codes if normalize_stock_code_value(code)})
    if not normalized_codes:
        return {}
    placeholders = ",".join(["?"] * len(normalized_codes))
    params: list[Any] = [start_date.strftime("%Y%m%d"), end_date.strftime("%Y%m%d"), *normalized_codes]
    query = f"""
        SELECT file_date_key, stock_code, close_price
        FROM daily_close_cache
        WHERE file_date_key BETWEEN ? AND ?
          AND stock_code IN ({placeholders})
    """
    with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
        rows = conn.execute(query, params).fetchall()
    result: dict[str, dict[str, float]] = {}
    for file_date_key, stock_code, close_price in rows:
        key = str(file_date_key or "")
        code = normalize_stock_code_value(stock_code)
        price = to_float(close_price)
        if not key or not code or price in (None, 0):
            continue
        result.setdefault(key, {})[code] = float(price)
    return result


@lru_cache(maxsize=1)
def ensure_us_screening_db_indexes() -> bool:
    if not US_SCREENING_FAST_DB_PATH.exists():
        return False
    with sqlite3.connect(str(US_SCREENING_FAST_DB_PATH)) as conn:
        conn.execute("CREATE INDEX IF NOT EXISTS idx_screening_date ON screening_rows(file_date_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_screening_code_date ON screening_rows(stock_code, file_date_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_screening_date_score_code ON screening_rows(file_date_key, score_s DESC, stock_code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_close_cache_date ON daily_close_cache(file_date_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_close_cache_code_date ON daily_close_cache(stock_code, file_date_key)")
        conn.commit()
    return True


def us_screening_available_file_entries(limit: int | None = None) -> list[dict[str, str]]:
    ensure_us_screening_db_indexes()
    if not US_SCREENING_FAST_DB_PATH.exists():
        return []
    query = """
        SELECT file_date_key, file_name
        FROM file_meta
        ORDER BY file_date_key DESC
    """
    if limit is not None:
        query += f" LIMIT {max(1, int(limit))}"
    with sqlite3.connect(str(US_SCREENING_FAST_DB_PATH)) as conn:
        rows = conn.execute(query).fetchall()
    result: list[dict[str, str]] = []
    for file_date_key, file_name in rows:
        date_key = str(file_date_key or "")
        if not re.fullmatch(r"20\d{6}", date_key):
            continue
        result.append(
            {
                "file_date_key": date_key,
                "file_date": f"{date_key[:4]}-{date_key[4:6]}-{date_key[6:]}",
                "file_name": str(file_name or f"{date_key}_us_daily_screening.xlsx"),
            }
        )
    return result


def load_us_screening_summaries_for_dates(file_dates: list[str]) -> list[dict[str, Any]]:
    ensure_us_screening_db_indexes()
    if not US_SCREENING_FAST_DB_PATH.exists():
        return []
    normalized_dates = []
    for file_date in file_dates:
        digits = re.sub(r"\D", "", str(file_date or ""))
        if re.fullmatch(r"20\d{6}", digits):
            normalized_dates.append(digits)
    normalized_dates = sorted(set(normalized_dates))
    if not normalized_dates:
        return []
    placeholders = ",".join(["?"] * len(normalized_dates))
    with sqlite3.connect(str(US_SCREENING_FAST_DB_PATH)) as conn:
        existing_columns = {
            str(row[1])
            for row in conn.execute("PRAGMA table_info(screening_rows)").fetchall()
            if row and len(row) > 1
        }
        rank_select = "rank" if "rank" in existing_columns else "NULL AS rank"
        query = f"""
        SELECT
            file_date,
            file_date_key,
            {rank_select},
            stock_code,
            stock_name,
            sector,
            industry,
            market_cap_100m,
            trading_value_100m,
            change_pct,
            score_o,
            score_s,
            sortino_norm,
            is_52w_high,
            avg_1w,
            avg_3m,
            note
        FROM screening_rows
        WHERE file_date_key IN ({placeholders})
        ORDER BY file_date_key ASC, score_s DESC, stock_code ASC
    """
        frame = pd.read_sql_query(query, conn, params=normalized_dates)
    if frame.empty:
        return []
    grouped: list[dict[str, Any]] = []
    for file_date, group in frame.groupby("file_date", sort=True):
        qualified_stocks: list[dict[str, Any]] = []
        for index_no, (_, row) in enumerate(group.iterrows(), start=1):
            major_sector = str(row.get("sector") or "").strip()
            detailed_sector = str(row.get("industry") or "").strip()
            qualified_stocks.append(
                {
                    "file_date": str(row.get("file_date") or ""),
                    "file_date_key": str(row.get("file_date_key") or ""),
                    "rank": int(to_float(row.get("rank")) or index_no),
                    "stock_code": str(row.get("stock_code") or "").strip().upper(),
                    "stock_name": str(row.get("stock_name") or ""),
                    "resolved_name": str(row.get("stock_name") or ""),
                    "manual_sector": detailed_sector or major_sector,
                    "theme": detailed_sector or major_sector,
                    "industry": major_sector,
                    "market_cap_100m": to_float(row.get("market_cap_100m")),
                    "trading_value_100m": to_float(row.get("trading_value_100m")),
                    "change_pct": to_float(row.get("change_pct")) or 0.0,
                    "score_o": to_float(row.get("score_o")) or 0.0,
                    "score": to_float(row.get("score_s")) or 0.0,
                    "sortino_norm": to_float(row.get("sortino_norm")) or 0.0,
                    "is_52w_high": int(to_float(row.get("is_52w_high")) or 0),
                    "avg_1w": to_float(row.get("avg_1w")) or 0.0,
                    "avg_3m": to_float(row.get("avg_3m")) or 0.0,
                    "note": str(row.get("note") or ""),
                }
            )
        grouped.append(
            {
                "file_date": str(file_date or ""),
                "file_date_key": datetime.strptime(str(file_date), "%Y-%m-%d").strftime("%Y%m%d") if file_date else "",
                "qualified_stocks": qualified_stocks,
            }
        )
    return grouped


def load_us_screening_close_map(
    start_date: date,
    end_date: date,
    stock_codes: list[str],
) -> dict[str, dict[str, float]]:
    if not US_SCREENING_FAST_DB_PATH.exists():
        return {}
    normalized_codes = sorted({str(code or "").strip().upper() for code in stock_codes if str(code or "").strip()})
    if not normalized_codes:
        return {}
    placeholders = ",".join(["?"] * len(normalized_codes))
    params: list[Any] = [start_date.strftime("%Y%m%d"), end_date.strftime("%Y%m%d"), *normalized_codes]
    query = f"""
        SELECT file_date_key, stock_code, close_price
        FROM daily_close_cache
        WHERE file_date_key BETWEEN ? AND ?
          AND stock_code IN ({placeholders})
    """
    with sqlite3.connect(str(US_SCREENING_FAST_DB_PATH)) as conn:
        rows = conn.execute(query, params).fetchall()
    result: dict[str, dict[str, float]] = {}
    for file_date_key, stock_code, close_price in rows:
        key = str(file_date_key or "")
        code = str(stock_code or "").strip().upper()
        price = to_float(close_price)
        if not key or not code or price in (None, 0):
            continue
        result.setdefault(key, {})[code] = float(price)
    return result


@lru_cache(maxsize=1)
def ensure_asia_screening_db_indexes() -> bool:
    if not ASIA_SCREENING_FAST_DB_PATH.exists():
        return False
    with sqlite3.connect(str(ASIA_SCREENING_FAST_DB_PATH)) as conn:
        conn.execute("CREATE INDEX IF NOT EXISTS idx_screening_date ON screening_rows(file_date_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_screening_code_date ON screening_rows(stock_code, file_date_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_screening_date_score_code ON screening_rows(file_date_key, score_s DESC, stock_code)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_close_cache_date ON daily_close_cache(file_date_key)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_daily_close_cache_code_date ON daily_close_cache(stock_code, file_date_key)")
        conn.commit()
    return True


def asia_screening_available_file_entries(limit: int | None = None) -> list[dict[str, str]]:
    ensure_asia_screening_db_indexes()
    if not ASIA_SCREENING_FAST_DB_PATH.exists():
        return []
    query = """
        SELECT file_date_key, file_name
        FROM file_meta
        ORDER BY file_date_key DESC
    """
    if limit is not None:
        query += f" LIMIT {max(1, int(limit))}"
    with sqlite3.connect(str(ASIA_SCREENING_FAST_DB_PATH)) as conn:
        rows = conn.execute(query).fetchall()
    result: list[dict[str, str]] = []
    for file_date_key, file_name in rows:
        date_key = str(file_date_key or "")
        if not re.fullmatch(r"20\d{6}", date_key):
            continue
        result.append(
            {
                "file_date_key": date_key,
                "file_date": f"{date_key[:4]}-{date_key[4:6]}-{date_key[6:]}",
                "file_name": str(file_name or f"{date_key}_asia_daily_screening.xlsx"),
            }
        )
    return result


def load_asia_screening_summaries_for_dates(file_dates: list[str]) -> list[dict[str, Any]]:
    ensure_asia_screening_db_indexes()
    if not ASIA_SCREENING_FAST_DB_PATH.exists():
        return []
    normalized_dates = []
    for file_date in file_dates:
        digits = re.sub(r"\D", "", str(file_date or ""))
        if re.fullmatch(r"20\d{6}", digits):
            normalized_dates.append(digits)
    normalized_dates = sorted(set(normalized_dates))
    if not normalized_dates:
        return []
    placeholders = ",".join(["?"] * len(normalized_dates))
    with sqlite3.connect(str(ASIA_SCREENING_FAST_DB_PATH)) as conn:
        existing_columns = {
            str(row[1])
            for row in conn.execute("PRAGMA table_info(screening_rows)").fetchall()
            if row and len(row) > 1
        }
        rank_select = "rank" if "rank" in existing_columns else "NULL AS rank"
        query = f"""
        SELECT
            file_date,
            file_date_key,
            {rank_select},
            stock_code,
            stock_name,
            sector,
            industry,
            market_cap_100m,
            trading_value_100m,
            change_pct,
            score_o,
            score_s,
            sortino_norm,
            is_52w_high,
            avg_1w,
            avg_3m,
            note
        FROM screening_rows
        WHERE file_date_key IN ({placeholders})
        ORDER BY file_date_key ASC, score_s DESC, stock_code ASC
    """
        frame = pd.read_sql_query(query, conn, params=normalized_dates)
    if frame.empty:
        return []
    grouped: list[dict[str, Any]] = []
    for file_date, group in frame.groupby("file_date", sort=True):
        qualified_stocks: list[dict[str, Any]] = []
        for index_no, (_, row) in enumerate(group.iterrows(), start=1):
            qualified_stocks.append(
                {
                    "file_date": str(row.get("file_date") or ""),
                    "file_date_key": str(row.get("file_date_key") or ""),
                    "rank": int(to_float(row.get("rank")) or index_no),
                    "stock_code": str(row.get("stock_code") or "").strip().upper(),
                    "stock_name": str(row.get("stock_name") or ""),
                    "resolved_name": str(row.get("stock_name") or ""),
                    "manual_sector": str(row.get("sector") or ""),
                    "theme": str(row.get("sector") or ""),
                    "industry": str(row.get("industry") or ""),
                    "market_cap_100m": to_float(row.get("market_cap_100m")),
                    "trading_value_100m": to_float(row.get("trading_value_100m")),
                    "change_pct": to_float(row.get("change_pct")) or 0.0,
                    "score_o": to_float(row.get("score_o")) or 0.0,
                    "score": to_float(row.get("score_s")) or 0.0,
                    "sortino_norm": to_float(row.get("sortino_norm")) or 0.0,
                    "is_52w_high": int(to_float(row.get("is_52w_high")) or 0),
                    "avg_1w": to_float(row.get("avg_1w")) or 0.0,
                    "avg_3m": to_float(row.get("avg_3m")) or 0.0,
                    "note": str(row.get("note") or ""),
                }
            )
        grouped.append(
            {
                "file_date": str(file_date or ""),
                "file_date_key": datetime.strptime(str(file_date), "%Y-%m-%d").strftime("%Y%m%d") if file_date else "",
                "qualified_stocks": qualified_stocks,
            }
        )
    return grouped


def load_strategy_benchmark_close_map(index_key: str, start_date: date, end_date: date) -> dict[str, float]:
    mapping = {"KS11": "1001", "KQ11": "2001"}
    if index_key not in mapping:
        return {}

    def close_map_from_frame(frame: pd.DataFrame, close_col_hint: str | None = None) -> dict[str, float]:
        if frame is None or frame.empty:
            return {}
        working = frame.reset_index()
        date_col = "Date" if "Date" in working.columns else working.columns[0]
        close_col = close_col_hint if close_col_hint and close_col_hint in working.columns else None
        if close_col is None:
            close_col = "Close" if "Close" in working.columns else ("종가" if "종가" in working.columns else working.columns[-1])
        output: dict[str, float] = {}
        for _, row in working.iterrows():
            dt = pd.to_datetime(row.get(date_col), errors="coerce")
            close_val = to_float(row.get(close_col))
            if pd.isna(dt) or close_val in (None, 0):
                continue
            output[dt.date().isoformat()] = float(close_val)
        return output

    if pykrx_stock is not None:
        try:
            frame = pykrx_stock.get_index_ohlcv_by_date(
                start_date.strftime("%Y%m%d"),
                end_date.strftime("%Y%m%d"),
                mapping[index_key],
            )
            output = close_map_from_frame(frame, close_col_hint="종가")
            if output:
                return output
        except Exception:
            pass

    symbol = str((STRATEGY_INDEXES.get(index_key) or {}).get("symbol") or "").strip()
    if not symbol:
        return {}
    try:
        frame = load_strategy_price_frame(symbol, start_date, end_date)
        return close_map_from_frame(frame, close_col_hint="Close")
    except Exception:
        return {}


def sector_rotation_sector_key(row: dict[str, Any], sector_db: dict[str, Any]) -> str:
    manual_sector = resolve_sector_for_stock(row.get("stock_code"), row.get("stock_name"), sector_db)
    if manual_sector:
        return manual_sector
    for key in ["theme", "industry", "market"]:
        value = str(row.get(key) or "").strip()
        if value and value != "-":
            return value
    return "미분류"


def sector_rotation_stock_key(row: dict[str, Any]) -> str:
    code = normalize_stock_code_value(row.get("stock_code"))
    if code:
        return code
    return normalize_text(row.get("stock_name"))


def build_sector_rotation_signals(summary: dict[str, Any], min_score: float, sector_db: dict[str, Any]) -> list[dict[str, Any]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for raw in summary.get("qualified_stocks", []):
        if not isinstance(raw, dict):
            continue
        score = to_float(raw.get("score"))
        market_cap = to_float(raw.get("market_cap_100m"))
        if score is None or score < min_score:
            continue
        if market_cap is not None and market_cap < SCREENING_MIN_MARKET_CAP_100M:
            continue
        row = dict(raw)
        row["sector"] = sector_rotation_sector_key(row, sector_db)
        grouped.setdefault(row["sector"], []).append(row)

    signals: list[dict[str, Any]] = []
    for sector, stocks in grouped.items():
        if not stocks:
            continue
        scores = [float(to_float(item.get("score")) or 0.0) for item in stocks]
        trading_values = [float(to_float(item.get("trading_value_100m")) or 0.0) for item in stocks]
        market_caps = [float(to_float(item.get("market_cap_100m")) or 0.0) for item in stocks]
        total_trading = sum(trading_values)
        total_market_cap = sum(value for value in market_caps if value > 0)
        turnover_ratio = (total_trading / total_market_cap) if total_market_cap > 0 else 0.0
        avg_score = float(np.mean(scores)) if scores else 0.0
        max_score = max(scores) if scores else 0.0
        strong_count = sum(1 for value in scores if value >= 70)
        strength_score = avg_score + max_score * 0.18 + min(len(stocks), 12) * 1.4 + min(turnover_ratio * 100.0, 30.0)
        sorted_stocks = sorted(
            stocks,
            key=lambda item: (
                float(to_float(item.get("score")) or -9999),
                float(to_float(item.get("trading_value_100m")) or 0),
            ),
            reverse=True,
        )
        signals.append(
            {
                "sector": sector,
                "stock_count": len(stocks),
                "avg_score": round(avg_score, 2),
                "max_score": round(max_score, 2),
                "strong_count": strong_count,
                "total_trading_value_100m": round(total_trading, 1),
                "total_market_cap_100m": round(total_market_cap, 1),
                "turnover_ratio_pct": round(turnover_ratio * 100, 3),
                "strength_score": round(strength_score, 3),
                "stocks": sorted_stocks,
            }
        )
    return sorted(signals, key=lambda item: item["strength_score"], reverse=True)


def max_drawdown_pct(nav_values: list[float]) -> float:
    if not nav_values:
        return 0.0
    peak = nav_values[0]
    max_dd = 0.0
    for value in nav_values:
        peak = max(peak, value)
        if peak > 0:
            max_dd = min(max_dd, value / peak - 1)
    return round(max_dd * 100, 2)


def annualized_return_pct(total_return_pct: float, start_date: str, end_date: str) -> float | None:
    try:
        start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
    except Exception:
        return None
    days = max(1, (end_dt - start_dt).days)
    total = 1 + total_return_pct / 100.0
    if total <= 0:
        return None
    return round((total ** (365.0 / days) - 1) * 100, 2)


def build_sector_rotation_backtest(
    start: str | None = None,
    end: str | None = None,
    min_score: float = 50.0,
    top_sectors: int = 4,
    stocks_per_sector: int = 4,
    leverage: float = 1.0,
    weight_method: str = "strength",
) -> dict[str, Any]:
    summaries = screening_backtest_source_summaries()
    if len(summaries) < 2:
        raise ValueError("섹터 로테이션 백테스트에 필요한 주도주 캐시가 부족합니다.")

    available_dates = [str(item.get("file_date") or "") for item in summaries if item.get("file_date")]
    fallback_start = available_dates[max(0, len(available_dates) - 30)]
    fallback_end = available_dates[-1]
    start_date = safe_strategy_date(start, datetime.strptime(fallback_start, "%Y-%m-%d").date()).isoformat()
    end_date = safe_strategy_date(end, datetime.strptime(fallback_end, "%Y-%m-%d").date()).isoformat()
    filtered = [item for item in summaries if start_date <= str(item.get("file_date") or "") <= end_date]
    if len(filtered) < 2:
        filtered = summaries[-min(30, len(summaries)) :]

    top_sectors = max(1, min(int(top_sectors or 4), 12))
    stocks_per_sector = max(1, min(int(stocks_per_sector or 4), 20))
    leverage = max(0.0, min(float(leverage or 1.0), 2.5))
    weight_method = weight_method if weight_method in {"equal_sector", "strength"} else "strength"

    sector_db = load_sector_db()
    nav = 100.0
    benchmark_nav = 100.0
    previous_weights: dict[str, float] = {}
    rows: list[dict[str, Any]] = []
    sector_performance: dict[str, dict[str, Any]] = {}
    latest_signals: list[dict[str, Any]] = []
    total_turnover = 0.0
    total_fee_pct = 0.0

    for index_no in range(len(filtered) - 1):
        current = filtered[index_no]
        next_summary = filtered[index_no + 1]
        current_date = str(current.get("file_date") or "")
        next_date = str(next_summary.get("file_date") or "")
        signals = build_sector_rotation_signals(current, min_score=min_score, sector_db=sector_db)
        latest_signals = signals
        selected = signals[:top_sectors]
        next_rows_by_key = {
            sector_rotation_stock_key(row): row
            for row in next_summary.get("qualified_stocks", [])
            if isinstance(row, dict) and sector_rotation_stock_key(row)
        }

        target_weights: dict[str, float] = {}
        stock_payload_by_key: dict[str, dict[str, Any]] = {}
        selected_sectors_payload: list[dict[str, Any]] = []
        sector_weight_inputs = [max(0.01, float(item.get("strength_score") or 0)) for item in selected]
        sector_weight_sum = sum(sector_weight_inputs) if weight_method == "strength" else len(selected)

        for sector_index, signal in enumerate(selected):
            sector_weight = (sector_weight_inputs[sector_index] / sector_weight_sum) if sector_weight_sum else 0.0
            if weight_method == "equal_sector" and selected:
                sector_weight = 1.0 / len(selected)
            chosen_stocks = signal.get("stocks", [])[:stocks_per_sector]
            if not chosen_stocks:
                continue
            stock_weight = leverage * sector_weight / len(chosen_stocks)
            selected_stock_rows = []
            for stock in chosen_stocks:
                stock_key = sector_rotation_stock_key(stock)
                if not stock_key:
                    continue
                target_weights[stock_key] = target_weights.get(stock_key, 0.0) + stock_weight
                stock_payload_by_key[stock_key] = stock
                selected_stock_rows.append(
                    {
                        "stock_key": stock_key,
                        "stock_name": stock.get("resolved_name") or stock.get("stock_name") or stock_key,
                        "stock_code": stock.get("stock_code") or "",
                        "score": stock.get("score"),
                        "weight_pct": round(stock_weight * 100, 2),
                    }
                )
            selected_sectors_payload.append(
                {
                    "sector": signal["sector"],
                    "strength_score": signal["strength_score"],
                    "avg_score": signal["avg_score"],
                    "stock_count": signal["stock_count"],
                    "target_weight_pct": round(sector_weight * leverage * 100, 2),
                    "stocks": selected_stock_rows,
                }
            )

        all_keys = set(previous_weights) | set(target_weights)
        turnover = sum(abs(target_weights.get(key, 0.0) - previous_weights.get(key, 0.0)) for key in all_keys)
        fee = turnover * STRATEGY_TRADE_FEE_RATE
        total_turnover += turnover
        total_fee_pct += fee * 100

        daily_return = -fee
        sector_return_map: dict[str, float] = {}
        sector_weight_map: dict[str, float] = {}
        for stock_key, weight in target_weights.items():
            next_row = next_rows_by_key.get(stock_key)
            stock_return = (to_float(next_row.get("change_pct")) if next_row else None) if next_row else None
            if stock_return is None:
                stock_return = 0.0
            contribution = weight * (stock_return / 100.0)
            daily_return += contribution
            stock_payload = stock_payload_by_key.get(stock_key, {})
            sector_name = sector_rotation_sector_key(stock_payload, sector_db) if stock_payload else "미분류"
            sector_return_map[sector_name] = sector_return_map.get(sector_name, 0.0) + contribution
            sector_weight_map[sector_name] = sector_weight_map.get(sector_name, 0.0) + weight

        benchmark_keys: list[str] = []
        for row in current.get("qualified_stocks", []):
            if not isinstance(row, dict):
                continue
            score = to_float(row.get("score"))
            market_cap = to_float(row.get("market_cap_100m"))
            if score is None or float(score) < min_score:
                continue
            if market_cap is not None and market_cap < SCREENING_MIN_MARKET_CAP_100M:
                continue
            stock_key = sector_rotation_stock_key(row)
            if stock_key:
                benchmark_keys.append(stock_key)
        benchmark_returns = []
        for stock_key in benchmark_keys:
            next_row = next_rows_by_key.get(stock_key)
            stock_return = (to_float(next_row.get("change_pct")) if next_row else None) if next_row else None
            benchmark_returns.append(float(stock_return or 0.0) / 100.0)
        benchmark_daily_return = float(np.mean(benchmark_returns)) if benchmark_returns else 0.0

        nav *= 1 + daily_return
        benchmark_nav *= 1 + benchmark_daily_return
        for sector_name, contribution in sector_return_map.items():
            perf = sector_performance.setdefault(
                sector_name,
                {"sector": sector_name, "selected_days": 0, "contribution_pct": 0.0, "avg_weight_pct": 0.0, "win_days": 0},
            )
            perf["selected_days"] += 1
            perf["contribution_pct"] += contribution * 100.0
            perf["avg_weight_pct"] += sector_weight_map.get(sector_name, 0.0) * 100.0
            if contribution > 0:
                perf["win_days"] += 1

        rows.append(
            {
                "date": next_date,
                "signal_date": current_date,
                "strategy_return_pct": round(nav - 100.0, 3),
                "benchmark_return_pct": round(benchmark_nav - 100.0, 3),
                "daily_return_pct": round(daily_return * 100.0, 3),
                "benchmark_daily_return_pct": round(benchmark_daily_return * 100.0, 3),
                "exposure_pct": round(sum(target_weights.values()) * 100.0, 2),
                "turnover_pct": round(turnover * 100.0, 2),
                "fee_pct": round(fee * 100.0, 3),
                "top_sector": selected_sectors_payload[0]["sector"] if selected_sectors_payload else "현금",
                "selected_sectors": selected_sectors_payload,
            }
        )
        previous_weights = target_weights

    nav_values = [100.0] + [100.0 + float(row.get("strategy_return_pct") or 0.0) for row in rows]
    benchmark_values = [100.0] + [100.0 + float(row.get("benchmark_return_pct") or 0.0) for row in rows]
    final_strategy = rows[-1]["strategy_return_pct"] if rows else 0.0
    final_benchmark = rows[-1]["benchmark_return_pct"] if rows else 0.0
    sector_rows = []
    for perf in sector_performance.values():
        selected_days = max(1, int(perf["selected_days"]))
        sector_rows.append(
            {
                "sector": perf["sector"],
                "selected_days": perf["selected_days"],
                "contribution_pct": round(perf["contribution_pct"], 3),
                "avg_weight_pct": round(perf["avg_weight_pct"] / selected_days, 2),
                "win_rate_pct": round(perf["win_days"] / selected_days * 100.0, 1),
            }
        )
    sector_rows.sort(key=lambda item: item["contribution_pct"], reverse=True)
    latest_signal_rows = [
        {
            "sector": item["sector"],
            "strength_score": item["strength_score"],
            "avg_score": item["avg_score"],
            "max_score": item["max_score"],
            "stock_count": item["stock_count"],
            "turnover_ratio_pct": item["turnover_ratio_pct"],
            "leaders": ", ".join([
                str(stock.get("resolved_name") or stock.get("stock_name") or "")
                for stock in item.get("stocks", [])[:5]
                if str(stock.get("resolved_name") or stock.get("stock_name") or "").strip()
            ]),
        }
        for item in latest_signals[:20]
    ]

    start_text = rows[0]["date"] if rows else filtered[0].get("file_date")
    end_text = rows[-1]["date"] if rows else filtered[-1].get("file_date")
    summary = {
        "strategy_return_pct": round(final_strategy, 2),
        "benchmark_return_pct": round(final_benchmark, 2),
        "excess_return_pct": round(final_strategy - final_benchmark, 2),
        "mdd_pct": max_drawdown_pct(nav_values),
        "benchmark_mdd_pct": max_drawdown_pct(benchmark_values),
        "cagr_pct": annualized_return_pct(final_strategy, str(start_text), str(end_text)),
        "avg_exposure_pct": round(float(np.mean([row["exposure_pct"] for row in rows])) if rows else 0.0, 2),
        "total_turnover_pct": round(total_turnover * 100.0, 2),
        "total_fee_pct_points": round(total_fee_pct, 2),
        "rebalance_count": len(rows),
    }
    return {
        "mode": "sector_rotation",
        "strategy_name": "주도주 섹터 로테이션",
        "benchmark_name": "시총 2000억 이상 주도주 평균",
        "start_date": start_text,
        "end_date": end_text,
        "params": {
            "min_score": min_score,
            "top_sectors": top_sectors,
            "stocks_per_sector": stocks_per_sector,
            "leverage": leverage,
            "weight_method": weight_method,
            "fee_rate_pct": round(STRATEGY_TRADE_FEE_RATE * 100.0, 2),
        },
        "rows": rows,
        "summary": summary,
        "sector_performance": sector_rows,
        "latest_sector_scores": latest_signal_rows,
        "available_dates": available_dates,
        "description": "오늘의 주도주 점수와 수동 섹터 DB를 이용해 D일 강한 섹터를 D+1일 보유한다고 가정합니다.",
    }


def price_close_on_or_before(frame: pd.DataFrame, target_date: str | date) -> float | None:
    if frame is None or frame.empty or "Close" not in frame.columns:
        return None
    try:
        target = pd.to_datetime(target_date).date()
    except Exception:
        return None
    working = frame.copy()
    if "Date" in working.columns:
        working["_date"] = pd.to_datetime(working["Date"], errors="coerce").dt.date
    else:
        working["_date"] = pd.to_datetime(working.index, errors="coerce").date
    working = working[working["_date"] <= target].dropna(subset=["_date", "Close"])
    if working.empty:
        return None
    return to_float(working.iloc[-1].get("Close"))


def index_return_between(symbol: str, from_date: str, to_date: str) -> float:
    try:
        start_dt = datetime.strptime(from_date, "%Y-%m-%d").date() - timedelta(days=320)
        end_dt = datetime.strptime(to_date, "%Y-%m-%d").date()
        frame = load_strategy_price_frame(symbol, start_dt, end_dt)
        start_close = price_close_on_or_before(frame, from_date)
        end_close = price_close_on_or_before(frame, to_date)
        if start_close in (None, 0) or end_close is None:
            return 0.0
        return float(end_close / start_close - 1.0)
    except Exception:
        return 0.0


def portfolio_series_return_between(series: list[dict[str, Any]], from_date: str, to_date: str) -> float:
    if not series:
        return 0.0

    def value_on_or_before(target_date: str) -> float | None:
        try:
            target = datetime.strptime(str(target_date), "%Y-%m-%d").date()
        except Exception:
            return None
        best_date: date | None = None
        best_value: float | None = None
        for item in series:
            try:
                item_date = datetime.strptime(str(item.get("date") or ""), "%Y-%m-%d").date()
            except Exception:
                continue
            if item_date > target:
                continue
            if best_date is None or item_date >= best_date:
                value = to_float(item.get("return_pct"))
                if value is not None:
                    best_date = item_date
                    best_value = 1.0 + value / 100.0
        return best_value

    start_value = value_on_or_before(from_date)
    end_value = value_on_or_before(to_date)
    if start_value in (None, 0) or end_value is None:
        return 0.0
    return float(end_value / start_value - 1.0)


def stock_return_between(stock_code: str, from_date: str, to_date: str, fallback_pct: float | None = None) -> float:
    code = normalize_stock_code_value(stock_code)
    if code:
        try:
            frame = fetch_price_frame(code)
            start_close = price_close_on_or_before(frame, from_date)
            end_close = price_close_on_or_before(frame, to_date)
            if start_close not in (None, 0) and end_close is not None:
                return float(end_close / start_close - 1.0)
        except Exception:
            pass
    if fallback_pct is not None:
        return float(fallback_pct) / 100.0
    return 0.0


def stock_technical_on_date(stock_code: str, target_date: str) -> dict[str, Any]:
    code = normalize_stock_code_value(stock_code)
    if not code:
        return {"above_ma20": None, "disparity": None}
    try:
        frame = fetch_price_frame(code)
        if frame.empty or "Close" not in frame.columns:
            return {"above_ma20": None, "disparity": None}
        working = frame.copy()
        working["_date"] = pd.to_datetime(working["Date"], errors="coerce").dt.date if "Date" in working.columns else pd.to_datetime(working.index, errors="coerce").date
        target = pd.to_datetime(target_date).date()
        working = working[working["_date"] <= target].sort_values("_date")
        if len(working) < 20:
            return {"above_ma20": None, "disparity": None}
        close = pd.to_numeric(working["Close"], errors="coerce")
        ma20 = close.rolling(20).mean()
        latest_close = to_float(close.iloc[-1])
        latest_ma20 = to_float(ma20.iloc[-1])
        if latest_close is None or latest_ma20 in (None, 0):
            return {"above_ma20": None, "disparity": None}
        disparity = latest_close / latest_ma20 * 100.0
        return {"above_ma20": latest_close > latest_ma20, "disparity": disparity}
    except Exception:
        return {"above_ma20": None, "disparity": None}


def advanced_stock_trend_score(stock: dict[str, Any], technical: dict[str, Any], beta: float | None = None, max_disparity: float = 110.0) -> float:
    score = float(to_float(stock.get("score")) or 0.0)
    trading_value = float(to_float(stock.get("trading_value_100m")) or 0.0)
    market_cap = float(to_float(stock.get("market_cap_100m")) or 0.0)
    turnover_pct = trading_value / market_cap * 100.0 if market_cap > 0 else 0.0
    above_ma20 = technical.get("above_ma20")
    disparity = to_float(technical.get("disparity"))
    beta_score = max(0.0, min(float(beta or 1.0), 2.5)) * 3.0
    trend_bonus = 14.0 if above_ma20 is True else -18.0 if above_ma20 is False else -6.0
    if disparity is None:
        disparity_penalty = 6.0
    elif disparity > max_disparity:
        disparity_penalty = (disparity - max_disparity) * 4.0
    else:
        disparity_penalty = abs(disparity - 104.0) * 0.7
    return round(
        score
        + min(turnover_pct, 12.0) * 2.0
        + trend_bonus
        + beta_score
        - disparity_penalty,
        3,
    )


def advanced_market_timing_state(signal_date: str) -> dict[str, Any]:
    markets = []
    for key, symbol in [("kospi", "KS11"), ("kosdaq", "KQ11")]:
        try:
            target = datetime.strptime(signal_date, "%Y-%m-%d").date()
            frame = load_strategy_price_frame(symbol, target - timedelta(days=120), target)
            frame = frame[frame["Date"] <= target]
            if frame.empty:
                markets.append({"key": key, "above_ma20": False, "gap_pct": 0.0})
                continue
            row = frame.iloc[-1]
            close = to_float(row.get("Close"))
            ma20 = to_float(row.get("ma20"))
            above = bool(close is not None and ma20 is not None and close > ma20)
            gap_pct = ((close / ma20 - 1.0) * 100.0) if close is not None and ma20 not in (None, 0) else 0.0
            markets.append({"key": key, "above_ma20": above, "gap_pct": round(gap_pct, 2)})
        except Exception:
            markets.append({"key": key, "above_ma20": False, "gap_pct": 0.0})
    above_count = sum(1 for item in markets if item["above_ma20"])
    if above_count == 0:
        multiplier = 0.0
        label = "KOSPI/KOSDAQ 모두 20일선 아래"
    elif above_count == 1:
        multiplier = 0.55
        label = "양 지수 중 1개만 20일선 위"
    else:
        avg_gap = sum(float(item.get("gap_pct") or 0.0) for item in markets) / 2.0
        multiplier = 1.0 if avg_gap >= 0 else 0.75
        label = "KOSPI/KOSDAQ 모두 20일선 위"
    return {"markets": markets, "above_count": above_count, "multiplier": multiplier, "label": label}


def build_advanced_sector_backtest(
    start: str | None = None,
    end: str | None = None,
    min_score: float = 50.0,
    top_sectors: int = 4,
    stocks_per_sector: int = 4,
    max_leverage: float = 1.5,
    benchmark: str = "KS11",
    weight_method: str = "entry_beta",
    beta_window: int = 63,
    min_breadth: float = 60.0,
    max_disparity: float = 110.0,
    trading_rank_limit: int = 20,
    stock_selection: str = "trend_strength",
) -> dict[str, Any]:
    summaries = screening_backtest_source_summaries()
    if len(summaries) < 2:
        raise ValueError("고급 백테스트에 필요한 주도주 캐시가 부족합니다.")

    available_dates = [str(item.get("file_date") or "") for item in summaries if item.get("file_date")]
    fallback_start = available_dates[max(0, len(available_dates) - 60)]
    fallback_end = available_dates[-1]
    start_date = safe_strategy_date(start, datetime.strptime(fallback_start, "%Y-%m-%d").date()).isoformat()
    end_date = safe_strategy_date(end, datetime.strptime(fallback_end, "%Y-%m-%d").date()).isoformat()
    filtered = [item for item in summaries if start_date <= str(item.get("file_date") or "") <= end_date]
    if len(filtered) < 2:
        filtered = summaries[-min(60, len(summaries)) :]

    min_score = max(0.0, min(float(min_score or 50.0), 100.0))
    top_sectors = max(1, min(int(top_sectors or 4), 10))
    stocks_per_sector = max(1, min(int(stocks_per_sector or 4), 12))
    max_leverage = max(0.0, min(float(max_leverage or 1.5), 2.5))
    beta_window = 63 if int(beta_window or 63) <= 100 else 252
    min_breadth = max(0.0, min(float(min_breadth or 60.0), 100.0))
    max_disparity = max(100.0, min(float(max_disparity or 110.0), 140.0))
    trading_rank_limit = max(1, min(int(trading_rank_limit or 20), 50))
    stock_selection = stock_selection if stock_selection in {"score", "trend_strength"} else "trend_strength"
    benchmark_key = str(benchmark or "KS11")
    if benchmark_key not in {"leader_avg", "portfolio_dashboard", "KS11", "KQ11", "KRX_BLEND", "IXIC", "US500"}:
        benchmark_key = "KS11"

    sector_db = load_sector_db()
    nav = 100.0
    benchmark_nav = 100.0
    previous_weights: dict[str, float] = {}
    previous_payload: dict[str, dict[str, Any]] = {}
    rows: list[dict[str, Any]] = []
    trade_log: list[dict[str, Any]] = []
    holdings_timeline: list[dict[str, Any]] = []
    rejected_signals: list[dict[str, Any]] = []
    active_trade_entries: dict[str, dict[str, Any]] = {}
    total_turnover = 0.0
    total_fee_pct = 0.0
    portfolio_benchmark_series: list[dict[str, Any]] = []
    if benchmark_key == "portfolio_dashboard":
        try:
            portfolio_benchmark_series = calculate_portfolio_performance().get("series", [])
        except Exception:
            portfolio_benchmark_series = []

    def strategy_stock_trade_price(payload: dict[str, Any], target_date: str) -> float | None:
        code = normalize_stock_code_value(payload.get("stock_code") or payload.get("stock_key") or "")
        if not code:
            return None
        try:
            return price_close_on_or_before(fetch_price_frame(code), target_date)
        except Exception:
            return None

    for index_no in range(len(filtered) - 1):
        current = filtered[index_no]
        next_summary = filtered[index_no + 1]
        current_date = str(current.get("file_date") or "")
        next_date = str(next_summary.get("file_date") or "")
        market_state = advanced_market_timing_state(current_date)
        market_exposure = max_leverage * float(market_state.get("multiplier") or 0.0)

        window = filtered[max(0, index_no - 4) : index_no + 1]
        trading_totals: dict[str, float] = {}
        for window_summary in window:
            for signal in build_sector_rotation_signals(window_summary, min_score=min_score, sector_db=sector_db):
                sector = str(signal.get("sector") or "").strip()
                if sector:
                    trading_totals[sector] = trading_totals.get(sector, 0.0) + float(signal.get("total_trading_value_100m") or 0.0)
        trading_rank_rows = sorted(trading_totals.items(), key=lambda item: item[1], reverse=True)
        rank_map = {sector: rank + 1 for rank, (sector, _value) in enumerate(trading_rank_rows)}

        candidates: list[dict[str, Any]] = []
        for signal in build_sector_rotation_signals(current, min_score=min_score, sector_db=sector_db):
            sector = str(signal.get("sector") or "").strip()
            trading_rank = int(rank_map.get(sector, 9999))
            if trading_rank > trading_rank_limit:
                continue
            source_stocks = signal.get("stocks", [])[: max(stocks_per_sector * 2, 6)]
            enriched_stocks: list[dict[str, Any]] = []
            tech_rows = []
            for stock in source_stocks:
                technical = stock_technical_on_date(stock.get("stock_code"), current_date)
                tech_rows.append(technical)
                enriched_stocks.append(dict(stock, above_ma20=technical.get("above_ma20"), disparity=technical.get("disparity")))
            valid_tech = [item for item in tech_rows if item.get("above_ma20") is not None and item.get("disparity") is not None]
            breadth = sum(1 for item in valid_tech if item.get("above_ma20")) / len(valid_tech) * 100.0 if valid_tech else 0.0
            avg_disparity = float(np.mean([float(item.get("disparity") or 0.0) for item in valid_tech])) if valid_tech else 999.0
            entry_score = (
                float(signal.get("strength_score") or 0.0)
                + max(0.0, (trading_rank_limit + 1 - trading_rank)) * 2.0
                + min(float(signal.get("turnover_ratio_pct") or 0.0), 10.0) * 1.2
            )
            payload = dict(signal)
            payload["stocks"] = enriched_stocks
            payload.update(
                {
                    "date": current_date,
                    "trading_rank": trading_rank,
                    "breadth_ratio": round(breadth, 1),
                    "avg_disparity": round(avg_disparity, 2),
                    "entry_score": round(entry_score, 2),
                    "entry_pass": breadth >= min_breadth and avg_disparity <= max_disparity,
                    "reject_reason": "",
                }
            )
            if market_exposure <= 0:
                payload["entry_pass"] = False
                payload["reject_reason"] = "양 지수 20일선 아래"
            elif breadth < min_breadth:
                payload["reject_reason"] = f"20일선 위 종목 비율 {breadth:.1f}%"
            elif avg_disparity > max_disparity:
                payload["reject_reason"] = f"이격도 과열 {avg_disparity:.1f}%"
            if payload["entry_pass"]:
                candidates.append(payload)
            else:
                rejected_signals.append(
                    {
                        "date": current_date,
                        "sector": sector,
                        "entry_score": round(entry_score, 2),
                        "reason": payload["reject_reason"],
                    }
                )

        selected = sorted(candidates, key=lambda item: item["entry_score"], reverse=True)[:top_sectors]
        target_weights: dict[str, float] = {}
        stock_payload_by_key: dict[str, dict[str, Any]] = {}
        selected_sectors_payload: list[dict[str, Any]] = []
        entry_sum = sum(max(0.01, float(item.get("entry_score") or 0.0)) for item in selected)
        for sector_signal in selected:
            sector_weight = max(0.01, float(sector_signal.get("entry_score") or 0.0)) / entry_sum if entry_sum else 0.0
            chosen = sector_signal.get("stocks", [])
            beta_items = []
            for stock in chosen:
                stock_code = normalize_stock_code_value(stock.get("stock_code"))
                listing = find_listing_row_by_code(stock_code) if stock_code else None
                market = str((listing or {}).get("market") or stock.get("market") or "")
                beta = None
                if stock_code:
                    try:
                        beta = compute_stock_beta(fetch_price_frame(stock_code), market, window=beta_window)
                    except Exception:
                        beta = None
                beta_weight = max(0.35, min(float(beta or 1.0), 3.0)) if weight_method == "entry_beta" else 1.0
                trend_score = advanced_stock_trend_score(stock, stock, beta=beta, max_disparity=max_disparity)
                beta_items.append((stock, beta, beta_weight, trend_score))
            if stock_selection == "trend_strength":
                beta_items = sorted(
                    beta_items,
                    key=lambda item: (
                        float(item[3] or -9999),
                        float(to_float(item[0].get("score")) or -9999),
                        float(to_float(item[0].get("trading_value_100m")) or 0.0),
                    ),
                    reverse=True,
                )
            beta_items = beta_items[:stocks_per_sector]
            beta_sum = sum(item[2] for item in beta_items) or len(beta_items) or 1.0
            stock_rows = []
            for stock, beta, beta_weight, trend_score in beta_items:
                stock_key = sector_rotation_stock_key(stock)
                if not stock_key:
                    continue
                weight = market_exposure * sector_weight * beta_weight / beta_sum
                target_weights[stock_key] = target_weights.get(stock_key, 0.0) + weight
                stock_payload_by_key[stock_key] = dict(stock, sector=sector_signal.get("sector"), beta=beta, trend_score=trend_score)
                stock_rows.append(
                    {
                        "stock_key": stock_key,
                        "stock_code": normalize_stock_code_value(stock.get("stock_code")),
                        "stock_name": stock.get("resolved_name") or stock.get("stock_name") or stock_key,
                        "score": stock.get("score"),
                        "beta": beta,
                        "trend_score": trend_score,
                        "above_ma20": stock.get("above_ma20"),
                        "disparity": stock.get("disparity"),
                        "target_weight_pct": round(weight * 100.0, 2),
                    }
                )
            selected_sectors_payload.append(
                {
                    "sector": sector_signal.get("sector"),
                    "entry_score": sector_signal.get("entry_score"),
                    "breadth_ratio": sector_signal.get("breadth_ratio"),
                    "avg_disparity": sector_signal.get("avg_disparity"),
                    "target_weight_pct": round(sum(row["target_weight_pct"] for row in stock_rows), 2),
                    "stocks": stock_rows,
                }
            )

        all_keys = set(previous_weights) | set(target_weights)
        turnover = sum(abs(target_weights.get(key, 0.0) - previous_weights.get(key, 0.0)) for key in all_keys)
        fee = turnover * STRATEGY_TRADE_FEE_RATE
        total_turnover += turnover
        total_fee_pct += fee * 100.0
        next_rows_by_key = {
            sector_rotation_stock_key(row): row
            for row in next_summary.get("qualified_stocks", [])
            if isinstance(row, dict) and sector_rotation_stock_key(row)
        }
        daily_return = -fee
        for stock_key, weight in target_weights.items():
            fallback = to_float((next_rows_by_key.get(stock_key) or {}).get("change_pct"))
            payload = stock_payload_by_key.get(stock_key, {})
            daily_return += weight * stock_return_between(payload.get("stock_code") or stock_key, current_date, next_date, fallback)

        if benchmark_key == "leader_avg":
            benchmark_returns = []
            for stock_key in [
                sector_rotation_stock_key(row)
                for row in current.get("qualified_stocks", [])
                if isinstance(row, dict)
                and to_float(row.get("score")) is not None
                and float(to_float(row.get("score")) or 0) >= min_score
                and float(to_float(row.get("market_cap_100m")) or SCREENING_MIN_MARKET_CAP_100M) >= SCREENING_MIN_MARKET_CAP_100M
            ]:
                next_row = next_rows_by_key.get(stock_key)
                benchmark_returns.append(float(to_float((next_row or {}).get("change_pct")) or 0.0) / 100.0)
            benchmark_daily_return = float(np.mean(benchmark_returns)) if benchmark_returns else 0.0
            benchmark_name = "시총 2000억 이상 주도주 평균"
        elif benchmark_key == "portfolio_dashboard":
            benchmark_daily_return = portfolio_series_return_between(portfolio_benchmark_series, current_date, next_date)
            benchmark_name = "포트폴리오 수익 페이지"
        elif benchmark_key == "KRX_BLEND":
            benchmark_daily_return = (index_return_between("KS11", current_date, next_date) + index_return_between("KQ11", current_date, next_date)) / 2.0
            benchmark_name = "KOSPI/KOSDAQ 50:50"
        else:
            benchmark_daily_return = index_return_between(STRATEGY_INDEXES.get(benchmark_key, STRATEGY_INDEXES["KS11"])["symbol"], current_date, next_date)
            benchmark_name = STRATEGY_INDEXES.get(benchmark_key, STRATEGY_INDEXES["KS11"])["name"]

        nav *= 1 + daily_return
        benchmark_nav *= 1 + benchmark_daily_return

        for stock_key in sorted(all_keys):
            before = previous_weights.get(stock_key, 0.0)
            after = target_weights.get(stock_key, 0.0)
            diff = after - before
            if abs(diff) < 0.001:
                continue
            payload = stock_payload_by_key.get(stock_key) or previous_payload.get(stock_key) or {}
            action = "매수" if before <= 0 and after > 0 else "매도" if after <= 0 else "증액" if diff > 0 else "감액"
            trade_price = strategy_stock_trade_price({**payload, "stock_key": stock_key}, next_date)
            entry = active_trade_entries.get(stock_key)
            entry_date = str((entry or {}).get("entry_date") or "")
            entry_price = to_float((entry or {}).get("entry_price"))
            trade_return_pct = None
            holding_days = None
            if diff < 0 and entry_price not in (None, 0) and trade_price is not None:
                trade_return_pct = round((float(trade_price) / float(entry_price) - 1.0) * 100.0, 2)
                try:
                    holding_days = (datetime.strptime(next_date, "%Y-%m-%d").date() - datetime.strptime(entry_date, "%Y-%m-%d").date()).days
                except Exception:
                    holding_days = None
            trade_log.append(
                {
                    "date": current_date,
                    "apply_date": next_date,
                    "action": action,
                    "stock_key": stock_key,
                    "stock_code": payload.get("stock_code") or normalize_stock_code_value(stock_key),
                    "stock_name": payload.get("resolved_name") or payload.get("stock_name") or stock_key,
                    "sector": payload.get("sector") or "",
                    "before_weight_pct": round(before * 100.0, 2),
                    "after_weight_pct": round(after * 100.0, 2),
                    "change_weight_pct": round(diff * 100.0, 2),
                    "entry_date": entry_date,
                    "entry_price": round(entry_price, 2) if entry_price is not None else None,
                    "exit_date": next_date if diff < 0 else "",
                    "exit_price": round(trade_price, 2) if diff < 0 and trade_price is not None else None,
                    "trade_return_pct": trade_return_pct,
                    "holding_days": holding_days,
                    "beta": payload.get("beta"),
                    "score": payload.get("score"),
                    "trend_score": payload.get("trend_score"),
                }
            )
            if diff > 0 and trade_price is not None:
                if entry and entry_price not in (None, 0) and after > 0:
                    entry_weight = max(0.0, float((entry or {}).get("entry_weight") or before))
                    next_weight = max(0.0, entry_weight + diff)
                    blended_price = ((float(entry_price) * entry_weight) + (float(trade_price) * diff)) / next_weight if next_weight > 0 else float(trade_price)
                    active_trade_entries[stock_key] = {
                        "entry_date": entry_date or next_date,
                        "entry_price": blended_price,
                        "entry_weight": next_weight,
                    }
                else:
                    active_trade_entries[stock_key] = {
                        "entry_date": next_date,
                        "entry_price": float(trade_price),
                        "entry_weight": max(0.0, after),
                    }
            elif diff < 0:
                if after <= 0:
                    active_trade_entries.pop(stock_key, None)
                elif entry:
                    active_trade_entries[stock_key] = {
                        **entry,
                        "entry_weight": max(0.0, after),
                    }

        holding_rows = []
        for stock_key, weight in sorted(target_weights.items(), key=lambda item: item[1], reverse=True):
            payload = stock_payload_by_key.get(stock_key, {})
            holding_rows.append(
                {
                    "stock_key": stock_key,
                    "stock_code": payload.get("stock_code") or normalize_stock_code_value(stock_key),
                    "stock_name": payload.get("resolved_name") or payload.get("stock_name") or stock_key,
                    "sector": payload.get("sector") or "",
                    "weight_pct": round(weight * 100.0, 2),
                    "beta": payload.get("beta"),
                    "score": payload.get("score"),
                    "trend_score": payload.get("trend_score"),
                }
            )
        holdings_timeline.append({"date": next_date, "signal_date": current_date, "holdings": holding_rows})
        rows.append(
            {
                "date": next_date,
                "signal_date": current_date,
                "strategy_return_pct": round(nav - 100.0, 3),
                "benchmark_return_pct": round(benchmark_nav - 100.0, 3),
                "daily_return_pct": round(daily_return * 100.0, 3),
                "benchmark_daily_return_pct": round(benchmark_daily_return * 100.0, 3),
                "exposure_pct": round(sum(target_weights.values()) * 100.0, 2),
                "turnover_pct": round(turnover * 100.0, 2),
                "fee_pct": round(fee * 100.0, 3),
                "market_state": market_state,
                "selected_sectors": selected_sectors_payload,
                "holding_count": len(holding_rows),
            }
        )
        previous_weights = target_weights
        previous_payload = stock_payload_by_key

    nav_values = [100.0] + [100.0 + float(row.get("strategy_return_pct") or 0.0) for row in rows]
    benchmark_values = [100.0] + [100.0 + float(row.get("benchmark_return_pct") or 0.0) for row in rows]
    final_strategy = rows[-1]["strategy_return_pct"] if rows else 0.0
    final_benchmark = rows[-1]["benchmark_return_pct"] if rows else 0.0
    latest_holdings = holdings_timeline[-1]["holdings"] if holdings_timeline else []
    return {
        "mode": "advanced_sector",
        "strategy_name": "고급 섹터 진입 신호 포트폴리오",
        "benchmark_name": benchmark_name if "benchmark_name" in locals() else "KOSPI",
        "start_date": rows[0]["date"] if rows else "",
        "end_date": rows[-1]["date"] if rows else "",
        "rows": rows,
        "trade_log": trade_log[-500:],
        "holdings_timeline": holdings_timeline[-120:],
        "latest_holdings": latest_holdings,
        "rejected_signals": rejected_signals[-120:],
        "summary": {
            "strategy_return_pct": round(final_strategy, 2),
            "benchmark_return_pct": round(final_benchmark, 2),
            "excess_return_pct": round(final_strategy - final_benchmark, 2),
            "mdd_pct": max_drawdown_pct(nav_values),
            "benchmark_mdd_pct": max_drawdown_pct(benchmark_values),
            "cagr_pct": annualized_return_pct(final_strategy, rows[0]["date"], rows[-1]["date"]) if rows else None,
            "avg_exposure_pct": round(float(np.mean([row["exposure_pct"] for row in rows])) if rows else 0.0, 2),
            "max_exposure_pct": round(max((row["exposure_pct"] for row in rows), default=0.0), 2),
            "total_turnover_pct": round(total_turnover * 100.0, 2),
            "total_fee_pct_points": round(total_fee_pct, 2),
            "rebalance_count": len(rows),
            "trade_count": len(trade_log),
        },
        "params": {
            "min_score": min_score,
            "top_sectors": top_sectors,
            "stocks_per_sector": stocks_per_sector,
            "max_leverage": max_leverage,
            "benchmark": benchmark_key,
            "weight_method": weight_method,
            "beta_window": beta_window,
            "min_breadth": min_breadth,
            "max_disparity": max_disparity,
            "trading_rank_limit": trading_rank_limit,
            "stock_selection": stock_selection,
            "fee_rate_pct": round(STRATEGY_TRADE_FEE_RATE * 100.0, 2),
        },
        "available_benchmarks": [
            {"key": "leader_avg", "name": "시총 2000억 이상 주도주 평균"},
            {"key": "portfolio_dashboard", "name": "포트폴리오 수익 페이지"},
            {"key": "KS11", "name": "KOSPI"},
            {"key": "KQ11", "name": "KOSDAQ"},
            {"key": "KRX_BLEND", "name": "KOSPI/KOSDAQ 50:50"},
            {"key": "IXIC", "name": "NASDAQ"},
            {"key": "US500", "name": "S&P 500"},
        ],
        "execution": {
            "kis": kis_status_payload(check_token=False),
            "locked": True,
            "message": "모의투자 주문 실행은 목표 포트폴리오 확인 후 별도 승인 단계에서만 열도록 잠금 처리했습니다.",
        },
        "description": "양 지수가 모두 20일선 아래면 주식 노출을 0%로 만들고, 섹터 진입 신호가 통과한 섹터만 베타 가중으로 편입하는 포트폴리오 백테스트입니다.",
    }


def build_sector_entry_signals(
    start: str | None = None,
    end: str | None = None,
    lookback_days: int = 80,
    min_score: float = 50.0,
    trading_rank_limit: int = 20,
    min_avg_score: float = 55.0,
    min_strong_count: int = 2,
    min_stock_count: int = 1,
    beta_window: int = 120,
) -> dict[str, Any]:
    summaries = screening_backtest_source_summaries()
    if not summaries:
        return {
            "rows": [],
            "latest_active": [],
            "summary": {"signal_count": 0, "date_count": 0},
            "message": "오늘의 주도주 캐시가 없습니다.",
        }

    available_dates = [str(item.get("file_date") or "") for item in summaries if item.get("file_date")]
    fallback_end = available_dates[-1]
    fallback_start = available_dates[max(0, len(available_dates) - max(1, int(lookback_days or 80)))]
    start_date = safe_strategy_date(start, datetime.strptime(fallback_start, "%Y-%m-%d").date()).isoformat()
    end_date = safe_strategy_date(end, datetime.strptime(fallback_end, "%Y-%m-%d").date()).isoformat()
    filtered = [item for item in summaries if start_date <= str(item.get("file_date") or "") <= end_date]
    if not filtered:
        filtered = summaries[-max(1, min(int(lookback_days or 80), len(summaries))) :]

    min_score = max(0.0, min(float(min_score or 50.0), 100.0))
    trading_rank_limit = max(1, min(int(trading_rank_limit or 20), 50))
    min_avg_score = max(0.0, min(float(min_avg_score or 55.0), 100.0))
    min_strong_count = max(0, min(int(min_strong_count or 0), 50))
    min_stock_count = max(1, min(int(min_stock_count or 1), 50))
    beta_window = 63 if int(beta_window or 120) <= 100 else 252

    sector_db = load_sector_db()
    previous_active: set[str] = set()
    rows: list[dict[str, Any]] = []
    latest_active: list[dict[str, Any]] = []

    for index_no, summary in enumerate(filtered):
        current_date = str(summary.get("file_date") or "")
        window = filtered[max(0, index_no - 4) : index_no + 1]
        trading_totals: dict[str, float] = {}
        for window_summary in window:
            for signal in build_sector_rotation_signals(window_summary, min_score=min_score, sector_db=sector_db):
                sector = str(signal.get("sector") or "").strip()
                if sector:
                    trading_totals[sector] = trading_totals.get(sector, 0.0) + float(signal.get("total_trading_value_100m") or 0.0)

        window_count = max(1, len(window))
        trading_rank_rows = sorted(
            [{"sector": sector, "avg_5d_trading_value_100m": value / window_count} for sector, value in trading_totals.items()],
            key=lambda item: item["avg_5d_trading_value_100m"],
            reverse=True,
        )
        rank_map = {
            item["sector"]: {"rank": rank + 1, "avg_5d_trading_value_100m": item["avg_5d_trading_value_100m"]}
            for rank, item in enumerate(trading_rank_rows)
        }

        active_payload: list[dict[str, Any]] = []
        for signal in build_sector_rotation_signals(summary, min_score=min_score, sector_db=sector_db):
            sector = str(signal.get("sector") or "").strip()
            rank_info = rank_map.get(sector, {})
            trading_rank = int(rank_info.get("rank") or 9999)
            if trading_rank > trading_rank_limit:
                continue
            if float(signal.get("avg_score") or 0.0) < min_avg_score:
                continue
            if int(signal.get("strong_count") or 0) < min_strong_count:
                continue
            if int(signal.get("stock_count") or 0) < min_stock_count:
                continue

            entry_score = (
                float(signal.get("strength_score") or 0.0)
                + max(0.0, (trading_rank_limit + 1 - trading_rank)) * 2.0
                + min(float(signal.get("turnover_ratio_pct") or 0.0), 10.0) * 1.2
            )
            signal_payload = {
                    "date": current_date,
                    "sector": sector,
                    "signal": "진입",
                    "signal_level": "강한 진입" if trading_rank <= 5 and float(signal.get("avg_score") or 0) >= 70 else "진입",
                    "trading_rank": trading_rank,
                    "avg_5d_trading_value_100m": round(float(rank_info.get("avg_5d_trading_value_100m") or 0.0), 1),
                    "avg_score": signal.get("avg_score"),
                    "max_score": signal.get("max_score"),
                    "strong_count": signal.get("strong_count"),
                    "stock_count": signal.get("stock_count"),
                    "turnover_ratio_pct": signal.get("turnover_ratio_pct"),
                    "strength_score": signal.get("strength_score"),
                    "entry_score": round(entry_score, 2),
                    "reason": f"5일 거래대금 {trading_rank}위 · 평균 종합점수 {signal.get('avg_score')} · 70점 이상 {signal.get('strong_count')}개",
                }
            signal_payload["entry_phase"] = "유지"
            active_payload.append(enrich_sector_entry_signal_leaders(signal_payload, signal.get("stocks", []), beta_window=beta_window))

        active_payload = sorted(active_payload, key=lambda item: item["entry_score"], reverse=True)
        active_sectors = {item["sector"] for item in active_payload}
        for item in active_payload:
            if item["sector"] not in previous_active:
                item["entry_phase"] = "신규진입"
                rows.append(dict(item, signal="신규 진입"))
            else:
                item["entry_phase"] = "유지"
        for exited in sorted(previous_active - active_sectors):
            rows.append(
                {
                    "date": current_date,
                    "sector": exited,
                    "signal": "이탈",
                    "signal_level": "이탈",
                    "entry_phase": "이탈",
                    "reason": f"종합점수 대역 이탈({min_avg_score:.0f} 미만) 또는 필터 미충족",
                }
            )
        previous_active = active_sectors
        latest_active = active_payload

    latest_top = latest_active[:6]
    weight_sum = sum(max(0.01, float(item.get("entry_score") or 0.0)) for item in latest_top)
    for item in latest_top:
        item["suggested_weight_pct"] = round(max(0.01, float(item.get("entry_score") or 0.0)) / weight_sum * 100.0, 1) if weight_sum else None

    rows = sorted(rows, key=lambda item: (item.get("date") or "", item.get("entry_score") or 0), reverse=True)
    return {
        "start_date": str(filtered[0].get("file_date") or "") if filtered else "",
        "end_date": str(filtered[-1].get("file_date") or "") if filtered else "",
        "params": {
            "min_score": min_score,
            "trading_rank_limit": trading_rank_limit,
            "min_avg_score": min_avg_score,
            "min_strong_count": min_strong_count,
            "min_stock_count": min_stock_count,
            "beta_window": beta_window,
            "beta_window_label": "3개월" if beta_window <= 100 else "1년",
            "lookback_window": "최근 5거래일 섹터 거래대금 평균",
        },
        "summary": {
            "signal_count": len(rows),
            "date_count": len(filtered),
            "latest_active_count": len(latest_active),
            "latest_date": str(filtered[-1].get("file_date") or "") if filtered else "",
        },
        "rows": rows,
        "latest_active": latest_top,
        "description": "최근 5거래일 섹터 거래대금 평균 순위와 오늘의 주도주 점수로, 꼬리 종목 대신 주도 섹터만 신규 진입 후보로 거르는 신호입니다.",
    }


def compute_sector_entry_technical_filters(source_stocks: list[dict[str, Any]]) -> dict[str, Any]:
    valid_count = 0
    above_ma20_count = 0
    disparities: list[float] = []
    for stock in source_stocks[:20]:
        stock_code = normalize_stock_code_value(stock.get("stock_code"))
        if not stock_code:
            continue
        try:
            price_frame = fetch_price_frame(stock_code)
            closes = pd.to_numeric(price_frame.get("Close"), errors="coerce").dropna() if not price_frame.empty else pd.Series(dtype=float)
            if len(closes) < 25:
                continue
            ma20 = closes.rolling(20).mean()
            latest_close = float(closes.iloc[-1])
            latest_ma20 = float(ma20.iloc[-1])
            if not np.isfinite(latest_close) or not np.isfinite(latest_ma20) or latest_ma20 <= 0:
                continue
            valid_count += 1
            if latest_close > latest_ma20:
                above_ma20_count += 1
            disparities.append(latest_close / latest_ma20 * 100.0)
        except Exception:
            continue
    breadth_ratio = above_ma20_count / valid_count * 100.0 if valid_count else None
    avg_disparity = float(np.mean(disparities)) if disparities else None
    leverage_status = "검증불가"
    leverage_label = "기술 필터 데이터 부족"
    if breadth_ratio is not None and avg_disparity is not None:
        if breadth_ratio >= 60 and avg_disparity <= 105:
            leverage_status = "full_leverage"
            leverage_label = "Breadth 60% 이상 · 20일선 근처"
        elif breadth_ratio >= 60 and avg_disparity <= 110:
            leverage_status = "no_leverage"
            leverage_label = "진입 가능하지만 이격도 높음"
        elif breadth_ratio < 60:
            leverage_status = "weak_breadth"
            leverage_label = "섹터 확산 부족"
        else:
            leverage_status = "overheated"
            leverage_label = "20일선 이격 과열"
    return {
        "breadth_ratio": round(breadth_ratio, 1) if breadth_ratio is not None else None,
        "avg_disparity": round(avg_disparity, 2) if avg_disparity is not None else None,
        "technical_stock_count": valid_count,
        "above_ma20_count": above_ma20_count,
        "leverage_status": leverage_status,
        "leverage_label": leverage_label,
        "entry_pass": leverage_status in {"full_leverage", "no_leverage"},
    }


def analyze_sector_entry_stock(stock: dict[str, Any], beta_window: int = 120) -> dict[str, Any]:
    stock_code = normalize_stock_code_value(stock.get("stock_code"))
    stock_name = str(stock.get("resolved_name") or stock.get("stock_name") or stock_code or "").strip()
    listing = find_listing_row_by_code(stock_code) if stock_code else None
    market = str((listing or {}).get("market") or stock.get("market") or "")
    beta: float | None = None
    trend_status = "normal"
    trend_label = ""
    try:
        if stock_code:
            price_frame = fetch_price_frame(stock_code)
            beta = compute_stock_beta(price_frame, market, window=beta_window)
            closes = pd.to_numeric(price_frame.get("Close"), errors="coerce").dropna() if not price_frame.empty else pd.Series(dtype=float)
            if len(closes) >= 205:
                ma20 = closes.rolling(20).mean()
                ma60 = closes.rolling(60).mean()
                ma200 = closes.rolling(200).mean()
                latest_close = float(closes.iloc[-1])
                latest_ma20 = float(ma20.iloc[-1])
                latest_ma60 = float(ma60.iloc[-1])
                latest_ma200 = float(ma200.iloc[-1])
                prev_ma20 = float(ma20.iloc[-2])
                prev_ma60 = float(ma60.iloc[-2])
                golden_cross = np.isfinite(prev_ma20) and np.isfinite(prev_ma60) and np.isfinite(latest_ma20) and np.isfinite(latest_ma60) and prev_ma20 <= prev_ma60 and latest_ma20 > latest_ma60
                aligned = (
                    np.isfinite(latest_close)
                    and np.isfinite(latest_ma20)
                    and np.isfinite(latest_ma60)
                    and np.isfinite(latest_ma200)
                    and latest_close > latest_ma20 > latest_ma60 > latest_ma200
                )
                if golden_cross:
                    trend_status = "golden_cross"
                    trend_label = "골든크로스"
                elif aligned:
                    trend_status = "aligned"
                    trend_label = "정배열"
    except Exception:
        pass
    return {
        "stock_code": stock_code,
        "stock_name": stock_name,
        "score": stock.get("score"),
        "beta": beta,
        "beta_window": beta_window,
        "beta_window_label": "3개월" if int(beta_window or 120) <= 100 else "1년",
        "trend_status": trend_status,
        "trend_label": trend_label,
    }


def enrich_sector_entry_signal_leaders(signal_payload: dict[str, Any], source_stocks: list[dict[str, Any]], beta_window: int = 120) -> dict[str, Any]:
    signal_payload.update(compute_sector_entry_technical_filters(source_stocks))
    leader_stocks = [analyze_sector_entry_stock(stock, beta_window=beta_window) for stock in source_stocks[:8]]
    leader_stocks = sorted(
        leader_stocks,
        key=lambda item: (
            to_float(item.get("beta")) if to_float(item.get("beta")) is not None else -9999,
            to_float(item.get("score")) if to_float(item.get("score")) is not None else -9999,
        ),
        reverse=True,
    )[:5]
    signal_payload["leader_stocks"] = leader_stocks
    signal_payload["leaders"] = [item.get("stock_name") for item in leader_stocks if item.get("stock_name")]
    return signal_payload


SIGNAL_RADAR_DEFINITIONS = {
    "smart_money_combo": {
        "label": "외인+기관 동반 순매수",
        "short_label": "쌍끌이",
        "description": "외국인과 기관이 같은 날 동시에 순매수한 종목입니다.",
    },
    "retail_overheat": {
        "label": "개인 순매수 과열",
        "short_label": "개인 과열",
        "description": "개인이 강하게 받고 외국인+기관은 매도 우위인 종목입니다.",
    },
    "flow_reversal": {
        "label": "어제 매도→오늘 매수 반전",
        "short_label": "수급 반전",
        "description": "전일 스마트머니 순매도에서 당일 순매수로 방향이 바뀐 종목입니다.",
    },
    "flow_5d_streak": {
        "label": "5일 연속 외인/기관 매수",
        "short_label": "5일 연속",
        "description": "최근 5거래일 동안 외국인 또는 기관 중 한 주체가 연속 순매수한 종목입니다.",
    },
    "high_52w": {
        "label": "52주 신고가",
        "short_label": "신고가",
        "description": "52주 신고가 또는 신고가권에 진입한 추세 강세 종목입니다.",
    },
    "golden_cross": {
        "label": "골든크로스",
        "short_label": "골든",
        "description": "5일선이 20일선을 상향 돌파한 종목입니다.",
    },
    "ma20_reclaim": {
        "label": "20일선 재돌파",
        "short_label": "20일선",
        "description": "주가가 20일선을 다시 상향 돌파한 종목입니다.",
    },
}


def signal_radar_cache_key(
    latest_date: str,
    lookback_days: int,
    max_stocks: int,
    min_score: float,
    max_history_events: int,
) -> str:
    return "|".join([
        latest_date,
        str(int(lookback_days)),
        str(int(max_stocks)),
        f"{float(min_score):.2f}",
        str(int(max_history_events)),
        "v6",
    ])


def load_signal_radar_cache() -> dict[str, Any]:
    if not SIGNAL_RADAR_CACHE_PATH.exists():
        return {}
    try:
        return json.loads(SIGNAL_RADAR_CACHE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_signal_radar_cache(cache: dict[str, Any]) -> dict[str, Any]:
    SIGNAL_RADAR_CACHE_PATH.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")
    return cache


def latest_portfolio_stock_codes() -> set[str]:
    try:
        blocks = parse_portfolio_blocks()
    except Exception:
        return set()
    if not blocks:
        return set()
    latest = blocks[-1]
    return {
        normalize_stock_code_value(item.get("stock_code"))
        for item in latest.get("holdings", [])
        if normalize_stock_code_value(item.get("stock_code"))
    }


def sector_db_stock_codes() -> set[str]:
    db = load_sector_db()
    stock_map = db.get("stock_map", {}) if isinstance(db, dict) else {}
    codes: set[str] = set()
    for item in stock_map.values():
        code = normalize_stock_code_value((item or {}).get("stock_code"))
        if code:
            codes.add(code)
    return codes


def price_close_after_trading_days(frame: pd.DataFrame, signal_date: str, horizon: int) -> tuple[str, float | None, list[float]]:
    if frame is None or frame.empty or "Date" not in frame.columns or "Close" not in frame.columns:
        return "", None, []
    try:
        target = pd.to_datetime(signal_date).date()
    except Exception:
        return "", None, []
    working = frame[["Date", "Close"]].copy()
    working["Date"] = pd.to_datetime(working["Date"], errors="coerce")
    working["Close"] = pd.to_numeric(working["Close"], errors="coerce")
    working = working.dropna(subset=["Date", "Close"]).sort_values("Date").reset_index(drop=True)
    candidates = working[working["Date"].dt.date >= target]
    if candidates.empty:
        return "", None, []
    start_index = int(candidates.index[0])
    end_index = min(len(working) - 1, start_index + max(0, int(horizon)))
    close_value = to_float(working.iloc[end_index].get("Close"))
    close_date = pd.Timestamp(working.iloc[end_index].get("Date")).strftime("%Y-%m-%d")
    path = [float(value) for value in working.iloc[start_index : end_index + 1]["Close"].tolist() if to_float(value) is not None]
    return close_date, close_value, path


def signal_forward_performance(stock_code: str, signal_date: str, horizons: tuple[int, ...] = (5, 20, 60)) -> dict[str, Any]:
    code = normalize_stock_code_value(stock_code)
    if not code:
        return {}
    try:
        frame = fetch_price_frame(code)
    except Exception:
        return {}
    base_date, base_close, _path = price_close_after_trading_days(frame, signal_date, 0)
    if base_close in (None, 0):
        return {}
    result: dict[str, Any] = {"base_date": base_date, "base_close": base_close}
    for horizon in horizons:
        end_date, end_close, path = price_close_after_trading_days(frame, signal_date, horizon)
        if end_close is None:
            continue
        returns = (float(end_close) / float(base_close) - 1.0) * 100.0
        if path:
            nav_path = [value / float(base_close) * 100.0 for value in path if base_close]
            mdd = max_drawdown_pct(nav_path)
        else:
            mdd = None
        result[f"return_{horizon}d_pct"] = round(returns, 2)
        result[f"mdd_{horizon}d_pct"] = mdd
        result[f"end_{horizon}d_date"] = end_date
    return result


def stock_technical_signal_flags(stock_code: str, row: dict[str, Any] | None = None, as_of: str | None = None) -> dict[str, Any]:
    code = normalize_stock_code_value(stock_code)
    row = row or {}
    if not code:
        return {"signals": [], "close": None, "ma20": None, "disparity": None}
    try:
        frame = fetch_price_frame(code)
    except Exception:
        frame = pd.DataFrame()
    if frame.empty or "Date" not in frame.columns or "Close" not in frame.columns:
        return {"signals": [], "close": None, "ma20": None, "disparity": None}
    working = frame.copy()
    working["Date"] = pd.to_datetime(working["Date"], errors="coerce")
    working["Close"] = pd.to_numeric(working["Close"], errors="coerce")
    working = working.dropna(subset=["Date", "Close"]).sort_values("Date").reset_index(drop=True)
    if as_of:
        try:
            target = pd.to_datetime(as_of).date()
            working = working[working["Date"].dt.date <= target].copy()
        except Exception:
            pass
    if len(working) < 25:
        return {"signals": [], "close": None, "ma20": None, "disparity": None}
    close = float(working["Close"].iloc[-1])
    ma5 = working["Close"].rolling(5).mean()
    ma20 = working["Close"].rolling(20).mean()
    latest_ma20 = to_float(ma20.iloc[-1])
    prev_close = to_float(working["Close"].iloc[-2]) if len(working) >= 2 else None
    prev_ma20 = to_float(ma20.iloc[-2]) if len(working) >= 2 else None
    latest_ma5 = to_float(ma5.iloc[-1])
    prev_ma5 = to_float(ma5.iloc[-2]) if len(working) >= 2 else None
    high_window = working["Close"].tail(252)
    high_52 = float(high_window.max()) if not high_window.empty else None
    raw_high = str(row.get("is_52w_high") or "").strip().upper()
    signals: list[str] = []
    if raw_high in {"O", "Y", "TRUE", "1"} or (high_52 and close >= high_52 * 0.99):
        signals.append("high_52w")
    if prev_ma5 is not None and prev_ma20 is not None and latest_ma5 is not None and latest_ma20 is not None:
        if prev_ma5 <= prev_ma20 and latest_ma5 > latest_ma20:
            signals.append("golden_cross")
    if prev_close is not None and prev_ma20 is not None and latest_ma20 is not None:
        if prev_close <= prev_ma20 and close > latest_ma20:
            signals.append("ma20_reclaim")
    disparity = (close / latest_ma20 * 100.0) if latest_ma20 else None
    return {
        "signals": signals,
        "close": round(close, 2),
        "ma20": round(latest_ma20, 2) if latest_ma20 else None,
        "disparity": round(disparity, 2) if disparity is not None else None,
    }


def flow_signal_flags_from_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    ordered = sorted([row for row in rows if row.get("date")], key=lambda item: str(item.get("date") or ""))
    if not ordered:
        return {"signals": [], "latest": {}, "source_rows": 0}
    latest = ordered[-1]
    previous = ordered[-2] if len(ordered) >= 2 else {}
    foreigner = int(latest.get("foreigner") or 0)
    institution = int(latest.get("institution") or 0)
    individual = int(latest.get("individual") or 0)
    smart = foreigner + institution
    prev_smart = int(previous.get("foreigner") or 0) + int(previous.get("institution") or 0)
    individual_abs_values = [abs(int(row.get("individual") or 0)) for row in ordered[-6:] if row.get("individual") is not None]
    individual_threshold = float(np.median(individual_abs_values)) if individual_abs_values else 0.0
    signals: list[str] = []
    if foreigner > 0 and institution > 0:
        signals.append("smart_money_combo")
    if individual > 0 and smart < 0 and (individual_threshold <= 0 or individual >= individual_threshold):
        signals.append("retail_overheat")
    if prev_smart <= 0 and smart > 0:
        signals.append("flow_reversal")
    recent5 = ordered[-5:]
    if len(recent5) >= 5:
        foreigner_streak = all(int(row.get("foreigner") or 0) > 0 for row in recent5)
        institution_streak = all(int(row.get("institution") or 0) > 0 for row in recent5)
        if foreigner_streak or institution_streak:
            signals.append("flow_5d_streak")
    return {
        "signals": signals,
        "latest": {
            "date": latest.get("date"),
            "individual": individual,
            "foreigner": foreigner,
            "institution": institution,
            "smart_money": smart,
            "previous_smart_money": prev_smart,
        },
        "source_rows": len(ordered),
    }


def stock_flow_signal_payload(stock: dict[str, Any], days: int = 12) -> dict[str, Any]:
    code = normalize_stock_code_value(stock.get("stock_code"))
    name = str(stock.get("resolved_name") or stock.get("stock_name") or code).strip()
    if not code:
        return {"signals": [], "flow_error": "종목코드 없음"}
    try:
        # The radar needs fast breadth, not full investor-detail history. Query Naver
        # directly here so one slow KIS request cannot stall the whole page.
        rows = fetch_naver_investor_flow_rows(code, pages=1)
        if days and rows:
            cutoff = (datetime.now().date() - timedelta(days=max(1, int(days)))).isoformat()
            rows = [row for row in rows if str(row.get("date") or "") >= cutoff] or rows[:12]
        flags = flow_signal_flags_from_rows(rows)
        return {
            **flags,
            "flow_source": "Naver Finance 외국인/기관 매매현황",
            "flow_note": "레이더 응답 속도를 위해 네이버 공개 수급표를 직접 사용합니다. 개인은 외국인+기관계 잔차 기준 추정치입니다.",
        }
    except Exception as exc:
        return {"signals": [], "flow_error": str(exc)}


def summarize_signal_performance(events: list[dict[str, Any]]) -> dict[str, Any]:
    by_signal: dict[str, list[dict[str, Any]]] = {}
    for event in events:
        for signal_key in event.get("signals") or []:
            by_signal.setdefault(signal_key, []).append(event)
    rows: list[dict[str, Any]] = []
    for signal_key in SIGNAL_RADAR_DEFINITIONS.keys():
        signal_events = by_signal.get(signal_key, [])
        definition = SIGNAL_RADAR_DEFINITIONS.get(signal_key, {})
        row: dict[str, Any] = {
            "signal_key": signal_key,
            "label": definition.get("label") or signal_key,
            "short_label": definition.get("short_label") or signal_key,
            "description": definition.get("description") or "",
            "event_count": len(signal_events),
        }
        for horizon in (5, 20, 60):
            values = [
                to_float(event.get(f"return_{horizon}d_pct"))
                for event in signal_events
                if to_float(event.get(f"return_{horizon}d_pct")) is not None
            ]
            mdds = [
                to_float(event.get(f"mdd_{horizon}d_pct"))
                for event in signal_events
                if to_float(event.get(f"mdd_{horizon}d_pct")) is not None
            ]
            if values:
                row[f"avg_return_{horizon}d_pct"] = round(float(np.mean(values)), 2)
                row[f"win_rate_{horizon}d_pct"] = round(sum(1 for value in values if value > 0) / len(values) * 100.0, 1)
            else:
                row[f"avg_return_{horizon}d_pct"] = None
                row[f"win_rate_{horizon}d_pct"] = None
            row[f"mdd_{horizon}d_pct"] = round(min(mdds), 2) if mdds else None
        rows.append(row)
    rows.sort(key=lambda item: (item.get("event_count") or 0, item.get("avg_return_20d_pct") or -9999), reverse=True)
    return {
        "rows": rows,
        "event_count": len(events),
    }


def build_historical_signal_events(
    summaries: list[dict[str, Any]],
    start_date: str,
    end_date: str,
    min_score: float,
    max_events: int = 900,
) -> list[dict[str, Any]]:
    events: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for summary in summaries:
        signal_date = str(summary.get("file_date") or "")
        if not signal_date or signal_date < start_date or signal_date > end_date:
            continue
        for row in summary.get("qualified_stocks", []):
            if len(events) >= max_events:
                break
            if not isinstance(row, dict):
                continue
            score = to_float(row.get("score"))
            if score is None or score < min_score:
                continue
            code = normalize_stock_code_value(row.get("stock_code"))
            if not code:
                continue
            technical = stock_technical_signal_flags(code, row, as_of=signal_date)
            signals = technical.get("signals") or []
            if not signals:
                continue
            perf = signal_forward_performance(code, signal_date)
            if not perf:
                continue
            sector = resolve_sector_for_stock(code, row.get("stock_name"), load_sector_db()) or str(row.get("manual_sector") or row.get("theme") or "")
            for signal_key in signals:
                key = (signal_date, code, signal_key)
                if key in seen:
                    continue
                seen.add(key)
            events.append(
                {
                    "date": signal_date,
                    "stock_code": code,
                    "stock_name": row.get("resolved_name") or row.get("stock_name") or code,
                    "sector": sector,
                    "score": score,
                    "signals": signals,
                    **perf,
                }
            )
        if len(events) >= max_events:
            break
    return events


def build_signal_radar(
    lookback_days: int = 120,
    max_stocks: int = 45,
    min_score: float = 50.0,
    max_history_events: int = 220,
) -> dict[str, Any]:
    summaries = screening_backtest_source_summaries()
    if not summaries:
        return {
            "summary": {"signal_count": 0, "stock_count": 0},
            "signals": [],
            "performance": {"rows": [], "event_count": 0},
            "message": "오늘의 주도주 캐시가 없습니다.",
        }
    available_dates = [str(item.get("file_date") or "") for item in summaries if item.get("file_date")]
    latest_date = available_dates[-1]
    lookback_days = max(20, min(int(lookback_days or 120), 260))
    max_stocks = max(10, min(int(max_stocks or 45), 80))
    min_score = max(0.0, min(float(min_score or 50.0), 100.0))
    max_history_events = max(60, min(int(max_history_events or 220), 900))
    cache_key = signal_radar_cache_key(latest_date, lookback_days, max_stocks, min_score, max_history_events)
    cache = load_signal_radar_cache()
    cached = cache.get(cache_key)
    if isinstance(cached, dict) and time.time() - float(cached.get("_cached_at") or 0) < 30 * 60:
        return {key: value for key, value in cached.items() if key != "_cached_at"}

    latest_summary = summaries[-1]
    sector_db = load_sector_db()
    portfolio_codes = latest_portfolio_stock_codes()
    watch_codes = sector_db_stock_codes()
    candidates = []
    for row in latest_summary.get("qualified_stocks", []):
        if not isinstance(row, dict):
            continue
        score = to_float(row.get("score"))
        if score is None or score < min_score:
            continue
        code = normalize_stock_code_value(row.get("stock_code"))
        if not code:
            continue
        enriched = dict(row)
        enriched["stock_code"] = code
        enriched["sector"] = resolve_sector_for_stock(code, row.get("stock_name"), sector_db) or str(row.get("manual_sector") or row.get("theme") or "")
        candidates.append(enriched)
    candidates.sort(
        key=lambda item: (
            to_float(item.get("score")) or 0.0,
            to_float(item.get("trading_value_100m")) or 0.0,
            to_float(item.get("market_cap_100m")) or 0.0,
        ),
        reverse=True,
    )
    candidates = candidates[:max_stocks]

    flow_payload_by_code: dict[str, dict[str, Any]] = {}
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {executor.submit(stock_flow_signal_payload, stock, 12): stock for stock in candidates}
        for future in as_completed(futures):
            stock = futures[future]
            code = normalize_stock_code_value(stock.get("stock_code"))
            try:
                flow_payload_by_code[code] = future.result()
            except Exception as exc:
                flow_payload_by_code[code] = {"signals": [], "flow_error": str(exc)}

    signal_rows: list[dict[str, Any]] = []
    for stock in candidates:
        code = normalize_stock_code_value(stock.get("stock_code"))
        technical = stock_technical_signal_flags(code, stock, as_of=latest_date)
        flow = flow_payload_by_code.get(code, {"signals": []})
        signal_keys = sorted(set((technical.get("signals") or []) + (flow.get("signals") or [])))
        if not signal_keys:
            continue
        labels = [SIGNAL_RADAR_DEFINITIONS.get(key, {}).get("short_label") or key for key in signal_keys]
        signal_rows.append(
            {
                "date": latest_date,
                "stock_code": code,
                "stock_name": stock.get("resolved_name") or stock.get("stock_name") or code,
                "sector": stock.get("sector") or "",
                "score": stock.get("score"),
                "change_pct": stock.get("change_pct"),
                "market_cap_100m": stock.get("market_cap_100m"),
                "trading_value_100m": stock.get("trading_value_100m"),
                "execution_strength": stock.get("execution_strength"),
                "signals": signal_keys,
                "signal_labels": labels,
                "signal_count": len(signal_keys),
                "is_portfolio": code in portfolio_codes,
                "is_watch": code in watch_codes,
                "technical": {key: technical.get(key) for key in ["close", "ma20", "disparity"]},
                "flow": flow.get("latest") or {},
                "flow_source": flow.get("flow_source") or "",
                "flow_error": flow.get("flow_error") or "",
            }
        )
    signal_rows.sort(
        key=lambda item: (
            1 if item.get("is_portfolio") else 0,
            1 if item.get("is_watch") else 0,
            item.get("signal_count") or 0,
            to_float(item.get("score")) or 0.0,
        ),
        reverse=True,
    )

    start_date = (datetime.strptime(latest_date, "%Y-%m-%d") - timedelta(days=lookback_days)).strftime("%Y-%m-%d")
    historical_events = build_historical_signal_events(
        summaries=summaries,
        start_date=start_date,
        end_date=latest_date,
        min_score=min_score,
        max_events=max_history_events,
    )
    performance = summarize_signal_performance(historical_events)
    signal_counts: dict[str, int] = {}
    for row in signal_rows:
        for signal_key in row.get("signals") or []:
            signal_counts[signal_key] = signal_counts.get(signal_key, 0) + 1
    payload = {
        "date": latest_date,
        "start_date": start_date,
        "summary": {
            "stock_count": len(signal_rows),
            "candidate_count": len(candidates),
            "signal_count": sum(len(row.get("signals") or []) for row in signal_rows),
            "portfolio_hit_count": sum(1 for row in signal_rows if row.get("is_portfolio")),
            "watch_hit_count": sum(1 for row in signal_rows if row.get("is_watch")),
        },
        "definitions": [
            {"key": key, **value, "today_count": signal_counts.get(key, 0)}
            for key, value in SIGNAL_RADAR_DEFINITIONS.items()
        ],
        "signals": signal_rows[:120],
        "performance": performance,
        "params": {
            "lookback_days": lookback_days,
            "max_stocks": max_stocks,
            "min_score": min_score,
            "max_history_events": max_history_events,
            "source": "오늘의 주도주 DB + FinanceDataReader + KIS/Naver 투자자 매매동향",
        },
        "description": "오늘의 주도주 후보를 기준으로 수급/차트 신호를 포착하고, 과거 동일 신호의 5/20/60거래일 성과를 함께 보여줍니다.",
    }
    cache[cache_key] = {**payload, "_cached_at": time.time()}
    cache = dict(list(cache.items())[-12:])
    save_signal_radar_cache(cache)
    return payload


def open_tradingview_desktop(stock_code: str | None, stock_name: str | None) -> dict[str, Any]:
    code, symbol, web_url = build_tradingview_symbol(stock_code, stock_name)
    protocol_url = f"tradingview://chart/?symbol={quote(symbol, safe='')}"
    executable = find_tradingview_executable()

    try:
        os.startfile(web_url)  # type: ignore[attr-defined]
        return {
            "ok": True,
            "method": "web-link-handler",
            "stock_code": code,
            "symbol": symbol,
            "url": web_url,
            "message": "TradingView 차트 링크를 열었습니다. Windows의 TradingView 앱 링크 처리가 켜져 있으면 앱에서 해당 종목 차트가 열립니다.",
        }
    except OSError:
        pass

    try:
        os.startfile(protocol_url)  # type: ignore[attr-defined]
        return {
            "ok": True,
            "method": "protocol",
            "stock_code": code,
            "symbol": symbol,
            "url": protocol_url,
            "message": "TradingView 앱 프로토콜로 차트를 열었습니다.",
        }
    except OSError:
        pass

    if executable:
        try:
            subprocess.Popen(
                [str(executable), web_url],
                cwd=str(executable.parent),
                stdout=subprocess.DEVNULL,
                stderr=subprocess.DEVNULL,
                close_fds=True,
            )
            return {
                "ok": True,
                "method": "exe",
                "stock_code": code,
                "symbol": symbol,
                "url": web_url,
                "message": "TradingView 앱을 실행했습니다. 앱이 종목 링크 인자를 무시하면 앱 설정의 링크 처리를 확인해 주세요.",
            }
        except OSError:
            pass

    app_id = "31178TradingViewInc.TradingView_q4jpyh43s5mv6!TradingView.Desktop"
    subprocess.Popen(
        ["explorer.exe", f"shell:AppsFolder\\{app_id}"],
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        close_fds=True,
    )
    return {
        "ok": True,
        "method": "appsfolder",
        "stock_code": code,
        "symbol": symbol,
        "url": web_url,
        "message": "TradingView 앱을 열었습니다. 특정 종목이 바로 열리지 않으면 앱의 링크 처리 제한 때문입니다.",
    }


@lru_cache(maxsize=1024)
def fetch_price_frame(code: str) -> pd.DataFrame:
    start = (datetime.today() - pd.Timedelta(days=430)).strftime("%Y-%m-%d")
    frame = fdr.DataReader(code, start)
    if frame is None or frame.empty:
        return pd.DataFrame()
    frame = frame.reset_index()
    if "Date" not in frame.columns:
        frame = frame.rename(columns={frame.columns[0]: "Date"})
    frame["Date"] = pd.to_datetime(frame["Date"])
    return frame.sort_values("Date").reset_index(drop=True)


def chart_preview_cache_path(code: str, months: int) -> Path:
    cache_seed = f"{str(code or '').strip().upper()}|{int(months or 3)}"
    return CHART_PREVIEW_CACHE_DIR / f"{hashlib.sha1(cache_seed.encode('utf-8')).hexdigest()[:24]}.json"


def load_chart_preview_disk_cache(code: str, months: int) -> dict[str, Any] | None:
    cache_path = chart_preview_cache_path(code, months)
    if not cache_path.exists():
        return None
    age_seconds = time.time() - cache_path.stat().st_mtime
    if age_seconds > 12 * 60 * 60:
        return None
    try:
        payload = json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception:
        return None
    if not isinstance(payload, dict) or not isinstance(payload.get("rows"), list):
        return None
    return payload


def save_chart_preview_disk_cache(code: str, months: int, payload: dict[str, Any]) -> None:
    CHART_PREVIEW_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path = chart_preview_cache_path(code, months)
    temp_path = cache_path.with_name(f"{cache_path.stem}_{uuid.uuid4().hex[:8]}.tmp")
    temp_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    temp_path.replace(cache_path)


def get_return_between(closes: pd.Series, from_offset: int, to_offset: int) -> float | None:
    if len(closes) <= from_offset or len(closes) <= to_offset:
        return None
    newer = to_float(closes.iloc[-(to_offset + 1)])
    older = to_float(closes.iloc[-(from_offset + 1)])
    if older in (None, 0) or newer is None:
        return None
    return round(((newer / older) - 1.0) * 100.0, 2)


def get_ytd_return_from_frame(price_frame: pd.DataFrame) -> float | None:
    if price_frame.empty or "Date" not in price_frame.columns or "Close" not in price_frame.columns:
        return None
    current_date = pd.to_datetime(price_frame["Date"]).max()
    if pd.isna(current_date):
        return None
    year_start = pd.Timestamp(year=current_date.year, month=1, day=1)
    frame = price_frame[pd.to_datetime(price_frame["Date"]) >= year_start].copy()
    if frame.empty:
        return None
    closes = pd.to_numeric(frame["Close"], errors="coerce").dropna()
    if len(closes) < 2:
        return None
    first = to_float(closes.iloc[0])
    last = to_float(closes.iloc[-1])
    if first in (None, 0) or last is None:
        return None
    return round(((last / first) - 1.0) * 100.0, 2)


def compute_gap_from_ma(closes: pd.Series, window: int) -> float | None:
    if len(closes) < window:
        return None
    current = to_float(closes.iloc[-1])
    moving_average = to_float(closes.rolling(window).mean().iloc[-1])
    if current is None or moving_average in (None, 0):
        return None
    return round(((current / moving_average) - 1.0) * 100.0, 2)


def compute_avg_trading_value_to_marcap(price_frame: pd.DataFrame, market_cap: float | None, window: int = 20) -> float | None:
    if market_cap in (None, 0) or price_frame.empty:
        return None
    if "Close" not in price_frame.columns or "Volume" not in price_frame.columns:
        return None
    values = pd.to_numeric(price_frame["Close"], errors="coerce") * pd.to_numeric(price_frame["Volume"], errors="coerce")
    values = values.dropna().tail(window)
    if values.empty:
        return None
    return round(float(values.mean()) / float(market_cap), 5)


def fetch_naver_investor_frame(code: str) -> pd.DataFrame:
    response = requests.get(
        f"https://finance.naver.com/item/frgn.naver?code={str(code).zfill(6)}&page=1",
        headers={"User-Agent": "Mozilla/5.0"},
        timeout=12,
    )
    response.raise_for_status()
    response.encoding = "cp949"
    tables = pd.read_html(StringIO(response.text))
    target: pd.DataFrame | None = None
    for table in tables:
        columns = [str(column) for column in table.columns]
        if any("외국인" in column for column in columns) and any("순매매" in column for column in columns):
            target = table.copy()
            break
    if target is None or target.empty:
        return pd.DataFrame()

    if isinstance(target.columns, pd.MultiIndex):
        target.columns = [
            "_".join([str(part).strip() for part in column if str(part).strip() and str(part) != "nan"]).strip("_")
            for column in target.columns
        ]
    else:
        target.columns = [str(column).strip() for column in target.columns]

    rename_map: dict[str, str] = {}
    for column in target.columns:
        if column.startswith("날짜"):
            rename_map[column] = "date"
        elif column.startswith("종가"):
            rename_map[column] = "close"
        elif "외국인" in column and "순매매" in column:
            rename_map[column] = "foreigner_net"
    target = target.rename(columns=rename_map)
    if not {"date", "close", "foreigner_net"}.issubset(set(target.columns)):
        return pd.DataFrame()

    target = target.dropna(subset=["date"]).copy()
    target = target[target["date"].astype(str).str.contains(r"\d{4}\.\d{2}\.\d{2}", na=False)].copy()
    if target.empty:
        return pd.DataFrame()

    target["date"] = pd.to_datetime(target["date"].astype(str).str.replace(".", "-", regex=False), errors="coerce")
    target = target.dropna(subset=["date"]).copy()
    target["close"] = target["close"].map(clean_numeric_text)
    target["foreigner_net"] = target["foreigner_net"].map(clean_numeric_text)
    return target.dropna(subset=["close", "foreigner_net"]).sort_values("date").reset_index(drop=True)


def flatten_table_columns(frame: pd.DataFrame) -> pd.DataFrame:
    frame = frame.copy()
    if isinstance(frame.columns, pd.MultiIndex):
        frame.columns = [
            "_".join([str(part).strip() for part in column if str(part).strip() and str(part) != "nan"]).strip("_")
            for column in frame.columns
        ]
    else:
        frame.columns = [str(column).strip() for column in frame.columns]
    return frame


def parse_percent_text(value: Any) -> float | None:
    parsed = clean_numeric_text(value)
    if parsed is None:
        return None
    return float(parsed)


def parse_naver_change_text(value: Any) -> float | None:
    text = str(value or "").strip()
    parsed = clean_numeric_text(text)
    if parsed is None:
        return None
    if "하락" in text or text.startswith("-"):
        return -abs(float(parsed))
    if "상승" in text or text.startswith("+"):
        return abs(float(parsed))
    return float(parsed)


def signed_number(value: Any) -> int | None:
    parsed = clean_numeric_text(value)
    if parsed is None:
        return None
    return int(round(float(parsed)))


def kis_get_json(path: str, tr_id: str, params: dict[str, Any], environment: str | None = None) -> dict[str, Any]:
    kis = get_kis_settings(environment)
    if not kis["app_key"] or not kis["app_secret"]:
        raise ValueError("한국투자증권 API Key/Secret이 설정되어 있지 않습니다.")
    token = get_kis_access_token()
    response = requests.get(
        kis["base_url"].rstrip("/") + path,
        headers={
            "content-type": "application/json; charset=utf-8",
            "authorization": "Bearer " + str(token.get("access_token", "")),
            "appkey": kis["app_key"],
            "appsecret": kis["app_secret"],
            "tr_id": tr_id,
            "custtype": "P",
        },
        params=params,
        timeout=20,
    )
    response.raise_for_status()
    payload = response.json()
    if str(payload.get("rt_cd", "")) != "0":
        raise ValueError(str(payload.get("msg1") or payload.get("msg_cd") or "한국투자증권 API 호출에 실패했습니다."))
    return payload


def kis_investor_value(row: dict[str, Any], key: str) -> int | None:
    value = signed_number(row.get(key))
    return value


def fetch_kis_investor_flow_rows(code: str, days: int = 31) -> list[dict[str, Any]]:
    if is_public_web_mode():
        return []
    end_date = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d")
    payload = kis_get_json(
        "/uapi/domestic-stock/v1/quotations/investor-trade-by-stock-daily",
        "FHPTJ04160001",
        {
            "FID_COND_MRKT_DIV_CODE": "J",
            "FID_INPUT_ISCD": str(code or "").zfill(6),
            "FID_INPUT_DATE_1": end_date,
            "FID_ORG_ADJ_PRC": "",
            "FID_ETC_CLS_CODE": "",
        },
    )
    output = payload.get("output2")
    raw_rows = output if isinstance(output, list) else ([output] if isinstance(output, dict) else [])
    rows: list[dict[str, Any]] = []
    cutoff = (datetime.now(timezone(timedelta(hours=9))).date() - timedelta(days=max(1, int(days or 31)))).strftime("%Y%m%d")
    for item in raw_rows:
        if not isinstance(item, dict):
            continue
        row_date = str(item.get("stck_bsop_date") or "")
        if days and row_date and row_date < cutoff:
            continue
        rows.append(
            {
                "date": f"{row_date[:4]}-{row_date[4:6]}-{row_date[6:8]}" if len(row_date) == 8 else row_date,
                "close": signed_number(item.get("stck_clpr")),
                "change": clean_numeric_text(item.get("prdy_vrss")),
                "change_text": str(item.get("prdy_vrss", "") or "").strip(),
                "change_pct": parse_percent_text(item.get("prdy_ctrt")),
                "volume": signed_number(item.get("acml_vol")),
                "individual": kis_investor_value(item, "prsn_ntby_qty"),
                "foreigner": kis_investor_value(item, "frgn_ntby_qty"),
                "institution": kis_investor_value(item, "orgn_ntby_qty"),
                "financial_investment": kis_investor_value(item, "scrt_ntby_qty"),
                "insurance": kis_investor_value(item, "insu_ntby_qty"),
                "trust": kis_investor_value(item, "ivtr_ntby_qty"),
                "bank": kis_investor_value(item, "bank_ntby_qty"),
                "other_finance": kis_investor_value(item, "mrbn_ntby_qty"),
                "pension": kis_investor_value(item, "fund_ntby_qty"),
                "private_fund": kis_investor_value(item, "pe_fund_ntby_vol"),
                "other_corp": kis_investor_value(item, "etc_corp_ntby_vol"),
                "other_foreigner": kis_investor_value(item, "frgn_nreg_ntby_qty"),
            }
        )
    return sorted(rows, key=lambda row: row.get("date") or "", reverse=True)


def fetch_naver_investor_flow_rows(code: str, pages: int = 2) -> list[dict[str, Any]]:
    code = str(code or "").zfill(6)
    rows: list[dict[str, Any]] = []
    for page in range(1, max(1, pages) + 1):
        response = requests.get(
            f"https://finance.naver.com/item/frgn.naver?code={code}&page={page}",
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=12,
        )
        response.raise_for_status()
        response.encoding = "cp949"
        tables = pd.read_html(StringIO(response.text))
        target: pd.DataFrame | None = None
        for table in tables:
            table = flatten_table_columns(table)
            if any("기관" in column and "순매매" in column for column in table.columns) and any(
                "외국인" in column and "순매매" in column for column in table.columns
            ):
                target = table
                break
        if target is None or target.empty:
            continue

        column_map: dict[str, str] = {}
        for column in target.columns:
            text = str(column)
            if text.startswith("날짜"):
                column_map[column] = "date"
            elif text.startswith("종가"):
                column_map[column] = "close"
            elif text.startswith("전일비"):
                column_map[column] = "change"
            elif text.startswith("등락률"):
                column_map[column] = "change_pct"
            elif text.startswith("거래량"):
                column_map[column] = "volume"
            elif "기관" in text and "순매매" in text:
                column_map[column] = "institution"
            elif "외국인" in text and "순매매" in text:
                column_map[column] = "foreigner"
        target = target.rename(columns=column_map)
        required = {"date", "close", "change_pct", "volume", "institution", "foreigner"}
        if not required.issubset(set(target.columns)):
            continue

        target = target.dropna(subset=["date"]).copy()
        target = target[target["date"].astype(str).str.contains(r"\d{4}\.\d{2}\.\d{2}", na=False)]
        for _, item in target.iterrows():
            institution = signed_number(item.get("institution"))
            foreigner = signed_number(item.get("foreigner"))
            individual = None
            if institution is not None and foreigner is not None:
                individual = int(-(institution + foreigner))
            rows.append(
                {
                    "date": str(item.get("date", "")).replace(".", "-"),
                    "close": signed_number(item.get("close")),
                    "change": parse_naver_change_text(item.get("change")),
                    "change_text": str(item.get("change", "") or "").strip(),
                    "change_pct": parse_percent_text(item.get("change_pct")),
                    "volume": signed_number(item.get("volume")),
                    "individual": individual,
                    "foreigner": foreigner,
                    "institution": institution,
                    "financial_investment": None,
                    "insurance": None,
                    "trust": None,
                    "bank": None,
                    "other_finance": None,
                    "pension": None,
                    "private_fund": None,
                    "other_corp": None,
                    "other_foreigner": None,
                }
            )
    deduped: dict[str, dict[str, Any]] = {}
    for row in rows:
        if row.get("date") and row["date"] not in deduped:
            deduped[row["date"]] = row
    return sorted(deduped.values(), key=lambda row: row.get("date") or "", reverse=True)


def build_stock_investor_flows(code: str | None, stock_name: str | None = None, days: int = 31) -> dict[str, Any]:
    resolved = resolve_stock_payload(code, stock_name)
    if not resolved:
        raise ValueError("수급을 조회할 종목을 찾지 못했습니다.")
    stock_code = str(resolved.get("code") or code or "").zfill(6)
    stock_label = str(resolved.get("name") or stock_name or stock_code)
    source = "한국투자증권 OpenAPI 종목별 투자자매매동향(일별)"
    note = "한국투자증권 OpenAPI 기준 일자별 주체별 순매수/순매도 수량입니다."
    try:
        rows = fetch_kis_investor_flow_rows(stock_code, days=days)
    except Exception:
        rows = []
    if not rows:
        pages = max(1, min(8, math.ceil(max(1, int(days or 31)) / 20) + 1))
        rows = fetch_naver_investor_flow_rows(stock_code, pages=pages)
        source = "Naver Finance 외국인/기관 매매현황"
        note = "한국투자증권 API 조회가 실패해 네이버 공개 표로 대체했습니다. 이 경우 기관 세부 주체별 수급은 제공되지 않고, 개인은 외국인+기관계 잔차 기준 추정치입니다."
    if days and rows:
        cutoff = (datetime.now().date() - timedelta(days=max(1, int(days)))).isoformat()
        filtered = [row for row in rows if str(row.get("date") or "") >= cutoff]
        rows = filtered or rows[:20]
    if not rows:
        raise ValueError(f"{stock_label} 수급 데이터를 찾지 못했습니다.")

    sum_keys = [
        "individual",
        "foreigner",
        "institution",
        "financial_investment",
        "insurance",
        "trust",
        "bank",
        "other_finance",
        "pension",
        "private_fund",
        "other_corp",
        "other_foreigner",
    ]
    totals: dict[str, int | None] = {}
    for key in sum_keys:
        values = [int(row[key]) for row in rows if row.get(key) is not None]
        totals[key] = int(sum(values)) if values else None
    return {
        "stock_code": stock_code,
        "stock_name": stock_label,
        "source": source,
        "unit": "주",
        "from_date": rows[-1].get("date"),
        "to_date": rows[0].get("date"),
        "note": note,
        "columns": [
            {"key": "individual", "label": "개인(추정)"},
            {"key": "foreigner", "label": "외국인"},
            {"key": "institution", "label": "기관계"},
            {"key": "financial_investment", "label": "금융투자"},
            {"key": "insurance", "label": "보험"},
            {"key": "trust", "label": "투신"},
            {"key": "bank", "label": "은행"},
            {"key": "other_finance", "label": "기타금융"},
            {"key": "pension", "label": "연기금등"},
            {"key": "private_fund", "label": "사모펀드"},
            {"key": "other_corp", "label": "기타법인"},
            {"key": "other_foreigner", "label": "기타외"},
        ],
        "totals": totals,
        "rows": rows,
    }


MARKET_INVESTOR_COLUMNS = [
    {"key": "individual", "label": "개인", "amount_key": "prsn_ntby_tr_pbmn"},
    {"key": "foreigner", "label": "외국인", "amount_key": "frgn_ntby_tr_pbmn"},
    {"key": "institution", "label": "기관계", "amount_key": "orgn_ntby_tr_pbmn"},
    {"key": "financial_investment", "label": "금융투자", "amount_key": "scrt_ntby_tr_pbmn"},
    {"key": "insurance", "label": "보험", "amount_key": "insu_ntby_tr_pbmn"},
    {"key": "trust", "label": "투신", "amount_key": "ivtr_ntby_tr_pbmn"},
    {"key": "bank", "label": "은행", "amount_key": "bank_ntby_tr_pbmn"},
    {"key": "other_finance", "label": "기타금융", "amount_key": "mrbn_ntby_tr_pbmn"},
    {"key": "pension", "label": "연기금등", "amount_key": "fund_ntby_tr_pbmn"},
    {"key": "private_fund", "label": "사모펀드", "amount_key": "pe_fund_ntby_tr_pbmn"},
    {"key": "other_corp", "label": "기타법인", "amount_key": "etc_corp_ntby_tr_pbmn"},
    {"key": "other_foreigner", "label": "기타외국인", "amount_key": "frgn_nreg_ntby_pbmn"},
]


def kis_market_amount_100m(row: dict[str, Any], key: str) -> float | None:
    value = clean_numeric_text(row.get(key))
    if value is None:
        return None
    # KIS 거래대금 필드는 백만원 단위라 억원 단위로 환산한다.
    return round(float(value) / 100.0, 1)


def fetch_kis_market_investor_flow(market_key: str, market_name: str, index_code: str) -> dict[str, Any]:
    today_text = datetime.now(timezone(timedelta(hours=9))).strftime("%Y%m%d")
    payload = kis_get_json(
        "/uapi/domestic-stock/v1/quotations/inquire-investor-daily-by-market",
        "FHPTJ04040000",
        {
            "FID_COND_MRKT_DIV_CODE": "U",
            "FID_INPUT_ISCD": index_code,
            "FID_INPUT_DATE_1": today_text,
            "FID_INPUT_ISCD_1": market_key,
            "FID_INPUT_DATE_2": today_text,
            "FID_INPUT_ISCD_2": index_code,
        },
    )
    output = payload.get("output")
    raw_rows = output if isinstance(output, list) else ([output] if isinstance(output, dict) else [])
    if not raw_rows:
        raise ValueError(f"{market_name} 시장 수급 데이터가 비어 있습니다.")
    item = raw_rows[0]
    if not isinstance(item, dict):
        raise ValueError(f"{market_name} 시장 수급 데이터 형식이 올바르지 않습니다.")
    row_date = str(item.get("stck_bsop_date") or "")
    values = {
        column["key"]: kis_market_amount_100m(item, str(column["amount_key"]))
        for column in MARKET_INVESTOR_COLUMNS
    }
    return {
        "market": market_key,
        "name": market_name,
        "index_code": index_code,
        "date": f"{row_date[:4]}-{row_date[4:6]}-{row_date[6:8]}" if len(row_date) == 8 else row_date,
        "index_value": clean_numeric_text(item.get("bstp_nmix_prpr")),
        "change_pct": parse_percent_text(item.get("bstp_nmix_prdy_ctrt")),
        "values": values,
    }


def build_market_investor_flow_summary() -> dict[str, Any]:
    columns = [{"key": column["key"], "label": column["label"]} for column in MARKET_INVESTOR_COLUMNS]
    if is_public_web_mode():
        return {
            "columns": columns,
            "markets": [],
            "unit": "억원",
            "source": "한국투자증권 OpenAPI",
            "error": "공개 웹 모드에서는 한국투자증권 API 수급 데이터가 잠겨 있습니다.",
        }
    if MARKET_INVESTOR_FLOW_CACHE_PATH.exists():
        try:
            cached = json.loads(MARKET_INVESTOR_FLOW_CACHE_PATH.read_text(encoding="utf-8"))
            cached_at = datetime.fromisoformat(str(cached.get("cached_at", "")))
            if cached_at > datetime.now() - timedelta(minutes=5) and cached.get("markets"):
                return {key: value for key, value in cached.items() if key != "cached_at"}
        except Exception:
            pass
    try:
        markets = [
            fetch_kis_market_investor_flow("KSP", "코스피", "0001"),
            fetch_kis_market_investor_flow("KSQ", "코스닥", "1001"),
        ]
        payload = {
            "columns": columns,
            "markets": markets,
            "unit": "억원",
            "source": "한국투자증권 OpenAPI 시장별 투자자매매동향(일별)",
            "as_of_date": max((market.get("date") or "" for market in markets), default=""),
        }
        MARKET_INVESTOR_FLOW_CACHE_PATH.write_text(
            json.dumps({**payload, "cached_at": datetime.now().isoformat(timespec="seconds")}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return payload
    except Exception as exc:
        if MARKET_INVESTOR_FLOW_CACHE_PATH.exists():
            try:
                cached = json.loads(MARKET_INVESTOR_FLOW_CACHE_PATH.read_text(encoding="utf-8"))
                if cached.get("markets"):
                    return {
                        **{key: value for key, value in cached.items() if key != "cached_at"},
                        "warning": "한국투자증권 API가 일시 실패해 마지막 정상 수급 데이터를 표시합니다.",
                    }
            except Exception:
                pass
        return {
            "columns": columns,
            "markets": [],
            "unit": "억원",
            "source": "한국투자증권 OpenAPI 시장별 투자자매매동향(일별)",
            "error": "시장별 수급 데이터를 잠시 가져오지 못했습니다. 현재가/등락률 새로고침을 다시 눌러주세요.",
        }


def compute_foreigner_net_value_to_marcap(code: str, market_cap: float | None, window: int = 20) -> float | None:
    if market_cap in (None, 0):
        return None
    investor_frame = fetch_naver_investor_frame(code)
    if investor_frame.empty:
        return None
    recent = investor_frame.tail(window)
    net_value = (pd.to_numeric(recent["foreigner_net"], errors="coerce") * pd.to_numeric(recent["close"], errors="coerce")).dropna()
    if net_value.empty:
        return None
    return round(float(net_value.sum()) / float(market_cap), 5)


def clean_numeric_text(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    text = str(value).strip()
    if not text or text in {"-", "N/A"}:
        return None
    negative = text.startswith("(") and text.endswith(")")
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if not cleaned:
        return None
    try:
        number = float(cleaned)
        return -number if negative and number > 0 else number
    except ValueError:
        return None


@lru_cache(maxsize=128)
def get_benchmark_return_series(symbol: str, start: str, end: str) -> pd.Series:
    try:
        frame = fdr.DataReader(symbol, start, end)
    except Exception:
        return pd.Series(dtype=float)
    if frame is None or frame.empty or "Close" not in frame.columns:
        return pd.Series(dtype=float)
    close = pd.to_numeric(frame["Close"], errors="coerce").dropna()
    if close.empty:
        return pd.Series(dtype=float)
    close.index = pd.to_datetime(close.index)
    return close.pct_change().dropna()


def compute_stock_beta(price_frame: pd.DataFrame, market: str | None, window: int = 120) -> float | None:
    if price_frame is None or price_frame.empty or "Close" not in price_frame.columns:
        return None
    frame = price_frame.tail(max(30, int(window or 120)) + 5).copy()
    if "Date" in frame.columns:
        frame.index = pd.to_datetime(frame["Date"], errors="coerce")
    else:
        frame.index = pd.to_datetime(frame.index, errors="coerce")
    close = pd.to_numeric(frame["Close"], errors="coerce").dropna()
    close = close[~close.index.isna()]
    if len(close) < 30:
        return None
    start = close.index.min().date().isoformat()
    end = close.index.max().date().isoformat()
    market_text = str(market or "").upper()
    benchmark_symbol = "KQ11" if "KOSDAQ" in market_text or "KQ" in market_text else "KS11"
    benchmark_returns = get_benchmark_return_series(benchmark_symbol, start, end)
    stock_returns = close.pct_change().dropna()
    aligned = pd.concat([stock_returns.rename("stock"), benchmark_returns.rename("market")], axis=1, join="inner").dropna()
    if len(aligned) < 25:
        return None
    market_var = float(aligned["market"].var())
    if not np.isfinite(market_var) or market_var == 0:
        return None
    beta = float(aligned["stock"].cov(aligned["market"]) / market_var)
    if not np.isfinite(beta):
        return None
    return round(beta, 2)


def build_sector_stock_row(sector: str, stock_input: SectorStockItem) -> dict[str, Any]:
    resolved = resolve_stock_payload(stock_input.code, stock_input.name)
    if not resolved:
        raise ValueError(f"종목을 찾을 수 없습니다: {stock_input.name or stock_input.code or ''}")

    code = resolved["code"]
    price_frame = fetch_price_frame(code)
    if price_frame.empty:
        raise ValueError(f"가격 데이터를 가져오지 못했습니다: {resolved['name']}")

    closes = pd.to_numeric(price_frame["Close"], errors="coerce").dropna().reset_index(drop=True)
    if closes.empty:
        raise ValueError(f"종가 데이터가 비어 있습니다: {resolved['name']}")

    as_of_date = price_frame["Date"].iloc[-1].strftime("%Y-%m-%d")
    current_price = to_float(closes.iloc[-1]) or resolved.get("close") or 0.0
    stocks_outstanding = to_float(resolved.get("stocks"))
    market_cap = to_float(resolved.get("marcap"))

    w_return_pct = get_return_between(closes, 5, 0)
    w1_return_pct = get_return_between(closes, 10, 5)
    m1_return_pct = get_return_between(closes, 20, 0)
    m3_return_pct = get_return_between(closes, 60, 0)
    ytd_return_pct = get_ytd_return_from_frame(price_frame)
    avg_trading_value_marcap_pct = compute_avg_trading_value_to_marcap(price_frame, market_cap)
    foreigner_net_value_marcap_pct = compute_foreigner_net_value_to_marcap(code, market_cap)
    beta_120d = compute_stock_beta(price_frame, resolved.get("market"))

    strength_components = [
        value
        for value in [w_return_pct, w1_return_pct, m1_return_pct, m3_return_pct, avg_trading_value_marcap_pct]
        if value is not None
    ]
    strength_score = round(sum(strength_components) / len(strength_components), 2) if strength_components else None

    return {
        "sector": sector,
        "stock_name": resolved["name"],
        "stock_code": code,
        "market": resolved.get("market", ""),
        "as_of_date": as_of_date,
        "current_price": round(current_price, 2),
        "market_cap_100m": round(market_cap / 100000000.0, 1) if market_cap is not None else None,
        "w_return_pct": w_return_pct,
        "w1_return_pct": w1_return_pct,
        "m1_return_pct": m1_return_pct,
        "m3_return_pct": m3_return_pct,
        "ytd_return_pct": ytd_return_pct,
        "foreigner_net_k": None,
        "institution_net_k": None,
        "foreigner_pct": None,
        "institution_pct": None,
        "avg_trading_value_marcap_pct": avg_trading_value_marcap_pct,
        "foreigner_net_value_marcap_pct": foreigner_net_value_marcap_pct,
        "beta_120d": beta_120d,
        "strength_score": strength_score,
        "market_cap_krw": market_cap,
        "shares_outstanding": stocks_outstanding,
    }


def build_sector_average_row(sector: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    numeric_fields = [
        "market_cap_100m",
        "w_return_pct",
        "w1_return_pct",
        "m1_return_pct",
        "m3_return_pct",
        "ytd_return_pct",
        "foreigner_net_k",
        "institution_net_k",
        "foreigner_pct",
        "institution_pct",
        "avg_trading_value_marcap_pct",
        "foreigner_net_value_marcap_pct",
        "beta_120d",
        "strength_score",
    ]
    result: dict[str, Any] = {
        "sector": sector,
        "stock_count": len(rows),
        "as_of_date": max((row.get("as_of_date") or "" for row in rows), default=""),
    }
    for field in numeric_fields:
        values = [to_float(row.get(field)) for row in rows if to_float(row.get(field)) is not None]
        digits = 5 if field in {"avg_trading_value_marcap_pct", "foreigner_net_value_marcap_pct"} else 2
        result[field] = round(sum(values) / len(values), digits) if values else None
    return result


def build_sector_snapshot(request: SectorSnapshotRequest) -> dict[str, Any]:
    normalized_groups: list[tuple[str, list[SectorStockItem]]] = []
    for group in request.groups:
        sector_name = str(group.sector or "").strip()
        stock_items = [item for item in group.stocks if str(item.name or item.code or "").strip()]
        if sector_name and stock_items:
            normalized_groups.append((sector_name, stock_items))

    if not normalized_groups:
        raise ValueError("섹터와 종목을 하나 이상 입력해 주세요.")

    tasks: list[tuple[int, str, SectorStockItem]] = []
    task_order = 0
    for sector_name, stock_items in normalized_groups:
        for stock_item in stock_items:
            tasks.append((task_order, sector_name, stock_item))
            task_order += 1

    stock_rows: list[dict[str, Any]] = []
    errors: list[str] = []
    with ThreadPoolExecutor(max_workers=min(8, max(1, len(tasks)))) as executor:
        futures = {
            executor.submit(build_sector_stock_row, sector_name, stock_item): (order, sector_name, stock_item)
            for order, sector_name, stock_item in tasks
        }
        for future in as_completed(futures):
            order, sector_name, stock_item = futures[future]
            try:
                row = future.result()
                row["_order"] = order
                stock_rows.append(row)
            except Exception as exc:
                label = stock_item.name or stock_item.code or "종목"
                errors.append(f"{sector_name} / {label}: {exc}")

    if not stock_rows:
        raise ValueError(errors[0] if errors else "종목 데이터를 불러오지 못했습니다.")

    stock_rows = sorted(
        stock_rows,
        key=lambda row: (
            str(row.get("sector") or ""),
            -(to_float(row.get("beta_120d")) if to_float(row.get("beta_120d")) is not None else -9999),
            row.get("_order", 0),
        ),
    )
    sector_groups: dict[str, list[dict[str, Any]]] = {}
    for row in stock_rows:
        sector_groups.setdefault(row["sector"], []).append(row)

    sector_rows = [build_sector_average_row(sector_name, rows) for sector_name, rows in sector_groups.items()]
    sector_rows = sorted(
        sector_rows,
        key=lambda row: (
            to_float(row.get("beta_120d")) if to_float(row.get("beta_120d")) is not None else -9999,
            to_float(row.get("strength_score")) if to_float(row.get("strength_score")) is not None else -9999,
        ),
        reverse=True,
    )
    as_of_date = max((row.get("as_of_date") or "" for row in stock_rows), default="")

    return {
        "as_of_date": as_of_date,
        "stock_rows": stock_rows,
        "sector_rows": sector_rows,
        "summary": {
            "sector_count": len(sector_rows),
            "stock_count": len(stock_rows),
            "error_count": len(errors),
        },
        "errors": errors,
    }


@lru_cache(maxsize=1)
def build_market_ytd_ranking_cached() -> dict[str, Any]:
    listing = get_listing_table().copy()
    listing = listing.dropna(subset=["Code", "Name"]).copy()
    listing["Code"] = listing["Code"].astype(str).str.zfill(6)
    listing["Marcap"] = pd.to_numeric(listing.get("Marcap"), errors="coerce")
    listing = listing.dropna(subset=["Marcap"])
    listing = listing.sort_values("Marcap", ascending=False).head(MARKET_YTD_UNIVERSE_LIMIT)
    today = date.today()
    start = date(today.year, 1, 1).isoformat()
    end = today.isoformat()
    codes = listing["Code"].drop_duplicates().tolist()
    latest_close_map = {
        str(row["Code"]).zfill(6): to_float(row.get("Close"))
        for _, row in listing.iterrows()
    }
    ytd_map: dict[str, float] = {}
    batch_size = 300
    for index in range(0, len(codes), batch_size):
        batch = codes[index:index + batch_size]
        try:
            price_frame = fdr.DataReader(batch, start, end)
        except Exception:
            continue
        if price_frame is None or price_frame.empty:
            continue
        if isinstance(price_frame, pd.Series):
            price_frame = price_frame.to_frame()
        for code in batch:
            if code not in price_frame.columns:
                continue
            closes = pd.to_numeric(price_frame[code], errors="coerce").dropna()
            if len(closes) < 2:
                continue
            first = to_float(closes.iloc[0])
            last = to_float(closes.iloc[-1]) or latest_close_map.get(code)
            if first in (None, 0) or last is None:
                continue
            ytd_map[code] = round(((last / first) - 1.0) * 100.0, 2)

    rows: list[dict[str, Any]] = []
    for _, row in listing.iterrows():
        code = str(row["Code"]).zfill(6)
        if code not in ytd_map:
            continue
        rows.append(
            {
                "stock_code": code,
                "stock_name": str(row.get("Name") or ""),
                "market": str(row.get("Market") or ""),
                "current_price": latest_close_map.get(code),
                "market_cap_100m": round(float(row.get("Marcap")) / 100000000.0, 1),
                "ytd_return_pct": ytd_map[code],
            }
        )
    rows.sort(key=lambda item: item.get("ytd_return_pct") if item.get("ytd_return_pct") is not None else -9999, reverse=True)
    for rank, row in enumerate(rows, start=1):
        row["rank"] = rank
    return {
        "as_of_date": today.isoformat(),
        "start_date": start,
        "rows": rows,
        "summary": {"stock_count": len(rows)},
        "universe_count": len(codes),
        "universe_label": f"시가총액 상위 {len(codes)}개 종목",
    }


def build_market_ytd_ranking(limit: int = 100) -> dict[str, Any]:
    payload = build_market_ytd_ranking_cached()
    safe_limit = max(10, min(int(limit or 100), 500))
    return {
        **payload,
        "rows": payload.get("rows", [])[:safe_limit],
        "summary": {
            **payload.get("summary", {}),
            "returned_count": min(safe_limit, len(payload.get("rows", []))),
            "universe_count": payload.get("universe_count"),
            "universe_label": payload.get("universe_label"),
        },
    }


def save_sector_watch_order(request: SectorWatchOrderRequest) -> dict[str, Any]:
    db = load_sector_db()
    current_sectors = [str(item or "").strip() for item in db.get("sectors", []) if str(item or "").strip()]
    next_sectors: list[str] = []
    for sector in request.sectors:
        sector_name = str(sector or "").strip()
        if sector_name and sector_name in current_sectors and sector_name not in next_sectors:
            next_sectors.append(sector_name)
    for sector in current_sectors:
        if sector not in next_sectors:
            next_sectors.append(sector)

    stock_map = dict(db.get("stock_map", {}))
    stocks_by_sector: dict[str, list[str]] = {}
    for key, item in stock_map.items():
        sector = str(item.get("sector") or "").strip()
        code = str(item.get("stock_code") or key or "").strip()
        if sector and code:
            stocks_by_sector.setdefault(sector, []).append(code)

    next_stock_order: dict[str, list[str]] = {}
    for sector in next_sectors:
        valid_codes = stocks_by_sector.get(sector, [])
        ordered: list[str] = []
        for code in request.stocks_by_sector.get(sector, []):
            code_text = str(code or "").strip()
            if code_text and code_text in valid_codes and code_text not in ordered:
                ordered.append(code_text)
        for code in valid_codes:
            if code not in ordered:
                ordered.append(code)
        if ordered:
            next_stock_order[sector] = ordered
    save_sector_db({"stock_map": stock_map, "sectors": next_sectors, "stock_order": next_stock_order})
    return build_sector_watch_board(limit_per_sector=80)


def build_sector_watch_board(limit_per_sector: int = 6) -> dict[str, Any]:
    groups = sector_db_groups()
    safe_limit = max(3, min(int(limit_per_sector or 6), 80))
    stock_inputs: list[dict[str, Any]] = []
    seen_codes: set[str] = set()
    for group in groups:
        sector = str(group.get("sector") or "").strip()
        if not sector:
            continue
        for stock in group.get("stocks", []):
            resolved = resolve_stock_payload(stock.get("code"), stock.get("name"))
            code = str((resolved or {}).get("code") or stock.get("code") or "").strip().zfill(6)
            name = str((resolved or {}).get("name") or stock.get("name") or "").strip()
            if not code or code == "000000" or code in seen_codes:
                continue
            seen_codes.add(code)
            stock_inputs.append({"sector": sector, "code": code, "name": name})

    codes = [item["code"] for item in stock_inputs]
    today = date.today()
    start = (today - timedelta(days=14)).isoformat()
    close_map: dict[str, float | None] = {}
    prev_close_map: dict[str, float | None] = {}
    date_map: dict[str, str] = {}
    if codes:
        batch_size = 250
        for index in range(0, len(codes), batch_size):
            batch = codes[index:index + batch_size]
            try:
                price_frame = fdr.DataReader(batch, start, today.isoformat())
            except Exception:
                price_frame = pd.DataFrame()
            if price_frame is None or price_frame.empty:
                continue
            if isinstance(price_frame, pd.Series):
                price_frame = price_frame.to_frame()
            for code in batch:
                if code not in price_frame.columns:
                    continue
                closes = pd.to_numeric(price_frame[code], errors="coerce").dropna()
                if closes.empty:
                    continue
                close_map[code] = to_float(closes.iloc[-1])
                prev_close_map[code] = to_float(closes.iloc[-2]) if len(closes) >= 2 else None
                last_index = closes.index[-1]
                date_map[code] = pd.to_datetime(last_index).strftime("%Y-%m-%d")

    listing = get_listing_table()
    listing_map = {
        str(row.get("Code") or "").zfill(6): row
        for _, row in listing.iterrows()
        if str(row.get("Code") or "").strip()
    }
    sector_rows: dict[str, list[dict[str, Any]]] = {}
    for item in stock_inputs:
        code = item["code"]
        listing_row = listing_map.get(code)
        current_price = close_map.get(code)
        if current_price is None and listing_row is not None:
            current_price = to_float(listing_row.get("Close"))
        prev_close = prev_close_map.get(code)
        change_pct = pct_change_between(current_price, prev_close)
        market = str(listing_row.get("Market") or "") if listing_row is not None else ""
        row = {
            "sector": item["sector"],
            "stock_code": code,
            "stock_name": item["name"],
            "market": market,
            "current_price": round(current_price or 0, 2) if current_price is not None else None,
            "prev_close": round(prev_close or 0, 2) if prev_close is not None else None,
            "change_pct": round(change_pct, 2) if change_pct is not None else None,
            "as_of_date": date_map.get(code) or today.isoformat(),
        }
        sector_rows.setdefault(item["sector"], []).append(row)

    sectors = []
    for group in groups:
        sector = str(group.get("sector") or "").strip()
        rows = sector_rows.get(sector, [])
        if not sector or not rows:
            continue
        displayed = rows[:safe_limit]
        valid_changes = [float(row["change_pct"]) for row in rows if row.get("change_pct") is not None]
        sectors.append(
            {
                "sector": sector,
                "stocks": displayed,
                "stock_count": len(rows),
                "avg_change_pct": round(sum(valid_changes) / len(valid_changes), 2) if valid_changes else None,
            }
        )

    return {
        "as_of_date": max((row.get("as_of_date") or "" for rows in sector_rows.values() for row in rows), default=today.isoformat()),
        "sectors": sectors,
        "market_investor_flows": build_market_investor_flow_summary(),
        "summary": {
            "sector_count": len(sectors),
            "stock_count": sum(len(rows) for rows in sector_rows.values()),
            "limit_per_sector": safe_limit,
        },
        "source": "섹터 DB / FinanceDataReader",
    }


def trade_data_api_key() -> str:
    settings = load_settings()
    public_data = settings.get("public_data", {}) if isinstance(settings.get("public_data"), dict) else {}
    return (
        os.getenv("STOCK_DASHBOARD_CUSTOMS_API_KEY")
        or os.getenv("CUSTOMS_API_KEY")
        or os.getenv("STOCK_DASHBOARD_DATAGO_API_KEY")
        or os.getenv("DATA_GO_KR_API_KEY")
        or os.getenv("DATAGO_API_KEY")
        or str(public_data.get("data_go_kr_api_key") or "")
        or str(public_data.get("customs_service_key") or "")
        or ""
    ).strip()


def motie_trade_service_key() -> str:
    settings = load_settings()
    public_data = settings.get("public_data", {}) if isinstance(settings.get("public_data"), dict) else {}
    return (
        os.getenv("STOCK_DASHBOARD_MOTIE_TRADE_API_KEY")
        or os.getenv("MOTIE_TRADE_API_KEY")
        or str(public_data.get("motie_trade_service_key") or "")
        or str(public_data.get("trade_service_key") or "")
        or ""
    ).strip()


def check_motie_trade_service(api_key: str) -> dict[str, Any]:
    if not api_key:
        return {"configured": False, "ok": False, "message": "산업부 무역정보서비스 API 키가 설정되어 있지 않습니다."}
    endpoint = "http://apis.data.go.kr/1450000/infoSrvcGuidanceService/getAreaExpimpInfoGuidance"
    try:
        response = requests.get(
            endpoint,
            params={
                "ServiceKey": api_key,
                "stYearMonth": "201408",
                "endYearMonth": "201409",
                "orgCd": "11000",
                "numOfRows": "1",
                "pageNo": "1",
            },
            timeout=8,
        )
        body = response.text or ""
        result_msg = ""
        result_code = ""
        try:
            root = ET.fromstring(body)
            result_code = root.findtext(".//resultCode") or ""
            result_msg = root.findtext(".//resultMsg") or ""
        except ET.ParseError:
            result_msg = body[:120]
        return {
            "configured": True,
            "ok": response.ok and (not result_code or result_code == "00"),
            "http_status": response.status_code,
            "result_code": result_code,
            "message": result_msg or response.reason,
            "endpoint": endpoint,
            "note": "이 API는 문서 기준 기업/거래물품/지역별 수출입 안내용이며, 라면·화장품 같은 품목별 월간 수출액은 관세청 품목별 수출입실적 API가 더 정확합니다.",
        }
    except Exception as exc:
        return {"configured": True, "ok": False, "message": str(exc), "endpoint": endpoint}


def trade_month_key(year: int, month: int) -> str:
    return f"{int(year):04d}{int(month):02d}"


def parse_trade_number(value: Any) -> float | None:
    text = str(value or "").replace(",", "").strip()
    if not text or text == "-":
        return None
    try:
        return float(text)
    except ValueError:
        return None


def trade_year_range(start_year: int = 2021) -> list[int]:
    return list(range(start_year, date.today().year + 1))


def build_trade_sample_payload() -> dict[str, Any]:
    years = trade_year_range(2021)
    items: list[dict[str, Any]] = []
    for item in TRADE_IMPORT_EXPORT_ITEMS:
        rows = []
        values_by_year = item.get("sample_values", {})
        for month in range(1, 13):
            month_values = []
            for year in years:
                raw_values = values_by_year.get(str(year), [None] * 12)
                value = raw_values[month - 1] if month - 1 < len(raw_values) else None
                month_values.append({"year": year, "value": value})
            rows.append({"month": month, "values": month_values})
        items.append({
            "key": item["key"],
            "name": item["name"],
            "hs_codes": item["hs_codes"],
            "companies": item.get("companies", []),
            "unit": item["unit"],
            "rows": rows,
            "metrics": trade_item_change_metrics(values),
        })
    return {
        "source": "sample",
        "source_label": "샘플 데이터",
        "message": "공공데이터포털 관세청 API 키가 없어 이미지 형태 확인용 샘플 데이터로 표시합니다.",
        "release_hint": "관세청 수출입무역통계는 매월 1일, 11일, 21일 잠정/확정치 확인 용도로 사용합니다.",
        "years": years,
        "items": items,
        "trade_items_version": TRADE_DATA_VERSION,
        "loaded_at": datetime.now().isoformat(timespec="seconds"),
    }


def parse_trade_api_response(payload_text: str, fallback_month: str = "") -> dict[str, float]:
    result: dict[str, float] = {}
    try:
        root = ET.fromstring(payload_text)
    except ET.ParseError:
        return result
    parsed_items: list[dict[str, Any]] = []
    for item in root.findall(".//item"):
        month_value = ""
        export_value = None
        has_detail_code = False
        is_total = False
        for child in list(item):
            tag = child.tag.split("}")[-1]
            text = child.text or ""
            if tag in {"year", "yearMm", "statCd", "period", "priodTitle"}:
                if "총계" in text:
                    is_total = True
                digits = re.sub(r"\D", "", text)
                if len(digits) >= 6:
                    month_value = digits[:6]
            if tag in {"expDlr", "expDlrAmt", "exportAmount", "expUsd", "expUsdAmt"}:
                export_value = parse_trade_number(text)
            if tag in {"hsCode", "hsSgn"} and re.sub(r"\D", "", text):
                has_detail_code = True
        if not month_value and fallback_month:
            month_value = fallback_month
        if month_value and export_value is not None:
            parsed_items.append({
                "month": month_value,
                "value": export_value,
                "has_detail_code": has_detail_code,
                "is_total": is_total,
            })
    use_detail_rows = any(item.get("has_detail_code") for item in parsed_items)
    for item in parsed_items:
        if use_detail_rows and not item.get("has_detail_code"):
            continue
        if item.get("is_total") and use_detail_rows:
            continue
        month_value = str(item.get("month") or "")
        export_value = parse_trade_number(item.get("value"))
        if month_value and export_value is not None:
            result[month_value] = result.get(month_value, 0) + export_value / 1000.0
    return result


def request_trade_item_payload(api_key: str, year: int, hs_code: str) -> str:
    new_gateway = requests.get(
        "https://apis.data.go.kr/1220000/Itemtrade/getItemtradeList",
        params={"ServiceKey": api_key, "strtYymm": trade_month_key(year, 1), "endYymm": trade_month_key(year, 12), "hsSgn": hs_code},
        timeout=20,
    )
    new_gateway.raise_for_status()
    if new_gateway.text:
        return new_gateway.text
    legacy_gateway = requests.get(
        "http://openapi.customs.go.kr/openapi/service/newTradestatistics/getitemtradeList",
        params={
            "ServiceKey": api_key,
            "searchBgnDe": trade_month_key(year, 1),
            "searchEndDe": trade_month_key(year, 12),
            "searchItemCd": hs_code,
        },
        timeout=20,
    )
    legacy_gateway.raise_for_status()
    return legacy_gateway.text


def trade_region_option(region: str) -> dict[str, str]:
    return next((item for item in TRADE_REGION_OPTIONS if item["code"] == region), TRADE_REGION_OPTIONS[0])


def request_sido_trade_item_payload(api_key: str, year: int, hs_code: str, region: str) -> str:
    return request_sido_trade_item_payload_for_month(api_key, trade_month_key(year, 1), trade_month_key(year, 12), hs_code, region)


def request_sido_trade_item_payload_for_month(api_key: str, start_month: str, end_month: str, hs_code: str, region: str) -> str:
    region_option = trade_region_option(region)
    sido_code = region_option.get("sido_code") or region_option.get("code") or ""
    response = requests.get(
        "https://apis.data.go.kr/1220000/sidoitemtrade/getSidoitemtradeList",
        params={
            "serviceKey": api_key,
            "numOfRows": "1000",
            "pageNo": "1",
            "strtYymm": start_month,
            "endYymm": end_month,
            "hsSgn": hs_code,
            "sidoCd": sido_code,
        },
        timeout=20,
    )
    response.raise_for_status()
    return response.text


def fetch_trade_item_real_data(item: dict[str, Any], years: list[int], api_key: str, region: str = "all") -> dict[str, float]:
    values: dict[str, float] = {}
    for hs_code in item["hs_codes"]:
        for year in years:
            # The public customs endpoint documents a maximum one-year query window.
            if region and region != "all":
                for month in range(1, 13):
                    month_key = trade_month_key(year, month)
                    raw_payload = request_sido_trade_item_payload_for_month(api_key, month_key, month_key, hs_code, region)
                    parsed = parse_trade_api_response(raw_payload, fallback_month=month_key)
                    for key, value in parsed.items():
                        values[key] = values.get(key, 0) + value
                continue
            else:
                raw_payload = request_trade_item_payload(api_key, year, hs_code)
            parsed = parse_trade_api_response(raw_payload)
            for key, value in parsed.items():
                values[key] = values.get(key, 0) + value
    return values


def trade_data_cache_path(region: str, item_key: str = "") -> Path:
    normalized_region = re.sub(r"[^A-Za-z0-9_-]", "", str(region or "all")) or "all"
    if normalized_region == "all":
        return TRADE_DATA_CACHE_PATH
    normalized_item = re.sub(r"[^A-Za-z0-9_-]", "", str(item_key or "")) or "all"
    return STATE_DIR / f"trade_import_export_cache_{normalized_region}_{normalized_item}.json"


def trade_item_change_metrics(values: dict[str, float]) -> dict[str, Any]:
    numeric_values = {
        str(key): float(value)
        for key, value in (values or {}).items()
        if re.fullmatch(r"\d{6}", str(key or "")) and value is not None and float(value) > 0
    }
    if not numeric_values:
        return {"latest_month": "", "latest_value": None, "qoq_pct": None, "yoy_pct": None}
    latest_month = max(numeric_values)
    year = int(latest_month[:4])
    month = int(latest_month[4:6])
    prev_year = year
    prev_month = month - 1
    if prev_month <= 0:
        prev_year -= 1
        prev_month = 12
    prev_key = f"{prev_year:04d}{prev_month:02d}"
    yoy_key = f"{year - 1:04d}{month:02d}"
    latest_value = numeric_values.get(latest_month)
    prev_value = numeric_values.get(prev_key)
    yoy_value = numeric_values.get(yoy_key)

    def pct(current: float | None, previous: float | None) -> float | None:
        if current is None or previous is None or previous == 0:
            return None
        return round((current / previous - 1) * 100, 1)

    return {
        "latest_month": latest_month,
        "latest_value": latest_value,
        "qoq_pct": pct(latest_value, prev_value),
        "yoy_pct": pct(latest_value, yoy_value),
    }


def trade_snapshot_slot(collected_on: date | None = None) -> tuple[str, str]:
    target = collected_on or date.today()
    if target.day <= 10:
        slot_day = 1
    elif target.day <= 20:
        slot_day = 11
    else:
        slot_day = 21
    return f"{target.year:04d}{target.month:02d}:{slot_day:02d}", f"{slot_day}일"


def trade_cache_policy_label() -> str:
    return "수출입 데이터는 매월 1일, 11일, 21일 발표 구간이 바뀔 때만 자동 재조회하고 그 전에는 DB 캐시를 사용합니다."


def load_trade_snapshot_history() -> dict[str, Any]:
    if not TRADE_SNAPSHOT_HISTORY_PATH.exists():
        return {"snapshots": {}}
    try:
        payload = json.loads(TRADE_SNAPSHOT_HISTORY_PATH.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else {"snapshots": {}}
    except Exception:
        return {"snapshots": {}}


def save_trade_snapshot_history(history: dict[str, Any]) -> None:
    try:
        temp_path = TRADE_SNAPSHOT_HISTORY_PATH.with_name(f"{TRADE_SNAPSHOT_HISTORY_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
        temp_path.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        temp_path.replace(TRADE_SNAPSHOT_HISTORY_PATH)
    except Exception:
        pass


def persist_trade_snapshot(payload: dict[str, Any], region: str) -> None:
    if not payload.get("items"):
        return
    slot_key, slot_label = trade_snapshot_slot()
    history = load_trade_snapshot_history()
    snapshots = history.setdefault("snapshots", {})
    snapshot = {
        "slot": slot_key,
        "slot_label": slot_label,
        "collected_at": datetime.now().isoformat(timespec="seconds"),
        "region": region,
        "source": payload.get("source"),
        "items": {},
    }
    for item in payload.get("items", []):
        item_payload: dict[str, Any] = {}
        for row in item.get("rows", []):
            month_key = f"{int(row.get('month') or 0):02d}"
            year_values = {}
            for entry in row.get("values", []):
                value = entry.get("value")
                if value is not None:
                    year_values[str(entry.get("year"))] = value
            if year_values:
                item_payload[month_key] = year_values
        if item_payload:
            snapshot["items"][item.get("key")] = item_payload
    snapshots[f"{region}:{slot_key}"] = snapshot
    history["updated_at"] = snapshot["collected_at"]
    save_trade_snapshot_history(history)


BUILDING_UNIT_LAYOUT = [
    {"floor": "B1", "units": ["B01", "B02", "B03", "B04"]},
    {"floor": "1F", "units": ["101", "102", "103"]},
    {"floor": "2F", "units": ["201", "202", "203"]},
]

BUILDING_UNIT_AREAS = {
    "B01": {"exclusive_m2": 20.94, "parking_m2": 6.69, "common_m2": 8.78, "exclusive_py": 6.3, "sale_py": 11.01},
    "B02": {"exclusive_m2": 21.31, "parking_m2": 6.8, "common_m2": 8.94, "exclusive_py": 6.5, "sale_py": 11.2},
    "B03": {"exclusive_m2": 21.25, "parking_m2": 6.78, "common_m2": 8.91, "exclusive_py": 6.4, "sale_py": 11.17},
    "B04": {"exclusive_m2": 24.99, "parking_m2": 7.98, "common_m2": 10.48, "exclusive_py": 7.5, "sale_py": 13.14},
    "101": {"exclusive_m2": 27.69, "parking_m2": 8.84, "common_m2": 11.61, "exclusive_py": 8.4, "sale_py": 14.56},
    "102": {"exclusive_m2": 28.17, "parking_m2": 9.0, "common_m2": 11.82, "exclusive_py": 8.5, "sale_py": 14.82},
    "103": {"exclusive_m2": 52.53, "parking_m2": 16.77, "common_m2": 22.03, "exclusive_py": 15.9, "sale_py": 27.62},
    "201": {"exclusive_m2": 30.59, "parking_m2": 9.77, "common_m2": 12.83, "exclusive_py": 9.25, "sale_py": 16.08},
    "202": {"exclusive_m2": 28.17, "parking_m2": 9.0, "common_m2": 11.81, "exclusive_py": 8.5, "sale_py": 14.8},
    "203": {"exclusive_m2": 52.53, "parking_m2": 16.77, "common_m2": 22.03, "exclusive_py": 15.9, "sale_py": 27.62},
}

BUILDING_UNIT_RENT_PLANS = {
    "B01": {"deposit_manwon": 2000, "monthly_rent_manwon": 85, "management_fee_manwon": 0},
    "B02": {"deposit_manwon": 2000, "monthly_rent_manwon": 90, "management_fee_manwon": 0},
    "B03": {"deposit_manwon": 2000, "monthly_rent_manwon": 90, "management_fee_manwon": 0},
    "B04": {"deposit_manwon": 2000, "monthly_rent_manwon": 110, "management_fee_manwon": 0},
    "101": {"deposit_manwon": 2000, "monthly_rent_manwon": 80, "management_fee_manwon": 0},
    "102": {"deposit_manwon": 2000, "monthly_rent_manwon": 80, "management_fee_manwon": 0},
    "103": {"deposit_manwon": 2000, "monthly_rent_manwon": 110, "management_fee_manwon": 0},
    "201": {"deposit_manwon": 2000, "monthly_rent_manwon": 85, "management_fee_manwon": 0},
    "202": {"deposit_manwon": 2000, "monthly_rent_manwon": 80, "management_fee_manwon": 0},
    "203": {"deposit_manwon": 2000, "monthly_rent_manwon": 110, "management_fee_manwon": 0},
}

BUILDING_WATER_DEFAULT_MONTHS = ["2026-01", "2026-03", "2026-05", "2026-07", "2026-09", "2026-11"]
BUILDING_WATER_DEFAULT_READINGS = {
    "2026-01": {"B01": 0.01, "B02": 0.69, "B03": 0.0, "B04": 0.01, "101": 0.02, "102": 0.0, "103": 0.0, "201": 0.0, "202": 0.0, "203": 0.01},
    "2026-03": {"B01": 0.01, "B02": 1.04, "B03": 0.0, "B04": 0.01, "101": 0.02, "102": 0.0, "103": 0.0, "201": 0.0, "202": 0.0, "203": 0.01},
    "2026-05": {"B01": 0.01, "B02": 1.57, "B03": 0.0, "B04": 0.01, "101": 0.02, "102": 0.0, "103": 0.0, "201": 0.0, "202": 0.0, "203": 0.01},
}

BUILDING_ELECTRIC_CUSTOMERS = {
    "common": {"label": "상가", "customer_no": "01 6035 7485"},
    "B01": {"label": "B01호", "customer_no": "01 6035 7519"},
    "B02": {"label": "B02호", "customer_no": "01 6035 7528"},
    "B03": {"label": "B03호", "customer_no": "01 6035 7537"},
    "B04": {"label": "B04호", "customer_no": "01 6035 7546"},
    "101": {"label": "101호", "customer_no": "01 6035 7555"},
    "102": {"label": "102호", "customer_no": "01 6035 7564"},
    "103": {"label": "103호", "customer_no": "01 6035 7573"},
    "201": {"label": "201호", "customer_no": "01 6035 7582"},
    "202": {"label": "202호", "customer_no": "01 6035 7591"},
    "203": {"label": "203호", "customer_no": "01 6035 7608"},
}

BUILDING_ELECTRIC_CUSTOMER_DIGITS = {
    re.sub(r"\D+", "", info["customer_no"]): unit_id
    for unit_id, info in BUILDING_ELECTRIC_CUSTOMERS.items()
}


REAL_ESTATE_PAYMENT_METHODS = {"현금인출", "하나카드", "삼성카드", "신한카드", "현대카드"}


def real_estate_payment_method(value: Any) -> str:
    text = str(value or "").strip()
    if text in REAL_ESTATE_PAYMENT_METHODS:
        return text
    compact = re.sub(r"\s+", "", text)
    if "하나" in compact:
        return "하나카드"
    if "삼성" in compact:
        return "삼성카드"
    if "신한" in compact:
        return "신한카드"
    if "현대" in compact:
        return "현대카드"
    return "현금인출"


def normalize_real_estate_transaction_list(items: Any) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    if not isinstance(items, list):
        return normalized
    for item in items:
        if not isinstance(item, dict):
            continue
        row = dict(item)
        row["payment_method"] = real_estate_payment_method(row.get("payment_method"))
        normalized.append(row)
    return normalized


def default_real_estate_db() -> dict[str, Any]:
    units = {}
    for floor in BUILDING_UNIT_LAYOUT:
        for unit_id in floor["units"]:
            units[unit_id] = {
                "unit_id": unit_id,
                "area": BUILDING_UNIT_AREAS.get(unit_id, {}),
                "rent_plan": BUILDING_UNIT_RENT_PLANS.get(unit_id, {}),
                "contract": {
                    "tenant": "",
                    "contract_date": "",
                    "balance_date": "",
                    "deposit": 0,
                    "discount_rate": 0,
                    "management_fee": 0,
                    "rent_free_months": 0,
                    "settlement_support_months": 0,
                    "rent_start_date": "",
                    "monthly_rent": 0,
                    "memo": "",
                },
                "water": {
                    "meter_start": 0,
                    "meter_end": 0,
                    "amount": 0,
                    "month": "",
                },
                "transactions": [],
            }
    return {
        "version": 1,
        "building_name": "상가 임대 관리",
        "investment": {
            "acquired_date": "2025-08-28",
            "purchase_price": 1250000000,
            "land_supply_price": 890000000,
            "building_supply_price": 360000000,
            "brokerage_fee": 11250000,
            "acquisition_tax_legal_fee": 62209922,
            "completion_date": "2024-07-29",
            "registration_date": "2025-07-02",
            "defect_warranty_years": 5,
        },
        "layout": BUILDING_UNIT_LAYOUT,
        "units": units,
        "service_contracts": [],
        "water_billing": {
            "source": "홀수달 1일 수도 검침",
            "months": [
                {
                    "month": month,
                    "total_usage_m3": 0,
                    "total_bill": 0,
                    "readings": BUILDING_WATER_DEFAULT_READINGS.get(month, {unit_id: 0 for unit_id in units}),
                    "memo": "",
                }
                for month in BUILDING_WATER_DEFAULT_MONTHS
            ],
        },
        "electricity_billing": {
            "source": "한국전력 모바일 청구서",
            "customers": BUILDING_ELECTRIC_CUSTOMERS,
            "months": [],
        },
        "operating_transactions": [],
        "bank_transactions": [],
        "bank_transaction_exclusions": [],
        "bank_memo_unit_map": {},
        "bank_memo_category_map": {},
        "bank_import": {
            "data_dir": str(REAL_ESTATE_BANK_IMPORT_DIR),
            "last_imported_at": "",
            "file_results": [],
            "total_files": 0,
            "total_transactions": 0,
        },
        "updated_at": "",
    }


def normalize_real_estate_db(payload: dict[str, Any] | None) -> dict[str, Any]:
    payload = repair_mojibake_recursive(payload)
    base = default_real_estate_db()
    if not isinstance(payload, dict):
        return base
    base["building_name"] = str(payload.get("building_name") or base["building_name"])
    base["updated_at"] = str(payload.get("updated_at") or "")
    if isinstance(payload.get("investment"), dict):
        base["investment"].update(payload.get("investment") or {})
    if isinstance(payload.get("service_contracts"), list):
        base["service_contracts"] = payload.get("service_contracts")
    if isinstance(payload.get("operating_transactions"), list):
        base["operating_transactions"] = normalize_real_estate_transaction_list(payload.get("operating_transactions"))
    if isinstance(payload.get("bank_transactions"), list):
        base["bank_transactions"] = normalize_real_estate_transaction_list(payload.get("bank_transactions"))
    if isinstance(payload.get("bank_transaction_exclusions"), list):
        base["bank_transaction_exclusions"] = [
            str(item)
            for item in payload.get("bank_transaction_exclusions")
            if str(item or "").strip()
        ]
    if isinstance(payload.get("bank_memo_unit_map"), dict):
        base["bank_memo_unit_map"] = {
            str(key): str(value)
            for key, value in payload.get("bank_memo_unit_map", {}).items()
            if str(key or "").strip() and str(value or "").strip()
        }
    if isinstance(payload.get("bank_memo_category_map"), dict):
        base["bank_memo_category_map"] = {
            str(key): str(value)
            for key, value in payload.get("bank_memo_category_map", {}).items()
            if str(key or "").strip() and str(value or "").strip()
        }
    if isinstance(payload.get("bank_import"), dict):
        base["bank_import"].update(payload.get("bank_import") or {})
    if isinstance(payload.get("water_billing"), dict):
        base["water_billing"].update(payload.get("water_billing") or {})
        base["water_billing"]["months"] = payload.get("water_billing", {}).get("months") if isinstance(payload.get("water_billing", {}).get("months"), list) else base["water_billing"]["months"]
    if isinstance(payload.get("electricity_billing"), dict):
        base["electricity_billing"].update(payload.get("electricity_billing") or {})
        base["electricity_billing"]["customers"] = BUILDING_ELECTRIC_CUSTOMERS
        base["electricity_billing"]["months"] = payload.get("electricity_billing", {}).get("months") if isinstance(payload.get("electricity_billing", {}).get("months"), list) else base["electricity_billing"]["months"]
    source_units = payload.get("units") if isinstance(payload.get("units"), dict) else {}
    for unit_id, unit in base["units"].items():
        incoming = source_units.get(unit_id) if isinstance(source_units, dict) else None
        if not isinstance(incoming, dict):
            continue
        unit["area"] = BUILDING_UNIT_AREAS.get(unit_id, incoming.get("area") if isinstance(incoming.get("area"), dict) else {})
        incoming_rent_plan = incoming.get("rent_plan") if isinstance(incoming.get("rent_plan"), dict) else {}
        unit["rent_plan"] = {
            **BUILDING_UNIT_RENT_PLANS.get(unit_id, {}),
            **incoming_rent_plan,
        }
        contract = incoming.get("contract") if isinstance(incoming.get("contract"), dict) else {}
        water = incoming.get("water") if isinstance(incoming.get("water"), dict) else {}
        unit["contract"].update(contract)
        unit["contract"]["vat_note"] = real_estate_management_vat_flag(unit["contract"].get("vat_note"))
        unit["water"].update(water)
        unit["transactions"] = incoming.get("transactions") if isinstance(incoming.get("transactions"), list) else []
    return base


def real_estate_date_text(value: Any) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return datetime.strptime(text[:10] if fmt != "%Y%m%d" else text[:8], fmt).date().isoformat()
        except Exception:
            continue
    return text


def real_estate_excel_date(value: Any) -> Any:
    text = real_estate_date_text(value)
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d")
    except Exception:
        return text


def real_estate_unit_id(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def parse_real_estate_area(area_m2: Any, area_py: Any) -> dict[str, float]:
    def parts(value: Any) -> list[float]:
        return [real_estate_number(part) for part in str(value or "").split("/")]

    m2 = parts(area_m2)
    py = parts(area_py)
    return {
        "exclusive_m2": m2[0] if len(m2) > 0 else 0,
        "parking_m2": m2[1] if len(m2) > 1 else 0,
        "common_m2": m2[2] if len(m2) > 2 else 0,
        "exclusive_py": py[0] if len(py) > 0 else 0,
        "sale_py": py[1] if len(py) > 1 else 0,
    }


def real_estate_numeric_cell(value: Any) -> bool:
    if value in (None, ""):
        return False
    text = str(value).strip().replace(",", "")
    return bool(re.fullmatch(r"-?\d+(?:\.\d+)?", text))


def real_estate_management_vat_flag(value: Any) -> str:
    text = str(value or "").strip().upper()
    compact = re.sub(r"\s+", "", text)
    if not compact:
        return "X"
    if compact in {"O", "Y", "YES", "TRUE", "1"}:
        return "O"
    if compact in {"X", "N", "NO", "FALSE", "0"}:
        return "X"
    if "미적용" in compact or "없음" in compact or "면세" in compact or "포함" in compact:
        return "X"
    if "적용" in compact or "별도" in compact:
        return "O"
    return "X"


def real_estate_monthly_charge_parts(rent: Any, management_fee: Any, management_vat_flag: Any) -> dict[str, float]:
    rent_amount = max(0.0, real_estate_number(rent))
    management_amount = max(0.0, real_estate_number(management_fee))
    rent_vat = round(rent_amount * 0.1)
    management_vat = round(management_amount * 0.1) if real_estate_management_vat_flag(management_vat_flag) == "O" else 0
    return {
        "rent": rent_amount,
        "rent_vat": rent_vat,
        "management_fee": management_amount,
        "management_vat": management_vat,
        "vat_total": rent_vat + management_vat,
        "total": rent_amount + management_amount + rent_vat + management_vat,
    }


def load_real_estate_excel_payload() -> dict[str, Any]:
    if not REAL_ESTATE_EXCEL_PATH.exists():
        return {"units": {}, "source_path": str(REAL_ESTATE_EXCEL_PATH), "loaded": False}
    workbook = None
    workbook_path = None
    try:
        workbook_path = safe_copy_to_temp(REAL_ESTATE_EXCEL_PATH)
        workbook = load_workbook(workbook_path, data_only=False)
        sheet = workbook["상가 임대 내역"] if "상가 임대 내역" in workbook.sheetnames else workbook.active
        units: dict[str, Any] = {}
        for row_index in range(4, sheet.max_row + 1):
            unit_id = real_estate_unit_id(sheet.cell(row=row_index, column=2).value)
            if not unit_id or unit_id == "합계" or unit_id not in BUILDING_UNIT_AREAS:
                continue
            plan_management_fee = real_estate_number(sheet.cell(row=row_index, column=7).value)
            raw_deposit = real_estate_number(sheet.cell(row=row_index, column=24).value)
            raw_contract_deposit = real_estate_number(sheet.cell(row=row_index, column=25).value)
            raw_balance_amount = real_estate_number(sheet.cell(row=row_index, column=26).value)
            raw_monthly_rent = real_estate_number(sheet.cell(row=row_index, column=27).value)
            raw_management_fee = real_estate_number(sheet.cell(row=row_index, column=28).value)
            raw_vat_note = real_estate_management_vat_flag(sheet.cell(row=row_index, column=29).value)
            rent_free_months_cell = sheet.cell(row=row_index, column=18).value
            legacy_rent_free_months_cell = sheet.cell(row=row_index, column=17).value
            settlement_months_cell = sheet.cell(row=row_index, column=20).value
            legacy_settlement_months_cell = sheet.cell(row=row_index, column=19).value
            rent_free_note_cell = sheet.cell(row=row_index, column=19).value
            legacy_rent_free_note_cell = sheet.cell(row=row_index, column=18).value
            settlement_note_cell = sheet.cell(row=row_index, column=21).value
            legacy_settlement_note_cell = sheet.cell(row=row_index, column=20).value
            rent_free_months = (
                int(real_estate_number(rent_free_months_cell))
                if real_estate_numeric_cell(rent_free_months_cell)
                else int(real_estate_number(legacy_rent_free_months_cell))
            )
            settlement_support_months = (
                int(real_estate_number(settlement_months_cell))
                if real_estate_numeric_cell(settlement_months_cell)
                else int(real_estate_number(legacy_settlement_months_cell))
            )
            rent_free_note = (
                str(rent_free_note_cell or "")
                if not real_estate_numeric_cell(rent_free_note_cell)
                else (str(legacy_rent_free_note_cell or "") if not real_estate_numeric_cell(legacy_rent_free_note_cell) else "")
            )
            settlement_support_note = (
                str(settlement_note_cell or "")
                if not real_estate_numeric_cell(settlement_note_cell)
                else (str(legacy_settlement_note_cell or "") if not real_estate_numeric_cell(legacy_settlement_note_cell) else "")
            )
            units[unit_id] = {
                "area": parse_real_estate_area(sheet.cell(row=row_index, column=3).value, sheet.cell(row=row_index, column=4).value),
                "rent_plan": {
                    "deposit_manwon": real_estate_number(sheet.cell(row=row_index, column=5).value),
                    "monthly_rent_manwon": real_estate_number(sheet.cell(row=row_index, column=6).value),
                    "management_fee_manwon": plan_management_fee,
                },
                "contract": {
                    "tenant_business": str(sheet.cell(row=row_index, column=8).value or ""),
                    "tenant": str(sheet.cell(row=row_index, column=9).value or ""),
                    "registration_no": str(sheet.cell(row=row_index, column=10).value or ""),
                    "phone": str(sheet.cell(row=row_index, column=11).value or ""),
                    "address": str(sheet.cell(row=row_index, column=12).value or ""),
                    "contract_date": real_estate_date_text(sheet.cell(row=row_index, column=13).value),
                    "balance_date": real_estate_date_text(sheet.cell(row=row_index, column=14).value),
                    "lease_end_date": real_estate_date_text(sheet.cell(row=row_index, column=15).value),
                    "lease_term": str(sheet.cell(row=row_index, column=16).value or ""),
                    "rent_payment_day": str(sheet.cell(row=row_index, column=17).value or ""),
                    "rent_free_months": rent_free_months,
                    "rent_free_note": rent_free_note,
                    "settlement_support_months": settlement_support_months,
                    "settlement_support_note": settlement_support_note,
                    "deposit_discount_formula": str(sheet.cell(row=row_index, column=22).value or ""),
                    "rent_discount_formula": str(sheet.cell(row=row_index, column=23).value or ""),
                    "deposit": raw_deposit * 10000,
                    "contract_deposit": raw_contract_deposit * 10000,
                    "balance_amount": raw_balance_amount * 10000,
                    "monthly_rent": raw_monthly_rent * 10000,
                    "management_fee": raw_management_fee * 10000,
                    "vat_note": raw_vat_note,
                },
                "excel_row": row_index,
            }
        return {"units": units, "source_path": str(REAL_ESTATE_EXCEL_PATH), "loaded": True}
    except Exception as exc:
        return {"units": {}, "source_path": str(REAL_ESTATE_EXCEL_PATH), "loaded": False, "error": str(exc)}
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass
        if workbook_path and workbook_path != REAL_ESTATE_EXCEL_PATH:
            try:
                workbook_path.unlink(missing_ok=True)
            except Exception:
                pass


def merge_real_estate_excel(db: dict[str, Any]) -> dict[str, Any]:
    excel_payload = load_real_estate_excel_payload()
    for unit_id, excel_unit in excel_payload.get("units", {}).items():
        if unit_id not in db.get("units", {}):
            continue
        unit = db["units"][unit_id]
        unit["area"] = excel_unit.get("area") or unit.get("area") or {}
        unit["rent_plan"] = excel_unit.get("rent_plan") or unit.get("rent_plan") or {}
        unit["excel_row"] = excel_unit.get("excel_row")
        unit["contract"].update(excel_unit.get("contract") or {})
        unit["contract"]["vat_note"] = real_estate_management_vat_flag(unit["contract"].get("vat_note"))
    db["excel_source"] = {
        "path": excel_payload.get("source_path"),
        "loaded": bool(excel_payload.get("loaded")),
        "error": excel_payload.get("error", ""),
    }
    return db


def real_estate_excel_rows_payload(db: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for unit_id, unit in (db.get("units") or {}).items():
        contract = unit.get("contract") or {}
        rent_plan = unit.get("rent_plan") or {}
        area = unit.get("area") or {}
        rows.append({
            "unit_id": unit_id,
            "area_m2": f"{area.get('exclusive_m2', 0)}/{area.get('parking_m2', 0)}/{area.get('common_m2', 0)}",
            "area_py": f"{area.get('exclusive_py', 0)}/{area.get('sale_py', 0)}",
            "plan_deposit": real_estate_number(rent_plan.get("deposit_manwon")),
            "plan_rent": real_estate_number(rent_plan.get("monthly_rent_manwon")),
            "plan_management_fee": real_estate_number(rent_plan.get("management_fee_manwon")),
            "tenant_business": contract.get("tenant_business") or "",
            "tenant": contract.get("tenant") or "",
            "registration_no": contract.get("registration_no") or "",
            "phone": contract.get("phone") or "",
            "address": contract.get("address") or "",
            "contract_date": real_estate_date_text(contract.get("contract_date")),
            "balance_date": real_estate_date_text(contract.get("balance_date")),
            "lease_end_date": real_estate_date_text(contract.get("lease_end_date")),
            "lease_term": contract.get("lease_term") or "",
            "rent_payment_day": contract.get("rent_payment_day") or "",
            "rent_free_months": int(real_estate_number(contract.get("rent_free_months"))),
            "rent_free_note": contract.get("rent_free_note") or "",
            "settlement_support_months": int(real_estate_number(contract.get("settlement_support_months"))),
            "settlement_support_note": contract.get("settlement_support_note") or "",
            "deposit": real_estate_number(contract.get("deposit")) / 10000 if real_estate_number(contract.get("deposit")) else None,
            "contract_deposit": real_estate_number(contract.get("contract_deposit")) / 10000 if real_estate_number(contract.get("contract_deposit")) else None,
            "balance_amount": real_estate_number(contract.get("balance_amount")) / 10000 if real_estate_number(contract.get("balance_amount")) else None,
            "monthly_rent": real_estate_number(contract.get("monthly_rent")) / 10000 if real_estate_number(contract.get("monthly_rent")) else None,
            "management_fee": real_estate_number(contract.get("management_fee")) / 10000 if real_estate_number(contract.get("management_fee")) else None,
            "vat_note": real_estate_management_vat_flag(contract.get("vat_note")),
        })
    return rows


def write_real_estate_excel_with_powershell_com(db: dict[str, Any]) -> dict[str, Any]:
    payload = {
        "path": str(REAL_ESTATE_EXCEL_PATH.resolve()),
        "sheet": "상가 임대 내역",
        "management_vat_header": "관리비 부가세 적용 여부",
        "rows": real_estate_excel_rows_payload(db),
    }
    script = r'''
param([string]$PayloadPath)
$ErrorActionPreference = "Stop"
$payload = Get-Content -LiteralPath $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json

function Unit-Text($value) {
  if ($null -eq $value) { return "" }
  $text = ([string]$value).Trim()
  if ($text -match "^(\d+)\.0+$") { return $Matches[1] }
  return $text
}

function Set-Cell($sheet, [int]$row, [int]$col, $value) {
  if ($null -eq $value) {
    $sheet.Cells.Item($row, $col).Value = $null
  } else {
    $sheet.Cells.Item($row, $col).Value = [string]$value
  }
}

$target = [System.IO.Path]::GetFullPath([string]$payload.path)
$excel = $null
$workbook = $null
$openedByApp = $false
$createdExcel = $false

try {
  try { $excel = [Runtime.InteropServices.Marshal]::GetActiveObject("Excel.Application") } catch { $excel = $null }
  if ($null -ne $excel) {
    foreach ($candidate in @($excel.Workbooks)) {
      $candidatePath = [System.IO.Path]::GetFullPath([string]$candidate.FullName)
      if ([string]::Equals($candidatePath, $target, [System.StringComparison]::OrdinalIgnoreCase)) {
        $workbook = $candidate
        break
      }
    }
  }
  if ($null -eq $workbook) {
    $excel = New-Object -ComObject Excel.Application
    $createdExcel = $true
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($target, 0, $false)
    $openedByApp = $true
  }

  $sheet = $workbook.Worksheets.Item([string]$payload.sheet)
  Set-Cell $sheet 2 29 $payload.management_vat_header
  $rowByUnit = @{}
  $lastRow = $sheet.UsedRange.Rows.Count
  for ($r = 4; $r -le $lastRow; $r++) {
    $unitId = Unit-Text $sheet.Cells.Item($r, 2).Text
    if ($unitId) { $rowByUnit[$unitId] = $r }
  }

  foreach ($item in @($payload.rows)) {
    $unitId = Unit-Text $item.unit_id
    if (!$rowByUnit.ContainsKey($unitId)) { continue }
    $r = [int]$rowByUnit[$unitId]
    Set-Cell $sheet $r 3 $item.area_m2
    Set-Cell $sheet $r 4 $item.area_py
    Set-Cell $sheet $r 5 $item.plan_deposit
    Set-Cell $sheet $r 6 $item.plan_rent
    Set-Cell $sheet $r 7 $item.plan_management_fee
    Set-Cell $sheet $r 8 $item.tenant_business
    Set-Cell $sheet $r 9 $item.tenant
    Set-Cell $sheet $r 10 $item.registration_no
    Set-Cell $sheet $r 11 $item.phone
    Set-Cell $sheet $r 12 $item.address
    Set-Cell $sheet $r 13 $item.contract_date
    Set-Cell $sheet $r 14 $item.balance_date
    Set-Cell $sheet $r 15 $item.lease_end_date
    Set-Cell $sheet $r 16 $item.lease_term
    Set-Cell $sheet $r 17 $item.rent_payment_day
    Set-Cell $sheet $r 18 $item.rent_free_months
    Set-Cell $sheet $r 19 $item.rent_free_note
    Set-Cell $sheet $r 20 $item.settlement_support_months
    Set-Cell $sheet $r 21 $item.settlement_support_note
    Set-Cell $sheet $r 24 $item.deposit
    Set-Cell $sheet $r 25 $item.contract_deposit
    Set-Cell $sheet $r 26 $item.balance_amount
    Set-Cell $sheet $r 27 $item.monthly_rent
    Set-Cell $sheet $r 28 $item.management_fee
    Set-Cell $sheet $r 29 $item.vat_note
  }

  $workbook.Save()
  [pscustomobject]@{ ok = $true; message = "열려 있는 엑셀 파일도 업데이트했습니다."; path = $target } | ConvertTo-Json -Compress
} finally {
  if ($openedByApp -and $null -ne $workbook) { $workbook.Close($true) }
  if ($createdExcel -and $null -ne $excel) { $excel.Quit() }
}
'''
    payload_path = None
    script_path = None
    try:
        with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as payload_file:
            json.dump(payload, payload_file, ensure_ascii=True)
            payload_path = Path(payload_file.name)
        with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as script_file:
            script_file.write(script)
            script_path = Path(script_file.name)
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(script_path), str(payload_path)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )
        if result.returncode != 0:
            return {"ok": False, "message": (result.stderr or result.stdout or "Excel COM 저장 실패").strip(), "path": str(REAL_ESTATE_EXCEL_PATH)}
        parsed = json.loads(result.stdout.strip() or "{}")
        if parsed.get("ok"):
            parsed["message"] = "열려 있는 엑셀 파일도 업데이트했습니다."
        return parsed
    except Exception as exc:
        return {"ok": False, "message": str(exc), "path": str(REAL_ESTATE_EXCEL_PATH)}
    finally:
        for path in (payload_path, script_path):
            if path:
                try:
                    path.unlink(missing_ok=True)
                except Exception:
                    pass


def write_real_estate_excel(db: dict[str, Any]) -> dict[str, Any]:
    if not REAL_ESTATE_EXCEL_PATH.exists():
        return {"ok": False, "message": "상가 관리 엑셀 파일을 찾지 못했습니다.", "path": str(REAL_ESTATE_EXCEL_PATH)}
    workbook = None
    try:
        workbook = load_workbook(REAL_ESTATE_EXCEL_PATH)
        sheet = workbook["상가 임대 내역"] if "상가 임대 내역" in workbook.sheetnames else workbook.active
        sheet.cell(row=2, column=29).value = "관리비 부가세 적용 여부"
        row_by_unit: dict[str, int] = {}
        for row_index in range(4, sheet.max_row + 1):
            unit_id = real_estate_unit_id(sheet.cell(row=row_index, column=2).value)
            if unit_id:
                row_by_unit[unit_id] = row_index
        for unit_id, unit in (db.get("units") or {}).items():
            row_index = row_by_unit.get(unit_id)
            if not row_index:
                continue
            contract = unit.get("contract") or {}
            rent_plan = unit.get("rent_plan") or {}
            area = unit.get("area") or {}
            sheet.cell(row=row_index, column=3).value = f"{area.get('exclusive_m2', 0)}/{area.get('parking_m2', 0)}/{area.get('common_m2', 0)}"
            sheet.cell(row=row_index, column=4).value = f"{area.get('exclusive_py', 0)}/{area.get('sale_py', 0)}"
            sheet.cell(row=row_index, column=5).value = real_estate_number(rent_plan.get("deposit_manwon"))
            sheet.cell(row=row_index, column=6).value = real_estate_number(rent_plan.get("monthly_rent_manwon"))
            sheet.cell(row=row_index, column=7).value = real_estate_number(rent_plan.get("management_fee_manwon"))
            sheet.cell(row=row_index, column=8).value = contract.get("tenant_business") or ""
            sheet.cell(row=row_index, column=9).value = contract.get("tenant") or ""
            sheet.cell(row=row_index, column=10).value = contract.get("registration_no") or ""
            sheet.cell(row=row_index, column=11).value = contract.get("phone") or ""
            sheet.cell(row=row_index, column=12).value = contract.get("address") or ""
            sheet.cell(row=row_index, column=13).value = real_estate_excel_date(contract.get("contract_date"))
            sheet.cell(row=row_index, column=14).value = real_estate_excel_date(contract.get("balance_date"))
            sheet.cell(row=row_index, column=15).value = real_estate_excel_date(contract.get("lease_end_date"))
            sheet.cell(row=row_index, column=16).value = contract.get("lease_term") or ""
            sheet.cell(row=row_index, column=17).value = contract.get("rent_payment_day") or ""
            sheet.cell(row=row_index, column=18).value = int(real_estate_number(contract.get("rent_free_months")))
            sheet.cell(row=row_index, column=19).value = contract.get("rent_free_note") or ""
            sheet.cell(row=row_index, column=20).value = int(real_estate_number(contract.get("settlement_support_months")))
            sheet.cell(row=row_index, column=21).value = contract.get("settlement_support_note") or ""
            sheet.cell(row=row_index, column=24).value = real_estate_number(contract.get("deposit")) / 10000 if real_estate_number(contract.get("deposit")) else None
            sheet.cell(row=row_index, column=25).value = real_estate_number(contract.get("contract_deposit")) / 10000 if real_estate_number(contract.get("contract_deposit")) else None
            sheet.cell(row=row_index, column=26).value = real_estate_number(contract.get("balance_amount")) / 10000 if real_estate_number(contract.get("balance_amount")) else None
            sheet.cell(row=row_index, column=27).value = real_estate_number(contract.get("monthly_rent")) / 10000 if real_estate_number(contract.get("monthly_rent")) else None
            sheet.cell(row=row_index, column=28).value = real_estate_number(contract.get("management_fee")) / 10000 if real_estate_number(contract.get("management_fee")) else None
            sheet.cell(row=row_index, column=29).value = contract.get("vat_note") or ""
        workbook.save(REAL_ESTATE_EXCEL_PATH)
        return {"ok": True, "message": "엑셀 파일도 업데이트했습니다.", "path": str(REAL_ESTATE_EXCEL_PATH)}
    except PermissionError:
        return write_real_estate_excel_with_powershell_com(db)
    except Exception as exc:
        return {"ok": False, "message": str(exc), "path": str(REAL_ESTATE_EXCEL_PATH)}
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def real_estate_vat_note_rows_payload(db: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for unit_id, unit in (db.get("units") or {}).items():
        contract = unit.get("contract") or {}
        rows.append({
            "unit_id": str(unit_id),
            "vat_note": contract.get("vat_note") or "",
        })
    return rows


def write_real_estate_vat_notes_with_powershell_com(db: dict[str, Any]) -> dict[str, Any]:
    payload = {
        "path": str(REAL_ESTATE_EXCEL_PATH.resolve()),
        "sheet": "상가 임대 내역",
        "rows": real_estate_vat_note_rows_payload(db),
    }
    script = r'''
param([string]$PayloadPath)
$ErrorActionPreference = "Stop"
$payload = Get-Content -LiteralPath $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json

function Unit-Text($value) {
  if ($null -eq $value) { return "" }
  $text = ([string]$value).Trim()
  if ($text -match "^(\d+)\.0+$") { return $Matches[1] }
  return $text
}

$target = [System.IO.Path]::GetFullPath([string]$payload.path)
$excel = $null
$workbook = $null
$openedByApp = $false
$createdExcel = $false

try {
  try { $excel = [Runtime.InteropServices.Marshal]::GetActiveObject("Excel.Application") } catch { $excel = $null }
  if ($null -ne $excel) {
    foreach ($candidate in @($excel.Workbooks)) {
      $candidatePath = [System.IO.Path]::GetFullPath([string]$candidate.FullName)
      if ([string]::Equals($candidatePath, $target, [System.StringComparison]::OrdinalIgnoreCase)) {
        $workbook = $candidate
        break
      }
    }
  }
  if ($null -eq $workbook) {
    $excel = New-Object -ComObject Excel.Application
    $createdExcel = $true
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($target, 0, $false)
    $openedByApp = $true
  }

  $sheet = $workbook.Worksheets.Item([string]$payload.sheet)
  $rowByUnit = @{}
  $lastRow = $sheet.UsedRange.Rows.Count
  for ($r = 4; $r -le $lastRow; $r++) {
    $unitId = Unit-Text $sheet.Cells.Item($r, 2).Text
    if ($unitId) { $rowByUnit[$unitId] = $r }
  }

  foreach ($item in @($payload.rows)) {
    $unitId = Unit-Text $item.unit_id
    if (!$rowByUnit.ContainsKey($unitId)) { continue }
    $r = [int]$rowByUnit[$unitId]
    $sheet.Cells.Item($r, 29).Value = [string]$item.vat_note
  }

  $workbook.Save()
  [pscustomobject]@{ ok = $true; message = "부가세 적용 여부를 엑셀 AC열에 기록했습니다."; path = $target } | ConvertTo-Json -Compress
} finally {
  if ($openedByApp -and $null -ne $workbook) { $workbook.Close($true) }
  if ($createdExcel -and $null -ne $excel) { $excel.Quit() }
}
'''
    payload_path = None
    script_path = None
    try:
        with tempfile.NamedTemporaryFile("w", suffix=".json", delete=False, encoding="utf-8") as payload_file:
            json.dump(payload, payload_file, ensure_ascii=True)
            payload_path = Path(payload_file.name)
        with tempfile.NamedTemporaryFile("w", suffix=".ps1", delete=False, encoding="utf-8") as script_file:
            script_file.write(script)
            script_path = Path(script_file.name)
        result = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", str(script_path), str(payload_path)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=30,
        )
        if result.returncode != 0:
            return {"ok": False, "message": (result.stderr or result.stdout or "Excel COM AC열 저장 실패").strip(), "path": str(REAL_ESTATE_EXCEL_PATH)}
        return json.loads(result.stdout.strip() or "{}")
    except Exception as exc:
        return {"ok": False, "message": str(exc), "path": str(REAL_ESTATE_EXCEL_PATH)}
    finally:
        for path in (payload_path, script_path):
            if path:
                try:
                    path.unlink(missing_ok=True)
                except Exception:
                    pass


def write_real_estate_vat_notes_excel(db: dict[str, Any]) -> dict[str, Any]:
    if not REAL_ESTATE_EXCEL_PATH.exists():
        return {"ok": False, "message": "상가 관리 엑셀 파일을 찾지 못했습니다.", "path": str(REAL_ESTATE_EXCEL_PATH)}
    workbook = None
    try:
        workbook = load_workbook(REAL_ESTATE_EXCEL_PATH)
        sheet = workbook["상가 임대 내역"] if "상가 임대 내역" in workbook.sheetnames else workbook.active
        row_by_unit: dict[str, int] = {}
        for row_index in range(4, sheet.max_row + 1):
            unit_id = real_estate_unit_id(sheet.cell(row=row_index, column=2).value)
            if unit_id:
                row_by_unit[unit_id] = row_index
        for item in real_estate_vat_note_rows_payload(db):
            row_index = row_by_unit.get(item["unit_id"])
            if row_index:
                sheet.cell(row=row_index, column=29).value = item.get("vat_note") or ""
        workbook.save(REAL_ESTATE_EXCEL_PATH)
        return {"ok": True, "message": "부가세 적용 여부를 엑셀 AC열에 기록했습니다.", "path": str(REAL_ESTATE_EXCEL_PATH)}
    except PermissionError:
        return write_real_estate_vat_notes_with_powershell_com(db)
    except Exception as exc:
        return {"ok": False, "message": str(exc), "path": str(REAL_ESTATE_EXCEL_PATH)}
    finally:
        if workbook is not None:
            try:
                workbook.close()
            except Exception:
                pass


def load_real_estate_db() -> dict[str, Any]:
    if not REAL_ESTATE_DB_PATH.exists():
        return enrich_real_estate_special_terms(merge_real_estate_excel(default_real_estate_db()))
    try:
        return enrich_real_estate_special_terms(merge_real_estate_excel(normalize_real_estate_db(json.loads(REAL_ESTATE_DB_PATH.read_text(encoding="utf-8")))))
    except Exception:
        return enrich_real_estate_special_terms(merge_real_estate_excel(default_real_estate_db()))


def save_real_estate_db(payload: dict[str, Any]) -> dict[str, Any]:
    normalized = enrich_real_estate_special_terms(normalize_real_estate_db(payload))
    normalized["updated_at"] = datetime.now().isoformat(timespec="seconds")
    try:
        temp_path = REAL_ESTATE_DB_PATH.with_name(f"{REAL_ESTATE_DB_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
        temp_path.write_text(json.dumps(normalized, ensure_ascii=False, indent=2), encoding="utf-8")
        temp_path.replace(REAL_ESTATE_DB_PATH)
    except Exception:
        pass
    excel_sync = write_real_estate_excel(normalized)
    normalized["excel_sync"] = {
        **excel_sync,
        "message": (excel_sync.get("message") or "DB 저장 완료") + " 기존 행 구조 안에서 엑셀도 함께 업데이트했습니다.",
    }
    return normalized


REAL_ESTATE_PRICE_MAPS = [
    {
        "key": "national",
        "title": "전국",
        "subtitle": "광역 시도",
        "regions": [
            ("서울", 54, 19), ("경기", 49, 27), ("인천", 38, 24), ("강원", 67, 21),
            ("충북", 56, 39), ("충남", 43, 44), ("대전", 50, 49), ("세종", 47, 43),
            ("전북", 46, 60), ("전남", 42, 77), ("광주", 36, 70), ("경북", 68, 53),
            ("대구", 65, 61), ("울산", 75, 70), ("부산", 70, 78), ("경남", 60, 76),
            ("제주", 34, 93),
        ],
    },
    {
        "key": "seoul",
        "title": "서울",
        "subtitle": "주요 구",
        "regions": [
            ("강남", 61, 67), ("서초", 51, 70), ("송파", 73, 68), ("용산", 45, 55),
            ("마포", 32, 47), ("성동", 56, 49), ("광진", 66, 53), ("양천", 24, 66),
            ("강동", 84, 60), ("영등포", 35, 66), ("동작", 43, 72), ("관악", 42, 82),
            ("노원", 67, 21), ("도봉", 59, 15), ("강북", 52, 21), ("은평", 31, 28),
            ("종로", 45, 37), ("중구", 49, 46),
        ],
    },
    {
        "key": "gyeonggi",
        "title": "경기도",
        "subtitle": "주요 시",
        "regions": [
            ("성남", 58, 54), ("과천", 48, 58), ("하남", 66, 48), ("용인", 58, 70),
            ("수원", 48, 72), ("고양", 35, 25), ("화성", 39, 82), ("광명", 36, 55),
            ("안양", 43, 61), ("의왕", 49, 64), ("구리", 61, 38), ("남양주", 68, 35),
            ("김포", 25, 32), ("파주", 31, 15), ("평택", 48, 91), ("안산", 32, 68),
            ("시흥", 29, 62), ("이천", 73, 76), ("광주", 67, 64), ("여주", 82, 51),
        ],
    },
]
R_ONE_WEEKLY_SALE_INDEX_TABLE = "T244183132827305"
R_ONE_WEEKLY_MAX_WORKERS = 8

SEOUL_LAWD_CODES = {
    "종로구": ["11110"], "중구": ["11140"], "용산구": ["11170"], "성동구": ["11200"],
    "광진구": ["11215"], "동대문구": ["11230"], "중랑구": ["11260"], "성북구": ["11290"],
    "강북구": ["11305"], "도봉구": ["11320"], "노원구": ["11350"], "은평구": ["11380"],
    "서대문구": ["11410"], "마포구": ["11440"], "양천구": ["11470"], "강서구": ["11500"],
    "구로구": ["11530"], "금천구": ["11545"], "영등포구": ["11560"], "동작구": ["11590"],
    "관악구": ["11620"], "서초구": ["11650"], "강남구": ["11680"], "송파구": ["11710"],
    "강동구": ["11740"],
}

GYEONGGI_LAWD_CODES = {
    "수원시": ["41111", "41113", "41115", "41117"],
    "장안구": ["41111"], "권선구": ["41113"], "팔달구": ["41115"], "영통구": ["41117"],
    "성남시": ["41131", "41133", "41135"],
    "수정구": ["41131"], "중원구": ["41133"], "분당구": ["41135"],
    "의정부시": ["41150"],
    "안양시": ["41171", "41173"], "만안구": ["41171"], "동안구": ["41173"],
    "부천시": ["41190"], "광명시": ["41210"], "평택시": ["41220"], "동두천시": ["41250"],
    "안산시": ["41271", "41273"], "상록구": ["41271"], "단원구": ["41273"],
    "고양시": ["41281", "41285", "41287"], "덕양구": ["41281"], "일산동구": ["41285"], "일산서구": ["41287"],
    "과천시": ["41290"], "구리시": ["41310"], "남양주시": ["41360"], "오산시": ["41370"],
    "시흥시": ["41390"], "군포시": ["41410"], "의왕시": ["41430"], "하남시": ["41450"],
    "용인시": ["41461", "41463", "41465"], "처인구": ["41461"], "기흥구": ["41463"], "수지구": ["41465"],
    "파주시": ["41480"], "이천시": ["41500"], "안성시": ["41550"], "김포시": ["41570"],
    "화성시": ["41590"], "광주시": ["41610"], "양주시": ["41630"], "포천시": ["41650"],
    "여주시": ["41670"], "연천군": ["41800"], "가평군": ["41820"], "양평군": ["41830"],
}

SEOUL_REGION_COORDS = {
    "종로구": (47, 37), "중구": (50, 45), "용산구": (45, 55), "성동구": (56, 49),
    "광진구": (66, 53), "동대문구": (58, 38), "중랑구": (70, 35), "성북구": (49, 29),
    "강북구": (52, 21), "도봉구": (59, 15), "노원구": (67, 21), "은평구": (31, 28),
    "서대문구": (36, 42), "마포구": (32, 47), "양천구": (24, 66), "강서구": (17, 55),
    "구로구": (27, 76), "금천구": (34, 84), "영등포구": (35, 66), "동작구": (43, 72),
    "관악구": (42, 82), "서초구": (51, 70), "강남구": (61, 67), "송파구": (73, 68),
    "강동구": (84, 60),
}

GYEONGGI_REGION_COORDS = {
    "연천군": (47, 4), "동두천시": (50, 10), "포천시": (63, 12), "파주시": (31, 15),
    "양주시": (46, 17), "의정부시": (54, 20), "가평군": (82, 24), "고양시": (35, 25),
    "김포시": (24, 32), "남양주시": (68, 35), "구리시": (61, 38), "부천시": (27, 51),
    "광명시": (36, 55), "하남시": (66, 48), "성남시": (58, 54), "과천시": (48, 58),
    "안양시": (43, 61), "의왕시": (49, 64), "군포시": (41, 66), "시흥시": (29, 62),
    "안산시": (32, 68), "수원시": (48, 72), "용인시": (58, 70), "광주시": (67, 64),
    "양평군": (84, 52), "화성시": (39, 82), "오산시": (46, 80), "평택시": (48, 91),
    "안성시": (59, 91), "이천시": (73, 76), "여주시": (84, 75),
    "장안구": (46, 69), "권선구": (44, 75), "팔달구": (49, 73), "영통구": (53, 73),
    "수정구": (55, 52), "중원구": (59, 54), "분당구": (61, 59),
    "만안구": (41, 60), "동안구": (45, 62), "상록구": (31, 66), "단원구": (29, 71),
    "덕양구": (35, 22), "일산동구": (32, 27), "일산서구": (28, 28),
    "처인구": (62, 77), "기흥구": (56, 68), "수지구": (53, 63),
}


def get_r_one_api_key() -> str:
    settings = load_settings()
    public_data = settings.get("public_data") if isinstance(settings.get("public_data"), dict) else {}
    return str(
        os.getenv("STOCK_DASHBOARD_R_ONE_API_KEY")
        or public_data.get("r_one_api_key")
        or settings.get("r_one_api_key")
        or ""
    ).strip()


def r_one_get_json(params: dict[str, Any], timeout: int = 30) -> dict[str, Any]:
    base_url = "https://www.reb.or.kr/r-one/openapi/SttsApiTblData.do"
    response = requests.get(base_url, params=params, timeout=timeout)
    response.raise_for_status()
    data = response.json()
    if isinstance(data, dict) and "RESULT" in data:
        result = data.get("RESULT") or {}
        raise RuntimeError(str(result.get("MESSAGE") or result.get("CODE") or "R-ONE API 오류"))
    return data


def r_one_extract_rows(payload: dict[str, Any]) -> tuple[int, list[dict[str, Any]]]:
    container = payload.get("SttsApiTblData")
    if not isinstance(container, list) or len(container) < 2:
        return 0, []
    total = 0
    head = container[0].get("head") if isinstance(container[0], dict) else []
    if isinstance(head, list) and head:
        try:
            total = int((head[0] or {}).get("list_total_count") or 0)
        except Exception:
            total = 0
    rows = container[1].get("row") if isinstance(container[1], dict) else []
    if isinstance(rows, dict):
        rows = [rows]
    return total, rows if isinstance(rows, list) else []


def fetch_r_one_weekly_sale_index_rows() -> list[dict[str, Any]]:
    api_key = get_r_one_api_key()
    if not api_key:
        raise RuntimeError("R-ONE 인증키가 설정되어 있지 않습니다.")
    page_size = 1000
    base_params = {
        "KEY": api_key,
        "Type": "json",
        "STATBL_ID": R_ONE_WEEKLY_SALE_INDEX_TABLE,
        "DTACYCLE_CD": "WK",
        "pSize": page_size,
    }
    first_payload = r_one_get_json({**base_params, "pIndex": 1})
    total_count, first_rows = r_one_extract_rows(first_payload)
    if total_count <= 0:
        return first_rows
    total_pages = max(1, math.ceil(total_count / page_size))
    latest_payload = first_payload if total_pages == 1 else r_one_get_json({**base_params, "pIndex": total_pages})
    _, latest_page_rows = r_one_extract_rows(latest_payload)
    latest_identifier = max((str(row.get("WRTTIME_IDTFR_ID") or "") for row in latest_page_rows), default="")
    if latest_identifier:
        latest_payload = r_one_get_json({**base_params, "WRTTIME_IDTFR_ID": latest_identifier, "pIndex": 1})
        _, latest_rows = r_one_extract_rows(latest_payload)
    else:
        latest_date = max((str(row.get("WRTTIME_DESC") or "") for row in latest_page_rows), default="")
        latest_rows = [row for row in latest_page_rows if str(row.get("WRTTIME_DESC") or "") == latest_date] or latest_page_rows
    target_cls_ids: set[str] = set()
    for map_info in REAL_ESTATE_PRICE_MAPS:
        for row in real_estate_region_display_rows(latest_rows, map_info["key"]):
            if row.get("CLS_ID") is not None:
                target_cls_ids.add(str(row.get("CLS_ID")))
    if not target_cls_ids:
        return latest_page_rows

    def fetch_region(cls_id: str) -> list[dict[str, Any]]:
        payload = r_one_get_json({**base_params, "CLS_ID": cls_id, "pIndex": 1})
        _, region_rows = r_one_extract_rows(payload)
        return region_rows

    rows: list[dict[str, Any]] = []
    with ThreadPoolExecutor(max_workers=R_ONE_WEEKLY_MAX_WORKERS) as executor:
        futures = [executor.submit(fetch_region, cls_id) for cls_id in sorted(target_cls_ids)]
        for future in as_completed(futures):
            rows.extend(future.result())
    return rows


def normalize_real_estate_region_name(value: Any) -> str:
    text = normalize_text(str(value or ""))
    for suffix in ("특별시", "광역시", "특별자치시", "특별자치도", "자치도", "시", "군", "구"):
        if text.endswith(suffix):
            text = text[: -len(suffix)]
            break
    return text


def r_one_region_matches(map_key: str, target_name: str, row: dict[str, Any]) -> bool:
    cls_name = str(row.get("CLS_NM") or "").strip()
    full_name = str(row.get("CLS_FULLNM") or cls_name).strip()
    normalized_target = normalize_real_estate_region_name(target_name)
    normalized_cls = normalize_real_estate_region_name(cls_name)
    if normalized_target != normalized_cls:
        return False
    if map_key == "national":
        return ">" not in full_name
    if map_key == "seoul":
        return full_name.startswith("서울>")
    if map_key == "gyeonggi":
        return full_name.startswith("경기>")
    return True


def real_estate_region_path(row: dict[str, Any]) -> list[str]:
    full_name = str(row.get("CLS_FULLNM") or row.get("CLS_NM") or "").strip()
    return [part.strip() for part in full_name.split(">") if part.strip()]


def real_estate_region_display_rows(latest_rows: list[dict[str, Any]], map_key: str) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in latest_rows:
        if str(row.get("ITM_NM") or "").strip() != "지수":
            continue
        cls_id = str(row.get("CLS_ID") or "")
        if not cls_id or cls_id in seen:
            continue
        path = real_estate_region_path(row)
        name = str(row.get("CLS_NM") or "").strip()
        if map_key == "national":
            include = len(path) == 1 and name != "전국"
        elif map_key == "seoul":
            include = len(path) >= 2 and path[0] == "서울" and name.endswith("구")
        elif map_key == "gyeonggi":
            include = len(path) >= 2 and path[0] == "경기" and (name.endswith("시") or name.endswith("군") or name.endswith("구"))
        else:
            include = False
        if include:
            selected.append(row)
            seen.add(cls_id)
    selected.sort(key=lambda item: str(item.get("CLS_FULLNM") or item.get("CLS_NM") or ""))
    return selected


def real_estate_region_position(map_key: str, name: str, index: int, total: int) -> tuple[float, float]:
    if map_key == "seoul" and name in SEOUL_REGION_COORDS:
        return SEOUL_REGION_COORDS[name]
    if map_key == "gyeonggi" and name in GYEONGGI_REGION_COORDS:
        return GYEONGGI_REGION_COORDS[name]
    for map_info in REAL_ESTATE_PRICE_MAPS:
        if map_info["key"] == map_key:
            for candidate, x, y in map_info["regions"]:
                if candidate == name or normalize_real_estate_region_name(candidate) == normalize_real_estate_region_name(name):
                    return float(x), float(y)
    if total <= 1:
        return 50.0, 50.0
    columns = max(1, math.ceil(math.sqrt(total)))
    rows = max(1, math.ceil(total / columns))
    col = index % columns
    row = index // columns
    x = 10.0 + (80.0 * col / max(1, columns - 1))
    y = 12.0 + (76.0 * row / max(1, rows - 1))
    return round(x, 2), round(y, 2)


def build_real_estate_price_series_for_cls_id(cls_id: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    points: dict[str, dict[str, Any]] = {}
    for row in rows:
        if str(row.get("ITM_NM") or "").strip() != "지수":
            continue
        if str(row.get("CLS_ID") or "") != str(cls_id):
            continue
        date_text = str(row.get("WRTTIME_DESC") or "").strip()
        try:
            index_value = float(row.get("DTA_VAL"))
        except Exception:
            continue
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_text):
            continue
        points[date_text] = {"date": date_text, "index": round(index_value, 2)}
    series = [points[key] for key in sorted(points.keys())][-110:]
    for index, item in enumerate(series):
        previous = series[index - 1]["index"] if index else None
        year_ago = series[index - 52]["index"] if index >= 52 else None
        item["wow_pct"] = round(((item["index"] / previous) - 1.0) * 100.0, 2) if previous else 0.0
        item["yoy_pct"] = round(((item["index"] / year_ago) - 1.0) * 100.0, 2) if year_ago else None
    return series


def real_estate_latest_series_item(series: list[dict[str, Any]]) -> dict[str, Any]:
    return series[-1] if series else {}


def real_estate_child_region_summaries(
    parent_row: dict[str, Any],
    latest_rows: list[dict[str, Any]],
    series_by_cls: dict[str, list[dict[str, Any]]],
) -> list[dict[str, Any]]:
    parent_path = real_estate_region_path(parent_row)
    if not parent_path:
        return []
    parent_name = str(parent_row.get("CLS_NM") or "").strip()
    children: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in latest_rows:
        cls_id = str(row.get("CLS_ID") or "")
        if not cls_id or cls_id in seen or cls_id == str(parent_row.get("CLS_ID") or ""):
            continue
        path = real_estate_region_path(row)
        name = str(row.get("CLS_NM") or "").strip()
        if not path or str(row.get("ITM_NM") or "").strip() != "지수":
            continue
        include = False
        if parent_name == "서울":
            include = path[0] == "서울" and name.endswith("구")
        elif parent_name == "경기":
            include = path[0] == "경기" and (name.endswith("시") or name.endswith("군"))
        else:
            include = len(path) == len(parent_path) + 1 and path[: len(parent_path)] == parent_path
        if not include or cls_id not in series_by_cls:
            continue
        series = series_by_cls.get(cls_id) or []
        latest = real_estate_latest_series_item(series)
        children.append({
            "key": f"{cls_id}:{name}",
            "name": name,
            "full_name": str(row.get("CLS_FULLNM") or name),
            "latest_index": latest.get("index"),
            "wow_pct": latest.get("wow_pct"),
            "yoy_pct": latest.get("yoy_pct"),
        })
        seen.add(cls_id)
    children.sort(key=lambda item: (-(item.get("yoy_pct") if item.get("yoy_pct") is not None else -999), item.get("name") or ""))
    return children[:80]


def build_real_estate_price_series_from_rows(map_key: str, target_name: str, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    points: dict[str, dict[str, Any]] = {}
    for row in rows:
        if str(row.get("ITM_NM") or "").strip() != "지수":
            continue
        if not r_one_region_matches(map_key, target_name, row):
            continue
        date_text = str(row.get("WRTTIME_DESC") or "").strip()
        try:
            index_value = float(row.get("DTA_VAL"))
        except Exception:
            continue
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", date_text):
            continue
        points[date_text] = {
            "date": date_text,
            "index": round(index_value, 2),
        }
    series = [points[key] for key in sorted(points.keys())][-110:]
    for index, item in enumerate(series):
        previous = series[index - 1]["index"] if index else None
        year_ago = series[index - 52]["index"] if index >= 52 else None
        item["wow_pct"] = round(((item["index"] / previous) - 1.0) * 100.0, 2) if previous else 0.0
        item["yoy_pct"] = round(((item["index"] / year_ago) - 1.0) * 100.0, 2) if year_ago else None
    return series


def build_real_estate_price_payload_from_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    latest_date = max((str(row.get("WRTTIME_DESC") or "") for row in rows if str(row.get("WRTTIME_DESC") or "")), default="")
    latest_rows = [row for row in rows if str(row.get("WRTTIME_DESC") or "") == latest_date] if latest_date else rows
    series_by_cls: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        cls_id = str(row.get("CLS_ID") or "")
        if cls_id and cls_id not in series_by_cls:
            series_by_cls[cls_id] = build_real_estate_price_series_for_cls_id(cls_id, rows)
    maps = []
    missing_regions = []
    for map_info in REAL_ESTATE_PRICE_MAPS:
        regions = []
        display_rows = real_estate_region_display_rows(latest_rows, map_info["key"])
        if not display_rows:
            display_rows = [
                {"CLS_NM": name, "CLS_FULLNM": name, "CLS_ID": f"fallback:{map_info['key']}:{name}"}
                for name, _, _ in map_info["regions"]
            ]
        for index, row in enumerate(display_rows):
            name = str(row.get("CLS_NM") or "").strip()
            full_name = str(row.get("CLS_FULLNM") or name).strip()
            cls_id = str(row.get("CLS_ID") or "")
            x, y = real_estate_region_position(map_info["key"], name, index, len(display_rows))
            series = series_by_cls.get(cls_id) or build_real_estate_price_series_from_rows(map_info["key"], name, rows)
            if not series:
                missing_regions.append(f"{map_info['title']}:{name}")
                series = build_real_estate_price_series(name, map_info["key"])
            latest = series[-1] if series else {}
            regions.append({
                "key": f"{map_info['key']}:{name}",
                "name": name,
                "full_name": full_name,
                "cls_id": cls_id,
                "x": x,
                "y": y,
                "latest_index": latest.get("index"),
                "wow_pct": latest.get("wow_pct"),
                "yoy_pct": latest.get("yoy_pct"),
                "series": series,
                "children": real_estate_child_region_summaries(row, latest_rows, series_by_cls),
            })
        maps.append({
            "key": map_info["key"],
            "title": map_info["title"],
            "subtitle": map_info["subtitle"],
            "regions": regions,
        })
    payload_latest_date = ""
    for map_info in maps:
        for region in map_info["regions"]:
            series = region.get("series") if isinstance(region.get("series"), list) else []
            if series and str(series[-1].get("date") or "") > payload_latest_date:
                payload_latest_date = str(series[-1].get("date") or "")
    return {
        "version": 3,
        "source": "한국부동산원 R-ONE 주간 매매가격지수",
        "table_id": R_ONE_WEEKLY_SALE_INDEX_TABLE,
        "as_of": payload_latest_date or date.today().isoformat(),
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "detail_note": "지역지수는 R-ONE 기준이며, 동·단지별 상승 주도는 국토교통부 실거래가 상세 API를 함께 조회해 추정합니다.",
        "maps": maps,
        "missing_regions": missing_regions,
    }


def real_estate_price_seed(text: str) -> int:
    return sum((index + 1) * ord(char) for index, char in enumerate(text))


def real_estate_price_week_dates(weeks: int = 110) -> list[date]:
    today = date.today()
    last_monday = today - timedelta(days=today.weekday())
    start = last_monday - timedelta(weeks=weeks - 1)
    return [start + timedelta(weeks=index) for index in range(weeks)]


def build_real_estate_price_series(name: str, map_key: str, weeks: int = 110) -> list[dict[str, Any]]:
    seed = real_estate_price_seed(map_key + ":" + name)
    dates = real_estate_price_week_dates(weeks)
    index_value = 92.0 + (seed % 24) * 0.55
    base_bias = ((seed % 13) - 5) * 0.006
    cycle_shift = (seed % 31) / 5.0
    series: list[dict[str, Any]] = []
    for idx, current_date in enumerate(dates):
        cycle = math.sin((idx / 8.0) + cycle_shift) * 0.045
        long_cycle = math.cos((idx / 23.0) + cycle_shift) * 0.025
        recent_lift = max(0.0, idx - weeks + 22) * (0.002 + (seed % 5) * 0.00045)
        weekly_pct = base_bias + cycle + long_cycle + recent_lift
        index_value = max(70.0, index_value * (1.0 + weekly_pct / 100.0))
        series.append({
            "date": current_date.isoformat(),
            "index": round(index_value, 2),
        })
    for idx, item in enumerate(series):
        previous = series[idx - 1]["index"] if idx else None
        year_ago = series[idx - 52]["index"] if idx >= 52 else None
        item["wow_pct"] = round(((item["index"] / previous) - 1.0) * 100.0, 2) if previous else 0.0
        item["yoy_pct"] = round(((item["index"] / year_ago) - 1.0) * 100.0, 2) if year_ago else None
    return series


def build_real_estate_price_payload(force_refresh: bool = False) -> dict[str, Any]:
    if not force_refresh and REAL_ESTATE_PRICE_CACHE_PATH.exists():
        try:
            cached = json.loads(REAL_ESTATE_PRICE_CACHE_PATH.read_text(encoding="utf-8"))
            fetched_at = parse_iso_datetime(cached.get("fetched_at")) if cached.get("fetched_at") else None
            cache_age = (datetime.now() - fetched_at).total_seconds() if fetched_at else 0
            if cached.get("version") == 3 and fetched_at and cache_age < 60 * 60 * 12:
                return cached
        except Exception:
            pass

    try:
        rows = fetch_r_one_weekly_sale_index_rows()
        payload = build_real_estate_price_payload_from_rows(rows)
        REAL_ESTATE_PRICE_CACHE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        return payload
    except Exception as exc:
        if REAL_ESTATE_PRICE_CACHE_PATH.exists():
            try:
                cached = json.loads(REAL_ESTATE_PRICE_CACHE_PATH.read_text(encoding="utf-8"))
                cached["warning"] = f"R-ONE 최신 데이터 갱신 실패: {exc}"
                return cached
            except Exception:
                pass

    maps = []
    for map_info in REAL_ESTATE_PRICE_MAPS:
        regions = []
        for name, x, y in map_info["regions"]:
            series = build_real_estate_price_series(name, map_info["key"])
            latest = series[-1] if series else {}
            regions.append({
                "key": f"{map_info['key']}:{name}",
                "name": name,
                "x": x,
                "y": y,
                "latest_index": latest.get("index"),
                "wow_pct": latest.get("wow_pct"),
                "yoy_pct": latest.get("yoy_pct"),
                "series": series,
            })
        maps.append({
            "key": map_info["key"],
            "title": map_info["title"],
            "subtitle": map_info["subtitle"],
            "regions": regions,
        })

    payload = {
        "version": 1,
        "source": "한국부동산원 R-ONE 예비 데이터",
        "note": "R-ONE API 호출에 실패해 화면 확인용 로컬 캐시 지수를 사용합니다.",
        "as_of": date.today().isoformat(),
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "maps": maps,
    }
    try:
        REAL_ESTATE_PRICE_CACHE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass
    return payload


def get_molit_real_estate_api_key() -> str:
    settings = load_settings()
    public_data = settings.get("public_data") if isinstance(settings.get("public_data"), dict) else {}
    return str(
        os.getenv("STOCK_DASHBOARD_MOLIT_APT_TRADE_API_KEY")
        or public_data.get("apartment_trade_api_key")
        or public_data.get("molit_api_key")
        or public_data.get("r_one_api_key")
        or settings.get("apartment_trade_api_key")
        or settings.get("molit_api_key")
        or ""
    ).strip()


def real_estate_region_lawd_codes(region_name: str, full_name: str = "") -> list[str]:
    name = str(region_name or "").strip()
    full = str(full_name or "").strip()
    candidates = [name]
    if name and not name.endswith(("시", "군", "구")):
        candidates.extend([name + "시", name + "군", name + "구"])
    for part in real_estate_region_path({"CLS_FULLNM": full}):
        candidates.append(part)
        if part and not part.endswith(("시", "군", "구")):
            candidates.extend([part + "시", part + "군", part + "구"])
    codes: list[str] = []
    for candidate in candidates:
        for code in SEOUL_LAWD_CODES.get(candidate, []) + GYEONGGI_LAWD_CODES.get(candidate, []):
            if code not in codes:
                codes.append(code)
    return codes


def real_estate_recent_months(month_count: int = 8) -> list[str]:
    today = date.today().replace(day=1)
    months: list[str] = []
    year = today.year
    month = today.month
    for _ in range(month_count):
        months.append(f"{year:04d}{month:02d}")
        month -= 1
        if month <= 0:
            month = 12
            year -= 1
    return months


def real_estate_trade_cache_key(region_name: str, full_name: str, lawd_codes: list[str]) -> str:
    raw = "|".join([region_name, full_name, ",".join(lawd_codes), date.today().strftime("%Y%m")])
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def load_real_estate_trade_cache() -> dict[str, Any]:
    try:
        if REAL_ESTATE_TRADE_CACHE_PATH.exists():
            payload = json.loads(REAL_ESTATE_TRADE_CACHE_PATH.read_text(encoding="utf-8"))
            return payload if isinstance(payload, dict) else {}
    except Exception:
        pass
    return {}


def save_real_estate_trade_cache(payload: dict[str, Any]) -> None:
    try:
        REAL_ESTATE_TRADE_CACHE_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


def parse_molit_trade_items(xml_text: str) -> list[dict[str, Any]]:
    try:
        root = ET.fromstring(xml_text)
    except Exception:
        return []
    header_code = root.findtext(".//header/resultCode")
    header_message = root.findtext(".//header/resultMsg")
    if header_code and header_code not in {"00", "000"}:
        raise RuntimeError(header_message or f"국토부 실거래 API 오류({header_code})")
    items = []
    for item in root.findall(".//item"):
        row = {child.tag: (child.text or "").strip() for child in list(item)}
        items.append(row)
    return items


def molit_number(value: Any) -> float:
    try:
        return float(str(value or "").replace(",", "").replace(" ", ""))
    except Exception:
        return 0.0


def normalize_molit_trade_row(row: dict[str, Any], lawd_code: str, deal_ym: str) -> dict[str, Any] | None:
    amount = molit_number(row.get("dealAmount"))
    area = molit_number(row.get("excluUseAr"))
    if amount <= 0 or area <= 0:
        return None
    year = str(row.get("dealYear") or deal_ym[:4])
    month = str(row.get("dealMonth") or deal_ym[4:]).zfill(2)
    day = str(row.get("dealDay") or "1").zfill(2)
    deal_date = f"{year}-{month}-{day}"
    apt_name = str(row.get("aptNm") or row.get("aptNm") or row.get("aptName") or row.get("mhouseNm") or "-").strip()
    dong = str(row.get("umdNm") or row.get("umdName") or row.get("dong") or "-").strip()
    return {
        "lawd_code": lawd_code,
        "deal_ym": deal_ym,
        "deal_date": deal_date,
        "dong": dong or "-",
        "apartment": apt_name or "-",
        "amount_manwon": amount,
        "area_m2": area,
        "price_per_m2": round(amount / area, 2),
        "floor": str(row.get("floor") or "").strip(),
        "build_year": str(row.get("buildYear") or "").strip(),
    }


def fetch_molit_apartment_trades_for_month(lawd_code: str, deal_ym: str, api_key: str) -> list[dict[str, Any]]:
    endpoint = "https://apis.data.go.kr/1613000/RTMSDataSvcAptTradeDev/getRTMSDataSvcAptTradeDev"
    params = {
        "serviceKey": api_key,
        "LAWD_CD": lawd_code,
        "DEAL_YMD": deal_ym,
        "numOfRows": 1000,
        "pageNo": 1,
    }
    response = requests.get(endpoint, params=params, timeout=18)
    response.raise_for_status()
    rows = parse_molit_trade_items(response.text)
    return [normalized for row in rows if (normalized := normalize_molit_trade_row(row, lawd_code, deal_ym))]


def median_number(values: list[float]) -> float | None:
    clean = sorted(float(value) for value in values if value is not None and math.isfinite(float(value)))
    if not clean:
        return None
    mid = len(clean) // 2
    if len(clean) % 2:
        return clean[mid]
    return (clean[mid - 1] + clean[mid]) / 2.0


def summarize_molit_leaders(trades: list[dict[str, Any]]) -> dict[str, Any]:
    if not trades:
        return {"dong_leaders": [], "apartment_leaders": [], "latest_trades": []}
    months = sorted({str(row.get("deal_ym") or "") for row in trades if row.get("deal_ym")})
    recent_months = set(months[-3:])
    previous_months = set(months[-6:-3])

    def summarize_group(key_name: str, group_name: str, rows: list[dict[str, Any]]) -> dict[str, Any] | None:
        recent = [row for row in rows if row.get("deal_ym") in recent_months]
        previous = [row for row in rows if row.get("deal_ym") in previous_months]
        recent_median = median_number([float(row.get("price_per_m2") or 0) for row in recent])
        previous_median = median_number([float(row.get("price_per_m2") or 0) for row in previous])
        if recent_median is None:
            return None
        change_pct = None
        if previous_median and previous_median > 0:
            change_pct = round(((recent_median / previous_median) - 1.0) * 100.0, 2)
        latest = sorted(recent or rows, key=lambda row: str(row.get("deal_date") or ""), reverse=True)[0]
        return {
            key_name: group_name,
            "recent_price_per_m2": round(recent_median, 2),
            "previous_price_per_m2": round(previous_median, 2) if previous_median is not None else None,
            "change_pct": change_pct,
            "recent_count": len(recent),
            "previous_count": len(previous),
            "latest_amount_manwon": latest.get("amount_manwon"),
            "latest_area_m2": latest.get("area_m2"),
            "latest_date": latest.get("deal_date"),
        }

    by_dong: dict[str, list[dict[str, Any]]] = {}
    by_apartment: dict[str, list[dict[str, Any]]] = {}
    for row in trades:
        dong = str(row.get("dong") or "-")
        apt = str(row.get("apartment") or "-")
        by_dong.setdefault(dong, []).append(row)
        by_apartment.setdefault(dong + "||" + apt, []).append(row)

    dong_leaders = [item for dong, rows in by_dong.items() if (item := summarize_group("dong", dong, rows))]
    apartment_leaders = []
    for key, rows in by_apartment.items():
        dong, apt = key.split("||", 1)
        item = summarize_group("apartment", apt, rows)
        if item:
            item["dong"] = dong
            apartment_leaders.append(item)
    dong_leaders.sort(key=lambda item: (-(item.get("change_pct") if item.get("change_pct") is not None else -999), -item.get("recent_count", 0)))
    apartment_leaders.sort(key=lambda item: (-(item.get("change_pct") if item.get("change_pct") is not None else -999), -item.get("recent_count", 0)))
    latest_trades = sorted(trades, key=lambda row: str(row.get("deal_date") or ""), reverse=True)[:80]
    return {
        "dong_leaders": dong_leaders[:50],
        "apartment_leaders": apartment_leaders[:80],
        "latest_trades": latest_trades,
    }


def build_real_estate_trade_detail(region_name: str, full_name: str = "", force_refresh: bool = False) -> dict[str, Any]:
    lawd_codes = real_estate_region_lawd_codes(region_name, full_name)
    if not lawd_codes:
        return {
            "ok": False,
            "region_name": region_name,
            "full_name": full_name,
            "message": "선택 지역의 법정동코드를 아직 매칭하지 못했습니다. 서울 구·경기도 시/구를 선택하면 실거래 상세를 볼 수 있습니다.",
            "housing_types": [{"key": "apartment", "label": "아파트", "dong_leaders": [], "apartment_leaders": [], "latest_trades": []}],
        }
    if len(lawd_codes) > 8:
        return {
            "ok": False,
            "region_name": region_name,
            "full_name": full_name,
            "lawd_codes": lawd_codes,
            "message": "범위가 넓어 실거래 상세 조회를 생략했습니다. 지도에서 구/시 단위 지역을 클릭해 주세요.",
            "housing_types": [{"key": "apartment", "label": "아파트", "dong_leaders": [], "apartment_leaders": [], "latest_trades": []}],
        }
    cache_key = real_estate_trade_cache_key(region_name, full_name, lawd_codes)
    cache = load_real_estate_trade_cache()
    cached = cache.get(cache_key) if isinstance(cache.get(cache_key), dict) else None
    if not force_refresh and cached:
        fetched_at = parse_iso_datetime(cached.get("fetched_at")) if cached.get("fetched_at") else None
        if fetched_at and (datetime.now() - fetched_at).total_seconds() < 60 * 60 * 12:
            return cached
    api_key = get_molit_real_estate_api_key()
    if not api_key:
        raise RuntimeError("국토부 실거래가 API 인증키가 설정되어 있지 않습니다.")
    months = real_estate_recent_months(8)
    trades: list[dict[str, Any]] = []
    errors: list[str] = []
    for lawd_code in lawd_codes:
        for deal_ym in months:
            try:
                trades.extend(fetch_molit_apartment_trades_for_month(lawd_code, deal_ym, api_key))
            except Exception as exc:
                errors.append(f"{lawd_code}/{deal_ym}: {exc}")
                if len(errors) >= 3:
                    break
        if len(errors) >= 3:
            break
    summary = summarize_molit_leaders(trades)
    payload = {
        "ok": True,
        "region_name": region_name,
        "full_name": full_name,
        "lawd_codes": lawd_codes,
        "months": months,
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "source": "국토교통부 아파트 매매 실거래가 상세 자료",
        "message": ("최근 실거래를 3개월 중앙값 기준으로 비교했습니다." if trades else "조회된 아파트 실거래가 없습니다."),
        "errors": errors[:5],
        "housing_types": [
            {
                "key": "apartment",
                "label": "아파트",
                **summary,
            },
            {
                "key": "villa",
                "label": "빌라·연립",
                "message": "현재는 아파트 실거래 상세 API부터 연결했습니다. 연립/다세대 API도 같은 구조로 추가 가능합니다.",
                "dong_leaders": [],
                "apartment_leaders": [],
                "latest_trades": [],
            },
        ],
    }
    cache[cache_key] = payload
    save_real_estate_trade_cache(cache)
    return payload


def real_estate_number(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(str(value).replace(",", ""))
    except Exception:
        return 0.0


def repair_mojibake_text(value: Any) -> Any:
    if not isinstance(value, str) or not value:
        return value
    if any("\uac00" <= char <= "\ud7a3" for char in value):
        return value
    if not any(0x80 <= ord(char) <= 0xFF for char in value):
        return value
    try:
        repaired = value.encode("latin1").decode("utf-8")
    except Exception:
        return value
    if sum(1 for char in repaired if "\uac00" <= char <= "\ud7a3") > sum(1 for char in value if "\uac00" <= char <= "\ud7a3"):
        return repaired
    return value


def repair_mojibake_recursive(value: Any) -> Any:
    if isinstance(value, dict):
        return {key: repair_mojibake_recursive(item) for key, item in value.items()}
    if isinstance(value, list):
        return [repair_mojibake_recursive(item) for item in value]
    return repair_mojibake_text(value)


def real_estate_bank_file_candidates() -> list[Path]:
    if not REAL_ESTATE_BANK_IMPORT_DIR.exists():
        return []
    files: list[tuple[tuple[int, str], Path]] = []
    for path in REAL_ESTATE_BANK_IMPORT_DIR.iterdir():
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        suffix = path.suffix.lower()
        if suffix not in {".xlsx", ".xlsm", ".xls", ".csv", ".tsv"}:
            continue
        stem = str(path.stem or "").strip()
        if not stem or "corrupt_before_restore" in stem:
            continue
        if re.fullmatch(r"거래내역조회_기본", stem):
            files.append(((0, stem), path))
            continue
        month_match = re.fullmatch(r"거래내역조회_(20\d{4})", stem)
        if month_match:
            files.append(((1, month_match.group(1)), path))
            continue
    return [path for _, path in sorted(files, key=lambda item: item[0])]


def read_bank_file_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv"}:
        separator = "\t" if suffix == ".tsv" else ","
        last_error: Exception | None = None
        for encoding in ("utf-8-sig", "cp949", "euc-kr"):
            try:
                return pd.read_csv(path, header=None, dtype=str, sep=separator, encoding=encoding, engine="python", names=list(range(80)))
            except Exception as exc:
                last_error = exc
        raise ValueError(str(last_error) if last_error else "CSV 파일을 읽지 못했습니다.")
    return pd.read_excel(path, header=None, dtype=str)


def normalize_bank_header(value: Any) -> str:
    return normalize_search_text(str(value or ""))


def bank_header_score(headers: list[str]) -> int:
    normalized = [normalize_bank_header(header) for header in headers]
    joined = " ".join(normalized)
    score = 0
    if any(token in joined for token in ("거래일", "거래일자", "일자", "거래일시")):
        score += 3
    if any(token in joined for token in ("입금", "입금액", "받은금액")):
        score += 3
    if any(token in joined for token in ("출금", "출금액", "지급금액")):
        score += 3
    if any(token in joined for token in ("거래금액", "금액")):
        score += 2
    if any(token in joined for token in ("내용", "적요", "거래내용", "거래기록사항", "의뢰인", "받는분", "보낸분")):
        score += 2
    return score


def detect_bank_header_row(frame: pd.DataFrame) -> int | None:
    best_index = None
    best_score = 0
    max_rows = min(len(frame.index), 30)
    for row_index in range(max_rows):
        headers = [str(value or "").strip() for value in frame.iloc[row_index].fillna("").tolist()]
        score = bank_header_score(headers)
        if score > best_score:
            best_index = row_index
            best_score = score
    return best_index if best_score >= 5 else None


def bank_column_lookup(headers: list[Any]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    normalized = [normalize_bank_header(header) for header in headers]

    def find(*tokens: str) -> int | None:
        for token in tokens:
            normalized_token = normalize_bank_header(token)
            for index, header in enumerate(normalized):
                if normalized_token and normalized_token in header:
                    return index
        return None

    mapping = {
        "date": find("거래일시", "거래일자", "거래일", "일자", "날짜"),
        "time": find("거래시간", "시간"),
        "kind": find("입출금", "구분", "거래구분"),
        "deposit": find("입금액", "입금", "받은금액"),
        "withdrawal": find("출금액", "출금", "지급금액"),
        "amount": find("거래금액", "금액"),
        "balance": find("잔액", "거래후잔액"),
        "memo": find("거래기록사항", "거래내용", "내용", "적요", "비고", "메모"),
        "counterparty": find("의뢰인", "보낸분", "받는분", "입금자", "수취인"),
        "branch": find("취급점", "거래점", "지점"),
    }
    for key, index in mapping.items():
        if index is not None:
            lookup[key] = index
    return lookup


def parse_bank_date(value: Any) -> str:
    if value is None or pd.isna(value):
        return ""
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat()
    text = str(value).strip()
    if not text:
        return ""
    match = re.search(r"(\d{4})[./-]?\s*(\d{1,2})[./-]?\s*(\d{1,2})", text)
    if match:
        year, month, day = (int(part) for part in match.groups())
        try:
            return date(year, month, day).isoformat()
        except Exception:
            return ""
    return real_estate_date_text(text)


def parse_bank_time(value: Any) -> str:
    text = str(value or "").strip()
    match = re.search(r"(\d{1,2})[:시]\s*(\d{1,2})(?:[:분]\s*(\d{1,2}))?", text)
    if match:
        hour = int(match.group(1))
        minute = int(match.group(2))
        second = int(match.group(3) or 0)
        return f"{hour:02d}:{minute:02d}:{second:02d}"
    digits = re.sub(r"\D+", "", text)
    if len(digits) >= 4:
        return f"{int(digits[:2]):02d}:{int(digits[2:4]):02d}:{int(digits[4:6] or '0'):02d}"
    return ""


def parse_bank_amount(value: Any) -> float:
    if value is None or pd.isna(value):
        return 0.0
    text = str(value).strip()
    if not text or text in {"-", "－"}:
        return 0.0
    negative = text.startswith("-") or text.startswith("△") or text.startswith("▲") or text.endswith("-")
    cleaned = re.sub(r"[^0-9.]", "", text)
    if not cleaned:
        return 0.0
    amount = float(cleaned)
    return -amount if negative else amount


def bank_transaction_id(row: dict[str, Any]) -> str:
    source = "|".join(
        str(row.get(key) or "")
        for key in ("source", "source_file", "source_row", "date", "time", "kind", "amount", "memo", "balance")
    )
    return hashlib.sha1(source.encode("utf-8")).hexdigest()[:24]


def normalize_bank_memo_key(value: Any) -> str:
    text = repair_mojibake_text(str(value or ""))
    text = " ".join(str(text or "").strip().split())
    if not text:
        return ""
    if " / " in text:
        left, _, _right = text.partition(" / ")
        text = left.strip() or text
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalize_bank_memo_match_key(value: Any) -> str:
    text = normalize_bank_memo_key(value)
    if not text:
        return ""
    text = re.sub(r"[^0-9A-Za-z가-힣]", "", text)
    text = re.sub(r"(\D)\d+$", r"\1", text)
    return text.strip().lower()


def default_bank_transaction_category(kind: str, target: str = "", category: str = "") -> str:
    source = str(category or "").strip()
    is_expense = str(kind or "") == "expense"
    if is_expense:
        return "세금" if source in {"", "은행출금", "수동출금", "기타출금"} else source
    if source in {"", "은행입금", "수동입금"}:
        return "월세+관리비+부가세" if str(target or "").strip() else "기타입금"
    return source


def parse_bank_file_transactions(path: Path) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    frame = read_bank_file_table(path)
    header_row = detect_bank_header_row(frame)
    if header_row is None:
        return [], {"file": path.name, "rows": 0, "imported": 0, "error": "거래내역 헤더를 찾지 못했습니다."}
    headers = frame.iloc[header_row].fillna("").tolist()
    columns = bank_column_lookup(headers)
    if "date" not in columns:
        return [], {"file": path.name, "rows": 0, "imported": 0, "error": "거래일자 열을 찾지 못했습니다."}
    transactions = []
    for row_index in range(header_row + 1, len(frame.index)):
        values = frame.iloc[row_index].fillna("").tolist()

        def cell(key: str) -> Any:
            index = columns.get(key)
            return values[index] if index is not None and index < len(values) else ""

        date_text = parse_bank_date(cell("date"))
        if not date_text:
            continue
        time_text = parse_bank_time(cell("time"))
        deposit = abs(parse_bank_amount(cell("deposit")))
        withdrawal = abs(parse_bank_amount(cell("withdrawal")))
        raw_amount = parse_bank_amount(cell("amount"))
        kind_text = str(cell("kind") or "")
        if deposit:
            kind = "income"
            amount = deposit
        elif withdrawal:
            kind = "expense"
            amount = withdrawal
        elif raw_amount:
            if raw_amount < 0 or any(token in kind_text for token in ("출금", "지급", "인출")):
                kind = "expense"
            else:
                kind = "income"
            amount = abs(raw_amount)
        else:
            continue
        memo_parts = [str(cell(key) or "").strip() for key in ("memo", "counterparty", "branch")]
        memo = " / ".join(part for part in memo_parts if part)
        tx = {
            "source": "bank_file",
            "date": date_text,
            "time": time_text,
            "kind": kind,
            "category": default_bank_transaction_category(kind),
            "target": "",
            "payment_method": "현금인출",
            "amount": round(amount),
            "balance": round(abs(parse_bank_amount(cell("balance")))) if "balance" in columns else 0,
            "memo": memo,
            "source_file": path.name,
            "source_row": row_index + 1,
        }
        tx["id"] = bank_transaction_id(tx)
        transactions.append(tx)
    return transactions, {"file": path.name, "rows": max(0, len(frame.index) - header_row - 1), "imported": len(transactions), "error": ""}


def import_real_estate_bank_files() -> dict[str, Any]:
    candidates = real_estate_bank_file_candidates()
    imported: list[dict[str, Any]] = []
    file_results = []
    for path in candidates:
        try:
            transactions, result = parse_bank_file_transactions(path)
            imported.extend(transactions)
            file_results.append(result)
        except Exception as exc:
            file_results.append({"file": path.name, "rows": 0, "imported": 0, "error": str(exc)})

    db = load_real_estate_db()
    existing = db.get("bank_transactions") if isinstance(db.get("bank_transactions"), list) else []
    memo_unit_map = db.get("bank_memo_unit_map") if isinstance(db.get("bank_memo_unit_map"), dict) else {}
    memo_category_map = db.get("bank_memo_category_map") if isinstance(db.get("bank_memo_category_map"), dict) else {}
    normalized_unit_map: dict[str, str] = {}
    normalized_category_map: dict[str, str] = {}
    match_unit_map: dict[str, str] = {}
    match_category_map: dict[str, str] = {}
    for memo_key, target in memo_unit_map.items():
        normalized_key = normalize_bank_memo_key(memo_key)
        match_key = normalize_bank_memo_match_key(memo_key)
        if normalized_key and str(target or "").strip():
            normalized_unit_map[normalized_key] = str(target or "").strip()
        if match_key and str(target or "").strip():
            match_unit_map[match_key] = str(target or "").strip()
    for memo_key, category in memo_category_map.items():
        normalized_key = normalize_bank_memo_key(memo_key)
        match_key = normalize_bank_memo_match_key(memo_key)
        if normalized_key and str(category or "").strip():
            normalized_category_map[normalized_key] = str(category or "").strip()
        if match_key and str(category or "").strip():
            match_category_map[match_key] = str(category or "").strip()
    candidate_names = {path.name for path in candidates}
    exclusions = {
        str(item)
        for item in db.get("bank_transaction_exclusions", [])
        if str(item or "").strip()
    }
    by_key: dict[str, dict[str, Any]] = {}
    historical_target_by_norm_key: dict[str, str] = {}
    historical_category_by_norm_key: dict[str, str] = {}
    historical_target_by_match_key: dict[str, str] = {}
    historical_category_by_match_key: dict[str, str] = {}
    for item in existing:
        if not isinstance(item, dict):
            continue
        memo_key = str(item.get("memo") or "").strip()
        normalized_memo_key = normalize_bank_memo_key(memo_key)
        match_memo_key = normalize_bank_memo_match_key(memo_key)
        if normalized_memo_key:
            target_value = str(item.get("target") or "").strip()
            category_value = str(item.get("category") or "").strip()
            if target_value and normalized_memo_key not in historical_target_by_norm_key:
                historical_target_by_norm_key[normalized_memo_key] = target_value
            if category_value and normalized_memo_key not in historical_category_by_norm_key:
                historical_category_by_norm_key[normalized_memo_key] = category_value
            if target_value and match_memo_key and match_memo_key not in historical_target_by_match_key:
                historical_target_by_match_key[match_memo_key] = target_value
            if category_value and match_memo_key and match_memo_key not in historical_category_by_match_key:
                historical_category_by_match_key[match_memo_key] = category_value
        if item.get("source") == "bank_file" and str(item.get("source_file") or "") in candidate_names:
            continue
        key = str(item.get("id") or "") or bank_transaction_id(item)
        item["id"] = key
        if key in exclusions:
            continue
        by_key[key] = item
    for item in imported:
        if not isinstance(item, dict):
            continue
        key = str(item.get("id") or "") or bank_transaction_id(item)
        item["id"] = key
        memo_key = str(item.get("memo") or "").strip()
        normalized_memo_key = normalize_bank_memo_key(memo_key)
        match_memo_key = normalize_bank_memo_match_key(memo_key)
        target_value = str(memo_unit_map.get(memo_key) or "").strip()
        if not target_value and normalized_memo_key:
            target_value = str(normalized_unit_map.get(normalized_memo_key) or historical_target_by_norm_key.get(normalized_memo_key) or "").strip()
        if not target_value and match_memo_key:
            target_value = str(match_unit_map.get(match_memo_key) or historical_target_by_match_key.get(match_memo_key) or "").strip()
        if target_value:
            item["target"] = target_value
        if normalized_memo_key and target_value:
            historical_target_by_norm_key.setdefault(normalized_memo_key, target_value)
        if match_memo_key and target_value:
            historical_target_by_match_key.setdefault(match_memo_key, target_value)
        category_value = str(memo_category_map.get(memo_key) or "").strip()
        if not category_value and normalized_memo_key:
            category_value = str(normalized_category_map.get(normalized_memo_key) or historical_category_by_norm_key.get(normalized_memo_key) or "").strip()
        if not category_value and match_memo_key:
            category_value = str(match_category_map.get(match_memo_key) or historical_category_by_match_key.get(match_memo_key) or "").strip()
        if category_value:
            item["category"] = category_value
        if normalized_memo_key and category_value:
            historical_category_by_norm_key.setdefault(normalized_memo_key, category_value)
        if match_memo_key and category_value:
            historical_category_by_match_key.setdefault(match_memo_key, category_value)
        item["category"] = default_bank_transaction_category(
            str(item.get("kind") or ""),
            str(item.get("target") or ""),
            str(item.get("category") or ""),
        )
        if key in exclusions:
            continue
        by_key[key] = item
    db["bank_transactions"] = sorted(by_key.values(), key=lambda row: (str(row.get("date") or ""), str(row.get("time") or "")), reverse=True)
    db["bank_import"] = {
        "data_dir": str(REAL_ESTATE_BANK_IMPORT_DIR),
        "last_imported_at": datetime.now().isoformat(timespec="seconds"),
        "file_results": file_results,
        "total_files": len(candidates),
        "total_transactions": len(db["bank_transactions"]),
    }
    saved = save_real_estate_db(db)
    saved["summary"] = build_real_estate_summary(saved)
    return {"ok": True, "imported": len(imported), "files": file_results, "real_estate": saved}


REAL_ESTATE_RENT_DISCOUNT_RULES = [
    {"rate": 15.0, "label": "학원·교습소·독서실·스터디카페·무인매장", "keywords": ["학원", "교습소", "독서실", "스터디카페", "무인매장"]},
    {"rate": 8.0, "label": "세탁소·빨래방·베이커리·카페·커피숍·미용실", "keywords": ["세탁소", "빨래방", "베이커리", "카페", "커피숍", "미용실"]},
    {"rate": 3.0, "label": "사무실·음식점·분식점", "keywords": ["사무실", "음식점", "분식점"]},
]


def real_estate_parse_date(value: Any) -> date | None:
    text = real_estate_date_text(value)
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except Exception:
        return None


def add_months(base_date: date, months: int) -> date:
    month_index = base_date.month - 1 + months
    year = base_date.year + month_index // 12
    month = month_index % 12 + 1
    day = min(base_date.day, [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1])
    return date(year, month, day)


def special_support_months_for_unit(unit_id: str) -> int:
    text = str(unit_id)
    if text.startswith("2"):
        return 4
    if text.startswith("B") or text.startswith("1"):
        return 2
    return 0


def real_estate_months_from_contract(contract: dict[str, Any], field: str, note_field: str, default_when_applied: int = 0) -> int:
    raw_value = contract.get(field)
    if raw_value not in (None, ""):
        return max(0, int(real_estate_number(raw_value)))
    return 0


def special_discount_for_contract(contract: dict[str, Any]) -> dict[str, Any]:
    corpus = " ".join(
        str(contract.get(key) or "")
        for key in ("tenant_business", "tenant", "memo", "vat_note", "rent_free_note", "settlement_support_note")
    ).lower()
    for rule in REAL_ESTATE_RENT_DISCOUNT_RULES:
        matched = [keyword for keyword in rule["keywords"] if keyword.lower() in corpus]
        if matched:
            return {"rate": rule["rate"], "label": rule["label"], "matched": matched}
    return {"rate": 0.0, "label": "할인 조건 미해당", "matched": []}


def special_lease_terms_for_unit(unit_id: str, unit: dict[str, Any]) -> dict[str, Any]:
    contract = unit.get("contract") or {}
    rent_plan = unit.get("rent_plan") or {}
    planned_rent = real_estate_number(rent_plan.get("monthly_rent_manwon")) * 10000
    actual_rent = real_estate_number(contract.get("monthly_rent"))
    management_fee = real_estate_number(contract.get("management_fee"))
    if planned_rent and actual_rent > planned_rent * 4:
        actual_rent = planned_rent
    if planned_rent and management_fee > max(planned_rent * 3, 2_000_000):
        management_fee = 0
    discount = special_discount_for_contract(contract)
    discounted_rent = round(planned_rent * (1 - discount["rate"] / 100.0)) if planned_rent else actual_rent
    contract_rent = actual_rent or discounted_rent
    rent_free_months = real_estate_months_from_contract(contract, "rent_free_months", "rent_free_note", 1)
    support_months = real_estate_months_from_contract(contract, "settlement_support_months", "settlement_support_note", special_support_months_for_unit(unit_id))
    start_date = real_estate_parse_date(contract.get("balance_date")) or real_estate_parse_date(contract.get("contract_date"))
    rent_free_end = None
    support_end = None
    paid_rent_start = None
    current_phase = "normal"
    current_rent_due = contract_rent
    today = date.today()
    if start_date:
        rent_free_end = add_months(start_date, rent_free_months) - timedelta(days=1)
        support_start = add_months(start_date, rent_free_months)
        support_end = add_months(start_date, rent_free_months + support_months) - timedelta(days=1)
        paid_rent_start = add_months(start_date, rent_free_months + support_months)
        if rent_free_months and today <= rent_free_end:
            current_phase = "rent_free"
            current_rent_due = 0
        elif support_months and today <= support_end:
            current_phase = "settlement_support"
            current_rent_due = 0
    return {
        "source": "안암해링턴애비뉴 특별임대조건",
        "rent_free_months": rent_free_months,
        "settlement_support_months": support_months,
        "total_zero_rent_months": rent_free_months + support_months,
        "discount_rate": discount["rate"],
        "discount_label": discount["label"],
        "discount_matched": discount["matched"],
        "planned_monthly_rent": round(planned_rent),
        "discounted_monthly_rent": round(discounted_rent),
        "contract_monthly_rent": round(contract_rent),
        "management_fee": round(management_fee),
        "current_phase": current_phase,
        "current_rent_due": round(current_rent_due),
        "current_monthly_due": round(current_rent_due + management_fee),
        "rent_free_end_date": rent_free_end.isoformat() if rent_free_end else "",
        "settlement_support_end_date": support_end.isoformat() if support_end else "",
        "paid_rent_start_date": paid_rent_start.isoformat() if paid_rent_start else "",
        "notes": [
            f"렌트프리 {rent_free_months}개월",
            f"창업정착지원 {support_months}개월",
            f"업종 할인 {discount['rate']:.0f}%",
        ],
    }


def enrich_real_estate_special_terms(db: dict[str, Any]) -> dict[str, Any]:
    for unit_id, unit in (db.get("units") or {}).items():
        if isinstance(unit, dict):
            terms = special_lease_terms_for_unit(unit_id, unit)
            contract = unit.setdefault("contract", {})
            contract["rent_free_months"] = terms.get("rent_free_months", 0)
            contract["settlement_support_months"] = terms.get("settlement_support_months", 0)
            if terms.get("paid_rent_start_date"):
                contract["rent_start_date"] = terms["paid_rent_start_date"]
            unit["special_terms"] = terms
    return db


def build_water_billing_rows(db: dict[str, Any]) -> list[dict[str, Any]]:
    unit_ids = [unit_id for floor in BUILDING_UNIT_LAYOUT for unit_id in floor["units"]]
    rows: list[dict[str, Any]] = []
    for item in (db.get("water_billing") or {}).get("months", []):
        if not isinstance(item, dict):
            continue
        readings = item.get("readings") if isinstance(item.get("readings"), dict) else {}
        unit_usage = {unit_id: real_estate_number(readings.get(unit_id)) for unit_id in unit_ids}
        units_usage_total = sum(unit_usage.values())
        total_usage = real_estate_number(item.get("total_usage_m3"))
        total_bill = real_estate_number(item.get("total_bill"))
        basis_usage = total_usage if total_usage > 0 else units_usage_total
        common_usage = max(0.0, total_usage - units_usage_total) if total_usage > 0 else 0.0
        unit_rows = []
        for unit_id in unit_ids:
            usage = unit_usage.get(unit_id, 0.0)
            ratio = usage / basis_usage if basis_usage > 0 else 0.0
            amount = total_bill * ratio if total_bill > 0 else 0.0
            unit_rows.append({
                "unit_id": unit_id,
                "usage_m3": round(usage, 4),
                "ratio": round(ratio * 100, 4),
                "amount": round(amount),
            })
        common_ratio = common_usage / basis_usage if basis_usage > 0 else 0.0
        rows.append({
            "month": str(item.get("month") or ""),
            "total_usage_m3": round(total_usage, 4),
            "units_usage_m3": round(units_usage_total, 4),
            "common_usage_m3": round(common_usage, 4),
            "basis_usage_m3": round(basis_usage, 4),
            "total_bill": round(total_bill),
            "common_amount": round(total_bill * common_ratio) if total_bill > 0 else 0,
            "common_ratio": round(common_ratio * 100, 4),
            "units": unit_rows,
            "memo": str(item.get("memo") or ""),
        })
    return rows


def real_estate_unit_ids() -> list[str]:
    return [unit_id for floor in BUILDING_UNIT_LAYOUT for unit_id in floor["units"]]


def normalize_real_estate_unit_target(value: Any) -> str:
    compact = re.sub(r"[^A-Z0-9]", "", str(value or "").upper())
    for unit_id in real_estate_unit_ids():
        unit_key = str(unit_id).upper()
        if compact == unit_key or compact.startswith(unit_key):
            return unit_id
    return ""


def build_water_yearly_rows(db: dict[str, Any], water_rows: list[dict[str, Any]] | None = None) -> list[dict[str, Any]]:
    unit_ids = real_estate_unit_ids()
    water_rows = water_rows if water_rows is not None else build_water_billing_rows(db)
    yearly: dict[str, dict[str, Any]] = {}
    for row in water_rows:
        year = str(row.get("month") or "")[:4]
        if not re.match(r"^\d{4}$", year):
            continue
        bucket = yearly.setdefault(
            year,
            {
                "year": year,
                "total_bill": 0.0,
                "common_amount": 0.0,
                "charge_amount": 0.0,
                "paid_total": 0.0,
                "unassigned_paid": 0.0,
                "units": {unit_id: {"unit_id": unit_id, "amount": 0.0, "paid_amount": 0.0} for unit_id in unit_ids},
            },
        )
        bucket["total_bill"] += real_estate_number(row.get("total_bill"))
        bucket["common_amount"] += real_estate_number(row.get("common_amount"))
        for unit_row in row.get("units", []):
            if not isinstance(unit_row, dict):
                continue
            unit_id = str(unit_row.get("unit_id") or "")
            if unit_id not in bucket["units"]:
                continue
            amount = real_estate_number(unit_row.get("amount"))
            bucket["units"][unit_id]["amount"] += amount
            bucket["charge_amount"] += amount

    for tx in db.get("bank_transactions", []):
        if not isinstance(tx, dict) or str(tx.get("kind") or "") == "expense":
            continue
        if str(tx.get("category") or "").strip() != "수도세":
            continue
        year = str(tx.get("date") or "")[:4]
        if year not in yearly:
            continue
        amount = real_estate_number(tx.get("amount"))
        unit_id = normalize_real_estate_unit_target(tx.get("target"))
        yearly[year]["paid_total"] += amount
        if unit_id and unit_id in yearly[year]["units"]:
            yearly[year]["units"][unit_id]["paid_amount"] += amount
        else:
            yearly[year]["unassigned_paid"] += amount

    rows = []
    for year, bucket in sorted(yearly.items()):
        unit_rows = []
        outstanding_total = 0.0
        for unit_id in unit_ids:
            unit = bucket["units"][unit_id]
            amount = real_estate_number(unit.get("amount"))
            paid_amount = real_estate_number(unit.get("paid_amount"))
            outstanding = max(0.0, amount - paid_amount)
            outstanding_total += outstanding
            unit_rows.append({
                "unit_id": unit_id,
                "amount": round(amount),
                "paid_amount": round(paid_amount),
                "outstanding": round(outstanding),
            })
        rows.append({
            "year": year,
            "total_bill": round(bucket["total_bill"]),
            "charge_amount": round(bucket["charge_amount"]),
            "common_amount": round(bucket["common_amount"]),
            "paid_total": round(bucket["paid_total"]),
            "unassigned_paid": round(bucket["unassigned_paid"]),
            "outstanding_total": round(outstanding_total),
            "units": unit_rows,
        })
    return rows


def build_electricity_billing_rows(db: dict[str, Any]) -> list[dict[str, Any]]:
    customer_ids = ["common"] + [unit_id for floor in BUILDING_UNIT_LAYOUT for unit_id in floor["units"]]
    rows: list[dict[str, Any]] = []
    for item in (db.get("electricity_billing") or {}).get("months", []):
        if not isinstance(item, dict):
            continue
        bills = item.get("bills") if isinstance(item.get("bills"), dict) else {}
        unit_rows = []
        total_bill = 0.0
        for unit_id in customer_ids:
            amount = real_estate_number(bills.get(unit_id))
            total_bill += amount
            info = BUILDING_ELECTRIC_CUSTOMERS.get(unit_id, {})
            unit_rows.append({
                "unit_id": unit_id,
                "label": str(info.get("label") or unit_id),
                "customer_no": str(info.get("customer_no") or ""),
                "amount": round(amount),
            })
        saved_total = real_estate_number(item.get("total_bill"))
        rows.append({
            "month": str(item.get("month") or ""),
            "due_date": str(item.get("due_date") or ""),
            "due_dates": item.get("due_dates") if isinstance(item.get("due_dates"), dict) else {},
            "total_bill": round(saved_total if saved_total > 0 else total_bill),
            "units": unit_rows,
            "memo": str(item.get("memo") or ""),
            "source": str(item.get("source") or ""),
        })
    return rows


def build_real_estate_summary(db: dict[str, Any]) -> dict[str, Any]:
    def is_profit_withdrawal(category: Any) -> bool:
        return re.sub(r"\s+", "", str(category or "")) == "수익출금"

    def effective_bank_category(tx: dict[str, Any]) -> str:
        memo = str(tx.get("memo") or "").strip()
        mapped = db.get("bank_memo_category_map", {})
        if str(tx.get("kind") or "") == "expense" and isinstance(mapped, dict) and memo and str(mapped.get(memo) or "").strip():
            return str(mapped.get(memo) or "").strip()
        return str(tx.get("category") or "").strip()

    def add_actual_flow(tx_date: str, signed: float, category: Any = "") -> None:
        nonlocal month_income, month_expense, month_net, year_income, year_expense, year_net
        if signed < 0 and is_profit_withdrawal(category):
            return
        if tx_date.startswith(current_month):
            month_net += signed
            if signed > 0:
                month_income += signed
            if signed < 0:
                month_expense += abs(signed)
        if tx_date.startswith(str(date.today().year)):
            year_net += signed
            if signed > 0:
                year_income += signed
            if signed < 0:
                year_expense += abs(signed)

    monthly_contract_income = 0.0
    monthly_rent_income_total = 0.0
    monthly_vat_income_total = 0.0
    monthly_management_income_total = 0.0
    total_deposit = 0.0
    occupied_count = 0
    monthly_water = 0.0
    monthly_electricity = 0.0
    monthly_service_expense = 0.0
    yearly_service_expense = 0.0
    area_totals = {"exclusive_m2": 0.0, "parking_m2": 0.0, "common_m2": 0.0, "exclusive_py": 0.0, "sale_py": 0.0}
    current_month = date.today().strftime("%Y-%m")
    month_income = 0.0
    month_expense = 0.0
    month_net = 0.0
    year_income = 0.0
    year_expense = 0.0
    year_net = 0.0
    water_billing_rows = build_water_billing_rows(db)
    water_yearly_rows = build_water_yearly_rows(db, water_billing_rows)
    electricity_billing_rows = build_electricity_billing_rows(db)
    latest_water_bill = 0.0
    latest_water_charge = 0.0
    for water_row in sorted(water_billing_rows, key=lambda row: row.get("month") or ""):
        if real_estate_number(water_row.get("total_bill")):
            latest_water_bill = real_estate_number(water_row.get("total_bill"))
            latest_water_charge = sum(real_estate_number(unit.get("amount")) for unit in water_row.get("units", []))
    current_year = str(date.today().year)
    current_year_water = next((row for row in water_yearly_rows if str(row.get("year") or "") == current_year), {})
    current_year_water_charge = real_estate_number(current_year_water.get("charge_amount"))
    current_year_water_bill = real_estate_number(current_year_water.get("total_bill"))
    latest_electricity_bill = 0.0
    for electric_row in sorted(electricity_billing_rows, key=lambda row: row.get("month") or ""):
        if real_estate_number(electric_row.get("total_bill")):
            latest_electricity_bill = real_estate_number(electric_row.get("total_bill"))
    for unit in db.get("units", {}).values():
        area = unit.get("area", {})
        if isinstance(area, dict):
            for key in area_totals:
                area_totals[key] += real_estate_number(area.get(key))
        contract = unit.get("contract", {})
        tenant = str(contract.get("tenant") or "").strip()
        tenant_business = str(contract.get("tenant_business") or "").strip()
        special_terms = unit.get("special_terms") or special_lease_terms_for_unit(str(unit.get("unit_id") or ""), unit)
        rent = real_estate_number(special_terms.get("current_rent_due"))
        management_fee = real_estate_number(special_terms.get("management_fee"))
        if not management_fee:
            current_due = real_estate_number(special_terms.get("current_monthly_due"))
            if current_due > rent:
                management_fee = current_due - rent
        if not management_fee:
            management_fee = real_estate_number(contract.get("management_fee"))
        if rent and management_fee > max(rent * 3, 2_000_000):
            management_fee = 0
        discount_rate = real_estate_number(contract.get("discount_rate"))
        deposit = real_estate_number(contract.get("deposit"))
        is_occupied = bool(tenant or tenant_business)
        if is_occupied:
            occupied_count += 1
            total_deposit += deposit
            monthly_rent_income = max(0.0, rent * (1 - discount_rate / 100.0))
            charge = real_estate_monthly_charge_parts(monthly_rent_income, management_fee, contract.get("vat_note"))
            monthly_rent_income_total += max(0.0, charge["rent"])
            monthly_vat_income_total += max(0.0, charge["vat_total"])
            monthly_management_income_total += max(0.0, charge["management_fee"])
            monthly_contract_income += max(0.0, charge["total"])
        for tx in unit.get("transactions", []):
            amount = real_estate_number(tx.get("amount"))
            kind = str(tx.get("kind") or "income")
            signed = amount if kind != "expense" else -amount
            tx_date = str(tx.get("date") or "")
            add_actual_flow(tx_date, signed, tx.get("category"))
    service_rows = []
    for service in db.get("service_contracts", []):
        if not isinstance(service, dict):
            continue
        amount = real_estate_number(service.get("amount"))
        cycle = str(service.get("cycle") or "monthly").strip() or "monthly"
        start_date = str(service.get("start_date") or "")
        payment_date = str(
            service.get("payment_date")
            or service.get("annual_payment_date")
            or service.get("yearly_payment_date")
            or ""
        )
        is_started = not (start_date and start_date[:7] > current_month)
        if cycle in {"yearly", "annual"}:
            annual_anchor = payment_date if len(payment_date) >= 10 else start_date
            annual_month = annual_anchor[5:7] if len(annual_anchor) >= 7 else ""
            monthly_amount = amount if is_started and annual_month and annual_month == current_month[5:7] else 0.0
            if not start_date or start_date[:4] <= current_year:
                yearly_service_expense += amount
        else:
            cycle = "monthly"
            monthly_amount = amount if is_started else 0.0
            if is_started:
                yearly_service_expense += amount * 12
        monthly_service_expense += monthly_amount
        service_rows.append({
            "category": str(service.get("category") or "기타"),
            "vendor": str(service.get("vendor") or ""),
            "amount": round(amount),
            "monthly_amount": round(monthly_amount),
            "cycle": cycle,
            "payment_day": int(real_estate_number(service.get("payment_day") or service.get("due_day") or 1)) or 1,
            "payment_date": payment_date,
            "start_date": start_date,
        })
    for tx in db.get("operating_transactions", []):
        if not isinstance(tx, dict):
            continue
        amount = real_estate_number(tx.get("amount"))
        kind = str(tx.get("kind") or "expense")
        signed = amount if kind == "income" else -amount
        tx_date = str(tx.get("date") or "")
        add_actual_flow(tx_date, signed, tx.get("category"))
    for tx in db.get("bank_transactions", []):
        if not isinstance(tx, dict):
            continue
        amount = real_estate_number(tx.get("amount"))
        kind = str(tx.get("kind") or "income")
        signed = amount if kind != "expense" else -amount
        tx_date = str(tx.get("date") or "")
        add_actual_flow(tx_date, signed, effective_bank_category(tx))
    units_count = len(db.get("units", {})) or 1
    monthly_water_income = latest_water_charge
    monthly_water_expense = latest_water_bill
    monthly_base_service_expense = monthly_service_expense
    if latest_electricity_bill:
        monthly_electricity = latest_electricity_bill
    monthly_service_expense += monthly_electricity + monthly_water_expense
    expected_gross_income = monthly_contract_income + monthly_water_income
    expected_monthly_net_income = expected_gross_income - monthly_service_expense
    expected_yearly_income = monthly_contract_income * 12 + current_year_water_charge
    expected_yearly_service_expense = yearly_service_expense + monthly_electricity * 12 + current_year_water_bill
    expected_yearly_net_income = expected_yearly_income - expected_yearly_service_expense
    return {
        "units_count": units_count,
        "occupied_count": occupied_count,
        "vacancy_count": units_count - occupied_count,
        "total_deposit": round(total_deposit),
        "expected_monthly_income": round(expected_gross_income),
        "expected_yearly_income": round(expected_yearly_income),
        "expected_monthly_service_expense": round(monthly_service_expense),
        "expected_yearly_service_expense": round(expected_yearly_service_expense),
        "expected_monthly_net_income": round(expected_monthly_net_income),
        "expected_yearly_net_income": round(expected_yearly_net_income),
        "actual_month_income": round(month_income),
        "actual_month_expense": round(month_expense),
        "actual_month_net_income": round(month_net),
        "actual_year_income": round(year_income),
        "actual_year_expense": round(year_expense),
        "actual_year_net_income": round(year_net),
        "occupancy_rate": round(occupied_count / units_count * 100, 1),
        "area_totals": {key: round(value, 2) for key, value in area_totals.items()},
        "water_billing_rows": water_billing_rows,
        "water_yearly_rows": water_yearly_rows,
        "electricity_billing_rows": electricity_billing_rows,
        "service_rows": service_rows,
        "profit_rows": [
            {"label": "월세", "kind": "plus", "monthly": round(monthly_rent_income_total), "yearly": round(monthly_rent_income_total * 12)},
            {"label": "부가세", "kind": "plus", "monthly": round(monthly_vat_income_total), "yearly": round(monthly_vat_income_total * 12)},
            {"label": "관리비", "kind": "plus", "monthly": round(monthly_management_income_total), "yearly": round(monthly_management_income_total * 12)},
            {"label": "수도세 청구", "kind": "plus", "monthly": round(monthly_water_income), "yearly": round(current_year_water_charge)},
            {"label": "수도세 납부", "kind": "minus", "monthly": round(monthly_water_expense), "yearly": round(current_year_water_bill)},
            {"label": "전기세 청구", "kind": "minus", "monthly": round(monthly_electricity), "yearly": round(monthly_electricity * 12)},
            {"label": "용역/서비스 비용", "kind": "minus", "monthly": round(monthly_base_service_expense), "yearly": round(yearly_service_expense)},
            {"label": "예상 순수익", "kind": "total", "monthly": round(expected_monthly_net_income), "yearly": round(expected_yearly_net_income)},
        ],
    }


def build_trade_real_payload(api_key: str, region: str = "all", item_key: str = "") -> dict[str, Any]:
    years = trade_year_range(2021)
    items: list[dict[str, Any]] = []
    has_any_value = False
    region_option = trade_region_option(region)
    item_values: dict[str, dict[str, float]] = {}
    errors: list[str] = []
    selected_item_key = str(item_key or "").strip() or TRADE_IMPORT_EXPORT_ITEMS[0]["key"]
    fetch_items = TRADE_IMPORT_EXPORT_ITEMS if region == "all" else [
        item for item in TRADE_IMPORT_EXPORT_ITEMS if item["key"] == selected_item_key
    ]
    if not fetch_items:
        fetch_items = [TRADE_IMPORT_EXPORT_ITEMS[0]]
        selected_item_key = TRADE_IMPORT_EXPORT_ITEMS[0]["key"]
    with ThreadPoolExecutor(max_workers=min(8, max(1, len(TRADE_IMPORT_EXPORT_ITEMS)))) as executor:
        futures = {
            executor.submit(fetch_trade_item_real_data, item, years, api_key, region): item
            for item in fetch_items
        }
        for future in as_completed(futures):
            item = futures[future]
            try:
                item_values[item["key"]] = future.result()
            except Exception as exc:
                item_values[item["key"]] = {}
                errors.append(f"{item['name']}: {exc}")
    for item in TRADE_IMPORT_EXPORT_ITEMS:
        values = item_values.get(item["key"], {})
        if values:
            has_any_value = True
        rows = []
        for month in range(1, 13):
            month_values = [{"year": year, "value": values.get(trade_month_key(year, month))} for year in years]
            rows.append({"month": month, "values": month_values})
        items.append({
            "key": item["key"],
            "name": item["name"],
            "hs_codes": item["hs_codes"],
            "companies": item.get("companies", []),
            "unit": item["unit"],
            "rows": rows,
            "metrics": trade_item_change_metrics(values),
        })
    if not has_any_value:
        raise ValueError("관세청 API에서 조회 가능한 품목 수출입 데이터가 없습니다.")
    return {
        "source": "customs_sido_api" if region != "all" else "customs_api",
        "source_label": "관세청 시도별 품목별 수출입실적 API" if region != "all" else "관세청 수출입무역통계 API",
        "message": (
            f"{region_option['name']} 지역의 시도별 품목별 수출금액입니다. 특정 기업의 지역 생산/수출 비중을 추정하는 보조 지표로 활용합니다."
            if region != "all"
            else "관세청 품목별 수출입실적 API에서 조회한 월별 수출금액입니다. 원 API의 달러 값을 1,000으로 나눠 천달러 단위로 표시합니다."
        ),
        "release_hint": "매월 1일, 11일, 21일 발표/갱신 데이터 확인용입니다.",
        "years": years,
        "items": items,
        "trade_items_version": TRADE_DATA_VERSION,
        "item_errors": errors[:20],
        "loaded_at": datetime.now().isoformat(timespec="seconds"),
        "selected_region": region,
        "selected_item_key": selected_item_key,
        "region_options": TRADE_REGION_OPTIONS,
        "region_note": f"{region_option['name']} 지역 기준입니다." if region != "all" else "전국 HS코드 기준 수출금액입니다.",
    }


def load_trade_import_export_payload(force_refresh: bool = False, region: str = "all", item_key: str = "") -> dict[str, Any]:
    region_codes = {str(item["code"]) for item in TRADE_REGION_OPTIONS}
    selected_region = region if region in region_codes else "all"
    selected_item_key = str(item_key or "").strip() or TRADE_IMPORT_EXPORT_ITEMS[0]["key"]
    item_keys = {str(item["key"]) for item in TRADE_IMPORT_EXPORT_ITEMS}
    if selected_item_key not in item_keys:
        selected_item_key = TRADE_IMPORT_EXPORT_ITEMS[0]["key"]
    cache_path = trade_data_cache_path(selected_region, selected_item_key)
    current_slot_key, current_slot_label = trade_snapshot_slot()
    if not force_refresh and cache_path.exists():
        try:
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
            if (
                str(cached.get("trade_snapshot_slot") or "") == current_slot_key
                and int(cached.get("trade_items_version") or 0) == TRADE_DATA_VERSION
                and str(cached.get("selected_region") or selected_region) == selected_region
                and str(cached.get("selected_item_key") or selected_item_key) == selected_item_key
                and not (selected_region != "all" and cached.get("region_fallback"))
            ):
                cached["cache_hit"] = True
                cached["trade_snapshot_label"] = cached.get("trade_snapshot_label") or current_slot_label
                cached["cache_policy"] = cached.get("cache_policy") or trade_cache_policy_label()
                return cached
        except Exception:
            pass
    api_key = trade_data_api_key()
    if api_key:
        try:
            payload = build_trade_real_payload(api_key, region=selected_region, item_key=selected_item_key)
        except Exception as exc:
            if selected_region != "all":
                payload = build_trade_real_payload(api_key, region="all", item_key=selected_item_key)
                payload["selected_region"] = selected_region
                payload["region_options"] = TRADE_REGION_OPTIONS
                payload["region_note"] = (
                    f"{trade_region_option(selected_region)['name']} 지역 API 호출 실패로 전국 데이터를 대신 표시합니다: {exc}"
                )
                payload["region_fallback"] = True
            else:
                payload = build_trade_sample_payload()
                payload["source"] = "sample_fallback"
                payload["source_label"] = "샘플 데이터(API 호출 실패)"
                payload["message"] = f"관세청 API 호출 실패로 샘플 데이터를 표시합니다: {exc}"
    else:
        payload = build_trade_sample_payload()
    motie_status = check_motie_trade_service(motie_trade_service_key())
    payload["motie_api"] = motie_status
    if payload.get("source") == "sample" and motie_status.get("configured"):
        payload["source_label"] = "샘플 데이터(산업부 API 키 저장됨)"
        payload["message"] = (
            "산업부 무역정보서비스 키는 저장했지만 문서상 해당 API는 기업/거래물품/지역별 수출입 안내용입니다. "
            "라면·화장품 품목별 월간 수출액 실데이터는 관세청 품목별 수출입실적 API 키가 필요해 현재는 이미지 형태 검증용 샘플로 표시합니다."
        )
    payload["selected_region"] = selected_region
    payload["selected_item_key"] = selected_item_key
    payload["region_options"] = TRADE_REGION_OPTIONS
    payload["trade_snapshot_slot"] = current_slot_key
    payload["trade_snapshot_label"] = current_slot_label
    payload["cache_policy"] = trade_cache_policy_label()
    payload["cache_hit"] = False
    if payload.get("source") in {"customs_api", "customs_sido_api"} and not payload.get("region_fallback"):
        persist_trade_snapshot(payload, selected_region)
    try:
        temp_path = cache_path.with_name(f"{cache_path.stem}_{uuid.uuid4().hex[:8]}.tmp")
        temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        temp_path.replace(cache_path)
    except Exception:
        pass
    return payload


def load_tourist_visitor_cache() -> dict[str, Any]:
    if not TOURIST_VISITOR_CACHE_PATH.exists():
        return {}
    try:
        payload = json.loads(TOURIST_VISITOR_CACHE_PATH.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def save_tourist_visitor_cache(payload: dict[str, Any]) -> None:
    try:
        temp_path = TOURIST_VISITOR_CACHE_PATH.with_name(f"{TOURIST_VISITOR_CACHE_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
        temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        temp_path.replace(TOURIST_VISITOR_CACHE_PATH)
    except Exception:
        pass


def parse_tourist_visitor_xml(xml_text: str) -> list[dict[str, Any]]:
    root = ET.fromstring(xml_text.encode("utf-8"))
    rows: list[dict[str, Any]] = []
    for row in root.findall(".//row"):
        label = (row.findtext("TS_NM") or "").strip()
        if label and label != "계":
            continue
        for child in row:
            match = re.fullmatch(r"CUR_(\d{4})(\d{2})", child.tag or "")
            if not match:
                continue
            value = clean_numeric_text(child.text)
            if value is None or value <= 0:
                continue
            rows.append(
                {
                    "year": int(match.group(1)),
                    "month": int(match.group(2)),
                    "date": f"{match.group(1)}-{match.group(2)}",
                    "visitors": int(value),
                }
            )
        if rows:
            break
    rows.sort(key=lambda item: item["date"])
    return rows


def fetch_tourist_inbound_visitors(start_year: int, end_year: int) -> dict[str, Any]:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0",
            "Referer": "https://know.tour.go.kr/stat/entryTourStatDis19Re.do",
        }
    )
    session.get("https://know.tour.go.kr/stat/entryTourStatDis19Re.do", timeout=15)
    response = session.post(
        "https://know.tour.go.kr/stat/entryTourStatDis_DataXML.do",
        data={
            "searchDivision": "entryBase",
            "searchDateDivision": "MONTH",
            "searchStartYear": str(start_year),
            "searchStartMonth": "01",
            "searchEndYear": str(end_year),
            "searchEndMonth": "12",
            "searchComponent": "N",
            "searchGrow": "N",
            "searchPrevMGrow": "N",
            "searchPrevYGrow": "N",
            "mergyYN": "Y",
            "advancedSearchYn": "N",
            "searchType": "",
            "searchXItemFirst": "",
            "searchXItemSecond": "",
            "searchYItem": "",
        },
        timeout=25,
    )
    response.raise_for_status()
    response.encoding = "utf-8"
    rows = parse_tourist_visitor_xml(response.text)
    if not rows:
        raise ValueError("방한 외래객 월별 데이터를 찾지 못했습니다.")
    previous_year_map = {(row["year"], row["month"]): row for row in rows}
    for row in rows:
        previous = previous_year_map.get((row["year"] - 1, row["month"]))
        row["yoy_pct"] = round(((row["visitors"] / previous["visitors"]) - 1) * 100, 1) if previous and previous.get("visitors") else None
    latest = rows[-1]
    return {
        "version": TOURIST_VISITOR_DATA_VERSION,
        "source": "know_tour",
        "source_label": "관광지식정보시스템 방한 외래객 통계",
        "source_url": "https://know.tour.go.kr/stat/entryTourStatDis19Re.do",
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "start_year": start_year,
        "end_year": end_year,
        "latest_month": latest["date"],
        "latest_visitors": latest["visitors"],
        "rows": rows,
    }


def build_tourist_inbound_visitors_payload(force_refresh: bool = False) -> dict[str, Any]:
    today = date.today()
    start_year = today.year - 5
    end_year = today.year
    cache = load_tourist_visitor_cache()
    cache_loaded_at = str(cache.get("fetched_at") or "")
    should_fetch = force_refresh or int(cache.get("version") or 0) != TOURIST_VISITOR_DATA_VERSION or not cache.get("rows")
    if not should_fetch and cache_loaded_at:
        try:
            loaded_date = datetime.fromisoformat(cache_loaded_at).date()
            should_fetch = loaded_date < today
        except Exception:
            should_fetch = True
    if should_fetch:
        try:
            payload = fetch_tourist_inbound_visitors(start_year, end_year)
            save_tourist_visitor_cache(payload)
            return payload
        except Exception as exc:
            if cache.get("rows"):
                return {**cache, "error": str(exc), "message": "관광객 통계 최신 조회에 실패해 저장된 데이터를 표시합니다."}
            raise
    return cache


DRAM_PRICE_URL = "https://www.trendforce.com/price/dram/dram_contract"
DRAM_PRICE_SECTIONS = {
    "DRAM Spot Price": "dram_spot",
    "DRAM Contract Price": "dram_contract",
    "Module Spot Price": "module_spot",
    "GDDR Spot Price": "gddr_spot",
    "LPDDR Spot Price": "lpddr_spot",
    "Mobile DRAM Contract Price": "mobile_dram_contract",
}
SSD_PRICE_URL = "https://www.trendforce.com/price/flash/ssd_street"
SSD_PRICE_SECTIONS = {
    "SSD Street Price": "ssd_street",
    "PC-Client OEM SSD Contract Price": "pc_client_oem_ssd_contract",
}


def parse_dram_price_number(value: str) -> float | None:
    cleaned = re.sub(r"[^\d.\-]", "", str(value or ""))
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_dram_change(value: str) -> float | None:
    text = str(value or "")
    number = parse_dram_price_number(text)
    if number is None:
        return None
    if "▼" in text or "↓" in text:
        return -abs(number)
    return number


def parse_trendforce_dram_prices(html_text: str) -> dict[str, Any]:
    soup = BeautifulSoup(html_text, "html.parser")
    tables = soup.select("table.price-table")
    items: list[dict[str, Any]] = []
    for table in tables:
        title = ""
        section = ""
        title_node = table.find_previous("div", class_="price-title")
        title_text = title_node.get_text(" ", strip=True) if title_node else ""
        for section_title, section_key in DRAM_PRICE_SECTIONS.items():
            if title_text.startswith(section_title):
                title = title_text
                section = section_key
                break
        if not section:
            continue
        update_node = table.find_previous("div", class_="price-last-update")
        source_update = ""
        if update_node:
            update_text = update_node.get_text(" ", strip=True)
            match = re.search(r"Last\s+Update\s+(.+)", update_text, flags=re.IGNORECASE)
            source_update = match.group(1).strip() if match else update_text
        headers = [cell.get_text(" ", strip=True) for cell in table.select("thead th")]
        header_index = {header: index for index, header in enumerate(headers)}

        def read_cell(cells: list[str], *names: str) -> str:
            for name in names:
                index = header_index.get(name)
                if index is not None and index < len(cells):
                    return cells[index]
            return ""

        for row in table.select("tbody tr"):
            cells = [cell.get_text(" ", strip=True) for cell in row.find_all("td")]
            if len(cells) < 4:
                continue
            item_name = read_cell(cells, "Item") or cells[0].strip()
            if not item_name:
                continue
            item = {
                "section": section,
                "section_label": title,
                "item": item_name,
                "source_update": source_update,
                "daily_high": parse_dram_price_number(read_cell(cells, "Daily High", "Weekly High")),
                "daily_low": parse_dram_price_number(read_cell(cells, "Daily Low", "Weekly Low")),
                "session_high": parse_dram_price_number(read_cell(cells, "Session High")),
                "session_low": parse_dram_price_number(read_cell(cells, "Session Low")),
                "session_average": parse_dram_price_number(read_cell(cells, "Session Average")),
                "change_pct": parse_dram_change(read_cell(cells, "Session Change", "Average Change", "Low Change")),
            }
            if item["session_average"] is None:
                continue
            item["key"] = normalize_search_text(f"{section}:{item_name}")
            items.append(item)
    return {
        "source": "trendforce_dramexchange",
        "source_label": "DramExchange / TrendForce 공개 가격표",
        "source_url": DRAM_PRICE_URL,
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "items": items,
    }


def load_dram_price_history() -> dict[str, Any]:
    if not DRAM_PRICE_HISTORY_PATH.exists():
        return {"items": {}}
    try:
        payload = json.loads(DRAM_PRICE_HISTORY_PATH.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else {"items": {}}
    except Exception:
        return {"items": {}}


def load_ssd_price_history() -> dict[str, Any]:
    if not SSD_PRICE_HISTORY_PATH.exists():
        return {"items": {}}
    try:
        payload = json.loads(SSD_PRICE_HISTORY_PATH.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else {"items": {}}
    except Exception:
        return {"items": {}}


def save_dram_price_history(history: dict[str, Any]) -> None:
    try:
        temp_path = DRAM_PRICE_HISTORY_PATH.with_name(f"{DRAM_PRICE_HISTORY_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
        temp_path.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        temp_path.replace(DRAM_PRICE_HISTORY_PATH)
    except Exception:
        pass


def save_ssd_price_history(history: dict[str, Any]) -> None:
    try:
        temp_path = SSD_PRICE_HISTORY_PATH.with_name(f"{SSD_PRICE_HISTORY_PATH.stem}_{uuid.uuid4().hex[:8]}.tmp")
        temp_path.write_text(json.dumps(history, ensure_ascii=False, indent=2), encoding="utf-8")
        temp_path.replace(SSD_PRICE_HISTORY_PATH)
    except Exception:
        pass


def merge_dram_price_history(snapshot: dict[str, Any]) -> dict[str, Any]:
    history = load_dram_price_history()
    items_by_key = history.setdefault("items", {})
    fetched_at = str(snapshot.get("fetched_at") or datetime.now().isoformat(timespec="seconds"))
    for item in snapshot.get("items", []):
        key = str(item.get("key") or "")
        avg = item.get("session_average")
        if not key or avg is None:
            continue
        current = items_by_key.setdefault(key, {"meta": {}, "points": []})
        current["meta"] = {
            "key": key,
            "section": item.get("section"),
            "section_label": item.get("section_label"),
            "item": item.get("item"),
        }
        source_update = str(item.get("source_update") or fetched_at)
        point = {
            "date": source_update,
            "fetched_at": fetched_at,
            "value": avg,
            "session_high": item.get("session_high"),
            "session_low": item.get("session_low"),
            "change_pct": item.get("change_pct"),
        }
        points = [p for p in current.get("points", []) if str(p.get("date")) != source_update]
        points.append(point)
        current["points"] = sorted(points, key=lambda p: str(p.get("date") or ""))[-180:]
    history["updated_at"] = fetched_at
    save_dram_price_history(history)
    return history


def merge_ssd_price_history(snapshot: dict[str, Any]) -> dict[str, Any]:
    history = load_ssd_price_history()
    items_by_key = history.setdefault("items", {})
    fetched_at = str(snapshot.get("fetched_at") or datetime.now().isoformat(timespec="seconds"))
    for item in snapshot.get("items", []):
        key = str(item.get("key") or "")
        avg = item.get("session_average")
        if not key or avg is None:
            continue
        current = items_by_key.setdefault(key, {"meta": {}, "points": []})
        current["meta"] = {
            "key": key,
            "section": item.get("section"),
            "section_label": item.get("section_label"),
            "item": item.get("item"),
        }
        source_update = str(item.get("source_update") or fetched_at)
        point = {
            "date": source_update,
            "fetched_at": fetched_at,
            "value": avg,
            "session_high": item.get("session_high"),
            "session_low": item.get("session_low"),
            "change_pct": item.get("change_pct"),
        }
        points = [p for p in current.get("points", []) if str(p.get("date")) != source_update]
        points.append(point)
        current["points"] = sorted(points, key=lambda p: str(p.get("date") or ""))[-180:]
    history["updated_at"] = fetched_at
    save_ssd_price_history(history)
    return history


def parse_trendforce_ssd_prices(html_text: str) -> dict[str, Any]:
    soup = BeautifulSoup(html_text, "html.parser")
    items: list[dict[str, Any]] = []
    for table in soup.select("table.price-table"):
        title_node = table.find_previous("div", class_="price-title")
        title_text = title_node.get_text(" ", strip=True) if title_node else ""
        title = ""
        section = ""
        for section_title, section_key in SSD_PRICE_SECTIONS.items():
            if title_text.startswith(section_title):
                title = title_text
                section = section_key
                break
        if not section:
            continue
        update_node = table.find_previous("div", class_="price-last-update")
        source_update = ""
        if update_node:
            update_text = update_node.get_text(" ", strip=True)
            match = re.search(r"Last\s+Update\s+(.+)", update_text, flags=re.IGNORECASE)
            source_update = match.group(1).strip() if match else update_text
        headers = [cell.get_text(" ", strip=True) for cell in table.select("thead th")]
        header_index = {header: index for index, header in enumerate(headers)}

        def read_cell(cells: list[str], *names: str) -> str:
            for name in names:
                index = header_index.get(name)
                if index is not None and index < len(cells):
                    return cells[index]
            return ""

        for row in table.select("tbody tr"):
            cells = [cell.get_text(" ", strip=True) for cell in row.find_all("td")]
            if len(cells) < 4:
                continue
            if section == "ssd_street":
                item_name = " ".join(filter(None, [
                    read_cell(cells, "Brand"),
                    read_cell(cells, "Series"),
                    read_cell(cells, "Capacity"),
                    read_cell(cells, "Interface"),
                ])).strip()
                high_text = read_cell(cells, "High")
                low_text = read_cell(cells, "Low")
                avg_text = read_cell(cells, "Average")
                change_text = read_cell(cells, "Change")
            else:
                item_name = read_cell(cells, "Item") or cells[0].strip()
                high_text = read_cell(cells, "High", "Session High")
                low_text = read_cell(cells, "Low", "Session Low")
                avg_text = read_cell(cells, "Average", "Session Average")
                change_text = read_cell(cells, "Average Change", "Session Change", "Change")
            avg = parse_dram_price_number(avg_text)
            if not item_name or avg is None:
                continue
            item = {
                "section": section,
                "section_label": title,
                "item": item_name,
                "source_update": source_update,
                "session_high": parse_dram_price_number(high_text),
                "session_low": parse_dram_price_number(low_text),
                "session_average": avg,
                "change_pct": parse_dram_change(change_text),
            }
            item["key"] = normalize_search_text(f"{section}:{item_name}")
            items.append(item)
    return {
        "source": "trendforce_ssd",
        "source_label": "DramExchange / TrendForce SSD 공개 가격표",
        "source_url": SSD_PRICE_URL,
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
        "items": items,
    }


def build_ssd_price_payload(force_refresh: bool = False) -> dict[str, Any]:
    history = load_ssd_price_history()
    should_fetch = force_refresh or not history.get("updated_at")
    if not should_fetch:
        try:
            updated_at = datetime.fromisoformat(str(history.get("updated_at")))
            should_fetch = updated_at < datetime.now() - timedelta(hours=3)
        except Exception:
            should_fetch = True
    snapshot: dict[str, Any] | None = None
    error_message = ""
    if should_fetch:
        try:
            response = requests.get(SSD_PRICE_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
            response.raise_for_status()
            snapshot = parse_trendforce_ssd_prices(response.text)
            if not snapshot.get("items"):
                raise ValueError("SSD 가격표 항목을 찾지 못했습니다.")
            history = merge_ssd_price_history(snapshot)
        except Exception as exc:
            error_message = str(exc)
    current_items = snapshot.get("items", []) if snapshot else []
    if not current_items:
        latest_by_key: dict[str, dict[str, Any]] = {}
        for key, entry in (history.get("items") or {}).items():
            points = entry.get("points") or []
            if not points:
                continue
            latest = points[-1]
            meta = entry.get("meta") or {}
            latest_by_key[key] = {
                **meta,
                "source_update": latest.get("date"),
                "session_average": latest.get("value"),
                "session_high": latest.get("session_high"),
                "session_low": latest.get("session_low"),
                "change_pct": latest.get("change_pct"),
            }
        current_items = list(latest_by_key.values())
    history_items = []
    for key, entry in (history.get("items") or {}).items():
        meta = entry.get("meta") or {}
        history_items.append({**meta, "points": entry.get("points") or []})
    history_items.sort(key=lambda item: (str(item.get("section") or ""), str(item.get("item") or "")))
    return {
        "source": "trendforce_ssd",
        "source_label": "DramExchange / TrendForce SSD 공개 가격표",
        "source_url": SSD_PRICE_URL,
        "fetched_at": snapshot.get("fetched_at") if snapshot else history.get("updated_at", ""),
        "items": current_items,
        "history": history_items,
        "error": error_message,
        "message": "공개 SSD 가격표를 읽어 평균가를 로컬에 누적 저장합니다. 과거 다운로드 데이터는 멤버십 영역이라 앱 실행 이후부터 추이가 쌓입니다.",
    }


def build_dram_price_payload(force_refresh: bool = False) -> dict[str, Any]:
    history = load_dram_price_history()
    should_fetch = force_refresh or not history.get("updated_at")
    if not should_fetch:
        try:
            updated_at = datetime.fromisoformat(str(history.get("updated_at")))
            should_fetch = updated_at < datetime.now() - timedelta(hours=3)
        except Exception:
            should_fetch = True
    snapshot: dict[str, Any] | None = None
    error_message = ""
    if should_fetch:
        try:
            response = requests.get(DRAM_PRICE_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=20)
            response.raise_for_status()
            snapshot = parse_trendforce_dram_prices(response.text)
            if not snapshot.get("items"):
                raise ValueError("가격표 항목을 찾지 못했습니다.")
            history = merge_dram_price_history(snapshot)
        except Exception as exc:
            error_message = str(exc)
    current_items = snapshot.get("items", []) if snapshot else []
    if not current_items:
        latest_by_key: dict[str, dict[str, Any]] = {}
        for key, entry in (history.get("items") or {}).items():
            points = entry.get("points") or []
            if not points:
                continue
            latest = points[-1]
            meta = entry.get("meta") or {}
            latest_by_key[key] = {
                **meta,
                "source_update": latest.get("date"),
                "session_average": latest.get("value"),
                "session_high": latest.get("session_high"),
                "session_low": latest.get("session_low"),
                "change_pct": latest.get("change_pct"),
            }
        current_items = list(latest_by_key.values())
    history_items = []
    for key, entry in (history.get("items") or {}).items():
        meta = entry.get("meta") or {}
        history_items.append({**meta, "points": entry.get("points") or []})
    history_items.sort(key=lambda item: (str(item.get("section") or ""), str(item.get("item") or "")))
    return {
        "source": "trendforce_dramexchange",
        "source_label": "DramExchange / TrendForce 공개 가격표",
        "source_url": DRAM_PRICE_URL,
        "fetched_at": snapshot.get("fetched_at") if snapshot else history.get("updated_at", ""),
        "items": current_items,
        "history": history_items,
        "error": error_message,
        "message": "공개 가격표를 읽어 세션 평균가를 로컬에 누적 저장합니다. 과거 다운로드 데이터는 멤버십 영역이라 앱 실행 이후부터 추이가 쌓입니다.",
    }


def autosize_worksheet_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        width = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[letter].width = min(max(width + 2, 12), 26)


def scale_ratio_for_display(value: Any) -> float | None:
    number = to_float(value)
    return round(number * 1000.0, 1) if number is not None else None


def create_sector_snapshot_workbook(payload: dict[str, Any]) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Sector Summary"
    stock_sheet = workbook.create_sheet("Stock Detail")
    input_sheet = workbook.create_sheet("Export Info")

    header_fill = PatternFill(fill_type="solid", fgColor="2D4F73")
    header_font = Font(color="FFFFFF", bold=True)

    summary_sheet["A1"] = "Sector Snapshot Export"
    summary_sheet["A1"].font = Font(size=16, bold=True)
    summary_sheet["A3"] = "As Of"
    summary_sheet["B3"] = payload.get("as_of_date", "")
    summary_sheet["A4"] = "Sector Count"
    summary_sheet["B4"] = payload.get("summary", {}).get("sector_count", 0)
    summary_sheet["A5"] = "Stock Count"
    summary_sheet["B5"] = payload.get("summary", {}).get("stock_count", 0)

    sector_columns = [
        ("A7", "Sector"),
        ("B7", "Stock Count"),
        ("C7", "Market Cap (100M KRW)"),
        ("D7", "Strength Score"),
        ("E7", "W Return %"),
        ("F7", "W-1 Return %"),
        ("G7", "1M Return %"),
        ("H7", "3M Return %"),
        ("I7", "YTD Return %"),
        ("J7", "Avg Trading Value / Market Cap x1000"),
        ("K7", "Foreigner Net Buy / Market Cap x1000"),
    ]
    for cell_ref, value in sector_columns:
        summary_sheet[cell_ref] = value
        summary_sheet[cell_ref].fill = header_fill
        summary_sheet[cell_ref].font = header_font

    for index, row in enumerate(payload.get("sector_rows", []), start=8):
        summary_sheet[f"A{index}"] = row.get("sector", "")
        summary_sheet[f"B{index}"] = row.get("stock_count")
        summary_sheet[f"C{index}"] = row.get("market_cap_100m")
        summary_sheet[f"D{index}"] = row.get("strength_score")
        summary_sheet[f"E{index}"] = row.get("w_return_pct")
        summary_sheet[f"F{index}"] = row.get("w1_return_pct")
        summary_sheet[f"G{index}"] = row.get("m1_return_pct")
        summary_sheet[f"H{index}"] = row.get("m3_return_pct")
        summary_sheet[f"I{index}"] = row.get("ytd_return_pct")
        summary_sheet[f"J{index}"] = scale_ratio_for_display(row.get("avg_trading_value_marcap_pct"))
        summary_sheet[f"K{index}"] = scale_ratio_for_display(row.get("foreigner_net_value_marcap_pct"))

    stock_headers = [
        "Sector",
        "Stock",
        "Code",
        "Market",
        "As Of",
        "Current Price",
        "Market Cap (100M KRW)",
        "Strength Score",
        "W Return %",
        "W-1 Return %",
        "1M Return %",
        "3M Return %",
        "YTD Return %",
        "Avg Trading Value / Market Cap x1000",
        "Foreigner Net Buy / Market Cap x1000",
    ]
    for column_index, header in enumerate(stock_headers, start=1):
        cell = stock_sheet.cell(row=1, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_index, row in enumerate(payload.get("stock_rows", []), start=2):
        values = [
            row.get("sector"),
            row.get("stock_name"),
            row.get("stock_code"),
            row.get("market"),
            row.get("as_of_date"),
            row.get("current_price"),
            row.get("market_cap_100m"),
            row.get("strength_score"),
            row.get("w_return_pct"),
            row.get("w1_return_pct"),
            row.get("m1_return_pct"),
            row.get("m3_return_pct"),
            row.get("ytd_return_pct"),
            scale_ratio_for_display(row.get("avg_trading_value_marcap_pct")),
            scale_ratio_for_display(row.get("foreigner_net_value_marcap_pct")),
        ]
        for column_index, value in enumerate(values, start=1):
            stock_sheet.cell(row=row_index, column=column_index, value=value)

    input_sheet["A1"] = "Errors"
    input_sheet["A1"].fill = header_fill
    input_sheet["A1"].font = header_font
    for index, error in enumerate(payload.get("errors", []), start=2):
        input_sheet[f"A{index}"] = error

    for worksheet in [summary_sheet, stock_sheet, input_sheet]:
        worksheet.freeze_panes = "A2"
        autosize_worksheet_columns(worksheet)
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = EXPORT_DIR / f"sector_snapshot_{timestamp}.xlsx"
    workbook.save(output_path)
    return output_path


def list_screening_files(limit: int | None = None) -> list[Path]:
    if SCREENING_SQL_ONLY and SCREENING_FAST_DB_PATH.exists():
        try:
            with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
                rows = conn.execute(
                    """
                    SELECT file_date_key, file_name
                    FROM file_meta
                    ORDER BY file_date_key DESC
                    """
                ).fetchall()
            candidates_sql = []
            for date_key, file_name in rows:
                if file_name:
                    candidates_sql.append(SCREENING_DIR / str(file_name))
                else:
                    candidates_sql.append(SCREENING_DIR / f"{date_key}_데일리_기업스크리닝.xlsx")
            if limit is None:
                return candidates_sql
            return candidates_sql[:limit]
        except Exception:
            # SQL 조회 실패 시에만 파일 시스템 경로로 폴백
            pass

    def _is_valid_screening_file(path: Path) -> bool:
        return is_valid_excel_file_header(path)

    def _name_date_key(path: Path) -> tuple[int, str]:
        match = re.search(r"(20\d{6})", path.name)
        if match:
            return (int(match.group(1)), path.name)
        return (0, path.name)

    candidates = sorted(
        [
            path
            for path in SCREENING_DIR.glob("*\ub370\uc77c\ub9ac_\uae30\uc5c5\uc2a4\ud06c\ub9ac\ub2dd.xls*")
            if (
                not path.name.startswith("~$")
                and path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}
                and _is_valid_screening_file(path)
            )
        ],
        key=_name_date_key,
        reverse=True,
    )
    if limit is None:
        return candidates
    return candidates[:limit]


def is_valid_excel_file_header(path: Path) -> bool:
    try:
        with path.open("rb") as f:
            sig = f.read(4)
        ext = path.suffix.lower()
        if ext in {".xlsx", ".xlsm"}:
            if not sig.startswith(b"PK"):
                return False
            try:
                with zipfile.ZipFile(path, "r") as zf:
                    names = set(zf.namelist())
                    if "[Content_Types].xml" not in names:
                        return False
                    bad_member = zf.testzip()
                    if bad_member is not None:
                        return False
            except Exception:
                return False
            return True
        if ext == ".xls":
            return sig == b"\xD0\xCF\x11\xE0"
        return False
    except Exception:
        return False


def get_latest_screening_file() -> Path:
    candidates = list_screening_files(limit=1)
    if not candidates:
        raise FileNotFoundError("\ub370\uc77c\ub9ac \uae30\uc5c5\uc2a4\ud06c\ub9ac\ub2dd \ud30c\uc77c\uc744 \ucc3e\uc744 \uc218 \uc5c6\uc2b5\ub2c8\ub2e4.")
    return candidates[0]


def list_recent_screening_files(limit: int = RECENT_SCREENING_LOOKBACK) -> list[Path]:
    return list_screening_files(limit=limit)


def parse_portfolio_blocks() -> list[dict[str, Any]]:
    workbook_path = safe_copy_to_temp(PORTFOLIO_PATH)
    df = pd.read_excel(workbook_path, sheet_name=PORTFOLIO_SHEET, header=None, engine=excel_engine_for_path(workbook_path))

    blocks: list[dict[str, Any]] = []
    current_date: datetime | None = None
    current_sector = ""
    current_mode = "initial"
    seed_capital: float | None = None
    rows: list[PositionRow] = []

    for _, row in df.iterrows():
        values = list(row.tolist())
        date_candidate = parse_date_label(values[5] if len(values) > 5 else None)
        if date_candidate:
            if current_date and rows:
                blocks.append(
                    {
                        "rebalance_date": current_date.strftime("%Y-%m-%d"),
                        "seed_capital": seed_capital,
                        "items": [item.__dict__ for item in rows],
                    }
                )
            current_date = date_candidate
            current_sector = ""
            current_mode = "initial"
            seed_capital = None
            rows = []
            continue

        if current_date is None:
            continue

        sector = values[1] if len(values) > 1 else None
        name = values[2] if len(values) > 2 else None
        col3 = values[3] if len(values) > 3 else None
        col4 = values[4] if len(values) > 4 else None
        note = values[5] if len(values) > 5 else None

        normalized_sector = normalize_text(sector)
        normalized_name = normalize_text(name)

        if normalized_sector == ROW_TOTAL:
            if rows:
                blocks.append(
                    {
                        "rebalance_date": current_date.strftime("%Y-%m-%d"),
                        "seed_capital": seed_capital,
                        "items": [item.__dict__ for item in rows],
                    }
                )
            current_date = None
            current_sector = ""
            current_mode = "initial"
            seed_capital = None
            rows = []
            continue

        if normalized_sector == ROW_SEED:
            seed_capital = to_float(name)
            continue

        if normalized_sector == ROW_SECTOR:
            col3_header = normalize_text(col3)
            if col3_header == HEADER_PREV:
                current_mode = "change"
            elif col3_header == HEADER_WEIGHT:
                current_mode = "initial"
            continue

        if sector is not None and not pd.isna(sector) and normalized_sector not in {ROW_SEED, ROW_TOTAL}:
            current_sector = str(sector).strip()

        if pd.isna(name) or not str(name).strip():
            continue

        if current_mode == "change":
            prev_weight = to_float(col3) or 0.0
            target_weight = to_float(col4) or 0.0
        else:
            prev_weight = 0.0
            target_weight = to_float(col3) or 0.0

        if prev_weight == 0 and target_weight == 0:
            continue

        code, resolved_name = resolve_stock(str(name).strip())
        rows.append(
            PositionRow(
                sector=current_sector,
                stock_name=str(name).strip(),
                stock_code=code,
                resolved_name=resolved_name,
                prev_weight=prev_weight,
                target_weight=target_weight,
                note="" if pd.isna(note) else str(note).strip(),
            )
        )

    deduped: list[dict[str, Any]] = []
    seen_dates: set[str] = set()
    for block in blocks:
        if block["rebalance_date"] in seen_dates:
            continue
        seen_dates.add(block["rebalance_date"])
        deduped.append(block)

    for index, block in enumerate(deduped):
        if index == 0:
            block["holdings"] = [item for item in block["items"] if item["target_weight"] > 0]
            continue

        all_prev_zero = all(abs(float(item["prev_weight"])) < 1e-12 for item in block["items"])
        if not all_prev_zero:
            block["holdings"] = [item for item in block["items"] if item["target_weight"] > 0]
            continue

        prev_block = deduped[index - 1]
        prev_map = {
            item["stock_code"]: item
            for item in prev_block["items"]
            if item.get("stock_code") and float(item.get("target_weight", 0.0)) > 0
        }
        current_map = {
            item["stock_code"]: dict(item)
            for item in block["items"]
            if item.get("stock_code")
        }

        rebuilt_items: list[dict[str, Any]] = []
        touched_codes = sorted(set(prev_map) | set(current_map))
        for code in touched_codes:
            if code in current_map:
                item = current_map[code]
                item["prev_weight"] = float(prev_map.get(code, {}).get("target_weight", 0.0))
                rebuilt_items.append(item)
            else:
                previous_item = prev_map[code]
                rebuilt_items.append(
                    {
                        "sector": previous_item.get("sector", ""),
                        "stock_name": previous_item.get("stock_name", ""),
                        "stock_code": previous_item.get("stock_code"),
                        "resolved_name": previous_item.get("resolved_name", previous_item.get("stock_name", "")),
                        "prev_weight": float(previous_item.get("target_weight", 0.0)),
                        "target_weight": 0.0,
                        "note": "",
                    }
                )

        block["items"] = rebuilt_items
        block["holdings"] = [item for item in rebuilt_items if item["target_weight"] > 0]

    for block in deduped:
        if "holdings" not in block:
            block["holdings"] = [item for item in block["items"] if item["target_weight"] > 0]
    sector_db = load_sector_db()
    for block in deduped:
        block["items"] = [enrich_portfolio_item_metadata(dict(item), sector_db) for item in block.get("items", [])]
        block["holdings"] = [
            enrich_portfolio_item_metadata(dict(item), sector_db)
            for item in block.get("holdings", [])
            if float(item.get("target_weight", 0.0) or 0.0) > 0
        ]
    return deduped


def fetch_price_panel(codes: list[str], start: str, end: str) -> tuple[dict[str, pd.DataFrame], pd.DatetimeIndex]:
    frames: dict[str, pd.DataFrame] = {}
    all_dates: pd.DatetimeIndex | None = None
    for code in sorted(set(codes)):
        try:
            data = fdr.DataReader(code, start, end)
        except Exception:
            continue
        if data.empty or "Open" not in data.columns or "Close" not in data.columns:
            continue
        data = data[["Open", "Close"]].sort_index().copy()
        frames[code] = data
        all_dates = data.index if all_dates is None else all_dates.union(data.index)
    return frames, (all_dates if all_dates is not None else pd.DatetimeIndex([]))


def align_rebalance_date(target_date: str, calendar: pd.DatetimeIndex) -> pd.Timestamp | None:
    if calendar.empty:
        return None
    ts = pd.Timestamp(target_date)
    later = calendar[calendar >= ts]
    if len(later) == 0:
        return None
    return pd.Timestamp(later[0])


def get_price(frame: dict[str, pd.DataFrame], code: str, date: pd.Timestamp, column: str) -> float | None:
    stock_frame = frame.get(code)
    if stock_frame is None or date not in stock_frame.index:
        return None
    value = stock_frame.at[date, column]
    if pd.isna(value):
        return None
    return float(value)


def sum_position_value(positions: dict[str, float], price_frames: dict[str, pd.DataFrame], date: pd.Timestamp, column: str) -> float:
    total = 0.0
    for code, shares in positions.items():
        price = get_price(price_frames, code, date, column)
        if price is not None:
            total += shares * price
    return total


def make_trade_items(
    items: list[dict[str, Any]],
    positions_before: dict[str, float],
    positions_after: dict[str, float],
    nav_open_before_trade: float,
    price_frames: dict[str, pd.DataFrame],
    executed_date: pd.Timestamp,
) -> list[dict[str, Any]]:
    trade_items: list[dict[str, Any]] = []
    by_code = {item["stock_code"]: item for item in items if item.get("stock_code")}
    for code, meta in by_code.items():
        open_price = get_price(price_frames, code, executed_date, "Open")
        if open_price is None:
            continue
        before_shares = positions_before.get(code, 0.0)
        after_shares = positions_after.get(code, 0.0)
        delta_shares = after_shares - before_shares
        if abs(delta_shares) < 1e-12:
            continue
        trade_items.append(
            {
                "stock_code": code,
                "stock_name": meta.get("stock_name", code),
                "resolved_name": meta.get("resolved_name", meta.get("stock_name", code)),
                "sector": meta.get("sector", ""),
                "note": meta.get("note", ""),
                "prev_weight": round(float(meta.get("prev_weight", 0.0)), 4),
                "target_weight": round(float(meta.get("target_weight", 0.0)), 4),
                "delta_weight": round(float(meta.get("target_weight", 0.0) - meta.get("prev_weight", 0.0)), 4),
                "trade_action": "buy" if delta_shares > 0 else "sell",
                "trade_price": round(open_price, 2),
                "delta_shares": round(delta_shares, 6),
                "trade_value": round(delta_shares * open_price, 2),
                "target_value": round(nav_open_before_trade * float(meta.get("target_weight", 0.0)), 2),
            }
        )
    return trade_items


def build_daily_contributors(
    date: pd.Timestamp,
    prev_close_nav: float,
    prev_positions: dict[str, float],
    post_trade_positions: dict[str, float],
    price_frames: dict[str, pd.DataFrame],
    prev_close_lookup: dict[str, float],
    row_meta: dict[str, dict[str, Any]],
) -> list[dict[str, Any]]:
    contributors: list[dict[str, Any]] = []
    for code in sorted(set(prev_positions) | set(post_trade_positions)):
        open_price = get_price(price_frames, code, date, "Open")
        close_price = get_price(price_frames, code, date, "Close")
        prev_close_price = prev_close_lookup.get(code)
        if open_price is None or close_price is None or prev_close_price is None:
            continue
        before_shares = prev_positions.get(code, 0.0)
        after_shares = post_trade_positions.get(code, 0.0)
        overnight_pnl = before_shares * (open_price - prev_close_price)
        intraday_pnl = after_shares * (close_price - open_price)
        total_pnl = overnight_pnl + intraday_pnl
        if abs(total_pnl) < 1e-10:
            continue
        meta = row_meta.get(code, {})
        contributors.append(
            {
                "stock_code": code,
                "stock_name": meta.get("stock_name", code),
                "resolved_name": meta.get("resolved_name", meta.get("stock_name", code)),
                "sector": meta.get("sector", ""),
                "note": meta.get("note", ""),
                "overnight_pct_points": round((overnight_pnl / prev_close_nav) * 100, 3),
                "intraday_pct_points": round((intraday_pnl / prev_close_nav) * 100, 3),
                "total_pct_points": round((total_pnl / prev_close_nav) * 100, 3),
                "close_change_pct": round(((close_price / prev_close_price) - 1) * 100, 2),
            }
        )
    contributors.sort(key=lambda item: abs(item["total_pct_points"]), reverse=True)
    return contributors


def build_portfolio_trade_analysis(
    trade_events: list[dict[str, Any]],
    daily_details: list[dict[str, Any]],
    open_positions: dict[str, float],
    position_meta: dict[str, dict[str, Any]],
    price_frames: dict[str, pd.DataFrame],
    last_date: pd.Timestamp | None,
) -> dict[str, Any]:
    lots_by_code: dict[str, list[dict[str, Any]]] = {}
    summary_by_code: dict[str, dict[str, Any]] = {}
    round_trips: list[dict[str, Any]] = []

    def stock_summary(code: str, meta: dict[str, Any] | None = None) -> dict[str, Any]:
        meta = meta or position_meta.get(code, {})
        if code not in summary_by_code:
            summary_by_code[code] = {
                "stock_code": code,
                "stock_name": meta.get("stock_name", code),
                "resolved_name": meta.get("resolved_name", meta.get("stock_name", code)),
                "sector": meta.get("sector", ""),
                "buy_value": 0.0,
                "sell_value": 0.0,
                "realized_pnl": 0.0,
                "realized_buy_cost": 0.0,
                "realized_sell_value": 0.0,
                "unrealized_pnl": 0.0,
                "open_cost": 0.0,
                "open_value": 0.0,
                "buy_shares": 0.0,
                "sell_shares": 0.0,
                "trade_count": 0,
                "first_buy_date": "",
                "last_sell_date": "",
                "avg_buy_price": None,
                "avg_sell_price": None,
                "contribution_pct_points": 0.0,
            }
        else:
            row = summary_by_code[code]
            if meta:
                row["stock_name"] = row.get("stock_name") or meta.get("stock_name", code)
                row["resolved_name"] = row.get("resolved_name") or meta.get("resolved_name", meta.get("stock_name", code))
                row["sector"] = row.get("sector") or meta.get("sector", "")
        return summary_by_code[code]

    for detail in daily_details:
        for contributor in detail.get("contributors") or []:
            code = str(contributor.get("stock_code") or "").strip()
            if not code:
                continue
            row = stock_summary(code, contributor)
            row["contribution_pct_points"] += float(contributor.get("total_pct_points") or 0.0)

    for trade in sorted(trade_events, key=lambda item: (str(item.get("executed_date") or ""), str(item.get("stock_code") or ""))):
        code = str(trade.get("stock_code") or "").strip()
        if not code:
            continue
        meta = trade
        row = stock_summary(code, meta)
        shares_delta = float(trade.get("delta_shares") or 0.0)
        price = float(trade.get("trade_price") or 0.0)
        if abs(shares_delta) < 1e-12 or price <= 0:
            continue
        trade_date = str(trade.get("executed_date") or "")
        row["trade_count"] += 1
        if shares_delta > 0:
            cost = shares_delta * price
            lots_by_code.setdefault(code, []).append({"shares": shares_delta, "price": price, "date": trade_date, "cost": cost})
            row["buy_value"] += cost
            row["buy_shares"] += shares_delta
            if not row["first_buy_date"] or trade_date < row["first_buy_date"]:
                row["first_buy_date"] = trade_date
            continue

        sell_shares_remaining = abs(shares_delta)
        sell_value_total = sell_shares_remaining * price
        realized_cost = 0.0
        matched_shares = 0.0
        buy_dates: list[str] = []
        weighted_buy_price_sum = 0.0
        lots = lots_by_code.setdefault(code, [])
        while sell_shares_remaining > 1e-12 and lots:
            lot = lots[0]
            match_shares = min(float(lot.get("shares") or 0.0), sell_shares_remaining)
            if match_shares <= 0:
                lots.pop(0)
                continue
            lot_price = float(lot.get("price") or 0.0)
            realized_cost += match_shares * lot_price
            weighted_buy_price_sum += match_shares * lot_price
            matched_shares += match_shares
            if lot.get("date"):
                buy_dates.append(str(lot.get("date")))
            lot["shares"] = float(lot.get("shares") or 0.0) - match_shares
            sell_shares_remaining -= match_shares
            if lot["shares"] <= 1e-12:
                lots.pop(0)
        if sell_shares_remaining > 1e-9:
            realized_cost += sell_shares_remaining * price
            weighted_buy_price_sum += sell_shares_remaining * price
            matched_shares += sell_shares_remaining
            sell_shares_remaining = 0.0
        realized_sell_value = matched_shares * price
        realized_pnl = realized_sell_value - realized_cost
        row["sell_value"] += sell_value_total
        row["sell_shares"] += abs(shares_delta)
        row["realized_buy_cost"] += realized_cost
        row["realized_sell_value"] += realized_sell_value
        row["realized_pnl"] += realized_pnl
        row["last_sell_date"] = trade_date
        round_trips.append(
            {
                "stock_code": code,
                "stock_name": row.get("stock_name") or code,
                "resolved_name": row.get("resolved_name") or row.get("stock_name") or code,
                "sector": row.get("sector") or "",
                "buy_date": min(buy_dates) if buy_dates else "",
                "sell_date": trade_date,
                "avg_buy_price": round(weighted_buy_price_sum / matched_shares, 2) if matched_shares else None,
                "sell_price": round(price, 2),
                "shares": round(matched_shares, 4),
                "buy_cost": round(realized_cost, 2),
                "sell_value": round(realized_sell_value, 2),
                "realized_pnl": round(realized_pnl, 2),
                "return_pct": round((realized_pnl / realized_cost) * 100, 2) if realized_cost else None,
            }
        )

    if last_date is not None:
        for code, shares in open_positions.items():
            if abs(float(shares or 0.0)) < 1e-12:
                continue
            row = stock_summary(code, position_meta.get(code, {}))
            close_price = get_price(price_frames, code, last_date, "Close")
            if close_price is None:
                continue
            open_cost = sum(float(lot.get("shares") or 0.0) * float(lot.get("price") or 0.0) for lot in lots_by_code.get(code, []))
            open_value = float(shares) * float(close_price)
            row["open_cost"] = open_cost
            row["open_value"] = open_value
            row["unrealized_pnl"] = open_value - open_cost

    stock_rows = []
    for row in summary_by_code.values():
        buy_shares = float(row.get("buy_shares") or 0.0)
        sell_shares = float(row.get("sell_shares") or 0.0)
        buy_value = float(row.get("buy_value") or 0.0)
        sell_value = float(row.get("sell_value") or 0.0)
        total_pnl = float(row.get("realized_pnl") or 0.0) + float(row.get("unrealized_pnl") or 0.0)
        total_cost = float(row.get("realized_buy_cost") or 0.0) + float(row.get("open_cost") or 0.0)
        stock_rows.append(
            {
                **row,
                "buy_value": round(buy_value, 2),
                "sell_value": round(sell_value, 2),
                "realized_pnl": round(float(row.get("realized_pnl") or 0.0), 2),
                "unrealized_pnl": round(float(row.get("unrealized_pnl") or 0.0), 2),
                "total_pnl": round(total_pnl, 2),
                "total_return_pct": round((total_pnl / total_cost) * 100, 2) if total_cost else None,
                "avg_buy_price": round(buy_value / buy_shares, 2) if buy_shares else None,
                "avg_sell_price": round(sell_value / sell_shares, 2) if sell_shares else None,
                "contribution_pct_points": round(float(row.get("contribution_pct_points") or 0.0), 3),
                "is_tail_candidate": total_pnl < 0 or float(row.get("contribution_pct_points") or 0.0) < 0,
            }
        )
    stock_rows.sort(key=lambda item: item.get("total_pnl") or 0)

    sector_map: dict[str, dict[str, Any]] = {}
    for row in stock_rows:
        sector = str(row.get("sector") or "기타")
        bucket = sector_map.setdefault(
            sector,
            {"sector": sector, "stock_count": 0, "total_pnl": 0.0, "realized_pnl": 0.0, "unrealized_pnl": 0.0, "contribution_pct_points": 0.0, "winners": 0, "losers": 0},
        )
        bucket["stock_count"] += 1
        bucket["total_pnl"] += float(row.get("total_pnl") or 0.0)
        bucket["realized_pnl"] += float(row.get("realized_pnl") or 0.0)
        bucket["unrealized_pnl"] += float(row.get("unrealized_pnl") or 0.0)
        bucket["contribution_pct_points"] += float(row.get("contribution_pct_points") or 0.0)
        if float(row.get("total_pnl") or 0.0) >= 0:
            bucket["winners"] += 1
        else:
            bucket["losers"] += 1
    sector_rows = [
        {
            **row,
            "total_pnl": round(float(row.get("total_pnl") or 0.0), 2),
            "realized_pnl": round(float(row.get("realized_pnl") or 0.0), 2),
            "unrealized_pnl": round(float(row.get("unrealized_pnl") or 0.0), 2),
            "contribution_pct_points": round(float(row.get("contribution_pct_points") or 0.0), 3),
        }
        for row in sector_map.values()
    ]
    sector_rows.sort(key=lambda item: item.get("total_pnl") or 0, reverse=True)

    tail_rows = [row for row in stock_rows if row.get("is_tail_candidate")]
    tail_loss_pnl = sum(min(0.0, float(row.get("total_pnl") or 0.0)) for row in tail_rows)
    tail_loss_contribution = sum(min(0.0, float(row.get("contribution_pct_points") or 0.0)) for row in tail_rows)
    return {
        "stocks": stock_rows,
        "sectors": sector_rows,
        "round_trips": sorted(round_trips, key=lambda item: item.get("realized_pnl") or 0)[:80],
        "tail_summary": {
            "tail_stock_count": len(tail_rows),
            "tail_loss_pnl": round(tail_loss_pnl, 2),
            "tail_loss_contribution_pct_points": round(tail_loss_contribution, 3),
            "worst_stocks": tail_rows[:12],
        },
        "top_contributors": sorted(stock_rows, key=lambda item: item.get("total_pnl") or 0, reverse=True)[:12],
    }


def calculate_portfolio_performance() -> dict[str, Any]:
    blocks = parse_portfolio_blocks()
    if not blocks:
        return {"series": [], "benchmark": [], "summary": {}, "rebalances": [], "daily_details": []}

    all_codes = sorted(
        {
            item["stock_code"]
            for block in blocks
            for item in block["items"]
            if item.get("stock_code")
        }
    )
    start_date = blocks[0]["rebalance_date"]
    end_date = datetime.today().strftime("%Y-%m-%d")
    price_frames, calendar = fetch_price_panel(all_codes, start_date, end_date)
    if calendar.empty:
        return {"series": [], "benchmark": [], "summary": {}, "rebalances": blocks, "daily_details": []}

    initial_capital = float(blocks[0].get("seed_capital") or 100000000)
    positions: dict[str, float] = {}
    position_meta: dict[str, dict[str, Any]] = {}
    cash = initial_capital
    prev_close_nav = initial_capital
    prev_close_lookup: dict[str, float] = {}

    block_schedule: dict[pd.Timestamp, dict[str, Any]] = {}
    for block in blocks:
        aligned = align_rebalance_date(block["rebalance_date"], calendar)
        if aligned is not None:
            copied = dict(block)
            copied["executed_date"] = aligned.strftime("%Y-%m-%d")
            block_schedule[aligned] = copied

    series: list[dict[str, Any]] = []
    rebalances_output: list[dict[str, Any]] = []
    daily_details: list[dict[str, Any]] = []
    daily_allocations: list[dict[str, Any]] = []
    all_trade_events: list[dict[str, Any]] = []
    current_allocation_items: list[dict[str, Any]] = blocks[0]["holdings"] if blocks else []

    for date in calendar:
        date = pd.Timestamp(date)
        positions_before_open = dict(positions)
        nav_open_before_trade = cash + sum_position_value(positions_before_open, price_frames, date, "Open")
        row_meta: dict[str, dict[str, Any]] = {
            code: enrich_portfolio_item_metadata(
                {
                    "stock_code": code,
                    "stock_name": position_meta.get(code, {}).get("stock_name", code),
                    "resolved_name": position_meta.get(code, {}).get("resolved_name", position_meta.get(code, {}).get("stock_name", code)),
                    "sector": position_meta.get(code, {}).get("sector", ""),
                    "note": position_meta.get(code, {}).get("note", ""),
                }
            )
            for code in positions_before_open
        }
        executed_trades: list[dict[str, Any]] = []
        block = block_schedule.get(date)

        if block:
            positions_after_trade = dict(positions_before_open)
            current_allocation_items = block["holdings"]
            block_items_for_trade = [dict(item) for item in block["items"]]
            block_codes = {
                str(item.get("stock_code") or "").strip()
                for item in block_items_for_trade
                if str(item.get("stock_code") or "").strip()
            }
            for code, shares in positions_before_open.items():
                if code in block_codes or abs(float(shares or 0.0)) < 1e-12:
                    continue
                open_price = get_price(price_frames, code, date, "Open")
                if open_price is None or open_price <= 0 or nav_open_before_trade == 0:
                    continue
                previous_meta = position_meta.get(code, {})
                prev_weight = (float(shares) * float(open_price)) / nav_open_before_trade
                block_items_for_trade.append(
                    {
                        "sector": previous_meta.get("sector", ""),
                        "stock_name": previous_meta.get("stock_name", code),
                        "stock_code": code,
                        "resolved_name": previous_meta.get("resolved_name", previous_meta.get("stock_name", code)),
                        "prev_weight": prev_weight,
                        "target_weight": 0.0,
                        "note": previous_meta.get("note", ""),
                    }
                )

            for item in block["holdings"]:
                code = item.get("stock_code")
                if not code:
                    continue
                row_meta[code] = item
                position_meta[code] = item
            for item in block_items_for_trade:
                code = item.get("stock_code")
                if not code:
                    continue
                row_meta[code] = item
                position_meta[code] = item
                open_price = get_price(price_frames, code, date, "Open")
                if open_price is None or open_price <= 0:
                    continue
                delta_weight = float(item.get("target_weight", 0.0) - item.get("prev_weight", 0.0))
                if abs(delta_weight) < 1e-12:
                    continue
                if float(item.get("target_weight", 0.0) or 0.0) <= 0 and positions_after_trade.get(code):
                    delta_shares = -float(positions_after_trade.get(code, 0.0) or 0.0)
                else:
                    delta_value = nav_open_before_trade * delta_weight
                    delta_shares = delta_value / open_price
                if abs(delta_shares) < 1e-12:
                    continue
                cash -= delta_shares * open_price
                next_shares = positions_after_trade.get(code, 0.0) + delta_shares
                if abs(next_shares) < 1e-12:
                    positions_after_trade.pop(code, None)
                    if float(item.get("target_weight", 0.0) or 0.0) <= 0:
                        position_meta.pop(code, None)
                else:
                    positions_after_trade[code] = next_shares

            executed_trades = make_trade_items(
                block_items_for_trade,
                positions_before_open,
                positions_after_trade,
                nav_open_before_trade,
                price_frames,
                date,
            )
            for trade in executed_trades:
                trade["executed_date"] = date.strftime("%Y-%m-%d")
                trade["nav_open_before_trade"] = round(nav_open_before_trade, 2)
            all_trade_events.extend(executed_trades)
            positions = positions_after_trade

        nav_close = cash + sum_position_value(positions, price_frames, date, "Close")
        daily_return_pct = ((nav_close / prev_close_nav) - 1) * 100 if prev_close_nav else 0.0
        contributors = build_daily_contributors(
            date,
            prev_close_nav,
            positions_before_open,
            positions,
            price_frames,
            prev_close_lookup,
            row_meta,
        )

        daily_details.append(
            {
                "date": date.strftime("%Y-%m-%d"),
                "daily_return_pct": round(daily_return_pct, 3),
                "nav_open_before_trade": round(nav_open_before_trade, 2),
                "nav_close": round(nav_close, 2),
                "cash_close": round(cash, 2),
                "trade_count": len(executed_trades),
                "trades": executed_trades,
                "contributors": contributors,
            }
        )
        series.append(
            {
                "date": date.strftime("%Y-%m-%d"),
                "value": round((nav_close / initial_capital) * 100, 2),
                "return_pct": round(((nav_close / initial_capital) - 1) * 100, 2),
                "nav": round(nav_close, 2),
                "daily_return_pct": round(daily_return_pct, 3),
            }
        )

        sector_weights: dict[str, float] = {}
        stock_weights: dict[str, float] = {}
        for item in current_allocation_items:
            weight_pct = float(item.get("target_weight", 0.0) or 0.0) * 100
            if abs(weight_pct) < 1e-9:
                continue
            resolved_name = item.get("resolved_name") or item.get("stock_name") or item.get("stock_code") or "기타"
            sector_name = item.get("sector") or "기타"
            stock_weights[resolved_name] = stock_weights.get(resolved_name, 0.0) + weight_pct
            sector_weights[sector_name] = sector_weights.get(sector_name, 0.0) + weight_pct
        daily_allocations.append(
            {
                "date": date.strftime("%Y-%m-%d"),
                "stock_weights": {key: round(value, 3) for key, value in stock_weights.items()},
                "sector_weights": {key: round(value, 3) for key, value in sector_weights.items()},
            }
        )

        if block:
            rebalances_output.append(
                {
                    "rebalance_date": block["rebalance_date"],
                    "executed_date": block["executed_date"],
                    "holdings": block["holdings"],
                    "items": block_items_for_trade,
                    "trades": executed_trades,
                }
            )

        prev_close_nav = nav_close
        prev_close_lookup = {
            code: get_price(price_frames, code, date, "Close")
            for code in positions
            if get_price(price_frames, code, date, "Close") is not None
        }

    benchmark_series: list[dict[str, Any]] = []
    try:
        benchmark_data = fdr.DataReader("KS11", start_date, end_date)
        if not benchmark_data.empty:
            base_close = float(benchmark_data["Close"].iloc[0])
            for idx, row in benchmark_data.iterrows():
                benchmark_series.append(
                    {
                        "date": pd.Timestamp(idx).strftime("%Y-%m-%d"),
                        "value": round((float(row["Close"]) / base_close) * 100, 2),
                        "return_pct": round(((float(row["Close"]) / base_close) - 1) * 100, 2),
                    }
                )
    except Exception:
        benchmark_series = []

    final_nav = series[-1]["nav"] if series else initial_capital
    summary = {
        "start_date": start_date,
        "end_date": series[-1]["date"] if series else start_date,
        "initial_capital": initial_capital,
        "final_nav": round(final_nav, 2),
        "final_value": round((final_nav / initial_capital) * 100, 2),
        "total_return_pct": round(((final_nav / initial_capital) - 1) * 100, 2),
        "rebalance_count": len(rebalances_output),
        "holding_count_latest": len(rebalances_output[-1]["holdings"]) if rebalances_output else 0,
    }
    last_trade_date = pd.Timestamp(calendar[-1]) if len(calendar) else None
    trade_analysis = build_portfolio_trade_analysis(
        all_trade_events,
        daily_details,
        positions,
        position_meta,
        price_frames,
        last_trade_date,
    )

    return {
        "series": series,
        "benchmark": benchmark_series,
        "summary": summary,
        "rebalances": rebalances_output,
        "daily_details": daily_details,
        "daily_allocations": daily_allocations,
        "trade_analysis": trade_analysis,
    }


def portfolio_performance_cache_key() -> str:
    try:
        portfolio_mtime = PORTFOLIO_PATH.stat().st_mtime if PORTFOLIO_PATH.exists() else 0
    except Exception:
        portfolio_mtime = 0
    try:
        sector_mtime = SECTOR_DB_PATH.stat().st_mtime if SECTOR_DB_PATH.exists() else 0
    except Exception:
        sector_mtime = 0
    return f"{portfolio_mtime:.0f}:{sector_mtime:.0f}:{date.today().isoformat()}"


def get_cached_portfolio_performance(force_refresh: bool = False) -> dict[str, Any]:
    cache_key = portfolio_performance_cache_key()
    now = time.time()
    with PORTFOLIO_PERFORMANCE_CACHE_LOCK:
        cached_key = PORTFOLIO_PERFORMANCE_CACHE.get("key")
        cached_at = float(PORTFOLIO_PERFORMANCE_CACHE.get("cached_at") or 0)
        cached_payload = PORTFOLIO_PERFORMANCE_CACHE.get("payload")
        if (
            not force_refresh
            and cached_key == cache_key
            and isinstance(cached_payload, dict)
            and now - cached_at < PORTFOLIO_PERFORMANCE_CACHE_TTL_SECONDS
        ):
            return cached_payload
        payload = calculate_portfolio_performance()
        PORTFOLIO_PERFORMANCE_CACHE.clear()
        PORTFOLIO_PERFORMANCE_CACHE.update({"key": cache_key, "cached_at": now, "payload": payload})
        return payload


def summarize_portfolio_diagnostic_series(
    key: str,
    name: str,
    nav_values: list[float],
    start_date: str,
    end_date: str,
    actual_final_return: float,
    description: str,
) -> dict[str, Any]:
    final_value = nav_values[-1] if nav_values else 100.0
    total_return = final_value - 100.0
    return {
        "key": key,
        "name": name,
        "total_return_pct": round(total_return, 2),
        "excess_vs_actual_pct": round(total_return - actual_final_return, 2),
        "mdd_pct": max_drawdown_pct(nav_values),
        "cagr_pct": annualized_return_pct(total_return, start_date, end_date),
        "description": description,
    }


def build_portfolio_tail_rule_context(screening_summaries: list[dict[str, Any]], sector_db: dict[str, Any]) -> dict[str, Any]:
    dates = [str(item.get("file_date") or "") for item in screening_summaries if item.get("file_date")]
    summary_by_date = {str(item.get("file_date") or ""): item for item in screening_summaries if item.get("file_date")}
    context_cache: dict[str, dict[str, Any]] = {}

    def latest_summary(signal_date: str) -> tuple[str, dict[str, Any] | None]:
        eligible = [item for item in dates if item <= signal_date]
        if not eligible:
            return "", None
        picked = eligible[-1]
        return picked, summary_by_date.get(picked)

    def context(signal_date: str) -> dict[str, Any]:
        picked_date, summary = latest_summary(signal_date)
        if not picked_date or not summary:
            return {"date": "", "stock_scores": {}, "stock_rows": {}, "active_sectors": set(), "sector_ranks": {}}
        if picked_date in context_cache:
            return context_cache[picked_date]
        stock_scores: dict[str, float] = {}
        stock_rows: dict[str, dict[str, Any]] = {}
        for row in summary.get("qualified_stocks", []):
            if not isinstance(row, dict):
                continue
            code = normalize_stock_code_value(row.get("stock_code"))
            name = normalize_text(row.get("stock_name") or row.get("resolved_name") or "")
            score = to_float(row.get("score"))
            if code:
                stock_rows[code] = row
                stock_scores[code] = float(score or 0.0)
            if name:
                stock_rows[name] = row
                stock_scores[name] = float(score or 0.0)
        signals = build_sector_rotation_signals(summary, min_score=50.0, sector_db=sector_db)
        sector_ranks = {str(signal.get("sector") or ""): index + 1 for index, signal in enumerate(signals)}
        active_sectors = {
            str(signal.get("sector") or "")
            for index, signal in enumerate(signals[:7])
            if float(signal.get("avg_score") or 0.0) >= 55.0
            and (
                int(signal.get("strong_count") or 0) >= 2
                or float(signal.get("turnover_ratio_pct") or 0.0) >= 0.35
            )
            and float(signal.get("strength_score") or 0.0) >= 50.0
        }
        payload = {
            "date": picked_date,
            "stock_scores": stock_scores,
            "stock_rows": stock_rows,
            "active_sectors": active_sectors,
            "sector_ranks": sector_ranks,
        }
        context_cache[picked_date] = payload
        return payload

    return {"context": context}


def classify_portfolio_tail_rule(contributor: dict[str, Any], signal_date: str, context: dict[str, Any]) -> dict[str, Any]:
    code = normalize_stock_code_value(contributor.get("stock_code"))
    name = str(contributor.get("resolved_name") or contributor.get("stock_name") or "").strip()
    sector = str(contributor.get("sector") or "").strip() or "기타"
    normalized_name = normalize_text(name)
    score = None
    stock_scores = context.get("stock_scores") if isinstance(context.get("stock_scores"), dict) else {}
    if code and code in stock_scores:
        score = to_float(stock_scores.get(code))
    elif normalized_name and normalized_name in stock_scores:
        score = to_float(stock_scores.get(normalized_name))
    technical = stock_technical_on_date(code, signal_date) if code else {"above_ma20": None, "disparity": None}
    above_ma20 = technical.get("above_ma20")
    disparity = to_float(technical.get("disparity"))
    active_sectors = context.get("active_sectors") if isinstance(context.get("active_sectors"), set) else set()
    sector_active = sector in active_sectors
    position_state = str(contributor.get("position_state") or "held")
    prior_contribution = float(contributor.get("prior_contribution_pct_points") or 0.0)
    has_profit_buffer = prior_contribution >= 2.0
    reasons: list[str] = []
    mode = "보유 유지"
    if not sector_active:
        reasons.append("주도 섹터 아님")
    if score is None:
        reasons.append("주도주 점수 없음")
    elif score < 55:
        reasons.append(f"점수 {score:.1f}<55")
    if above_ma20 is False:
        reasons.append("20일선 아래")
    elif above_ma20 is None:
        reasons.append("20일선 확인불가")
    if disparity is not None and disparity > 112:
        reasons.append(f"이격도 {disparity:.1f}%")
    if disparity is not None and disparity < 96:
        reasons.append(f"20일선 하방 이격 {disparity:.1f}%")
    if position_state in {"new", "increase"}:
        reasons.append("신규/증액 엄격 기준")
    if has_profit_buffer:
        reasons.append(f"수익 버퍼 {prior_contribution:.1f}%p")

    # New money should be strict: avoid adding to names without sector/stock
    # confirmation. Existing winners get more room so a normal pullback does not
    # erase the trend-following edge.
    is_tail = False
    if position_state in {"new", "increase"}:
        mode = "신규/증액 차단"
        if above_ma20 is False:
            is_tail = True
        elif not sector_active:
            is_tail = True
        elif score is None or score < 65:
            is_tail = True
        elif disparity is not None and disparity < 98:
            is_tail = True
        elif disparity is not None and disparity > 118 and (score is None or score < 75):
            is_tail = True
    elif has_profit_buffer:
        mode = "수익 보유 완충"
        if above_ma20 is False and disparity is not None and disparity < 94 and (score is None or score < 55):
            is_tail = True
        elif not sector_active and score is not None and score < 40:
            is_tail = True
        elif score is not None and score < 35:
            is_tail = True
    elif above_ma20 is False and (not sector_active or score is None or score < 65):
        mode = "기존 보유 점검"
        is_tail = True
    elif not sector_active and score is not None and score < 60:
        mode = "기존 보유 점검"
        is_tail = True
    elif score is not None and score < 50:
        mode = "기존 보유 점검"
        is_tail = True
    elif disparity is not None and disparity > 115 and score is not None and score < 70 and not sector_active:
        mode = "과열 감액 후보"
        is_tail = True
    return {
        "is_tail": is_tail,
        "reasons": reasons,
        "score": score,
        "above_ma20": above_ma20,
        "disparity": disparity,
        "sector_active": sector_active,
        "position_state": position_state,
        "prior_contribution_pct_points": prior_contribution,
        "has_profit_buffer": has_profit_buffer,
        "mode": mode,
    }


def build_portfolio_diagnostic() -> dict[str, Any]:
    def as_list(value: Any) -> list[Any]:
        return value if isinstance(value, list) else []

    performance = calculate_portfolio_performance()
    series = as_list(performance.get("series"))
    daily_details = as_list(performance.get("daily_details"))
    daily_allocations = as_list(performance.get("daily_allocations"))
    trade_analysis = performance.get("trade_analysis") or {}
    summary = performance.get("summary") or {}
    if not series or not daily_details:
        return {
            "rows": [],
            "scenarios": [],
            "summary": {},
            "diagnosis": {},
            "message": "포트폴리오 수익 데이터가 없습니다.",
        }

    tail_stocks = [
        row for row in as_list(trade_analysis.get("stocks"))
        if row.get("is_tail_candidate")
    ]
    tail_codes = {str(row.get("stock_code") or "").strip() for row in tail_stocks if row.get("stock_code")}
    tail_names = {
        str(row.get("resolved_name") or row.get("stock_name") or "").strip()
        for row in tail_stocks
        if str(row.get("resolved_name") or row.get("stock_name") or "").strip()
    }
    allocation_by_date = {
        str(row.get("date") or ""): row
        for row in daily_allocations
        if row.get("date")
    }

    navs = {
        "actual": 100.0,
        "tail_cash": 100.0,
        "tail_index": 100.0,
        "rule_tail_cash": 100.0,
        "rule_tail_index": 100.0,
        "market_filter": 100.0,
        "sector_gate": 100.0,
        "combined": 100.0,
    }
    nav_history = {key: [100.0] for key in navs}
    rows: list[dict[str, Any]] = []
    previous_date = str(daily_details[0].get("date") or series[0].get("date") or "")
    tail_weight_sum = 0.0
    tail_weight_count = 0
    market_off_days = 0
    one_market_days = 0
    sector_gate_days = 0
    rule_tail_weight_sum = 0.0
    rule_tail_weight_count = 0
    rule_tail_hit_count = 0
    rule_tail_examples: list[dict[str, Any]] = []
    rule_tail_mode_counts: dict[str, int] = {}
    prior_contribution_by_code: dict[str, float] = {}
    previous_holding_codes: set[str] = set()
    signal_active_by_date: dict[str, bool] = {}
    tail_rule_context_builder: dict[str, Any] = {"context": lambda _date: {}}
    try:
        sector_db = load_sector_db()
        screening_summaries = screening_backtest_source_summaries()
        tail_rule_context_builder = build_portfolio_tail_rule_context(screening_summaries, sector_db)
        for screening_summary in screening_summaries:
            signal_date = str(screening_summary.get("file_date") or "")
            if not signal_date:
                continue
            signals = build_sector_rotation_signals(screening_summary, min_score=50, sector_db=sector_db)
            signal_active_by_date[signal_date] = any(
                float(signal.get("avg_score") or 0.0) >= 55.0
                and int(signal.get("stock_count") or 0) >= 2
                and float(signal.get("strength_score") or 0.0) >= 50.0
                for signal in signals[:5]
            )
    except Exception:
        signal_active_by_date = {}

    for detail in daily_details:
        date_text = str(detail.get("date") or "")
        if not date_text:
            continue
        actual_daily_return = float(detail.get("daily_return_pct") or 0.0) / 100.0
        tail_contribution = 0.0
        for contributor in as_list(detail.get("contributors")):
            code = str(contributor.get("stock_code") or "").strip()
            name = str(contributor.get("resolved_name") or contributor.get("stock_name") or "").strip()
            if code in tail_codes or name in tail_names:
                tail_contribution += float(contributor.get("total_pct_points") or 0.0) / 100.0

        allocation = allocation_by_date.get(date_text) or {}
        stock_weights = allocation.get("stock_weights") if isinstance(allocation.get("stock_weights"), dict) else {}
        tail_weight = 0.0
        for name, weight_pct in stock_weights.items():
            if str(name or "").strip() in tail_names:
                tail_weight += max(0.0, float(weight_pct or 0.0)) / 100.0
        if tail_weight > 0:
            tail_weight_sum += tail_weight
            tail_weight_count += 1

        signal_date_for_rule = previous_date or date_text
        rule_context = tail_rule_context_builder.get("context", lambda _date: {})(signal_date_for_rule)
        rule_tail_contribution = 0.0
        rule_tail_names: set[str] = set()
        trades_by_code: dict[str, dict[str, Any]] = {}
        for trade in as_list(detail.get("trades")):
            code = normalize_stock_code_value(trade.get("stock_code"))
            if code:
                trades_by_code[code] = trade
        for contributor in as_list(detail.get("contributors")):
            contributor_for_rule = dict(contributor)
            code = normalize_stock_code_value(contributor_for_rule.get("stock_code"))
            trade = trades_by_code.get(code or "")
            if trade:
                prev_weight = float(trade.get("prev_weight") or 0.0)
                delta_weight = float(trade.get("delta_weight") or 0.0)
                if prev_weight <= 0 and delta_weight > 0:
                    position_state = "new"
                elif delta_weight > 0:
                    position_state = "increase"
                elif delta_weight < 0:
                    position_state = "decrease"
                else:
                    position_state = "held"
            elif code and code in previous_holding_codes:
                position_state = "held"
            else:
                position_state = "new"
            contributor_for_rule["position_state"] = position_state
            contributor_for_rule["prior_contribution_pct_points"] = prior_contribution_by_code.get(code or "", 0.0)
            rule = classify_portfolio_tail_rule(contributor_for_rule, signal_date_for_rule, rule_context)
            if not rule.get("is_tail"):
                continue
            rule_tail_contribution += float(contributor.get("total_pct_points") or 0.0) / 100.0
            stock_name = str(contributor.get("resolved_name") or contributor.get("stock_name") or "").strip()
            if stock_name:
                rule_tail_names.add(stock_name)
            rule_tail_hit_count += 1
            rule_mode = str(rule.get("mode") or "실전 꼬리룰")
            rule_tail_mode_counts[rule_mode] = rule_tail_mode_counts.get(rule_mode, 0) + 1
            if len(rule_tail_examples) < 80:
                rule_tail_examples.append(
                    {
                        "date": date_text,
                        "signal_date": signal_date_for_rule,
                        "stock_code": contributor.get("stock_code"),
                        "stock_name": stock_name,
                        "sector": contributor.get("sector") or "기타",
                        "score": rule.get("score"),
                        "above_ma20": rule.get("above_ma20"),
                        "disparity": rule.get("disparity"),
                        "position_state": rule.get("position_state"),
                        "prior_contribution_pct_points": round(float(rule.get("prior_contribution_pct_points") or 0.0), 3),
                        "rule_mode": rule.get("mode"),
                        "contribution_pct_points": round(float(contributor.get("total_pct_points") or 0.0), 3),
                        "reason": ", ".join(rule.get("reasons") or []) or "약한 꼬리 조건",
                    }
                )
        rule_tail_weight = 0.0
        for name, weight_pct in stock_weights.items():
            if str(name or "").strip() in rule_tail_names:
                rule_tail_weight += max(0.0, float(weight_pct or 0.0)) / 100.0
        if rule_tail_weight > 0:
            rule_tail_weight_sum += rule_tail_weight
            rule_tail_weight_count += 1

        index_daily_return = 0.0
        if previous_date and previous_date != date_text:
            index_daily_return = (
                index_return_between("KS11", previous_date, date_text)
                + index_return_between("KQ11", previous_date, date_text)
            ) / 2.0
        market_state = advanced_market_timing_state(previous_date or date_text)
        market_multiplier = float(market_state.get("multiplier") or 0.0)
        if market_multiplier <= 0:
            market_off_days += 1
        elif market_multiplier < 1:
            one_market_days += 1
        sector_signal_active = bool(signal_active_by_date.get(previous_date or date_text))
        if sector_signal_active:
            sector_gate_days += 1

        scenario_returns = {
            "actual": actual_daily_return,
            "tail_cash": actual_daily_return - tail_contribution,
            "tail_index": actual_daily_return - tail_contribution + tail_weight * index_daily_return,
            "rule_tail_cash": actual_daily_return - rule_tail_contribution,
            "rule_tail_index": actual_daily_return - rule_tail_contribution + rule_tail_weight * index_daily_return,
            "market_filter": actual_daily_return * market_multiplier,
            "sector_gate": actual_daily_return * market_multiplier if sector_signal_active else 0.0,
            "combined": (actual_daily_return - rule_tail_contribution + rule_tail_weight * index_daily_return) * market_multiplier,
        }
        row = {
            "date": date_text,
            "actual_daily_return_pct": round(scenario_returns["actual"] * 100.0, 3),
            "tail_contribution_pct": round(tail_contribution * 100.0, 3),
            "tail_weight_pct": round(tail_weight * 100.0, 2),
            "rule_tail_contribution_pct": round(rule_tail_contribution * 100.0, 3),
            "rule_tail_weight_pct": round(rule_tail_weight * 100.0, 2),
            "index_replacement_return_pct": round(index_daily_return * 100.0, 3),
            "market_multiplier": round(market_multiplier, 2),
            "market_label": market_state.get("label") or "",
            "sector_signal_active": sector_signal_active,
        }
        for key, daily_return in scenario_returns.items():
            navs[key] *= 1.0 + daily_return
            nav_history[key].append(navs[key])
            row[f"{key}_return_pct"] = round(navs[key] - 100.0, 3)
            row[f"{key}_daily_return_pct"] = round(daily_return * 100.0, 3)
        rows.append(row)
        for contributor in as_list(detail.get("contributors")):
            code = normalize_stock_code_value(contributor.get("stock_code"))
            if not code:
                continue
            prior_contribution_by_code[code] = prior_contribution_by_code.get(code, 0.0) + float(contributor.get("total_pct_points") or 0.0)
        next_holding_codes = set(previous_holding_codes)
        for trade in as_list(detail.get("trades")):
            code = normalize_stock_code_value(trade.get("stock_code"))
            if not code:
                continue
            if float(trade.get("target_weight") or 0.0) > 0:
                next_holding_codes.add(code)
            else:
                next_holding_codes.discard(code)
        previous_holding_codes = next_holding_codes
        previous_date = date_text

    start_date = str(rows[0].get("date") or summary.get("start_date") or "")
    end_date = str(rows[-1].get("date") or summary.get("end_date") or "")
    actual_final_return = float(rows[-1].get("actual_return_pct") or 0.0) if rows else 0.0
    scenarios = [
        summarize_portfolio_diagnostic_series(
            "actual",
            "현재 실제 방식",
            nav_history["actual"],
            start_date,
            end_date,
            actual_final_return,
            "엑셀 비중 그대로 시초가 매수/매도, 보유 수량 유지 기준입니다.",
        ),
        summarize_portfolio_diagnostic_series(
            "tail_cash",
            "꼬리 제거 · 현금 보유",
            nav_history["tail_cash"],
            start_date,
            end_date,
            actual_final_return,
            "손실 또는 음의 기여도가 누적된 꼬리 종목 기여도를 제거하고 남는 비중은 현금으로 둔 근사치입니다.",
        ),
        summarize_portfolio_diagnostic_series(
            "tail_index",
            "꼬리 제거 · 지수 대체",
            nav_history["tail_index"],
            start_date,
            end_date,
            actual_final_return,
            "꼬리 종목 비중을 KOSPI/KOSDAQ 50:50 수익으로 대체한 근사치입니다.",
        ),
        summarize_portfolio_diagnostic_series(
            "rule_tail_cash",
            "실전 꼬리룰 · 현금 보유",
            nav_history["rule_tail_cash"],
            start_date,
            end_date,
            actual_final_return,
            "당시 알 수 있던 섹터/점수/20일선/이격도 규칙으로 약한 종목을 제외하고 남는 비중은 현금으로 둔 시나리오입니다.",
        ),
        summarize_portfolio_diagnostic_series(
            "rule_tail_index",
            "실전 꼬리룰 · 지수 대체",
            nav_history["rule_tail_index"],
            start_date,
            end_date,
            actual_final_return,
            "당시 꼬리룰에 걸린 종목 비중을 KOSPI/KOSDAQ 50:50으로 대체한 실전형 시나리오입니다.",
        ),
        summarize_portfolio_diagnostic_series(
            "market_filter",
            "시장 20일선 필터",
            nav_history["market_filter"],
            start_date,
            end_date,
            actual_final_return,
            "양 지수 모두 20일선 아래면 주식 노출 0%, 하나만 위면 55%만 노출하는 필터를 적용한 근사치입니다.",
        ),
        summarize_portfolio_diagnostic_series(
            "sector_gate",
            "섹터신호 게이트",
            nav_history["sector_gate"],
            start_date,
            end_date,
            actual_final_return,
            "시장 필터가 켜져 있고 강한 섹터 신호가 있는 날에만 현재 추세추종 노출을 허용한 근사치입니다.",
        ),
        summarize_portfolio_diagnostic_series(
            "combined",
            "꼬리 지수대체 + 시장필터",
            nav_history["combined"],
            start_date,
            end_date,
            actual_final_return,
            "약한 꼬리는 지수로 대체하고 시장 필터까지 적용한 보수적 개선 시나리오입니다.",
        ),
    ]

    tail_summary = trade_analysis.get("tail_summary") or {}
    avg_tail_weight = (tail_weight_sum / tail_weight_count * 100.0) if tail_weight_count else 0.0
    avg_rule_tail_weight = (rule_tail_weight_sum / rule_tail_weight_count * 100.0) if rule_tail_weight_count else 0.0
    diagnosis = {
        "tail_summary": tail_summary,
        "tail_rule": {
            "name": "실전 꼬리 판별 규칙",
            "avg_rule_tail_weight_pct": round(avg_rule_tail_weight, 2),
            "hit_count": rule_tail_hit_count,
            "examples": rule_tail_examples[:40],
            "mode_counts": [
                {"mode": mode, "count": count}
                for mode, count in sorted(rule_tail_mode_counts.items(), key=lambda item: item[1], reverse=True)
            ],
            "rules": [
                "신규 편입·증액 종목은 주도 섹터, 20일선, 주도주 점수를 모두 엄격하게 확인합니다.",
                "이미 누적 기여가 2%p 이상인 보유 주도주는 정상 조정 구간에서 바로 자르지 않고 완충 기준을 적용합니다.",
                "수익 버퍼가 없는 기존 보유 종목은 20일선 이탈, 약한 점수, 비주도 섹터 조건을 조합해 감액 후보로 봅니다.",
                "과열 이격도 종목은 신규 진입에는 엄격하게, 기존 수익 종목에는 감액 후보로만 해석합니다.",
                "이 규칙은 매도 이후 손익을 보지 않고 해당 날짜에 알 수 있던 주도주/가격 데이터만 사용합니다.",
            ],
        },
        "top_contributors": as_list(trade_analysis.get("top_contributors"))[:12],
        "worst_stocks": as_list(tail_summary.get("worst_stocks"))[:12],
        "sector_contribution": as_list(trade_analysis.get("sectors"))[:20],
        "round_trips": as_list(trade_analysis.get("round_trips"))[:80],
        "avg_tail_weight_pct": round(avg_tail_weight, 2),
        "avg_rule_tail_weight_pct": round(avg_rule_tail_weight, 2),
        "market_off_days": market_off_days,
        "one_market_days": one_market_days,
        "sector_gate_days": sector_gate_days,
        "feedback": [
            "사후 손실 종목 제거는 참고용이고, 실제 운용에는 섹터/점수/20일선/이격도 기반의 실전 꼬리룰을 우선 봐야 합니다.",
            "현재 방식은 주도주를 길게 가져가는 장점이 있으므로, 전량매도보다 당시 약한 꼬리 비중을 줄이고 남는 비중을 지수로 대체하는 쪽이 성향에 더 잘 맞습니다.",
            "20일선 아래 종목은 추세추종 관점에서 가장 명확한 꼬리 후보입니다. 다만 과열 이격도 종목은 이미 보유 중이면 전량 제외보다 감액 룰로 발전시키는 편이 좋습니다.",
            "신규 편입 기준과 기존 보유 청산 기준을 분리했기 때문에, 실전 꼬리룰은 이제 '나쁜 종목을 전부 제거'보다 '새 돈은 엄격하게, 기존 승자는 느슨하게'에 가깝습니다.",
        ],
    }
    return {
        "mode": "portfolio_diagnostic",
        "strategy_name": "현재 방식 진단",
        "description": "포트폴리오 수익 페이지의 실제 비중 데이터를 기준으로 꼬리 제거, 지수 대체, 시장 필터 적용 시나리오를 비교합니다.",
        "start_date": start_date,
        "end_date": end_date,
        "rows": rows,
        "scenarios": scenarios,
        "summary": {
            **summary,
            "scenario_count": len(scenarios),
            "avg_tail_weight_pct": round(avg_tail_weight, 2),
            "avg_rule_tail_weight_pct": round(avg_rule_tail_weight, 2),
            "tail_stock_count": tail_summary.get("tail_stock_count"),
            "tail_loss_contribution_pct_points": tail_summary.get("tail_loss_contribution_pct_points"),
            "rule_tail_hit_count": rule_tail_hit_count,
        },
        "diagnosis": diagnosis,
        "sector_signal": {
            "mode": "gate",
            "message": "정밀한 섹터별 종목 편입 백테스트는 고급 섹터 신호 탭에서 별도로 계산합니다.",
        },
    }


def latest_stock_alert_holdings(min_weight_pct: float = 0.0, latest_non_empty: bool = True) -> list[dict[str, Any]]:
    performance = get_cached_portfolio_performance()
    allocations = performance.get("daily_allocations") or []
    rebalances = performance.get("rebalances") or []
    if not allocations:
        return []
    selected_allocation = allocations[-1] if isinstance(allocations[-1], dict) else {}
    if latest_non_empty and not (selected_allocation.get("stock_weights") or {}):
        for allocation in reversed(allocations):
            if isinstance(allocation, dict) and (allocation.get("stock_weights") or {}):
                selected_allocation = allocation
                break
    stock_weights = selected_allocation.get("stock_weights") or {}
    meta_by_name: dict[str, dict[str, Any]] = {}
    for rebalance in reversed(rebalances):
        for item in rebalance.get("holdings") or []:
            name = str(item.get("resolved_name") or item.get("stock_name") or item.get("stock_code") or "").strip()
            if name and name not in meta_by_name:
                meta_by_name[name] = item
        if meta_by_name:
            break
    holdings: list[dict[str, Any]] = []
    seen: set[str] = set()
    for name, weight in sorted(stock_weights.items(), key=lambda item: float(item[1] or 0), reverse=True):
        try:
            weight_pct = float(weight or 0.0)
        except Exception:
            weight_pct = 0.0
        if weight_pct < min_weight_pct:
            continue
        meta = meta_by_name.get(str(name)) or {}
        listing = resolve_stock_payload(name=str(name)) or {}
        code = str(meta.get("stock_code") or listing.get("code") or "").strip()
        key = code or str(name)
        if not key or key in seen:
            continue
        seen.add(key)
        holdings.append(
            {
                "name": str(name),
                "code": code,
                "weight_pct": round(weight_pct, 3),
                "sector": str(meta.get("sector") or ""),
                "source_date": selected_allocation.get("date") or "",
            }
        )
    return holdings


def github_api_request(method: str, repo: str, path: str, token: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
    if "/" not in repo:
        raise ValueError("GitHub repository는 owner/name 형식이어야 합니다.")
    data = json.dumps(payload).encode("utf-8") if payload is not None else None
    headers = {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {token}",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "stock-dashboard-local",
    }
    if data is not None:
        headers["Content-Type"] = "application/json"
    response = requests.request(
        method,
        f"https://api.github.com/repos/{repo}{path}",
        headers=headers,
        data=data,
        timeout=25,
    )
    if response.status_code >= 400:
        raise ValueError(f"GitHub API 오류 {response.status_code}: {response.text[:300]}")
    return response.json() if response.text else {}


def encrypt_github_actions_secret(public_key: str, value: str) -> str:
    try:
        from nacl import encoding, public
    except Exception as exc:
        raise RuntimeError("PyNaCl 패키지가 필요합니다. requirements.txt 설치 후 다시 시도해 주세요.") from exc
    key = public.PublicKey(public_key.encode("utf-8"), encoding.Base64Encoder())
    sealed_box = public.SealedBox(key)
    encrypted = sealed_box.encrypt(value.encode("utf-8"))
    return base64.b64encode(encrypted).decode("utf-8")


def update_github_actions_secret(repo: str, token: str, name: str, value: str) -> dict[str, Any]:
    key_payload = github_api_request("GET", repo, "/actions/secrets/public-key", token)
    encrypted_value = encrypt_github_actions_secret(str(key_payload.get("key") or ""), value)
    github_api_request(
        "PUT",
        repo,
        f"/actions/secrets/{quote(name)}",
        token,
        {"encrypted_value": encrypted_value, "key_id": key_payload.get("key_id")},
    )
    return {"secret_name": name, "repository": repo}


def sync_stock_alert_holdings_secret() -> dict[str, Any]:
    settings = stock_alert_settings()
    if not settings["github_repository"] or not settings["github_token"]:
        raise ValueError("GitHub repository/token 설정이 필요합니다.")
    holdings = latest_stock_alert_holdings(latest_non_empty=True)
    if not holdings:
        raise ValueError("동기화할 보유 종목이 없습니다.")
    value = json.dumps(holdings, ensure_ascii=False, separators=(",", ":"))
    update_github_actions_secret(settings["github_repository"], settings["github_token"], "STOCK_ALERT_HOLDINGS_JSON", value)
    snapshot = {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "repository": settings["github_repository"],
        "source_date": holdings[0].get("source_date") or "",
        "holdings": holdings,
    }
    STOCK_ALERT_HOLDINGS_SNAPSHOT_PATH.write_text(json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8")
    return {
        "ok": True,
        "repository": settings["github_repository"],
        "secret_name": "STOCK_ALERT_HOLDINGS_JSON",
        "holding_count": len(holdings),
        "source_date": holdings[0].get("source_date") or "",
        "updated_at": snapshot["updated_at"],
    }


def detect_telegram_chat_id() -> dict[str, Any]:
    settings = stock_alert_settings()
    bot_token = settings["telegram_bot_token"]
    if not bot_token:
        raise ValueError("Telegram bot token 설정이 필요합니다.")
    response = requests.get(f"https://api.telegram.org/bot{bot_token}/getUpdates", timeout=20)
    response.raise_for_status()
    payload = response.json()
    if not payload.get("ok"):
        raise ValueError(str(payload))
    candidates: list[dict[str, Any]] = []
    for update in payload.get("result") or []:
        message = update.get("message") or update.get("edited_message") or update.get("channel_post")
        if not isinstance(message, dict):
            continue
        chat = message.get("chat") if isinstance(message.get("chat"), dict) else {}
        chat_id = str(chat.get("id") or "").strip()
        if not chat_id:
            continue
        candidates.append(
            {
                "chat_id": chat_id,
                "type": chat.get("type") or "",
                "title": chat.get("title") or chat.get("username") or chat.get("first_name") or "",
                "date": message.get("date") or 0,
                "text": str(message.get("text") or "")[:80],
            }
        )
    if not candidates:
        raise ValueError("봇에 들어온 메시지가 없습니다. Telegram에서 봇에게 /start를 보낸 뒤 다시 시도해 주세요.")
    candidates.sort(key=lambda item: int(item.get("date") or 0), reverse=True)
    picked = candidates[0]
    settings_payload = load_settings()
    stock_alert = settings_payload.get("stock_alert") if isinstance(settings_payload.get("stock_alert"), dict) else {}
    stock_alert["telegram_chat_id"] = picked["chat_id"]
    settings_payload["stock_alert"] = stock_alert
    save_settings(settings_payload)
    return {"ok": True, "picked": picked, "candidates": candidates[:5]}


def sync_stock_alert_telegram_secrets() -> dict[str, Any]:
    settings = stock_alert_settings()
    if not settings["github_repository"] or not settings["github_token"]:
        raise ValueError("GitHub repository/token 설정이 필요합니다.")
    if not settings["telegram_bot_token"] or not settings["telegram_chat_id"]:
        raise ValueError("Telegram bot token/chat id 설정이 필요합니다.")
    update_github_actions_secret(
        settings["github_repository"],
        settings["github_token"],
        "TELEGRAM_BOT_TOKEN",
        settings["telegram_bot_token"],
    )
    update_github_actions_secret(
        settings["github_repository"],
        settings["github_token"],
        "TELEGRAM_CHAT_ID",
        settings["telegram_chat_id"],
    )
    return {
        "ok": True,
        "repository": settings["github_repository"],
        "synced": ["TELEGRAM_BOT_TOKEN", "TELEGRAM_CHAT_ID"],
    }


def split_theme_tokens(note: Any) -> list[str]:
    if note is None or (isinstance(note, float) and math.isnan(note)):
        return []
    text = str(note).strip()
    if not text:
        return []
    text = re.sub(r"[\?\!\(\)\[\]]", "", text)
    tokens = [token.strip() for token in re.split(r"[,/\n;]|쨌", text) if token.strip()]
    return tokens[:3] if tokens else [text]


def build_portfolio_export_payload() -> dict[str, Any]:
    performance = calculate_portfolio_performance()
    series = performance.get("series", [])
    benchmark_lookup = {item["date"]: item for item in performance.get("benchmark", [])}
    details_lookup = {item["date"]: item for item in performance.get("daily_details", [])}
    allocations_lookup = {item["date"]: item for item in performance.get("daily_allocations", [])}

    daily_rows: list[dict[str, Any]] = []
    for item in series:
        date_key = item["date"]
        benchmark_item = benchmark_lookup.get(date_key, {})
        detail_item = details_lookup.get(date_key, {})
        allocation_item = allocations_lookup.get(date_key, {})
        daily_rows.append(
            {
                "date": date_key,
                "portfolio_return_pct": item.get("return_pct", 0),
                "benchmark_return_pct": benchmark_item.get("return_pct", 0),
                "daily_return_pct": item.get("daily_return_pct", 0),
                "nav": item.get("nav", 0),
                "trade_count": detail_item.get("trade_count", 0),
                "stock_total_pct": round(
                    sum(float(value or 0) for value in (allocation_item.get("stock_weights") or {}).values()),
                    3,
                ),
                "sector_total_pct": round(
                    sum(float(value or 0) for value in (allocation_item.get("sector_weights") or {}).values()),
                    3,
                ),
            }
        )

    sector_keys = sorted(
        {
            sector
            for item in performance.get("daily_allocations", [])
            for sector, value in (item.get("sector_weights") or {}).items()
            if abs(float(value or 0)) > 1e-9
        }
    )
    stock_keys = sorted(
        {
            stock
            for item in performance.get("daily_allocations", [])
            for stock, value in (item.get("stock_weights") or {}).items()
            if abs(float(value or 0)) > 1e-9
        }
    )

    sector_rows: list[dict[str, Any]] = []
    stock_rows: list[dict[str, Any]] = []
    for item in performance.get("daily_allocations", []):
        sector_weights = item.get("sector_weights") or {}
        stock_weights = item.get("stock_weights") or {}
        sector_row = {"date": item["date"]}
        stock_row = {"date": item["date"]}
        for key in sector_keys:
            sector_row[key] = float(sector_weights.get(key, 0) or 0)
        for key in stock_keys:
            stock_row[key] = float(stock_weights.get(key, 0) or 0)
        sector_rows.append(sector_row)
        stock_rows.append(stock_row)

    return {
        "summary": performance.get("summary", {}),
        "daily_rows": daily_rows,
        "sector_rows": sector_rows,
        "stock_rows": stock_rows,
        "sector_keys": sector_keys,
        "stock_keys": stock_keys,
    }


def ensure_workspace_node_modules_link() -> None:
    target = BASE_DIR / "node_modules"
    if target.exists():
        return
    if os.name != "nt":
        return
    if not WORKSPACE_NODE_MODULES.exists():
        raise FileNotFoundError(f"Node modules path not found: {WORKSPACE_NODE_MODULES}")
    subprocess.run(
        [
            "powershell",
            "-NoProfile",
            "-Command",
            f"New-Item -ItemType Junction -Path '{target}' -Target '{WORKSPACE_NODE_MODULES}' | Out-Null",
        ],
        check=True,
        cwd=BASE_DIR,
    )


def create_portfolio_export_workbook_openpyxl(payload: dict[str, Any], output_path: Path) -> Path:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    daily_sheet = workbook.create_sheet("Daily")
    sector_sheet = workbook.create_sheet("Sector Weights")
    stock_sheet = workbook.create_sheet("Stock Weights")

    summary = payload.get("summary", {})
    summary_rows = [
        ("Start Date", summary.get("start_date", "")),
        ("End Date", summary.get("end_date", "")),
        ("Initial Capital", summary.get("initial_capital", "")),
        ("Final NAV", summary.get("final_nav", "")),
        ("Total Return %", summary.get("total_return_pct", "")),
        ("Rebalance Count", summary.get("rebalance_count", "")),
    ]
    for row_index, (label, value) in enumerate(summary_rows, start=1):
        summary_sheet.cell(row=row_index, column=1, value=label)
        summary_sheet.cell(row=row_index, column=2, value=value)
    summary_sheet["A1"].font = Font(bold=True)

    daily_headers = [
        "Date",
        "Portfolio Return %",
        "Benchmark Return %",
        "Daily Return %",
        "NAV",
        "Trade Count",
        "Stock Weight Total %",
        "Sector Weight Total %",
    ]
    daily_sheet.append(daily_headers)
    for item in payload.get("daily_rows", []):
        daily_sheet.append(
            [
                item.get("date"),
                item.get("portfolio_return_pct"),
                item.get("benchmark_return_pct"),
                item.get("daily_return_pct"),
                item.get("nav"),
                item.get("trade_count"),
                item.get("stock_total_pct"),
                item.get("sector_total_pct"),
            ]
        )

    def fill_weight_sheet(sheet, rows: list[dict[str, Any]], keys: list[str]) -> None:
        sheet.append(["Date", *keys])
        for item in rows:
            sheet.append([item.get("date"), *[item.get(key, 0) for key in keys]])

    fill_weight_sheet(sector_sheet, payload.get("sector_rows", []), payload.get("sector_keys", []))
    fill_weight_sheet(stock_sheet, payload.get("stock_rows", []), payload.get("stock_keys", []))

    for sheet in [summary_sheet, daily_sheet, sector_sheet, stock_sheet]:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center")
        for column in sheet.columns:
            column_letter = column[0].column_letter
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 28)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")

    if daily_sheet.max_row >= 3:
        line_chart = LineChart()
        line_chart.title = "Portfolio vs KOSPI Return"
        line_chart.y_axis.title = "Return %"
        line_chart.x_axis.title = "Date"
        data = Reference(daily_sheet, min_col=2, max_col=3, min_row=1, max_row=daily_sheet.max_row)
        cats = Reference(daily_sheet, min_col=1, min_row=2, max_row=daily_sheet.max_row)
        line_chart.add_data(data, titles_from_data=True)
        line_chart.set_categories(cats)
        line_chart.height = 9
        line_chart.width = 22
        summary_sheet.add_chart(line_chart, "D2")

    if sector_sheet.max_row >= 3 and sector_sheet.max_column >= 2:
        bar_chart = BarChart()
        bar_chart.type = "col"
        bar_chart.style = 10
        bar_chart.title = "Sector Weight"
        bar_chart.y_axis.title = "Weight %"
        data = Reference(sector_sheet, min_col=2, max_col=min(sector_sheet.max_column, 12), min_row=1, max_row=sector_sheet.max_row)
        cats = Reference(sector_sheet, min_col=1, min_row=2, max_row=sector_sheet.max_row)
        bar_chart.add_data(data, titles_from_data=True)
        bar_chart.set_categories(cats)
        bar_chart.height = 9
        bar_chart.width = 22
        summary_sheet.add_chart(bar_chart, "D20")

    workbook.save(output_path)
    return output_path


def create_portfolio_export_workbook() -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    payload_path = EXPORT_DIR / "portfolio_export_payload.json"
    output_path = EXPORT_DIR / f"portfolio_backtest_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    payload = build_portfolio_export_payload()
    payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    try:
        ensure_workspace_node_modules_link()
        subprocess.run(
            [
                str(WORKSPACE_NODE_EXE),
                str(PORTFOLIO_EXPORT_SCRIPT),
                str(payload_path),
                str(output_path),
            ],
            check=True,
            cwd=BASE_DIR,
        )
    except Exception:
        create_portfolio_export_workbook_openpyxl(payload, output_path)
    return output_path


def canonicalize_theme(note: Any, stock_name: str = "") -> tuple[str, list[str]]:
    candidates = split_theme_tokens(note)
    note_corpus = " ".join(candidates).lower()
    stock_corpus = stock_name.lower()
    for rule in THEME_RULES:
        matched = [keyword for keyword in rule["keywords"] if keyword.lower() in note_corpus]
        if matched:
            return rule["theme"], matched[:3]
    for rule in STOCK_THEME_RULES:
        matched = [keyword for keyword in rule["keywords"] if keyword.lower() in stock_corpus]
        if matched:
            return rule["theme"], matched[:3]
    if candidates:
        return candidates[0], candidates[:3]
    return "\uae30\ud0c0", []


def load_screening_frame(path: Path) -> pd.DataFrame:
    def _date_key_from_path(target_path: Path) -> str:
        m = re.match(r"^(20\d{6})_", target_path.name)
        return m.group(1) if m else ""

    def _try_load_fast_db(target_path: Path) -> pd.DataFrame | None:
        date_key = _date_key_from_path(target_path)
        if not date_key or not SCREENING_FAST_DB_PATH.exists():
            return None
        try:
            with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
                q = """
                SELECT
                    stock_code AS 종목코드,
                    stock_name AS 종목명,
                    sector AS 섹터,
                    industry AS 업종구분,
                    market_cap_100m AS 시가총액,
                    trading_value_100m AS 거래대금,
                    trading_value_100m AS [거래대금.1],
                    change_pct AS 등락률,
                    score_o AS O열점수,
                    avg_1w AS [1W 평균],
                    avg_1m AS [1M 평균],
                    avg_3m AS [3M 평균],
                    sortino_norm AS 소르티노_정규화,
                    score_s AS [종합 점수],
                    note AS 비고
                FROM screening_rows
                WHERE file_date_key = ?
                """
                fast_df = pd.read_sql_query(q, conn, params=[date_key])
            if fast_df is None or fast_df.empty:
                return None
            fast_df["점수"] = pd.to_numeric(fast_df["종합 점수"], errors="coerce")
            fast_df["_market_cap_100m"] = pd.to_numeric(fast_df["시가총액"], errors="coerce")
            fast_df["_code_norm"] = fast_df["종목코드"].map(normalize_stock_code_value)
            return fast_df
        except Exception:
            return None

    fast_df = _try_load_fast_db(path)
    if fast_df is not None and not fast_df.empty:
        return fast_df
    if SCREENING_SQL_ONLY:
        date_key = _date_key_from_path(path)
        raise FileNotFoundError(f"SQL 캐시에 해당 날짜 데이터가 없습니다: {date_key or path.name}")

    try:
        screening_path = safe_copy_to_temp(path)
    except PermissionError:
        # Excel이 파일을 점유 중이면 임시 복사 대신 원본 직접 읽기를 시도한다.
        screening_path = path
    try:
        df = pd.read_excel(screening_path, sheet_name=SCREENING_SHEET, engine=excel_engine_for_path(screening_path))
    except Exception:
        # 신형 Stock_Daily 파일은 첫 번째 시트 기준으로 로드
        df = pd.read_excel(screening_path, sheet_name=0, engine=excel_engine_for_path(screening_path))
    df = df.rename(columns=lambda col: str(col).strip())

    # 신형 Stock_Daily 스키마를 내부 표준 컬럼으로 정규화
    renamed = {
        "종목 이름": "종목명",
        "업종": "업종구분",
        "시총 (억원)": "시가총액",
        "거래대금 (억원)": "거래대금",
        "1W 평균 점수": "1W 평균",
        "1M 평균 점수": "1M 평균",
        "3M 평균 점수": "1M 평균",
        "60일 기준 Sortino 정규화 점수": "소르티노_정규화",
    }
    apply_rename = {k: v for k, v in renamed.items() if k in df.columns}
    if apply_rename:
        df = df.rename(columns=apply_rename)
    has_new_schema_score = "종합 점수" in df.columns
    if has_new_schema_score:
        df["점수"] = pd.to_numeric(df["종합 점수"], errors="coerce")

    if "\uc885\ubaa9\ucf54\ub4dc" in df.columns:
        df["_code_norm"] = df["\uc885\ubaa9\ucf54\ub4dc"].map(normalize_stock_code_value)
    elif df.shape[1] >= 3:
        df["_code_norm"] = df.iloc[:, 2].map(normalize_stock_code_value)
    else:
        df["_code_norm"] = ""

    daily_map: dict[str, dict[str, Any]] = {}
    if not has_new_schema_score:
        # 구형 스키마는 데일리데이터/S열 캐시 기반 로직 사용
        try:
            daily_df = pd.read_excel(screening_path, sheet_name="\ub370\uc77c\ub9ac\ub370\uc774\ud130", engine=excel_engine_for_path(screening_path))
            daily_df = daily_df.rename(columns=lambda col: str(col).strip())
            for _, row in daily_df.iterrows():
                code = normalize_stock_code_value(row.get("\uc885\ubaa9\ucf54\ub4dc"))
                if not code:
                    continue
                daily_map[code] = {
                    "name": row.get("\uc885\ubaa9\uba85"),
                    "industry": row.get("\uc5c5\uc885\uad6c\ubd84"),
                    "market": row.get("\ub300\uc0c1"),
                    "change_pct": row.get("\ub4f1\ub77d\ub960"),
                    "trading_value": row.get("\uac70\ub798\ub300\uae08"),
                    "market_cap": row.get("\uc2dc\uac00\ucd1d\uc561"),
                }
        except Exception:
            daily_map = {}

        # 구형은 S열 계산 결과(value cache)를 점수로 사용
        s_score_map = read_screening_s_values(screening_path)
        df["\uc810\uc218"] = pd.to_numeric(df["_code_norm"].map(s_score_map), errors="coerce")

    def _safe_fill_string_column(
        column_name: str,
        resolver: Callable[[str], Any],
    ) -> None:
        if column_name not in df.columns:
            df[column_name] = ""
        try:
            df[column_name] = df[column_name].astype("object")
        except Exception:
            df[column_name] = df[column_name].astype(str)
        missing = df[column_name].isna() | (df[column_name].astype(str).str.strip() == "")
        if not missing.any():
            return
        for idx in df.index[missing]:
            code = normalize_stock_code_value(df.at[idx, "_code_norm"])
            if not code:
                continue
            try:
                value = resolver(code)
            except Exception:
                continue
            if value is None:
                continue
            text_value = str(value).strip()
            if not text_value:
                continue
            try:
                df.at[idx, column_name] = text_value
            except Exception:
                # 일부 행/셀 dtype 충돌이 있어도 전체 로드는 계속 진행한다.
                continue

    _safe_fill_string_column(
        "\uc885\ubaa9\uba85",
        lambda code: daily_map.get(code, {}).get("name") if daily_map else None,
    )

    missing_name = df["\uc885\ubaa9\uba85"].isna() | (df["\uc885\ubaa9\uba85"].astype(str).str.strip() == "")
    if missing_name.any():
        def _resolve_name_from_code(code: Any) -> str:
            normalized = normalize_stock_code_value(code)
            if not normalized:
                return ""
            resolved_code, resolved_name = resolve_stock(normalized)
            return str(resolved_name or "").strip()

        for idx in df.index[missing_name]:
            code = normalize_stock_code_value(df.at[idx, "_code_norm"])
            if not code:
                continue
            try:
                resolved_name = _resolve_name_from_code(code)
            except Exception:
                continue
            if not resolved_name:
                continue
            try:
                df.at[idx, "\uc885\ubaa9\uba85"] = resolved_name
            except Exception:
                continue

    if "\uc5c5\uc885\uad6c\ubd84" in df.columns:
        _safe_fill_string_column(
            "\uc5c5\uc885\uad6c\ubd84",
            lambda code: daily_map.get(code, {}).get("industry") if daily_map else None,
        )
    if "\ub300\uc0c1" in df.columns:
        _safe_fill_string_column(
            "\ub300\uc0c1",
            lambda code: daily_map.get(code, {}).get("market") if daily_map else None,
        )

    if "\ub4f1\ub77d\ub960" in df.columns:
        change_series = pd.to_numeric(df["\ub4f1\ub77d\ub960"], errors="coerce")
        if daily_map:
            mapped_change = pd.to_numeric(
                df["_code_norm"].map(lambda code: daily_map.get(code, {}).get("change_pct")),
                errors="coerce",
            )
            change_series = change_series.fillna(mapped_change)
        df["\ub4f1\ub77d\ub960"] = change_series
    if "\uc2dc\uac00\ucd1d\uc561.1" in df.columns:
        primary_cap = pd.to_numeric(df["\uc2dc\uac00\ucd1d\uc561.1"], errors="coerce")
        fallback_cap = None
        if "\uc2dc\uac00\ucd1d\uc561" in df.columns:
            if daily_map:
                missing_cap = df["\uc2dc\uac00\ucd1d\uc561"].isna() | (df["\uc2dc\uac00\ucd1d\uc561"].astype(str).str.strip() == "")
                if missing_cap.any():
                    df.loc[missing_cap, "\uc2dc\uac00\ucd1d\uc561"] = df.loc[missing_cap, "_code_norm"].map(lambda code: daily_map.get(code, {}).get("market_cap"))
            fallback_cap = df["\uc2dc\uac00\ucd1d\uc561"].map(parse_korean_number)
        if fallback_cap is not None:
            df["_market_cap_100m"] = primary_cap.where(primary_cap.notna(), fallback_cap)
        else:
            df["_market_cap_100m"] = primary_cap
    elif "\uc2dc\uac00\ucd1d\uc561" in df.columns:
        if daily_map:
            missing_cap = df["\uc2dc\uac00\ucd1d\uc561"].isna() | (df["\uc2dc\uac00\ucd1d\uc561"].astype(str).str.strip() == "")
            if missing_cap.any():
                df.loc[missing_cap, "\uc2dc\uac00\ucd1d\uc561"] = df.loc[missing_cap, "_code_norm"].map(lambda code: daily_map.get(code, {}).get("market_cap"))
        df["_market_cap_100m"] = df["\uc2dc\uac00\ucd1d\uc561"].map(parse_korean_number)
    else:
        df["_market_cap_100m"] = np.nan

    if "\uac70\ub798\ub300\uae08" in df.columns and daily_map:
        missing_tv = df["\uac70\ub798\ub300\uae08"].isna() | (df["\uac70\ub798\ub300\uae08"].astype(str).str.strip() == "")
        if missing_tv.any():
            df.loc[missing_tv, "\uac70\ub798\ub300\uae08"] = df.loc[missing_tv, "_code_norm"].map(lambda code: daily_map.get(code, {}).get("trading_value"))

    return df.dropna(subset=["\uc885\ubaa9\uba85", "\uc810\uc218"]).copy()


def parse_screening_date(path: Path) -> str:
    date_match = re.search(r"(\d{8})", path.name)
    if not date_match:
        return ""
    return datetime.strptime(date_match.group(1), "%Y%m%d").strftime("%Y-%m-%d")


def get_screening_file_by_date(file_date: str | None = None, fallback_latest: bool = False) -> Path:
    if not file_date:
        return get_latest_screening_file()

    digits = re.sub(r"\D", "", str(file_date))
    iso_date = ""
    if len(digits) == 8:
        iso_date = datetime.strptime(digits, "%Y%m%d").strftime("%Y-%m-%d")
    else:
        iso_date = str(file_date).strip()

    for path in list_screening_files():
        if parse_screening_date(path) == iso_date:
            return path

    if fallback_latest:
        return get_latest_screening_file()

    raise FileNotFoundError(f"해당 일자의 데일리 기업스크리닝 파일을 찾을 수 없습니다: {file_date}")


SCREENING_AVERAGE_FILE_RE = re.compile(r"^(20\d{6})_\ub370\uc77c\ub9ac_\uae30\uc5c5\uc2a4\ud06c\ub9ac\ub2dd\.(?:xlsx|xlsm)$", re.IGNORECASE)


def screening_compact_date_from_path(path: Path) -> str:
    match = SCREENING_AVERAGE_FILE_RE.match(path.name)
    if match:
        return match.group(1)
    digits = re.sub(r"\D", "", parse_screening_date(path))
    return digits if re.fullmatch(r"20\d{6}", digits) else ""


def list_screening_average_source_files() -> list[Path]:
    files: list[tuple[str, Path]] = []
    for path in SCREENING_DIR.glob("*\ub370\uc77c\ub9ac_\uae30\uc5c5\uc2a4\ud06c\ub9ac\ub2dd.xls*"):
        if path.name.startswith("~$"):
            continue
        if not is_valid_excel_file_header(path):
            continue
        compact = screening_compact_date_from_path(path)
        if compact:
            files.append((compact, path))
    return [path for _, path in sorted(files, key=lambda item: item[0])]


def screening_header_positions(values: list[Any]) -> dict[str, list[int]]:
    positions: dict[str, list[int]] = {}
    for index, value in enumerate(values, start=1):
        key = str(value or "").strip()
        if not key:
            continue
        positions.setdefault(key, []).append(index)
    return positions


def find_screening_header_row_and_columns(sheet: Any) -> tuple[int, dict[str, list[int]]]:
    max_header_scan = min(getattr(sheet, "max_row", 30) or 30, 30)
    for row_index, values in enumerate(sheet.iter_rows(min_row=1, max_row=max_header_scan, values_only=True), start=1):
        header_values = list(values or [])
        positions = screening_header_positions(header_values)
        if "\uc885\ubaa9\ucf54\ub4dc" in positions and "\uc810\uc218" in positions:
            return row_index, positions
    raise ValueError(f"엑셀에서 '{SCREENING_SHEET}' 시트의 종목코드/점수 헤더를 찾지 못했습니다.")


def screening_score_candidate_columns(header_positions: dict[str, list[int]]) -> list[int]:
    score_columns = header_positions.get("\uc810\uc218", [])
    q_columns = header_positions.get("1M \ud3c9\uade0", [])
    q_column = q_columns[0] if q_columns else 17
    before_average = [column for column in score_columns if column < q_column]
    after_average = [column for column in score_columns if column >= q_column]
    # 기존 VBA는 Q/R 앞의 점수 열을 사용한다. 새 테스트 양식처럼 위치가 바뀐 경우는 뒤쪽 점수 열을 보조로 사용한다.
    return list(reversed(before_average)) + list(reversed(after_average)) or [15]


def extract_excel_digits_number(value: Any) -> float | None:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (int, float, np.integer, np.floating)):
        return float(value)
    digits = re.sub(r"\D", "", str(value))
    return float(digits) if digits else None


def read_screening_scores_from_daily_data_formula(path: Path) -> dict[str, float]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if "\ub370\uc77c\ub9ac\ub370\uc774\ud130" not in workbook.sheetnames:
            return {}
        daily_sheet = workbook["\ub370\uc77c\ub9ac\ub370\uc774\ud130"]
        high_names: set[str] = set()
        if "52\uc8fc\uc2e0\uace0\uac00" in workbook.sheetnames:
            high_sheet = workbook["52\uc8fc\uc2e0\uace0\uac00"]
            for values in high_sheet.iter_rows(min_row=1, values_only=True):
                row_values = list(values or [])
                if len(row_values) >= 3:
                    normalized_name = normalize_text(row_values[2])
                    if normalized_name:
                        high_names.add(normalized_name)

        scores: dict[str, float] = {}
        for values in daily_sheet.iter_rows(min_row=2, values_only=True):
            row_values = list(values or [])
            if len(row_values) < 10:
                continue
            stock_code = normalize_stock_code_value(row_values[1])
            stock_name = normalize_text(row_values[2])
            change_rate = to_float(row_values[5])
            trading_amount_digits = extract_excel_digits_number(row_values[7])
            market_cap_digits = extract_excel_digits_number(row_values[9])
            if not stock_code or change_rate is None or not trading_amount_digits or not market_cap_digits:
                continue
            trading_amount_score_unit = trading_amount_digits / 100000
            if trading_amount_score_unit <= 0 or market_cap_digits <= 0 or (1.1 + change_rate) <= 0:
                continue
            high_adjustment = -4 if stock_name and stock_name in high_names else 4
            try:
                score = (
                    math.log(
                        (trading_amount_score_unit * trading_amount_score_unit)
                        / (market_cap_digits**0.8)
                        * ((1.1 + change_rate) ** 4),
                        1.1,
                    )
                    + high_adjustment
                    - 13
                )
            except (ValueError, ZeroDivisionError, OverflowError):
                continue
            if math.isfinite(score):
                scores[stock_code] = score
        return scores
    finally:
        workbook.close()


def read_screening_s_values(path: Path) -> dict[str, float]:
    """
    '주도주 찾기' 시트 S열(종합점수)의 계산 결과(value cache)만 읽는다.
    수식 문자열은 읽지 않는다.
    """
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    try:
        if SCREENING_SHEET in workbook.sheetnames:
            sheet = workbook[SCREENING_SHEET]
        else:
            fallback_name = next((name for name in workbook.sheetnames if "주도주" in str(name)), None)
            if not fallback_name:
                return {}
            sheet = workbook[fallback_name]

        header_row, header_positions = find_screening_header_row_and_columns(sheet)
        code_columns = header_positions.get("\uc885\ubaa9\ucf54\ub4dc", [])
        if not code_columns:
            return {}
        code_column = code_columns[0]

        s_column = SCREENING_SCORE_COLUMN_INDEX + 1  # pandas 0-based index -> excel 1-based col index
        scores: dict[str, float] = {}
        for values in sheet.iter_rows(min_row=header_row + 1, values_only=True):
            row_values = list(values or [])
            code_value = row_values[code_column - 1] if len(row_values) >= code_column else None
            stock_code = normalize_stock_code_value(code_value)
            if not stock_code:
                continue
            raw_s = row_values[s_column - 1] if len(row_values) >= s_column else None
            score = to_float(raw_s)
            if score is None or not math.isfinite(score):
                continue
            scores[stock_code] = float(score)
        return scores
    finally:
        workbook.close()


def update_screening_score_average_columns(selected_file: Path) -> dict[str, Any]:
    return {
        "qr_updated": False,
        "file_name": selected_file.name,
        "file_date": parse_screening_date(selected_file),
        "updated_rows": 0,
        "source_files": 0,
        "numeric_score_files": 0,
        "db_source_files": 0,
        "db_changed_files": 0,
        "db_total_scores": 0,
        "disabled": True,
        "mode": "sql_only",
    }


def find_screening_note_cell(workbook: Any, stock_code: str | None, stock_name: str) -> tuple[Any, int, int]:
    try:
        sheet = workbook[SCREENING_SHEET]
    except KeyError:
        raise ValueError(f"엑셀에서 '{SCREENING_SHEET}' 시트를 찾지 못했습니다.")

    header_row = 0
    header_map: dict[str, int] = {}
    for row_index in range(1, min(sheet.max_row, 30) + 1):
        values = [str(cell.value or "").strip() for cell in sheet[row_index]]
        if "\uc885\ubaa9\uba85" in values and "\ube44\uace0" in values:
            header_row = row_index
            header_map = {value: index + 1 for index, value in enumerate(values) if value}
            break
    if not header_row:
        raise ValueError("엑셀에서 종목명/비고 헤더 행을 찾지 못했습니다.")

    name_col = header_map.get("\uc885\ubaa9\uba85")
    note_col = header_map.get("\ube44\uace0")
    code_col = header_map.get("\uc885\ubaa9\ucf54\ub4dc")
    if not name_col or not note_col:
        raise ValueError("엑셀에 종목명 또는 비고 열이 없습니다.")

    target_code = normalize_stock_code_value(stock_code)
    target_name = normalize_text(stock_name)
    for row_index in range(header_row + 1, sheet.max_row + 1):
        row_code = normalize_stock_code_value(sheet.cell(row=row_index, column=code_col).value) if code_col else ""
        row_name = normalize_text(sheet.cell(row=row_index, column=name_col).value)
        if (target_code and row_code == target_code) or (target_name and row_name == target_name):
            return sheet, row_index, note_col
    raise ValueError(f"엑셀에서 비고를 수정할 종목을 찾지 못했습니다: {stock_name}")


def write_excel_workbook_with_powershell_com(
    path: Path,
    stock_code: str | None,
    stock_name: str,
    note: str,
) -> dict[str, Any]:
    payload = {
        "path": str(path.resolve()),
        "sheet": SCREENING_SHEET,
        "stock_code": stock_code or "",
        "stock_name": stock_name or "",
        "note": note or "",
    }
    script = r'''
param([string]$PayloadPath)
$ErrorActionPreference = "Stop"
$payload = Get-Content -LiteralPath $PayloadPath -Raw -Encoding UTF8 | ConvertFrom-Json

function Normalize-Code($value) {
  if ($null -eq $value) { return "" }
  $digits = ([string]$value) -replace "\D", ""
  if (!$digits) { return "" }
  return $digits.PadLeft(6, "0")
}

function Normalize-Text($value) {
  if ($null -eq $value) { return "" }
  return (([string]$value).Trim().ToLowerInvariant() -replace "\s+", "")
}

$target = [System.IO.Path]::GetFullPath([string]$payload.path)
$excel = $null
$workbook = $null
$openedByApp = $false
$createdExcel = $false

try {
  try {
    $excel = [Runtime.InteropServices.Marshal]::GetActiveObject("Excel.Application")
  } catch {
    $excel = $null
  }

  if ($null -ne $excel) {
    foreach ($candidate in @($excel.Workbooks)) {
      $candidatePath = [System.IO.Path]::GetFullPath([string]$candidate.FullName)
      if ([string]::Equals($candidatePath, $target, [System.StringComparison]::OrdinalIgnoreCase)) {
        $workbook = $candidate
        break
      }
    }
  }

  if ($null -eq $workbook) {
    $excel = New-Object -ComObject Excel.Application
    $createdExcel = $true
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $workbook = $excel.Workbooks.Open($target, 0, $false)
    $openedByApp = $true
  }

  $sheet = $workbook.Worksheets.Item([string]$payload.sheet)
  $used = $sheet.UsedRange
  $maxRow = [Math]::Min([int]($used.Row + $used.Rows.Count - 1), 30)
  $maxCol = [Math]::Min([Math]::Max([int]($used.Column + $used.Columns.Count - 1), 20), 300)
  $headerRow = 0
  $headerMap = @{}

  for ($r = 1; $r -le $maxRow; $r++) {
    $headerMap = @{}
    $values = @()
    for ($c = 1; $c -le $maxCol; $c++) {
      $text = ([string]($sheet.Cells.Item($r, $c).Value2)).Trim()
      $values += $text
      if ($text -and !$headerMap.ContainsKey($text)) {
        $headerMap[$text] = $c
      }
    }
    if ($values -contains "종목명" -and $values -contains "비고") {
      $headerRow = $r
      break
    }
  }

  if (!$headerRow) { throw "엑셀에서 종목명/비고 헤더 행을 찾지 못했습니다." }
  $nameCol = $headerMap["종목명"]
  $noteCol = $headerMap["비고"]
  $codeCol = $headerMap["종목코드"]
  if (!$nameCol -or !$noteCol) { throw "엑셀에 종목명 또는 비고 열이 없습니다." }

  $targetCode = Normalize-Code $payload.stock_code
  $targetName = Normalize-Text $payload.stock_name
  $lastRow = [int]($used.Row + $used.Rows.Count - 1)

  for ($r = $headerRow + 1; $r -le $lastRow; $r++) {
    $rowCode = ""
    if ($codeCol) { $rowCode = Normalize-Code $sheet.Cells.Item($r, $codeCol).Value2 }
    $rowName = Normalize-Text $sheet.Cells.Item($r, $nameCol).Value2
    if (($targetCode -and $rowCode -eq $targetCode) -or ($targetName -and $rowName -eq $targetName)) {
      $sheet.Cells.Item($r, $noteCol).Value2 = [string]$payload.note
      $workbook.Save()
      [PSCustomObject]@{
        row = $r
        column = $noteCol
        written_open_excel = -not $openedByApp
        written_with_excel = $true
        closed_excel = $false
      } | ConvertTo-Json -Compress
      exit 0
    }
  }
  throw ("엑셀에서 비고를 수정할 종목을 찾지 못했습니다: " + [string]$payload.stock_name)
} finally {
  if ($openedByApp -and $null -ne $workbook) {
    $workbook.Close($true)
  }
  if ($createdExcel -and $null -ne $excel) {
    $excel.Quit()
  }
}
'''
    with tempfile.TemporaryDirectory(prefix="stock-dashboard-excel-com-") as tmp:
        tmp_path = Path(tmp)
        payload_path = tmp_path / "payload.json"
        script_path = tmp_path / "write_note.ps1"
        payload_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        script_path.write_text(script, encoding="utf-8-sig")
        result = subprocess.run(
            [
                "powershell.exe",
                "-NoProfile",
                "-ExecutionPolicy",
                "Bypass",
                "-File",
                str(script_path),
                str(payload_path),
            ],
            cwd=BASE_DIR,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=60,
        )
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip()
        raise RuntimeError(detail or "Excel COM 저장에 실패했습니다.")
    output = (result.stdout or "").strip().splitlines()[-1]
    return json.loads(output)


def write_open_excel_workbook(
    path: Path,
    stock_code: str | None,
    stock_name: str,
    note: str,
) -> dict[str, Any] | None:
    try:
        import win32com.client  # type: ignore
    except Exception as exc:
        return write_excel_workbook_with_powershell_com(path, stock_code, stock_name, note)

    target = os.path.normcase(os.path.abspath(path))
    excel = None
    workbook = None
    opened_by_app = False
    created_excel = False

    try:
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            excel = None

        if excel is not None:
            for candidate in list(excel.Workbooks):
                full_name = os.path.normcase(os.path.abspath(str(candidate.FullName)))
                if full_name == target:
                    workbook = candidate
                    break

        if workbook is None:
            excel = win32com.client.DispatchEx("Excel.Application")
            created_excel = True
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(target, 0, False)
            opened_by_app = True

        try:
            sheet = workbook.Worksheets(SCREENING_SHEET)
        except Exception as exc:
            raise ValueError(f"엑셀에서 '{SCREENING_SHEET}' 시트를 찾지 못했습니다.") from exc

        used_range = sheet.UsedRange
        max_row = min(int(used_range.Row + used_range.Rows.Count - 1), 30)
        max_col = min(max(int(used_range.Column + used_range.Columns.Count - 1), 20), 300)
        header_row = 0
        header_map: dict[str, int] = {}
        for row_index in range(1, max_row + 1):
            values: list[str] = []
            header_map = {}
            for col_index in range(1, max_col + 1):
                value = sheet.Cells(row_index, col_index).Value
                text = str(value or "").strip()
                values.append(text)
                if text and text not in header_map:
                    header_map[text] = col_index
            if "종목명" in values and "비고" in values:
                header_row = row_index
                break
        if not header_row:
            raise ValueError("엑셀에서 종목명/비고 헤더 행을 찾지 못했습니다.")

        name_col = header_map.get("종목명")
        note_col = header_map.get("비고")
        code_col = header_map.get("종목코드")
        if not name_col or not note_col:
            raise ValueError("엑셀에 종목명 또는 비고 열이 없습니다.")

        target_code = normalize_stock_code_value(stock_code)
        target_name = normalize_text(stock_name)
        last_row = int(used_range.Row + used_range.Rows.Count - 1)
        for row_index in range(header_row + 1, last_row + 1):
            row_code = normalize_stock_code_value(sheet.Cells(row_index, code_col).Value) if code_col else ""
            row_name = normalize_text(sheet.Cells(row_index, name_col).Value)
            if (target_code and row_code == target_code) or (target_name and row_name == target_name):
                sheet.Cells(row_index, note_col).Value = note
                workbook.Save()
                return {
                    "row": row_index,
                    "column": note_col,
                    "written_open_excel": not opened_by_app,
                    "written_with_excel": True,
                    "closed_excel": False,
                }
        raise ValueError(f"엑셀에서 비고를 수정할 종목을 찾지 못했습니다: {stock_name}")
    finally:
        if opened_by_app and workbook is not None:
            workbook.Close(SaveChanges=True)
        if created_excel and excel is not None:
            excel.Quit()


def close_open_excel_workbook(path: Path) -> bool:
    try:
        import win32com.client  # type: ignore
    except Exception as exc:
        raise RuntimeError("열려 있는 Excel 파일을 닫으려면 pywin32/Excel COM 연결이 필요합니다.") from exc

    target = os.path.normcase(os.path.abspath(path))
    excel = win32com.client.GetActiveObject("Excel.Application")
    for workbook in list(excel.Workbooks):
        full_name = os.path.normcase(os.path.abspath(str(workbook.FullName)))
        if full_name == target:
            workbook.Save()
            workbook.Close(SaveChanges=True)
            return True
    return False


def write_screening_note_to_excel(
    selected_file: Path,
    stock_code: str | None,
    stock_name: str,
    note: str,
    close_open_excel: bool = False,
    write_open_excel: bool = False,
) -> dict[str, Any]:
    if selected_file.suffix.lower() in {".xlsm", ".xlsb"}:
        result = write_open_excel_workbook(selected_file, stock_code, stock_name, note)
        if result:
            return result
        raise PermissionError("Excel COM으로 파일을 열어 비고를 저장하지 못했습니다.")

    def write_once() -> dict[str, Any]:
        workbook = load_workbook(selected_file)
        try:
            sheet, row_index, note_col = find_screening_note_cell(workbook, stock_code, stock_name)
            sheet.cell(row=row_index, column=note_col).value = note
            workbook.save(selected_file)
        finally:
            workbook.close()
        return {"row": row_index, "column": note_col}

    try:
        result = write_once()
        result["closed_excel"] = False
        result["written_open_excel"] = False
        return result
    except PermissionError:
        if not (close_open_excel or write_open_excel):
            raise
        open_result = write_open_excel_workbook(selected_file, stock_code, stock_name, note)
        if open_result:
            return open_result
        if close_open_excel:
            closed = close_open_excel_workbook(selected_file)
            if not closed:
                raise PermissionError("엑셀 파일이 열려 있지만 열린 Excel 워크북을 찾지 못했습니다.")
            result = write_once()
            result["closed_excel"] = True
            result["written_open_excel"] = False
            return result
        raise PermissionError("엑셀 파일이 열려 있지만 열린 Excel 워크북을 찾지 못했습니다.")


def update_screening_note_cache(file_date: str, stock_code: str | None, stock_name: str, note: str) -> None:
    cache = load_screening_cache()
    summaries = cache.setdefault("summaries", {})
    target_code = normalize_stock_code_value(stock_code)
    target_name = normalize_text(stock_name)
    changed = False
    for key, summary in list(summaries.items()):
        if not isinstance(summary, dict):
            continue
        summary_file_date = str(summary.get("file_date") or "")
        key_text = str(key)
        if summary_file_date != file_date and file_date not in key_text:
            continue
        rows = summary.get("qualified_stocks")
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            row_code = normalize_stock_code_value(row.get("stock_code"))
            row_names = {normalize_text(row.get("stock_name")), normalize_text(row.get("resolved_name"))}
            if (target_code and row_code == target_code) or (target_name and target_name in row_names):
                row["note"] = note
                changed = True
    if changed:
        cache["recent_leaders"] = {}
        save_screening_cache(cache)


def update_screening_note_sql(file_date: str, stock_code: str | None, stock_name: str, note: str) -> None:
    if not SCREENING_FAST_DB_PATH.exists():
        return
    target_code = normalize_stock_code_value(stock_code)
    target_name = str(stock_name or "").strip()
    with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
        updated = 0
        if target_code:
            updated = conn.execute(
                """
                UPDATE screening_rows
                SET note = ?
                WHERE file_date_key = ?
                  AND stock_code = ?
                """,
                (note, re.sub(r"\D", "", file_date), target_code),
            ).rowcount or 0
        if updated <= 0 and target_name:
            updated = conn.execute(
                """
                UPDATE screening_rows
                SET note = ?
                WHERE file_date_key = ?
                  AND stock_name = ?
                """,
                (note, re.sub(r"\D", "", file_date), target_name),
            ).rowcount or 0
        conn.commit()


def resolve_screening_file_date(file_date: str | None = None) -> tuple[str, str]:
    available_entries = screening_available_file_entries(limit=None)
    if not available_entries:
        raise FileNotFoundError("SQL 캐시에 주도주 데이터가 없습니다.")
    available_map = {str(item.get("file_date") or ""): str(item.get("file_name") or "") for item in available_entries}
    requested_date = ""
    if file_date:
        digits = re.sub(r"\D", "", str(file_date))
        if len(digits) == 8:
            requested_date = datetime.strptime(digits, "%Y%m%d").strftime("%Y-%m-%d")
        else:
            requested_date = str(file_date).strip()
    selected_date = requested_date if requested_date in available_map else str(available_entries[0].get("file_date") or "")
    if not selected_date:
        raise FileNotFoundError("SQL 캐시에 주도주 데이터가 없습니다.")
    return selected_date, str(available_map.get(selected_date) or f"{selected_date.replace('-', '')}_데일리_기업스크리닝.xlsx")


def update_screening_note(request: ThemeNoteUpdateRequest) -> dict[str, Any]:
    file_date, file_name = resolve_screening_file_date(request.file_date)
    note = str(request.note or "").strip()
    stock_name = str(request.stock_name or "").strip()
    if not stock_name and not request.stock_code:
        raise ValueError("비고를 수정할 종목 정보가 없습니다.")
    update_screening_note_sql(file_date, request.stock_code, stock_name, note)
    update_screening_note_cache(file_date, request.stock_code, stock_name, note)
    return {
        "ok": True,
        "locked": False,
        "file_name": file_name,
        "file_date": file_date,
        "stock_code": normalize_stock_code_value(request.stock_code),
        "stock_name": stock_name,
        "note": note,
        "closed_excel": False,
        "written_open_excel": False,
        "written_with_excel": False,
        "row": None,
    }


def build_recent_leader_stats(files: list[Path], min_score: float) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for file in files:
        file_date = parse_screening_date(file)
        try:
            df = load_screening_frame(file)
        except Exception:
            continue
        qualified = df[df["\uc810\uc218"] >= min_score].copy()
        for _, row in qualified.iterrows():
            stock_name = str(row.get("\uc885\ubaa9\uba85")).strip()
            canonical_theme, _ = canonicalize_theme(row.get("\ube44\uace0"), stock_name)
            rows.append(
                {
                    "date": file_date,
                    "stock_name": stock_name,
                    "score": float(row.get("\uc810\uc218")),
                    "change_pct": float(row.get("\ub4f1\ub77d\ub960", 0.0) or 0.0) * 100,
                    "note": "" if pd.isna(row.get("\ube44\uace0")) else str(row.get("\ube44\uace0")).strip(),
                    "theme": canonical_theme,
                }
            )

    if not rows:
        return []

    recent_df = pd.DataFrame(rows)
    grouped = (
        recent_df.groupby("stock_name", as_index=False)
        .agg(
            appearances=("date", "count"),
            strong_days=("score", lambda s: int((s >= 80).sum())),
            avg_score=("score", "mean"),
            max_score=("score", "max"),
            avg_change_pct=("change_pct", "mean"),
            latest_date=("date", "max"),
            theme_signature=("theme", lambda s: ", ".join(pd.Series(s).value_counts().head(3).index.tolist())),
            note_signature=("note", lambda s: ", ".join([item for item in pd.Series(s).value_counts().index.tolist() if item][:2])),
        )
        .sort_values(["appearances", "strong_days", "avg_score", "max_score"], ascending=[False, False, False, False])
        .head(20)
    )
    grouped["avg_score"] = grouped["avg_score"].round(2)
    grouped["max_score"] = grouped["max_score"].round(2)
    grouped["avg_change_pct"] = grouped["avg_change_pct"].round(2)
    grouped["stock_code"] = grouped["stock_name"].map(lambda name: resolve_stock(str(name))[0] or "")
    return grouped.to_dict(orient="records")


def build_recent_leader_stats_from_summaries(
    summaries: list[dict[str, Any]],
    min_score: float,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for summary in summaries:
        file_date = str(summary.get("file_date") or "")
        for item in summary.get("qualified_stocks", []):
            score = to_float(item.get("score"))
            if score is None or score < min_score:
                continue
            stock_name = str(item.get("stock_name") or "").strip()
            stock_code = normalize_stock_code_value(item.get("stock_code"))
            canonical_theme, _ = canonicalize_theme(item.get("note"), stock_name)
            rows.append(
                {
                    "date": file_date,
                    "stock_code": stock_code or "",
                    "stock_name": stock_name,
                    "score": float(score),
                    "change_pct": float(to_float(item.get("change_pct")) or 0.0),
                    "note": str(item.get("note") or "").strip(),
                    "theme": canonical_theme,
                }
            )

    if not rows:
        return []

    recent_df = pd.DataFrame(rows)
    grouped = (
        recent_df.groupby("stock_name", as_index=False)
        .agg(
            stock_code=("stock_code", lambda s: next((str(item).strip() for item in s if str(item).strip()), "")),
            appearances=("date", "count"),
            strong_days=("score", lambda s: int((s >= 80).sum())),
            avg_score=("score", "mean"),
            max_score=("score", "max"),
            avg_change_pct=("change_pct", "mean"),
            latest_date=("date", "max"),
            theme_signature=("theme", lambda s: ", ".join(pd.Series(s).value_counts().head(3).index.tolist())),
            note_signature=("note", lambda s: ", ".join([item for item in pd.Series(s).value_counts().index.tolist() if item][:2])),
        )
        .sort_values(["appearances", "strong_days", "avg_score", "max_score"], ascending=[False, False, False, False])
        .head(20)
    )
    grouped["avg_score"] = grouped["avg_score"].round(2)
    grouped["max_score"] = grouped["max_score"].round(2)
    grouped["avg_change_pct"] = grouped["avg_change_pct"].round(2)
    return grouped.to_dict(orient="records")


def build_screening_summary_from_sql(file_date: str, min_score: float) -> dict[str, Any]:
    summaries = load_screening_summaries_for_dates([file_date])
    selected_summary = summaries[0] if summaries else None
    if not selected_summary:
        raise FileNotFoundError(f"SQL 캐시에 해당 날짜 데이터가 없습니다: {file_date}")

    available_file_map = {item["file_date"]: item["file_name"] for item in screening_available_file_entries(limit=None)}
    qualified_rows: list[dict[str, Any]] = []
    for row in selected_summary.get("qualified_stocks", []):
        market_cap_100m = to_float(row.get("market_cap_100m"))
        if market_cap_100m is None or market_cap_100m < SCREENING_MIN_MARKET_CAP_100M:
            continue
        stock_name = str(row.get("stock_name") or "").strip()
        stock_code = normalize_stock_code_value(row.get("stock_code"))
        resolved_code = stock_code or ""
        resolved_name = stock_name
        if not resolved_code:
            resolved_code, resolved_name = resolve_stock(stock_name)
        canonical_theme, matched_keywords = canonicalize_theme(row.get("note"), stock_name)
        qualified_rows.append(
            {
                "theme": canonical_theme,
                "theme_keywords": matched_keywords,
                "stock_name": stock_name,
                "stock_code": stock_code or resolved_code or "",
                "resolved_name": resolved_name or stock_name,
                "score": round(float(to_float(row.get("score")) or 0.0), 2),
                "score_o": round(float(to_float(row.get("score_o")) or 0.0), 2),
                "change_pct": round(float(to_float(row.get("change_pct")) or 0.0), 2),
                "is_52w_high": int(to_float(row.get("is_52w_high")) or 0),
                "market_cap_100m": round(float(market_cap_100m), 1),
                "trading_value_100m": round(float(to_float(row.get("trading_value_100m")) or 0.0), 1),
                "market": "" if pd.isna(row.get("market")) else str(row.get("market") or ""),
                "industry": "" if pd.isna(row.get("industry")) else str(row.get("industry") or "").strip(),
                "execution_strength": None,
                "avg_1m": round(float(to_float(row.get("avg_1m")) or 0.0), 2) if to_float(row.get("avg_1m")) is not None else None,
                "avg_1w": round(float(to_float(row.get("avg_1w")) or 0.0), 2) if to_float(row.get("avg_1w")) is not None else None,
                "avg_3m": round(float(to_float(row.get("avg_3m")) or 0.0), 2) if to_float(row.get("avg_3m")) is not None else None,
                "sortino_norm": round(float(to_float(row.get("sortino_norm")) or 0.0), 4),
                "note": str(row.get("note") or "").strip(),
                "lead_count": int(row.get("lead_count") or 0),
                "avg_lead_score": round(float(row.get("avg_lead_score") or 0.0), 2),
                "rank": len(qualified_rows) + 1,
            }
        )

    theme_df = pd.DataFrame(qualified_rows)
    grouped_records: list[dict[str, Any]] = []
    if not theme_df.empty:
        grouped = (
            theme_df.groupby("theme", as_index=False)
            .agg(
                count=("stock_name", "count"),
                avg_score=("score", "mean"),
                max_score=("score", "max"),
                leaders=("stock_name", lambda items: ", ".join(list(items)[:5])),
                keywords=("theme_keywords", lambda items: ", ".join(sorted({token for sub in items for token in sub if token})[:5])),
            )
            .sort_values(["count", "avg_score", "max_score"], ascending=[False, False, False])
            .head(15)
        )
        grouped["avg_score"] = grouped["avg_score"].round(2)
        grouped["max_score"] = grouped["max_score"].round(2)
        grouped_records = grouped.to_dict(orient="records")

    return {
        "file_name": available_file_map.get(file_date, f"{str(file_date).replace('-', '')}_데일리_기업스크리닝.xlsx"),
        "file_date": file_date,
        "min_score": min_score,
        "score_basis": "sql_score_s",
        "score_column": "screening_rows.score_s",
        "qualified_count": len(qualified_rows),
        "qualified_stocks": qualified_rows,
        "theme_summary": grouped_records,
    }


def build_us_screening_summary_from_sql(file_date: str, min_score: float) -> dict[str, Any]:
    summaries = load_us_screening_summaries_for_dates([file_date])
    selected_summary = summaries[0] if summaries else None
    if not selected_summary:
        raise FileNotFoundError(f"US SQL 캐시에 해당 날짜 데이터가 없습니다: {file_date}")

    available_file_map = {item["file_date"]: item["file_name"] for item in us_screening_available_file_entries(limit=None)}
    qualified_rows: list[dict[str, Any]] = []
    for row in selected_summary.get("qualified_stocks", []):
        market_cap_100m = to_float(row.get("market_cap_100m"))
        if market_cap_100m is None or market_cap_100m <= 0:
            continue
        stock_name = str(row.get("stock_name") or "").strip()
        stock_code = str(row.get("stock_code") or "").strip().upper()
        theme_name = str(row.get("manual_sector") or row.get("theme") or row.get("industry") or "Other").strip() or "Other"
        industry_name = str(row.get("industry") or "").strip()
        qualified_rows.append(
            {
                "theme": theme_name,
                "theme_keywords": [theme_name] if theme_name else [],
                "stock_name": stock_name,
                "stock_code": stock_code,
                "resolved_name": stock_name,
                "score": round(float(to_float(row.get("score")) or 0.0), 2),
                "score_o": round(float(to_float(row.get("score_o")) or 0.0), 2),
                "change_pct": round(float(to_float(row.get("change_pct")) or 0.0), 2),
                "is_52w_high": int(to_float(row.get("is_52w_high")) or 0),
                "market_cap_100m": round(float(market_cap_100m), 2),
                "trading_value_100m": round(float(to_float(row.get("trading_value_100m")) or 0.0), 2),
                "market_cap_usd": round(float(market_cap_100m) * 100000000.0, 2),
                "trading_value_usd": round(float(to_float(row.get("trading_value_100m")) or 0.0) * 100000000.0, 2),
                "display_currency": "USD",
                "local_currency": "USD",
                "market": "US",
                "industry": industry_name,
                "execution_strength": None,
                "avg_1m": round(float(to_float(row.get("avg_1m")) or 0.0), 2) if to_float(row.get("avg_1m")) is not None else None,
                "avg_3m": round(float(to_float(row.get("avg_3m")) or 0.0), 2) if to_float(row.get("avg_3m")) is not None else None,
                "avg_1w": round(float(to_float(row.get("avg_1w")) or 0.0), 2) if to_float(row.get("avg_1w")) is not None else None,
                "sortino_norm": round(float(to_float(row.get("sortino_norm")) or 0.0), 4),
                "note": str(row.get("note") or "").strip(),
                "lead_count": int(row.get("lead_count") or 0),
                "avg_lead_score": round(float(row.get("avg_lead_score") or 0.0), 2),
                "rank": len(qualified_rows) + 1,
                "manual_sector": theme_name,
            }
        )
    theme_df = pd.DataFrame(qualified_rows)
    grouped_records: list[dict[str, Any]] = []
    if not theme_df.empty:
        grouped = (
            theme_df.groupby("theme", as_index=False)
            .agg(
                count=("stock_name", "count"),
                avg_score=("score", "mean"),
                max_score=("score", "max"),
                leaders=("stock_name", lambda items: ", ".join(list(items)[:5])),
                keywords=("theme_keywords", lambda items: ", ".join(sorted({token for sub in items for token in sub if token})[:5])),
            )
            .sort_values(["count", "avg_score", "max_score"], ascending=[False, False, False])
            .head(15)
        )
        grouped["avg_score"] = grouped["avg_score"].round(2)
        grouped["max_score"] = grouped["max_score"].round(2)
        grouped_records = grouped.to_dict(orient="records")
    return {
        "file_name": available_file_map.get(file_date, f"{str(file_date).replace('-', '')}_us_daily_screening.xlsx"),
        "file_date": file_date,
        "min_score": min_score,
        "score_basis": "sql_score_s",
        "score_column": "screening_rows.score_s",
        "qualified_count": len(qualified_rows),
        "qualified_stocks": qualified_rows,
        "theme_summary": grouped_records,
    }


ASIA_THEME_REGION_OPTIONS = [
    {"code": "jp", "label": "일본"},
    {"code": "cn", "label": "중국"},
    {"code": "tw", "label": "대만"},
]


def normalize_asia_theme_region(region: str | None) -> str:
    value = str(region or "").strip().lower()
    if value in {"jp", "japan", "tse"}:
        return "jp"
    if value in {"cn", "china", "sse", "szse"}:
        return "cn"
    if value in {"tw", "taiwan", "twse"}:
        return "tw"
    return "jp"


def infer_asia_theme_region(stock_code: Any, industry: Any = "", manual_sector: Any = "") -> str:
    code = str(stock_code or "").strip().upper()
    industry_text = str(industry or "").strip().upper()
    sector_text = str(manual_sector or "").strip().upper()
    combined = " ".join(part for part in [code, industry_text, sector_text] if part)
    if combined.startswith("TWSE:") or "TWSE" in combined:
        return "tw"
    if combined.startswith("TSE:") or "TSE" in combined:
        return "jp"
    if combined.startswith("SSE:") or combined.startswith("SZSE:") or "SSE" in combined or "SZSE" in combined:
        return "cn"
    return "jp"


def asia_theme_region_currency(region: str) -> str:
    normalized = normalize_asia_theme_region(region)
    if normalized == "tw":
        return "TWD"
    if normalized == "cn":
        return "CNY"
    return "JPY"


def screening_value_100m_to_usd(value_100m: Any, currency: str) -> float | None:
    amount_100m = to_float(value_100m)
    if amount_100m is None:
        return None
    usd_value = convert_currency_value(float(amount_100m) * 100000000.0, currency, "USD")
    if usd_value is None:
        return None
    return round(float(usd_value), 2)


def build_asia_screening_summary_from_sql(file_date: str, min_score: float, region: str = "jp") -> dict[str, Any]:
    normalized_region = normalize_asia_theme_region(region)
    region_currency = asia_theme_region_currency(normalized_region)
    summaries = load_asia_screening_summaries_for_dates([file_date])
    selected_summary = summaries[0] if summaries else None
    if not selected_summary:
        raise FileNotFoundError(f"ASIA SQL 캐시에 해당 날짜 데이터가 없습니다: {file_date}")

    available_file_map = {item["file_date"]: item["file_name"] for item in asia_screening_available_file_entries(limit=None)}
    qualified_rows: list[dict[str, Any]] = []
    for row in selected_summary.get("qualified_stocks", []):
        market_cap_100m = to_float(row.get("market_cap_100m"))
        if market_cap_100m is None or market_cap_100m <= 0:
            continue
        if infer_asia_theme_region(row.get("stock_code"), row.get("industry"), row.get("manual_sector") or row.get("theme")) != normalized_region:
            continue
        stock_name = str(row.get("stock_name") or "").strip()
        stock_code = str(row.get("stock_code") or "").strip().upper()
        theme_name = str(row.get("manual_sector") or row.get("theme") or "Other").strip() or "Other"
        market_name = str(row.get("industry") or "").strip()
        market_cap_usd = screening_value_100m_to_usd(market_cap_100m, region_currency)
        trading_value_usd = screening_value_100m_to_usd(row.get("trading_value_100m"), region_currency)
        qualified_rows.append(
            {
                "theme": theme_name,
                "theme_keywords": [theme_name] if theme_name else [],
                "stock_name": stock_name,
                "stock_code": stock_code,
                "resolved_name": stock_name,
                "score": round(float(to_float(row.get("score")) or 0.0), 2),
                "score_o": round(float(to_float(row.get("score_o")) or 0.0), 2),
                "change_pct": round(float(to_float(row.get("change_pct")) or 0.0), 2),
                "is_52w_high": int(to_float(row.get("is_52w_high")) or 0),
                "market_cap_100m": round(float(market_cap_100m), 2),
                "trading_value_100m": round(float(to_float(row.get("trading_value_100m")) or 0.0), 2),
                "market_cap_usd": market_cap_usd,
                "trading_value_usd": trading_value_usd,
                "display_currency": "USD",
                "local_currency": region_currency,
                "market": "ASIA",
                "industry": market_name,
                "execution_strength": None,
                "avg_1m": round(float(to_float(row.get("avg_1m")) or 0.0), 2) if to_float(row.get("avg_1m")) is not None else None,
                "avg_3m": round(float(to_float(row.get("avg_3m")) or 0.0), 2) if to_float(row.get("avg_3m")) is not None else None,
                "avg_1w": round(float(to_float(row.get("avg_1w")) or 0.0), 2) if to_float(row.get("avg_1w")) is not None else None,
                "sortino_norm": round(float(to_float(row.get("sortino_norm")) or 0.0), 4),
                "note": str(row.get("note") or "").strip(),
                "lead_count": int(row.get("lead_count") or 0),
                "avg_lead_score": round(float(row.get("avg_lead_score") or 0.0), 2),
                "rank": len(qualified_rows) + 1,
                "manual_sector": theme_name,
            }
        )
    theme_df = pd.DataFrame(qualified_rows)
    grouped_records: list[dict[str, Any]] = []
    if not theme_df.empty:
        grouped = (
            theme_df.groupby("theme", as_index=False)
            .agg(
                count=("stock_name", "count"),
                avg_score=("score", "mean"),
                max_score=("score", "max"),
                leaders=("stock_name", lambda items: ", ".join(list(items)[:5])),
                keywords=("theme_keywords", lambda items: ", ".join(sorted({token for sub in items for token in sub if token})[:5])),
            )
            .sort_values(["count", "avg_score", "max_score"], ascending=[False, False, False])
            .head(15)
        )
        grouped["avg_score"] = grouped["avg_score"].round(2)
        grouped["max_score"] = grouped["max_score"].round(2)
        grouped_records = grouped.to_dict(orient="records")
    return {
        "file_name": available_file_map.get(file_date, f"{str(file_date).replace('-', '')}_asia_daily_screening.xlsx"),
        "file_date": file_date,
        "min_score": min_score,
        "selected_region": normalized_region,
        "region_options": ASIA_THEME_REGION_OPTIONS,
        "score_basis": "sql_score_s",
        "score_column": "screening_rows.score_s",
        "qualified_count": len(qualified_rows),
        "qualified_stocks": qualified_rows,
        "theme_summary": grouped_records,
    }


def build_screening_summary_from_excel(selected_file: Path, min_score: float) -> dict[str, Any]:
    df = load_screening_frame(selected_file)
    if "_market_cap_100m" not in df.columns:
        if "시가총액" in df.columns:
            df["_market_cap_100m"] = pd.to_numeric(df["시가총액"], errors="coerce")
            if df["_market_cap_100m"].isna().all():
                df["_market_cap_100m"] = df["시가총액"].map(parse_korean_number)
        else:
            df["_market_cap_100m"] = np.nan
    if "점수" not in df.columns:
        if "종합 점수" in df.columns:
            df["점수"] = pd.to_numeric(df["종합 점수"], errors="coerce")
        elif "O열점수" in df.columns:
            df["점수"] = pd.to_numeric(df["O열점수"], errors="coerce")
        else:
            df["점수"] = np.nan
    # 점수 컷은 적용하지 않고, 시총 필터(2000억 이상)만 유지한다.
    df = df.sort_values("\uc810\uc218", ascending=False)
    df = df[df["_market_cap_100m"] >= SCREENING_MIN_MARKET_CAP_100M]

    qualified_rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        stock_name = str(row.get("\uc885\ubaa9\uba85")).strip()
        raw_stock_code = normalize_stock_code_value(row.get("\uc885\ubaa9\ucf54\ub4dc"))
        stock_code, resolved_name = resolve_stock(stock_name)
        stock_code = raw_stock_code or stock_code
        canonical_theme, matched_keywords = canonicalize_theme(row.get("\ube44\uace0"), stock_name)
        trading_value = to_float(row.get("\uac70\ub798\ub300\uae08.1"))
        if trading_value is None:
            trading_value = parse_korean_number(row.get("\uac70\ub798\ub300\uae08")) or 0.0
        execution_strength = to_float(row.get("\uccb4\uacb0\uac15\ub3c4"))
        avg_1m = to_float(row.get("1M \ud3c9\uade0"))
        avg_1w = to_float(row.get("1W \ud3c9\uade0"))
        raw_change_pct = to_float(row.get("\ub4f1\ub77d\ub960"))
        if raw_change_pct is None:
            normalized_change_pct = 0.0
        else:
            # 신형 Stock_Daily(종합 점수 컬럼 존재)는 등락률이 이미 퍼센트 단위다.
            # 구형 스키마에서만 소수 비율 표기를 퍼센트로 보정한다.
            if "종합 점수" in df.columns:
                normalized_change_pct = raw_change_pct
            elif abs(raw_change_pct) <= 1.0:
                normalized_change_pct = raw_change_pct * 100.0
            else:
                normalized_change_pct = raw_change_pct

        qualified_rows.append(
            {
                "theme": canonical_theme,
                "theme_keywords": matched_keywords,
                "stock_name": stock_name,
                "stock_code": stock_code or "",
                "resolved_name": resolved_name,
                "score": round(float(row.get("\uc810\uc218")), 2),
                "score_o": round(float(to_float(row.get("O열점수")) or 0.0), 2),
                "change_pct": round(float(normalized_change_pct), 2),
                "is_52w_high": str(row.get("52\uc2e0\uace0") or "").strip(),
                "market_cap_100m": round(float(row.get("_market_cap_100m")), 1),
                "trading_value_100m": round(float(trading_value or 0.0), 1),
                "market": "" if pd.isna(row.get("\ub300\uc0c1")) else str(row.get("\ub300\uc0c1")),
                "industry": "" if pd.isna(row.get("\uc5c5\uc885\uad6c\ubd84")) else str(row.get("\uc5c5\uc885\uad6c\ubd84")).strip(),
                "execution_strength": round(execution_strength * 100, 2) if execution_strength is not None else None,
                "avg_1m": round(avg_1m, 2) if avg_1m is not None else None,
                "avg_1w": round(avg_1w, 2) if avg_1w is not None else None,
                "sortino_norm": round(float(to_float(row.get("소르티노_정규화")) or 0.0), 4),
                "note": "" if pd.isna(row.get("\ube44\uace0")) else str(row.get("\ube44\uace0")).strip(),
                "lead_count": int(row.get("\uc8fc\ub3c4\ud69f\uc218", 0) or 0),
                "avg_lead_score": round(float(row.get("\ud3c9\uade0\uc810\uc218", 0) or 0), 2),
                "rank": len(qualified_rows) + 1,
            }
        )

    theme_df = pd.DataFrame(qualified_rows)
    grouped_records: list[dict[str, Any]] = []
    if not theme_df.empty:
        grouped = (
            theme_df.groupby("theme", as_index=False)
            .agg(
                count=("stock_name", "count"),
                avg_score=("score", "mean"),
                max_score=("score", "max"),
                leaders=("stock_name", lambda items: ", ".join(list(items)[:5])),
                keywords=("theme_keywords", lambda items: ", ".join(sorted({token for sub in items for token in sub if token})[:5])),
            )
            .sort_values(["count", "avg_score", "max_score"], ascending=[False, False, False])
            .head(15)
        )
        grouped["avg_score"] = grouped["avg_score"].round(2)
        grouped["max_score"] = grouped["max_score"].round(2)
        grouped_records = grouped.to_dict(orient="records")

    return {
        "file_name": selected_file.name,
        "file_date": parse_screening_date(selected_file),
        "min_score": min_score,
        "score_basis": "excel_column_s",
        "score_column": SCREENING_SCORE_COLUMN_NAME,
        "qualified_count": len(qualified_rows),
        "qualified_stocks": qualified_rows,
        "theme_summary": grouped_records,
    }


def attach_screening_runtime_fields(summary: dict[str, Any]) -> dict[str, Any]:
    payload = json.loads(json.dumps(summary, ensure_ascii=False))
    sector_db = load_sector_db()
    file_date = str(payload.get("file_date") or "")
    signal_state_map = build_stock_entry_state_map(file_date, threshold=65.0, exit_threshold=50.0, lookback_days=365) if file_date else {}
    prev_scores_by_key: dict[str, float] = {}
    has_prev_day_scores = False
    if file_date:
        available_entries = screening_available_file_entries(limit=None)
        available_dates = [item["file_date"] for item in reversed(available_entries)]
        prev_summary: dict[str, Any] | None = None
        for index_no, date_text in enumerate(available_dates):
            if date_text == file_date and index_no > 0:
                previous_date = available_dates[index_no - 1]
                summaries = load_screening_summaries_for_dates([previous_date])
                prev_summary = summaries[0] if summaries else None
                break
        if isinstance(prev_summary, dict):
            has_prev_day_scores = True
            for prev_row in prev_summary.get("qualified_stocks", []):
                if not isinstance(prev_row, dict):
                    continue
                prev_key = sector_rotation_stock_key(prev_row)
                prev_score = to_float(prev_row.get("score"))
                if prev_key and prev_score is not None:
                    prev_scores_by_key[prev_key] = float(prev_score)
    for row in payload.get("qualified_stocks", []):
        row["manual_sector"] = resolve_sector_for_stock(row.get("stock_code"), row.get("stock_name"), sector_db)
        stock_key = sector_rotation_stock_key(row)
        current_score = to_float(row.get("score"))
        prev_score = prev_scores_by_key.get(stock_key)
        entry_signal_today = (
            has_prev_day_scores
            and current_score is not None
            and float(current_score) > 65.0
            and (prev_score is None or float(prev_score) <= 65.0)
        )
        signal_state = signal_state_map.get(stock_key, {}) if stock_key else {}
        exit_signal_today = (
            bool(signal_state)
            and not bool(signal_state.get("active"))
            and str(signal_state.get("end_date") or "") == file_date
            and current_score is not None
            and float(current_score) <= 50.0
        )
        signal_type = "entry" if entry_signal_today else "exit" if exit_signal_today else ""
        row["entry_signal_type"] = signal_type
        row["entry_signal_active"] = bool(signal_type)
        row["entry_signal_start_date"] = file_date if entry_signal_today else str(signal_state.get("start_date") or "")
        row["entry_signal_end_date"] = file_date if exit_signal_today else str(signal_state.get("end_date") or "")
        row["entry_signal_label"] = "진입 시그널" if entry_signal_today else "편출 시그널" if exit_signal_today else ""
    return payload


@lru_cache(maxsize=128)
def build_stock_entry_state_map(
    end_date: str,
    threshold: float = 70.0,
    lookback_days: int = 140,
    exit_threshold: float | None = None,
) -> dict[str, dict[str, Any]]:
    if not end_date:
        return {}
    try:
        end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
    except Exception:
        return {}
    lookback_span = max(10, int(lookback_days or 140))
    start_dt = end_dt - timedelta(days=lookback_span + 10)
    summaries = screening_backtest_source_summaries(start_date=start_dt, end_date=end_dt)
    filtered = [item for item in summaries if str(item.get("file_date") or "") <= end_date]
    if not filtered:
        return {}
    filtered = filtered[-max(10, min(lookback_span, len(filtered))):]
    state_by_stock: dict[str, dict[str, Any]] = {}
    for summary in filtered:
        date_text = str(summary.get("file_date") or "")
        day_scores: dict[str, float] = {}
        for row in summary.get("qualified_stocks", []):
            score = to_float(row.get("score"))
            if score is None:
                continue
            stock_key = sector_rotation_stock_key(row)
            if stock_key:
                day_scores[stock_key] = float(score)
        all_keys = set(state_by_stock.keys()) | set(day_scores.keys())
        for stock_key in all_keys:
            prev = state_by_stock.get(stock_key, {"active": False, "start_date": "", "end_date": ""})
            score = day_scores.get(stock_key)
            if prev.get("active"):
                active_now = score is not None and score > (float(exit_threshold) if exit_threshold is not None else threshold)
            else:
                active_now = score is not None and score >= threshold
            if active_now and not prev.get("active"):
                prev["active"] = True
                prev["start_date"] = date_text
                prev["end_date"] = ""
            elif (not active_now) and prev.get("active"):
                prev["active"] = False
                prev["end_date"] = date_text
            state_by_stock[stock_key] = prev
    for stock_key, state in state_by_stock.items():
        if state.get("active") and state.get("start_date"):
            try:
                start_dt = datetime.strptime(str(state["start_date"]), "%Y-%m-%d").date()
                end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
                days = (end_dt - start_dt).days + 1
                state["label"] = f"진입 {max(days, 1)}일차"
            except Exception:
                state["label"] = "진입중"
        elif state.get("end_date"):
            state["label"] = f"이탈({state.get('end_date')})"
        else:
            state["label"] = ""
    return state_by_stock


def screening_summary_has_table_fields(summary: dict[str, Any]) -> bool:
    rows = summary.get("qualified_stocks", [])
    if not rows:
        return True
    required_fields = {"is_52w_high", "execution_strength", "avg_1m", "avg_1w", "sortino_norm", "score_o"}
    return required_fields.issubset(set(rows[0].keys()))


def load_screening_summary(
    min_score: float = 50.0,
    recent_limit: int = RECENT_SCREENING_LOOKBACK,
    file_date: str | None = None,
    force_reload: bool = False,
) -> dict[str, Any]:
    requested_date = ""
    if file_date:
        digits = re.sub(r"\D", "", str(file_date))
        requested_date = datetime.strptime(digits, "%Y%m%d").strftime("%Y-%m-%d") if len(digits) == 8 else str(file_date).strip()
    available_entries = screening_available_file_entries(limit=None)
    available_dates = [item["file_date"] for item in available_entries]
    if not available_dates:
        raise FileNotFoundError("SQL 캐시에 주도주 데이터가 없습니다.")
    selected_date = requested_date if requested_date in available_dates else available_dates[0]
    fallback_used = bool(requested_date and requested_date != selected_date)
    cache = load_screening_cache()
    summaries_cache = cache.setdefault("summaries", {})
    payload_cache_key = screening_summary_payload_cache_key(selected_date, min_score, recent_limit)
    if not force_reload:
        cached_payload = summaries_cache.get(payload_cache_key)
        if isinstance(cached_payload, dict):
            payload = json.loads(json.dumps(cached_payload, ensure_ascii=False))
            payload["requested_file_date"] = requested_date
            payload["fallback_file_date"] = selected_date if fallback_used else ""
            payload["fallback_reason"] = "requested_file_missing" if fallback_used else ""
            payload["cache_source"] = "sql_cache"
            payload["cache_loaded_at"] = datetime.now().isoformat(timespec="seconds")
            return payload

    payload = attach_screening_runtime_fields(build_screening_summary_from_sql(selected_date, min_score))
    recent_dates = [item["file_date"] for item in reversed(available_entries[: max(1, int(recent_limit))])]
    recent_summaries = load_screening_summaries_for_dates(recent_dates)
    payload["recent_leaders"] = build_recent_leader_stats_from_summaries(recent_summaries, min_score=min_score)
    payload["available_files"] = [
        {
            "file_name": item["file_name"],
            "file_date": item["file_date"],
        }
        for item in available_entries
    ]
    payload["cache_loaded_at"] = datetime.now().isoformat(timespec="seconds")
    payload["cache_source"] = "sql"
    payload["requested_file_date"] = requested_date
    payload["fallback_file_date"] = selected_date if fallback_used else ""
    payload["fallback_reason"] = "requested_file_missing" if fallback_used else ""
    summaries_cache[payload_cache_key] = json.loads(json.dumps(payload, ensure_ascii=False))
    save_screening_cache(cache)
    return payload


def attach_us_screening_runtime_fields(summary: dict[str, Any]) -> dict[str, Any]:
    payload = json.loads(json.dumps(summary, ensure_ascii=False))
    file_date = str(payload.get("file_date") or "")
    prev_scores_by_key: dict[str, float] = {}
    has_prev_day_scores = False
    if file_date:
        available_entries = us_screening_available_file_entries(limit=None)
        available_dates = [item["file_date"] for item in reversed(available_entries)]
        prev_summary: dict[str, Any] | None = None
        for index_no, date_text in enumerate(available_dates):
            if date_text == file_date and index_no > 0:
                previous_date = available_dates[index_no - 1]
                summaries = load_us_screening_summaries_for_dates([previous_date])
                prev_summary = summaries[0] if summaries else None
                break
        if isinstance(prev_summary, dict):
            has_prev_day_scores = True
            for prev_row in prev_summary.get("qualified_stocks", []):
                if not isinstance(prev_row, dict):
                    continue
                prev_key = sector_rotation_stock_key(prev_row)
                prev_score = to_float(prev_row.get("score"))
                if prev_key and prev_score is not None:
                    prev_scores_by_key[prev_key] = float(prev_score)
    for row in payload.get("qualified_stocks", []):
        row["manual_sector"] = str(row.get("manual_sector") or row.get("theme") or "Other").strip() or "Other"
        stock_key = sector_rotation_stock_key(row)
        current_score = to_float(row.get("score"))
        prev_score = prev_scores_by_key.get(stock_key)
        first_breakout_80 = (
            has_prev_day_scores
            and current_score is not None
            and float(current_score) >= 80.0
            and (prev_score is None or float(prev_score) < 80.0)
        )
        row["entry_signal_active"] = bool(first_breakout_80)
        row["entry_signal_start_date"] = file_date if first_breakout_80 else ""
        row["entry_signal_end_date"] = file_date if first_breakout_80 else ""
        row["entry_signal_label"] = "80 신규 돌파" if first_breakout_80 else ""
    return payload


def load_us_screening_summary(
    min_score: float = 50.0,
    recent_limit: int = RECENT_SCREENING_LOOKBACK,
    file_date: str | None = None,
) -> dict[str, Any]:
    requested_date = ""
    if file_date:
        digits = re.sub(r"\D", "", str(file_date))
        requested_date = datetime.strptime(digits, "%Y%m%d").strftime("%Y-%m-%d") if len(digits) == 8 else str(file_date).strip()
    available_entries = us_screening_available_file_entries(limit=None)
    available_dates = [item["file_date"] for item in available_entries]
    if not available_dates:
        raise FileNotFoundError("US SQL 캐시에 주도주 데이터가 없습니다.")
    selected_date = requested_date if requested_date in available_dates else available_dates[0]
    fallback_used = bool(requested_date and requested_date != selected_date)
    payload = attach_us_screening_runtime_fields(build_us_screening_summary_from_sql(selected_date, min_score))
    recent_dates = [item["file_date"] for item in reversed(available_entries[: max(1, int(recent_limit))])]
    recent_summaries = load_us_screening_summaries_for_dates(recent_dates)
    payload["recent_leaders"] = build_recent_leader_stats_from_summaries(recent_summaries, min_score=min_score)
    payload["available_files"] = [{"file_name": item["file_name"], "file_date": item["file_date"]} for item in available_entries]
    payload["cache_loaded_at"] = datetime.now().isoformat(timespec="seconds")
    payload["cache_source"] = "sql"
    payload["requested_file_date"] = requested_date
    payload["fallback_file_date"] = selected_date if fallback_used else ""
    payload["fallback_reason"] = "requested_file_missing" if fallback_used else ""
    return payload


def build_us_theme_sector_calendar(min_score: float = 50.0, limit: int = 60) -> dict[str, Any]:
    cache = load_us_screening_cache()
    calendar_cache = cache.setdefault("calendar", {})
    calendar_key = f"v{US_SCREENING_CALENDAR_CACHE_VERSION}|{float(min_score):.4f}|{int(limit)}"
    cached_calendar = calendar_cache.get(calendar_key)
    if isinstance(cached_calendar, dict):
        return cached_calendar
    available_entries = us_screening_available_file_entries(limit=max(1, int(limit)))
    source_summaries = load_us_screening_summaries_for_dates([item["file_date"] for item in reversed(available_entries)])
    days: list[dict[str, Any]] = []
    for summary in source_summaries:
        file_date = str(summary.get("file_date") or "")
        rows = summary.get("qualified_stocks") or []
        if not file_date or not rows:
            continue
        sector_map: dict[str, dict[str, Any]] = {}
        for row in rows:
            sector = str(row.get("manual_sector") or row.get("theme") or "Other").strip() or "Other"
            bucket = sector_map.setdefault(
                sector,
                {
                    "sector": sector,
                    "count": 0,
                    "score_total": 0.0,
                    "rank_total": 0.0,
                    "top20_count": 0,
                    "top50_count": 0,
                    "trading_value_100m": 0.0,
                    "market_cap_100m": 0.0,
                    "leaders": [],
                },
            )
            rank_value = int(to_float(row.get("rank")) or (len(rows) + 1))
            bucket["count"] += 1
            bucket["score_total"] += float(row.get("score") or 0.0)
            bucket["rank_total"] += rank_value
            if rank_value <= 20:
                bucket["top20_count"] += 1
            if rank_value <= 50:
                bucket["top50_count"] += 1
            bucket["trading_value_100m"] += float(row.get("trading_value_100m") or 0.0)
            bucket["market_cap_100m"] += float(row.get("market_cap_100m") or 0.0)
            if len(bucket["leaders"]) < 4:
                bucket["leaders"].append(str(row.get("stock_name") or ""))
        sectors = []
        for item in sector_map.values():
            turnover_ratio = item["trading_value_100m"] / item["market_cap_100m"] if item["market_cap_100m"] else 0.0
            top20_ratio = item["top20_count"] / item["count"] if item["count"] else 0.0
            top50_ratio = item["top50_count"] / item["count"] if item["count"] else 0.0
            sector_strength = (
                (item["score_total"] / item["count"] if item["count"] else 0.0) * 0.6
                + top20_ratio * 100.0 * 0.2
                + min(max(turnover_ratio * 100.0, 0.0), 10.0) * 10.0 * 0.2
            )
            sectors.append(
                {
                    "sector": item["sector"],
                    "count": item["count"],
                    "avg_score": round(item["score_total"] / item["count"], 2) if item["count"] else 0.0,
                    "rank_strength": 0.0,
                    "sector_strength": round(sector_strength, 2),
                    "rank_power": 0.0,
                    "avg_rank": round(item["rank_total"] / item["count"], 1) if item["count"] else 0.0,
                    "top20_count": item["top20_count"],
                    "top50_count": item["top50_count"],
                    "top20_ratio": round(top20_ratio, 4),
                    "top50_ratio": round(top50_ratio, 4),
                    "trading_value_100m": round(item["trading_value_100m"], 2),
                    "market_cap_100m": round(item["market_cap_100m"], 2),
                    "turnover_ratio": round(turnover_ratio, 6),
                    "turnover_score": round(min(max(turnover_ratio * 100.0, 0.0), 10.0) * 10.0, 2),
                    "confidence": "높음" if item["count"] >= 3 else "보통" if item["count"] == 2 else "낮음",
                    "leaders": item["leaders"],
                }
            )
        sectors.sort(key=lambda item: (-(item["sector_strength"] or 0.0), -(item["avg_score"] or 0.0), item["sector"]))
        days.append(
            {
                "date": file_date,
                "file_name": next((entry["file_name"] for entry in available_entries if entry["file_date"] == file_date), f"{file_date.replace('-', '')}_us_daily_screening.xlsx"),
                "qualified_count": len(rows),
                "assigned_count": len(rows),
                "top50_avg_score": round(float(np.mean([float(row.get("score") or 0.0) for row in rows[:50]])) if rows else 0.0, 2),
                "sectors": sectors[:5],
            }
        )
    payload = {"days": days}
    calendar_cache[calendar_key] = payload
    save_us_screening_cache(cache)
    return payload


def reload_us_screening_cache(request: ThemeReloadRequest) -> dict[str, Any]:
    cache = load_us_screening_cache()
    cache["summaries"] = {}
    cache["recent_leaders"] = {}
    cache["calendar"] = {}
    save_us_screening_cache(cache)
    return load_us_screening_summary(min_score=request.min_score, recent_limit=request.recent_limit, file_date=request.file_date)


def resolve_us_screening_market_date(target_date: str | None = None) -> str:
    date_key = re.sub(r"\D", "", str(target_date or datetime.now().strftime("%Y%m%d")))
    if not re.fullmatch(r"20\d{6}", date_key):
        raise ValueError("invalid market date")
    probe_start = datetime.strptime(date_key, "%Y%m%d").date()
    best_date = ""
    try:
        yahoo_response = requests.get(
            "https://query1.finance.yahoo.com/v8/finance/chart/AAPL",
            params={"range": "1mo", "interval": "1d"},
            headers={"User-Agent": "Mozilla/5.0"},
            timeout=20,
        )
        yahoo_response.raise_for_status()
        result = (yahoo_response.json().get("chart", {}).get("result") or [{}])[0]
        meta = result.get("meta") or {}
        market_tz = ZoneInfo(str(meta.get("exchangeTimezoneName") or "").strip() or "America/New_York")
        local_dates = [
            datetime.fromtimestamp(int(value), timezone.utc).astimezone(market_tz).strftime("%Y%m%d")
            for value in (result.get("timestamp") or [])
            if value and datetime.fromtimestamp(int(value), timezone.utc).astimezone(market_tz).date() <= probe_start
        ]
        if local_dates:
            best_date = max(local_dates)
    except Exception:
        best_date = ""
    for offset in range(8):
        probe = probe_start - timedelta(days=offset)
        start = (probe - timedelta(days=7)).isoformat()
        try:
            frame = fdr.DataReader("AAPL", start, probe.isoformat())
        except Exception:
            continue
        if frame is None or frame.empty:
            continue
        frame = frame.reset_index()
        if "Date" not in frame.columns:
            frame = frame.rename(columns={frame.columns[0]: "Date"})
        frame["Date"] = pd.to_datetime(frame["Date"], errors="coerce")
        frame = frame.dropna(subset=["Date"])
        if frame.empty:
            continue
        candidate = frame["Date"].max().strftime("%Y%m%d")
        if candidate > best_date:
            best_date = candidate
        break
    return best_date or date_key


def _load_screening_existing_date_keys(db_path: Path) -> set[str]:
    if not db_path.exists():
        return set()
    try:
        with sqlite3.connect(str(db_path)) as conn:
            rows = conn.execute("SELECT DISTINCT file_date_key FROM file_meta ORDER BY file_date_key ASC").fetchall()
    except Exception:
        return set()
    return {str(row[0] or "").strip() for row in rows if row and re.fullmatch(r"20\d{6}", str(row[0] or "").strip())}


def _collect_missing_market_dates(
    *,
    db_path: Path,
    target_compact: str,
    resolve_market_date: Callable[[str | None], str],
) -> list[str]:
    existing_dates = _load_screening_existing_date_keys(db_path)
    if existing_dates:
        start_date = datetime.strptime(max(existing_dates), "%Y%m%d").date() + timedelta(days=1)
    else:
        start_date = datetime.strptime(target_compact, "%Y%m%d").date()
    end_date = datetime.strptime(target_compact, "%Y%m%d").date()
    if start_date > end_date:
        return []

    missing_dates: list[str] = []
    seen_dates = set(existing_dates)
    probe = start_date
    while probe <= end_date:
        if probe.weekday() < 5:
            resolved_date = resolve_market_date(probe.strftime("%Y%m%d"))
            if re.fullmatch(r"20\d{6}", resolved_date) and resolved_date <= target_compact and resolved_date not in seen_dates:
                missing_dates.append(resolved_date)
                seen_dates.add(resolved_date)
        probe += timedelta(days=1)
    return missing_dates


def _run_daily_builder(script_path: Path, date_keys: list[str], *, timeout: int = 1800) -> list[dict[str, Any]]:
    build_results: list[dict[str, Any]] = []
    for date_key in date_keys:
        command = [sys.executable, "-u", str(script_path), "--date", date_key]
        result = subprocess.run(
            command,
            cwd=str(BASE_DIR),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=timeout,
        )
        if result.returncode != 0:
            raise RuntimeError((result.stderr or result.stdout or f"일자 데이터 생성 실패: {date_key}").strip())
        build_results.append(
            {
                "requested_date": date_key,
                "stdout": (result.stdout or "").strip(),
            }
        )
    return build_results


def create_us_theme_today_data_and_reload(request: ThemeBuildTodayExcelRequest) -> dict[str, Any]:
    requested_compact = datetime.now().strftime("%Y%m%d")
    today_compact = resolve_us_screening_market_date(requested_compact)
    today_iso = f"{today_compact[:4]}-{today_compact[4:6]}-{today_compact[6:]}"
    existing_date_keys = _load_screening_existing_date_keys(US_SCREENING_FAST_DB_PATH)
    script_path = BASE_DIR / "tools" / "build_us_stock_daily_single.py"
    missing_dates = _collect_missing_market_dates(
        db_path=US_SCREENING_FAST_DB_PATH,
        target_compact=today_compact,
        resolve_market_date=resolve_us_screening_market_date,
    )
    if today_compact not in existing_date_keys and today_compact not in missing_dates:
        missing_dates.append(today_compact)
    if not missing_dates:
        existing_rows = 0
        if US_SCREENING_FAST_DB_PATH.exists():
            try:
                with sqlite3.connect(str(US_SCREENING_FAST_DB_PATH)) as conn:
                    row = conn.execute(
                        "SELECT COUNT(*) FROM screening_rows WHERE file_date_key = ?",
                        (today_compact,),
                    ).fetchone()
                    existing_rows = int(row[0] or 0) if row else 0
            except Exception:
                existing_rows = 0
        payload = attach_us_screening_runtime_fields(build_us_screening_summary_from_sql(today_iso, float(request.min_score)))
        recent_limit = max(1, int(request.recent_limit))
        available_entries = us_screening_available_file_entries(limit=None)
        recent_dates = [item["file_date"] for item in reversed(available_entries[:recent_limit])]
        recent_summaries = load_us_screening_summaries_for_dates(recent_dates)
        payload["recent_leaders"] = build_recent_leader_stats_from_summaries(recent_summaries, min_score=float(request.min_score))
        payload["available_files"] = [{"file_name": item["file_name"], "file_date": item["file_date"]} for item in available_entries]
        payload["cache_loaded_at"] = datetime.now().isoformat(timespec="seconds")
        payload["cache_source"] = "sql"
        payload["requested_file_date"] = today_iso
        payload["fallback_file_date"] = ""
        payload["fallback_reason"] = ""
        payload["today_excel_build"] = {
            "ok": True,
            "date": today_iso,
            "requested_date": f"{requested_compact[:4]}-{requested_compact[4:6]}-{requested_compact[6:]}",
            "mode": "sql_cached",
            "rows": existing_rows,
            "built_dates": [],
        }
        return payload

    build_results = _run_daily_builder(script_path, missing_dates, timeout=1800)
    cache = load_us_screening_cache()
    cache["summaries"] = {}
    cache["recent_leaders"] = {}
    cache["calendar"] = {}
    save_us_screening_cache(cache)
    payload = load_us_screening_summary(min_score=float(request.min_score), recent_limit=int(request.recent_limit), file_date=today_iso)
    payload["today_excel_build"] = {
        "ok": True,
        "date": today_iso,
        "requested_date": f"{requested_compact[:4]}-{requested_compact[4:6]}-{requested_compact[6:]}",
        "mode": "sql_backfill",
        "built_dates": build_results,
    }
    return payload


def attach_asia_screening_runtime_fields(summary: dict[str, Any]) -> dict[str, Any]:
    payload = json.loads(json.dumps(summary, ensure_ascii=False))
    for row in payload.get("qualified_stocks", []):
        row["manual_sector"] = str(row.get("manual_sector") or row.get("theme") or "Other").strip() or "Other"
    return payload


def load_asia_screening_summary(
    min_score: float = 50.0,
    recent_limit: int = RECENT_SCREENING_LOOKBACK,
    file_date: str | None = None,
    region: str = "jp",
) -> dict[str, Any]:
    normalized_region = normalize_asia_theme_region(region)
    requested_date = ""
    if file_date:
        digits = re.sub(r"\D", "", str(file_date))
        requested_date = datetime.strptime(digits, "%Y%m%d").strftime("%Y-%m-%d") if len(digits) == 8 else str(file_date).strip()
    available_entries = asia_screening_available_file_entries(limit=None)
    available_dates = [item["file_date"] for item in available_entries]
    if not available_dates:
        raise FileNotFoundError("ASIA SQL 캐시에 주도주 데이터가 없습니다.")
    selected_date = requested_date if requested_date in available_dates else available_dates[0]
    fallback_used = bool(requested_date and requested_date != selected_date)
    payload = attach_asia_screening_runtime_fields(build_asia_screening_summary_from_sql(selected_date, min_score, region=normalized_region))
    recent_dates = [item["file_date"] for item in reversed(available_entries[: max(1, int(recent_limit))])]
    recent_summaries = load_asia_screening_summaries_for_dates(recent_dates)
    payload["recent_leaders"] = build_recent_leader_stats_from_summaries(
        [
            {
                **summary,
                "qualified_stocks": [
                    row
                    for row in list(summary.get("qualified_stocks") or [])
                    if infer_asia_theme_region(row.get("stock_code"), row.get("industry"), row.get("manual_sector") or row.get("theme")) == normalized_region
                ],
            }
            for summary in recent_summaries
        ],
        min_score=min_score,
    )
    payload["available_files"] = [{"file_name": item["file_name"], "file_date": item["file_date"]} for item in available_entries]
    payload["cache_loaded_at"] = datetime.now().isoformat(timespec="seconds")
    payload["cache_source"] = "sql"
    payload["requested_file_date"] = requested_date
    payload["fallback_file_date"] = selected_date if fallback_used else ""
    payload["fallback_reason"] = "requested_file_missing" if fallback_used else ""
    payload["selected_region"] = normalized_region
    payload["region_options"] = ASIA_THEME_REGION_OPTIONS
    return payload


def build_asia_theme_sector_calendar(min_score: float = 50.0, limit: int = 60, region: str = "jp") -> dict[str, Any]:
    normalized_region = normalize_asia_theme_region(region)
    cache = load_asia_screening_cache()
    calendar_cache = cache.setdefault("calendar", {})
    calendar_key = f"v{ASIA_SCREENING_CALENDAR_CACHE_VERSION}|{float(min_score):.4f}|{int(limit)}|{normalized_region}"
    cached_calendar = calendar_cache.get(calendar_key)
    if isinstance(cached_calendar, dict):
        return cached_calendar
    available_entries = asia_screening_available_file_entries(limit=max(1, int(limit)))
    source_summaries = load_asia_screening_summaries_for_dates([item["file_date"] for item in reversed(available_entries)])
    days: list[dict[str, Any]] = []
    for summary in source_summaries:
        file_date = str(summary.get("file_date") or "")
        rows = summary.get("qualified_stocks") or []
        if not file_date or not rows:
            continue
        sector_map: dict[str, dict[str, Any]] = {}
        for row in rows:
            if infer_asia_theme_region(row.get("stock_code"), row.get("industry"), row.get("manual_sector") or row.get("theme")) != normalized_region:
                continue
            sector = str(row.get("manual_sector") or row.get("theme") or "Other").strip() or "Other"
            bucket = sector_map.setdefault(
                sector,
                {
                    "sector": sector,
                    "count": 0,
                    "score_total": 0.0,
                    "rank_total": 0.0,
                    "top20_count": 0,
                    "top50_count": 0,
                    "trading_value_100m": 0.0,
                    "market_cap_100m": 0.0,
                    "leaders": [],
                },
            )
            rank_value = int(to_float(row.get("rank")) or (len(rows) + 1))
            bucket["count"] += 1
            bucket["score_total"] += float(row.get("score") or 0.0)
            bucket["rank_total"] += rank_value
            if rank_value <= 20:
                bucket["top20_count"] += 1
            if rank_value <= 50:
                bucket["top50_count"] += 1
            bucket["trading_value_100m"] += float(row.get("trading_value_100m") or 0.0)
            bucket["market_cap_100m"] += float(row.get("market_cap_100m") or 0.0)
            if len(bucket["leaders"]) < 4:
                bucket["leaders"].append(str(row.get("stock_name") or ""))
        sectors = []
        for item in sector_map.values():
            turnover_ratio = item["trading_value_100m"] / item["market_cap_100m"] if item["market_cap_100m"] else 0.0
            top20_ratio = item["top20_count"] / item["count"] if item["count"] else 0.0
            sector_strength = (
                (item["score_total"] / item["count"] if item["count"] else 0.0) * 0.6
                + top20_ratio * 100.0 * 0.2
                + min(max(turnover_ratio * 100.0, 0.0), 10.0) * 10.0 * 0.2
            )
            sectors.append(
                {
                    "sector": item["sector"],
                    "count": item["count"],
                    "avg_score": round(item["score_total"] / item["count"], 2) if item["count"] else 0.0,
                    "rank_strength": 0.0,
                    "sector_strength": round(sector_strength, 2),
                    "rank_power": 0.0,
                    "avg_rank": round(item["rank_total"] / item["count"], 1) if item["count"] else 0.0,
                    "top20_count": item["top20_count"],
                    "top50_count": item["top50_count"],
                    "top20_ratio": round(top20_ratio, 4),
                    "top50_ratio": round(item["top50_count"] / item["count"], 4) if item["count"] else 0.0,
                    "trading_value_100m": round(item["trading_value_100m"], 2),
                    "market_cap_100m": round(item["market_cap_100m"], 2),
                    "turnover_ratio": round(turnover_ratio, 6),
                    "turnover_score": round(min(max(turnover_ratio * 100.0, 0.0), 10.0) * 10.0, 2),
                    "confidence": "높음" if item["count"] >= 3 else "보통" if item["count"] == 2 else "낮음",
                    "leaders": item["leaders"],
                }
            )
        sectors.sort(key=lambda item: (-(item["sector_strength"] or 0.0), -(item["avg_score"] or 0.0), item["sector"]))
        filtered_rows = [
            row
            for row in rows
            if infer_asia_theme_region(row.get("stock_code"), row.get("industry"), row.get("manual_sector") or row.get("theme")) == normalized_region
        ]
        if not filtered_rows:
            continue
        days.append(
            {
                "date": file_date,
                "file_name": next((entry["file_name"] for entry in available_entries if entry["file_date"] == file_date), f"{file_date.replace('-', '')}_asia_daily_screening.xlsx"),
                "qualified_count": len(filtered_rows),
                "assigned_count": len(filtered_rows),
                "top50_avg_score": round(float(np.mean([float(row.get("score") or 0.0) for row in filtered_rows[:50]])) if filtered_rows else 0.0, 2),
                "sectors": sectors[:5],
            }
        )
    payload = {"days": days, "selected_region": normalized_region, "region_options": ASIA_THEME_REGION_OPTIONS}
    calendar_cache[calendar_key] = payload
    save_asia_screening_cache(cache)
    return payload


def reload_asia_screening_cache(request: ThemeReloadRequest) -> dict[str, Any]:
    cache = load_asia_screening_cache()
    cache["summaries"] = {}
    cache["recent_leaders"] = {}
    cache["calendar"] = {}
    save_asia_screening_cache(cache)
    return load_asia_screening_summary(
        min_score=request.min_score,
        recent_limit=request.recent_limit,
        file_date=request.file_date,
        region=request.region or "jp",
    )


def resolve_asia_screening_market_date(target_date: str | None = None) -> str:
    date_key = re.sub(r"\D", "", str(target_date or datetime.now().strftime("%Y%m%d")))
    if not re.fullmatch(r"20\d{6}", date_key):
        raise ValueError("invalid market date")
    probe_start = datetime.strptime(date_key, "%Y%m%d").date()
    sentinels = ["TSE:7203", "SSE:601288", "SZSE:000333"]
    best_date = ""
    for yahoo_symbol, default_tz in [("7203.T", "Asia/Tokyo"), ("601288.SS", "Asia/Shanghai"), ("000333.SZ", "Asia/Shanghai"), ("2330.TW", "Asia/Taipei")]:
        try:
            yahoo_response = requests.get(
                f"https://query1.finance.yahoo.com/v8/finance/chart/{yahoo_symbol}",
                params={"range": "1mo", "interval": "1d"},
                headers={"User-Agent": "Mozilla/5.0"},
                timeout=20,
            )
            yahoo_response.raise_for_status()
            result = (yahoo_response.json().get("chart", {}).get("result") or [{}])[0]
            meta = result.get("meta") or {}
            market_tz = ZoneInfo(str(meta.get("exchangeTimezoneName") or "").strip() or default_tz)
            local_dates = [
                datetime.fromtimestamp(int(value), timezone.utc).astimezone(market_tz).strftime("%Y%m%d")
                for value in (result.get("timestamp") or [])
                if value and datetime.fromtimestamp(int(value), timezone.utc).astimezone(market_tz).date() <= probe_start
            ]
            if local_dates:
                candidate = max(local_dates)
                if candidate > best_date:
                    best_date = candidate
        except Exception:
            continue
    for sentinel in sentinels:
        for offset in range(8):
            probe = probe_start - timedelta(days=offset)
            start = (probe - timedelta(days=7)).isoformat()
            try:
                frame = fdr.DataReader(sentinel, start, probe.isoformat())
            except Exception:
                continue
            if frame is None or frame.empty:
                continue
            frame = frame.reset_index()
            if "Date" not in frame.columns:
                frame = frame.rename(columns={frame.columns[0]: "Date"})
            frame["Date"] = pd.to_datetime(frame["Date"], errors="coerce")
            frame = frame.dropna(subset=["Date"])
            if frame.empty:
                continue
            candidate = frame["Date"].max().strftime("%Y%m%d")
            if candidate > best_date:
                best_date = candidate
            break
    return best_date or date_key


def create_asia_theme_today_data_and_reload(request: ThemeBuildTodayExcelRequest) -> dict[str, Any]:
    normalized_region = normalize_asia_theme_region(request.region or "jp")
    requested_compact = datetime.now().strftime("%Y%m%d")
    today_compact = resolve_asia_screening_market_date(requested_compact)
    today_iso = f"{today_compact[:4]}-{today_compact[4:6]}-{today_compact[6:]}"
    existing_date_keys = _load_screening_existing_date_keys(ASIA_SCREENING_FAST_DB_PATH)
    script_path = BASE_DIR / "tools" / "build_asia_stock_daily_single.py"
    missing_dates = _collect_missing_market_dates(
        db_path=ASIA_SCREENING_FAST_DB_PATH,
        target_compact=today_compact,
        resolve_market_date=resolve_asia_screening_market_date,
    )
    if today_compact not in existing_date_keys and today_compact not in missing_dates:
        missing_dates.append(today_compact)
    if not missing_dates:
        existing_rows = 0
        if ASIA_SCREENING_FAST_DB_PATH.exists():
            try:
                with sqlite3.connect(str(ASIA_SCREENING_FAST_DB_PATH)) as conn:
                    row = conn.execute(
                        "SELECT COUNT(*) FROM screening_rows WHERE file_date_key = ?",
                        (today_compact,),
                    ).fetchone()
                    existing_rows = int(row[0] or 0) if row else 0
            except Exception:
                existing_rows = 0
        payload = attach_asia_screening_runtime_fields(build_asia_screening_summary_from_sql(today_iso, float(request.min_score), region=normalized_region))
        recent_limit = max(1, int(request.recent_limit))
        available_entries = asia_screening_available_file_entries(limit=None)
        recent_dates = [item["file_date"] for item in reversed(available_entries[:recent_limit])]
        recent_summaries = load_asia_screening_summaries_for_dates(recent_dates)
        payload["recent_leaders"] = build_recent_leader_stats_from_summaries(
            [
                {
                    **summary,
                    "qualified_stocks": [
                        row
                        for row in list(summary.get("qualified_stocks") or [])
                        if infer_asia_theme_region(row.get("stock_code"), row.get("industry"), row.get("manual_sector") or row.get("theme")) == normalized_region
                    ],
                }
                for summary in recent_summaries
            ],
            min_score=float(request.min_score),
        )
        payload["available_files"] = [{"file_name": item["file_name"], "file_date": item["file_date"]} for item in available_entries]
        payload["cache_loaded_at"] = datetime.now().isoformat(timespec="seconds")
        payload["cache_source"] = "sql"
        payload["requested_file_date"] = today_iso
        payload["fallback_file_date"] = ""
        payload["fallback_reason"] = ""
        payload["selected_region"] = normalized_region
        payload["region_options"] = ASIA_THEME_REGION_OPTIONS
        payload["today_excel_build"] = {
            "ok": True,
            "date": today_iso,
            "requested_date": f"{requested_compact[:4]}-{requested_compact[4:6]}-{requested_compact[6:]}",
            "mode": "sql_cached",
            "rows": existing_rows,
            "built_dates": [],
        }
        return payload

    build_results = _run_daily_builder(script_path, missing_dates, timeout=1800)
    cache = load_asia_screening_cache()
    cache["summaries"] = {}
    cache["recent_leaders"] = {}
    cache["calendar"] = {}
    save_asia_screening_cache(cache)
    payload = load_asia_screening_summary(
        min_score=float(request.min_score),
        recent_limit=int(request.recent_limit),
        file_date=today_iso,
        region=normalized_region,
    )
    payload["today_excel_build"] = {
        "ok": True,
        "date": today_iso,
        "requested_date": f"{requested_compact[:4]}-{requested_compact[4:6]}-{requested_compact[6:]}",
        "mode": "sql_backfill",
        "built_dates": build_results,
    }
    return payload


def reload_screening_cache(request: ThemeReloadRequest) -> dict[str, Any]:
    cache = load_screening_cache()
    cache["summaries"] = {}
    cache["recent_leaders"] = {}
    cache["calendar"] = {}
    save_screening_cache(cache)
    _screening_backtest_source_summaries_cached.cache_clear()
    build_stock_sector_entry_markers.cache_clear()
    build_sector_entry_markers_for_sector.cache_clear()
    load_stock_chart_preview_cached.cache_clear()
    ensure_screening_db_indexes.cache_clear()
    return load_screening_summary(min_score=request.min_score, recent_limit=request.recent_limit, file_date=request.file_date, force_reload=True)


def invalidate_screening_runtime_caches() -> None:
    _screening_backtest_source_summaries_cached.cache_clear()
    build_stock_sector_entry_markers.cache_clear()
    build_sector_entry_markers_for_sector.cache_clear()
    load_stock_chart_preview_cached.cache_clear()
    build_stock_entry_state_map.cache_clear()
    ensure_screening_db_indexes.cache_clear()


def normalize_screening_compact_date(file_date: str | None) -> str:
    if file_date:
        digits = re.sub(r"\D", "", str(file_date))
        if re.fullmatch(r"20\d{6}", digits):
            return digits
    selected_file = get_latest_screening_file()
    parsed = parse_screening_date(selected_file)
    digits = re.sub(r"\D", "", parsed)
    if not re.fullmatch(r"20\d{6}", digits):
        raise ValueError(f"조회 날짜를 YYYYMMDD로 변환하지 못했습니다: {file_date or parsed}")
    return digits


def create_theme_test_excel(request: ThemeTestExcelRequest) -> dict[str, Any]:
    target_date = normalize_screening_compact_date(request.file_date)
    script_path = BASE_DIR / "tools" / "generate_leader_screening_test.py"
    if not script_path.exists():
        raise FileNotFoundError(f"테스트 엑셀 생성 스크립트를 찾지 못했습니다: {script_path}")
    suffix = request.suffix or "_test_52w"
    command = [sys.executable, "-u", str(script_path), target_date, suffix]
    result = subprocess.run(
        command,
        cwd=str(BASE_DIR),
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=420,
    )
    output_path = SCREENING_DIR / f"{target_date}_데일리_기업스크리닝{suffix if suffix.startswith('_') else '_' + suffix}.xlsm"
    if result.returncode != 0:
        raise RuntimeError((result.stderr or result.stdout or "테스트 엑셀 생성 실패").strip())
    if not output_path.exists():
        created_match = re.search(r"created=(.+)", result.stdout or "")
        if created_match:
            output_path = Path(created_match.group(1).strip())
    return {
        "ok": True,
        "file_date": f"{target_date[:4]}-{target_date[4:6]}-{target_date[6:]}",
        "file_name": output_path.name,
        "path": str(output_path),
        "stdout": result.stdout,
    }


def resolve_screening_market_date(target_date: str | None = None) -> str:
    date_key = re.sub(r"\D", "", str(target_date or datetime.now().strftime("%Y%m%d")))
    if not re.fullmatch(r"20\d{6}", date_key):
        raise ValueError("invalid market date")
    if pykrx_stock is None:
        return date_key
    start_dt = datetime.strptime(date_key, "%Y%m%d")
    checked_dates: set[str] = set()
    for offset in range(8):
        probe_key = (start_dt - timedelta(days=offset)).strftime("%Y%m%d")
        try:
            effective_date = pykrx_stock.get_nearest_business_day_in_a_week(probe_key) or probe_key
        except Exception:
            effective_date = probe_key
        if effective_date in checked_dates:
            continue
        checked_dates.add(effective_date)
        try:
            frame = pykrx_stock.get_market_ohlcv_by_ticker(effective_date, market="ALL")
        except Exception:
            continue
        if frame is not None and not frame.empty:
            return effective_date
    return date_key


def create_theme_today_excel_and_reload(request: ThemeBuildTodayExcelRequest) -> dict[str, Any]:
    requested_compact = datetime.now().strftime("%Y%m%d")
    today_compact = resolve_screening_market_date(requested_compact)
    existing_rows = 0
    if SCREENING_FAST_DB_PATH.exists():
        try:
            with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
                row = conn.execute(
                    "SELECT COUNT(*) FROM screening_rows WHERE file_date_key = ?",
                    (today_compact,),
                ).fetchone()
                existing_rows = int(row[0] or 0) if row else 0
        except Exception:
            existing_rows = 0

    today_iso = f"{today_compact[:4]}-{today_compact[4:6]}-{today_compact[6:]}"
    if existing_rows > 0:
        return {
            "ok": True,
            "file_date": today_iso,
            "today_excel_build": {
                "ok": True,
                "date": today_iso,
                "requested_date": f"{requested_compact[:4]}-{requested_compact[4:6]}-{requested_compact[6:]}",
                "mode": "sql_cached",
                "rows": existing_rows,
            },
        }

    script_path = BASE_DIR / "tools" / "build_stock_daily_single.py"
    if not script_path.exists():
        raise FileNotFoundError(f"오늘자 데이터 생성 스크립트를 찾지 못했습니다: {script_path}")

    command = [sys.executable, "-u", str(script_path), "--date", today_compact, "--sql-only"]
    env = os.environ.copy()
    krx = get_krx_settings()
    if krx.get("id"):
        env["KRX_ID"] = str(krx["id"])
    if krx.get("password"):
        env["KRX_PW"] = str(krx["password"])
    result = subprocess.run(
        command,
        cwd=str(BASE_DIR),
        env=env,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
        timeout=900,
    )
    if result.returncode != 0:
        raise RuntimeError((result.stderr or result.stdout or "오늘자 데이터 생성 실패").strip())
    invalidate_screening_runtime_caches()
    return {
        "ok": True,
        "file_date": today_iso,
        "today_excel_build": {
            "ok": True,
            "date": today_iso,
            "requested_date": f"{requested_compact[:4]}-{requested_compact[4:6]}-{requested_compact[6:]}",
            "stdout": result.stdout,
            "mode": "sql_direct",
        },
    }


def build_stock_score_history(
    stock_code: str | None = None,
    stock_name: str | None = None,
    end_date: str | None = None,
    days: int = 31,
) -> dict[str, Any]:
    target_code = normalize_stock_code_value(stock_code)
    target_name = str(stock_name or "").strip()
    normalized_target_name = normalize_text(target_name)
    available_entries = screening_available_file_entries(limit=None)
    available_dates = [item["file_date"] for item in available_entries]
    if not available_dates:
        raise FileNotFoundError("SQL 캐시에 주도주 데이터가 없습니다.")
    requested_date = ""
    if end_date:
        digits = re.sub(r"\D", "", str(end_date))
        requested_date = datetime.strptime(digits, "%Y%m%d").strftime("%Y-%m-%d") if len(digits) == 8 else str(end_date).strip()
    selected_date = requested_date if requested_date in available_dates else available_dates[0]
    end_dt = datetime.strptime(selected_date, "%Y-%m-%d")
    start_dt = end_dt - timedelta(days=max(7, min(int(days or 31), 90)) - 1)
    rows: list[dict[str, Any]] = []

    if SCREENING_FAST_DB_PATH.exists():
        start_key = start_dt.strftime("%Y%m%d")
        end_key = end_dt.strftime("%Y%m%d")
        where_sql = ""
        params: list[Any] = [start_key, end_key]
        if target_code:
            where_sql = "AND r.stock_code = ?"
            params.append(target_code)
        elif normalized_target_name:
            where_sql = "AND REPLACE(LOWER(r.stock_name), ' ', '') = ?"
            params.append(normalized_target_name)
        else:
            return {
                "stock_code": "",
                "stock_name": target_name,
                "end_date": selected_date,
                "start_date": start_dt.strftime("%Y-%m-%d"),
                "rows": [],
                "summary": {"count": 0, "latest_score": None, "max_score": None, "avg_score": None},
            }

        with sqlite3.connect(str(SCREENING_FAST_DB_PATH)) as conn:
            query = f"""
                WITH ranked_rows AS (
                    SELECT
                        r.file_date AS date,
                        r.file_date_key,
                        r.stock_code,
                        r.stock_name,
                        r.score_s AS score,
                        RANK() OVER (PARTITION BY r.file_date_key ORDER BY r.score_s DESC, r.stock_code ASC) AS day_rank,
                        r.change_pct,
                        r.avg_1m AS avg_1m,
                        r.avg_3m AS avg_3m,
                        r.avg_1w,
                        r.note,
                        r.market_cap_100m,
                        r.trading_value_100m
                    FROM screening_rows r
                    WHERE r.file_date_key BETWEEN ? AND ?
                )
                SELECT
                    rr.date,
                    rr.stock_code,
                    rr.stock_name,
                    rr.score,
                    rr.day_rank,
                    rr.change_pct,
                    rr.avg_1m,
                    rr.avg_1w,
                    rr.note,
                    rr.market_cap_100m,
                    rr.trading_value_100m,
                    COALESCE(c.close_price, NULL) AS close_price
                FROM ranked_rows rr
                LEFT JOIN daily_close_cache c
                  ON c.file_date_key = rr.file_date_key
                 AND c.stock_code = rr.stock_code
                WHERE 1=1
                {where_sql.replace('r.', 'rr.')}
                ORDER BY rr.file_date_key
            """
            history_df = pd.read_sql_query(query, conn, params=params)

        if not history_df.empty:
            history_df["score"] = pd.to_numeric(history_df["score"], errors="coerce")
            history_df["day_rank"] = pd.to_numeric(history_df["day_rank"], errors="coerce")
            history_df["change_pct"] = pd.to_numeric(history_df["change_pct"], errors="coerce")
            history_df["avg_1m"] = pd.to_numeric(history_df["avg_1m"], errors="coerce")
            history_df["avg_1w"] = pd.to_numeric(history_df["avg_1w"], errors="coerce")
            history_df["close_price"] = pd.to_numeric(history_df["close_price"], errors="coerce")
            history_df = history_df.dropna(subset=["score"]).copy()
            rows = [
                {
                    "date": str(row["date"]),
                    "score": None if pd.isna(row["score"]) else round(float(row["score"]), 2),
                    "change_pct": None if pd.isna(row["change_pct"]) else round(float(row["change_pct"]), 2),
                    "rank": 0 if pd.isna(row["day_rank"]) else int(row["day_rank"]),
                    "avg_1m": None if pd.isna(row["avg_1m"]) else round(float(row["avg_1m"]), 2),
                    "avg_1w": None if pd.isna(row["avg_1w"]) else round(float(row["avg_1w"]), 2),
                    "note": str(row["note"] or ""),
                    "stock_code": str(row["stock_code"] or ""),
                    "stock_name": str(row["stock_name"] or target_name),
                    "resolved_name": str(row["stock_name"] or target_name),
                    "close": None if pd.isna(row["close_price"]) else round(float(row["close_price"]), 2),
                }
                for _, row in history_df.iterrows()
            ]

    rows = sorted(rows, key=lambda item: item["date"])
    if rows:
        normalized_close = 100.0
        for index, item in enumerate(rows):
            change_pct = to_float(item.get("change_pct"))
            if index == 0:
                item["close_normalized"] = round(normalized_close, 2)
                continue
            if change_pct is None:
                item["close_normalized"] = round(normalized_close, 2)
                continue
            normalized_close *= 1.0 + (float(change_pct) / 100.0)
            item["close_normalized"] = round(normalized_close, 2)
    display_name = target_name
    display_code = target_code
    for item in reversed(rows):
        display_name = item.get("resolved_name") or item.get("stock_name") or display_name
        display_code = item.get("stock_code") or display_code
        if display_name:
            break

    # SQL close cache가 없을 때만 FDR fallback.
    if rows and display_code and any(item.get("close") is None for item in rows):
        try:
            price_frame = fetch_price_frame(display_code)
            if not price_frame.empty:
                for item in rows:
                    if item.get("close") is None:
                        close_on_date = price_close_on_or_before(price_frame, item["date"])
                        item["close"] = round(float(close_on_date), 2) if close_on_date is not None else None
        except Exception:
            pass

    return {
        "stock_code": display_code,
        "stock_name": display_name,
        "end_date": selected_date,
        "start_date": start_dt.strftime("%Y-%m-%d"),
        "rows": rows,
        "summary": {
            "count": len(rows),
            "latest_score": rows[-1].get("score") if rows else None,
            "max_score": max((float(row.get("score") or 0) for row in rows), default=None),
            "avg_score": round(
                sum(float(row.get("score") or 0) for row in rows) / len(rows),
                2,
            ) if rows else None,
        },
    }


def build_theme_sector_calendar(
    min_score: float = 50.0,
    limit: int = 60,
    force_refresh: bool = False,
    score_basis: str = "score",
) -> dict[str, Any]:
    sector_db = load_sector_db()
    cache = load_screening_cache()
    calendar_cache = cache.setdefault("calendar", {})
    sector_db_version = SECTOR_DB_PATH.stat().st_mtime_ns if SECTOR_DB_PATH.exists() else 0
    screening_version = screening_data_version_token()
    normalized_basis = "score_o" if str(score_basis or "").strip().lower() in {"score_o", "daily", "today"} else "score"
    calendar_key = f"v{SCREENING_CALENDAR_CACHE_VERSION}|{screening_version}|{sector_db_version}|{float(min_score):.4f}|{int(limit)}|{normalized_basis}"
    cached_calendar = calendar_cache.get(calendar_key)
    if not force_refresh and isinstance(cached_calendar, dict):
        return cached_calendar
    available_entries = screening_available_file_entries(limit=max(1, int(limit)))
    source_summaries = load_screening_summaries_for_dates([item["file_date"] for item in reversed(available_entries)])
    days: list[dict[str, Any]] = []
    for summary in source_summaries:
        file_date = str(summary.get("file_date") or "")
        if not file_date:
            continue
        rows = summary.get("qualified_stocks", [])
        if not rows:
            continue
        ranked_rows = sorted(
            rows,
            key=lambda row: (
                float(to_float(row.get(normalized_basis)) or -999999.0),
                float(to_float(row.get("score")) or -999999.0),
                str(row.get("stock_code") or ""),
            ),
            reverse=True,
        )
        rank_by_stock_key: dict[str, int] = {}
        for index_no, ranked_row in enumerate(ranked_rows, start=1):
            stock_key = sector_rotation_stock_key(ranked_row)
            if stock_key:
                rank_by_stock_key[stock_key] = index_no
        sector_map: dict[str, dict[str, Any]] = {}
        assigned_count = 0
        for row in rows:
            stock_name = str(row.get("stock_name") or "").strip()
            stock_code = str(row.get("stock_code") or "").strip()
            sector = resolve_sector_for_stock(stock_code, stock_name, sector_db)
            if not sector:
                continue
            trading_value = float(row.get("trading_value_100m") or 0.0)
            assigned_count += 1
            bucket = sector_map.setdefault(
                sector,
                {
                    "sector": sector,
                    "count": 0,
                    "score_total": 0.0,
                    "rank_strength": 0.0,
                    "rank_total": 0.0,
                    "rank_scores": [],
                    "top20_count": 0,
                    "top50_count": 0,
                    "trading_value_100m": 0.0,
                    "market_cap_100m": 0.0,
                    "leaders": [],
                },
            )
            bucket["count"] += 1
            bucket["score_total"] += float(to_float(row.get(normalized_basis)) or 0.0)
            stock_key = sector_rotation_stock_key(row)
            rank_value = int(rank_by_stock_key.get(stock_key) or len(ranked_rows) + 1)
            rank_weight = max(0, 121 - rank_value)
            rank_score = max(0.0, (121.0 - float(rank_value)) / 120.0 * 100.0)
            bucket["rank_strength"] += float(rank_weight)
            bucket["rank_total"] += float(rank_value)
            bucket["rank_scores"].append(rank_score)
            if rank_value <= 20:
                bucket["top20_count"] += 1
            if rank_value <= 50:
                bucket["top50_count"] += 1
            bucket["trading_value_100m"] += float(trading_value or 0.0)
            bucket["market_cap_100m"] += float(row.get("market_cap_100m") or 0.0)
            if len(bucket["leaders"]) < 4:
                bucket["leaders"].append(stock_name)
        sectors = []
        for item in sector_map.values():
            turnover_ratio = (
                item["trading_value_100m"] / item["market_cap_100m"]
                if item["market_cap_100m"]
                else 0.0
            )
            rank_scores = sorted(item.get("rank_scores") or [], reverse=True)
            top_rank_scores = rank_scores[:3] if rank_scores else [0.0]
            rank_power = sum(top_rank_scores) / len(top_rank_scores)
            top20_ratio = item["top20_count"] / item["count"] if item["count"] else 0.0
            top50_ratio = item["top50_count"] / item["count"] if item["count"] else 0.0
            turnover_ratio_pct = turnover_ratio * 100
            turnover_score = min(max(turnover_ratio_pct, 0.0), 10.0) * 10.0
            sector_strength = (
                rank_power * 0.45
                + (item["score_total"] / item["count"] if item["count"] else 0.0) * 0.25
                + top20_ratio * 100.0 * 0.15
                + turnover_score * 0.15
            )
            confidence = "높음" if item["count"] >= 3 else "보통" if item["count"] == 2 else "낮음"
            sectors.append(
                {
                    "sector": item["sector"],
                    "count": item["count"],
                    "avg_score": round(item["score_total"] / item["count"], 2) if item["count"] else 0,
                    "rank_strength": round(item["rank_strength"], 2),
                    "sector_strength": round(sector_strength, 2),
                    "rank_power": round(rank_power, 2),
                    "avg_rank": round(item["rank_total"] / item["count"], 1) if item["count"] else 0,
                    "top20_count": item["top20_count"],
                    "top50_count": item["top50_count"],
                    "top20_ratio": round(top20_ratio, 4),
                    "top50_ratio": round(top50_ratio, 4),
                    "trading_value_100m": round(item["trading_value_100m"], 1),
                    "market_cap_100m": round(item["market_cap_100m"], 1),
                    "turnover_ratio": round(turnover_ratio, 6),
                    "turnover_score": round(turnover_score, 2),
                    "confidence": confidence,
                    "leaders": item["leaders"],
                }
            )
        sectors.sort(
            key=lambda item: (
                item["sector_strength"],
                item["rank_power"],
                item["avg_score"],
                item["turnover_score"],
                item["top20_ratio"],
            ),
            reverse=True,
        )
        if assigned_count <= 0:
            continue
        top10_rows = ranked_rows[:10]
        top10_scores = [float(to_float(row.get(normalized_basis)) or 0.0) for row in top10_rows]
        top10_avg_score = float(np.mean(top10_scores)) if top10_scores else 0.0
        days.append(
            {
                "date": file_date,
                "file_name": next((item["file_name"] for item in available_entries if item["file_date"] == file_date), f"{file_date.replace('-', '')}_데일리_기업스크리닝.xlsx"),
                "qualified_count": int(len(rows)),
                "assigned_count": assigned_count,
                "top10_avg_score": round(top10_avg_score, 2),
                "top50_avg_score": round(top10_avg_score, 2),
                "sectors": sectors[:6],
            }
        )
    payload = {
        "days": days,
        "sector_count": len(sector_db.get("sectors", [])),
        "stock_count": len(sector_db.get("stock_map", {})),
        "cache_refreshed": bool(force_refresh),
        "score_basis": normalized_basis,
    }
    calendar_cache[calendar_key] = payload
    save_screening_cache(cache)
    return payload


def split_sector_group_detail(sector_name: str) -> tuple[str, str]:
    raw = str(sector_name or "").strip()
    if not raw:
        return ("미분류", "")
    # 섹터 그룹 분해 기능 롤백: 섹터 원문 자체를 그룹으로 취급한다.
    return (raw, raw)


def build_sector_cycle_clock_payload(min_score: float = 50.0, limit: int = 40, force_refresh: bool = False) -> dict[str, Any]:
    calendar = build_theme_sector_calendar(min_score=min_score, limit=max(20, min(limit, 120)), force_refresh=force_refresh)
    days = sorted([item for item in (calendar.get("days") or []) if item.get("date")], key=lambda item: str(item.get("date")))
    if len(days) < 3:
        return {"groups": [], "latest_date": "", "error": "섹터 순환 시계를 계산할 데이터가 부족합니다."}
    latest = days[-1]
    prev = days[-2]
    prev2 = days[-3]
    prev_map = {str(item.get("sector") or "").strip(): item for item in (prev.get("sectors") or [])}
    prev2_map = {str(item.get("sector") or "").strip(): item for item in (prev2.get("sectors") or [])}

    group_bucket: dict[str, dict[str, Any]] = {}
    for item in (latest.get("sectors") or []):
        sector = str(item.get("sector") or "").strip()
        if not sector:
            continue
        group, detail = split_sector_group_detail(sector)
        prev_item = prev_map.get(sector, {})
        prev2_item = prev2_map.get(sector, {})
        cur_strength = float(item.get("sector_strength") or 0.0)
        prev_strength = float(prev_item.get("sector_strength") or cur_strength)
        prev2_strength = float(prev2_item.get("sector_strength") or prev_strength)
        momentum = (cur_strength - prev_strength) * 0.65 + (prev_strength - prev2_strength) * 0.35
        grp = group_bucket.setdefault(group, {"group": group, "details": [], "strength_values": [], "momentum_values": []})
        grp["details"].append({
            "sector": sector,
            "detail": detail or sector,
            "strength": round(cur_strength, 2),
            "momentum": round(momentum, 2),
            "avg_score": float(item.get("avg_score") or 0.0),
            "leaders": item.get("leaders") or [],
            "count": int(item.get("count") or 0),
        })
        grp["strength_values"].append(cur_strength)
        grp["momentum_values"].append(momentum)

    def phase_of(x: float, y: float) -> str:
        if x >= 0 and y >= 0:
            return "상승"
        if x < 0 and y >= 0:
            return "회복"
        if x >= 0 and y < 0:
            return "둔화"
        return "하강"

    groups = []
    for group_name, grp in group_bucket.items():
        strength = float(np.mean(grp["strength_values"])) if grp["strength_values"] else 0.0
        momentum = float(np.mean(grp["momentum_values"])) if grp["momentum_values"] else 0.0
        x = max(-4.0, min(4.0, (strength - 50.0) / 8.0))
        y = max(-4.0, min(4.0, momentum / 4.0))
        details = sorted(grp["details"], key=lambda d: (float(d.get("strength") or 0.0), float(d.get("momentum") or 0.0)), reverse=True)
        groups.append({
            "group": group_name,
            "x": round(x, 3),
            "y": round(y, 3),
            "phase": phase_of(x, y),
            "strength": round(strength, 2),
            "momentum": round(momentum, 2),
            "detail_count": len(details),
            "details": details,
        })
    groups.sort(key=lambda g: (float(g.get("strength") or 0.0), float(g.get("momentum") or 0.0)), reverse=True)
    return {
        "latest_date": latest.get("date") or "",
        "groups": groups,
        "source_note": "종합점수 기반 섹터 강도(sector_strength)와 3일 변화량 모멘텀으로 그룹 섹터 순환 위치를 계산합니다.",
    }


def get_dart_client() -> OpenDartReader.OpenDartReader | None:
    api_key = get_dart_api_key()
    if not api_key:
        return None
    try:
        return OpenDartReader(api_key)
    except Exception:
        return None


DART_SUMMARY_KEYWORDS = [
    "매출",
    "영업",
    "순이익",
    "손익",
    "계약",
    "수주",
    "공급",
    "투자",
    "취득",
    "처분",
    "증자",
    "배당",
    "자기주식",
    "합병",
    "분할",
    "최대주주",
    "전환사채",
    "신규시설",
    "영업정지",
]


def classify_dart_report(report_name: str) -> str:
    name = str(report_name or "")
    rules = [
        ("실적/손익", ["손익구조", "영업실적", "잠정", "매출액", "실적"]),
        ("수주/계약", ["단일판매", "공급계약", "수주", "계약"]),
        ("투자/취득", ["시설투자", "타법인", "취득", "처분", "투자"]),
        ("자금조달", ["유상증자", "무상증자", "전환사채", "신주인수권", "자금조달"]),
        ("배당/자사주", ["배당", "자기주식", "자사주", "소각"]),
        ("지배구조", ["최대주주", "대표이사", "임원", "주주총회"]),
        ("정기보고", ["사업보고서", "분기보고서", "반기보고서"]),
    ]
    for label, tokens in rules:
        if any(token in name for token in tokens):
            return label
    return "일반공시"


def clean_dart_document_text(raw_text: Any) -> str:
    text = str(raw_text or "")
    text = re.sub(r"<!\[CDATA\[(.*?)\]\]>", r"\1", text, flags=re.DOTALL)
    text = re.sub(r"<br\s*/?>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"</(p|tr|table|section|div|title)>", "\n", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html_lib.unescape(text)
    text = re.sub(r"[ \t\r\f\v]+", " ", text)
    text = re.sub(r"\n\s+", "\n", text)
    text = re.sub(r"\n{2,}", "\n", text)
    return text.strip()


def extract_dart_highlights(document_text: str, report_name: str) -> list[str]:
    text = clean_dart_document_text(document_text)
    if not text:
        return []
    fragments = re.split(r"(?<=[.。!?])\s+|\n+", text)
    highlights: list[str] = []
    seen: set[str] = set()
    for fragment in fragments:
        cleaned = re.sub(r"\s+", " ", fragment).strip(" -ㆍ·")
        if len(cleaned) < 12 or len(cleaned) > 260:
            continue
        normalized = normalize_text(cleaned)
        if normalized in seen:
            continue
        if any(token in cleaned for token in DART_SUMMARY_KEYWORDS) or any(token in cleaned for token in re.findall(r"[가-힣A-Za-z0-9]+", report_name)[:4]):
            seen.add(normalized)
            highlights.append(cleaned)
        if len(highlights) >= 4:
            break
    if not highlights:
        for fragment in fragments:
            cleaned = re.sub(r"\s+", " ", fragment).strip(" -ㆍ·")
            if 20 <= len(cleaned) <= 220:
                highlights.append(cleaned)
            if len(highlights) >= 2:
                break
    return highlights


def summarize_dart_item(client: OpenDartReader.OpenDartReader, rcept_no: str, report_name: str) -> dict[str, Any]:
    category = classify_dart_report(report_name)
    link = f"https://dart.fss.or.kr/dsaf001/main.do?rcpNo={rcept_no}" if rcept_no else ""
    fallback = f"{category} 관련 공시입니다. 제목과 원문 링크를 함께 확인해 주세요."
    if not rcept_no:
        return {"category": category, "summary": fallback, "highlights": [], "link": link, "summary_source": "title"}
    try:
        document = client.document(rcept_no, cache=True)
        highlights = extract_dart_highlights(document, report_name)
        summary = " / ".join(highlights[:2]) if highlights else fallback
        return {
            "category": category,
            "summary": summary[:500],
            "highlights": highlights,
            "link": link,
            "summary_source": "document" if highlights else "title",
        }
    except Exception:
        return {"category": category, "summary": fallback, "highlights": [], "link": link, "summary_source": "title"}


def load_dart_summary(
    allowed_stock_names: list[str] | None = None,
    min_score: float = 50.0,
    file_date: str | None = None,
) -> dict[str, Any]:
    client = get_dart_client()
    if client is None:
        return {
            "enabled": False,
            "message": "DART API \ud0a4\uac00 \uc5c6\uc5b4 \uacf5\uc2dc \uc5f0\ub3d9\uc774 \ube44\ud65c\uc131\ud654\ub418\uc5b4 \uc788\uc2b5\ub2c8\ub2e4.",
            "items": [],
        }

    try:
        target_names = allowed_stock_names
        if target_names is None:
            try:
                target_names = [
                    str(item.get("stock_name", "")).strip()
                    for item in load_screening_summary(min_score=min_score, file_date=file_date).get("qualified_stocks", [])
                    if str(item.get("stock_name", "")).strip()
                ]
            except Exception:
                target_names = []
        target_keys = {normalize_text(name) for name in target_names if normalize_text(name)}

        today = datetime.today().strftime("%Y-%m-%d")
        recent = client.list(start=today, end=today)
        if recent is None or len(recent) == 0:
            return {
                "enabled": True,
                "message": "\uc624\ub298\uc790 \uacf5\uc2dc\uac00 \uc5c6\uac70\ub098 \uc870\ud68c\ub41c \ub0b4\uc5ed\uc774 \uc5c6\uc2b5\ub2c8\ub2e4.",
                "items": [],
            }

        recent = recent.fillna("")
        items = []
        for _, row in recent.iterrows():
            corp_name = str(row.get("corp_name", "")).strip()
            corp_key = normalize_text(corp_name)
            if target_keys and not any(key and (key in corp_key or corp_key in key) for key in target_keys):
                continue
            report_name = str(row.get("report_nm", "")).strip()
            rcept_no = str(row.get("rcept_no", "") or row.get("rcept_no", "")).strip()
            summary_payload = summarize_dart_item(client, rcept_no, report_name)
            items.append(
                {
                    "corp_name": corp_name,
                    "report_name": report_name,
                    "date": row.get("rcept_dt", ""),
                    "rcept_no": rcept_no,
                    "category": summary_payload.get("category", ""),
                    "summary": summary_payload.get("summary", ""),
                    "highlights": summary_payload.get("highlights", []),
                    "link": summary_payload.get("link", ""),
                    "summary_source": summary_payload.get("summary_source", ""),
                }
            )
        items = items[:50]
        target_message = f"시총 2000억 이상 종목 {len(target_keys)}개 기준" if target_keys else "시총 2000억 이상 종목 기준"
        return {
            "enabled": True,
            "message": f"오늘 공시 중 {target_message}으로 {len(items)}건을 불러왔습니다.",
            "items": items,
        }
    except Exception as exc:
        return {
            "enabled": True,
            "message": f"\uacf5\uc2dc \uc870\ud68c \uc911 \uc624\ub958\uac00 \ubc1c\uc0dd\ud588\uc2b5\ub2c8\ub2e4: {exc}",
            "items": [],
        }


async def telegram_status_payload() -> dict[str, Any]:
    telegram = get_telegram_settings()
    configured = bool(telegram["api_id"] and telegram["api_hash"] and telegram["phone"])
    if not configured:
        return {
            "configured": False,
            "authorized": False,
            "phone": telegram["phone"],
            "needs_code": False,
            "needs_password": False,
            "dialogs": [],
            "message": "Enter Telegram API ID, API Hash, and phone number to connect.",
        }

    client, temp_dir = build_telegram_readonly_client()
    try:
        await client.connect()
        authorized = await client.is_user_authorized()
        dialogs = []
        if authorized:
            async for dialog in client.iter_dialogs():
                if dialog_matches_earnings_channel(str(dialog.name or "")):
                    TELEGRAM_EARNINGS_DIALOG_CACHE["id"] = int(dialog.id)
                    TELEGRAM_EARNINGS_DIALOG_CACHE["name"] = str(dialog.name or "")
                    TELEGRAM_EARNINGS_DIALOG_CACHE["entity"] = getattr(dialog, "input_entity", None) or dialog.entity
                dialogs.append(
                    {
                        "id": int(dialog.id),
                        "name": dialog.name,
                        "is_channel": bool(dialog.is_channel),
                        "is_group": bool(dialog.is_group),
                        "unread_count": int(dialog.unread_count or 0),
                    }
                )
        return {
            "configured": True,
            "authorized": bool(authorized),
            "phone": telegram["phone"],
            "needs_code": False,
            "needs_password": False,
            "dialogs": dialogs,
            "message": "Telegram account connected." if authorized else "Telegram login verification is required.",
        }
    finally:
        await client.disconnect()
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)


def normalize_keyword_list(keywords: list[str]) -> list[str]:
    normalized: list[str] = []
    for keyword in keywords:
        for chunk in re.split(r"[\n,;]+", str(keyword or "")):
            parts = [part.strip() for part in re.split(r"\s+", chunk.strip()) if part.strip()]
            index = 0
            while index < len(parts):
                token = parts[index]
                next_token = parts[index + 1] if index + 1 < len(parts) else ""
                if (
                    normalize_search_text(token) in TELEGRAM_SEARCH_PREFIX_ALIASES
                    and next_token
                    and re.search(r"[가-힣]", next_token)
                ):
                    token = token + next_token
                    index += 2
                else:
                    index += 1
                if token and token not in normalized:
                    normalized.append(token)
    return normalized


def normalize_search_text(value: str) -> str:
    text = str(value or "").lower()
    return re.sub(r"[\s\-_·•.,:/\\()\[\]{}<>|]+", "", text)


TELEGRAM_SEARCH_PREFIX_ALIASES = {
    "lg": "엘지",
    "sk": "에스케이",
    "hd": "에이치디",
    "cj": "씨제이",
    "db": "디비",
    "kt": "케이티",
}


def unique_texts(values: list[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        text = str(value or "").strip()
        if not text:
            continue
        key = text.lower()
        if key and key not in seen:
            seen.add(key)
            result.append(text)
    return result


def telegram_keyword_match_needles(keyword: str, exact_phrase: bool) -> list[str]:
    base = normalize_search_text(keyword)
    needles = [base]
    for english, korean in TELEGRAM_SEARCH_PREFIX_ALIASES.items():
        if base.startswith(english) and len(base) > len(english):
            needles.append(normalize_search_text(korean + base[len(english) :]))
        if base.startswith(korean) and len(base) > len(korean):
            needles.append(normalize_search_text(english + base[len(korean) :]))
    return unique_texts(needles)


def telegram_search_query_variants(keyword: str) -> list[str]:
    raw = str(keyword or "").strip()
    normalized = normalize_search_text(raw)
    parts = re.findall(r"[A-Za-z]+|[가-힣]+|\d+", normalized)
    spaced = " ".join(parts)
    variants = [raw, normalized, normalized.upper(), spaced, spaced.upper()]
    has_latin = bool(re.search(r"[A-Za-z]", normalized))
    has_korean = bool(re.search(r"[가-힣]", normalized))
    if has_latin and has_korean:
        variants.extend([part for part in parts if re.search(r"[가-힣]", part)])
    for english, korean in TELEGRAM_SEARCH_PREFIX_ALIASES.items():
        if normalized.startswith(english) and len(normalized) > len(english):
            alias = korean + normalized[len(english) :]
            variants.extend([alias, " ".join(re.findall(r"[A-Za-z]+|[가-힣]+|\d+", alias))])
    return unique_texts(variants)


def telegram_search_queries_for_keywords(keywords: list[str], match_mode: str) -> list[str]:
    search_terms = keywords if match_mode != "all" else keywords[:1]
    variants: list[str] = []
    for keyword in search_terms:
        variants.extend(telegram_search_query_variants(keyword))
    return unique_texts(variants)


def message_has_file(message: Any) -> bool:
    if getattr(message, "file", None):
        return True
    if getattr(message, "media", None):
        return True
    return False


def get_message_file_name(message: Any) -> str:
    return str(getattr(getattr(message, "file", None), "name", None) or "").strip()


def get_message_search_blob(message: Any) -> str:
    text = str(getattr(message, "message", None) or "").strip()
    file_name = get_message_file_name(message)
    return " ".join(part for part in [text, file_name] if part).strip()


def extract_message_links(message: Any, text: str) -> list[str]:
    links: list[str] = []
    seen: set[str] = set()

    def add_link(value: str | None) -> None:
        if not value:
            return
        link = value.strip()
        if not link:
            return
        normalized = link if re.match(r"^[a-z]+://", link, flags=re.IGNORECASE) else f"https://{link}"
        if normalized not in seen:
            seen.add(normalized)
            links.append(normalized)

    for entity in getattr(message, "entities", None) or []:
        entity_name = entity.__class__.__name__
        if entity_name == "MessageEntityTextUrl":
            add_link(getattr(entity, "url", None))
            continue
        if entity_name == "MessageEntityUrl":
            offset = getattr(entity, "offset", 0)
            length = getattr(entity, "length", 0)
            add_link(text[offset : offset + length])

    for match in re.finditer(r"(?:(?:https?://)|(?:www\.))[^\s<>()]+", text, flags=re.IGNORECASE):
        add_link(match.group(0).rstrip(".,);!?]}>\"'"))

    return links


def build_attachment_metadata(chat_id: int, message_id: int, message: Any) -> tuple[str, str, str]:
    original_name = getattr(getattr(message, "file", None), "name", None)
    extension = getattr(getattr(message, "file", None), "ext", None) or ""
    if not original_name:
        original_name = f"attachment_{chat_id}_{message_id}{extension}"
    display_name = sanitize_filename(original_name)
    stored_name = f"{chat_id}_{message_id}_{display_name}"
    return display_name, stored_name, extension


def parse_search_date(value: str | None) -> date | None:
    if not value:
        return None
    return datetime.fromisoformat(value).date()


def build_search_result_payload(dialog: Any, message: Any, text: str, matched_keywords: list[str]) -> dict[str, Any]:
    file_flag = message_has_file(message)
    file_name = get_message_file_name(message)
    return {
        "chat_id": int(dialog.id),
        "chat_name": dialog.name,
        "message_id": int(message.id),
        "date": message.date.isoformat() if message.date else "",
        "sender_id": int(message.sender_id) if getattr(message, "sender_id", None) else None,
        "text": text,
        "has_file": file_flag,
        "file_name": file_name or None,
        "attachment_url": f"/api/telegram/attachment/{int(dialog.id)}/{int(message.id)}" if file_flag else None,
        "links": extract_message_links(message, text),
        "matched_keywords": matched_keywords,
        "link_hint": f"{dialog.name} / {message.id}",
    }


async def dialog_from_global_message(client: TelegramClient, message: Any) -> Any:
    chat = getattr(message, "chat", None)
    if chat is None:
        try:
            chat = await message.get_chat()
        except Exception:
            chat = None
    chat_id = getattr(message, "chat_id", None) or getattr(chat, "id", None) or 0
    first_name = str(getattr(chat, "first_name", "") or "").strip()
    last_name = str(getattr(chat, "last_name", "") or "").strip()
    chat_name = (
        str(getattr(chat, "title", "") or "").strip()
        or " ".join(part for part in [first_name, last_name] if part).strip()
        or str(getattr(chat, "username", "") or "").strip()
        or f"Chat {chat_id}"
    )
    return SimpleNamespace(id=int(chat_id), name=chat_name, entity=chat)


async def collect_global_search_results(
    client: TelegramClient,
    keywords: list[str],
    request: TelegramSearchRequest,
    result_limit: int,
    start_date: date | None,
    end_date: date | None,
    job: dict[str, Any] | None = None,
) -> dict[str, dict[str, Any]]:
    collected: dict[str, dict[str, Any]] = {}
    search_terms = telegram_search_queries_for_keywords(keywords, request.match_mode)
    checked_count = 0
    per_term_limit = max(300, min(3000, result_limit * 8))
    if job is not None:
        job["selected_chat_count"] = 0
        job["total_chat_count"] = len(search_terms)
        job["message"] = "텔레그램 전체 방 전역 검색 중..."
    for search_index, search_query in enumerate(search_terms):
        if not search_query:
            continue
        scanned_count = 0
        async for message in client.iter_messages(None, search=search_query, limit=per_term_limit):
            if job is not None and job_is_cancelled(job):
                job["status"] = "cancelled"
                job["message"] = "Search cancelled."
                return collected
            scanned_count += 1
            checked_count += 1
            text = get_message_search_blob(message)
            message_date = message.date.date() if getattr(message, "date", None) else None
            if end_date and message_date and message_date > end_date:
                continue
            if start_date and message_date and message_date < start_date:
                break
            has_match, matched_keywords = keyword_match(text, keywords, request.exact_phrase, request.match_mode)
            if not has_match:
                continue
            file_flag = message_has_file(message)
            if request.has_file is not None and file_flag != request.has_file:
                continue
            dialog = await dialog_from_global_message(client, message)
            key = f"{int(dialog.id)}:{int(message.id)}"
            existing = collected.get(key)
            merged_keywords = matched_keywords
            if existing:
                merged_keywords = sorted(set(existing["matched_keywords"] + matched_keywords))
            collected[key] = build_search_result_payload(dialog, message, text, merged_keywords)
            if job is not None:
                job["processed_chat_count"] = min(job.get("total_chat_count", 0), search_index + 1)
                job["results"] = sorted(collected.values(), key=lambda item: item["date"], reverse=True)[:result_limit]
                job["result_count"] = len(job["results"])
                job["message"] = (
                    f"전체 방 전역 검색 '{search_query}': 최근 결과 {scanned_count}개 확인, "
                    f"{len(job['results'])}개 발견"
                )
            if len(collected) >= result_limit:
                break
        if len(collected) >= result_limit:
            break
    if job is not None:
        job["processed_chat_count"] = job.get("total_chat_count", len(search_terms))
        job["selected_chat_count"] = len({item.get("chat_id") for item in collected.values()})
        job["message"] = f"전체 방 전역 검색 완료: {checked_count}개 후보 확인, {len(collected)}개 발견"
    return collected


def parse_water_notification_text(text: str, message_date: datetime | None = None) -> dict[str, Any] | None:
    source = str(text or "").replace("\u00a0", " ")
    normalized = normalize_search_text(source)
    if "중부수도사업소" not in source and "중부수도사업소" not in normalized:
        return None
    if "알림" not in source and "[중부수도사업소" not in source:
        return None

    def number_from_match(patterns: list[str], flags: int = re.IGNORECASE) -> float | None:
        for pattern in patterns:
            match = re.search(pattern, source, flags=flags)
            if match:
                try:
                    return float(match.group(1).replace(",", ""))
                except Exception:
                    continue
        return None

    usage = number_from_match(
        [
            r"(?:총\s*)?(?:사용량|검침량|수도사용량|상수도사용량|당월사용량)\s*[:：]?\s*([0-9,.]+)\s*(?:m3|㎥|톤|세제곱미터)",
            r"([0-9,.]+)\s*(?:m3|㎥|톤|세제곱미터)\s*(?:사용|사용량|검침)",
        ]
    )
    amount = number_from_match(
        [
            r"(?:총\s*)?(?:요금|수도요금|상하수도요금|청구금액|납부금액|고지금액|납입금액)\s*[:：]?\s*([0-9,]+)\s*원",
            r"([0-9,]+)\s*원\s*(?:입니다|납부|청구|고지)",
        ]
    )
    if amount is None:
        won_values = []
        for match in re.finditer(r"([0-9][0-9,]{2,})\s*원", source):
            try:
                won_values.append(float(match.group(1).replace(",", "")))
            except Exception:
                pass
        amount = max(won_values) if won_values else None

    month = ""
    month_patterns = [
        r"(\d{4})\s*[년./-]\s*(\d{1,2})\s*(?:월|[./-])?\s*(?:분|고지|청구|검침)?",
        r"(\d{4})(\d{2})\s*(?:분|고지|청구|검침)?",
    ]
    for pattern in month_patterns:
        match = re.search(pattern, source)
        if match:
            year = int(match.group(1))
            month_num = int(match.group(2))
            if 1 <= month_num <= 12:
                month = f"{year:04d}-{month_num:02d}"
                break
    if not month and message_date:
        kst = timezone(timedelta(hours=9))
        base_date = message_date if message_date.tzinfo else message_date.replace(tzinfo=timezone.utc)
        month = base_date.astimezone(kst).strftime("%Y-%m")

    if usage is None and amount is None:
        return None
    return {
        "month": month,
        "total_usage_m3": round(float(usage or 0), 4),
        "total_bill": round(float(amount or 0)),
        "text_excerpt": source[:700],
    }


def parse_electric_notification_text(text: str, message_date: datetime | None = None) -> dict[str, Any] | None:
    source = str(text or "").replace("\u00a0", " ")
    normalized = normalize_search_text(source)
    if not any(token in source or token in normalized for token in ("한국전력", "한전", "KEPCO", "kepco", "전기요금", "전기세")):
        return None

    compact_digits = re.sub(r"\D+", "", source)
    unit_id = ""
    customer_no = ""
    for customer_digits, candidate_unit_id in BUILDING_ELECTRIC_CUSTOMER_DIGITS.items():
        if customer_digits and customer_digits in compact_digits:
            unit_id = candidate_unit_id
            customer_no = BUILDING_ELECTRIC_CUSTOMERS[candidate_unit_id]["customer_no"]
            break
    if not unit_id:
        customer_match = re.search(r"(?:고객번호|전기사용계약번호)\s*[:：]?\s*([0-9\s-]{8,})", source)
        if customer_match:
            matched_digits = re.sub(r"\D+", "", customer_match.group(1))
            unit_id = BUILDING_ELECTRIC_CUSTOMER_DIGITS.get(matched_digits, "")
            if unit_id:
                customer_no = BUILDING_ELECTRIC_CUSTOMERS[unit_id]["customer_no"]
    if not unit_id:
        return None

    won_values: list[float] = []
    preferred_patterns = [
        r"(?:청구액|청구요금)\s*(?:은|는)?\s*([0-9,]+)\s*원",
        r"(?:청구금액|납부금액|납입금액|전기요금|당월요금|이번달요금|고지금액)\s*[:：]?\s*([0-9,]+)\s*원",
        r"([0-9][0-9,]{2,})\s*원\s*(?:입니다|납부|청구|고지)",
    ]
    for pattern in preferred_patterns:
        for match in re.finditer(pattern, source, flags=re.IGNORECASE):
            try:
                won_values.append(float(match.group(1).replace(",", "")))
            except Exception:
                pass
        if won_values:
            break
    if not won_values:
        for match in re.finditer(r"([0-9][0-9,]{2,})\s*원", source):
            try:
                value = float(match.group(1).replace(",", ""))
                if value >= 100:
                    won_values.append(value)
            except Exception:
                pass
    amount = max(won_values) if won_values else 0.0
    if amount <= 0:
        return None

    month = ""
    due_date = ""
    due_match = re.search(r"(?:납부기한|납기일|납부마감일)\s*[:：]?\s*(\d{4})\s*[년./-]\s*(\d{1,2})\s*[월./-]\s*(\d{1,2})\s*(?:일)?", source)
    if due_match:
        due_date = f"{int(due_match.group(1)):04d}-{int(due_match.group(2)):02d}-{int(due_match.group(3)):02d}"
    month_patterns = [
        r"청구년월\s*[:：]?\s*(\d{4})\s*년\s*(\d{1,2})\s*월",
        r"(\d{4})\s*[년./-]\s*(\d{1,2})\s*(?:월|[./-])?\s*(?:분|고지|청구|사용|납기)",
        r"\b(\d{4})(\d{2})\b\s*(?:분|고지|청구|사용|납기)",
    ]
    for pattern in month_patterns:
        match = re.search(pattern, source)
        if match:
            year = int(match.group(1))
            month_num = int(match.group(2))
            if 2000 <= year <= 2100 and 1 <= month_num <= 12:
                month = f"{year:04d}-{month_num:02d}"
                break
    if not month:
        usage_match = re.search(
            r"사용기간\s*[:：]?\s*(\d{2,4})\s*[년./-]\s*(\d{1,2})\s*[월./-]\s*(\d{1,2})\s*(?:일)?\s*~\s*(\d{2,4})\s*[년./-]\s*(\d{1,2})\s*[월./-]\s*(\d{1,2})",
            source,
        )
        if usage_match:
            year = int(usage_match.group(4))
            if year < 100:
                year += 2000
            month_num = int(usage_match.group(5))
            if 2000 <= year <= 2100 and 1 <= month_num <= 12:
                month = f"{year:04d}-{month_num:02d}"
    if not month:
        title_month_match = re.search(r"(?:제목\s*)?(\d{1,2})월\s*전기요금", source)
        if title_month_match:
            month_num = int(title_month_match.group(1))
            if 1 <= month_num <= 12:
                inferred_year = 0
                if message_date:
                    kst = timezone(timedelta(hours=9))
                    base_date = message_date if message_date.tzinfo else message_date.replace(tzinfo=timezone.utc)
                    local_date = base_date.astimezone(kst)
                    inferred_year = local_date.year - (1 if month_num > local_date.month + 1 else 0)
                if 2000 <= inferred_year <= 2100:
                    month = f"{inferred_year:04d}-{month_num:02d}"
    if not month and message_date:
        kst = timezone(timedelta(hours=9))
        base_date = message_date if message_date.tzinfo else message_date.replace(tzinfo=timezone.utc)
        month = base_date.astimezone(kst).strftime("%Y-%m")

    return {
        "month": month,
        "unit_id": unit_id,
        "label": BUILDING_ELECTRIC_CUSTOMERS.get(unit_id, {}).get("label", unit_id),
        "customer_no": customer_no,
        "amount": round(amount),
        "due_date": due_date,
        "text_excerpt": source[:700],
    }


async def find_telegram_dialog_by_name(client: TelegramClient, query: str) -> Any:
    normalized_query = normalize_search_text(query)
    fallback = None
    async for dialog in client.iter_dialogs():
        name = str(dialog.name or "")
        normalized_name = normalize_search_text(name)
        if normalized_name == normalized_query:
            return dialog
        if normalized_query and normalized_query in normalized_name and fallback is None:
            fallback = dialog
    if fallback is not None:
        return fallback
    raise ValueError(f"'{query}' 채팅방을 찾지 못했습니다.")


async def sync_real_estate_water_from_telegram(request: RealEstateWaterTelegramSyncRequest) -> dict[str, Any]:
    chat_name = str(request.chat_name or "엄마").strip() or "엄마"
    scan_limit = max(20, min(int(request.limit or 300), 2000))
    target_month = str(request.month or "").strip()[:7]
    client, temp_dir = build_telegram_readonly_client()
    try:
        await client.connect()
        if not await client.is_user_authorized():
            raise ValueError("Telegram login is required first.")
        dialog = await find_telegram_dialog_by_name(client, chat_name)
        candidates = []
        try:
            async for message in client.iter_messages(dialog.entity, search="중부수도사업소", limit=scan_limit):
                candidates.append(message)
        except Exception:
            candidates = []
        if not candidates:
            async for message in client.iter_messages(dialog.entity, limit=scan_limit):
                text = get_message_search_blob(message)
                if "중부수도사업소" in text or "[중부수도사업소" in text:
                    candidates.append(message)
        parsed_messages: list[dict[str, Any]] = []
        for message in candidates:
            text = get_message_search_blob(message)
            parsed = parse_water_notification_text(text, getattr(message, "date", None))
            if not parsed:
                continue
            if target_month and parsed.get("month") and parsed["month"] != target_month:
                continue
            month_key = str(parsed.get("month") or "").strip()[:7]
            if not month_key:
                continue
            parsed_messages.append({
                "month": month_key,
                "parsed": parsed,
                "message": message,
            })
        if not parsed_messages:
            raise ValueError(f"'{chat_name}' 방에서 중부수도사업소 알림 메시지를 찾지 못했습니다.")

        # Newer Telegram messages can include corrections. Keep the newest message per billing month.
        parsed_by_month: dict[str, dict[str, Any]] = {}
        for item in parsed_messages:
            month_key = item["month"]
            current = parsed_by_month.get(month_key)
            current_date = getattr(current.get("message"), "date", None) if current else None
            item_date = getattr(item.get("message"), "date", None)
            if current is None or (item_date and current_date and item_date > current_date):
                parsed_by_month[month_key] = item
            elif current is None:
                parsed_by_month[month_key] = item

        db = load_real_estate_db()
        water_billing = db.setdefault("water_billing", {})
        months = water_billing.get("months") if isinstance(water_billing.get("months"), list) else []
        unit_ids = [unit_id for floor in BUILDING_UNIT_LAYOUT for unit_id in floor["units"]]
        synced = []
        for month, item in sorted(parsed_by_month.items()):
            parsed_payload = item["parsed"]
            selected_message = item["message"]
            row = next((candidate for candidate in months if isinstance(candidate, dict) and str(candidate.get("month") or "")[:7] == month), None)
            if row is None:
                row = {
                    "month": month,
                    "total_usage_m3": 0,
                    "total_bill": 0,
                    "readings": {unit_id: 0 for unit_id in unit_ids},
                    "memo": "",
                }
                months.append(row)
            row["month"] = month
            if parsed_payload.get("total_usage_m3"):
                row["total_usage_m3"] = parsed_payload["total_usage_m3"]
            if parsed_payload.get("total_bill"):
                row["total_bill"] = parsed_payload["total_bill"]
            row["source"] = "telegram"
            row["source_chat"] = str(dialog.name or chat_name)
            row["source_message_id"] = int(getattr(selected_message, "id", 0) or 0) if selected_message else 0
            row["source_message_date"] = selected_message.date.isoformat() if selected_message and selected_message.date else ""
            row["source_excerpt"] = parsed_payload.get("text_excerpt", "")
            synced.append({
                "month": month,
                "total_usage_m3": row.get("total_usage_m3"),
                "total_bill": row.get("total_bill"),
                "message_id": row.get("source_message_id"),
                "message_date": row.get("source_message_date"),
            })
        water_billing["months"] = sorted(months, key=lambda item: str(item.get("month") or ""))
        db["water_billing"] = water_billing
        saved = save_real_estate_db(db)
        saved["summary"] = build_real_estate_summary(saved)
        latest = synced[-1] if synced else {}
        return {
            "ok": True,
            "month": latest.get("month", ""),
            "parsed": parsed_by_month.get(latest.get("month", ""), {}).get("parsed", {}) if latest else {},
            "message_id": latest.get("message_id", 0),
            "chat_name": str(dialog.name or chat_name),
            "synced_count": len(synced),
            "synced_months": synced,
            "real_estate": saved,
        }
    finally:
        await client.disconnect()
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)


async def sync_real_estate_electricity_from_telegram(request: RealEstateElectricTelegramSyncRequest) -> dict[str, Any]:
    chat_name = str(request.chat_name or "엄마").strip() or "엄마"
    scan_limit = max(20, min(int(request.limit or 500), 3000))
    target_month = str(request.month or "").strip()[:7]
    client, temp_dir = build_telegram_readonly_client()
    try:
        await client.connect()
        if not await client.is_user_authorized():
            raise ValueError("Telegram login is required first.")
        dialog = await find_telegram_dialog_by_name(client, chat_name)
        candidates = []
        search_terms = ["한국전력", "한전", "전기요금"]
        for term in search_terms:
            try:
                async for message in client.iter_messages(dialog.entity, search=term, limit=scan_limit):
                    if message not in candidates:
                        candidates.append(message)
            except Exception:
                pass
        if not candidates:
            async for message in client.iter_messages(dialog.entity, limit=scan_limit):
                text = get_message_search_blob(message)
                if any(term in text for term in ("한국전력", "한전", "전기요금", "전기세")):
                    candidates.append(message)

        parsed_messages: list[dict[str, Any]] = []
        for message in candidates:
            text = get_message_search_blob(message)
            parsed = parse_electric_notification_text(text, getattr(message, "date", None))
            if not parsed:
                continue
            if target_month and parsed.get("month") and parsed["month"] != target_month:
                continue
            month_key = str(parsed.get("month") or "").strip()[:7]
            if not month_key:
                continue
            parsed_messages.append({
                "month": month_key,
                "unit_id": parsed["unit_id"],
                "parsed": parsed,
                "message": message,
            })
        if not parsed_messages:
            raise ValueError(f"'{chat_name}' 방에서 한국전력 청구서 메시지를 찾지 못했습니다.")

        parsed_by_month_unit: dict[tuple[str, str], dict[str, Any]] = {}
        for item in parsed_messages:
            key = (item["month"], item["unit_id"])
            current = parsed_by_month_unit.get(key)
            current_date = getattr(current.get("message"), "date", None) if current else None
            item_date = getattr(item.get("message"), "date", None)
            if current is None or (item_date and current_date and item_date > current_date):
                parsed_by_month_unit[key] = item

        db = load_real_estate_db()
        electricity_billing = db.setdefault("electricity_billing", {})
        electricity_billing["customers"] = BUILDING_ELECTRIC_CUSTOMERS
        months = electricity_billing.get("months") if isinstance(electricity_billing.get("months"), list) else []
        synced = []
        for (month, unit_id), item in sorted(parsed_by_month_unit.items()):
            parsed_payload = item["parsed"]
            selected_message = item["message"]
            row = next((candidate for candidate in months if isinstance(candidate, dict) and str(candidate.get("month") or "")[:7] == month), None)
            if row is None:
                row = {"month": month, "total_bill": 0, "bills": {}, "memo": ""}
                months.append(row)
            row["month"] = month
            row["bills"] = row.get("bills") if isinstance(row.get("bills"), dict) else {}
            row["due_dates"] = row.get("due_dates") if isinstance(row.get("due_dates"), dict) else {}
            row["bills"][unit_id] = parsed_payload["amount"]
            if parsed_payload.get("due_date"):
                row["due_dates"][unit_id] = parsed_payload["due_date"]
                existing_due_dates = [str(value) for value in row["due_dates"].values() if str(value or "").strip()]
                if existing_due_dates:
                    row["due_date"] = sorted(existing_due_dates)[-1]
            row["total_bill"] = round(sum(real_estate_number(value) for value in row["bills"].values()))
            row["source"] = "telegram"
            row["source_chat"] = str(dialog.name or chat_name)
            row["source_message_id"] = int(getattr(selected_message, "id", 0) or 0) if selected_message else 0
            row["source_message_date"] = selected_message.date.isoformat() if selected_message and selected_message.date else ""
            row["source_excerpt"] = parsed_payload.get("text_excerpt", "")
            synced.append({
                "month": month,
                "unit_id": unit_id,
                "label": parsed_payload.get("label"),
                "customer_no": parsed_payload.get("customer_no"),
                "amount": parsed_payload.get("amount"),
                "due_date": parsed_payload.get("due_date"),
                "message_id": row.get("source_message_id"),
                "message_date": row.get("source_message_date"),
            })
        electricity_billing["months"] = sorted(months, key=lambda item: str(item.get("month") or ""))
        db["electricity_billing"] = electricity_billing
        saved = save_real_estate_db(db)
        saved["summary"] = build_real_estate_summary(saved)
        latest = synced[-1] if synced else {}
        return {
            "ok": True,
            "month": latest.get("month", ""),
            "chat_name": str(dialog.name or chat_name),
            "synced_count": len(synced),
            "synced_items": synced,
            "real_estate": saved,
        }
    finally:
        await client.disconnect()
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)


def is_earnings_disclosure_text(text: str) -> bool:
    source = str(text or "")
    has_identity = "기업명:" in source and "보고서명:" in source
    if not has_identity:
        return False
    report_name = extract_earnings_field(source, "보고서명")
    normalized_report = normalize_search_text(report_name)
    if any(normalize_search_text(token) in normalized_report for token in TELEGRAM_EARNINGS_REPORT_EXCLUDE_TOKENS):
        return False
    report_looks_like_earnings = any(
        normalize_search_text(token) in normalized_report
        for token in TELEGRAM_EARNINGS_REPORT_INCLUDE_TOKENS
    )
    has_sales_line = bool(extract_earnings_field(source, "매출액"))
    has_profit_line = bool(
        extract_earnings_field(source, "영업익")
        or extract_earnings_field(source, "영업이익")
        or extract_earnings_field(source, "순이익")
        or extract_earnings_field(source, "순익")
    )
    has_trend = "최근 실적 추이" in source
    has_estimate = "예상치" in source
    return report_looks_like_earnings and has_sales_line and has_profit_line and (has_trend or has_estimate)


def normalize_disclosure_category(value: str | None) -> str:
    category = str(value or "earnings").strip()
    return category if category in TELEGRAM_DISCLOSURE_CATEGORIES else "earnings"


def disclosure_report_name(text: str) -> str:
    return extract_earnings_field(str(text or ""), "보고서명")


def is_orders_disclosure_text(text: str) -> bool:
    source = str(text or "")
    report = disclosure_report_name(source)
    normalized_report = normalize_search_text(report)
    normalized_source = normalize_search_text(source)
    exclude_tokens = ["대량보유", "주식등의대량보유", "보유상황", "소유상황", "주요계약체결", "장내매매", "주요주주"]
    if any(normalize_search_text(token) in normalized_report or normalize_search_text(token) in normalized_source for token in exclude_tokens):
        return False
    include_report_tokens = ["단일판매공급계약", "판매공급계약", "공급계약", "수주", "계약수주"]
    include_body_tokens = ["계약금액", "계약상대", "계약기간", "판매ㆍ공급", "판매공급", "공급계약", "수주"]
    return any(normalize_search_text(token) in normalized_report for token in include_report_tokens) or any(
        normalize_search_text(token) in normalized_source for token in include_body_tokens
    )


def is_warning_disclosure_text(text: str) -> bool:
    source = str(text or "")
    report = disclosure_report_name(source)
    normalized = normalize_search_text(report + "\n" + source)
    include_tokens = ["투자경고", "투자주의", "투자위험", "단기과열", "매매거래정지", "관리종목", "불성실공시"]
    exclude_tokens = ["회사분할", "분할결정", "합병", "대량보유", "주식등의대량보유"]
    return any(normalize_search_text(token) in normalized for token in include_tokens) and not any(
        normalize_search_text(token) in normalized for token in exclude_tokens
    )


def is_investment_disclosure_text(text: str) -> bool:
    source = str(text or "")
    report = disclosure_report_name(source)
    normalized = normalize_search_text(report + "\n" + source)
    include_tokens = TELEGRAM_DISCLOSURE_CATEGORIES["investment"]["tokens"]
    return any(normalize_search_text(token) in normalized for token in include_tokens)


def is_ownership_disclosure_text(text: str) -> bool:
    source = str(text or "")
    report = disclosure_report_name(source)
    normalized = normalize_search_text(report + "\n" + source)
    include_tokens = TELEGRAM_DISCLOSURE_CATEGORIES["ownership"]["tokens"]
    return any(normalize_search_text(token) in normalized for token in include_tokens)


def is_disclosure_category_text(text: str, category: str) -> bool:
    source = str(text or "")
    if category == "earnings":
        return is_earnings_disclosure_text(source)
    if "기업명:" not in source or "보고서명:" not in source:
        return False
    if category == "all":
        return True
    if category == "orders":
        return is_orders_disclosure_text(source)
    if category == "warning":
        return is_warning_disclosure_text(source)
    if category == "investment":
        return is_investment_disclosure_text(source)
    if category == "ownership":
        return is_ownership_disclosure_text(source)
    tokens = TELEGRAM_DISCLOSURE_CATEGORIES.get(category, {}).get("tokens", [])
    return any(token in source for token in tokens)


def extract_earnings_field(text: str, label: str) -> str:
    match = re.search(rf"^{re.escape(label)}\s*:\s*(.+)$", text, flags=re.MULTILINE)
    return match.group(1).strip() if match else ""


def parse_earnings_line_parts(value: str) -> dict[str, str]:
    text = str(value or "").strip()
    actual = re.sub(r"\s*\([^)]*\)\s*", "", text).strip()
    expected = ""
    surprise = ""
    opm = ""
    expected_match = re.search(r"예상치\s*:\s*([^)]+)", text)
    if expected_match:
        expected_text = expected_match.group(1).strip()
        surprise_match = re.search(r"[-+]?\d+(?:\.\d+)?\s*%", expected_text)
        if surprise_match:
            surprise = surprise_match.group(0).replace(" ", "")
            expected = expected_text[: surprise_match.start()].rstrip(" ,/+")
        else:
            expected = expected_text.rstrip(" ,/+")
    opm_match = re.search(r"\(opm\s+([^)]+)\)", text, flags=re.IGNORECASE)
    if opm_match:
        opm = opm_match.group(1).strip()
    return {
        "actual": actual,
        "expected": expected,
        "surprise": surprise,
        "opm": opm,
    }


def parse_earnings_amount(value: Any) -> float | None:
    text = str(value or "")
    match = re.search(r"[-+]?\d[\d,]*(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return float(match.group(0).replace(",", ""))
    except ValueError:
        return None


def format_delta_percent(current: float | None, base: float | None) -> str:
    if current is None or base is None or base == 0:
        return ""
    delta = (current - base) / abs(base) * 100
    if not math.isfinite(delta):
        return ""
    return f"{delta:+.1f}%"


def parse_earnings_trend_rows(text: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    pattern = re.compile(
        r"^\s*(?P<year>\d{4})\.(?P<quarter>[1-4])Q\s+"
        r"(?P<sales>[-+]?\d[\d,]*(?:\.\d+)?[^\s/]*)\s*/\s*"
        r"(?P<operating>[-+]?\d[\d,]*(?:\.\d+)?[^\s/]*)\s*/\s*"
        r"(?P<net>-|[-+]?\d[\d,]*(?:\.\d+)?[^\s/]*)",
        flags=re.IGNORECASE,
    )
    for line in str(text or "").splitlines():
        match = pattern.match(line)
        if not match:
            continue
        rows.append(
            {
                "label": f"{match.group('year')}.{match.group('quarter')}Q",
                "year": int(match.group("year")),
                "quarter": int(match.group("quarter")),
                "sales": parse_earnings_amount(match.group("sales")),
                "operating_profit": parse_earnings_amount(match.group("operating")),
                "net_income": parse_earnings_amount(match.group("net")),
            }
        )
    return rows


def build_earnings_growth_metrics(text: str) -> dict[str, dict[str, str]]:
    trend_rows = parse_earnings_trend_rows(text)
    if not trend_rows:
        return {"sales": {}, "operating_profit": {}, "net_income": {}}
    current = trend_rows[0]
    previous = trend_rows[1] if len(trend_rows) > 1 else None
    prior_year = next(
        (
            row
            for row in trend_rows[1:]
            if row.get("year") == current.get("year") - 1 and row.get("quarter") == current.get("quarter")
        ),
        None,
    )
    metrics: dict[str, dict[str, str]] = {}
    for key in ["sales", "operating_profit", "net_income"]:
        metrics[key] = {
            "qoq": format_delta_percent(current.get(key), previous.get(key) if previous else None),
            "yoy": format_delta_percent(current.get(key), prior_year.get(key) if prior_year else None),
        }
    return metrics


def format_opm(revenue: float | None, operating_profit: float | None) -> str:
    if revenue is None or operating_profit is None or revenue == 0:
        return ""
    opm = operating_profit / revenue * 100
    if not math.isfinite(opm):
        return ""
    rounded = round(opm, 1)
    if abs(rounded - round(rounded)) < 0.05:
        return f"(opm {int(round(rounded))}%)"
    return f"(opm {rounded:.1f}%)"


def append_opm_to_telegram_earnings_text(text: str) -> str:
    source = str(text or "")
    sales_value = parse_earnings_amount(extract_earnings_field(source, "매출액"))
    operating_value = parse_earnings_amount(
        extract_earnings_field(source, "영업익") or extract_earnings_field(source, "영업이익")
    )
    top_opm = format_opm(sales_value, operating_value)
    lines = source.splitlines()
    enriched_lines: list[str] = []
    trend_line_pattern = re.compile(
        r"^(\s*\d{4}\.\dQ\s+)([-+]?\d[\d,]*(?:\.\d+)?\s*억?)\s*/\s*"
        r"([-+]?\d[\d,]*(?:\.\d+)?\s*억?)\s*/\s*"
        r"([-+]?\d[\d,]*(?:\.\d+)?\s*억?)(.*)$"
    )
    for line in lines:
        stripped = line.strip()
        if top_opm and re.match(r"^영업(?:익|이익)\s*:", stripped) and "opm" not in stripped.lower():
            enriched_lines.append(line.rstrip() + " " + top_opm)
            continue
        trend_match = trend_line_pattern.match(line)
        if trend_match and "opm" not in line.lower():
            revenue = parse_earnings_amount(trend_match.group(2))
            operating_profit = parse_earnings_amount(trend_match.group(3))
            opm_text = format_opm(revenue, operating_profit)
            if opm_text:
                enriched_lines.append(line.rstrip() + " " + opm_text)
                continue
        enriched_lines.append(line)
    return "\n".join(enriched_lines)


def parse_telegram_earnings_payload(payload: dict[str, Any]) -> dict[str, Any]:
    text = append_opm_to_telegram_earnings_text(str(payload.get("text") or ""))
    company_line_match = re.search(r"^기업명\s*:\s*(.+)$", text, flags=re.MULTILINE)
    company_line = company_line_match.group(1).strip() if company_line_match else ""
    company_match = re.search(r"^기업명\s*:\s*(.+?)(?:\(|\sA\d{6}|\n|$)", text, flags=re.MULTILINE)
    company_code_match = re.search(r"\bA(\d{6})\b", company_line, flags=re.IGNORECASE)
    explicit_company_code = company_code_match.group(1).strip() if company_code_match else ""
    report_name = extract_earnings_field(text, "보고서명")
    disclosure_match = re.search(r"공시링크\s*:\s*(https?://\S+)", text)
    company_info_match = re.search(r"회사정보\s*:\s*(https?://\S+)", text)
    company_info_link = company_info_match.group(1).strip() if company_info_match else ""
    company_info_code = ""
    if company_info_link:
        try:
            raw_company_info_code = str(parse_qs(urlparse(company_info_link).query).get("code", [""])[0]).strip()
            company_info_code = raw_company_info_code.zfill(6) if re.fullmatch(r"\d{1,6}", raw_company_info_code) else ""
        except Exception:
            company_info_code = ""
    sales = extract_earnings_field(text, "매출액")
    operating_profit = extract_earnings_field(text, "영업익") or extract_earnings_field(text, "영업이익")
    net_income = extract_earnings_field(text, "순이익") or extract_earnings_field(text, "순익")
    growth_metrics = build_earnings_growth_metrics(text)
    parsed = {
        "text": text,
        "company": company_match.group(1).strip() if company_match else "",
        "company_code": explicit_company_code or company_info_code,
        "company_info_code": company_info_code,
        "company_code_conflict": bool(explicit_company_code and company_info_code and explicit_company_code != company_info_code),
        "report_name": report_name,
        "sales": sales,
        "operating_profit": operating_profit,
        "net_income": net_income,
        "metrics": {
            "sales": {**parse_earnings_line_parts(sales), **growth_metrics.get("sales", {})},
            "operating_profit": {
                **parse_earnings_line_parts(operating_profit),
                **growth_metrics.get("operating_profit", {}),
            },
            "net_income": {**parse_earnings_line_parts(net_income), **growth_metrics.get("net_income", {})},
        },
        "disclosure_link": disclosure_match.group(1).strip() if disclosure_match else "",
        "company_info_link": company_info_link,
    }
    return {**payload, **parsed}


def screening_market_cap_universe(min_market_cap_100m: float = SCREENING_MIN_MARKET_CAP_100M) -> dict[str, dict[str, Any]]:
    listing = get_listing_table().copy()
    listing["Marcap"] = pd.to_numeric(listing.get("Marcap"), errors="coerce")
    listing = listing.dropna(subset=["Marcap"])
    listing = listing[listing["Marcap"] >= float(min_market_cap_100m) * 100000000.0]
    result: dict[str, dict[str, Any]] = {}
    for _, row in listing.iterrows():
        name = str(row.get("Name") or "").strip()
        code = str(row.get("Code") or "").strip().zfill(6)
        key = normalize_search_text(name)
        if not key:
            continue
        item = {
            "stock_name": name,
            "stock_code": code,
            "market": str(row.get("Market") or ""),
            "market_cap_100m": round(float(row.get("Marcap")) / 100000000.0, 1),
        }
        result[key] = item
        if code:
            result[normalize_search_text(code)] = item
    return result


def dialog_matches_earnings_channel(dialog_name: str) -> bool:
    normalized_name = normalize_search_text(dialog_name)
    normalized_target = normalize_search_text(TELEGRAM_EARNINGS_CHANNEL_NAME)
    if normalized_name == normalized_target:
        return True
    if normalized_target in normalized_name or normalized_name in normalized_target:
        return True
    return "awake" in normalized_name and "실시간주식공시정리채널" in normalized_name


async def find_earnings_dialog(client: TelegramClient) -> Any:
    if TELEGRAM_EARNINGS_CHANNEL_USERNAME:
        return SimpleNamespace(
            id=TELEGRAM_EARNINGS_CHANNEL_ID,
            name=TELEGRAM_EARNINGS_CHANNEL_NAME,
            entity=TELEGRAM_EARNINGS_CHANNEL_USERNAME,
        )
    cached_entity = TELEGRAM_EARNINGS_DIALOG_CACHE.get("entity")
    if cached_entity is not None:
        return SimpleNamespace(
            id=int(TELEGRAM_EARNINGS_DIALOG_CACHE.get("id") or getattr(cached_entity, "id", 0) or 0),
            name=str(TELEGRAM_EARNINGS_DIALOG_CACHE.get("name") or TELEGRAM_EARNINGS_CHANNEL_NAME),
            entity=cached_entity,
        )
    cached_id = TELEGRAM_EARNINGS_DIALOG_CACHE.get("id")
    if cached_id:
        try:
            numeric_id = int(cached_id)
            channel_id = int(str(abs(numeric_id))[3:]) if str(abs(numeric_id)).startswith("100") else abs(numeric_id)
            entity = await client.get_entity(PeerChannel(channel_id))
            return SimpleNamespace(
                id=numeric_id,
                name=str(TELEGRAM_EARNINGS_DIALOG_CACHE.get("name") or TELEGRAM_EARNINGS_CHANNEL_NAME),
                entity=entity,
            )
        except Exception:
            TELEGRAM_EARNINGS_DIALOG_CACHE.clear()
    fallback = None
    async for dialog in client.iter_dialogs():
        name = str(dialog.name or "")
        if dialog_matches_earnings_channel(name):
            TELEGRAM_EARNINGS_DIALOG_CACHE["id"] = int(dialog.id)
            TELEGRAM_EARNINGS_DIALOG_CACHE["name"] = name
            TELEGRAM_EARNINGS_DIALOG_CACHE["entity"] = getattr(dialog, "input_entity", None) or dialog.entity
            return dialog
        normalized = normalize_search_text(name)
        if fallback is None and "awake" in normalized and "공시" in normalized:
            fallback = dialog
    if fallback is not None:
        TELEGRAM_EARNINGS_DIALOG_CACHE["id"] = int(fallback.id)
        TELEGRAM_EARNINGS_DIALOG_CACHE["name"] = str(fallback.name or TELEGRAM_EARNINGS_CHANNEL_NAME)
        TELEGRAM_EARNINGS_DIALOG_CACHE["entity"] = getattr(fallback, "input_entity", None) or fallback.entity
        return fallback
    raise ValueError(f"'{TELEGRAM_EARNINGS_CHANNEL_NAME}' 방을 찾지 못했습니다. 텔레그램에서 해당 채널에 들어가 있는지 확인해 주세요.")


def build_disclosure_company_target(company: str) -> dict[str, Any]:
    raw = str(company or "").strip()
    normalized_raw = normalize_search_text(raw)
    target_code = ""
    target_name = ""
    if re.fullmatch(r"A?\d{6}", raw, flags=re.IGNORECASE):
        normalized_code = re.sub(r"^\D", "", raw).zfill(6)
        row = (get_screening_stock_lookup().get("by_code") or {}).get(normalized_code)
        if row is None:
            try:
                row = find_listing_row_by_code(normalized_code)
            except Exception:
                row = None
        if row:
            target_code = str(row.get("code") or normalized_code).zfill(6)
            target_name = str(row.get("name") or "").strip()
    if not target_name:
        resolved_code, resolved_name = resolve_stock_from_screening_cache(raw)
        if not resolved_code:
            try:
                resolved_code, resolved_name = resolve_stock(raw)
            except Exception:
                resolved_code, resolved_name = None, raw
        if resolved_name and normalize_search_text(resolved_name) != normalized_raw:
            target_name = resolved_name
            target_code = str(resolved_code or "").zfill(6) if resolved_code else ""
        elif resolved_code:
            target_name = resolved_name or raw
            target_code = str(resolved_code).zfill(6)
    normalized_names = {normalized_raw}
    if target_name:
        normalized_names.add(normalize_search_text(target_name))
    normalized_names = {name for name in normalized_names if name}
    query_terms = []
    for value in [target_code, f"A{target_code}" if target_code else "", target_name, raw]:
        cleaned = str(value or "").strip()
        if cleaned and cleaned not in query_terms:
            query_terms.append(cleaned)
    return {
        "raw": raw,
        "code": target_code,
        "name": target_name or raw,
        "normalized_names": normalized_names,
        "query_terms": query_terms or [raw],
    }


def disclosure_result_matches_company(parsed: dict[str, Any], target: dict[str, Any]) -> bool:
    if parsed.get("company_code_conflict"):
        return False
    target_code = str(target.get("code") or "").zfill(6) if target.get("code") else ""
    parsed_code = str(parsed.get("company_code") or "").zfill(6) if parsed.get("company_code") else ""
    parsed_info_code = str(parsed.get("company_info_code") or "").zfill(6) if parsed.get("company_info_code") else ""
    if target_code and parsed_info_code and parsed_info_code != target_code:
        return False
    if target_code and parsed_code and parsed_code != target_code:
        return False
    if target_code and parsed_code == target_code:
        return True
    if target_code and parsed_info_code == target_code:
        return True
    parsed_company = normalize_search_text(parsed.get("company") or "")
    target_names = target.get("normalized_names") or set()
    if parsed_company and parsed_company in target_names:
        return True
    if not target_code:
        raw = normalize_search_text(target.get("raw") or "")
        return bool(raw and parsed_company and (raw in parsed_company or parsed_company in raw))
    return False


def match_market_earnings_stock(parsed: dict[str, Any], universe: dict[str, dict[str, Any]]) -> dict[str, Any] | None:
    if parsed.get("company_code_conflict"):
        return None
    parsed_info_code = normalize_search_text(parsed.get("company_info_code") or "")
    parsed_code = normalize_search_text(parsed.get("company_code") or "")
    for code in [parsed_info_code, parsed_code]:
        if code:
            stock_info = universe.get(code)
            return stock_info if stock_info else None
    parsed_company_key = normalize_search_text(parsed.get("company") or "")
    if not parsed_company_key:
        return None
    stock_info = universe.get(parsed_company_key)
    if not stock_info:
        return None
    # Without an explicit stock code, only exact normalized company-name matches are trusted.
    if normalize_search_text(stock_info.get("stock_name") or "") != parsed_company_key:
        return None
    return stock_info


def parse_dart_amount_to_100m(value: Any) -> float | None:
    text = str(value or "").replace(",", "").strip()
    if not text or text in {"-", "nan", "None"}:
        return None
    match = re.search(r"[-+]?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return round(float(match.group(0)) / 100000000.0, 2)
    except ValueError:
        return None


def select_dart_financial_statement_rows(frame: pd.DataFrame, statement: str = "IS") -> pd.DataFrame:
    if frame is None or frame.empty:
        return pd.DataFrame()
    data = frame.copy()
    if "sj_div" in data.columns:
        data = data[data["sj_div"].astype(str).str.upper() == statement.upper()]
    if "fs_div" in data.columns and not data.empty:
        cfs = data[data["fs_div"].astype(str).str.upper() == "CFS"]
        if not cfs.empty:
            data = cfs
    return data


def dart_metric_value(data: pd.DataFrame, metric: str, column: str = "thstrm_amount") -> float | None:
    if data is None or data.empty or column not in data.columns:
        return None
    account = data.get("account_nm", pd.Series(dtype=str)).astype(str)
    if metric == "sales":
        candidates = data[account.str.fullmatch("매출액|영업수익|수익", na=False)]
        if candidates.empty:
            candidates = data[account.str.contains("매출액|영업수익", na=False)]
    elif metric == "operating_profit":
        candidates = data[account.str.contains("영업이익", na=False)]
    elif metric == "net_income":
        candidates = data[account.str.contains("당기순이익", na=False) & ~account.str.contains("법인세|지배|비지배", na=False)]
        if candidates.empty:
            candidates = data[account.str.contains("분기순이익|반기순이익|순이익", na=False) & ~account.str.contains("법인세|지배|비지배", na=False)]
    else:
        candidates = data[account.str.fullmatch("자본총계|자본총액|총자본", na=False)]
        if candidates.empty:
            candidates = data[account.str.contains("자본총계|자본총액|총자본", na=False)]
    if candidates.empty:
        return None
    for _, row in candidates.iterrows():
        value = parse_dart_amount_to_100m(row.get(column))
        if value is not None:
            return value
    return None


def dart_financial_row(client: OpenDartReader.OpenDartReader, stock_code: str, year: int, reprt_code: str) -> dict[str, Any] | None:
    try:
        frame = client.finstate(stock_code, year, reprt_code=reprt_code)
    except Exception:
        return None
    income_data = select_dart_financial_statement_rows(frame, "IS")
    balance_data = select_dart_financial_statement_rows(frame, "BS")
    if income_data.empty and balance_data.empty:
        return None
    row = {
        "sales": dart_metric_value(income_data, "sales"),
        "operating_profit": dart_metric_value(income_data, "operating_profit"),
        "net_income": dart_metric_value(income_data, "net_income"),
        "equity": dart_metric_value(balance_data, "equity"),
    }
    if all(row.get(key) is None for key in ["sales", "operating_profit", "net_income"]):
        return None
    return row


def build_valuation_band_rows(stock_code: str, annual_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    listing_row = (get_screening_stock_lookup().get("by_code") or {}).get(str(stock_code or "").zfill(6)) or {}
    if not listing_row:
        try:
            listing_row = find_listing_row_by_code(stock_code) or {}
        except Exception:
            listing_row = {}
    shares = to_float(listing_row.get("stocks"))
    if not shares or shares <= 0 or not annual_rows:
        return []
    start_year = min(int(row.get("year") or date.today().year) for row in annual_rows)
    end_year = max(int(row.get("year") or date.today().year) for row in annual_rows)
    try:
        prices = fdr.DataReader(stock_code, f"{start_year}-01-01", f"{end_year + 1}-01-10")
    except Exception:
        prices = pd.DataFrame()
    if prices is None or prices.empty:
        return []
    frame = prices.copy().reset_index()
    if "Date" not in frame.columns:
        frame = frame.rename(columns={frame.columns[0]: "Date"})
    frame["Date"] = pd.to_datetime(frame["Date"]).dt.date
    frame["Close"] = pd.to_numeric(frame.get("Close"), errors="coerce")
    rows = []
    for annual in annual_rows:
        year = int(annual.get("year") or 0)
        if year <= 0:
            continue
        year_frame = frame[(frame["Date"] >= date(year, 1, 1)) & (frame["Date"] <= date(year, 12, 31))].dropna(subset=["Close"])
        if year_frame.empty:
            continue
        close = to_float(year_frame.iloc[-1].get("Close"))
        net_income_100m = to_float(annual.get("net_income"))
        equity_100m = to_float(annual.get("equity"))
        market_cap_100m = (close * shares / 100000000.0) if close is not None else None
        per = market_cap_100m / net_income_100m if market_cap_100m and net_income_100m and net_income_100m > 0 else None
        pbr = market_cap_100m / equity_100m if market_cap_100m and equity_100m and equity_100m > 0 else None
        rows.append(
            {
                "year": year,
                "label": str(year),
                "close": round(close, 2) if close is not None else None,
                "market_cap_100m": round(market_cap_100m, 1) if market_cap_100m else None,
                "per": round(per, 2) if per is not None and math.isfinite(per) else None,
                "pbr": round(pbr, 2) if pbr is not None and math.isfinite(pbr) else None,
            }
        )
    return rows


def latest_annual_metric_by_date(annual_rows: list[dict[str, Any]], row_date: date, metric: str) -> float | None:
    available = [
        row for row in annual_rows
        if int(row.get("year") or 0) <= row_date.year and to_float(row.get(metric)) is not None
    ]
    if not available:
        return None
    return to_float(sorted(available, key=lambda row: int(row.get("year") or 0))[-1].get(metric))


def interpolated_annual_metric_by_date(annual_rows: list[dict[str, Any]], row_date: date, metric: str) -> float | None:
    points = []
    for row in annual_rows:
        year = int(row.get("year") or 0)
        value = to_float(row.get(metric))
        if year > 0 and value is not None:
            points.append((date(year, 12, 31), value))
    points = sorted(points, key=lambda item: item[0])
    if not points:
        return None
    if row_date <= points[0][0]:
        return points[0][1]
    if row_date >= points[-1][0]:
        if len(points) < 2:
            return points[-1][1]
        previous_date, previous_value = points[-2]
        last_date, last_value = points[-1]
        total_days = max(1, (last_date - previous_date).days)
        daily_slope = (last_value - previous_value) / total_days
        extrapolated = last_value + daily_slope * max(0, (row_date - last_date).days)
        if metric in {"net_income", "equity"} and extrapolated <= 0:
            return last_value
        return extrapolated
    for index in range(1, len(points)):
        previous_date, previous_value = points[index - 1]
        next_date, next_value = points[index]
        if previous_date <= row_date <= next_date:
            total_days = max(1, (next_date - previous_date).days)
            elapsed_days = max(0, (row_date - previous_date).days)
            weight = elapsed_days / total_days
            return previous_value + (next_value - previous_value) * weight
    return points[-1][1]


def build_price_band_series(stock_code: str, annual_rows: list[dict[str, Any]]) -> dict[str, Any]:
    listing_row = (get_screening_stock_lookup().get("by_code") or {}).get(str(stock_code or "").zfill(6)) or {}
    if not listing_row:
        try:
            listing_row = find_listing_row_by_code(stock_code) or {}
        except Exception:
            listing_row = {}
    shares = to_float(listing_row.get("stocks"))
    if not shares or shares <= 0 or not annual_rows:
        return {"rows": [], "per_multiples": [], "pbr_multiples": []}
    today = date.today()
    start_date = today - timedelta(days=365 * 4)
    try:
        prices = fdr.DataReader(stock_code, start_date.isoformat(), today.isoformat())
    except Exception:
        prices = pd.DataFrame()
    if prices is None or prices.empty:
        return {"rows": [], "per_multiples": [], "pbr_multiples": []}
    frame = prices.copy().reset_index()
    if "Date" not in frame.columns:
        frame = frame.rename(columns={frame.columns[0]: "Date"})
    frame["Date"] = pd.to_datetime(frame["Date"]).dt.date
    frame["Close"] = pd.to_numeric(frame.get("Close"), errors="coerce")
    rows = []
    per_values: list[float] = []
    pbr_values: list[float] = []
    MAX_REASONABLE_PER = 200.0
    MAX_REASONABLE_PBR = 30.0
    for _, price_row in frame.dropna(subset=["Close"]).iterrows():
        row_date = price_row["Date"]
        close = to_float(price_row.get("Close"))
        if close is None:
            continue
        net_income_100m = interpolated_annual_metric_by_date(annual_rows, row_date, "net_income")
        equity_100m = interpolated_annual_metric_by_date(annual_rows, row_date, "equity")
        eps = net_income_100m * 100000000.0 / shares if net_income_100m and net_income_100m > 0 else None
        bps = equity_100m * 100000000.0 / shares if equity_100m and equity_100m > 0 else None
        per = close / eps if eps and eps > 0 else None
        pbr = close / bps if bps and bps > 0 else None
        if per is not None and math.isfinite(per) and 0 < per <= MAX_REASONABLE_PER:
            per_values.append(per)
        if pbr is not None and math.isfinite(pbr) and 0 < pbr <= MAX_REASONABLE_PBR:
            pbr_values.append(pbr)
        rows.append(
            {
                "date": row_date.isoformat(),
                "close": round(close, 2),
                "eps": round(eps, 2) if eps is not None else None,
                "bps": round(bps, 2) if bps is not None else None,
                "per": round(per, 2) if per is not None and math.isfinite(per) else None,
                "pbr": round(pbr, 2) if pbr is not None and math.isfinite(pbr) else None,
            }
        )

    def multiples(values: list[float], fallback: list[float], max_value: float) -> list[float]:
        clean = sorted(
            value for value in values
            if value is not None and math.isfinite(value) and 0 < value <= max_value
        )
        if len(clean) < 3:
            return fallback
        positions = [0.10, 0.30, 0.50, 0.70, 0.90]
        result = []
        for position in positions:
            index = min(len(clean) - 1, max(0, int(round((len(clean) - 1) * position))))
            result.append(round(clean[index], 2))
        deduped = []
        for value in result:
            if value not in deduped:
                deduped.append(value)
        return deduped or fallback

    per_multiples = multiples(per_values, [5, 10, 15, 20, 25], MAX_REASONABLE_PER)
    pbr_multiples = multiples(pbr_values, [0.5, 1, 1.5, 2, 2.5], MAX_REASONABLE_PBR)
    for row in rows:
        eps = to_float(row.get("eps"))
        bps = to_float(row.get("bps"))
        row["per_lines"] = [round(eps * multiple, 2) if eps is not None else None for multiple in per_multiples]
        row["pbr_lines"] = [round(bps * multiple, 2) if bps is not None else None for multiple in pbr_multiples]
    return {
        "rows": rows,
        "per_multiples": per_multiples,
        "pbr_multiples": pbr_multiples,
    }


def valuation_band_summary(rows: list[dict[str, Any]], metric: str) -> dict[str, float | None]:
    values = [to_float(row.get(metric)) for row in rows]
    values = [value for value in values if value is not None and math.isfinite(value)]
    if not values:
        return {"min": None, "avg": None, "max": None}
    return {
        "min": round(min(values), 2),
        "avg": round(sum(values) / len(values), 2),
        "max": round(max(values), 2),
    }


def build_dart_earnings_trend(company: str) -> dict[str, Any]:
    raw_company = str(company or "").strip()
    if not raw_company:
        raise ValueError("기업명을 입력해 주세요.")
    stock_code = ""
    stock_name = raw_company
    if re.fullmatch(r"A?\d{6}", raw_company, flags=re.IGNORECASE):
        stock_code = re.sub(r"\D", "", raw_company).zfill(6)
        row = (get_screening_stock_lookup().get("by_code") or {}).get(stock_code)
        if row is None:
            try:
                row = find_listing_row_by_code(stock_code)
            except Exception:
                row = None
        if row:
            stock_name = str(row.get("name") or stock_name)
    if not stock_code:
        resolved_code, resolved_name = resolve_stock_from_screening_cache(raw_company)
        if not resolved_code:
            try:
                resolved_code, resolved_name = resolve_stock(raw_company)
            except Exception:
                resolved_code, resolved_name = None, raw_company
        if resolved_code:
            stock_code = str(resolved_code).zfill(6)
            stock_name = resolved_name or stock_name
    if not stock_code:
        raise ValueError(f"'{raw_company}' 종목코드를 찾지 못했습니다. 자동완성에서 종목을 선택하거나 종목코드로 검색해 주세요.")

    cache_key = re.sub(r"[^0-9A-Za-z_-]", "_", stock_code)
    cache_path = DART_EARNINGS_TREND_CACHE_DIR / f"{cache_key}.json"
    if cache_path.exists():
        try:
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
            cached_at = datetime.fromisoformat(str(cached.get("cached_at") or ""))
            if cached.get("cache_version") == 10 and datetime.now() - cached_at < timedelta(hours=12):
                return cached
        except Exception:
            pass

    client = get_dart_client()
    if client is None:
        raise ValueError("DART API 키가 없어 실적 추이를 불러올 수 없습니다.")

    current_year = date.today().year
    annual_rows: list[dict[str, Any]] = []
    annual_by_year: dict[int, dict[str, Any]] = {}
    for year in range(current_year, current_year - 11, -1):
        row = dart_financial_row(client, stock_code, year, "11011")
        if not row:
            continue
        annual_row = {"year": year, "label": str(year), **row}
        annual_rows.append(annual_row)
        annual_by_year[year] = annual_row
        if len(annual_rows) >= 8:
            break
    annual_rows = sorted(annual_rows, key=lambda item: item["year"])[-8:]
    valuation_annual_rows = list(annual_rows)

    quarter_rows: list[dict[str, Any]] = []
    quarter_codes = [(1, "11013"), (2, "11012"), (3, "11014")]
    quarter_by_year: dict[int, dict[int, dict[str, Any]]] = {}
    for year in range(current_year, current_year - 5, -1):
        quarter_by_year[year] = {}
        for quarter, report_code in quarter_codes:
            row = dart_financial_row(client, stock_code, year, report_code)
            if row:
                quarter_by_year[year][quarter] = row
                quarter_rows.append({"year": year, "quarter": quarter, "label": f"{year}.{quarter}Q", **row})
        annual = annual_by_year.get(year) or dart_financial_row(client, stock_code, year, "11011")
        if annual and all(q in quarter_by_year[year] for q in [1, 2, 3]):
            q4 = {}
            for metric in ["sales", "operating_profit", "net_income"]:
                annual_value = annual.get(metric)
                partial_sum = sum(
                    quarter_by_year[year][q].get(metric) or 0
                    for q in [1, 2, 3]
                    if quarter_by_year[year][q].get(metric) is not None
                )
                q4[metric] = round(annual_value - partial_sum, 2) if annual_value is not None else None
            if any(q4.get(metric) is not None for metric in ["sales", "operating_profit", "net_income"]):
                quarter_rows.append({"year": year, "quarter": 4, "label": f"{year}.4Q", **q4})
    quarter_rows = sorted(quarter_rows, key=lambda item: (int(item.get("year") or 0), int(item.get("quarter") or 0)))[-12:]

    if current_year not in annual_by_year and quarter_by_year.get(current_year):
        latest_quarter = max(quarter_by_year[current_year])
        latest_ytd = quarter_by_year[current_year][latest_quarter]
        ytd_row = {
            "year": current_year,
            "label": f"{current_year} YTD",
            "quarter": latest_quarter,
            "is_ytd": True,
            **latest_ytd,
        }
        if any(ytd_row.get(metric) is not None for metric in ["sales", "operating_profit", "net_income"]):
            annual_rows = sorted(annual_rows + [ytd_row], key=lambda item: (int(item.get("year") or 0), 1 if item.get("is_ytd") else 0))[-9:]

    valuation_rows = build_valuation_band_rows(stock_code, valuation_annual_rows)
    price_band = build_price_band_series(stock_code, valuation_annual_rows)

    payload = {
        "cache_version": 10,
        "company": stock_name,
        "stock_code": stock_code,
        "unit": "억원",
        "source": "OpenDartReader / DART finstate",
        "cached_at": datetime.now().isoformat(timespec="seconds"),
        "quarters": quarter_rows,
        "annual": annual_rows,
        "valuation": {
            "source": "DART annual net income/equity + FinanceDataReader year-end close",
            "rows": valuation_rows,
            "per_band": valuation_band_summary(valuation_rows, "per"),
            "pbr_band": valuation_band_summary(valuation_rows, "pbr"),
            "price_band": price_band,
        },
    }
    DART_EARNINGS_TREND_CACHE_DIR.mkdir(parents=True, exist_ok=True)
    cache_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    return payload


def earnings_report_priority(item: dict[str, Any]) -> tuple[int, str]:
    report_name = str(item.get("report_name") or "")
    score = 0
    if "연결" in report_name:
        score += 30
    if "잠정" in report_name:
        score += 10
    if "정정" in report_name:
        score += 5
    if item.get("metrics", {}).get("sales", {}).get("expected"):
        score += 3
    if item.get("metrics", {}).get("operating_profit", {}).get("expected"):
        score += 3
    return score, str(item.get("date") or "")


def dedupe_market_earnings_results(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    selected: dict[str, dict[str, Any]] = {}
    for item in items:
        stock_code = str(item.get("stock_code") or item.get("company_code") or "").zfill(6)
        item_date = str(item.get("date") or "")[:10]
        key = f"{stock_code}:{item_date}" if stock_code and item_date else f"{item.get('company')}:{item_date}:{item.get('message_id')}"
        current = selected.get(key)
        if current is None or earnings_report_priority(item) > earnings_report_priority(current):
            selected[key] = item
    return sorted(selected.values(), key=lambda item: item.get("date") or "", reverse=True)


async def telegram_earnings_search_payload(request: TelegramEarningsSearchRequest) -> dict[str, Any]:
    company = str(request.company or "").strip()
    if not company:
        raise ValueError("검색할 기업명을 입력해 주세요.")

    target = build_disclosure_company_target(company)
    category = normalize_disclosure_category(request.category)
    category_label = TELEGRAM_DISCLOSURE_CATEGORIES[category]["label"]
    result_limit = max(1, min(request.limit if request.limit and request.limit > 0 else 30, 60))
    search_limit = max(result_limit * 5, 80)
    offset_id = int(request.offset_id or 0) or None
    oldest_date = date.today() - timedelta(days=365 * 3 + 1)
    client, temp_dir = build_telegram_readonly_client()
    try:
        await client.connect()
        if not await client.is_user_authorized():
            raise ValueError("Telegram login is required first.")

        dialog = await find_earnings_dialog(client)
        queries = []
        for query in target.get("query_terms", [company]):
            cleaned = query.strip()
            if cleaned and cleaned not in queries:
                queries.append(cleaned)

        collected: dict[str, dict[str, Any]] = {}
        scanned = 0
        oldest_message_id = offset_id
        for query in queries:
            async for message in client.iter_messages(
                dialog.entity,
                search=query,
                limit=search_limit,
                offset_id=offset_id or 0,
            ):
                scanned += 1
                message_date = message.date.date() if getattr(message, "date", None) else None
                if message_date and message_date < oldest_date:
                    break
                text = get_message_search_blob(message)
                if not is_disclosure_category_text(text, category):
                    continue
                key = f"{int(dialog.id)}:{int(message.id)}"
                if key in collected:
                    continue
                base = build_search_result_payload(dialog, message, text, [company])
                parsed = parse_telegram_earnings_payload(base)
                if not disclosure_result_matches_company(parsed, target):
                    continue
                parsed["disclosure_category"] = category
                parsed["disclosure_category_label"] = category_label
                collected[key] = parsed
                oldest_message_id = int(message.id)
                if len(collected) >= result_limit:
                    break
            if len(collected) >= result_limit:
                break

        results = sorted(collected.values(), key=lambda item: item["date"], reverse=True)[:result_limit]
        next_offset_id = min([int(item["message_id"]) for item in results], default=0)
        has_more = bool(results) and next_offset_id > 0
        return {
            "company": company,
            "resolved_company": target.get("name") or company,
            "stock_code": target.get("code") or "",
            "category": category,
            "category_label": category_label,
            "channel_name": dialog.name,
            "result_count": len(results),
            "results": results,
            "next_offset_id": next_offset_id,
            "has_more": has_more,
            "scanned_count": scanned,
            "oldest_date": oldest_date.isoformat(),
            "message": f"{dialog.name}에서 {len(results)}개의 {category_label} 공시를 찾았습니다.",
        }
    finally:
        await client.disconnect()
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)


def compact_earnings_search_job(job: dict[str, Any]) -> dict[str, Any]:
    return {
        "job_id": job["job_id"],
        "status": job["status"],
        "company": job.get("company", ""),
        "resolved_company": job.get("resolved_company", ""),
        "stock_code": job.get("stock_code", ""),
        "category": job.get("category", ""),
        "category_label": job.get("category_label", ""),
        "channel_name": job.get("channel_name", ""),
        "result_count": len(job.get("results", [])),
        "results": job.get("results", []),
        "next_offset_id": job.get("next_offset_id", 0),
        "has_more": bool(job.get("has_more")),
        "scanned_count": int(job.get("scanned_count", 0)),
        "oldest_date": job.get("oldest_date", ""),
        "message": job.get("message", ""),
        "error": job.get("error", ""),
        "finished": job["status"] in {"completed", "failed", "cancelled"},
    }


async def run_telegram_earnings_search_job(job_id: str, request: TelegramEarningsSearchRequest) -> None:
    job = TELEGRAM_EARNINGS_SEARCH_JOBS[job_id]
    company = str(request.company or "").strip()
    try:
        if not company:
            raise ValueError("검색할 기업명을 입력해 주세요.")
        target = build_disclosure_company_target(company)
        category = normalize_disclosure_category(request.category)
        category_label = TELEGRAM_DISCLOSURE_CATEGORIES[category]["label"]
        result_limit = max(1, min(request.limit if request.limit and request.limit > 0 else 30, 60))
        search_limit = max(result_limit * 5, 80)
        offset_id = int(request.offset_id or 0) or None
        oldest_date = date.today() - timedelta(days=365 * 3 + 1)
        job.update({
            "company": company,
            "resolved_company": target.get("name") or company,
            "stock_code": target.get("code") or "",
            "category": category,
            "category_label": category_label,
            "oldest_date": oldest_date.isoformat(),
            "message": "텔레그램 공시 채널 연결 중...",
        })
        client = build_telegram_client()
        temp_dir = None
        try:
            await client.connect()
            if not await client.is_user_authorized():
                raise ValueError("Telegram login is required first.")
            dialog = await find_earnings_dialog(client)
            job["channel_name"] = dialog.name
            queries: list[str] = []
            for query in target.get("query_terms", [company]):
                cleaned = str(query or "").strip()
                if cleaned and cleaned not in queries:
                    queries.append(cleaned)
            collected: dict[str, dict[str, Any]] = {}
            scanned = 0
            job["message"] = f"{dialog.name}에서 {category_label} 공시 검색 중..."
            for query in queries:
                if job_is_cancelled(job):
                    job["status"] = "cancelled"
                    job["message"] = "검색이 중단되었습니다."
                    return
                job["message"] = f"{dialog.name}에서 '{query}' 검색 중..."
                async for message in client.iter_messages(
                    dialog.entity,
                    search=query,
                    limit=search_limit,
                    offset_id=offset_id or 0,
                ):
                    if job_is_cancelled(job):
                        job["status"] = "cancelled"
                        job["message"] = "검색이 중단되었습니다."
                        return
                    scanned += 1
                    job["scanned_count"] = scanned
                    message_date = message.date.date() if getattr(message, "date", None) else None
                    if message_date and message_date < oldest_date:
                        break
                    text = get_message_search_blob(message)
                    if not is_disclosure_category_text(text, category):
                        continue
                    key = f"{int(dialog.id)}:{int(message.id)}"
                    if key in collected:
                        continue
                    base = build_search_result_payload(dialog, message, text, [company])
                    parsed = parse_telegram_earnings_payload(base)
                    if not disclosure_result_matches_company(parsed, target):
                        continue
                    parsed["disclosure_category"] = category
                    parsed["disclosure_category_label"] = category_label
                    collected[key] = parsed
                    results = sorted(collected.values(), key=lambda item: item["date"], reverse=True)[:result_limit]
                    job["results"] = results
                    job["next_offset_id"] = min([int(item["message_id"]) for item in results], default=0)
                    job["has_more"] = bool(results) and int(job["next_offset_id"] or 0) > 0
                    job["message"] = f"{dialog.name}에서 {len(results)}개 발견 · {scanned}개 후보 확인"
                    if len(collected) >= result_limit:
                        break
                if len(collected) >= result_limit:
                    break
            results = sorted(collected.values(), key=lambda item: item["date"], reverse=True)[:result_limit]
            job["results"] = results
            job["next_offset_id"] = min([int(item["message_id"]) for item in results], default=0)
            job["has_more"] = bool(results) and int(job["next_offset_id"] or 0) > 0
            job["status"] = "completed"
            job["message"] = f"{dialog.name}에서 {len(results)}개의 {category_label} 공시를 찾았습니다."
        finally:
            await client.disconnect()
            if temp_dir:
                shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception as exc:
        job["status"] = "failed"
        job["error"] = str(exc)
        job["message"] = "실적 공시 검색 중 오류가 발생했습니다."


def launch_telegram_earnings_search_job(job_id: str, request_payload: dict[str, Any]) -> None:
    def runner() -> None:
        request = TelegramEarningsSearchRequest(**request_payload)
        job = TELEGRAM_EARNINGS_SEARCH_JOBS.get(job_id)
        if job and job_is_cancelled(job):
            job["status"] = "cancelled"
            job["message"] = "검색이 시작되기 전에 중단되었습니다."
            return
        acquired = TELEGRAM_EARNINGS_LOCK.acquire(timeout=8)
        if not acquired:
            if job:
                job["status"] = "failed"
                job["error"] = "다른 공시 검색이 아직 정리되지 않았습니다. 잠시 후 다시 검색해 주세요."
                job["message"] = "공시 검색 준비 시간이 길어져 중단했습니다."
            return
        try:
            if job and job_is_cancelled(job):
                job["status"] = "cancelled"
                job["message"] = "검색이 시작되기 전에 중단되었습니다."
                return
            asyncio.run(run_telegram_earnings_search_job(job_id, request))
        finally:
            TELEGRAM_EARNINGS_LOCK.release()

    thread = threading.Thread(target=runner, name=f"telegram-earnings-{job_id[:8]}", daemon=True)
    thread.start()


async def telegram_market_earnings_payload(request: TelegramMarketEarningsRequest) -> dict[str, Any]:
    days = max(1, min(int(request.days or 1095), 365 * 5))
    result_limit = max(1, min(int(request.limit or 100), 1000))
    scan_limit = max(result_limit, min(int(request.scan_limit or 5000), 20000))
    oldest_date = date.today() - timedelta(days=days)
    universe = screening_market_cap_universe()
    client, temp_dir = build_telegram_readonly_client()
    try:
        await client.connect()
        if not await client.is_user_authorized():
            raise ValueError("Telegram login is required first.")

        dialog = await find_earnings_dialog(client)
        results: list[dict[str, Any]] = []
        seen: set[str] = set()
        scanned = 0
        async for message in client.iter_messages(dialog.entity, limit=scan_limit):
            scanned += 1
            message_date = message.date.date() if getattr(message, "date", None) else None
            if message_date and message_date < oldest_date:
                break
            text = get_message_search_blob(message)
            if not is_earnings_disclosure_text(text):
                continue
            base = build_search_result_payload(dialog, message, text, [])
            parsed = parse_telegram_earnings_payload(base)
            stock_info = match_market_earnings_stock(parsed, universe)
            if not stock_info:
                continue
            key = f"{int(dialog.id)}:{int(message.id)}"
            if key in seen:
                continue
            seen.add(key)
            parsed["stock_name"] = stock_info.get("stock_name")
            parsed["stock_code"] = stock_info.get("stock_code")
            parsed["market"] = stock_info.get("market")
            parsed["market_cap_100m"] = stock_info.get("market_cap_100m")
            parsed["disclosure_category"] = "earnings"
            parsed["disclosure_category_label"] = TELEGRAM_DISCLOSURE_CATEGORIES["earnings"]["label"]
            results.append(parsed)
            if len(results) >= result_limit:
                break
        results = dedupe_market_earnings_results(results)[:result_limit]
        return {
            "channel_name": dialog.name,
            "days": days,
            "scan_limit": scan_limit,
            "scanned_count": scanned,
            "universe_count": len({item.get("stock_code") for item in universe.values() if item.get("stock_code")}),
            "result_count": len(results),
            "results": results,
            "oldest_date": oldest_date.isoformat(),
            "message": f"시총 2000억 이상 종목의 실적 공시 {len(results)}개를 찾았습니다.",
        }
    finally:
        await client.disconnect()
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)


def compact_search_job(job: dict[str, Any]) -> dict[str, Any]:
    return {
        "job_id": job["job_id"],
        "status": job["status"],
        "keywords": job["keywords"],
        "match_mode": job["match_mode"],
        "exact_phrase": job["exact_phrase"],
        "has_file": job["has_file"],
        "start_date": job["start_date"],
        "end_date": job["end_date"],
        "selected_chat_count": job["selected_chat_count"],
        "result_count": len(job["results"]),
        "results": job["results"],
        "processed_chat_count": job["processed_chat_count"],
        "total_chat_count": job["total_chat_count"],
        "message": job.get("message", ""),
        "error": job.get("error", ""),
        "finished": job["status"] in {"completed", "failed", "cancelled"},
    }


def job_is_cancelled(job: dict[str, Any]) -> bool:
    return bool(job.get("cancel_requested"))


def request_cancel_running_telegram_jobs(
    message: str,
    exclude_search_job_id: str | None = None,
    exclude_earnings_job_id: str | None = None,
) -> dict[str, int]:
    cancelled_search = 0
    cancelled_earnings = 0
    for job_id, job in TELEGRAM_SEARCH_JOBS.items():
        if job_id == exclude_search_job_id:
            continue
        if job.get("status") == "running":
            job["cancel_requested"] = True
            job["message"] = message
            cancelled_search += 1
    for job_id, job in TELEGRAM_EARNINGS_SEARCH_JOBS.items():
        if job_id == exclude_earnings_job_id:
            continue
        if job.get("status") == "running":
            job["cancel_requested"] = True
            job["message"] = message
            cancelled_earnings += 1
    return {"search_jobs": cancelled_search, "earnings_jobs": cancelled_earnings}


def keyword_match(text: str, keywords: list[str], exact_phrase: bool, match_mode: str) -> tuple[bool, list[str]]:
    if not keywords:
        return False, []
    haystack = normalize_search_text(text)
    matched: list[str] = []
    for keyword in keywords:
        needles = telegram_keyword_match_needles(keyword, exact_phrase)
        if any(needle and needle in haystack for needle in needles):
            matched.append(keyword)
    if match_mode == "all":
        return len(matched) == len(keywords), matched
    return len(matched) > 0, matched


async def telegram_search_payload(request: TelegramSearchRequest) -> dict[str, Any]:
    keywords = normalize_keyword_list(request.keywords)
    if not keywords:
        raise ValueError("Please enter at least one keyword.")

    client, temp_dir = build_telegram_readonly_client()
    try:
        await client.connect()
        if not await client.is_user_authorized():
            raise ValueError("Telegram login is required first.")

        dialogs = []
        async for dialog in client.iter_dialogs():
            dialogs.append(dialog)

        selected_dialogs = dialogs
        if request.chat_ids:
            selected_ids = set(request.chat_ids)
            selected_dialogs = [dialog for dialog in dialogs if int(dialog.id) in selected_ids]

        start_date = parse_search_date(request.start_date)
        end_date = parse_search_date(request.end_date)
        if start_date and end_date and start_date > end_date:
            raise ValueError("Start date cannot be later than end date.")

        candidate_messages: dict[str, dict[str, Any]] = {}
        result_limit = request.limit if request.limit and request.limit > 0 else 200

        if request.chat_ids:
            for dialog in selected_dialogs:
                for search_query in telegram_search_queries_for_keywords(keywords, request.match_mode):
                    async for message in client.iter_messages(dialog.entity, search=search_query, limit=1000):
                        text = get_message_search_blob(message)
                        message_date = message.date.date() if getattr(message, "date", None) else None
                        if end_date and message_date and message_date > end_date:
                            continue
                        if start_date and message_date and message_date < start_date:
                            break
                        has_match, matched_keywords = keyword_match(text, keywords, request.exact_phrase, request.match_mode)
                        if not has_match:
                            continue
                        file_flag = message_has_file(message)
                        if request.has_file is not None and file_flag != request.has_file:
                            continue
                        key = f"{int(dialog.id)}:{int(message.id)}"
                        existing = candidate_messages.get(key)
                        merged_keywords = matched_keywords
                        if existing:
                            merged_keywords = sorted(set(existing["matched_keywords"] + matched_keywords))
                        candidate_messages[key] = build_search_result_payload(dialog, message, text, merged_keywords)
                        if len(candidate_messages) >= result_limit:
                            break
                    if len(candidate_messages) >= result_limit:
                        break

                if len(candidate_messages) >= result_limit:
                    break
        else:
            candidate_messages = await collect_global_search_results(
                client,
                keywords,
                request,
                result_limit,
                start_date,
                end_date,
            )

        results = list(candidate_messages.values())
        results.sort(key=lambda item: item["date"], reverse=True)
        results = results[:result_limit]

        return {
            "keywords": keywords,
            "match_mode": request.match_mode,
            "exact_phrase": request.exact_phrase,
            "has_file": request.has_file,
            "start_date": request.start_date,
            "end_date": request.end_date,
            "selected_chat_count": len(selected_dialogs) if request.chat_ids else len({item.get("chat_id") for item in results}),
            "result_count": len(results),
            "results": results,
        }
    finally:
        await client.disconnect()
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)


async def run_telegram_search_job(job_id: str, request: TelegramSearchRequest) -> None:
    job = TELEGRAM_SEARCH_JOBS[job_id]
    keywords = normalize_keyword_list(request.keywords)
    start_date = parse_search_date(request.start_date)
    end_date = parse_search_date(request.end_date)

    try:
        client, temp_dir = build_telegram_readonly_client()
        try:
            await client.connect()
            if not await client.is_user_authorized():
                raise ValueError("Telegram login is required first.")

            dialogs = []
            async for dialog in client.iter_dialogs():
                dialogs.append(dialog)

            selected_dialogs = dialogs
            if request.chat_ids:
                selected_ids = set(request.chat_ids)
                selected_dialogs = [dialog for dialog in dialogs if int(dialog.id) in selected_ids]

            result_limit = request.limit if request.limit and request.limit > 0 else 200
            recent_limit = 400
            job["selected_chat_count"] = len(selected_dialogs)
            job["total_chat_count"] = len(selected_dialogs)
            job["message"] = "Searching recent messages..."
            collected: dict[str, dict[str, Any]] = {}

            if not request.chat_ids:
                collected = await collect_global_search_results(
                    client,
                    keywords,
                    request,
                    result_limit,
                    start_date,
                    end_date,
                    job,
                )
            else:
                for dialog in selected_dialogs:
                    for search_query in telegram_search_queries_for_keywords(keywords, request.match_mode):
                        scanned_count = 0
                        async for message in client.iter_messages(dialog.entity, search=search_query, limit=1000):
                            if job_is_cancelled(job):
                                job["status"] = "cancelled"
                                job["message"] = "Search cancelled."
                                return
                            scanned_count += 1
                            text = get_message_search_blob(message)
                            message_date = message.date.date() if getattr(message, "date", None) else None
                            if end_date and message_date and message_date > end_date:
                                continue
                            if start_date and message_date and message_date < start_date:
                                break
                            has_match, matched_keywords = keyword_match(text, keywords, request.exact_phrase, request.match_mode)
                            if not has_match:
                                continue
                            file_flag = message_has_file(message)
                            if request.has_file is not None and file_flag != request.has_file:
                                continue
                            key = f"{int(dialog.id)}:{int(message.id)}"
                            existing = collected.get(key)
                            merged_keywords = matched_keywords
                            if existing:
                                merged_keywords = sorted(set(existing["matched_keywords"] + matched_keywords))
                            collected[key] = build_search_result_payload(dialog, message, text, merged_keywords)
                            job["results"] = sorted(collected.values(), key=lambda item: item["date"], reverse=True)[:result_limit]
                            job["message"] = (
                                f"Searching {dialog.name or 'chat'} for '{search_query}': {scanned_count} hits checked, "
                                f"{len(job['results'])} results found."
                            )
                            if len(collected) >= result_limit:
                                break
                        if len(collected) >= result_limit:
                            break

                    job["processed_chat_count"] += 1
                    job["message"] = (
                        f"Checked {job['processed_chat_count']} / {job['total_chat_count']} chats, "
                        f"{len(job['results'])} results so far."
                    )
                    if len(collected) >= result_limit:
                        break

            job["results"] = sorted(collected.values(), key=lambda item: item["date"], reverse=True)[:result_limit]
            job["status"] = "completed"
            job["message"] = "Search completed."
        finally:
            await client.disconnect()
            if temp_dir:
                shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception as exc:
        job["status"] = "failed"
        job["error"] = str(exc)
        job["message"] = "An error occurred during search."


def launch_telegram_search_job(job_id: str, request_payload: dict[str, Any]) -> None:
    def runner() -> None:
        request = TelegramSearchRequest(**request_payload)
        job = TELEGRAM_SEARCH_JOBS.get(job_id)
        if job and job_is_cancelled(job):
            job["status"] = "cancelled"
            job["message"] = "Search cancelled before it started."
            return
        asyncio.run(run_telegram_search_job(job_id, request))

    thread = threading.Thread(target=runner, name=f"telegram-search-{job_id[:8]}", daemon=True)
    thread.start()


app = FastAPI(title="\uc8fc\uc2dd \ub300\uc2dc\ubcf4\ub4dc")


@app.middleware("http")
async def prevent_stale_frontend_cache(request, call_next):
    response = await call_next(request)
    path = request.url.path
    if path == "/" or path.startswith("/static/") or path.startswith("/vendor/"):
        response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
    return response


app.mount("/static/vendor", StaticFiles(directory=VENDOR_FRONTEND_DIR), name="static-vendor")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")
app.mount("/vendor", StaticFiles(directory=VENDOR_FRONTEND_DIR), name="vendor")


@app.get("/api/health")
def health_check() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/app-config")
def app_config() -> dict[str, Any]:
    public_web = is_public_web_mode()
    return {
        "public_web": public_web,
        "locked_features": ["telegram", "kis"] if public_web else [],
        "message": "Public web mode" if public_web else "Local desktop mode",
    }


@app.get("/api/kis/status")
def kis_status(check_token: bool = False) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Korea Investment API")
    try:
        return JSONResponse(kis_status_payload(check_token=check_token))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/tradingview/open")
def tradingview_open(request: TradingViewOpenRequest) -> JSONResponse:
    try:
        return JSONResponse(open_tradingview_desktop(request.stock_code, request.stock_name))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/stocks/investor-flows")
def stock_investor_flows(code: str | None = None, name: str | None = None, days: int = 31) -> JSONResponse:
    try:
        return JSONResponse(build_stock_investor_flows(code=code, stock_name=name, days=days))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/portfolio/performance")
def portfolio_performance(refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(get_cached_portfolio_performance(force_refresh=refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/stock-alert/status")
def stock_alert_status() -> JSONResponse:
    try:
        return JSONResponse(stock_alert_status_payload())
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/stock-alert/github-settings")
def stock_alert_github_settings(request: StockAlertGitHubSettingsRequest) -> JSONResponse:
    try:
        repository = request.repository.strip()
        token = request.token.strip()
        if "/" not in repository:
            return JSONResponse({"error": "repository는 owner/name 형식이어야 합니다."}, status_code=400)
        if not token:
            return JSONResponse({"error": "GitHub token이 필요합니다."}, status_code=400)
        settings = load_settings()
        stock_alert = settings.get("stock_alert") if isinstance(settings.get("stock_alert"), dict) else {}
        stock_alert["github_repository"] = repository
        stock_alert["github_token"] = token
        settings["stock_alert"] = stock_alert
        save_settings(settings)
        return JSONResponse(stock_alert_status_payload())
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/stock-alert/telegram-settings")
def stock_alert_telegram_settings(request: StockAlertTelegramSettingsRequest) -> JSONResponse:
    try:
        bot_token = request.bot_token.strip()
        chat_id = str(request.chat_id or "").strip()
        if not bot_token:
            return JSONResponse({"error": "Telegram bot token이 필요합니다."}, status_code=400)
        settings = load_settings()
        stock_alert = settings.get("stock_alert") if isinstance(settings.get("stock_alert"), dict) else {}
        stock_alert["telegram_bot_token"] = bot_token
        if chat_id:
            stock_alert["telegram_chat_id"] = chat_id
        settings["stock_alert"] = stock_alert
        save_settings(settings)
        return JSONResponse(stock_alert_status_payload())
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/stock-alert/detect-telegram-chat")
def stock_alert_detect_telegram_chat() -> JSONResponse:
    try:
        return JSONResponse(detect_telegram_chat_id())
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/stock-alert/sync-telegram-secrets")
def stock_alert_sync_telegram_secrets() -> JSONResponse:
    try:
        return JSONResponse(sync_stock_alert_telegram_secrets())
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/stock-alert/holdings")
def stock_alert_holdings() -> JSONResponse:
    try:
        holdings = latest_stock_alert_holdings(latest_non_empty=True)
        return JSONResponse(
            {
                "holdings": holdings,
                "holding_count": len(holdings),
                "source_date": holdings[0].get("source_date") if holdings else "",
            }
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/stock-alert/sync-holdings")
def stock_alert_sync_holdings() -> JSONResponse:
    try:
        return JSONResponse(sync_stock_alert_holdings_secret())
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/portfolio/export.xlsx", response_model=None)
def portfolio_export_xlsx():
    try:
        output_path = create_portfolio_export_workbook()
        return FileResponse(
            path=output_path,
            filename=output_path.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/strategy/backtest")
def strategy_backtest(
    index: str = "KS11",
    strategy: str = "ma20_cross",
    start: str | None = None,
    end: str | None = None,
    top_n: int | None = None,
    entry_threshold: float | None = None,
    exit_threshold: float | None = None,
    allocation_mode: str = "score_weight",
) -> JSONResponse:
    try:
        return JSONResponse(
            build_strategy_backtest(
                index=index,
                strategy=strategy,
                start=start,
                end=end,
                top_n=top_n,
                entry_threshold=entry_threshold,
                exit_threshold=exit_threshold,
                allocation_mode=allocation_mode,
            )
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/strategy/sector-rotation")
def strategy_sector_rotation(
    start: str | None = None,
    end: str | None = None,
    min_score: float = 50.0,
    top_sectors: int = 4,
    stocks_per_sector: int = 4,
    leverage: float = 1.0,
    weight_method: str = "strength",
) -> JSONResponse:
    try:
        return JSONResponse(
            build_sector_rotation_backtest(
                start=start,
                end=end,
                min_score=min_score,
                top_sectors=top_sectors,
                stocks_per_sector=stocks_per_sector,
                leverage=leverage,
                weight_method=weight_method,
            )
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/strategy/advanced-sector")
def strategy_advanced_sector(
    start: str | None = None,
    end: str | None = None,
    min_score: float = 50.0,
    top_sectors: int = 4,
    stocks_per_sector: int = 4,
    max_leverage: float = 1.5,
    benchmark: str = "KS11",
    weight_method: str = "entry_beta",
    beta_window: int = 63,
    min_breadth: float = 60.0,
    max_disparity: float = 110.0,
    trading_rank_limit: int = 20,
    stock_selection: str = "trend_strength",
) -> JSONResponse:
    try:
        return JSONResponse(
            build_advanced_sector_backtest(
                start=start,
                end=end,
                min_score=min_score,
                top_sectors=top_sectors,
                stocks_per_sector=stocks_per_sector,
                max_leverage=max_leverage,
                benchmark=benchmark,
                weight_method=weight_method,
                beta_window=beta_window,
                min_breadth=min_breadth,
                max_disparity=max_disparity,
                trading_rank_limit=trading_rank_limit,
                stock_selection=stock_selection,
            )
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/strategy/portfolio-diagnostic")
def strategy_portfolio_diagnostic() -> JSONResponse:
    try:
        portfolio_mtime = PORTFOLIO_PATH.stat().st_mtime if PORTFOLIO_PATH.exists() else 0
        sector_mtime = SECTOR_DB_PATH.stat().st_mtime if SECTOR_DB_PATH.exists() else 0
        cache_key = f"{portfolio_mtime:.0f}:{sector_mtime:.0f}:{date.today().isoformat()}"
        cache = getattr(strategy_portfolio_diagnostic, "_cache", {})
        if cache.get("key") == cache_key and isinstance(cache.get("payload"), dict):
            return JSONResponse(cache["payload"])
        payload = build_portfolio_diagnostic()
        setattr(strategy_portfolio_diagnostic, "_cache", {"key": cache_key, "payload": payload})
        return JSONResponse(payload)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/themes/today")
def themes_today(
    min_score: float = 50.0,
    recent_limit: int = RECENT_SCREENING_LOOKBACK,
    file_date: str | None = None,
) -> JSONResponse:
    try:
        return JSONResponse(load_screening_summary(min_score=min_score, recent_limit=recent_limit, file_date=file_date))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/themes/reload")
def themes_reload(request: ThemeReloadRequest) -> JSONResponse:
    try:
        return JSONResponse(reload_screening_cache(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/us-themes/today")
def us_themes_today(
    min_score: float = 50.0,
    recent_limit: int = RECENT_SCREENING_LOOKBACK,
    file_date: str | None = None,
) -> JSONResponse:
    try:
        return JSONResponse(load_us_screening_summary(min_score=min_score, recent_limit=recent_limit, file_date=file_date))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/us-themes/reload")
def us_themes_reload(request: ThemeReloadRequest) -> JSONResponse:
    try:
        return JSONResponse(reload_us_screening_cache(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/asia-themes/today")
def asia_themes_today(
    min_score: float = 50.0,
    recent_limit: int = RECENT_SCREENING_LOOKBACK,
    file_date: str | None = None,
    region: str = "jp",
) -> JSONResponse:
    try:
        return JSONResponse(load_asia_screening_summary(min_score=min_score, recent_limit=recent_limit, file_date=file_date, region=region))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/asia-themes/reload")
def asia_themes_reload(request: ThemeReloadRequest) -> JSONResponse:
    try:
        return JSONResponse(reload_asia_screening_cache(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/themes/test-excel")
def themes_test_excel(request: ThemeTestExcelRequest) -> JSONResponse:
    try:
        return JSONResponse(create_theme_test_excel(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/themes/build-today-excel")
def themes_build_today_excel(request: ThemeBuildTodayExcelRequest) -> JSONResponse:
    try:
        return JSONResponse(create_theme_today_excel_and_reload(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/themes/build-today-data")
def themes_build_today_data(request: ThemeBuildTodayExcelRequest) -> JSONResponse:
    try:
        return JSONResponse(create_theme_today_excel_and_reload(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/us-themes/build-today-data")
def us_themes_build_today_data(request: ThemeBuildTodayExcelRequest) -> JSONResponse:
    try:
        return JSONResponse(create_us_theme_today_data_and_reload(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/asia-themes/build-today-data")
def asia_themes_build_today_data(request: ThemeBuildTodayExcelRequest) -> JSONResponse:
    try:
        return JSONResponse(create_asia_theme_today_data_and_reload(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/themes/note")
def themes_note_update(request: ThemeNoteUpdateRequest) -> JSONResponse:
    try:
        payload = update_screening_note(request)
        return JSONResponse(payload, status_code=423 if payload.get("locked") else 200)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/themes/score-history")
def themes_score_history(
    code: str | None = None,
    name: str | None = None,
    end_date: str | None = None,
    days: int = 31,
) -> JSONResponse:
    try:
        return JSONResponse(
            build_stock_score_history(
                stock_code=code,
                stock_name=name,
                end_date=end_date,
                days=max(7, min(days, 90)),
            )
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/stocks/autocomplete")
def stocks_autocomplete(q: str = "", limit: int = 12) -> JSONResponse:
    try:
        return JSONResponse({"query": q, "items": autocomplete_stocks(q, limit=max(1, min(limit, 20)))})
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/news/search")
def news_search(q: str = "", limit: int = 30, days: int = 365) -> JSONResponse:
    try:
        return JSONResponse(search_stock_news(q, limit=limit, days=days))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/stocks/chart-preview")
def stocks_chart_preview(
    code: str | None = None,
    name: str | None = None,
    months: int = 3,
) -> JSONResponse:
    try:
        return JSONResponse(load_stock_chart_preview(stock_code=code, stock_name=name, months=months))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/themes/sector-market-cap-chart")
def themes_sector_market_cap_chart(request: SectorMarketCapChartRequest) -> JSONResponse:
    try:
        return JSONResponse(build_sector_market_cap_chart(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/global-stocks/search")
def global_stocks_search(q: str = "", limit: int = 12) -> JSONResponse:
    try:
        return JSONResponse({"query": q, "items": search_global_companies(q, limit=max(1, min(limit, 20)))})
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/global-stocks/detail")
def global_stocks_detail(symbol: str) -> JSONResponse:
    try:
        return JSONResponse(build_global_company_detail(symbol))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/global-stocks/ai-brief")
def global_stocks_ai_brief(symbol: str, force_refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(build_global_company_ai_brief(symbol, force_refresh=force_refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/market-calendar")
def market_calendar(start: str | None = None, end: str | None = None, refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(market_calendar_payload(start=start, end=end, refresh=refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/market-calendar/reload")
def market_calendar_reload(start: str | None = None, end: str | None = None) -> JSONResponse:
    try:
        return JSONResponse(market_calendar_payload(start=start, end=end, refresh=True))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/market-calendar/events")
def market_calendar_add_event(request: MarketCalendarEventRequest) -> JSONResponse:
    try:
        current = load_market_calendar_events()
        event = normalize_market_calendar_event(request.model_dump())
        if not event:
            return JSONResponse({"error": "일정 날짜와 제목을 확인해 주세요."}, status_code=400)
        current = [item for item in current if item.get("id") != event.get("id")]
        save_market_calendar_events(current + [event])
        return JSONResponse(market_calendar_payload())
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/market-calendar.ics")
def market_calendar_ics() -> Response:
    try:
        payload = market_calendar_payload(
            start=(date.today() - timedelta(days=7)).isoformat(),
            end=(date.today() + timedelta(days=90)).isoformat(),
        )
        content = build_market_calendar_ics(payload.get("events", []))
        return Response(
            content=content,
            media_type="text/calendar; charset=utf-8",
            headers={"Content-Disposition": 'attachment; filename="stock_dashboard_market_calendar.ics"'},
        )
    except Exception as exc:
        return Response(content=str(exc), status_code=500, media_type="text/plain; charset=utf-8")


@app.get("/api/global-indices")
def global_indices(group: str | None = None) -> JSONResponse:
    try:
        cache_key = f"{GLOBAL_INDICES_PAYLOAD_CACHE_VERSION}|{str(group or '')}"
        cached = GLOBAL_INDICES_PAYLOAD_CACHE.get(cache_key)
        if cached and (time.time() - cached[0]) < 10 * 60:
            return JSONResponse(cached[1])
        disk_key = hashlib.md5(cache_key.encode("utf-8")).hexdigest() or "default"
        disk_cache_path = GLOBAL_INDICES_PAYLOAD_CACHE_DIR / f"{disk_key}.json"
        if disk_cache_path.exists() and (time.time() - disk_cache_path.stat().st_mtime) < 6 * 60 * 60:
            payload = json.loads(disk_cache_path.read_text(encoding="utf-8"))
            GLOBAL_INDICES_PAYLOAD_CACHE[cache_key] = (time.time(), payload)
            return JSONResponse(payload)
        payload = build_global_indices_payload(group=group)
        GLOBAL_INDICES_PAYLOAD_CACHE[cache_key] = (time.time(), payload)
        GLOBAL_INDICES_PAYLOAD_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        disk_cache_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        return JSONResponse(payload)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/sector-watch-board")
def sector_watch_board(limit_per_sector: int = 6) -> JSONResponse:
    try:
        return JSONResponse(build_sector_watch_board(limit_per_sector=limit_per_sector))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/sector-watch-board/order")
def sector_watch_board_order(request: SectorWatchOrderRequest) -> JSONResponse:
    try:
        return JSONResponse(save_sector_watch_order(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/sector-db")
def sector_db_read() -> JSONResponse:
    try:
        db = load_sector_db()
        return JSONResponse(
            {
                "stock_map": db.get("stock_map", {}),
                "sectors": db.get("sectors", []),
                "groups": sector_db_groups(),
            }
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/sector-db/assign")
def sector_db_assign(request: SectorAssignmentRequest) -> JSONResponse:
    try:
        db = upsert_sector_assignment(request.stock_code, request.stock_name, request.sector)
        return JSONResponse(
            {
                "stock_map": db.get("stock_map", {}),
                "sectors": db.get("sectors", []),
                "groups": sector_db_groups(),
            }
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/sector-db/groups")
def sector_db_save_groups(request: SectorDatabaseSaveRequest) -> JSONResponse:
    try:
        db = save_sector_groups_to_db(request.groups)
        return JSONResponse(
            {
                "stock_map": db.get("stock_map", {}),
                "sectors": db.get("sectors", []),
                "groups": sector_db_groups(),
            }
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/theme-sector-calendar")
def theme_sector_calendar(min_score: float = 50.0, limit: int = 60, force_refresh: bool = False, score_basis: str = "score") -> JSONResponse:
    try:
        return JSONResponse(build_theme_sector_calendar(min_score=min_score, limit=max(1, min(limit, 120)), force_refresh=force_refresh, score_basis=score_basis))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/us-theme-sector-calendar")
def us_theme_sector_calendar(min_score: float = 50.0, limit: int = 60) -> JSONResponse:
    try:
        return JSONResponse(build_us_theme_sector_calendar(min_score=min_score, limit=max(1, min(limit, 120))))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/asia-theme-sector-calendar")
def asia_theme_sector_calendar(min_score: float = 50.0, limit: int = 60, region: str = "jp") -> JSONResponse:
    try:
        return JSONResponse(build_asia_theme_sector_calendar(min_score=min_score, limit=max(1, min(limit, 120)), region=region))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/sector-snapshot/preview")
def sector_snapshot_preview(request: SectorSnapshotRequest) -> JSONResponse:
    try:
        return JSONResponse(build_sector_snapshot(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/sector-snapshot/market-ytd")
def sector_snapshot_market_ytd(limit: int = 100) -> JSONResponse:
    try:
        return JSONResponse(build_market_ytd_ranking(limit=limit))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/sector-snapshot/entry-signals")
def sector_snapshot_entry_signals(
    start: str | None = None,
    end: str | None = None,
    lookback_days: int = 80,
    min_score: float = 50.0,
    trading_rank_limit: int = 20,
    min_avg_score: float = 55.0,
    min_strong_count: int = 2,
    min_stock_count: int = 1,
    beta_window: int = 120,
) -> JSONResponse:
    try:
        return JSONResponse(
            build_sector_entry_signals(
                start=start,
                end=end,
                lookback_days=lookback_days,
                min_score=min_score,
                trading_rank_limit=trading_rank_limit,
                min_avg_score=min_avg_score,
                min_strong_count=min_strong_count,
                min_stock_count=min_stock_count,
                beta_window=beta_window,
            )
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/sector-snapshot/signal-radar")
def sector_snapshot_signal_radar(
    lookback_days: int = 120,
    max_stocks: int = 45,
    min_score: float = 50.0,
    max_history_events: int = 220,
) -> JSONResponse:
    try:
        return JSONResponse(
            build_signal_radar(
                lookback_days=lookback_days,
                max_stocks=max_stocks,
                min_score=min_score,
                max_history_events=max_history_events,
            )
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/trade/import-export")
def trade_import_export(force_refresh: bool = False, region: str = "all", item_key: str = "") -> JSONResponse:
    try:
        payload = load_trade_import_export_payload(force_refresh=force_refresh, region=region, item_key=item_key)
        region_codes = {str(item["code"]) for item in TRADE_REGION_OPTIONS}
        selected_region = region if region in region_codes else "all"
        payload["region_options"] = TRADE_REGION_OPTIONS
        payload["selected_region"] = selected_region
        payload["selected_item_key"] = item_key or payload.get("selected_item_key") or TRADE_IMPORT_EXPORT_ITEMS[0]["key"]
        payload["region_note"] = payload.get("region_note") or "전국 HS코드 기준 수출금액입니다."
        return JSONResponse(payload)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/dram/prices")
def dram_prices(force_refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(build_dram_price_payload(force_refresh=force_refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/ssd/prices")
def ssd_prices(force_refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(build_ssd_price_payload(force_refresh=force_refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/tourism/inbound-visitors")
def tourism_inbound_visitors(force_refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(build_tourist_inbound_visitors_payload(force_refresh=force_refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/economy/cycle-clock")
def economy_cycle_clock(force_refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(build_economy_cycle_payload(force_refresh=force_refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc), "indicators": []}, status_code=200)


@app.get("/api/economy/sector-cycle-clock")
def economy_sector_cycle_clock(min_score: float = 50.0, limit: int = 40, force_refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(build_sector_cycle_clock_payload(min_score=min_score, limit=limit, force_refresh=force_refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc), "groups": []}, status_code=200)


@app.get("/api/real-estate/prices")
def real_estate_prices(force_refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(build_real_estate_price_payload(force_refresh=force_refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/real-estate/trade-detail")
def real_estate_trade_detail(region_name: str, full_name: str = "", force_refresh: bool = False) -> JSONResponse:
    try:
        return JSONResponse(build_real_estate_trade_detail(region_name=region_name, full_name=full_name, force_refresh=force_refresh))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/real-estate/building")
def real_estate_building() -> JSONResponse:
    try:
        db = load_real_estate_db()
        db["summary"] = build_real_estate_summary(db)
        return JSONResponse(db)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/real-estate/building")
def real_estate_building_save(request: RealEstateSaveRequest) -> JSONResponse:
    try:
        db = save_real_estate_db(request.data)
        db["summary"] = build_real_estate_summary(db)
        return JSONResponse(db)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/real-estate/water/sync-telegram")
async def real_estate_water_sync_telegram(request: RealEstateWaterTelegramSyncRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram water billing")
    try:
        return JSONResponse(await sync_real_estate_water_from_telegram(request))
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)


@app.post("/api/real-estate/electricity/sync-telegram")
async def real_estate_electricity_sync_telegram(request: RealEstateElectricTelegramSyncRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram electricity billing")
    try:
        return JSONResponse(await sync_real_estate_electricity_from_telegram(request))
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)


@app.post("/api/real-estate/bank/import-files")
def real_estate_bank_import_files() -> JSONResponse:
    try:
        return JSONResponse(import_real_estate_bank_files())
    except Exception as exc:
        return JSONResponse({"ok": False, "error": str(exc)}, status_code=500)


@app.post("/api/sector-snapshot/export.xlsx", response_model=None)
def sector_snapshot_export(request: SectorSnapshotRequest):
    try:
        payload = build_sector_snapshot(request)
        output_path = create_sector_snapshot_workbook(payload)
        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=output_path.name,
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/dart/today")
def dart_today(
    min_score: float = 50.0,
    file_date: str | None = None,
) -> JSONResponse:
    return JSONResponse(load_dart_summary(min_score=min_score, file_date=file_date))


@app.get("/api/telegram/status")
async def telegram_status() -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram")
    try:
        return JSONResponse(await telegram_status_payload())
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/telegram/ui_state")
def telegram_ui_state_read() -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram UI state")
    try:
        settings = load_settings()
        state = settings.get("telegram_ui_state", {})
        return JSONResponse({"state": state if isinstance(state, dict) else {}})
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/telegram/ui_state")
def telegram_ui_state_save(request: TelegramUiStateRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram UI state")
    try:
        settings = load_settings()
        settings["telegram_ui_state"] = request.state
        save_settings(settings)
        return JSONResponse({"ok": True})
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/telegram/send_code")
async def telegram_send_code(request: TelegramSendCodeRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram login")
    try:
        save_telegram_settings(request.api_id, request.api_hash, request.phone)
        with TELEGRAM_LOCK:
            client = build_telegram_client()
            try:
                await client.connect()
                result = await client.send_code_request(request.phone)
                save_telegram_login_state(request.phone, result.phone_code_hash)
            finally:
                await client.disconnect()
        return JSONResponse(
            {
                "ok": True,
                "message": "Verification code sent. Enter the code from Telegram.",
                "needs_code": True,
            }
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/telegram/verify_code")
async def telegram_verify_code(request: TelegramVerifyCodeRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram login")
    try:
        login_state = load_telegram_login_state()
        phone_code_hash = str(login_state.get("phone_code_hash", "")).strip()
        if not phone_code_hash:
            return JSONResponse(
                {
                    "error": "No login request is waiting. Please request a verification code again first."
                },
                status_code=400,
            )
        with TELEGRAM_LOCK:
            client = build_telegram_client()
            try:
                await client.connect()
                try:
                    await client.sign_in(phone=request.phone, code=request.code, phone_code_hash=phone_code_hash)
                except SessionPasswordNeededError:
                    if not request.password:
                        return JSONResponse(
                            {
                                "ok": False,
                                "needs_password": True,
                                "message": "This account uses two-step verification. Please enter your password.",
                            }
                        )
                    await client.sign_in(password=request.password)
            finally:
                await client.disconnect()
        clear_telegram_login_state()
        return JSONResponse(
            {
                "ok": True,
                "message": "Telegram login completed.",
            }
        )
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/telegram/search")
async def telegram_search(request: TelegramSearchRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram search")
    try:
        request_cancel_running_telegram_jobs("새 텔레그램 검색을 위해 이전 검색을 중단했습니다.")
        return JSONResponse(await telegram_search_payload(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/telegram/earnings_search")
async def telegram_earnings_search(request: TelegramEarningsSearchRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram disclosure search")
    try:
        request_cancel_running_telegram_jobs("새 공시 검색을 위해 이전 텔레그램 검색을 중단했습니다.")
        return JSONResponse(await telegram_earnings_search_payload(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/telegram/earnings_search_jobs")
async def telegram_earnings_search_jobs(request: TelegramEarningsSearchRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram disclosure search")
    try:
        category = normalize_disclosure_category(request.category)
        job_id = uuid.uuid4().hex
        request_cancel_running_telegram_jobs(
            "새 공시 검색을 위해 이전 텔레그램 검색을 중단했습니다.",
            exclude_earnings_job_id=job_id,
        )
        TELEGRAM_EARNINGS_SEARCH_JOBS[job_id] = {
            "job_id": job_id,
            "status": "running",
            "company": request.company,
            "resolved_company": "",
            "stock_code": "",
            "category": category,
            "category_label": TELEGRAM_DISCLOSURE_CATEGORIES[category]["label"],
            "channel_name": "",
            "results": [],
            "next_offset_id": 0,
            "has_more": False,
            "scanned_count": 0,
            "oldest_date": "",
            "message": "실적 공시 검색을 시작했습니다.",
            "error": "",
            "cancel_requested": False,
        }
        launch_telegram_earnings_search_job(
            job_id,
            {
                "company": request.company,
                "category": request.category,
                "limit": request.limit,
                "offset_id": request.offset_id,
            },
        )
        return JSONResponse(compact_earnings_search_job(TELEGRAM_EARNINGS_SEARCH_JOBS[job_id]))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/telegram/jobs/cancel_all")
async def telegram_cancel_all_jobs() -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram search")
    cancelled = request_cancel_running_telegram_jobs("모든 텔레그램 검색을 중단했습니다.")
    return JSONResponse({"ok": True, "cancelled": cancelled})


@app.get("/api/telegram/earnings_search_jobs/{job_id}")
async def telegram_earnings_search_job_status(job_id: str) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram disclosure search")
    job = TELEGRAM_EARNINGS_SEARCH_JOBS.get(job_id)
    if not job:
        return JSONResponse({"error": "Earnings search job not found."}, status_code=404)
    return JSONResponse(compact_earnings_search_job(job))


@app.post("/api/telegram/earnings_search_jobs/{job_id}/cancel")
async def telegram_earnings_search_job_cancel(job_id: str) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram disclosure search")
    job = TELEGRAM_EARNINGS_SEARCH_JOBS.get(job_id)
    if not job:
        return JSONResponse({"error": "Earnings search job not found."}, status_code=404)
    if job["status"] == "running":
        job["cancel_requested"] = True
        job["message"] = "검색 중단을 요청했습니다."
    return JSONResponse(compact_earnings_search_job(job))


@app.get("/api/dart/earnings-trend")
def dart_earnings_trend(company: str = "") -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("DART earnings trend")
    try:
        return JSONResponse(build_dart_earnings_trend(company))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/kind/latest-business-report")
def kind_latest_business_report(company: str = "") -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("KIND business report")
    try:
        return JSONResponse(find_latest_kind_business_report(company))
    except LookupError as exc:
        return JSONResponse({"error": str(exc)}, status_code=404)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/kind/latest-periodic-report")
def kind_latest_periodic_report(company: str = "") -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("KIND periodic report")
    try:
        return JSONResponse(find_latest_kind_periodic_report(company))
    except LookupError as exc:
        return JSONResponse({"error": str(exc)}, status_code=404)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/kind/business-segments")
def kind_business_segments(company: str = "") -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("KIND business segments")
    try:
        return JSONResponse(load_kind_business_segments(company))
    except LookupError as exc:
        return JSONResponse({"error": str(exc)}, status_code=404)
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/telegram/market_earnings")
async def telegram_market_earnings(request: TelegramMarketEarningsRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram market earnings search")
    try:
        return JSONResponse(await telegram_market_earnings_payload(request))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.post("/api/telegram/search_jobs")
async def telegram_search_jobs(request: TelegramSearchRequest) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram search")
    try:
        keywords = normalize_keyword_list(request.keywords)
        if not keywords:
            raise ValueError("Please enter at least one keyword.")
        job_id = uuid.uuid4().hex
        TELEGRAM_SEARCH_JOBS[job_id] = {
            "job_id": job_id,
            "status": "running",
            "keywords": keywords,
            "match_mode": request.match_mode,
            "exact_phrase": request.exact_phrase,
            "has_file": request.has_file,
            "start_date": request.start_date,
            "end_date": request.end_date,
            "selected_chat_count": 0,
            "processed_chat_count": 0,
            "total_chat_count": 0,
            "results": [],
            "message": "Preparing search...",
            "error": "",
            "cancel_requested": False,
        }
        launch_telegram_search_job(
            job_id,
            {
                "keywords": list(request.keywords),
                "chat_ids": list(request.chat_ids or []),
                "has_file": request.has_file,
                "match_mode": request.match_mode,
                "exact_phrase": request.exact_phrase,
                "start_date": request.start_date,
                "end_date": request.end_date,
                "limit": request.limit,
            },
        )
        return JSONResponse(compact_search_job(TELEGRAM_SEARCH_JOBS[job_id]))
    except Exception as exc:
        return JSONResponse({"error": str(exc)}, status_code=500)


@app.get("/api/telegram/search_jobs/{job_id}")
async def telegram_search_job_status(job_id: str) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram search")
    job = TELEGRAM_SEARCH_JOBS.get(job_id)
    if not job:
        return JSONResponse({"error": "Search job not found."}, status_code=404)
    return JSONResponse(compact_search_job(job))


@app.post("/api/telegram/search_jobs/{job_id}/cancel")
async def telegram_search_job_cancel(job_id: str) -> JSONResponse:
    if is_public_web_mode():
        return public_web_lock_response("Telegram search")
    job = TELEGRAM_SEARCH_JOBS.get(job_id)
    if not job:
        return JSONResponse({"error": "Search job not found."}, status_code=404)
    if job["status"] == "running":
        job["cancel_requested"] = True
        job["message"] = "Cancel requested."
    return JSONResponse(compact_search_job(job))


@app.get("/api/telegram/attachment/{chat_id}/{message_id}")
async def telegram_attachment(chat_id: int, message_id: int) -> FileResponse:
    if is_public_web_mode():
        raise HTTPException(status_code=403, detail="Telegram attachments are locked in public web mode.")
    with TELEGRAM_LOCK:
        client, temp_dir = build_telegram_readonly_client()
        try:
            await client.connect()
            if not await client.is_user_authorized():
                raise HTTPException(status_code=401, detail="?붾젅洹몃옩 濡쒓렇?몄씠 ?꾩슂?⑸땲??")

            entity = await client.get_input_entity(chat_id)
            message = await client.get_messages(entity, ids=message_id)
            if not message:
                raise HTTPException(status_code=404, detail="硫붿떆吏瑜?李얠쓣 ???놁뒿?덈떎.")
            if not message_has_file(message):
                raise HTTPException(status_code=404, detail="泥⑤??뚯씪???녿뒗 硫붿떆吏?낅땲??")

            display_name, stored_name, _ = build_attachment_metadata(chat_id, message_id, message)
            TELEGRAM_ATTACHMENT_DIR.mkdir(parents=True, exist_ok=True)
            target_path = TELEGRAM_ATTACHMENT_DIR / stored_name
            if not target_path.exists():
                downloaded = await client.download_media(message, file=str(target_path))
                if not downloaded:
                    raise HTTPException(status_code=500, detail="泥⑤??뚯씪??遺덈윭?ㅼ? 紐삵뻽?듬땲??")
                target_path = Path(downloaded)

            media_type = mimetypes.guess_type(display_name)[0] or "application/octet-stream"
            return FileResponse(path=target_path, filename=display_name, media_type=media_type)
        except HTTPException:
            raise
        except Exception as exc:
            raise HTTPException(status_code=500, detail=str(exc)) from exc
        finally:
            await client.disconnect()
            if temp_dir:
                shutil.rmtree(temp_dir, ignore_errors=True)


def render_shared_dashboard_html() -> str:
    try:
        payload = load_screening_summary(min_score=0, recent_limit=RECENT_SCREENING_LOOKBACK, file_date=None)
        calendar = build_theme_sector_calendar(min_score=0, limit=10)
        rows = payload.get("qualified_stocks", []) if isinstance(payload, dict) else []
        sectors = payload.get("manual_sector_summary") or payload.get("theme_summary") or []
        file_date = html_lib.escape(str(payload.get("file_date") or ""))
        file_name = html_lib.escape(str(payload.get("file_name") or ""))
        error_message = ""
    except Exception as exc:
        rows = []
        sectors = []
        calendar = {"days": []}
        file_date = ""
        file_name = ""
        error_message = html_lib.escape(str(exc))

    def safe(value: Any) -> str:
        return html_lib.escape("" if value is None else str(value))

    def fmt(value: Any, digits: int = 2) -> str:
        try:
            number = float(value)
            return f"{number:,.{digits}f}"
        except Exception:
            return "-"

    top_rows = rows[:80]
    sector_chips = "".join(
        f"<span class='chip'><b>{safe(item.get('sector') or item.get('theme') or '-')}</b> "
        f"{fmt(item.get('avg_score'), 1)}점</span>"
        for item in sectors[:12]
        if isinstance(item, dict)
    )
    flow_days = calendar.get("days", []) if isinstance(calendar, dict) else []
    flow_html = "".join(
        "<div class='flow-day'>"
        f"<div class='flow-date'>{safe(day.get('date'))}</div>"
        + "".join(
            f"<span>{safe(sector.get('sector'))}</span>"
            for sector in (day.get("sectors") or [])[:5]
            if isinstance(sector, dict)
        )
        + "</div>"
        for day in flow_days[-5:]
        if isinstance(day, dict)
    )
    body_rows = "".join(
        "<tr>"
        f"<td class='rank'>{safe(index + 1)}</td>"
        f"<td>{safe(row.get('sector') or row.get('theme') or '-')}</td>"
        f"<td><b>{safe(row.get('resolved_name') or row.get('stock_name'))}</b> <small>{safe(row.get('stock_code'))}</small></td>"
        f"<td>{safe(row.get('industry') or '-')}</td>"
        f"<td class='num {'up' if float(row.get('change_pct') or 0) >= 0 else 'down'}'>{fmt(row.get('change_pct'), 2)}%</td>"
        f"<td class='num'>{fmt(row.get('score'), 2)}</td>"
        f"<td>{safe(row.get('note') or '-')}</td>"
        "</tr>"
        for index, row in enumerate(top_rows)
        if isinstance(row, dict)
    )
    if not body_rows:
        body_rows = f"<tr><td colspan='7' class='empty'>{error_message or '표시할 데이터가 없습니다.'}</td></tr>"

    return f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Stock Dashboard Shared</title>
  <style>
    body {{ margin: 0; background: #f6f8fb; color: #172033; font-family: Arial, sans-serif; }}
    .wrap {{ max-width: 1180px; margin: 0 auto; padding: 22px; }}
    .hero, .panel {{ background: #fff; border: 1px solid #dbe3ef; border-radius: 18px; padding: 18px; box-shadow: 0 10px 30px rgba(15, 23, 42, .06); }}
    .hero {{ margin-bottom: 14px; }}
    h1 {{ margin: 0 0 8px; font-size: 26px; }}
    p {{ margin: 6px 0; color: #64748b; }}
    .actions {{ display: flex; gap: 8px; flex-wrap: wrap; margin-top: 12px; }}
    a.button {{ display: inline-block; padding: 9px 12px; border-radius: 999px; background: #2563eb; color: #fff; text-decoration: none; font-weight: 700; }}
    a.button.secondary {{ background: #e8eef8; color: #1e3a8a; }}
    .chips {{ display: flex; gap: 8px; flex-wrap: wrap; margin: 12px 0; }}
    .chip {{ display: inline-flex; gap: 6px; align-items: center; padding: 7px 10px; border-radius: 999px; background: #eef6ff; border: 1px solid #cfe1ff; font-size: 13px; }}
    .flow {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(160px, 1fr)); gap: 8px; margin: 12px 0; }}
    .flow-day {{ padding: 10px; border-radius: 14px; background: #f8fafc; border: 1px solid #e2e8f0; }}
    .flow-date {{ font-weight: 800; margin-bottom: 6px; color: #475569; }}
    .flow-day span {{ display: block; margin-top: 5px; padding: 5px 8px; border-radius: 999px; background: #16a34a; color: #fff; font-size: 12px; font-weight: 700; }}
    .table-box {{ overflow-x: auto; margin-top: 12px; }}
    table {{ width: 100%; border-collapse: collapse; min-width: 920px; font-size: 13px; }}
    th {{ background: #f1f5f9; color: #334155; text-align: left; padding: 9px; position: sticky; top: 0; }}
    td {{ border-top: 1px solid #e2e8f0; padding: 8px 9px; vertical-align: middle; }}
    small {{ color: #64748b; font-weight: 700; }}
    .rank {{ font-weight: 900; text-align: center; }}
    .num {{ text-align: right; font-variant-numeric: tabular-nums; }}
    .up {{ color: #dc2626; font-weight: 800; }}
    .down {{ color: #2563eb; font-weight: 800; }}
    .empty {{ text-align: center; color: #64748b; padding: 28px; }}
  </style>
</head>
<body>
  <div class="wrap">
    <section class="hero">
      <h1>오늘의 주도주</h1>
      <p>{file_date} · {file_name}</p>
      <p>이 화면은 외부 공유용 경량 페이지입니다. 데이터 계산은 모두 로컬 PC에서 실행되고, Cloudflare Tunnel은 결과 화면만 전달합니다.</p>
      <div class="actions">
        <a class="button" href="/shared">새로고침</a>
        <a class="button secondary" href="/?shared=1">전체 대시보드 앱 열기</a>
      </div>
    </section>
    <section class="panel">
      <h2>수동 섹터 요약</h2>
      <div class="chips">{sector_chips or "<span class='chip'>섹터 데이터 없음</span>"}</div>
      <h2>최근 주도 섹터 흐름</h2>
      <div class="flow">{flow_html or "<p>흐름 데이터가 없습니다.</p>"}</div>
      <h2>엑셀 원본 순서 테이블</h2>
      <div class="table-box">
        <table>
          <thead>
            <tr>
              <th>순위</th><th>섹터</th><th>종목</th><th>업종</th><th>등락률</th><th>종합점수</th><th>비고</th>
            </tr>
          </thead>
          <tbody>{body_rows}</tbody>
        </table>
      </div>
    </section>
  </div>
</body>
</html>"""


@app.get("/shared")
def shared_dashboard() -> FileResponse:
    return FileResponse(FRONTEND_DIR / "index.html")


@app.get("/")
def read_index() -> FileResponse:
    return FileResponse(FRONTEND_DIR / "index.html")


def main() -> None:
    uvicorn.run(
        "backend.app:app",
        host=os.getenv("STOCK_DASHBOARD_HOST", "127.0.0.1"),
        port=int(os.getenv("STOCK_DASHBOARD_PORT", "8123")),
        reload=False,
        app_dir=str(BASE_DIR),
    )


if __name__ == "__main__":
    main()

